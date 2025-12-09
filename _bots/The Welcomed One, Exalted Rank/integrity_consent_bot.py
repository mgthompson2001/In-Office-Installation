# integrity_consent_bot.py
import sys, os, re, time, tempfile, threading, queue
import datetime
from datetime import datetime as DT
import json

# --- GUI imports (keep these first so we fail early if tkinter is missing) ---
try:
    import tkinter as tk
    from tkinter import filedialog, scrolledtext
    from tkinter import ttk
except Exception as e:
    print("Tkinter failed to load. You may need to install/repair Python with Tk support.")
    print("Error:", e)
    sys.exit(1)

from welcome_uploader_bridge import add_uploader_button
APP_TITLE = "ISWS Consent Form Bot"
HELP_TEXT = """ISWS Consent Form Bot — Quick Help

WHAT IT DOES
• Consent Form Bot (TherapyNotes):
  - Logs in, finds each client, pulls/sets Date (from Intake Note if missing),
    generates "{First Last} Consent NPP.pdf", flattens it, uploads it to Documents,
    and writes status back to column D in an _UPDATED.xlsx.

WHERE FILES GO
• Consent PDFs:  Desktop/Consent Form Bot/Consent Forms Complete/
• Updated Excel: saved beside your original as {OriginalName}_UPDATED.xlsx
• Consent PDF name: "{First} {Last} Consent NPP.pdf"

THERAPYNOTES (Consent) — KEY PARTS
• Login page: https://www.therapynotes.com/app/login/IntegritySWS/
  - Username:  #Login__UsernameField
  - Password:  #Login__Password
  - Log in:    #Login__LogInButton
• Patients list reset: click "Patients" (link text) or go to /app/patients/list
  - Global search: //input[@placeholder='Name, Acct #, Phone, or Ins ID']
• Open Documents tab on a patient:
  - a[data-tab-id='Documents']  (fallbacks: //a[@data-tab-id='Documents'] or text 'Documents')
• Upload dialog:
  - Open button (one of): "Upload Patient File" buttons (multiple fallbacks)
  - Date:           #PatientFile__Date
  - File chooser:   #InputUploader
  - Document Name:  #PatientFile__DocumentName
  - Submit:         //input[@type='button' and @value='Add Document']
• Upload success: dialog closes + new row with .documentNameSpan, or a visible "document added" toast.

OPTIONS (GUI)
• Write-back column: D (fixed)
• Skip if Date present (Consent): ON by default
• Dry run: simulate upload; still makes PDFs and statuses
• Flexible name match: ON by default (ignores middle initials/punct; tokens in order)

GUARDS & TEACH MODE
• Won't click destructive actions (delete/close/archive/save/complete…)
• Teach Mode prompts you to click when a control isn't found; it learns the locator and frame.
• Detects new tabs and switches automatically.

TROUBLESHOOTING
• Search box missing (TN): app auto-navigates to /app/patients/list and retries.
• Documents tab won't open: tries 3 locators, then skips with a clear status message.
• Upload unclear: logs "Upload uncertain — no confirmation found."
"""
WELCOME_TEACH_MODE = True   # set False later to disable manual assist
MAROON = "#800000"
# Task labels

# Default PN (Penelope) URL
PN_DEFAULT_URL = "https://integrityseniorservices.athena-us.com/acm_loginControl"

STATUS_COL_LETTER = "D"  # Write-back column requested by user
TASK_CONSENT = "ISWS Consent Form Bot"
# Track PN navigation state for safety guards
CTX = {"opened_task_label": None}

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import json

LEARN_DB_PATH = os.path.join(os.getcwd(), "learned_selectors.json")

class SelectorLearner:
    def __init__(self, db_path=LEARN_DB_PATH):
        self.db_path = db_path
        self.data = {}
        self._load()

    def _load(self):
        try:
            with open(self.db_path, "r", encoding="utf-8") as f:
                self.data = json.load(f) or {}
        except Exception:
            self.data = {}

    def _save(self):
        try:
            with open(self.db_path, "w", encoding="utf-8") as f:
                json.dump(self.data, f, indent=2)
        except Exception:
            pass

    def get(self, key):
        return (self.data.get(key) or {}).get("locators") or []

    def remember(self, key, locators, frame_hint=None):
        # locators is a list of {"type":"css"|"xpath","value": "..."} (ordered by preference)
        self.data[key] = {"locators": locators, "frame": (frame_hint or {})}
        self._save()

    def frame_hint(self, key):
        return (self.data.get(key) or {}).get("frame") or {}

    def _enter_frame_if_needed(self, driver, frame_hint):
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
        if not frame_hint:
            return True
        fid = frame_hint.get("id")
        fname = frame_hint.get("name")
        # Try ID, then NAME, else try any iframe that matches either
        try:
            if fid:
                fr = driver.find_element(By.ID, fid)
                driver.switch_to.frame(fr); return True
        except Exception:
            pass
        try:
            if fname:
                fr = driver.find_element(By.NAME, fname)
                driver.switch_to.frame(fr); return True
        except Exception:
            pass
        # Last resort: scan all iframes and match id/name
        try:
            frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
        except Exception:
            frames = []
        for fr in frames:
            try:
                if (fr.get_attribute("id") or "") == (fid or "") or (fr.get_attribute("name") or "") == (fname or ""):
                    driver.switch_to.frame(fr); return True
            except Exception:
                continue
        # Couldn’t find a matching frame—stay in default
        return True

    def locate(self, driver, key, timeout=8):
        locs = self.get(key)
        if not locs:
            return None
        fh = self.frame_hint(key)
        self._enter_frame_if_needed(driver, fh)
        for loc in locs:
            try:
                if loc["type"] == "css":
                    el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.CSS_SELECTOR, loc["value"])))
                else:
                    el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, loc["value"])))
                return el
            except Exception:
                continue
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
        return None

def ensure_click_capture_js(driver):
    """
    Injects a click listener into the page (and best-effort into same-origin iframes)
    so we can read the last element you clicked from Python.
    """
    js = r"""
(function install(win){
  try {
    if (!win._shadowTeachInstalled) {
      win._lastClickedElement = null;
      win.document.addEventListener('click', function(ev){
        try { win._lastClickedElement = ev.target; } catch(e){}
      }, true);
      win._shadowTeachInstalled = true;
    }
  } catch(e){}
  try {
    var frames = win.document.querySelectorAll('iframe,frame');
    for (var i=0;i<frames.length;i++){
      try { if (frames[i].contentWindow) install(frames[i].contentWindow); } catch(e){}
    }
  } catch(e){}
})(window);
return true;
"""
    try: driver.execute_script(js)
    except Exception: pass

def _unique_css_for_el(driver, el):
    getlen = lambda css: driver.execute_script("try { return document.querySelectorAll(arguments[0]).length; } catch(e){return 0;}", css)

    tag = (el.tag_name or "div").lower()
    _id = el.get_attribute("id") or ""
    if _id:
        css = f"#{_id}"
        if getlen(css) == 1:
            return css

    # Try common attributes
    attrs = ["name", "aria-label", "title", "data-testid", "role", "type", "value"]
    for a in attrs:
        v = el.get_attribute(a) or ""
        if v:
            css = f"{tag}[{a}=\"{v}\"]"
            if getlen(css) == 1:
                return css

    # Try class subset
    classes = (el.get_attribute("class") or "").strip().split()
    if classes:
        # limit to 2 classes to keep it stable
        css = tag + "".join("." + c for c in classes[:2])
        if getlen(css) == 1:
            return css

    # Scoped nth-of-type path (fallback)
    path = driver.execute_script("""
      function cssPath(el){
        if (!el) return null;
        if (el.id) return '#' + el.id;
        var parts=[];
        while (el && el.nodeType===1 && parts.length<8){
          var idx=1, sib=el;
          while ((sib = sib.previousElementSibling) != null) {
            if (sib.tagName===el.tagName) idx++;
          }
          parts.unshift(el.tagName.toLowerCase() + ':nth-of-type(' + idx + ')');
          el = el.parentElement;
          if (el && el.id) { parts.unshift('#'+el.id); break; }
        }
        return parts.join(' > ');
      }
      return cssPath(arguments[0]);
    """, el) or ""
    if path and getlen(path) == 1:
        return path
    return None

def _xpath_for_el(driver, el):
    return driver.execute_script(r"""
      function getXPath(el) {
        if (el.id) return '//*[@id="'+el.id+'"]';
        var parts=[];
        while(el && el.nodeType===1){
          var ix=1, sib=el.previousSibling;
          while(sib){ if(sib.nodeType===1 && sib.nodeName===el.nodeName) ix++; sib=sib.previousSibling; }
          parts.unshift(el.nodeName.toLowerCase()+'['+ix+']');
          el=el.parentNode;
        }
        return '/'+parts.join('/');
      }
      return getXPath(arguments[0]);
    """, el)

def _frame_hint_for_current_context(driver):
    # Are we inside a frame? If so, return id/name of the frame element.
    try:
        # If this throws, we’re likely at top; that’s fine.
        return driver.execute_script("""
          try {
            if (window === window.top) return null;
            var fe = window.frameElement;
            if (!fe) return null;
            return {id: fe.id || null, name: fe.name || null};
          } catch(e) { return null; }
        """) or None
    except Exception:
        return None

def _build_locators_from_element(driver, el):
    locs = []
    try:
        css = _unique_css_for_el(driver, el)
        if css:
            locs.append({"type": "css", "value": css})
    except Exception:
        pass
    try:
        xp = _xpath_for_el(driver, el)
        if xp:
            locs.append({"type": "xpath", "value": xp})
    except Exception:
        pass
    # Ensure we have at least one
    if not locs:
        locs = [{"type": "xpath", "value": "//body"}]  # placeholder, won't match; just to be safe
    return locs

def teach_and_learn_click(driver, learner, key, prompt_text, log, safety_check=None):
    """
    1) Injects click-capture JS.
    2) Pauses with your prompt.
    3) Grabs the element you just clicked, builds robust locators, saves them.
    4) Returns that WebElement (if still present) or None.
    """
    ensure_click_capture_js(driver)
    teach_pause(prompt_text, log)

    info = driver.execute_script("""
      return {
        element: window._lastClickedElement || document.activeElement || null,
        frame: (function(){
          try{
            if (window===window.top) return null;
            var fe = window.frameElement;
            if (!fe) return null;
            return {id: fe.id || null, name: fe.name || null};
          }catch(e){ return null; }
        })()
      };
    """)
    el = None
    try:
        el = info.get("element", None)
    except Exception:
        pass

    if not el:
        log("[PN][TEACH] I didn’t detect a click. I’ll continue, but nothing was learned.")
        return None

    # Optional safety
    if safety_check and safety_check(el):
        log("[PN][SAFEGUARD] That element looks destructive. I won’t learn/click it.")
        return None

    # Build & persist locators
    frame_hint = info.get("frame") or _frame_hint_for_current_context(driver)
    locs = _build_locators_from_element(driver, el)
    learner.remember(key, locs, frame_hint=frame_hint)
    log(f"[PN][LEARNED] Stored locator for {key}: {locs[0]['type']} -> {locs[0]['value']}")
    return el

def smart_click(driver, learner, key, default_locators, teach_text, log, timeout=10, is_harmful_fn=None):
    """
    Try learned locator → default locators → teach-and-learn. Clicks once if safe.
    """
    # 1) Learned first
    el = learner.locate(driver, key, timeout=timeout)
    if el:
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        except Exception:
            pass
        if not is_harmful_fn or not is_harmful_fn(el):
            try:
                el.click()
            except Exception:
                try: driver.execute_script("arguments[0].click();", el)
                except Exception: pass
            return True

    # 2) Try your defaults
    for by, sel in default_locators:
        try:
            try: driver.switch_to.default_content()
            except Exception: pass
            el = WebDriverWait(driver, 4).until(EC.element_to_be_clickable((by, sel)))
            if not is_harmful_fn or not is_harmful_fn(el):
                try: el.click()
                except Exception:
                    try: driver.execute_script("arguments[0].click();", el)
                    except Exception: continue
                # Learn this default for next time
                try:
                    fh = _frame_hint_for_current_context(driver)
                    if by == By.CSS_SELECTOR:
                        learner.remember(key, [{"type":"css","value":sel}], frame_hint=fh)
                    elif by == By.XPATH:
                        learner.remember(key, [{"type":"xpath","value":sel}], frame_hint=fh)
                except Exception:
                    pass
                return True
        except Exception:
            continue

    # 3) Teach-and-learn
    el = teach_and_learn_click(driver, learner, key, teach_text, log, safety_check=is_harmful_fn)
    if not el:
        return False
    try:
        el.click()
    except Exception:
        try: driver.execute_script("arguments[0].click();", el)
        except Exception: return False
    return True

# ========= TEACH KIT v2 (drop-in) =========

def teach_pause(msg, log_fn=None, title="Teach Mode"):
    """
    Show a real modal OK dialog without blocking on stdin() afterward.
    No input() call = no freeze.
    """
    import time
    # Log to your log pane or stdout
    try:
        (log_fn or print)(msg)
    except Exception:
        print(msg)

    # Best: Tkinter modal (works well with your Tk UI)
    try:
        import tkinter as tk
        from tkinter import messagebox
        root = tk._default_root
        if root is None:
            root = tk.Tk()
            root.withdraw()
        # bring to front
        try:
            root.after(0, root.lift)
            root.attributes("-topmost", True)
        except Exception:
            pass
        messagebox.showinfo(title, msg, parent=root)
        try:
            root.attributes("-topmost", False)
        except Exception:
            pass
        return
    except Exception:
        pass

    # Fallback: Windows MessageBox
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, msg, title, 0x00001040)  # MB_ICONINFORMATION
        return
    except Exception:
        pass

    # Last-resort: short sleep (no stdin wait!)
    time.sleep(2.0)


def ensure_click_capture_js_all_windows(driver):
    """
    Inject click-capture into EVERY window and best-effort into same-origin iframes.
    This is more robust than injecting only into the current tab.
    """
    js = r"""
(function install(win){
  try {
    if (!win._shadowTeachInstalled) {
      win._lastClickedElement = null;
      win.document.addEventListener('click', function(ev){
        try { win._lastClickedElement = ev.target; } catch(e){}
      }, true);
      win._shadowTeachInstalled = true;
    }
  } catch(e){}
  try {
    var frames = win.document.querySelectorAll('iframe,frame');
    for (var i=0;i<frames.length;i++){
      try { if (frames[i].contentWindow) install(frames[i].contentWindow); } catch(e){}
    }
  } catch(e){}
})(window);
return true;
"""
    try:
        current = driver.current_window_handle
    except Exception:
        current = None
    try:
        handles = driver.window_handles or []
    except Exception:
        handles = []
    for h in handles:
        try:
            driver.switch_to.window(h)
            driver.execute_script(js)
        except Exception:
            continue
    if current:
        try:
            driver.switch_to.window(current)
        except Exception:
            pass


def _get_last_clicked_info_from_all_windows(driver):
    """
    Read the last clicked element and frame hint from whichever window recorded it.
    Returns (handle, info_dict) or (None, None) if nothing found.
    """
    read_js = r"""
return {
  element: (window._lastClickedElement || document.activeElement || null),
  frame: (function(){
    try {
      if (window === window.top) return null;
      var fe = window.frameElement;
      if (!fe) return null;
      return {id: fe.id || null, name: fe.name || null};
    } catch(e) { return null; }
  })()
};
"""
    try:
        current = driver.current_window_handle
        handles = driver.window_handles or []
    except Exception:
        return (None, None)

    for h in handles:
        try:
            driver.switch_to.window(h)
            info = driver.execute_script(read_js)
            if info and info.get("element", None):
                # Found a click recorded in this handle
                if current:
                    try: driver.switch_to.window(current)
                    except Exception: pass
                return (h, info)
        except Exception:
            continue

    if current:
        try: driver.switch_to.window(current)
        except Exception: pass
    return (None, None)


def _maybe_switch_to_new_tab(driver, log=lambda m: None):
    """
    If a teach click opened a new tab/window, switch to the newest.
    """
    try:
        handles = driver.window_handles
        if len(handles) >= 2:
            driver.switch_to.window(handles[-1])
            log("[PN] Switched to newly opened tab.")
            return True
    except Exception:
        pass
    return False


def teach_and_learn_click(driver, learner, key, prompt_text, log, safety_check=None):
    """
    NEW: robust, non-freezing teach step.
    - Injects click capture into ALL windows/iframes
    - Prompts you to click the target
    - Records a stable CSS/XPath + frame hint
    - If your click opened a new tab, switches to it so the flow continues
    """
    # 1) Install capture everywhere first
    ensure_click_capture_js_all_windows(driver)

    # 2) Prompt (non-blocking)
    teach_pause(prompt_text, log)

    # 3) Read last-clicked from whichever window recorded it
    handle, info = _get_last_clicked_info_from_all_windows(driver)
    if not info:
        log("[PN][TEACH] I didn’t detect a click. I’ll continue, but nothing was learned.")
        return None

    el = None
    try:
        el = info.get("element", None)
    except Exception:
        el = None
    if not el:
        log("[PN][TEACH] I didn’t detect a usable element. Continuing.")
        return None

    # 4) Optional safety
    try:
        if safety_check and safety_check(el):
            log("[PN][SAFEGUARD] That element looks destructive. I won’t learn/click it.")
            return None
    except Exception:
        pass

    # 5) Build & persist locators, including frame hint
    try:
        frame_hint = info.get("frame") or _frame_hint_for_current_context(driver)
    except Exception:
        frame_hint = _frame_hint_for_current_context(driver)
    locs = _build_locators_from_element(driver, el)
    try:
        learner.remember(key, locs, frame_hint=frame_hint)
        log(f"[PN][LEARNED] Stored locator for {key}: {locs[0]['type']} -> {locs[0]['value']}")
    except Exception as e:
        log(f"[PN][LEARNED] Could not persist locator for {key}: {e}")

    # 6) If your click opened something in a new tab, go there so the pipeline can proceed
    _maybe_switch_to_new_tab(driver, log)

    # We do NOT click again here—the human already clicked during teach.
    # Return a *fresh* handle on the element if still present in this context
    # (not strictly required for your flow, but harmless if available).
    try:
        return learner.locate(driver, key, timeout=4)
    except Exception:
        return None

# ========= END TEACH KIT v2 =========

# ===================== Utilities =====================
def sanitize_filename(name: str) -> str:
    return re.sub(r'[<>:\"/\\|?*]+', "_", name).strip()

def to_mmddyyyy(val):
    """Normalize many date values (Excel serials, strings, datetime) to m/d/YYYY."""
    if val is None:
        return ""
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.strftime("%m/%d/%Y")
    # Excel serial dates
    try:
        if isinstance(val, (int, float)) and float(val) > 20000:
            base = datetime.datetime(1899, 12, 30)  # Excel on Windows base
            return (base + datetime.timedelta(days=float(val))).strftime("%m/%d/%Y")
    except Exception:
        pass
    s = str(val).strip()
    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d"):
        try:
            return DT.strptime(s, fmt).strftime("%m/%d/%Y")
        except Exception:
            continue
    return s  # last resort

def make_pdf_name(first, last):
    return f"{first.strip()} {last.strip()} Consent NPP.pdf".replace("  ", " ").strip()

def ensure_output_dir():
    """Desktop/Consent Form Bot/Consent Forms Complete (create if missing)."""
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    base_dir = os.path.join(desktop_dir, "Consent Form Bot")
    out_dir = os.path.join(base_dir, "Consent Forms Complete")
    os.makedirs(out_dir, exist_ok=True)
    return out_dir

def ensure_welcome_patient_dir(base_dir, full_name):
    """
    Create a per-patient folder inside the chosen base folder.
    If base_dir is empty/missing, default to Desktop/Welcome Bot Output.
    """
    if not base_dir:
        desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        base_dir = os.path.join(desktop_dir, "Welcome Bot Output")
    os.makedirs(base_dir, exist_ok=True)
    patient_dir = os.path.join(base_dir, sanitize_filename(full_name))
    os.makedirs(patient_dir, exist_ok=True)
    return patient_dir, base_dir



# ===================== PDF: Fill & Lock =====================
def fill_pdf_locked(template_path, output_path, field_values, log):
    """
    Fill a fillable PDF (by fuzzy field names), set NeedAppearances, and mark fields read-only.
    field_values e.g.: {"Client name print": "John Doe", "Date": "08/01/2025"}
    """
    try:
        from pdfrw import PdfReader, PdfWriter, PdfDict, PdfObject, PdfName, PdfString

        def norm(s): return " ".join(str(s or "").strip().lower().split())
        wanted = {norm(k): (k, v) for k, v in field_values.items()}

        pdf = PdfReader(template_path)
        # Ask viewers to regenerate appearances
        try:
            if pdf.Root and pdf.Root.AcroForm:
                pdf.Root.AcroForm.update(PdfDict(NeedAppearances=PdfObject("true")))
            else:
                pdf.Root.AcroForm = PdfDict(NeedAppearances=PdfObject("true"))
        except Exception:
            pass

        # Collect fields
        norm_to_annots = {}
        for page in pdf.pages:
            annots = getattr(page, "Annots", None)
            if not annots:
                continue
            for annot in annots:
                if getattr(annot, "Subtype", None) != PdfName.Widget or not getattr(annot, "T", None):
                    continue
                raw_name = annot.T.to_unicode() if hasattr(annot, "to_unicode") else str(annot.T)
                key_raw = str(raw_name).strip("()")
                key_norm = norm(key_raw)
                norm_to_annots.setdefault(key_norm, []).append(annot)

        # Set values (fuzzy)
        filled_any = False
        for want_norm, (orig_key, value) in wanted.items():
            if value is None:
                continue
            value_str = str(value)
            targets = []
            if want_norm in norm_to_annots:
                targets.extend(norm_to_annots[want_norm])
            if not targets:
                for k_norm, ann_list in norm_to_annots.items():
                    if want_norm in k_norm or k_norm in want_norm:
                        targets.extend(ann_list)
            if not targets:
                log(f'  [WARN] No PDF field matched "{orig_key}".')
                continue
            for annot in targets:
                try:
                    pdf_value = PdfString.encode(value_str)
                    annot.update(PdfDict(V=pdf_value, DV=pdf_value))
                    filled_any = True
                except Exception as e:
                    log(f'  [WARN] Failed to set "{orig_key}": {e}')

        # Make all fields read-only
        READONLY_BIT = 1
        for page in pdf.pages:
            annots = getattr(page, "Annots", None)
            if not annots:
                continue
            for annot in annots:
                try:
                    if getattr(annot, "Subtype", None) == PdfName.Widget:
                        ff = 0
                        try:
                            ff_raw = annot.get("/Ff", 0)
                            if ff_raw is not None:
                                ff = int(str(ff_raw))
                        except Exception:
                            ff = 0
                        annot.update(PdfDict(Ff=ff | READONLY_BIT))
                except Exception as e:
                    log(f"  [WARN] Could not set ReadOnly on a field: {e}")

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        PdfWriter().write(output_path, pdf)

        if not filled_any:
            log("  [INFO] No fields were filled (names may not have matched).")
        else:
            log("  [INFO] Fields set to read-only (not fillable).")
        return True
    except Exception as e:
        log(f"  [ERROR] fill_pdf failed: {e}")
        return False

# ===================== PDF: Flatten for Upload =====================
def flatten_pdf_preserving_values(input_pdf, output_pdf, log):
    """
    Draws field values onto pages and strips form fields, so the PDF is no longer fillable.
    """
    try:
        from pdfrw import PdfReader, PdfWriter, PdfName, PageMerge
        from reportlab.pdfgen import canvas as rl_canvas

        pdf = PdfReader(input_pdf)
        pages = pdf.pages or []

        # Build overlay with drawn text where fields are
        tmp_overlay = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp_overlay_path = tmp_overlay.name
        tmp_overlay.close()
        c = rl_canvas.Canvas(tmp_overlay_path)

        for page in pages:
            mb = page.MediaBox
            llx, lly, urx, ury = [float(x) for x in (mb[0], mb[1], mb[2], mb[3])]
            width, height = (urx - llx), (ury - lly)
            try:
                c.setPageSize((width, height))
            except Exception:
                pass

            annots = getattr(page, "Annots", None)
            if annots:
                for annot in annots:
                    try:
                        if getattr(annot, "Subtype", None) != PdfName.Widget:
                            continue
                        rect = annot.Rect
                        if not rect or len(rect) != 4:
                            continue
                        x0, y0, x1, y1 = [float(v) for v in rect]
                        h = max(1.0, y1 - y0)
                        val = None
                        if getattr(annot, "V", None) is not None:
                            raw = annot.V
                            try:
                                val = raw.to_unicode() if hasattr(raw, "to_unicode") else str(raw)
                            except Exception:
                                val = str(raw)
                        if val is None:
                            continue
                        val = str(val).strip().strip("()")
                        font_size = max(8, min(14, h * 0.6))
                        c.setFont("Helvetica", font_size)
                        y_baseline = y0 + max(2, (h - font_size) * 0.75)
                        c.drawString(x0 + 2, y_baseline, val)
                    except Exception:
                        continue
            c.showPage()
        c.save()

        # Merge overlay and remove form fields
        overlay_pdf = PdfReader(tmp_overlay_path)
        for i, page in enumerate(pages):
            try:
                if i < len(overlay_pdf.pages):
                    PageMerge(page).add(overlay_pdf.pages[i]).render()
                if hasattr(page, "Annots"):
                    try:
                        del page.Annots
                    except Exception:
                        page.Annots = None
            except Exception:
                continue

        try:
            if pdf.Root and getattr(pdf.Root, "AcroForm", None):
                try:
                    del pdf.Root.AcroForm
                except Exception:
                    pdf.Root.AcroForm = None
        except Exception:
            pass

        PdfWriter().write(output_pdf, pdf)
        try:
            os.remove(tmp_overlay_path)
        except Exception:
            pass
        return True
    except Exception as e:
        log(f"  [WARN] Flatten failed ({e}); uploading original (might appear blank).")
        return False

# ===================== Excel helpers =====================
def col_letter_to_index(letter: str) -> int:
    letter = (letter or "").strip().upper()
    if not letter:
        return None
    total = 0
    for ch in letter:
        if not ("A" <= ch <= "Z"):
            return None
        total = total * 26 + (ord(ch) - ord('A') + 1)
    return total - 1

# ===================== One-by-one Pipeline =====================
def run_pipeline(username, password, excel_path, sheet_name, idx_last, idx_first, idx_date,
                 pdf_template_path, log_callback, stop_flag_cb=None,
                 skip_if_date: bool = True, dry_run: bool = False, allow_loose_match: bool = True):
    def log(msg): log_callback(msg)

    # Lazy imports to avoid crashing GUI at import time
    try:
        from openpyxl import load_workbook
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.chrome.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.common.exceptions import TimeoutException
    except Exception as e:
        log(f"[FATAL] Missing package: {e}. Did you run pip install?")
        return

    # ---------- helpers ----------
    from selenium.common.exceptions import StaleElementReferenceException

    def stable_find(locator, desc="element", timeout=12, attempts=4):
        """
        Re-acquires a fresh element if the DOM re-rendered (avoids stale refs).
        Usage: el = stable_find((By.ID,"Login__UsernameField"), "username")
        """
        for i in range(attempts):
            try:
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located(locator))
                _ = el.tag_name  # touch to verify it's not stale
                try:
                    el = WebDriverWait(driver, 4).until(EC.element_to_be_clickable(locator))
                except Exception:
                    pass
                return el
            except StaleElementReferenceException:
                if i == attempts - 1:
                    raise
                time.sleep(0.15)
            except Exception:
                if i == attempts - 1:
                    raise
                time.sleep(0.15)
        return None

    def safe_return_to_patients(wait, driver):
        """Return to Patients list reliably."""
        try:
            link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Patients")))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", link)
            try:
                link.click()
            except Exception:
                driver.execute_script("arguments[0].click();", link)
            wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Name, Acct #, Phone, or Ins ID']")))
            time.sleep(0.4)
            return True
        except Exception:
            try:
                driver.get("https://www.therapynotes.com/app/patients/list")
                WebDriverWait(driver, 12).until(EC.presence_of_element_located(
                    (By.XPATH, "//input[@placeholder='Name, Acct #, Phone, or Ins ID']")))
                time.sleep(0.4)
                return True
            except Exception:
                return False

    def safe_click(wait, locator, driver, desc="element", timeout=12):
        try:
            el = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable(locator))
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            except Exception:
                pass
            try:
                el.click()
            except Exception:
                driver.execute_script("arguments[0].click();", el)
            return True
        except Exception:
            log(f"   - Could not click {desc}.")
            return False

    def open_documents_tab(wait, driver):
        locators = [
            (By.CSS_SELECTOR, "a[data-tab-id='Documents']"),
            (By.XPATH, "//a[@data-tab-id='Documents']"),
            (By.XPATH, "//a[normalize-space(text())='Documents']"),
        ]
        for by, sel in locators:
            if safe_click(wait, (by, sel), driver, desc="Documents tab"):
                try:
                    WebDriverWait(driver, 12).until(
                        EC.any_of(
                            EC.presence_of_element_located((By.XPATH, "//h2[contains(., 'Documents')]")),
                            EC.presence_of_element_located((By.XPATH, "//span[contains(@class,'documentNameSpan')]")),
                            EC.presence_of_element_located((By.XPATH, "//table[contains(., 'Document')]"))
                        )
                    )
                    time.sleep(0.4)
                    return True
                except Exception:
                    pass
        return False

    def open_upload_dialog(wait, driver, log):
        """
        Ensure we're on Documents tab and open the 'Upload Patient File' dialog.
        """
        if not open_documents_tab(wait, driver):
            log("   - Cannot open Documents tab.")
            return False

        locators = [
            (By.XPATH, "//button[.//span[contains(@class,'upload')] and contains(normalize-space(.), 'Upload Patient File')]") ,
            (By.XPATH, "//button[contains(normalize-space(.), 'Upload Patient File')]") ,
            (By.XPATH, "//button[contains(., 'Upload')]") ,
        ]
        for by, sel in locators:
            try:
                btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((by, sel)))
                try:
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                except Exception:
                    pass
                try:
                    btn.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.6)
                return True
            except Exception:
                continue
        log("   - Could not open 'Upload Patient File' dialog.")
        return False

    def select_patient_by_name(wait, driver, full_name, log, allow_loose=False, timeout=12):
        """
        Types into the global patient search and selects from the dropdown.
        - Exact (default): must exactly equal the name text (case/space/punct normalized)
        - Flexible (if allow_loose=True): accept candidates that contain all target tokens
          (first/last/etc.) in order, allowing middle names/initials and punctuation differences.
        Retries once if the search box isn't available.
        """
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support import expected_conditions as EC
        import time, re

        def norm(s):
            s = (s or "").lower()
            s = re.sub(r"[^a-z0-9 ]+", " ", s)       # remove punctuation
            s = re.sub(r"\s+", " ", s).strip()       # collapse spaces
            return s

        def tokens_in_order(candidate_norm, target_tokens):
            """All target tokens must appear in order in the candidate string."""
            pos = 0
            for tok in target_tokens:
                i = candidate_norm.find(tok, pos)
                if i == -1:
                    return False
                pos = i + len(tok)
            return True

        target_norm = norm(full_name)
        target_tokens = [t for t in target_norm.split(" ") if t]

        # inner routine so we can retry once if the box isn't ready
        def try_once():
            sb = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//input[@placeholder='Name, Acct #, Phone, or Ins ID']")))
            sb.clear()

            # Normalize the outgoing query (strip dots like "T." → "T")
            search_query = re.sub(r'\s*\.\s*', ' ', full_name).strip()
            search_query = re.sub(r'\s+', ' ', search_query)

            # Build a "no middle-initial" variant and matching targets
            # (drop single-letter tokens from the *target* too so matching can succeed)
            def strip_single_letter_tokens(tokens):
                return [t for t in tokens if len(t) > 1]

            tokens_nomiddle = strip_single_letter_tokens(target_tokens)
            target_norm_nomiddle = " ".join(tokens_nomiddle)

            # Remove any standalone single-letter between words (e.g., "John T Patterson" → "John Patterson")
            search_query_nomiddle = re.sub(r'(^|\s)[A-Za-z](?=\s)', ' ', search_query).strip()
            search_query_nomiddle = re.sub(r'\s+', ' ', search_query_nomiddle)

            def evaluate_candidates(cands, target_norm_current, tokens_current):
                # 1) strict exact (normalized equality)
                for rdiv in cands:
                    try:
                        cand_text = norm(rdiv.text)
                        if cand_text == target_norm_current:
                            try:
                                rdiv.click()
                            except Exception:
                                driver.execute_script("arguments[0].click();", rdiv)
                            time.sleep(0.6)
                            return True
                    except Exception:
                        continue

                # 2) flexible: tokens in order (ignores middle initials/punct)
                if allow_loose:
                    for rdiv in cands:
                        try:
                            cand_text = norm(rdiv.text)
                            if tokens_in_order(cand_text, tokens_current):
                                try:
                                    rdiv.click()
                                except Exception:
                                    driver.execute_script("arguments[0].click();", rdiv)
                                time.sleep(0.6)
                                return True
                        except Exception:
                            continue

                # 3) last resort: startswith or contains
                for rdiv in cands:
                    try:
                        cand_text = norm(rdiv.text)
                        if cand_text.startswith(target_norm_current) or target_norm_current in cand_text:
                            try:
                                rdiv.click()
                            except Exception:
                                driver.execute_script("arguments[0].click();", rdiv)
                            time.sleep(0.6)
                            return True
                    except Exception:
                        continue

                return False

            # ---- First attempt: as typed ----
            sb.send_keys(search_query)
            time.sleep(1.0)  # allow dropdown to populate
            try:
                wait.until(EC.visibility_of_element_located((By.ID, "ContentBubbleResultsContainer")))
                time.sleep(0.3)
            except Exception:
                pass
            candidates = driver.find_elements(By.CSS_SELECTOR, "#ContentBubbleResultsContainer > div")

            if evaluate_candidates(candidates, target_norm, target_tokens):
                return True

            # ---- Second attempt: retry without a middle initial ----
            if search_query_nomiddle != search_query:
                log(f"   - Retrying dropdown search without middle initial: '{search_query_nomiddle}'")
                sb.clear()
                sb.send_keys(search_query_nomiddle)
                time.sleep(1.0)
                try:
                    wait.until(EC.visibility_of_element_located((By.ID, "ContentBubbleResultsContainer")))
                    time.sleep(0.3)
                except Exception:
                    pass
                candidates = driver.find_elements(By.CSS_SELECTOR, "#ContentBubbleResultsContainer > div")

                if evaluate_candidates(candidates, target_norm_nomiddle, tokens_nomiddle):
                    return True

            return False

        # Try once; if we can't even get the search box, return False (with a retry)
        try:
            if try_once():
                return True
            log(f"   - No {'flexible ' if allow_loose else ''}dropdown match for '{full_name}'.")
            return False
        except Exception:
            # Retry once by forcing Patients page, then try again
            log("   - Patient search box not available. Retrying…")
            try:
                driver.get("https://www.therapynotes.com/app/patients/list")
                WebDriverWait(driver, 12).until(EC.presence_of_element_located(
                    (By.XPATH, "//input[@placeholder='Name, Acct #, Phone, or Ins ID']")))
                time.sleep(0.4)
                if try_once():
                    return True
            except Exception:
                pass
            log("   - Patient search still not available.")
            return False


    def wait_for_upload_success(wait, driver, doc_name_text, timeout=25):
        """
        After clicking 'Add Document', consider the upload successful if either:
          - the upload dialog disappears AND a row containing the new document name appears
          - we see a success toast/snackbar that implies it was added
        """
        from selenium.webdriver.common.by import By

        end = time.time() + timeout
        doc_row_found = False
        dialog_gone = False
        toast_seen = False
        doc_name_lower = (doc_name_text or "").strip().lower()

        while time.time() < end:
            # 1) Is the dialog gone?
            try:
                driver.find_element(By.ID, "InputUploader")
                dialog_gone = False
            except Exception:
                dialog_gone = True

            # 2) Document row present?
            try:
                rows = driver.find_elements(
                    By.XPATH,
                    "//tr[.//span[contains(@class,'documentNameSpan')]]"
                )
                for row in rows:
                    try:
                        name_el = row.find_element(By.XPATH, ".//span[contains(@class,'documentNameSpan')]")
                        if doc_name_lower in name_el.text.strip().lower():
                            doc_row_found = True
                            break
                    except Exception:
                        continue
            except Exception:
                pass

            # 3) Toast/snackbar?
            try:
                toast = driver.find_elements(
                    By.XPATH,
                    "//*[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'document') "
                    "and contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'added')]"
                )
                if toast:
                    toast_seen = True
            except Exception:
                pass

            if (dialog_gone and doc_row_found) or toast_seen:
                return True

            time.sleep(0.5)

        return False

    # ---------- start ----------
    driver = None
    try:
        if not os.path.isfile(excel_path) or not excel_path.lower().endswith(".xlsx"):
            log("Please select a valid .xlsx Excel file.")
            return
        if not os.path.isfile(pdf_template_path):
            log("PDF template not found.")
            return

        log(f"Loading Excel (writeable): {excel_path}")
        wb = load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            log(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            return
        ws = wb[sheet_name]

        # status col index from letter
        status_col_idx = col_letter_to_index(STATUS_COL_LETTER)  # 0-based

        base = os.path.splitext(os.path.basename(excel_path))[0]
        out_dir = os.path.dirname(excel_path)
        updated_xlsx = os.path.join(out_dir, f"{base}_UPDATED.xlsx")

        log("Starting Chrome… (fast mode)")
        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--blink-settings=imagesEnabled=false")
        options.add_argument("--disable-features=Translate,CalculateNativeWinOcclusion")
        options.page_load_strategy = 'eager'  # don't wait for images/ads/etc

        prefs = {
            "profile.managed_default_content_settings.images": 2,  # block images
            "profile.default_content_setting_values.notifications": 2,
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False,
        }
        options.add_experimental_option("prefs", prefs)

        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(15)
        wait = WebDriverWait(driver, 8, poll_frequency=0.2)

        login_url = "https://www.therapynotes.com/app/login/IntegritySWS/"
        driver.get(login_url)
        log("Opening login page…")

        # Extra: wait a beat for the first paint, then use stale-proof lookups
        try:
            WebDriverWait(driver, 6).until(lambda d: d.execute_script("return document.readyState") in ("interactive","complete"))
        except Exception:
            pass

        try:
            u = stable_find((By.ID, "Login__UsernameField"), "username")
            try: u.clear()
            except Exception: pass
            u.send_keys(username)

            p = stable_find((By.ID, "Login__Password"), "password")
            try: p.clear()
            except Exception: pass
            p.send_keys(password)

            b = stable_find((By.ID, "Login__LogInButton"), "login button")
            try:
                b.click()
            except StaleElementReferenceException:
                b = stable_find((By.ID, "Login__LogInButton"), "login button")
                b.click()
        except Exception as e:
            log(f"[FATAL] {e}")
            return

        # Post-submit: wait for something stable on the landing page
        try:
            wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Patients")))
        except Exception:
            # One more try: the page might still be settling
            try:
                time.sleep(0.3)
                wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Patients")))
            except Exception:
                pass
        log("Login successful.")

        # Initial reset to Patients
        safe_return_to_patients(wait, driver)
        pdf_out_dir = ensure_output_dir()
        log(f"PDF output folder: {pdf_out_dir}")

        # Rows (write-access)
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            log("No rows found.")
            return

        # Assume row 1 is header if it looks like one
        start_idx = 2
        try:
            header_tokens = ("first", "last", "date", "client", "date from tn ia")
            def looks_like_header(v):
                return any(tok in str(v).strip().lower() for tok in header_tokens)
            if not (looks_like_header(rows[0][idx_first].value) or
                    looks_like_header(rows[0][idx_last].value) or
                    looks_like_header(rows[0][idx_date].value)):
                start_idx = 1
        except Exception:
            start_idx = 2

        processed = uploaded = skipped = 0

        def write_status(row_idx_one_based, text):
            try:
                if status_col_idx is not None:
                    ws.cell(row=row_idx_one_based, column=status_col_idx + 1).value = text
                    wb.save(updated_xlsx)
            except Exception as e:
                log(f"   - [WARN] Could not write status to Excel: {e}")

        for ridx in range(start_idx, len(rows) + 1):
            if stop_flag_cb and stop_flag_cb():
                log("Stop requested… finishing current client and exiting.")
                break

            try:
                row = ws[ridx]
                cell_last = row[idx_last]
                cell_first = row[idx_first]
                cell_date = row[idx_date]

                first_name = str((cell_first.value or "")).strip()
                last_name  = str((cell_last.value or "")).strip()
                if not first_name and not last_name:
                    continue

                full_name = f"{first_name} {last_name}".strip()
                log(f"— Processing: {full_name}")

                # Compute date in cell once
                date_str = to_mmddyyyy(cell_date.value).strip()

                # OPTION: Skip if Date already present (from GUI checkbox)
                if skip_if_date and date_str:
                    msg = f"Skipped — date present ({date_str})"
                    log(f"   - {msg}")
                    skipped += 1
                    write_status(ridx, msg)
                    safe_return_to_patients(wait, driver)
                    continue

                # 1) Ensure we have a Date in Excel; if not, find it on Documents tab
                if not date_str:
                    # Search for patient (EXACT MATCH)
                    if not select_patient_by_name(wait, driver, full_name, log, allow_loose_match):
                        skipped += 1
                        write_status(ridx, "Skipped — no exact patient match")
                        safe_return_to_patients(wait, driver)
                        continue

                    # Go to Documents
                    if not open_documents_tab(wait, driver):
                        log("   - Cannot open Documents tab. Skipping.")
                        skipped += 1
                        write_status(ridx, "Skipped — cannot open Documents tab")
                        safe_return_to_patients(wait, driver)
                        continue

                    # Scan for Intake Note date
                    try:
                        wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "documentNameSpan")))
                        doc_elems = driver.find_elements(By.CLASS_NAME, "documentNameSpan")
                        found_date = None
                        for elem in doc_elems:
                            name = (elem.text or "").strip()
                            if "intake note" in name.lower():
                                parent_row = elem.find_element(By.XPATH, "./ancestor::tr")
                                date_cells = parent_row.find_elements(
                                    By.XPATH, ".//td[contains(@class, 'v-align-top') and contains(@style, 'padding-bottom: 6px')]"
                                )
                                for ccc in date_cells:
                                    t = (ccc.text or "").strip()
                                    if t and any(ch.isdigit() for ch in t) and ('/' in t or '-' in t):
                                        found_date = to_mmddyyyy(t)
                                        break
                                if found_date:
                                    break

                        if found_date:
                            cell_date.value = found_date
                            date_str = found_date
                            wb.save(updated_xlsx)
                            log(f"   ✓ Date found & saved: {found_date}")
                        else:
                            log("   - No Intake Note date; skipping client.")
                            skipped += 1
                            write_status(ridx, "Skipped — no Intake Note date")
                            safe_return_to_patients(wait, driver)
                            continue
                    except Exception as e:
                        log(f"   - Error scanning documents: {e}; skipping.")
                        skipped += 1
                        write_status(ridx, "Skipped — error scanning documents")
                        safe_return_to_patients(wait, driver)
                        continue

                # 2) Create/confirm PDF FIRST (required before uploading)
                out_pdf_name = sanitize_filename(make_pdf_name(first_name, last_name))
                out_pdf_path = os.path.join(pdf_out_dir, out_pdf_name)
                field_values = {"Client name print": full_name, "Date": date_str}

                if not os.path.isfile(out_pdf_path):
                    log(f"   Creating PDF: {out_pdf_name}")
                    ok = fill_pdf_locked(pdf_template_path, out_pdf_path, field_values, log)
                    if not ok:
                        log("   - Failed to generate PDF; skipping client.")
                        skipped += 1
                        write_status(ridx, "Skipped — failed to generate PDF")
                        safe_return_to_patients(wait, driver)
                        continue
                    log(f"   ✓ PDF saved: {out_pdf_path}")
                else:
                    log(f"   (Using existing PDF): {out_pdf_path}")

                # 3) Flatten it for reliable rendering in TherapyNotes
                fd, tmp_flat = tempfile.mkstemp(suffix=".pdf"); os.close(fd)
                flattened_ok = flatten_pdf_preserving_values(out_pdf_path, tmp_flat, log)
                pdf_to_upload = tmp_flat if flattened_ok else out_pdf_path

                processed += 1

                if dry_run:
                    # Simulate upload but do not actually upload
                    msg = f"DRY RUN — would upload '{first_name} {last_name} Consent NPP' (Date: {date_str})"
                    log("   " + msg)
                    write_status(ridx, msg)
                    try: os.remove(tmp_flat)
                    except Exception: pass
                    safe_return_to_patients(wait, driver)
                    continue

                # 4) Return to Patients to guarantee the global search box is available
                if not safe_return_to_patients(wait, driver):
                    try:
                        driver.get("https://www.therapynotes.com/app/patients/list")
                        WebDriverWait(driver, 12).until(EC.presence_of_element_located(
                            (By.XPATH, "//input[@placeholder='Name, Acct #, Phone, or Ins ID']")))
                    except Exception:
                        log("   - Could not reset to Patients list. Skipping.")
                        skipped += 1
                        write_status(ridx, "Skipped — could not reset to Patients list")
                        try: os.remove(tmp_flat)
                        except Exception: pass
                        continue

                # 5) Search & open patient (upload step) EXACT MATCH
                if not select_patient_by_name(wait, driver, full_name, log, allow_loose_match):

                    skipped += 1
                    write_status(ridx, "Skipped — no exact patient match (upload step)")
                    try: os.remove(tmp_flat)
                    except Exception: pass
                    continue

                # 6) Open Documents and open 'Upload Patient File' dialog
                if not open_upload_dialog(wait, driver, log):
                    skipped += 1
                    write_status(ridx, "Skipped — cannot open Upload dialog")
                    try: os.remove(tmp_flat)
                    except Exception: pass
                    continue

                # 7) Fill dialog fields: Date, File, Document Name
                try:
                    din = WebDriverWait(driver, 12).until(EC.element_to_be_clickable((By.ID, "PatientFile__Date")))
                    din.clear(); din.send_keys(date_str); time.sleep(0.2)
                except TimeoutException:
                    log("   - Date input not found in dialog. Skipping.")
                    skipped += 1
                    write_status(ridx, "Skipped — Date input not found in dialog")
                    try: os.remove(tmp_flat)
                    except Exception: pass
                    continue

                try:
                    fin = WebDriverWait(driver, 12).until(EC.presence_of_element_located((By.ID, "InputUploader")))
                    fin.send_keys(os.path.abspath(pdf_to_upload))
                    time.sleep(0.6)
                except TimeoutException:
                    log("   - File input (choose file) not found. Skipping.")
                    skipped += 1
                    write_status(ridx, "Skipped — File input not found")
                    try: os.remove(tmp_flat)
                    except Exception: pass
                    continue

                try:
                    dn = WebDriverWait(driver, 12).until(EC.element_to_be_clickable((By.ID, "PatientFile__DocumentName")))
                    dn.clear(); dn.send_keys(f"{first_name} {last_name} Consent NPP"); time.sleep(0.2)
                except TimeoutException:
                    log("   - Document Name input not found. Skipping.")
                    skipped += 1
                    write_status(ridx, "Skipped — Document Name input not found")
                    try: os.remove(tmp_flat)
                    except Exception: pass
                    continue

                # Optional: click "Patient" label to move focus
                try:
                    lbl = WebDriverWait(driver, 6).until(
                        EC.element_to_be_clickable((By.XPATH, "//label[normalize-space(text())='Patient']"))
                    )
                    try: lbl.click()
                    except Exception: driver.execute_script("arguments[0].click();", lbl)
                    time.sleep(0.2)
                except Exception:
                    pass

                # 8) Add Document
                try:
                    add_btn = WebDriverWait(driver, 12).until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@type='button' and @value='Add Document']"))
                    )
                    try:
                        add_btn.click()
                    except Exception:
                        driver.execute_script("arguments[0].click();", add_btn)
                except TimeoutException:
                    log("   - Couldn’t click 'Add Document'. Skipping.")
                    skipped += 1
                    write_status(ridx, "Skipped — could not click Add Document")
                    try: os.remove(tmp_flat)
                    except Exception: pass
                    continue

                # Confirm success: dialog closes and/or row appears with the exact doc name
                doc_name_text = f"{first_name} {last_name} Consent NPP"
                if wait_for_upload_success(wait, driver, doc_name_text, timeout=25):
                    uploaded += 1
                    msg = f"Uploaded on {date_str}"
                    log(f"   ✓ {msg}: {doc_name_text}")
                    write_status(ridx, msg)
                else:
                    log("   - Upload may not have completed (no confirmation found).")
                    write_status(ridx, "Upload uncertain — no confirmation found")

                # Cleanup temp
                try: os.remove(tmp_flat)
                except Exception: pass

                # Back to Patients for next loop (guaranteed)
                if not safe_return_to_patients(wait, driver):
                    try:
                        driver.get("https://www.therapynotes.com/app/patients/list")
                        WebDriverWait(driver, 12).until(
                            EC.presence_of_element_located(
                                (By.XPATH, "//input[@placeholder='Name, Acct #, Phone, or Ins ID']")
                            )
                        )
                    except Exception:
                        pass

            except Exception as e:
                log(f"   [ERROR] Unexpected error on {full_name}: {e}")
                skipped += 1
                write_status(ridx, "Error — see log")
                # Always try to get back to Patients
                safe_return_to_patients(wait, driver)

        # === end of for ridx loop ===
        log(f"Done. Processed: {processed}, Uploaded: {uploaded}, Skipped: {skipped}")
        log(f"Updated Excel saved at: {updated_xlsx}")

    except Exception as e:
        log(f"[FATAL] {e}")
    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass

def run_welcome_login_only(pn_url, pn_user, pn_pass, log_callback):
    def log(msg): log_callback(msg)

    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.common.keys import Keys
        from webdriver_manager.chrome import ChromeDriverManager
        import time
    except Exception as e:
        log(f"[FATAL][PN] Missing package: {e}. Did you run pip install selenium webdriver-manager?")
        return

    # -------- helpers --------
    def norm(s):
        return " ".join((s or "").split()).strip()

    def visible_text(el):
        txt = norm(el.text)
        if not txt:
            for attr in ("value", "aria-label", "title", "alt"):
                v = el.get_attribute(attr)
                if v and norm(v):
                    txt = norm(v); break
        return txt

    def dump_clickables(ctx, where):
        # <a>, <button>, role=button, anything with onclick
        items = []
        try:
            nodes = ctx.find_elements(By.XPATH, "//*[self::a or self::button or @role='button' or @onclick]")
        except Exception:
            nodes = []
        for i, el in enumerate(nodes):
            try:
                txt = visible_text(el)
                if not txt:
                    continue
                items.append(txt)
            except Exception:
                continue
        if items:
            log(f"[PN] {where} — {len(items)} clickable item(s):")
            # dedupe while keeping order, print the first ~200
            seen = set()
            count = 0
            for t in items:
                tl = t.lower()
                if tl in seen:
                    continue
                seen.add(tl)
                log(f"[PN][{where}] Clickable: {t}")
                count += 1
                if count >= 200:
                    break
        else:
            log(f"[PN] {where} — no clickables found.")

    # -------- main flow --------
    driver = None
    try:
        log("Welcome Bot — opening Penelope…")
        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        wait = WebDriverWait(driver, 20)

        # --- helper: find and click a "Workflow(s)" button/link across frames ---
        def _click_workflows_button():
            targets = []

            # search strategies we will try in each context
            def find_candidates(ctx):
                found = []
                # any element whose visible text contains "workflow"
                try:
                    found += ctx.find_elements(By.XPATH, "//*[contains(translate(normalize-space(text()),'WORKFLO','workflo'),'workflow')]")
                except Exception:
                    pass
                # anything with title/aria-label containing workflow
                xp = "//*[@title[contains(translate(.,'WORKFLO','workflo'),'workflow')] or contains(translate(@aria-label,'WORKFLO','workflo'),'workflow')]"
                try:
                    found += ctx.find_elements(By.XPATH, xp)
                except Exception:
                    pass
                # ids/names that look like workflow
                try:
                    found += ctx.find_elements(By.XPATH, "//*[@id[contains(translate(.,'WORKFLO','workflo'),'workflow')] or @name[contains(translate(.,'WORKFLO','workflo'),'workflow')]]")
                except Exception:
                    pass
                # buttons/links with text that might be the right one
                try:
                    found += ctx.find_elements(By.XPATH, "//a[contains(translate(normalize-space(.),'WORKFLO','workflo'),'workflow')]|//button[contains(translate(normalize-space(.),'WORKFLO','workflo'),'workflow')]")
                except Exception:
                    pass
                # de-dup
                uniq = []
                seen = set()
                for el in found:
                    try:
                        key = (el.tag_name, el.get_attribute("id"), el.get_attribute("name"), el.get_attribute("class"))
                    except Exception:
                        key = (id(el),)
                    if key in seen:
                        continue
                    seen.add(key)
                    uniq.append(el)
                return uniq

            # 1) try in main document
            targets = find_candidates(driver)
            if not targets:
                # 2) search each iframe
                frames = driver.find_elements(By.TAG_NAME, "iframe")
                for fr in frames:
                    try:
                        driver.switch_to.default_content()
                        driver.switch_to.frame(fr)
                        cand = find_candidates(driver)
                        if cand:
                            targets = cand
                            break
                    except Exception:
                        continue
                driver.switch_to.default_content()

            if not targets:
                log("[PN] Could not find a 'Workflows' control by text/label/id. Tell me the exact label on that button if you can.")
                return False

            # Prefer visible/clickable
            btn = None
            for el in targets:
                try:
                    if el.is_displayed() and el.is_enabled():
                        btn = el
                        break
                except Exception:
                    continue

            if not btn:
                log("[PN] Found potential 'Workflow' elements, but none look clickable.")
                return False

            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            except Exception:
                pass
            try:
                btn.click()
            except Exception:
                try:
                    driver.execute_script("arguments[0].click();", btn)
                except Exception:
                    log("[PN] Failed to click the 'Workflows' control.")
                    return False

            log("[PN] Clicked the 'Workflows' control.")
            return True

        driver.get(pn_url or PN_DEFAULT_URL)
        log(f"[PN] Opened: {driver.current_url}")
        log(f"[PN] Page title: {driver.title or '(no title)'}")

        # --- Try to log in on MAIN first
        def try_fill_in_context(ctx):
            # collect inputs
            try:
                inputs = ctx.find_elements(By.TAG_NAME, "input")
            except Exception:
                inputs = []
            if not inputs:
                return False

            # choose likely username + password fields
            user_el = None; pwd_el = None; text_like = []
            for el in inputs:
                t = (el.get_attribute("type") or "").lower()
                if t == "password" and pwd_el is None:
                    pwd_el = el
                elif t in ("text", "email"):
                    text_like.append(el)

            def score_username(el):
                s = 1
                idv   = (el.get_attribute("id") or "").lower()
                namev = (el.get_attribute("name") or "").lower()
                ph    = (el.get_attribute("placeholder") or "").lower()
                ar    = (el.get_attribute("aria-label") or "").lower()
                for token in ("user", "email", "login", "uname"):
                    if token in idv: s += 3
                    if token in namev: s += 3
                    if token in ph: s += 2
                    if token in ar: s += 2
                return s

            if text_like:
                text_like.sort(key=score_username, reverse=True)
                user_el = text_like[0]

            if not (user_el and pwd_el):
                return False

            # type credentials
            try:
                try: user_el.clear()
                except Exception: pass
                user_el.send_keys(pn_user or "")
                try: pwd_el.clear()
                except Exception: pass
                pwd_el.send_keys(pn_pass or "")
            except Exception:
                return False

            # click submit or press ENTER
            candidates = [
                (By.XPATH, "//button[@type='submit']"),
                (By.XPATH, "//input[@type='submit']"),
                (By.XPATH, "//*[self::button or self::input][contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in') or contains(translate(@value,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in') or contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'login') or contains(translate(@value,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'login') or contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in') or contains(translate(@value,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]"),
            ]
            for by, sel in candidates:
                try:
                    btn = WebDriverWait(ctx, 3).until(EC.element_to_be_clickable((by, sel)))
                    try:
                        btn.click()
                    except Exception:
                        try: ctx.execute_script("arguments[0].click();", btn)
                        except Exception: pass
                    return True
                except Exception:
                    continue
            # fallback
            try:
                pwd_el.send_keys(Keys.ENTER)
                return True
            except Exception:
                return False

        # Try MAIN
        logged = try_fill_in_context(driver)
        if logged:
            log("[PN] Submitted login (MAIN).")
        else:
            # Try each IFRAME
            frames = driver.find_elements(By.TAG_NAME, "iframe")
            log(f"[PN] Found {len(frames)} iframe(s).")
            for idx, fr in enumerate(frames):
                try:
                    driver.switch_to.frame(fr)
                    if try_fill_in_context(driver):
                        log(f"[PN] Submitted login (IFRAME #{idx}).")
                        driver.switch_to.default_content()
                        logged = True
                        break
                    driver.switch_to.default_content()
                except Exception:
                    driver.switch_to.default_content()
            if not logged:
                log("[PN] Could not find login fields. Please share the login page HTML details.")
                return

        # Wait for some post-login signal
        try:
            wait.until(
                EC.any_of(
                    EC.url_changes(pn_url or PN_DEFAULT_URL),
                    EC.presence_of_element_located((By.XPATH, "//*[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'logout')]")),
                    EC.presence_of_element_located((By.XPATH, "//*[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'home')]"))
                )
            )
            log("[PN] Login appears successful (post-login signal detected).")
        except Exception:
            log("[PN] No clear post-login signal yet; proceeding to dump UI anyway.")

        # 2) open the Workflows panel
        if not _click_workflows_button():
            return


        # ==== NEW: Dump clickables on MAIN and inside each iframe ====
        dump_clickables(driver, "MAIN")
        frames = driver.find_elements(By.TAG_NAME, "iframe")
        log(f"[PN] Post-login iframe count: {len(frames)}")
        for idx, fr in enumerate(frames):
            fid = fr.get_attribute("id") or ""
            fname = fr.get_attribute("name") or ""
            fsrc = fr.get_attribute("src") or ""
            log(f"[PN] IFRAME #{idx}: id={fid!r} name={fname!r} src={fsrc!r}")
            try:
                driver.switch_to.frame(fr)
                dump_clickables(driver, f"IFRAME #{idx}")
                driver.switch_to.default_content()
            except Exception:
                try: driver.switch_to.default_content()
                except Exception: pass

        # Keep open briefly so you can see the page
        time.sleep(5)

    except Exception as e:
        log(f"[PN][ERROR] {e}")
    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass

def switch_to_content_frame(driver, log=None):
    # local imports so helpers have what they need
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    try:
        driver.switch_to.default_content()
    except Exception:
        pass

    candidates = [
        (By.NAME, "frm_content"),
        (By.ID, "frm_content_id"),
        (By.XPATH, "//iframe[contains(@src,'homepage') or contains(@src,'content')]"),
    ]
    for by, sel in candidates:
        try:
            WebDriverWait(driver, 12).until(EC.frame_to_be_available_and_switch_to_it((by, sel)))
            if log: log("[PN] Switched into content frame.")
            return True
        except Exception:
            pass

    # Fallback: choose largest iframe
    try:
        frames = driver.find_elements(By.TAG_NAME, "iframe")
    except Exception:
        frames = []
    largest, area = None, 0
    for fr in frames:
        try:
            size = fr.size or {}
            w = size.get("width", 0) or 0
            h = size.get("height", 0) or 0
            if (w * h) > area:
                area = w * h
                largest = fr
        except Exception:
            continue
    if largest:
        try:
            driver.switch_to.frame(largest)
            if log: log("[PN] Switched into largest iframe.")
            return True
        except Exception:
            pass

    try:
        driver.switch_to.default_content()
    except Exception:
        pass
    if log: log("[PN] Could not switch to content frame (will search all frames anyway).")
    return False
def open_workflows_flyout(driver, log=lambda m: None):
    """
    Click the right-edge 'Workflow' (singular) vertical tab. We try MAIN first
    (default content), then the content frame, then every iframe. We specifically
    target id=#taskHandle because your debug shows that's the real control.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver import ActionChains
    import time

    def click_in_context(ctx, label="CTX"):
        # 1) Exact element seen in your debug: id='taskHandle'
        try:
            el = ctx.find_element(By.ID, "taskHandle")
            try:
                ctx.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            except Exception:
                pass
            try:
                el.click()
            except Exception:
                try:
                    ctx.execute_script("arguments[0].click();", el)
                except Exception:
                    try:
                        ActionChains(ctx).move_to_element(el).pause(0.1).click(el).perform()
                    except Exception:
                        el = None
            if el:
                log(f"[PN] Clicked 'Workflow' handle via {label} (id=#taskHandle).")
                time.sleep(0.5)
                return True
        except Exception:
            pass

        # 2) Fallback locators containing 'workflow'
        locators = [
            (By.CSS_SELECTOR, "div.drawerHandle[aria*='workflow']"),
            (By.XPATH, "//*[@id='handleContainer']//*[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'workflow')]"),
            (By.XPATH, "//*[contains(@class,'drawerHandle')][contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'workflow')]"),
            (By.XPATH, "//*[@aria-label][contains(translate(@aria-label,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'workflow')]"),
            (By.XPATH, "//*[@title][contains(translate(@title,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'workflow')]"),
            (By.XPATH, "//*[self::a or self::button or self::span or self::div]"
                       "[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'workflow')]"),
        ]
        for by, sel in locators:
            try:
                el = WebDriverWait(ctx, 6).until(EC.element_to_be_clickable((by, sel)))
                try:
                    el.click()
                except Exception:
                    try:
                        ctx.execute_script("arguments[0].click();", el)
                    except Exception:
                        try:
                            ActionChains(ctx).move_to_element(el).pause(0.1).click(el).perform()
                        except Exception:
                            continue
                log(f"[PN] Clicked 'Workflow' via {label}.")
                time.sleep(0.5)
                return True
            except Exception:
                continue
        return False

    # Try MAIN (default content) FIRST — your tab lives here
    try:
        driver.switch_to.default_content()
    except Exception:
        pass
    try:
        driver.execute_script("window.scrollTo(document.body.scrollWidth, 0);")
    except Exception:
        pass
    if click_in_context(driver, "MAIN"):
        return True

    # Then try the 'content' frame
    try:
        if switch_to_content_frame(driver, log):
            if click_in_context(driver, "CONTENT FRAME"):
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                return True
    except Exception:
        pass

    # Finally try every iframe
    try:
        frames = driver.find_elements(By.TAG_NAME, "iframe")
    except Exception:
        frames = []
    for idx, fr in enumerate(frames):
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame(fr)
            if click_in_context(driver, f"IFRAME #{idx}"):
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                return True
        except Exception:
            continue

    try:
        driver.switch_to.default_content()
    except Exception:
        pass
    log("[PN] Could not click the 'Workflow' handle even after targeting #taskHandle.")
    return False
    def pn_search_and_open_first_welcome(driver, log, start_mmddyyyy, end_mmddyyyy, timeout=20):
        """
        Uses stable selectors from your recording to:
          - Open the Workflow drawer
          - Click 'Advanced'
          - Set the date range (prefer inputs; fallback to calendar icons)
          - Click 'Search'
          - Open the first 'Send Welcome Letter and NPP' result
        Returns True if we clicked into a task page (same tab or new tab), else False.
        """
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import time, re

        wait = WebDriverWait(driver, timeout)

        def _switch_default():
            try: driver.switch_to.default_content()
            except Exception: pass

        def _maybe_click(by, sel, safe=True):
            try:
                el = wait.until(EC.element_to_be_clickable((by, sel)))
                try: driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                except Exception: pass
                try:
                    el.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", el)
                return True
            except Exception:
                return False

        def _maybe_switch_to_new_tab():
            try:
                hs = driver.window_handles
                if len(hs) >= 2:
                    driver.switch_to.window(hs[-1])
                    log("[PN] Switched to newly opened tab.")
                    return True
            except Exception:
                pass
            return False

        def _fill_dates_prefer_inputs():
            """
            Try to find two date inputs in the Advanced filter area (#dueWithinDays),
            set mm/dd/yyyy directly and fire change events.
            """
            try:
                _switch_default()
                wrap = wait.until(EC.presence_of_element_located((By.ID, "dueWithinDays")))
                # common patterns: text inputs that take mm/dd/yyyy or <input type=date>
                inputs = wrap.find_elements(By.CSS_SELECTOR, "input[type='text'],input[type='date']")
                inputs = [i for i in inputs if i.is_displayed()]
                # grab first two visible
                if len(inputs) >= 2:
                    for el, val in ((inputs[0], start_mmddyyyy), (inputs[1], end_mmddyyyy)):
                        try: el.clear()
                        except Exception: pass
                        el.send_keys(val)
                        try:
                            driver.execute_script(
                                "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", el
                            )
                        except Exception:
                            pass
                    return True
            except Exception:
                pass
            return False

        def _pick_day_via_calendar(icon_css, mmddyyyy):
            """Fallback: click the calendar icon then choose the day-of-month link."""
            try:
                day = str(int(mmddyyyy.split('/')[1]))  # mm/dd/yyyy → dd
            except Exception:
                day = None
            if not day:
                return False
            _switch_default()
            if not _maybe_click(By.CSS_SELECTOR, icon_css):
                return False
            try:
                # click <a> whose visible text is the target day
                day_el = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, f"//table//a[normalize-space()='{day}']")
                ))
                try:
                    day_el.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", day_el)
                return True
            except Exception:
                return False

        # 1) Open the Workflow drawer (id from recording)
        _switch_default()
        if _maybe_click(By.ID, "taskHandle"):
            log("[PN] Clicked Workflow handle (#taskHandle).")
        else:
            log("[PN] Could not click Workflow handle; maybe already open — continuing.")

        # 2) Click 'Advanced' (id from recording)
        _switch_default()
        if _maybe_click(By.ID, "taskAdvancedView"):
            log("[PN] Clicked Advanced (#taskAdvancedView).")
        else:
            log("[PN] Could not click Advanced; continuing anyway.")

        # 3) Date range: prefer direct inputs; fallback to your two calendar icons from the recording
        if _fill_dates_prefer_inputs():
            log(f"[PN] Entered date range: {start_mmddyyyy} .. {end_mmddyyyy} (inputs).")
        else:
            # recorded icons under #dueWithinDays
            start_icon_css = "#dueWithinDays > span:nth-of-type(1) > span:nth-of-type(2) > span:nth-of-type(1) > img:nth-of-type(1)"
            end_icon_css   = "#dueWithinDays > span:nth-of-type(1) > span:nth-of-type(3) > span:nth-of-type(1) > img:nth-of-type(1)"
            ok1 = _pick_day_via_calendar(start_icon_css, start_mmddyyyy)
            ok2 = _pick_day_via_calendar(end_icon_css, end_mmddyyyy)
            if ok1 and ok2:
                log(f"[PN] Entered date range: {start_mmddyyyy} .. {end_mmddyyyy} (calendar).")
            else:
                log("[PN] Could not set both dates; continuing to Search anyway.")

        # 4) Click Search (id from recording)
        _switch_default()
        if not _maybe_click(By.ID, "advancedTaskSearchBtn"):
            log("[PN] Could not click Search (#advancedTaskSearchBtn).")
            return False
        log("[PN] Clicked Search.")

        # 5) Open first 'Send Welcome Letter and NPP' result
        _switch_default()
        try:
            link = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//a[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),"
                           " 'send welcome letter and npp')]")
            ))
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", link)
            except Exception:
                pass
            try:
                link.click()
            except Exception:
                driver.execute_script("arguments[0].click();", link)
            # follow new tab if it opened one
            _maybe_switch_to_new_tab()
            log("[PN] Opened a 'Send Welcome Letter and NPP' task.")
            return True
        except Exception:
            log("[PN] No 'Send Welcome Letter and NPP' result link found after search.")
            return False

def debug_log_all_workflowish_things(driver, log):
    """List anything that *looks* like a Workflow control across doc + iframes."""
    from selenium.webdriver.common.by import By

    def dump_ctx(ctx, label):
        try:
            nodes = ctx.find_elements(By.XPATH,
                "//*[@" 
                "aria-label or @title or @alt or @data-tooltip or contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'workflow')"
                "]"
            )
        except Exception:
            nodes = []
        count = 0
        log(f"[PN][DEBUG] Scanning {label}: {len(nodes)} potential 'workflow' nodes found.")
        for el in nodes:
            try:
                txt  = (el.text or "").strip()
                aria = (el.get_attribute('aria-label') or '')
                titl = (el.get_attribute('title') or '')
                alt  = (el.get_attribute('alt') or '')
                cls  = (el.get_attribute('class') or '')
                idv  = (el.get_attribute('id') or '')
                name = (el.get_attribute('name') or '')
                if any("workflow" in s.lower() for s in (txt, aria, titl, alt, cls, idv, name)):
                    log(f"   tag={el.tag_name} id={idv!r} name={name!r} class={cls!r} "
                        f"aria={aria!r} title={titl!r} alt={alt!r} text={txt[:60]!r}")
                    count += 1
                    if count >= 40:
                        log("   ...truncated...")
                        break
            except Exception:
                continue

    # main doc
    dump_ctx(driver, "MAIN")
    # iframes
    try:
        from selenium.webdriver.common.by import By
        frames = driver.find_elements(By.TAG_NAME, "iframe")
    except Exception:
        frames = []
    for i, fr in enumerate(frames):
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame(fr)
            dump_ctx(driver, f"IFRAME #{i}")
        except Exception:
            continue
    try:
        driver.switch_to.default_content()
    except Exception:
        pass

def run_welcome_collect_workflow_dates(pn_url, pn_user, pn_pass, log_callback):
    """
    Opens Penelope, logs in, opens the Workflows panel, and collects dates from
    items titled 'Send Welcome Letter and NPP'. No Excel required.
    """
    def log(msg): 
        log_callback(msg)
    CTX["opened_task_label"] = None


    # --- imports (local so GUI import won't break if selenium is missing) ---
    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.chrome.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.common.keys import Keys
    except Exception as e:
        log(f"[FATAL][WELCOME] Missing Selenium packages: {e}")
        return

    # --- quick, robust login (tries main doc then iframes) ---
    def try_login(driver, username, password):
        def try_in_context(ctx):
            # Prefer obvious fields first
            candidates_user = [
                (By.ID, "login_userName"),
                (By.NAME, "userName"),
                (By.CSS_SELECTOR, "input[type='text']"),
                (By.CSS_SELECTOR, "input[type='email']")
            ]
            candidates_pass = [
                (By.ID, "login_password"),
                (By.NAME, "password"),
                (By.CSS_SELECTOR, "input[type='password']")
            ]
            u = p = None
            for by, sel in candidates_user:
                try:
                    u = WebDriverWait(ctx, 3).until(EC.presence_of_element_located((by, sel)))
                    if u: break
                except Exception:
                    u = None
            for by, sel in candidates_pass:
                try:
                    p = WebDriverWait(ctx, 3).until(EC.presence_of_element_located((by, sel)))
                    if p: break
                except Exception:
                    p = None
            if not (u and p):
                return False
            try:
                try: u.clear()
                except Exception: pass
                u.send_keys(username or "")
                try: p.clear()
                except Exception: pass
                p.send_keys(password or "")
            except Exception:
                return False

            # Click submit or press ENTER
            submit_locators = [
                (By.CSS_SELECTOR, "button[type='submit']"),
                (By.CSS_SELECTOR, "input[type='submit']"),
                (By.NAME, "sbm"),  # Penelope often uses this
                (By.XPATH, "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'login') or contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]"),
                (By.XPATH, "//input[@type='image' or @type='submit']")
            ]
            for by, sel in submit_locators:
                try:
                    btn = WebDriverWait(ctx, 2).until(EC.element_to_be_clickable((by, sel)))
                    try: btn.click()
                    except Exception: 
                        try: ctx.execute_script("arguments[0].click();", btn)
                        except Exception: pass
                    return True
                except Exception:
                    continue
            try:
                p.send_keys(Keys.ENTER); 
                return True
            except Exception:
                return False

        # Try main doc
        if try_in_context(driver):
            log("[PN] Submitted login (MAIN).")
            return True

        # Try each iframe
        frames = driver.find_elements(By.TAG_NAME, "iframe")
        for idx, fr in enumerate(frames):
            try:
                driver.switch_to.default_content()
                driver.switch_to.frame(fr)
                if try_in_context(driver):
                    log(f"[PN] Submitted login (IFRAME #{idx}).")
                    driver.switch_to.default_content()
                    return True
            except Exception:
                pass
            finally:
                try: driver.switch_to.default_content()
                except Exception: pass
        return False

    # --- extract matching titles + dates in current context ---
    def extract_dates_from_workflow_panel(ctx):
        results = []
        title_pat = re.compile(r"send welcome letter and npp\s*(?!-)", re.I)  # skip variants with trailing " - ..."
        date_pat  = re.compile(r"\b(\d{1,2}/\d{1,2}/\d{4})\b")
        try:
            nodes = ctx.find_elements(By.XPATH, "//*[self::div or self::li or self::tr or self::a or self::span]")
        except Exception:
            nodes = []
        for n in nodes:
            try:
                txt = (n.text or "").strip()
                if not txt:
                    continue
                if title_pat.search(txt):
                    # search same node then its nearest ancestor for a date
                    m = date_pat.search(txt)
                    if not m:
                        try:
                            parent = n.find_element(By.XPATH, "./ancestor::*[1]")
                            m = date_pat.search(parent.text or "")
                        except Exception:
                            m = None
                    results.append({"title": txt, "date": (m.group(1) if m else None)})
            except Exception:
                continue
        return results

    # --- main flow ---
    driver = None
    try:
        if not pn_user or not pn_pass:
            log("[PN] Missing PN username/password."); 
            return

        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        wait = WebDriverWait(driver, 20)

        target_url = (pn_url or PN_DEFAULT_URL)
        log("Welcome Bot — opening Penelope…")
        driver.get(target_url)
        log(f"[PN] Opened: {driver.current_url}")
        log(f"[PN] Page title: {driver.title or '(no title)'}")

        if not try_login(driver, pn_user, pn_pass):
            log("[PN] Could not locate login fields. Aborting.")
            return

        # Some signal that we moved beyond login
        try:
            wait.until(
                EC.any_of(
                    EC.url_changes(target_url),
                    EC.presence_of_element_located((By.XPATH, "//*[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'logout')]"))
                )
            )
            log("[PN] Login appears successful.")
        except Exception:
            log("[PN] No clear post-login signal; continuing anyway.")

        # Open the Workflows flyout
        if not open_workflows_flyout(driver, log):
            debug_log_all_workflowish_things(driver, log)
            if WELCOME_TEACH_MODE:
                teach_pause("[PN][TEACH] Please manually click the right-edge 'Workflow' tab in the browser now, then click OK here to continue.", log)
                if not open_workflows_flyout(driver, log):
                    log("[PN] Still couldn't open the Workflow drawer after teach step.")
                    return
            else:
                return

        # Try to read items where we are
        items = extract_dates_from_workflow_panel(driver)

        # If none, peek inside each iframe too (some flyouts are framed)
        if not items:
            try:
                frames = driver.find_elements(By.TAG_NAME, "iframe")
            except Exception:
                frames = []
            for idx, fr in enumerate(frames):
                try:
                    driver.switch_to.frame(fr)
                    items = extract_dates_from_workflow_panel(driver)
                    driver.switch_to.default_content()
                    if items:
                        break
                except Exception:
                    try: driver.switch_to.default_content()
                    except Exception: pass

        if not items:
            log("[PN] No 'Send Welcome Letter and NPP' items found in the Workflows panel.")
            return

        log(f"[PN] Found {len(items)} matching workflow item(s):")
        good_dates = []
        for i, it in enumerate(items, 1):
            log(f"  #{i}: title={it['title']!r}  date={it['date']!r}")
            if it.get("date"):
                try:
                    good_dates.append(datetime.datetime.strptime(it["date"], "%m/%d/%Y").date())
                except Exception:
                    pass

        # Build the date window from the visible 'Send Welcome Letter and NPP' items
        import datetime as _dt

        # Merge dates we parsed per-item + any dates/tokens (“Today/Tomorrow/Yesterday”) in the drawer text
        panel_text = ""
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
        try:
            panel = driver.find_element(By.ID, "drawer")
            panel_text = panel.text or ""
        except Exception:
            panel_text = ""

        def _dates_from_text(txt):
            out = []
            if not txt:
                return out
            t = txt.lower()
            today = _dt.date.today()
            if "today" in t: out.append(today)
            if "tomorrow" in t: out.append(today + _dt.timedelta(days=1))
            if "yesterday" in t: out.append(today - _dt.timedelta(days=1))
            for m in re.findall(r"\b(\d{1,2}/\d{1,2}/\d{4})\b", txt):
                try:
                    out.append(_dt.datetime.strptime(m, "%m/%d/%Y").date())
                except Exception:
                    pass
            return out

        # good_dates already contains any mm/dd/yyyy we pulled beside each Welcome item
        _merged = set(good_dates)
        for d in _dates_from_text(panel_text):
            _merged.add(d)

        if _merged:
            # Use the earliest visible Welcome date; cap the span to 14 days and keep it inclusive of any later dates we saw
            start = min(_merged)
            end   = min(max(_merged), start + _dt.timedelta(days=13))
            log(f"[PN] Using dates from drawer: {start.strftime('%m/%d/%Y')} .. {end.strftime('%m/%d/%Y')}")
        else:
            # Last-ditch fallback: the next 30 days so we don't miss near-term tasks
            start = _dt.date.today()
            end   = start + _dt.timedelta(days=30)
            log(f"[PN] No dates found; using: {start.strftime('%m/%d/%Y')} .. {end.strftime('%m/%d/%Y')}")

        # Use recorded selectors to open Advanced, set dates, search, and open first Welcome item
        adv_start = start.strftime("%m/%d/%Y")
        adv_end   = end.strftime("%m/%d/%Y")

        if not pn_search_and_open_first_welcome(driver, log, adv_start, adv_end):
            if WELCOME_TEACH_MODE:
                teach_pause(
                    "[PN][TEACH] I couldn’t auto-open a 'Send Welcome Letter and NPP' task.\n"
                    "Please click one manually (it may open in a NEW TAB). Then click OK here.",
                    log
                )
                # Follow any newly opened tab so the next step (Case link) can see it
                try:
                    handles = driver.window_handles
                    if len(handles) >= 2:
                        driver.switch_to.window(handles[-1])
                        log("[PN] Followed newly opened tab after manual click.")
                except Exception:
                    pass
            else:
                log("[PN] Could not navigate to a 'Send Welcome Letter and NPP' task. Stopping.")
                return

        # 6) On the task page, find the 'Record' box → 'Case …(digits)' link and click it
        # Robustly find and click the "Case ..." link on the task page (scan default content + all iframes)
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait

        def _all_contexts_for_task():
            ctxs = [None]  # default content first
            try:
                frames = driver.find_elements(By.CSS_SELECTOR, "iframe, frame")
            except Exception:
                frames = []
            for fr in frames:
                try:
                    if fr.is_displayed():
                        ctxs.append(fr)
                except Exception:
                    continue
            return ctxs

        def _switch_to_ctx(ctx):
            try:
                driver.switch_to.default_content()
            except Exception:
                pass
            if ctx is not None:
                try:
                    driver.switch_to.frame(ctx)
                except Exception:
                    pass

        def _find_case_link_here():
            """
            Try several shapes for the Case link.
            Priorities:
              1) Anchors whose text includes 'Case'
              2) First anchor inside anything labeled 'Record'
              3) Any anchor that looks like '... (digits)'
            """
            xpaths = [
                # Most direct: an anchor that literally mentions 'Case'
                "//a[contains(normalize-space(.), 'Case')]",
                # Inside a container labeled 'Record'
                "//*[.//legend[contains(normalize-space(.), 'Record')] or contains(@class,'record') or contains(@id,'record') or .//*[contains(normalize-space(.),'Record')]]//a[contains(normalize-space(.), 'Case')]",
                # Label 'Record' followed by a link
                "(//*[self::label or self::div or self::span][contains(normalize-space(.), 'Record')])[1]/following::a[1]",
                # Any visible link with trailing (digits)
                "//a[matches(normalize-space(.), '\\(\\d+\\)\\s*$')]",
            ]
            for xp in xpaths:
                try:
                    links = driver.find_elements(By.XPATH, xp)
                except Exception:
                    links = []
                for a in links:
                    try:
                        if a.is_displayed():
                            return a, xp
                    except Exception:
                        continue
            return None, None

        case_clicked = False
        for ctx in _all_contexts_for_task():
            try:
                _switch_to_ctx(ctx)
                # small wait to let the frame render
                try:
                    WebDriverWait(driver, 3).until(lambda d: True)
                except Exception:
                    pass

                a, used_xpath = _find_case_link_here()
                if not a:
                    continue

                try:
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", a)
                except Exception:
                    pass

                if not _safe_click(a):
                    try:
                        driver.execute_script("arguments[0].click();", a)
                    except Exception as e:
                        log(f"[PN] JS click on Case link failed in this context: {e}")
                        continue

                log(f"[PN] Clicked the Case link (ctx={'default' if ctx is None else 'iframe'}) via XPath: {used_xpath}")
                case_clicked = True
                break
            except Exception as e:
                log(f"[PN] Context scan error: {e}")
                continue

        # If we still didn't click it, pause (teach mode) and dump HTML for you to inspect
        if not case_clicked:
            # Save current HTML so we can see what's actually on the page
            try:
                import os
                dump_path = os.path.join(os.getcwd(), "last_task_page.html")
                with open(dump_path, "w", encoding="utf-8") as f:
                    f.write(driver.page_source)
                log(f"[PN] Could not find/click the Case link. Saved HTML to: {dump_path}")
            except Exception as e:
                log(f"[PN] Failed to save HTML dump: {e}")

            if WELCOME_TEACH_MODE:
                teach_pause("[PN][TEACH] Please click the correct 'Case …' link in the browser now. If it opens in a NEW TAB, I'll follow it. Click OK here to continue.", log)
                # NEW-TAB HOP: follow any newly opened tab
                try:
                    hs = driver.window_handles
                    if len(hs) > 1:
                        driver.switch_to.window(hs[-1])
                        log("[PN] Followed newly opened Case tab.")
                except Exception:
                    pass
            else:
                return

        # Always return to default content before continuing to next step
        try:
            driver.switch_to.default_content()
        except Exception:
            pass

        # 7) On the Case page, open Members and click the client name (robust)
        def _text_looks_like_name(s: str) -> bool:
            s = (s or "").strip()
            if not s:
                return False
            # Must contain letters and at least one space (e.g., "Jane Doe")
            return bool(re.search(r"[A-Za-z]", s)) and (" " in s) and len(s) <= 80

        def _open_client_from_members(timeout=12):
            end = time.time() + timeout
            while time.time() < end:
                try:
                    # Locate the Members section container
                    hdr = _find_anywhere([
                        (By.XPATH, "//*[self::h1 or self::h2 or self::h3 or self::span or self::div]"
                                   "[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'members')]")
                    ], timeout=4)
                    box = None
                    if hdr:
                        try:
                            box = hdr.find_element(By.XPATH, "./ancestor::*[self::section or self::div][1]")
                        except Exception:
                            box = None

                    # If there's a toggle to expand Members, click it
                    if box:
                        try:
                            toggle = box.find_element(By.XPATH, ".//a|.//button")
                            label = (toggle.text or "").lower()
                            if "show" in label or "expand" in label or "open" in label:
                                _safe_click(toggle)
                                time.sleep(0.3)
                        except Exception:
                            pass

                    # Gather candidate person/client links inside the section first
                    candidates = []
                    scopes = [box] if box else []
                    scopes.append(driver)
                    for scope in scopes:
                        if not scope:
                            continue
                        try:
                            anchors = scope.find_elements(
                                By.XPATH,
                                ".//a[contains(translate(@href,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'person') or "
                                "     contains(translate(@href,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'client')]"
                            ) if scope is box else scope.find_elements(
                                By.XPATH,
                                "//a[contains(translate(@href,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'person') or "
                                "    contains(translate(@href,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'client')]"
                            )
                        except Exception:
                            anchors = []

                        for a in anchors:
                            try:
                                txt = (a.text or "").strip()
                                if _is_harmful(a):
                                    continue
                                if not _text_looks_like_name(txt):
                                    continue
                                candidates.append(a)
                            except Exception:
                                continue

                    # Prefer those inside the Members box, then others; click the first safe one
                    for a in candidates:
                        try:
                            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", a)
                        except Exception:
                            pass
                        if _safe_click(a, desc="client name"):
                            log("[PN] Opened client profile from Members.")
                            return True

                    # As a last resort, click any link near the 'Members' header
                    if hdr:
                        try:
                            near = hdr.find_elements(By.XPATH, ".//following::a[1]")
                        except Exception:
                            near = []
                        for a in near:
                            try:
                                if _is_harmful(a):
                                    continue
                                if _safe_click(a, desc="near Members"):
                                    log("[PN] Opened client profile (near Members).")
                                    return True
                            except Exception:
                                continue

                except Exception:
                    pass

                time.sleep(0.4)
            return False

        if not _open_client_from_members():
            if WELCOME_TEACH_MODE:
                teach_pause("[PN][TEACH] On the Case page, click the client's name under 'Members' now, then click OK here to continue.", log)
                _maybe_switch_to_new_tab()
            else:
                log("[PN] Could not open a member/client link; stopping here.")
                return

        # 8) On the client page, extract the Main Address block
        #    We’ll look for text node 'Main Address' and read the surrounding block
        address_text = ""
        try:
            time.sleep(0.6)
            # container that has 'Main Address'
            addr_container = _find_anywhere([
                (By.XPATH, "//*[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'main address')]")
            ], timeout=10)
            if addr_container:
                # read a slightly larger context (ancestor) to catch all lines
                try:
                    box = addr_container.find_element(By.XPATH, "./ancestor::*[1]")
                except Exception:
                    box = addr_container
                raw = (box.text or "").strip()
                # Strip the heading and '(show map)'
                raw = re.sub(r"(?i)\bmain address\b", "", raw)
                raw = raw.replace("(show map)", "")
                # collapse repeated blank lines; keep the address lines
                lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
                address_text = "\n".join(lines).strip()
        except Exception:
            address_text = ""

        if address_text:
            log("[PN] Extracted address:")
            for ln in address_text.splitlines():
                log(f"    {ln}")
        else:
            if WELCOME_TEACH_MODE:
                teach_pause("[PN][TEACH] If the address is hidden/collapsed, expand the 'Main Address' section now, then click OK here to continue.", log)
                # try extraction again
                address_text = ""
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
                try:
                    addr_container = _find_anywhere([
                        (By.XPATH, "//*[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'main address')]")
                    ], timeout=8)
                    if addr_container:
                        try:
                            box = addr_container.find_element(By.XPATH, "./ancestor::*[1]")
                        except Exception:
                            box = addr_container
                        raw = (box.text or "").strip()
                        raw = re.sub(r"(?i)\bmain address\b", "", raw).replace("(show map)", "")
                        lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
                        address_text = "\n".join(lines).strip()
                except Exception:
                    address_text = ""
                if address_text:
                    log("[PN] Extracted address:")
                    for ln in address_text.splitlines():
                        log(f"    {ln}")
                else:
                    log("[PN] Still couldn't extract the address.")
            else:
                log("[PN] Could not extract the address block.")

        time.sleep(2)

    except Exception as e:
        log(f"[PN][ERROR] {e}")
    finally:
        try:
            if driver:
                driver.quit()
        except Exception:
            pass

def run_welcome_pipeline(
    username, password,
    excel_path, sheet_name,
    idx_last, idx_first, idx_date,
    welcome_letter_template, welcome_packet_template,
    base_save_dir, initials, skip_penelope,
    penelope_url, penelope_username, penelope_password,
    log_callback, stop_flag_cb=None,
    allow_loose_match=True, dry_run=False
):
    def log(msg):
        log_callback(msg)

    # ---- PN info (now OUTSIDE the log() function) ----
    log("Welcome Bot — starting (skeleton).")
    log(f"  PN URL: {penelope_url or PN_DEFAULT_URL}")
    log(f"  PN Username: {penelope_username or '(blank)'}")

    # Minimal skeleton: read Excel, loop rows, create per-patient folders.
    # (No PDF generation or uploads yet — this is just the folder creation piece.)
    try:
        from openpyxl import load_workbook
    except Exception as e:
        log(f"[FATAL][WELCOME] Missing package: {e}. Did you run pip install openpyxl?")
        return

    try:
        if not os.path.isfile(excel_path) or not excel_path.lower().endswith(".xlsx"):
            log("Please select a valid .xlsx Excel file.")
            return

        wb = load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            log(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            return
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            log("No rows found.")
            return

        # Same header detection logic as the Consent bot
        start_idx = 2
        try:
            header_tokens = ("first", "last", "date", "client", "date from tn ia", "welcome")
            def looks_like_header(v):
                return any(tok in str(v).strip().lower() for tok in header_tokens)
            if not (looks_like_header(rows[0][idx_first].value) or
                    looks_like_header(rows[0][idx_last].value) or
                    looks_like_header(rows[0][idx_date].value)):
                start_idx = 1
        except Exception:
            start_idx = 2

        prepared = 0
        base_out = None

        for ridx in range(start_idx, len(rows) + 1):
            if stop_flag_cb and stop_flag_cb():
                log("Stop requested… finishing current client and exiting.")
                break

            row = ws[ridx]
            first_name = str((row[idx_first].value or "")).strip()
            last_name  = str((row[idx_last].value or "")).strip()
            if not first_name and not last_name:
                continue

            full_name = f"{first_name} {last_name}".strip()

            # >>> THIS is the call that creates the base folder (if needed)
            # >>> and the per-patient folder using your helper.
            patient_dir, base_out = ensure_welcome_patient_dir(base_dir=base_save_dir, full_name=full_name)

            log(f"— Prepared folder: {patient_dir}")
            prepared += 1

        if base_out is None:
            # ensure_welcome_patient_dir wasn’t called (no rows), so compute default for clarity
            if not base_save_dir:
                desktop_dir = os.path.join(os.path.expanduser('~'), 'Desktop')
                base_out = os.path.join(desktop_dir, "Welcome Bot Output")
            else:
                base_out = base_save_dir

        log(f"Welcome Bot prep done. Per-patient folders prepared: {prepared}")
        log(f"Base output folder: {base_out}")

        # Next steps you’ll add later:
        # - fill welcome letter + packet PDFs with each patient's info
        # - flatten them
        # - save inside patient_dir
        # - upload to TherapyNotes (keep current date field untouched)
        # - (optionally) log a note to Penelope

    except Exception as e:
        log(f"[FATAL][WELCOME] {e}")


# ===================== GUI =====================
class OneByOneGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("ISWS Consent Form Bot - Version 3.1.0, Last Updated 12/04/2025")
        self.root.geometry("1020x820")
        self.root.configure(bg=MAROON)

        # --- ttk theme & styles for a sleeker look ---
        style = ttk.Style()
        try:
            style.theme_use('clam')
        except Exception:
            pass
        style.configure('TLabel', background=MAROON, foreground='#ffffff', font=("Helvetica", 11))
        style.configure('Header.TLabel', font=("Helvetica", 22, 'bold'))
        style.configure('TButton', font=("Helvetica", 11, 'bold'))
        style.map('TButton', focuscolor=[('!disabled', MAROON)])
        style.configure('Card.TFrame', background='#faf7f7', relief='flat')
        style.configure('TCheckbutton', background=MAROON, foreground='#ffffff', font=("Helvetica", 11, 'bold'))
        style.configure('TEntry', fieldbackground='#ffffff')
        style.configure('TCombobox', fieldbackground='#ffffff')

        # rounded-ish card container helper (padding gives a modern card feel)
        def card(parent):
            frame = ttk.Frame(parent, style='Card.TFrame')
            frame.configure(padding=(16, 14))
            return frame

        self.log_queue = queue.Queue()
        self.stop_flag = False

        self.username = tk.StringVar()
        self.password = tk.StringVar()
        self.excel_path = ""
        self.pdf_template_path = ""
        self.selected_sheet = tk.StringVar()
        self.col_last = tk.StringVar(value="A")
        self.col_first = tk.StringVar(value="B")
        self.col_date = tk.StringVar(value="C")
        self.skip_if_date = tk.BooleanVar(value=True)   # default ON
        self.dry_run = tk.BooleanVar(value=False)       # NEW: dry run toggle
        self.loose_match = tk.BooleanVar(value=True)    # NEW: allow flexible matching (default ON)


        # Header
        self.header = ttk.Label(root, text=APP_TITLE, style='Header.TLabel', anchor='center')
        self.header.pack(pady=(14, 6), fill='x')

        # ===== SCROLLABLE CONTAINER (everything below the header goes inside this) =====
        container = ttk.Frame(root)
        container.pack(fill="both", expand=True, padx=0, pady=0)

        # Canvas + vertical scrollbar
        self.scroll_canvas = tk.Canvas(
            container,
            background=MAROON,
            highlightthickness=0
        )
        vscroll = ttk.Scrollbar(container, orient="vertical", command=self.scroll_canvas.yview)
        self.scroll_canvas.configure(yscrollcommand=vscroll.set)

        self.scroll_canvas.pack(side="left", fill="both", expand=True)
        vscroll.pack(side="right", fill="y")

        # Frame that will actually hold all your content
        self.scroll_frame = ttk.Frame(self.scroll_canvas, style='Card.TFrame')
        self.scroll_window = self.scroll_canvas.create_window(
             (0, 0), window=self.scroll_frame, anchor="nw"
        )

        # Keep scrollregion synced to content size
        def _on_frame_configure(event):
            self.scroll_canvas.configure(scrollregion=self.scroll_canvas.bbox("all"))

        self.scroll_frame.bind("<Configure>", _on_frame_configure)

        # Make the canvas width follow the container width
        def _on_canvas_configure(event):
            # Keep the inner frame the same width as the canvas
            self.scroll_canvas.itemconfig(self.scroll_window, width=event.width)

        self.scroll_canvas.bind("<Configure>", _on_canvas_configure)

        # Mouse wheel scrolling (Windows)
        def _on_mousewheel(event):
            self.scroll_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        self.scroll_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        # ============================================================================


        # Cards grid
        grid = ttk.Frame(self.scroll_frame, style='Card.TFrame')
        grid.pack(padx=14, pady=8, fill='x')

        # Credentials Card
        cred = card(grid)
        cred.grid(row=0, column=0, sticky='ew', padx=(0, 10))
        cred.columnconfigure(1, weight=1)
        ttk.Label(cred, text="TN Username:").grid(row=0, column=0, sticky='e', padx=6, pady=4)
        ttk.Entry(cred, textvariable=self.username, width=35).grid(row=0, column=1, sticky='ew', padx=6, pady=4)
        ttk.Label(cred, text="TN Password:").grid(row=1, column=0, sticky='e', padx=6, pady=4)
        ttk.Entry(cred, textvariable=self.password, width=35, show='*').grid(row=1, column=1, sticky='ew', padx=6, pady=4)

        # Files Card
        files = card(grid)
        files.grid(row=0, column=1, sticky='ew')
        ttk.Button(files, text="Select Excel", command=self.pick_excel).grid(row=0, column=0, padx=6, pady=6)
        ttk.Button(files, text="Select PDF Template", command=self.pick_pdf_template).grid(row=0, column=1, padx=6, pady=6)
        self.excel_label = ttk.Label(files, text="No Excel selected")
        self.excel_label.grid(row=1, column=0, columnspan=2, sticky='w')
        self.pdf_template_label = ttk.Label(files, text="No PDF template selected")
        self.pdf_template_label.grid(row=2, column=0, columnspan=2, sticky='w')

        # Make sure we add these cards inside the same scrollable area as the other controls
        parent_for_cards = self.scroll_frame if hasattr(self, "scroll_frame") else root


        # Mapping Card
        self.mapping_card = card(self.scroll_frame)
        self.mapping_card.pack(padx=14, pady=6, fill='x')
        ttk.Label(self.mapping_card, text="Sheet:").grid(row=0, column=0, sticky='e', padx=5, pady=4)
        self.sheet_combo = ttk.Combobox(self.mapping_card, textvariable=self.selected_sheet, width=32, state="readonly")
        self.sheet_combo.grid(row=0, column=1, sticky='w', padx=5)
        ttk.Label(self.mapping_card, text="Last Name col:").grid(row=1, column=0, sticky='e', padx=5, pady=4)
        ttk.Entry(self.mapping_card, textvariable=self.col_last, width=6).grid(row=1, column=1, sticky='w', padx=5)
        ttk.Label(self.mapping_card, text="First Name col:").grid(row=1, column=2, sticky='e', padx=5, pady=4)
        ttk.Entry(self.mapping_card, textvariable=self.col_first, width=6).grid(row=1, column=3, sticky='w', padx=5)
        ttk.Label(self.mapping_card, text="Date col:").grid(row=1, column=4, sticky='e', padx=5, pady=4)
        ttk.Entry(self.mapping_card, textvariable=self.col_date, width=6).grid(row=1, column=5, sticky='w', padx=5)

        # Options Card
        options = card(self.scroll_frame)
        options.pack(padx=14, pady=6, fill='x')
        ttk.Checkbutton(options, text="Skip any row that already has a Date in Excel",
                        variable=self.skip_if_date).grid(row=0, column=0, sticky='w', padx=6, pady=2)
        ttk.Checkbutton(options, text="Dry run (simulate, no uploads)",
                        variable=self.dry_run).grid(row=0, column=1, sticky='w', padx=12, pady=2)
        ttk.Checkbutton(options, text="Allow flexible name match (ignore middle initials, punctuation)",
                variable=self.loose_match).grid(row=0, column=2, sticky='w', padx=12, pady=2)
        ttk.Label(options, text=f"Write-back column: {STATUS_COL_LETTER} (status messages)").grid(row=1, column=0, columnspan=2, sticky='w', padx=6)

        # Controls Card
        controls = card(self.scroll_frame)
        controls.pack(padx=14, pady=6, fill='x')
        self.start_btn = ttk.Button(controls, text="Start", command=self.start)
        self.start_btn.grid(row=0, column=0, padx=6)
        self.stop_btn = ttk.Button(controls, text="Stop", command=self.stop)
        self.stop_btn.grid(row=0, column=1, padx=6)
        self.help_btn = ttk.Button(controls, text="Help", command=self.show_help)
        self.help_btn.grid(row=0, column=2, padx=6)


        # Log area
        log_card = card(self.scroll_frame)
        log_card.pack(padx=14, pady=10, fill='both', expand=True)
        self.log_area = scrolledtext.ScrolledText(log_card, width=120, height=28,
                                                  bg="#f5f5f5", fg="#000000", font=("Courier New", 10))
        self.log_area.pack(padx=4, pady=4, fill="both", expand=True)
        
        
        self.root.after(100, self.flush_log)

    def pick_excel(self):
        try:
            from openpyxl import load_workbook
        except Exception:
            self.enqueue_log("[ERROR] openpyxl is not installed. Run: pip install openpyxl")
            return
        f = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not f:
            return
        self.excel_path = f
        self.enqueue_log(f"Selected Excel: {f}")
        self.excel_label.config(text=f)

        try:
            wb = load_workbook(f, data_only=True)
            sheetnames = wb.sheetnames
            self.sheet_combo["values"] = sheetnames
            if sheetnames:
                self.selected_sheet.set(sheetnames[0])
        except Exception as e:
            self.enqueue_log(f"[ERROR] Could not load workbook: {e}")
            self.sheet_combo["values"] = []
            self.selected_sheet.set("")

    def pick_pdf_template(self):
        f = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if f:
            self.pdf_template_path = f
            self.enqueue_log(f"Selected PDF template: {f}")
            self.pdf_template_label.config(text=f)

    


    def pick_base_output_dir(self):
        d = filedialog.askdirectory()
        if d:
            self.base_save_dir.set(d)
            self.lbl_base_save.config(text=d)
            self.enqueue_log(f"Base save folder: {d}")

    def enqueue_log(self, msg):
        self.log_queue.put(msg)

    def flush_log(self):
        while not self.log_queue.empty():
            msg = self.log_queue.get()
            self.log_area.insert(tk.END, msg + "\n")
            self.log_area.see(tk.END)
        self.root.after(100, self.flush_log)

    def show_help(self):
        help_win = tk.Toplevel(self.root)
        help_win.title("ISWS Bot Help - Version 3.1.0, Last Updated 12/04/2025")
        help_win.geometry("860x640")

        # top bar
        top = ttk.Frame(help_win)
        top.pack(fill="x", padx=8, pady=8)
        ttk.Label(top, text="ISWS Consent Form Bot — Help", style="Header.TLabel").pack(side="left")

        # scrolling text
        text_frame = ttk.Frame(help_win)
        text_frame.pack(fill="both", expand=True, padx=8, pady=(0,8))
        txt = scrolledtext.ScrolledText(text_frame, wrap="word", font=("Segoe UI", 10))
        txt.pack(fill="both", expand=True)
        txt.insert("1.0", HELP_TEXT)
        txt.configure(state="disabled")  # read-only

        # bottom buttons
        bottom = ttk.Frame(help_win)
        bottom.pack(fill="x", padx=8, pady=(0,8))
        def copy_all():
            help_win.clipboard_clear()
            help_win.clipboard_append(HELP_TEXT)
            self.enqueue_log("[Help] Copied help text to clipboard.")
        ttk.Button(bottom, text="Copy to clipboard", command=copy_all).pack(side="left")
        ttk.Button(bottom, text="Close", command=help_win.destroy).pack(side="right")

    def stop(self):
        self.stop_flag = True
        self.enqueue_log("Stop requested… finishing current client and exiting.")

    def start(self):
        # ---------- Gather inputs ----------
        tn_username = self.username.get().strip()
        tn_password = self.password.get().strip()
        excel_path  = self.excel_path
        pdf_template = self.pdf_template_path
        sheet = self.selected_sheet.get()

        # ---------- Validation ----------
        if not tn_username or not tn_password or not excel_path or not sheet or not pdf_template:
            self.enqueue_log("Please enter TN username/password, select Excel + sheet, and select a PDF template.")
            return

        # Convert A/B/C letters to indices
        idx_last  = col_letter_to_index(self.col_last.get())
        idx_first = col_letter_to_index(self.col_first.get())
        idx_date  = col_letter_to_index(self.col_date.get())
        if None in (idx_last, idx_first, idx_date):
            self.enqueue_log("Invalid column letters. Use A, B, C, ...")
            return

        # ---------- Start background worker ----------
        self.stop_flag = False
        self.start_btn.config(state="disabled")
        self.enqueue_log("Starting…")

        def worker():
            try:
                # Run Consent bot pipeline
                run_pipeline(
                    tn_username, tn_password,
                    excel_path, sheet,
                    idx_last, idx_first, idx_date,
                    pdf_template,
                    self.enqueue_log, lambda: self.stop_flag,
                    skip_if_date=self.skip_if_date.get(),
                    dry_run=self.dry_run.get(),
                    allow_loose_match=self.loose_match.get()
                )
            finally:
                self.enqueue_log("Process finished.")
                self.start_btn.config(state="normal")

        threading.Thread(target=worker, daemon=True).start()

# ------------------ RUN ------------------
if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = OneByOneGUI(root)
        root.mainloop()
    except Exception as e:
        # Last-resort: show something in console if GUI init explodes
        print("[FATAL] GUI failed to start:", e)
        sys.exit(1)
