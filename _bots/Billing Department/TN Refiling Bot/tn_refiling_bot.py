#!/usr/bin/env python3
"""
TN Refiling Bot - Therapy Notes Refiling Bot
This bot automates refiling claims in Therapy Notes by reading Excel/CSV input data
and searching for clients using DOB, name, and date of service.
Also extracts claim numbers from scanned PDF documents using OCR.
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, simpledialog, filedialog
import pyautogui
import pywinauto
import time
import logging
from pathlib import Path
import subprocess
import sys
import threading
import json
import os
import re
import queue
import importlib
from logging.handlers import RotatingFileHandler

# Try to import keyboard module (optional - needed for hotkeys)
try:
    import keyboard
    KEYBOARD_AVAILABLE = True
except ImportError:
    KEYBOARD_AVAILABLE = False
    keyboard = None

# Try to import OpenCV (optional - needed for confidence parameter in image recognition)
try:
    import cv2
    OPENCV_AVAILABLE = True
except ImportError:
    OPENCV_AVAILABLE = False
    cv2 = None

# Try to import pdfplumber (required for parsing PDFs)
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False
    pdfplumber = None

# Try to import OCR libraries for scanned PDFs
try:
    import pytesseract
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    pytesseract = None

# Try to import Excel/CSV reading libraries
try:
    import pandas as pd
    EXCEL_AVAILABLE = True
    # Try to import openpyxl for .xlsx files
    try:
        import openpyxl
        OPENPYXL_AVAILABLE = True
    except ImportError:
        OPENPYXL_AVAILABLE = False
except ImportError:
    EXCEL_AVAILABLE = False
    pd = None
    OPENPYXL_AVAILABLE = False

# Try to import Selenium WebDriver for browser automation
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import WebDriverException, TimeoutException
    SELENIUM_AVAILABLE = True
    
    # Try to import webdriver-manager for automatic ChromeDriver management
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.chrome.service import Service as ChromeService
        WEBDRIVER_MANAGER_AVAILABLE = True
    except ImportError:
        WEBDRIVER_MANAGER_AVAILABLE = False
        ChromeDriverManager = None
        ChromeService = None
except ImportError:
    SELENIUM_AVAILABLE = False
    WEBDRIVER_MANAGER_AVAILABLE = False
    webdriver = None

# Try to import keyring for secure credential storage
try:
    keyring = importlib.import_module("keyring")
    keyring_errors = importlib.import_module("keyring.errors")
    KeyringError = getattr(keyring_errors, "KeyringError", Exception)
    KEYRING_AVAILABLE = True
except ImportError:
    KEYRING_AVAILABLE = False
    keyring = None
    KeyringError = Exception

# Configure logging with comprehensive debug information and rotation
LOG_FILE_PATH = Path(__file__).parent / 'tn_refiling_bot.log'
LOG_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)

file_log_handler = RotatingFileHandler(
    filename=str(LOG_FILE_PATH),
    maxBytes=5 * 1024 * 1024,  # 5 MB per log file
    backupCount=5,
    encoding='utf-8'
)

logging.basicConfig(
    level=logging.DEBUG,  # Set to DEBUG for comprehensive logging
    format='%(asctime)s - %(name)s - %(levelname)s - %(filename)s:%(lineno)d - %(funcName)s() - %(message)s',
    handlers=[
        file_log_handler,
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def _configure_ocr_paths():
    """Detect and configure Tesseract and Poppler paths on any Windows machine.
    
    Precedence (highest to lowest):
    1) Environment variables: TESSERACT_PATH, POPPLER_PATH
    2) Local vendor folder next to this script: vendor/Tesseract-OCR, vendor/poppler/Library/bin
    3) Common install locations (Program Files, LocalAppData, Conda)
    """
    if not (OCR_AVAILABLE and pytesseract):
        return
    try:
        script_dir = Path(__file__).parent
        # 1) Env vars
        tesseract_env = os.environ.get('TESSERACT_PATH')
        poppler_env = os.environ.get('POPPLER_PATH')
        if tesseract_env and Path(tesseract_env).exists():
            pytesseract.pytesseract.tesseract_cmd = tesseract_env
            logger.info(f"Configured Tesseract from env: {tesseract_env}")
        if poppler_env and Path(poppler_env).exists():
            os.environ['POPPLER_PATH'] = poppler_env
            logger.info(f"Configured Poppler from env: {poppler_env}")
        # 2) Vendor folder
        if not getattr(pytesseract.pytesseract, 'tesseract_cmd', None):
            vend_tess = script_dir / 'vendor' / 'Tesseract-OCR' / 'tesseract.exe'
            if vend_tess.exists():
                pytesseract.pytesseract.tesseract_cmd = str(vend_tess)
                logger.info(f"Configured Tesseract from vendor: {vend_tess}")
        if 'POPPLER_PATH' not in os.environ:
            vend_poppler = script_dir / 'vendor' / 'poppler' / 'Library' / 'bin'
            if vend_poppler.exists():
                os.environ['POPPLER_PATH'] = str(vend_poppler)
                logger.info(f"Configured Poppler from vendor: {vend_poppler}")
        # 3) Common locations
        current_cmd = getattr(pytesseract.pytesseract, 'tesseract_cmd', None)
        if not current_cmd or not Path(current_cmd).exists() if current_cmd else True:
            candidates = [
                Path('C:/Program Files/Tesseract-OCR/tesseract.exe'),
                Path('C:/Program Files (x86)/Tesseract-OCR/tesseract.exe'),
                Path.home() / 'AppData/Local/Programs/Tesseract-OCR/tesseract.exe',
                Path(r'C:\Users\mthompson\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'),  # Your specific path
            ]
            conda_prefix = os.environ.get('CONDA_PREFIX')
            if conda_prefix:
                candidates.append(Path(conda_prefix) / 'Library/bin/tesseract.exe')
            for c in candidates:
                if c.exists():
                    pytesseract.pytesseract.tesseract_cmd = str(c)
                    logger.info(f"Configured Tesseract from common path: {c}")
                    break
            # Final check: verify tesseract is callable
            try:
                pytesseract.get_tesseract_version()
                logger.info("Tesseract verified successfully")
            except Exception as e:
                logger.warning(f"Tesseract path configured but not working: {e}")
        if 'POPPLER_PATH' not in os.environ:
            poppler_candidates = [
                Path('C:/Program Files/poppler/Library/bin'),
                Path('C:/Program Files (x86)/poppler/Library/bin'),
                Path.home() / 'AppData/Local/poppler/Library/bin',
                Path(r'C:\Users\mthompson\Downloads\Release-25.07.0-0\poppler-25.07.0\Library\bin'),  # Your specific path
            ]
            if conda_prefix:
                poppler_candidates.append(Path(conda_prefix) / 'Library/bin')
            for d in poppler_candidates:
                if d.exists() and (d / 'pdftoppm.exe').exists():
                    os.environ['POPPLER_PATH'] = str(d)
                    logger.info(f"Configured Poppler from common path: {d}")
                    break
    except Exception as e:
        logger.warning(f"OCR path auto-config failed: {e}")

_configure_ocr_paths()

# Safety settings for PyAutoGUI
pyautogui.PAUSE = 1  # Add 1 second pause between actions
pyautogui.FAILSAFE = True  # Move mouse to corner to abort


class TNRefilingBot:
    """Main bot class for Therapy Notes refiling automation"""
    
    def __init__(self):
        # Therapy Notes URL
        self.therapy_notes_url = "https://www.therapynotes.com/app/login/IntegritySWS/?r=%2fapp%2fpatients%2f"
        
        self.app = None
        self.login_window = None
        self.main_window = None
        self.log_text = None
        self.root = None
        self.gui_log_queue = queue.Queue()
        self._gui_log_dispatcher_active = False
        self.is_logged_in = False  # Track login status
        self.users_file = Path(__file__).parent / "tn_users.json"  # Store users in same directory as script
        self.keyring_service = "TN Refiling Bot"
        self._keyring_warning_shown = False
        self._thread_lock = threading.Lock()
        self._active_threads = set()
        self._shutdown_requested = False
        self._last_waiting_thread_count = None
        self.users = {}  # Dictionary to store users: {"name": {"username": "...", "password": "..."}}
        self.load_users()  # Load existing users on startup
        
        # Coordinate storage file for training
        self.coords_file = Path(__file__).parent / "tn_coordinates.json"
        self.coordinates = {}  # Dictionary to store coordinates: {"element_name": {"x": 100, "y": 50}, ...}
        self.load_coordinates()  # Load saved coordinates
        
        # Selected PDF file for claim number extraction
        self.selected_pdf_path = None
        
        # Selected Excel/CSV file for batch processing (multiple clients)
        self.selected_excel_path = None
        self.excel_client_data = []  # List of dicts: [{"client_name": "...", "dob": "...", "date_of_service": "..."}, ...]
        
        # Manual column mapping settings
        self.use_manual_mapping = False
        self.manual_column_map = {
            'last_name_col': '',      # Column letter/name for Last Name
            'first_name_col': '',     # Column letter/name for First Name
            'name_col': '',           # Column letter/name for Full Name (if not split)
            'dob_col': '',            # Column letter/name for DOB
            'dos_col': ''             # Column letter/name for Date of Service
        }
        
        # Selenium WebDriver for browser automation
        self.driver = None
        self.wait = None
        
        # Processing statistics
        self.processing_stats = {
            'total': 0,
            'successful': 0,
            'failed': 0,
            'current': 0
        }
        
        # Output Excel file path (user-selectable)
        self.output_excel_path = None  # User-selected path for output Excel file
        
        # Comprehensive client tracking (for detailed output Excel)
        # Tracks ALL clients with full details: refiled, skipped, correct modifier, etc.
        self.tracked_clients = []  # List of dicts with comprehensive data:
        # {
        #   'client_name': str,
        #   'dob': str,
        #   'date_of_service': str,
        #   'excel_row': int,
        #   'status': str,  # 'Refiled', 'Not Refiled - No View ERA', 'Not Refiled - Modifier Correct', 'Error', etc.
        #   'original_modifier': str,  # '93' or '95' or 'N/A'
        #   'new_modifier': str,  # What it was changed to, or 'No Change' if unchanged
        #   'session_medium': str,  # 'video' or 'phone' or 'N/A'
        #   'expected_modifier': str,  # What modifier should be based on session medium
        #   'payer_claim_control': str,  # Payer Claim Control # from ERA
        #   'modifier_action': str,  # 'Updated', 'Already Correct', 'Not Checked', 'N/A'
        #   'reason': str,  # Detailed reason for status
        #   'processing_date': str,  # Timestamp when processed
        #   'error_message': str  # Error details if any
        # }
        
        # Legacy tracking (kept for backward compatibility)
        self.skipped_clients = []  # List of dicts: [{"client_name": "...", "dob": "...", "date_of_service": "...", "reason": "..."}, ...]
        
        # Clients with correct modifiers tracking (for output Excel)
        self.correct_modifier_clients = []  # List of dicts: [{"client_name": "...", "dob": "...", "date_of_service": "...", "reason": "Modifier was already correct, Not Resubmitted"}, ...]
        
        # Current processing subset (populated when running real processing)
        self.current_processing_clients = []
        
        # Current popup capture data
        self.current_primary_policy_name = None
        self.current_primary_policy_member_id = None
        self.current_primary_payment_description = None
        
        # Stop control
        self.stop_requested = False
        
        # Test mode flag (to stop before saving changes during testing)
        self.is_testing_mode = False
        # Complex testing mode flag (to run full flow but stop before final Submit Claims)
        self.is_complex_testing_mode = False
        self.test_row_number = None  # Store current test row number for status updates
    
    def _ensure_keyring_available(self):
        """Warn once if keyring is unavailable for secure storage."""
        if KEYRING_AVAILABLE:
            return True
        if not self._keyring_warning_shown:
            warning_msg = (
                "Secure credential storage is unavailable because the 'keyring' package is not installed.\n"
                "Passwords will not be saved between sessions.\n"
                "Install it with: pip install keyring"
            )
            self.gui_log(warning_msg, level="WARNING", include_context=False)
            try:
                if self.root:
                    messagebox.showwarning("Secure Storage Unavailable", warning_msg)
            except Exception:
                # Ignore messagebox failures if called before Tk is ready
                pass
            self._keyring_warning_shown = True
        return False

    def _store_secure_password(self, username, password):
        """Persist password in OS credential manager via keyring."""
        if not username or not password:
            return False
        if not KEYRING_AVAILABLE:
            self._ensure_keyring_available()
            return False
        try:
            current_secret = keyring.get_password(self.keyring_service, username)
            if current_secret != password:
                keyring.set_password(self.keyring_service, username, password)
            return True
        except KeyringError as e:
            self.log_error(f"Failed to store password securely for user '{username}'", exception=e, include_traceback=False)
        except Exception as e:
            self.log_error(f"Unexpected error storing password for '{username}'", exception=e, include_traceback=False)
        return False

    def _retrieve_secure_password(self, username):
        """Fetch password from credential manager if available."""
        if not username or not KEYRING_AVAILABLE:
            if not KEYRING_AVAILABLE:
                self._ensure_keyring_available()
            return None
        try:
            return keyring.get_password(self.keyring_service, username)
        except KeyringError as e:
            self.log_error(f"Failed to retrieve password for user '{username}'", exception=e, include_traceback=False)
        except Exception as e:
            self.log_error(f"Unexpected error retrieving password for '{username}'", exception=e, include_traceback=False)
        return None

    def _delete_secure_password(self, username):
        """Remove stored password when user is deleted or renamed."""
        if not username or not KEYRING_AVAILABLE:
            return
        try:
            keyring.delete_password(self.keyring_service, username)
        except KeyringError:
            # Ignore missing entries; no need to log noise
            pass
        except Exception as e:
            self.log_error(f"Unexpected error removing password for '{username}'", exception=e, include_traceback=False)

    def load_users(self):
        """Load users from JSON file and hydrate passwords from keyring."""
        try:
            if self.users_file.exists():
                with open(self.users_file, 'r') as f:
                    stored_users = json.load(f)
            else:
                stored_users = {}

            sanitized_users = {}
            migrated = False

            for display_name, creds in stored_users.items():
                username = (creds or {}).get('username', '')
                password = None

                if username:
                    password = self._retrieve_secure_password(username)
                    legacy_password = (creds or {}).get('password')
                    if not password and legacy_password:
                        if self._store_secure_password(username, legacy_password):
                            password = legacy_password
                            migrated = True
                        else:
                            password = legacy_password if not KEYRING_AVAILABLE else None

                sanitized_users[display_name] = {
                    'username': username,
                    'password': password or ''
                }

            self.users = sanitized_users

            if migrated:
                self.gui_log("Migrated legacy credentials into secure storage", level="INFO", include_context=False)
                self.save_users()

            logger.info(f"Loaded {len(self.users)} users from {self.users_file}")

            if not self.users_file.exists():
                # Ensure file exists even if empty
                self.save_users()

        except Exception as e:
            logger.error(f"Error loading users: {e}")
            self.users = {}
    
    def save_users(self):
        """Persist user metadata to disk and secrets to keyring."""
        try:
            data_to_save = {}

            for display_name, creds in (self.users or {}).items():
                username = (creds or {}).get('username', '').strip()
                password = (creds or {}).get('password', '').strip()

                data_to_save[display_name] = {'username': username}

                if password:
                    if not self._store_secure_password(username, password):
                        # If secure storage failed, keep password only in memory for current session
                        self.gui_log(
                            f"Password for user '{display_name}' was not stored securely. It will need to be re-entered next session.",
                            level="WARNING",
                            include_context=False
                        )

            with open(self.users_file, 'w', encoding='utf-8') as f:
                json.dump(data_to_save, f, indent=2)

            logger.info(f"Saved {len(data_to_save)} users to {self.users_file}")
            return True
        except Exception as e:
            logger.error(f"Error saving users: {e}")
            try:
                messagebox.showerror("Error", f"Failed to save users:\n{e}")
            except Exception:
                pass
            return False
    
    def load_coordinates(self):
        """Load saved coordinates from JSON file"""
        try:
            if self.coords_file.exists():
                with open(self.coords_file, 'r') as f:
                    self.coordinates = json.load(f)
                logger.info(f"Loaded {len(self.coordinates)} coordinates from {self.coords_file}")
            else:
                self.coordinates = {}
                self.save_coordinates()
        except Exception as e:
            logger.error(f"Error loading coordinates: {e}")
            self.coordinates = {}
    
    def save_coordinates(self):
        """Save coordinates to JSON file"""
        try:
            with open(self.coords_file, 'w') as f:
                json.dump(self.coordinates, f, indent=2)
            return True
        except Exception as e:
            logger.error(f"Error saving coordinates: {e}")
            return False
    
    def get_coordinate(self, element_name):
        """Get saved coordinates for an element, return (x, y) tuple or None"""
        if element_name in self.coordinates:
            coords = self.coordinates[element_name]
            if 'x' in coords and 'y' in coords:
                return (coords['x'], coords['y'])
        return None
    
    def _get_alive_worker_threads(self):
        """Return a list of currently alive managed worker threads."""
        with self._thread_lock:
            alive = []
            finished = []
            for thread in self._active_threads:
                if thread.is_alive():
                    alive.append(thread)
                else:
                    finished.append(thread)
            for thread in finished:
                self._active_threads.discard(thread)
        return alive

    def _start_worker_thread(self, name, target, args=(), kwargs=None):
        """Start a non-daemon worker thread and track it for shutdown."""
        if self._shutdown_requested:
            self.gui_log(
                f"Cannot start background task '{name}' while shutdown is in progress",
                level="WARNING",
                include_context=False
            )
            return None

        if kwargs is None:
            kwargs = {}

        thread_ref = {}

        def runner():
            try:
                target(*args, **kwargs)
            finally:
                with self._thread_lock:
                    thread_obj = thread_ref.get('thread', threading.current_thread())
                    self._active_threads.discard(thread_obj)

        thread = threading.Thread(target=runner, name=f"TNBot-{name}", daemon=False)
        thread_ref['thread'] = thread

        with self._thread_lock:
            self._active_threads.add(thread)

        thread.start()
        return thread

    def _wait_for_threads_and_close(self):
        """Wait for managed threads to finish before closing the application."""
        alive_threads = self._get_alive_worker_threads()

        if not alive_threads:
            if self.driver:
                self._close_driver()
            self.gui_log("✅ All background tasks completed. Shutting down bot.", level="INFO", include_context=False)
            if self.root:
                try:
                    self.root.quit()
                    self.root.destroy()
                except Exception:
                    pass
            return

        count = len(alive_threads)
        if self._last_waiting_thread_count != count:
            thread_names = ', '.join(t.name or f"Thread-{i}" for i, t in enumerate(alive_threads[:3], 1))
            if count > 3:
                thread_names += ", ..."
            self.gui_log(
                f"Waiting for {count} background task(s) to finish before closing ({thread_names})...",
                level="INFO",
                include_context=False
            )
            self._last_waiting_thread_count = count

        if self.root:
            try:
                self.root.after(250, self._wait_for_threads_and_close)
            except Exception:
                pass

    def _root_is_ready(self):
        """Return True if the Tk root window exists and is responsive."""
        if not self.root:
            return False
        try:
            return bool(self.root.winfo_exists())
        except tk.TclError:
            return False

    def _start_gui_log_dispatcher(self):
        """Ensure the log queue dispatcher is running on the Tk main loop."""
        if self._gui_log_dispatcher_active:
            return
        if not self._root_is_ready():
            return
        self._gui_log_dispatcher_active = True
        self.root.after(50, self._process_gui_log_queue)

    def _process_gui_log_queue(self):
        """Flush queued log entries onto the GUI from the main thread."""
        if not self._root_is_ready():
            self._gui_log_dispatcher_active = False
            return

        schedule_next = False
        while True:
            try:
                log_entry, level = self.gui_log_queue.get_nowait()
            except queue.Empty:
                break

            if not self.log_text:
                # GUI not yet ready; requeue and try again shortly.
                self.gui_log_queue.put((log_entry, level))
                schedule_next = True
                break

            self._append_gui_log_entry(log_entry, level)
            schedule_next = True

        if schedule_next or not self.gui_log_queue.empty():
            self.root.after(50, self._process_gui_log_queue)
        else:
            self._gui_log_dispatcher_active = False

    def _append_gui_log_entry(self, log_entry, level):
        """Append a single log entry to the GUI widget with styling."""
        if not self.log_text:
            return

        self.log_text.insert(tk.END, log_entry)

        if level == "ERROR":
            self.log_text.tag_add("error", f"end-{len(log_entry)}c", "end-1c")
            self.log_text.tag_config("error", foreground="red")
        elif level == "WARNING":
            self.log_text.tag_add("warning", f"end-{len(log_entry)}c", "end-1c")
            self.log_text.tag_config("warning", foreground="orange")
        elif level == "DEBUG":
            self.log_text.tag_add("debug", f"end-{len(log_entry)}c", "end-1c")
            self.log_text.tag_config("debug", foreground="blue")

        self.log_text.see(tk.END)

    def _drain_gui_log_queue(self):
        """Flush any queued log entries immediately (main thread only)."""
        if not self.log_text or threading.current_thread() is not threading.main_thread():
            return

        while True:
            try:
                log_entry, level = self.gui_log_queue.get_nowait()
            except queue.Empty:
                break
            self._append_gui_log_entry(log_entry, level)

    def gui_log(self, message, level="INFO", include_context=True):
        """Log a message to the GUI log window with comprehensive debugging."""
        timestamp = time.strftime("%H:%M:%S")

        if include_context and level in ["DEBUG", "ERROR", "WARNING"]:
            import inspect
            try:
                frame = inspect.currentframe().f_back
                filename = Path(frame.f_code.co_filename).name
                line_num = frame.f_lineno
                function = frame.f_code.co_name
                context = f"[{filename}:{line_num}:{function}]"
            except Exception:
                context = "[context unavailable]"
            log_entry = f"[{timestamp}] [{level}] {context} {message}\n"
        else:
            log_entry = f"[{timestamp}] [{level}] {message}\n"

        if self.log_text and threading.current_thread() is threading.main_thread():
            self._drain_gui_log_queue()
            self._append_gui_log_entry(log_entry, level)
        else:
            self.gui_log_queue.put((log_entry, level))
            if self._root_is_ready():
                self._start_gui_log_dispatcher()

        # Also log to file logger
        if level == "ERROR":
            logger.error(message)
        elif level == "WARNING":
            logger.warning(message)
        elif level == "DEBUG":
            logger.debug(message)
        else:
            logger.info(message)
    
    def log_error(self, message, exception=None, include_traceback=True):
        """Log an error with comprehensive debugging information"""
        self.gui_log(f"ERROR: {message}", level="ERROR", include_context=True)
        
        if exception:
            error_type = type(exception).__name__
            error_msg = str(exception)
            self.gui_log(f"  Exception Type: {error_type}", level="ERROR", include_context=False)
            self.gui_log(f"  Exception Message: {error_msg}", level="ERROR", include_context=False)
            
            if include_traceback:
                import traceback
                tb_lines = traceback.format_exc().split('\n')
                self.gui_log("  Traceback:", level="ERROR", include_context=False)
                for line in tb_lines:
                    if line.strip():
                        self.gui_log(f"    {line}", level="ERROR", include_context=False)
        
        # Also log to file
        if exception:
            logger.exception(message)
        else:
            logger.error(message)
    
    def log_debug(self, message):
        """Log a debug message with context"""
        self.gui_log(message, level="DEBUG", include_context=True)
    
    def request_stop(self):
        """Request bot to stop all operations"""
        self.stop_requested = True
        self.gui_log("⛔ STOP REQUESTED - Bot will stop after current operation", level="WARNING")
        self.update_status("Stop requested...", "#dc3545")
    
    def check_stop_requested(self):
        """Check if stop was requested, log if so"""
        if self.stop_requested:
            self.gui_log("⛔ Stop requested - aborting current operation", level="WARNING")
            return True
        return False
    
    def update_status(self, message, color="#0066cc"):
        """Update the status label in the GUI"""
        if hasattr(self, 'status_label'):
            self.status_label.config(text=f"Status: {message}", fg=color)
    
    def extract_claim_number_from_pdf(self, pdf_path):
        """Extract claim number from a scanned PDF document using OCR"""
        if not PDFPLUMBER_AVAILABLE:
            self.gui_log("pdfplumber not available - cannot parse PDF")
            return None
        
        self.gui_log(f"Extracting claim number from PDF: {pdf_path}")
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Extract text from all pages
                full_text = ""
                pages_text = []
                
                for page_num, page in enumerate(pdf.pages, 1):
                    page_text = page.extract_text() or ""
                    full_text += page_text
                    full_text += "\n"
                    pages_text.append((page_num, page_text))
                
                # If no text extracted, PDF might be scanned - try OCR if available
                if not full_text.strip() and OCR_AVAILABLE:
                    self.gui_log("PDF appears to be scanned (no text found) - attempting OCR...")
                    try:
                        from pdf2image import convert_from_path
                        poppler_path = os.environ.get('POPPLER_PATH')
                        if not poppler_path:
                            # Fallback to default location
                            poppler_path = r"C:\Users\mthompson\Downloads\Release-25.07.0-0\poppler-25.07.0\Library\bin"
                        
                        # OCR all pages (up to 5 pages to avoid long processing)
                        max_ocr_pages = min(len(pages_text), 5)
                        self.gui_log(f"Running OCR on pages 1-{max_ocr_pages}...")
                        images = convert_from_path(str(pdf_path), first_page=1, last_page=max_ocr_pages, poppler_path=poppler_path, dpi=300)
                        
                        if images:
                            # Combine OCR text from all pages
                            all_ocr_text = ""
                            for page_num, img in enumerate(images, 1):
                                page_ocr = pytesseract.image_to_string(img)
                                all_ocr_text += f"\n[PAGE {page_num} OCR]\n{page_ocr}\n"
                            full_text += all_ocr_text
                            self.gui_log(f"OCR extraction completed from {len(images)} page(s)")
                    except Exception as ocr_error:
                        self.gui_log(f"OCR not available or failed: {ocr_error}")
                        self.gui_log("Tip: Install Tesseract OCR for scanned PDFs (https://github.com/tesseract-ocr/tesseract)")
                
                # Search for claim number patterns
                claim_number = None
                lines = full_text.split('\n')
                
                # Common claim number patterns:
                # - "Claim #: 123456789" or "Claim Number: 123456789"
                # - "Ref #: 123456789" or "Reference: 123456789"
                # - Alphanumeric patterns like "CLM-123456" or "C123456789"
                
                claim_patterns = [
                    r'(?:claim|ref|reference)\s*(?:#|number|no|num)?\s*:?\s*([A-Z0-9\-]+)',
                    r'(?:claim|ref|reference)\s*(?:#|number|no|num)?\s*:?\s*(\d+)',
                    r'CLM[-_]?(\d+)',
                    r'claim\s*#?\s*(\d{8,12})',  # 8-12 digit claim numbers
                ]
                
                for line in lines:
                    line_lower = line.lower()
                    for pattern in claim_patterns:
                        match = re.search(pattern, line_lower, re.IGNORECASE)
                        if match:
                            potential_claim = match.group(1).strip()
                            # Validate: should be alphanumeric and reasonable length
                            if len(potential_claim) >= 6 and len(potential_claim) <= 20:
                                claim_number = potential_claim.upper()
                                self.gui_log(f"Found potential claim number: {claim_number}")
                                break
                    if claim_number:
                        break
                
                result = {
                    'claim_number': claim_number,
                    'raw_text': full_text[:2000]  # First 2000 chars for debugging
                }
                
                self.gui_log(f"Extracted from PDF:")
                self.gui_log(f"   Claim Number: {claim_number or 'Not found'}")
                
                return result
                
        except Exception as e:
            self.gui_log(f"Error extracting claim number from PDF: {e}")
            import traceback
            self.gui_log(f"   Traceback: {traceback.format_exc()}")
            return None
    
    def extract_payer_claim_control(self, pdf_path, target_client_name):
        """Extract Payer Claim Control # from PDF for a specific client
        
        Args:
            pdf_path: Path to the PDF file
            target_client_name: Name of the client (e.g., "Vernell E Luckey")
        
        Returns:
            dict with 'payer_claim_control' (the number) and 'matched' (bool), or None on error
        """
        self.gui_log(f"Extracting Payer Claim Control # from PDF for client: {target_client_name}")
        
        if not PDFPLUMBER_AVAILABLE:
            self.gui_log("pdfplumber not available - cannot extract from PDF", level="ERROR")
            return None
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Extract text from all pages
                full_text = ""
                for page_num, page in enumerate(pdf.pages, 1):
                    page_text = page.extract_text() or ""
                    full_text += page_text + "\n"
                
                # If no text, try OCR if available
                if not full_text.strip() and OCR_AVAILABLE:
                    self.gui_log("PDF appears to be scanned - attempting OCR...")
                    try:
                        from pdf2image import convert_from_path
                        poppler_path = os.environ.get('POPPLER_PATH')
                        if not poppler_path:
                            poppler_path = r"C:\Users\mthompson\Downloads\Release-25.07.0-0\poppler-25.07.0\Library\bin"
                        
                        images = convert_from_path(str(pdf_path), poppler_path=poppler_path, dpi=300)
                        for img in images:
                            page_ocr = pytesseract.image_to_string(img)
                            full_text += f"\n{page_ocr}\n"
                        self.gui_log("OCR extraction completed")
                    except Exception as ocr_error:
                        self.gui_log(f"OCR failed: {ocr_error}")
                
                lines = [line.strip() for line in full_text.split('\n') if line.strip()]
                
                # Normalize target client name
                target_parts = target_client_name.lower().strip().split()
                target_first = target_parts[0] if target_parts else None
                target_last = target_parts[-1] if len(target_parts) > 1 else target_parts[0] if target_parts else None
                
                self.gui_log(f"Searching for client: {target_client_name} (first: {target_first}, last: {target_last})", level="DEBUG")
                
                # STEP 1: Find all "Payer Claim Control #:" entries
                payer_claim_control_pattern = r'(?:payer\s*claim\s*control\s*#?\s*:?\s*)(.+)'
                all_payer_controls = []
                
                for i, line in enumerate(lines):
                    if 'payer claim control' in line.lower() and ':' in line:
                        match = re.search(payer_claim_control_pattern, line, re.IGNORECASE)
                        if match:
                            control_number = match.group(1).strip()
                            # Clean up - keep numbers, spaces, dashes, dots
                            control_number = re.sub(r'[^\d\s\-\.]+.*$', '', control_number).strip()
                            all_payer_controls.append({
                                'line_num': i + 1,
                                'line': line,
                                'control_number': control_number
                            })
                            self.gui_log(f"Found 'Payer Claim Control #:' on line {i+1}: {control_number}", level="DEBUG")
                
                if not all_payer_controls:
                    self.gui_log("No 'Payer Claim Control #:' entries found in PDF", level="WARNING")
                    return {
                        'payer_claim_control': None,
                        'matched': False,
                        'error': 'No Payer Claim Control # found in PDF'
                    }
                
                # STEP 2: Find patient ID and claim for target client
                patient_id_pattern = r'K(\d{10,})'
                claim_pattern = r'claim\s*#?\s*(\d+)'
                
                target_patient_id = None
                target_claim = None
                
                # Search for target client's last name
                for i, line in enumerate(lines):
                    line_lower = line.lower()
                    if target_last and target_last in line_lower:
                        # Look for patient ID nearby
                        context_start = max(0, i - 5)
                        context_end = min(len(lines), i + 6)
                        
                        # Collect patient IDs in context
                        candidate_ids = []
                        for j in range(context_start, context_end):
                            patient_match = re.search(patient_id_pattern, lines[j], re.IGNORECASE)
                            if patient_match:
                                potential_id = patient_match.group(1)
                                distance = abs(j - i)
                                is_direct_before = (j == i - 1)
                                candidate_ids.append({
                                    'id': potential_id,
                                    'distance': distance,
                                    'is_direct_before': is_direct_before
                                })
                        
                        if candidate_ids:
                            # Sort: prioritize direct before, then closest
                            candidate_ids.sort(key=lambda x: (not x['is_direct_before'], x['distance']))
                            
                            # Verify first name in context
                            context_text = ' '.join(lines[context_start:context_end]).lower()
                            first_name_found = False
                            if target_first:
                                # Handle corruption (e.g., "VerCnoedlle" for "Vernell")
                                if (target_first in context_text or 
                                    'vernell' in context_text or
                                    ('ver' in context_text and ('cnoedlle' in context_text or 'nel' in context_text))):
                                    first_name_found = True
                            
                            # Select best candidate
                            for candidate in candidate_ids:
                                if candidate['is_direct_before'] or first_name_found or not target_first:
                                    target_patient_id = candidate['id']
                                    self.gui_log(f"Found Patient ID: K{target_patient_id}", level="DEBUG")
                                    break
                        
                        if target_patient_id:
                            break
                
                # Find claim number
                if target_patient_id:
                    for i, line in enumerate(lines):
                        if target_patient_id and f"K{target_patient_id}" in line:
                            # Look for claim number nearby
                            context_start = max(0, i - 10)
                            context_end = min(len(lines), i + 10)
                            
                            for j in range(context_start, context_end):
                                claim_match = re.search(claim_pattern, lines[j], re.IGNORECASE)
                                if claim_match:
                                    target_claim = claim_match.group(1)
                                    self.gui_log(f"Found Claim #: {target_claim}", level="DEBUG")
                                    break
                            
                            if target_claim:
                                break
                
                # STEP 3: Find Payer Claim Control # associated with this claim/client
                target_payer_claim_control = None
                
                # Strategy: Look for "Payer Claim Control #:" after "Claim Totals" (common pattern)
                for i, line in enumerate(lines):
                    line_lower = line.lower()
                    
                    # Check if this line contains "Claim Totals"
                    if 'claim totals' in line_lower:
                        self.gui_log(f"Found 'Claim Totals' on line {i+1}, searching for 'Payer Claim Control #:' after it...", level="DEBUG")
                        
                        # Look for "Payer Claim Control #:" in the next 10 lines after "Claim Totals"
                        for j in range(i + 1, min(i + 11, len(lines))):
                            next_line = lines[j]
                            next_line_lower = next_line.lower()
                            
                            if 'payer claim control' in next_line_lower and ':' in next_line:
                                match = re.search(payer_claim_control_pattern, next_line, re.IGNORECASE)
                                if match:
                                    control_number = match.group(1).strip()
                                    control_number = re.sub(r'[^\d\s\-\.]+.*$', '', control_number).strip()
                                    target_payer_claim_control = control_number
                                    self.gui_log(f"✅ Found Payer Claim Control # after 'Claim Totals': {target_payer_claim_control}", level="INFO")
                                    break
                        
                        if target_payer_claim_control:
                            break
                
                # If not found after "Claim Totals", try other strategies
                if not target_payer_claim_control and target_claim:
                    # Find claim section
                    claim_start_line = None
                    for i, line in enumerate(lines):
                        if f"Claim #{target_claim}" in line or f"Claim # {target_claim}" in line:
                            claim_start_line = i + 1
                            break
                    
                    if claim_start_line:
                        # Search for Payer Claim Control # near the claim
                        target_line = None
                        for i, line in enumerate(lines):
                            if target_last and target_last in line.lower():
                                target_line = i + 1
                                break
                        
                        search_start = max(0, claim_start_line - 10)
                        search_end = min(len(lines), (target_line + 5) if target_line else claim_start_line + 30)
                        
                        for i in range(search_start - 1, search_end):
                            line_lower = lines[i].lower()
                            if 'payer claim control' in line_lower and ':' in lines[i]:
                                match = re.search(payer_claim_control_pattern, lines[i], re.IGNORECASE)
                                if match:
                                    control_number = match.group(1).strip()
                                    control_number = re.sub(r'[^\d\s\-\.]+.*$', '', control_number).strip()
                                    target_payer_claim_control = control_number
                                    self.gui_log(f"✅ Found Payer Claim Control # near claim: {target_payer_claim_control}", level="INFO")
                                    break
                
                # If not found near claim, try searching near patient ID
                if not target_payer_claim_control and target_patient_id:
                    for i, line in enumerate(lines):
                        if target_patient_id and f"K{target_patient_id}" in line:
                            patient_id_line = i + 1
                            search_start = max(0, patient_id_line - 20)
                            search_end = min(len(lines), patient_id_line + 20)
                            
                            for j in range(search_start - 1, search_end):
                                line_lower = lines[j].lower()
                                if 'payer claim control' in line_lower and ':' in lines[j]:
                                    match = re.search(payer_claim_control_pattern, lines[j], re.IGNORECASE)
                                    if match:
                                        control_number = match.group(1).strip()
                                        control_number = re.sub(r'[^\d\s\-\.]+.*$', '', control_number).strip()
                                        target_payer_claim_control = control_number
                                        self.gui_log(f"✅ Found Payer Claim Control #: {target_payer_claim_control}", level="INFO")
                                        break
                            
                            if target_payer_claim_control:
                                break
                
                # Return result
                if target_payer_claim_control:
                    self.gui_log(f"✅ Successfully extracted Payer Claim Control # for {target_client_name}: {target_payer_claim_control}", level="INFO")
                    return {
                        'payer_claim_control': target_payer_claim_control,
                        'client_name': target_client_name,
                        'patient_id': target_patient_id,
                        'claim_number': target_claim,
                        'matched': True
                    }
                else:
                    self.gui_log(f"⚠️ Payer Claim Control # not found for {target_client_name}", level="WARNING")
                    self.gui_log(f"   Found {len(all_payer_controls)} total 'Payer Claim Control #:' entries in PDF", level="WARNING")
                    return {
                        'payer_claim_control': None,
                        'client_name': target_client_name,
                        'matched': False,
                        'error': 'Payer Claim Control # not found for this client'
                    }
        
        except Exception as e:
            self.log_error(f"Error extracting Payer Claim Control # from PDF: {e}")
            return None
    
    def read_excel_csv(self, file_path):
        """Read Excel or CSV file and extract client data (Name, DOB, Date of Service)
        
        Uses manual column mapping if enabled, otherwise uses automatic detection.
        """
        if not EXCEL_AVAILABLE:
            self.gui_log("pandas not available - cannot read Excel/CSV files")
            return []
        
        self.gui_log(f"Reading Excel/CSV file: {file_path}")
        if self.use_manual_mapping:
            self.gui_log("Using manual column mapping", level="INFO")
        else:
            self.gui_log("Using automatic column detection", level="INFO")
        
        client_data = []
        
        try:
            file_path_obj = Path(file_path)
            
            # Determine file type and read accordingly
            if file_path_obj.suffix.lower() == '.csv':
                df = pd.read_csv(file_path)
                # CSV files don't have filters, so we read all rows
                filtered_rows = None
            elif file_path_obj.suffix.lower() in ['.xlsx', '.xls']:
                # For Excel files, we need to respect filters
                # Read the file with openpyxl to check for hidden/filtered rows
                filtered_rows = None
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, data_only=True, keep_links=False)
                    ws = wb.active
                    
                    # Check if auto_filter is enabled (indicating filters may be applied)
                    has_auto_filter = ws.auto_filter is not None
                    
                    if has_auto_filter:
                        self.gui_log("Excel file has auto-filter enabled - Checking for visible rows...", level="INFO")
                        
                        # Get visible (non-hidden) rows
                        # Note: Excel filters don't actually "hide" rows in the file structure,
                        # but we can check for manually hidden rows and get all rows if filter is active
                        # The user's filters are stored in the file, but we need to respect them
                        
                        # Strategy: Check which rows are NOT hidden
                        # For filtered Excel files, all non-hidden rows should be visible
                        visible_data_rows = []
                        for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                            # Check if row is hidden
                            row_dim = ws.row_dimensions[row_idx]
                            if not row_dim.hidden:
                                visible_data_rows.append(row_idx)
                        
                        if visible_data_rows:
                            # Convert to 0-based pandas index
                            # Row 2 in Excel = index 0 in pandas (row 1 is header, row 2 is first data)
                            filtered_rows = [r - 2 for r in visible_data_rows]  # Subtract 2: one for header, one for 0-based
                            total_data_rows = ws.max_row - 1  # Exclude header
                            self.gui_log(f"Respecting Excel filters - Processing {len(visible_data_rows)} visible row(s) out of {total_data_rows} total data rows", level="INFO")
                        else:
                            self.gui_log("Warning: No visible rows found (all rows may be filtered or hidden)", level="WARNING")
                    else:
                        self.gui_log("Excel file has no filters applied - Processing all rows", level="INFO")
                    
                    wb.close()
                except ImportError:
                    self.gui_log("openpyxl not available - Cannot check Excel filters. Processing all rows.", level="WARNING")
                    filtered_rows = None
                except Exception as e:
                    self.gui_log(f"Could not check Excel filters: {e}. Processing all rows.", level="WARNING")
                    filtered_rows = None
                
                # Read Excel file with pandas
                df = pd.read_excel(file_path)
                
                # If we have filtered rows, keep only those
                if filtered_rows is not None and len(filtered_rows) > 0:
                    original_count = len(df)
                    # Keep only visible rows
                    df = df.iloc[filtered_rows].reset_index(drop=True).copy()
                    self.gui_log(f"✅ Respecting Excel filters: Processing {len(df)} visible row(s) out of {original_count} total rows", level="INFO")
                elif filtered_rows is not None and len(filtered_rows) == 0:
                    self.gui_log("⚠️ Warning: All rows are filtered out in Excel. No data to process.", level="WARNING")
                    return []
            else:
                self.gui_log(f"Unsupported file type: {file_path_obj.suffix}")
                return []
            
            self.gui_log(f"File loaded: {len(df)} rows, {len(df.columns)} columns")
            self.gui_log(f"Columns: {list(df.columns)}", level="DEBUG")
            
            # Initialize column variables
            name_col = None
            last_name_col = None
            first_name_col = None
            dob_col = None
            dos_col = None  # Date of Service
            insurance_col = None  # Insurance provider
            claim_status_col = None  # Insurance Balance Status (how claim was originally submitted)
            
            if self.use_manual_mapping:
                # Use manual column mapping
                self.gui_log("Applying manual column mapping...", level="DEBUG")
                
                # Helper function to convert column letter/name to actual column
                def find_column(col_spec):
                    """Find column by letter (e.g., 'C') or by name"""
                    if not col_spec:
                        return None
                    
                    # Try as column letter first (convert to index)
                    if len(col_spec) == 1 and col_spec.isalpha():
                        try:
                            # Convert column letter to 0-based index (A=0, B=1, C=2, etc.)
                            col_idx = ord(col_spec.upper()) - ord('A')
                            if 0 <= col_idx < len(df.columns):
                                return df.columns[col_idx]
                        except Exception as e:
                            self.gui_log(f"Could not convert column letter '{col_spec}': {e}", level="DEBUG")
                    
                    # Try as column name (exact match)
                    for col in df.columns:
                        if str(col).upper() == col_spec.upper():
                            return col
                    
                    # Try as partial name match
                    for col in df.columns:
                        if col_spec.upper() in str(col).upper():
                            return col
                    
                    return None
                
                # Get manual mappings
                last_name_spec = self.manual_column_map.get('last_name_col', '')
                first_name_spec = self.manual_column_map.get('first_name_col', '')
                full_name_spec = self.manual_column_map.get('name_col', '')
                dob_spec = self.manual_column_map.get('dob_col', '')
                dos_spec = self.manual_column_map.get('dos_col', '')
                
                # Find columns using manual mapping
                if last_name_spec and first_name_spec:
                    last_name_col = find_column(last_name_spec)
                    first_name_col = find_column(first_name_spec)
                    if last_name_col and first_name_col:
                        name_col = (last_name_col, first_name_col)
                        self.gui_log(f"Manual mapping - Last Name: {last_name_col}, First Name: {first_name_col}", level="DEBUG")
                    elif last_name_col or first_name_col:
                        self.gui_log(f"Warning: Found only partial name columns - Last: {last_name_col}, First: {first_name_col}", level="WARNING")
                elif full_name_spec:
                    name_col = find_column(full_name_spec)
                    self.gui_log(f"Manual mapping - Full Name: {name_col}", level="DEBUG")
                
                if dob_spec:
                    dob_col = find_column(dob_spec)
                    self.gui_log(f"Manual mapping - DOB: {dob_col}", level="DEBUG")
                
                if dos_spec:
                    dos_col = find_column(dos_spec)
                    self.gui_log(f"Manual mapping - DOS: {dos_col}", level="DEBUG")
                
                # Verify all required columns found
                if not name_col:
                    self.gui_log("Warning: Name column not found with manual mapping", level="WARNING")
                if not dob_col and dob_spec:
                    self.gui_log(f"Warning: DOB column '{dob_spec}' not found with manual mapping", level="WARNING")
                if not dos_col and dos_spec:
                    self.gui_log(f"Warning: DOS column '{dos_spec}' not found with manual mapping", level="WARNING")
            
            else:
                # Automatic column detection (existing logic)
                self.gui_log("Auto-detecting columns...", level="DEBUG")
                
                # Search for name columns
                name_keywords = ['last name', 'first name', 'patient name', 'client name', 'name']
                for col in df.columns:
                    col_lower = str(col).lower()
                    if any(keyword in col_lower for keyword in name_keywords):
                        if 'last' in col_lower or ('first' in col_lower and name_col is None):
                            # If we find both first and last, combine them
                            if 'last' in col_lower:
                                # Look for first name column
                                for other_col in df.columns:
                                    if 'first' in str(other_col).lower() and other_col != col:
                                        name_col = (col, other_col)  # Tuple for combined names
                                        break
                            if name_col is None:
                                name_col = col
                        elif name_col is None:
                            name_col = col
                
                # Search for DOB column
                dob_keywords = ['dob', 'date of birth', 'birth date', 'birthdate']
                for col in df.columns:
                    col_lower = str(col).lower()
                    if any(keyword in col_lower for keyword in dob_keywords):
                        dob_col = col
                        break
                
                # Search for Date of Service column
                dos_keywords = ['date', 'service date', 'date of service', 'dos', 'appointment date']
                for col in df.columns:
                    col_lower = str(col).lower()
                    if any(keyword in col_lower for keyword in dos_keywords):
                        dos_col = col
                        break
                
                # Search for Insurance column
                # First check column U (index 20) - user specified this is where insurance is typically located
                if len(df.columns) > 20:
                    # Column U is index 20 (A=0, B=1, ..., U=20)
                    potential_insurance_col = df.columns[20]
                    col_lower = str(potential_insurance_col).lower()
                    # Check if this column has insurance-related keywords or contains actual insurance data
                    insurance_keywords = ['insurance', 'provider', 'payer', 'insurance provider', 'insurance name', 'carrier']
                    if any(keyword in col_lower for keyword in insurance_keywords):
                        insurance_col = potential_insurance_col
                        self.gui_log(f"Found insurance column at Column U (index 20): {insurance_col}", level="INFO")
                
                # If column U wasn't identified as insurance, search for keywords
                if insurance_col is None:
                    insurance_keywords = ['insurance provider', 'insurance', 'payer', 'carrier', 'insurance name', 'insurance company']
                    for col in df.columns:
                        col_lower = str(col).lower()
                        if any(keyword in col_lower for keyword in insurance_keywords):
                            insurance_col = col
                            self.gui_log(f"Found insurance column by keyword: {insurance_col}", level="INFO")
                            break
                    
                    # If still not found but column U exists, try using it anyway (user said it's usually there)
                    if insurance_col is None and len(df.columns) > 20:
                        insurance_col = df.columns[20]
                        self.gui_log(f"Using Column U (index 20) as insurance column: {insurance_col}", level="INFO")
                
                # Search for Insurance Balance Status column (claim submission status)
                # First check column AK (index 36) - user specified this is where it's typically located
                if len(df.columns) > 36:
                    # Column AK is index 36 (A=0, B=1, ..., Z=25, AA=26, AB=27, ..., AK=36)
                    potential_claim_status_col = df.columns[36]
                    col_lower = str(potential_claim_status_col).lower()
                    # Check if this column has insurance balance status keywords
                    claim_status_keywords = ['insurance balance status', 'insurance balance', 'claim status', 'submission status', 'balance status']
                    if any(keyword in col_lower for keyword in claim_status_keywords):
                        claim_status_col = potential_claim_status_col
                        self.gui_log(f"Found Insurance Balance Status column at Column AK (index 36): {claim_status_col}", level="INFO")
                
                # If column AK wasn't identified, search for keywords
                if claim_status_col is None:
                    claim_status_keywords = ['insurance balance status', 'insurance balance', 'claim status', 'submission status', 'balance status', 'submitted', 'submission']
                    for col in df.columns:
                        col_lower = str(col).lower()
                        if any(keyword in col_lower for keyword in claim_status_keywords):
                            claim_status_col = col
                            self.gui_log(f"Found Insurance Balance Status column by keyword: {claim_status_col}", level="INFO")
                            break
                    
                    # If still not found but column AK exists, try using it anyway (user said it's usually there)
                    if claim_status_col is None and len(df.columns) > 36:
                        claim_status_col = df.columns[36]
                        self.gui_log(f"Using Column AK (index 36) as Insurance Balance Status column: {claim_status_col}", level="INFO")
            
            self.gui_log(f"Final column mapping - Name: {name_col}, DOB: {dob_col}, DOS: {dos_col}, Insurance: {insurance_col}, Claim Status: {claim_status_col}")
            
            # Extract data row by row
            skipped_rows = 0
            # Track original Excel row numbers (for filtered files)
            original_row_map = {}  # Maps visible DataFrame index to original Excel row number
            if filtered_rows is not None and file_path_obj.suffix.lower() in ['.xlsx', '.xls']:
                # Rebuild the mapping from visible rows back to original Excel rows
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, data_only=True, keep_links=False)
                    ws = wb.active
                    visible_excel_rows = []
                    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                        row_dim = ws.row_dimensions[row_idx]
                        if not row_dim.hidden:
                            visible_excel_rows.append(row_idx)
                    wb.close()
                    
                    # Map DataFrame index (0-based) to Excel row number
                    for df_idx, excel_row in enumerate(visible_excel_rows):
                        original_row_map[df_idx] = excel_row
                except:
                    # If mapping fails, we'll calculate based on filtered_rows
                    # filtered_rows contains 0-based indices, so we add 2 (header + 0-based to 1-based)
                    for df_idx in range(len(df)):
                        # Try to reconstruct from filtered_rows if available
                        if df_idx < len(filtered_rows):
                            original_row_map[df_idx] = filtered_rows[df_idx] + 2  # +2 for header and 0-based to 1-based
            
            for idx, row in df.iterrows():
                client_name = None
                dob = None
                date_of_service = None
                insurance = None
                claim_status = None  # Insurance Balance Status (how claim was originally submitted)
                
                # Get original Excel row number (for testing and reference)
                # idx is the DataFrame index (0-based), excel_row is the actual Excel row number (1-based)
                excel_row = original_row_map.get(idx, idx + 2)  # Default: idx + 2 (DataFrame index + header + 0-to-1-based)
                
                # Get client name
                if name_col:
                    if isinstance(name_col, tuple):
                        # Combine first and last name
                        first = str(row.get(name_col[1], '')).strip() if len(name_col) > 1 else ''
                        last = str(row.get(name_col[0], '')).strip()
                        client_name = f"{first} {last}".strip() if first else last
                        
                        # Check if this row should be skipped (counselor entry, not a client)
                        # Skip if "New Referrals" appears in first name or last name columns
                        if 'new referrals' in first.lower() or 'new referrals' in last.lower():
                            skipped_rows += 1
                            self.gui_log(f"Skipping row {idx + 2} (Excel row {excel_row}): Contains 'New Referrals' - This is a counselor entry, not a client", level="INFO")
                            continue
                    else:
                        client_name = str(row.get(name_col, '')).strip()
                        
                        # Check if "New Referrals" appears in the name
                        if 'new referrals' in client_name.lower():
                            skipped_rows += 1
                            self.gui_log(f"Skipping row {idx + 2} (Excel row {excel_row}): Contains 'New Referrals' - This is a counselor entry, not a client", level="INFO")
                            continue
                else:
                    # If no name column found, skip this row
                    skipped_rows += 1
                    self.gui_log(f"Skipping row {idx + 2} (Excel row {excel_row}): No name column found", level="DEBUG")
                    continue
                
                # Get DOB
                if dob_col:
                    dob_value = row.get(dob_col)
                    if pd.notna(dob_value):
                        dob = str(dob_value).strip()
                
                # Get Date of Service
                if dos_col:
                    dos_value = row.get(dos_col)
                    if pd.notna(dos_value):
                        date_of_service = str(dos_value).strip()
                
                # Get Insurance
                if insurance_col:
                    insurance_value = row.get(insurance_col)
                    if pd.notna(insurance_value):
                        insurance = str(insurance_value).strip()
                        # Clean up the insurance value if needed
                        if insurance and insurance.lower() in ['nan', 'none', '']:
                            insurance = None
                
                # Get Insurance Balance Status (claim submission status)
                if claim_status_col:
                    claim_status_value = row.get(claim_status_col)
                    if pd.notna(claim_status_value):
                        claim_status = str(claim_status_value).strip()
                        # Clean up the claim status value if needed
                        if claim_status and claim_status.lower() in ['nan', 'none', '']:
                            claim_status = None
                
                # Only add if we have at least a name
                if client_name and client_name not in ['', 'nan', 'None']:
                    client_data.append({
                        'client_name': client_name,
                        'dob': dob,
                        'date_of_service': date_of_service,
                        'insurance': insurance,  # Add insurance to client data
                        'claim_status': claim_status,  # Add claim submission status
                        'excel_row': excel_row  # Store original Excel row number
                    })
            
            if skipped_rows > 0:
                self.gui_log(f"Skipped {skipped_rows} row(s) containing 'New Referrals' (counselor entries)", level="INFO")
            
            self.gui_log(f"Extracted {len(client_data)} client record(s) from file")
            return client_data
            
        except Exception as e:
            self.gui_log(f"Error reading Excel/CSV file: {e}")
            import traceback
            self.gui_log(f"   Traceback: {traceback.format_exc()}")
            return []
    
    def _update_user_dropdown(self):
        """Update the user dropdown with current users"""
        if hasattr(self, 'user_dropdown'):
            user_names = sorted(self.users.keys())
            if user_names:
                self.user_dropdown['values'] = user_names
            else:
                self.user_dropdown['values'] = []
    
    def _on_user_selected(self, event=None):
        """Handle user selection from dropdown"""
        selected_user = self.user_dropdown.get()
        if selected_user and selected_user in self.users:
            user_data = self.users[selected_user]
            username = user_data.get('username', '')
            password = self._retrieve_secure_password(username) or user_data.get('password', '')
            if password:
                self.users[selected_user]['password'] = password
            else:
                self.gui_log(
                    f"No stored password found for user '{selected_user}'. Please enter it manually.",
                    level="WARNING",
                    include_context=False
                )

            self.username_entry.delete(0, tk.END)
            self.username_entry.insert(0, username)
            self.password_entry.delete(0, tk.END)
            self.password_entry.insert(0, password)
            self._enable_login()
            self.gui_log(f"Loaded credentials for user: {selected_user}")
    
    def _add_user(self):
        """Add a new user to saved users"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Add New User - Version 3.1.0, Last Updated 12/04/2025")
        dialog.geometry("400x200")
        dialog.configure(bg="#f0f0f0")
        dialog.transient(self.root)
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        tk.Label(dialog, text="User Name:", font=("Arial", 10), bg="#f0f0f0").pack(pady=(20, 5))
        name_entry = tk.Entry(dialog, font=("Arial", 10), width=30)
        name_entry.pack(pady=(0, 10))
        name_entry.focus()
        
        tk.Label(dialog, text="Username:", font=("Arial", 10), bg="#f0f0f0").pack(pady=(0, 5))
        username_entry = tk.Entry(dialog, font=("Arial", 10), width=30)
        username_entry.pack(pady=(0, 10))
        
        tk.Label(dialog, text="Password:", font=("Arial", 10), bg="#f0f0f0").pack(pady=(0, 5))
        password_entry = tk.Entry(dialog, font=("Arial", 10), width=30, show="*")
        password_entry.pack(pady=(0, 15))
        
        def save_user():
            name = name_entry.get().strip()
            username = username_entry.get().strip()
            password = password_entry.get().strip()
            
            if not name or not username or not password:
                messagebox.showwarning("Invalid Input", "All fields are required")
                return
            
            if name in self.users:
                if not messagebox.askyesno("User Exists", f"User '{name}' already exists. Overwrite?"):
                    return
            
            self._store_secure_password(username, password)
            self.users[name] = {
                'username': username,
                'password': password
            }
            self.save_users()
            self._update_user_dropdown()
            self.gui_log(f"Added user: {name}")
            dialog.destroy()
            messagebox.showinfo("Success", f"User '{name}' added successfully")
        
        button_frame = tk.Frame(dialog, bg="#f0f0f0")
        button_frame.pack(pady=10)
        
        tk.Button(button_frame, text="Save", command=save_user,
                  bg="#800020", fg="white", font=("Arial", 10),
                  padx=15, pady=5, cursor="hand2", relief="flat", bd=0).pack(side="left", padx=5)
        tk.Button(button_frame, text="Cancel", command=dialog.destroy,
                  bg="#666666", fg="white", font=("Arial", 10),
                  padx=15, pady=5, cursor="hand2", relief="flat", bd=0).pack(side="left", padx=5)
        
        # Bind Enter key to save
        name_entry.bind("<Return>", lambda e: username_entry.focus())
        username_entry.bind("<Return>", lambda e: password_entry.focus())
        password_entry.bind("<Return>", lambda e: save_user())
    
    def _update_credentials(self):
        """Update credentials for selected user"""
        selected_user = self.user_dropdown.get()
        if not selected_user or selected_user not in self.users:
            messagebox.showwarning("No User Selected", "Please select a user from the dropdown")
            return
        
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        
        if not username or not password:
            messagebox.showwarning("Invalid Input", "Username and password are required")
            return
        
        existing_username = self.users[selected_user].get('username') if selected_user in self.users else None
        if existing_username and existing_username != username:
            self._delete_secure_password(existing_username)

        self._store_secure_password(username, password)

        self.users[selected_user] = {
            'username': username,
            'password': password
        }
        self.save_users()
        self.gui_log(f"Updated credentials for user: {selected_user}")
        messagebox.showinfo("Success", f"Credentials updated for '{selected_user}'")
    
    def _enable_login(self):
        """Enable login button if credentials are entered"""
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        if username and password:
            self.login_button.config(state="normal")
        else:
            self.login_button.config(state="disabled")
    
    def _browse_pdf(self):
        """Browse for PDF file"""
        file_path = filedialog.askopenfilename(
            title="Select PDF File",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if file_path:
            self.selected_pdf_path = file_path
            self.pdf_entry.config(state="normal")
            self.pdf_entry.delete(0, tk.END)
            self.pdf_entry.insert(0, file_path)
            self.pdf_entry.config(state="readonly")
            self.pdf_status_label.config(text=f"Selected: {Path(file_path).name}")
            self.gui_log(f"Selected PDF: {Path(file_path).name}")
    
    def _clear_pdf(self):
        """Clear selected PDF"""
        self.selected_pdf_path = None
        self.pdf_entry.config(state="normal")
        self.pdf_entry.delete(0, tk.END)
        self.pdf_entry.config(state="readonly")
        self.pdf_status_label.config(text="No PDF selected")
        self.gui_log("Cleared PDF selection")
    
    def _browse_excel(self):
        """Browse for Excel/CSV file"""
        file_path = filedialog.askopenfilename(
            title="Select Excel/CSV File",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            self._load_excel_file(file_path)
    
    def _paste_excel_path(self):
        """Paste Excel/CSV file path from clipboard"""
        try:
            clipboard_content = self.root.clipboard_get()
            if clipboard_content and (clipboard_content.endswith(('.xlsx', '.xls', '.csv')) or Path(clipboard_content).exists()):
                self._load_excel_file(clipboard_content)
            else:
                messagebox.showwarning("Invalid Path", "Clipboard does not contain a valid Excel/CSV file path")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to get clipboard content: {e}")
    
    def _load_excel_file(self, file_path):
        """Load Excel/CSV file and extract client data"""
        try:
            self.gui_log(f"Loading Excel/CSV file: {Path(file_path).name}", level="DEBUG")
            
            self.selected_excel_path = file_path
            self.excel_entry.config(state="normal")
            self.excel_entry.delete(0, tk.END)
            self.excel_entry.insert(0, file_path)
            self.excel_entry.config(state="readonly")
            
            # Read the file
            self.gui_log("Reading Excel/CSV file...", level="DEBUG")
            self.excel_client_data = self.read_excel_csv(file_path)
            
            if self.excel_client_data:
                self.excel_status_label.config(text=f"Loaded: {len(self.excel_client_data)} client(s)")
                self.gui_log(f"Successfully loaded {len(self.excel_client_data)} client record(s) from Excel/CSV")
                # Show preview
                preview_text = "Preview:\n"
                for i, client in enumerate(self.excel_client_data[:5], 1):  # Show first 5
                    preview_text += f"  {i}. {client.get('client_name', 'N/A')} - DOB: {client.get('dob', 'N/A')} - DOS: {client.get('date_of_service', 'N/A')}\n"
                if len(self.excel_client_data) > 5:
                    preview_text += f"  ... and {len(self.excel_client_data) - 5} more"
                messagebox.showinfo("File Loaded", preview_text)
                
                # Enable processing button if logged in
                self._enable_processing_button()
            else:
                self.excel_status_label.config(text="No valid client data found")
                self.gui_log("No valid client data found in Excel/CSV file", level="WARNING")
        
        except Exception as e:
            self.log_error(f"Error loading Excel/CSV file: {e}")
            messagebox.showerror("Error", f"Failed to load Excel/CSV file:\n{e}")
            self.excel_status_label.config(text="Error loading file")
    
    def _toggle_manual_mapping(self):
        """Toggle manual column mapping fields visibility"""
        if self.use_manual_mapping_var.get() == 1:
            self.manual_mapping_frame.pack(fill="x", pady=(0, 5))
            self.gui_log("Manual column mapping enabled - specify columns manually", level="INFO")
        else:
            self.manual_mapping_frame.pack_forget()
            self.use_manual_mapping = False
            self.gui_log("Manual column mapping disabled - using automatic detection", level="INFO")
    
    def _toggle_processing_range(self):
        """Enable or disable processing row range inputs based on checkbox state."""
        if not hasattr(self, 'processing_start_row_entry') or not hasattr(self, 'processing_end_row_entry'):
            return
        
        if self.run_all_rows_var.get() == 1:
            state = "disabled"
            self.processing_start_row_entry.delete(0, tk.END)
            self.processing_end_row_entry.delete(0, tk.END)
        else:
            state = "normal"
        
        self.processing_start_row_entry.config(state=state)
        self.processing_end_row_entry.config(state=state)
        
        if state == "normal":
            self.processing_start_row_entry.focus_set()
    
    def _apply_manual_mapping(self):
        """Apply manual column mapping settings"""
        # Get values from entry fields
        last_name_col = self.last_name_col_entry.get().strip()
        first_name_col = self.first_name_col_entry.get().strip()
        full_name_col = self.full_name_col_entry.get().strip()
        dob_col = self.dob_col_entry.get().strip()
        dos_col = self.dos_col_entry.get().strip()
        
        # Store mapping
        self.manual_column_map = {
            'last_name_col': last_name_col.upper() if last_name_col else '',
            'first_name_col': first_name_col.upper() if first_name_col else '',
            'name_col': full_name_col.upper() if full_name_col else '',
            'dob_col': dob_col.upper() if dob_col else '',
            'dos_col': dos_col.upper() if dos_col else ''
        }
        
        self.use_manual_mapping = True
        self.gui_log("Manual column mapping applied", level="INFO")
        self.gui_log(f"  Last Name: {self.manual_column_map['last_name_col'] or 'Not set'}", level="DEBUG")
        self.gui_log(f"  First Name: {self.manual_column_map['first_name_col'] or 'Not set'}", level="DEBUG")
        self.gui_log(f"  Full Name: {self.manual_column_map['name_col'] or 'Not set'}", level="DEBUG")
        self.gui_log(f"  DOB: {self.manual_column_map['dob_col'] or 'Not set'}", level="DEBUG")
        self.gui_log(f"  DOS: {self.manual_column_map['dos_col'] or 'Not set'}", level="DEBUG")
        
        # If Excel/CSV is already loaded, reload it with new mapping
        if self.selected_excel_path:
            self.gui_log("Reloading Excel/CSV file with new column mapping...", level="INFO")
            self._load_excel_file(self.selected_excel_path)
        else:
            messagebox.showinfo("Mapping Applied", "Column mapping saved. Load an Excel/CSV file to apply the mapping.")
    
    def _clear_excel(self):
        """Clear selected Excel/CSV file"""
        self.selected_excel_path = None
        self.excel_client_data = []
        self.excel_entry.config(state="normal")
        self.excel_entry.delete(0, tk.END)
        self.excel_entry.config(state="readonly")
        self.excel_status_label.config(text="No Excel/CSV selected")
        self.gui_log("Cleared Excel/CSV selection")
    
    def _browse_output_excel(self):
        """Browse for output Excel file save location"""
        # Suggest a default filename based on input file or timestamp
        default_filename = "tn_refiling_output"
        if self.selected_excel_path:
            input_file = Path(self.selected_excel_path)
            default_filename = f"{input_file.stem}_output"
        
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        default_filename = f"{default_filename}_{timestamp}.xlsx"
        
        # Get directory from input file if available, otherwise use current directory
        default_dir = Path(self.selected_excel_path).parent if self.selected_excel_path else Path.cwd()
        default_path = default_dir / default_filename
        
        file_path = filedialog.asksaveasfilename(
            title="Select Output Excel File Location",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ],
            initialfile=default_filename,
            initialdir=str(default_dir)
        )
        if file_path:
            # Ensure .xlsx extension
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            
            self.output_excel_path = file_path
            self.output_excel_entry.config(state="normal")
            self.output_excel_entry.delete(0, tk.END)
            self.output_excel_entry.insert(0, file_path)
            self.output_excel_entry.config(state="readonly")
            self.output_excel_status_label.config(text=f"Selected: {Path(file_path).name}", fg="#28a745")
            self.gui_log(f"Output Excel path set: {Path(file_path).name}")
    
    def _clear_output_excel(self):
        """Clear selected output Excel path"""
        self.output_excel_path = None
        self.output_excel_entry.config(state="normal")
        self.output_excel_entry.delete(0, tk.END)
        self.output_excel_entry.config(state="readonly")
        self.output_excel_status_label.config(text="No output file selected (will use default location)", fg="#666666")
        self.gui_log("Cleared output Excel path - will use default location")
    
    def _request_stop(self):
        """Request bot to stop - called from GUI"""
        self.request_stop()
        # Optionally close browser when stop is requested
        if self.driver:
            self.gui_log("Stop requested - Browser will remain open for inspection", level="WARNING")
        # Show confirmation
        self.root.after(0, lambda: messagebox.showinfo("Stop Requested", 
                                                       "Bot stop requested. Current operation will complete, then bot will stop.\n\n"
                                                       "Browser will remain open for inspection."))
    
    def _close_window(self):
        """Close the main window"""
        if self._shutdown_requested:
            return

        self._shutdown_requested = True
        self.request_stop()

        self.gui_log(
            "Shutdown requested - waiting for background tasks to finish...",
            level="WARNING",
            include_context=False
        )

        # Disable interactive buttons to prevent additional actions during shutdown
        for widget_name in [
            'stop_button', 'start_processing_button', 'login_button',
            'test_navigation_button', 'complex_test_navigation_button', 'close_button'
        ]:
            widget = getattr(self, widget_name, None)
            if widget:
                try:
                    widget.config(state="disabled")
                except Exception:
                    pass

        # Kick off wait loop for background threads and eventual shutdown
        self._last_waiting_thread_count = None
        self._wait_for_threads_and_close()
    
    def create_main_window(self):
        """Create the main window with login fields and activity log"""
        self.root = tk.Tk()
        self.root.title("TN Refiling Bot - Version 3.1.0, Last Updated 12/04/2025")
        self.root.geometry("900x800")
        self.root.configure(bg="#f0f0f0")
        
        # Center the window
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (self.root.winfo_width() // 2)
        y = (self.root.winfo_screenheight() // 2) - (self.root.winfo_height() // 2)
        self.root.geometry(f"+{x}+{y}")
        
        # Header
        header_frame = tk.Frame(self.root, bg="#800000", height=70)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(header_frame, text="TN Refiling Bot", 
                              font=("Arial", 18, "bold"), bg="#800000", fg="white")
        title_label.pack(expand=True)
        
        subtitle_label = tk.Label(header_frame, text="Enter credentials and monitor bot progress in real-time", 
                                 font=("Arial", 10), bg="#800000", fg="#cccccc")
        subtitle_label.pack()
        
        # Scrollable canvas container
        canvas_container = tk.Frame(self.root, bg="#f0f0f0")
        canvas_container.pack(fill="both", expand=True)
        
        # Create canvas and scrollbar
        self.canvas = tk.Canvas(canvas_container, bg="#f0f0f0", highlightthickness=0)
        scrollbar = tk.Scrollbar(canvas_container, orient="vertical", command=self.canvas.yview)
        scrollable_frame = tk.Frame(self.canvas, bg="#f0f0f0")
        
        # Configure scrollable frame
        scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        # Create window on canvas
        canvas_window = self.canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        # Configure canvas scrolling
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Make canvas resizeable
        def on_canvas_configure(event):
            canvas_width = event.width
            self.canvas.itemconfig(canvas_window, width=canvas_width)
        
        self.canvas.bind("<Configure>", on_canvas_configure)
        
        # Enable mousewheel scrolling
        def on_mousewheel(event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
        self.canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Main content frame
        main_frame = tk.Frame(scrollable_frame, bg="#f0f0f0")
        main_frame.pack(fill="both", expand=True, padx=20, pady=15)
        
        # ========== Login Section ==========
        login_section = tk.LabelFrame(main_frame, text="Therapy Notes Login Credentials", 
                                     font=("Arial", 11, "bold"), bg="#f0f0f0", fg="#800000")
        login_section.pack(fill="x", pady=(0, 15))
        
        login_content = tk.Frame(login_section, bg="#f0f0f0")
        login_content.pack(fill="both", expand=True, padx=15, pady=10)
        
        # User Selection Dropdown
        user_row = tk.Frame(login_content, bg="#f0f0f0")
        user_row.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        
        tk.Label(user_row, text="User:", font=("Arial", 10), bg="#f0f0f0",
                anchor="w").pack(side="left", padx=(0, 10))
        
        self.user_dropdown = ttk.Combobox(user_row, font=("Arial", 10), width=25, state="readonly")
        self.user_dropdown.pack(side="left", padx=(0, 10))
        self.user_dropdown.bind("<<ComboboxSelected>>", self._on_user_selected)
        self._update_user_dropdown()
        
        # Add User button
        self.add_user_button = tk.Button(user_row, text="Add User", 
                                         command=self._add_user,
                                         bg="#800020", fg="white", font=("Arial", 9),
                                         padx=12, pady=4, cursor="hand2", relief="flat", bd=0)
        self.add_user_button.pack(side="left", padx=(0, 5))
        
        # Update Credentials button
        self.update_creds_button = tk.Button(user_row, text="Update Credentials", 
                                            command=self._update_credentials,
                                            bg="#800020", fg="white", font=("Arial", 9),
                                            padx=12, pady=4, cursor="hand2", relief="flat", bd=0)
        self.update_creds_button.pack(side="left", padx=(0, 5))
        
        # Login button
        self.login_button = tk.Button(user_row, text="Login", 
                                     command=self._start_login,
                                    bg="#800020", fg="white", font=("Arial", 9),
                                    padx=12, pady=4, cursor="hand2", state="disabled",
                                    relief="flat", bd=0)
        self.login_button.pack(side="left")
        
        # Username
        tk.Label(login_content, text="Username:", font=("Arial", 10), bg="#f0f0f0",
                anchor="w").grid(row=1, column=0, sticky="w", pady=(0, 5))
        self.username_entry = tk.Entry(login_content, font=("Arial", 10), width=40)
        self.username_entry.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        self.username_entry.focus()
        
        # Password
        tk.Label(login_content, text="Password:", font=("Arial", 10), bg="#f0f0f0",
                anchor="w").grid(row=3, column=0, sticky="w", pady=(0, 5))
        self.password_entry = tk.Entry(login_content, font=("Arial", 10), width=40, show="*")
        self.password_entry.grid(row=4, column=0, sticky="ew", pady=(0, 10))
        self.password_entry.bind("<Return>", lambda e: self._start_login() if self.username_entry.get().strip() and self.password_entry.get().strip() else None)
        
        # Status label
        self.status_label = tk.Label(login_content, text="Status: Select a user or enter credentials to begin...", 
                                    font=("Arial", 9, "bold"), bg="#f0f0f0", fg="#0066cc")
        self.status_label.grid(row=5, column=0, sticky="w")
        
        login_content.grid_columnconfigure(0, weight=1)
        
        # ========== PDF Selection Section ==========
        pdf_section = tk.LabelFrame(main_frame, text="PDF Claim Number Extraction", 
                                   font=("Arial", 11, "bold"), bg="#f0f0f0", fg="#800000")
        pdf_section.pack(fill="x", pady=(0, 15))
        
        pdf_content = tk.Frame(pdf_section, bg="#f0f0f0")
        pdf_content.pack(fill="both", expand=True, padx=15, pady=10)
        
        # Instructions
        pdf_instruction = "Select a scanned PDF document to extract claim number using OCR."
        tk.Label(pdf_content, text=pdf_instruction, 
                font=("Arial", 9), bg="#f0f0f0", fg="#666666", wraplength=600, 
                justify="left").pack(pady=(0, 10))
        
        # PDF selection row
        pdf_row = tk.Frame(pdf_content, bg="#f0f0f0")
        pdf_row.pack(fill="x")
        
        tk.Label(pdf_row, text="PDF File:", font=("Arial", 10), bg="#f0f0f0",
                anchor="w").pack(side="left", padx=(0, 10))
        
        self.pdf_entry = tk.Entry(pdf_row, font=("Arial", 9), width=50, state="readonly")
        self.pdf_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        browse_pdf_button = tk.Button(pdf_row, text="Browse...", 
                                     command=self._browse_pdf,
                                     bg="#800020", fg="white", font=("Arial", 9),
                                     padx=12, pady=4, cursor="hand2", relief="flat", bd=0)
        browse_pdf_button.pack(side="left", padx=(0, 5))
        
        extract_claim_button = tk.Button(pdf_row, text="Extract Claim #", 
                                     command=self._extract_claim,
                                     bg="#800020", fg="white", font=("Arial", 9),
                                     padx=12, pady=4, cursor="hand2", relief="flat", bd=0)
        extract_claim_button.pack(side="left", padx=(0, 5))
        
        clear_pdf_button = tk.Button(pdf_row, text="Clear", 
                                     command=self._clear_pdf,
                                     bg="#800020", fg="white", font=("Arial", 9),
                                     padx=12, pady=4, cursor="hand2", relief="flat", bd=0)
        clear_pdf_button.pack(side="left")
        
        # PDF status label
        self.pdf_status_label = tk.Label(pdf_content, text="No PDF selected", 
                                        font=("Arial", 9, "italic"), bg="#f0f0f0", fg="#666666")
        self.pdf_status_label.pack(pady=(5, 0))
        
        # ========== Excel/CSV Batch Processing Section ==========
        excel_section = tk.LabelFrame(main_frame, text="Client Data Input (Excel/CSV)", 
                                     font=("Arial", 11, "bold"), bg="#f0f0f0", fg="#800000")
        excel_section.pack(fill="x", pady=(0, 15))
        
        excel_content = tk.Frame(excel_section, bg="#f0f0f0")
        excel_content.pack(fill="both", expand=True, padx=15, pady=10)
        
        # Instructions
        excel_instruction = "Select an Excel (.xlsx) or CSV file containing client data (Name, DOB, Date of Service)."
        tk.Label(excel_content, text=excel_instruction, 
                font=("Arial", 9), bg="#f0f0f0", fg="#666666", wraplength=600, 
                justify="left").pack(pady=(0, 10))
        
        # Excel/CSV selection row
        excel_row = tk.Frame(excel_content, bg="#f0f0f0")
        excel_row.pack(fill="x")
        
        tk.Label(excel_row, text="Excel/CSV File:", font=("Arial", 10), bg="#f0f0f0",
                anchor="w").pack(side="left", padx=(0, 10))
        
        self.excel_entry = tk.Entry(excel_row, font=("Arial", 9), width=50, state="readonly")
        self.excel_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        browse_excel_button = tk.Button(excel_row, text="Browse...", 
                                       command=self._browse_excel,
                                       bg="#800020", fg="white", font=("Arial", 9),
                                       padx=12, pady=4, cursor="hand2", relief="flat", bd=0)
        browse_excel_button.pack(side="left", padx=(0, 5))
        
        paste_excel_button = tk.Button(excel_row, text="Paste path...",
                                      command=self._paste_excel_path,
                                      bg="#800020", fg="white", font=("Arial", 9),
                                      padx=12, pady=4, cursor="hand2", relief="flat", bd=0)
        paste_excel_button.pack(side="left", padx=(0, 5))
        
        clear_excel_button = tk.Button(excel_row, text="Clear", 
                                      command=self._clear_excel,
                                      bg="#800020", fg="white", font=("Arial", 9),
                                      padx=12, pady=4, cursor="hand2", relief="flat", bd=0)
        clear_excel_button.pack(side="left")
        
        # Excel/CSV status label
        self.excel_status_label = tk.Label(excel_content, text="No Excel/CSV selected", 
                                          font=("Arial", 9, "italic"), bg="#f0f0f0", fg="#666666")
        self.excel_status_label.pack(pady=(5, 0))
        
        # ========== Manual Column Mapping Section ==========
        mapping_section = tk.LabelFrame(excel_content, text="Manual Column Mapping (Optional)", 
                                        font=("Arial", 10, "bold"), bg="#f0f0f0", fg="#800000")
        mapping_section.pack(fill="x", pady=(10, 0))
        
        mapping_content = tk.Frame(mapping_section, bg="#f0f0f0")
        mapping_content.pack(fill="x", padx=10, pady=8)
        
        # Checkbox to enable manual mapping
        self.use_manual_mapping_var = tk.IntVar(value=0)
        manual_mapping_checkbox = tk.Checkbutton(mapping_content, 
                                                  text="Use manual column mapping (specify columns manually)",
                                                  variable=self.use_manual_mapping_var,
                                                  command=self._toggle_manual_mapping,
                                                  bg="#f0f0f0", font=("Arial", 9),
                                                  onvalue=1, offvalue=0)
        manual_mapping_checkbox.pack(anchor="w", pady=(0, 8))
        
        # Manual mapping fields frame (initially hidden)
        self.manual_mapping_frame = tk.Frame(mapping_content, bg="#f0f0f0")
        self.manual_mapping_frame.pack(fill="x", pady=(0, 5))
        
        # Instructions for manual mapping
        mapping_instruction = "Enter column letter (e.g., 'C') or column name from Excel/CSV file."
        tk.Label(self.manual_mapping_frame, text=mapping_instruction, 
                font=("Arial", 8), bg="#f0f0f0", fg="#666666", wraplength=550,
                justify="left").pack(anchor="w", pady=(0, 8))
        
        # Name mapping row
        name_row = tk.Frame(self.manual_mapping_frame, bg="#f0f0f0")
        name_row.pack(fill="x", pady=(0, 5))
        
        tk.Label(name_row, text="Option 1 - Split Name:", font=("Arial", 9), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        tk.Label(name_row, text="Last Name:", font=("Arial", 9), bg="#f0f0f0").pack(side="left", padx=(0, 5))
        self.last_name_col_entry = tk.Entry(name_row, font=("Arial", 9), width=8)
        self.last_name_col_entry.pack(side="left", padx=(0, 10))
        
        tk.Label(name_row, text="First Name:", font=("Arial", 9), bg="#f0f0f0").pack(side="left", padx=(0, 5))
        self.first_name_col_entry = tk.Entry(name_row, font=("Arial", 9), width=8)
        self.first_name_col_entry.pack(side="left", padx=(0, 10))
        
        tk.Label(name_row, text="OR Option 2 - Full Name:", font=("Arial", 9), bg="#f0f0f0").pack(side="left", padx=(0, 5))
        self.full_name_col_entry = tk.Entry(name_row, font=("Arial", 9), width=8)
        self.full_name_col_entry.pack(side="left", padx=(0, 10))
        
        # DOB and DOS mapping row
        dob_dos_row = tk.Frame(self.manual_mapping_frame, bg="#f0f0f0")
        dob_dos_row.pack(fill="x", pady=(0, 5))
        
        tk.Label(dob_dos_row, text="DOB:", font=("Arial", 9), bg="#f0f0f0").pack(side="left", padx=(0, 5))
        self.dob_col_entry = tk.Entry(dob_dos_row, font=("Arial", 9), width=8)
        self.dob_col_entry.pack(side="left", padx=(0, 20))
        
        tk.Label(dob_dos_row, text="Date of Service:", font=("Arial", 9), bg="#f0f0f0").pack(side="left", padx=(0, 5))
        self.dos_col_entry = tk.Entry(dob_dos_row, font=("Arial", 9), width=8)
        self.dos_col_entry.pack(side="left")
        
        # Apply mapping button
        apply_mapping_button = tk.Button(self.manual_mapping_frame, text="Apply Mapping",
                                         command=self._apply_manual_mapping,
                                         bg="#800020", fg="white", font=("Arial", 9),
                                         padx=12, pady=4, cursor="hand2", relief="flat", bd=0)
        apply_mapping_button.pack(anchor="w", pady=(5, 0))
        
        # Initially hide manual mapping fields
        self.manual_mapping_frame.pack_forget()
        
        # ========== Processing Section ==========
        processing_section = tk.LabelFrame(main_frame, text="Processing", 
                                          font=("Arial", 11, "bold"), bg="#f0f0f0", fg="#800000")
        processing_section.pack(fill="x", pady=(0, 15))
        
        processing_content = tk.Frame(processing_section, bg="#f0f0f0")
        processing_content.pack(fill="both", expand=True, padx=15, pady=10)
        
        # Instructions
        processing_instruction = "After logging in and loading client data, start the processing workflow to refile claims in Therapy Notes."
        tk.Label(processing_content, text=processing_instruction, 
                font=("Arial", 9), bg="#f0f0f0", fg="#666666", wraplength=600, 
                justify="left").pack(pady=(0, 10))
        
        # Processing scope controls
        self.run_all_rows_var = tk.IntVar(value=1)
        run_all_checkbox = tk.Checkbutton(processing_content,
                                          text="Run all rows from loaded Excel/CSV",
                                          variable=self.run_all_rows_var,
                                          command=self._toggle_processing_range,
                                          bg="#f0f0f0",
                                          font=("Arial", 9),
                                          onvalue=1,
                                          offvalue=0)
        run_all_checkbox.pack(anchor="w", pady=(0, 6))
        
        self.processing_range_frame = tk.Frame(processing_content, bg="#f0f0f0")
        self.processing_range_frame.pack(fill="x", pady=(0, 10))
        
        tk.Label(self.processing_range_frame, text="Row range (inclusive):", font=("Arial", 9), bg="#f0f0f0").pack(side="left", padx=(0, 8))
        tk.Label(self.processing_range_frame, text="From", font=("Arial", 9), bg="#f0f0f0").pack(side="left")
        self.processing_start_row_entry = tk.Entry(self.processing_range_frame, font=("Arial", 9), width=8, state="disabled")
        self.processing_start_row_entry.pack(side="left", padx=(4, 12))
        tk.Label(self.processing_range_frame, text="To", font=("Arial", 9), bg="#f0f0f0").pack(side="left")
        self.processing_end_row_entry = tk.Entry(self.processing_range_frame, font=("Arial", 9), width=8, state="disabled")
        self.processing_end_row_entry.pack(side="left", padx=(4, 12))
        tk.Label(self.processing_range_frame, text="Leave blank to process a single row.", font=("Arial", 8), bg="#f0f0f0", fg="#666666").pack(side="left")
        self._toggle_processing_range()
        
        # Processing controls row
        processing_row = tk.Frame(processing_content, bg="#f0f0f0")
        processing_row.pack(fill="x")
        
        # Start Processing button
        self.start_processing_button = tk.Button(processing_row, text="Start Processing", 
                                                command=self._start_processing,
                                                bg="#800020", fg="white", font=("Arial", 10, "bold"),
                                                padx=18, pady=6, cursor="hand2", state="disabled",
                                                relief="flat", bd=0)
        self.start_processing_button.pack(side="left", padx=(0, 10))
        
        # Processing status label
        self.processing_status_label = tk.Label(processing_row, text="Status: Not started", 
                                               font=("Arial", 9, "bold"), bg="#f0f0f0", fg="#666666")
        self.processing_status_label.pack(side="left", padx=(10, 0))
        
        # Processing progress frame
        progress_frame = tk.Frame(processing_content, bg="#f0f0f0")
        progress_frame.pack(fill="x", pady=(10, 0))
        
        tk.Label(progress_frame, text="Progress:", font=("Arial", 9), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.processing_progress_label = tk.Label(progress_frame, text="0/0 clients processed", 
                                                  font=("Arial", 9), bg="#f0f0f0", fg="#666666")
        self.processing_progress_label.pack(side="left")
        
        # Processing stats frame
        stats_frame = tk.Frame(processing_content, bg="#f0f0f0")
        stats_frame.pack(fill="x", pady=(5, 0))
        
        tk.Label(stats_frame, text="Statistics:", font=("Arial", 9), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.processing_stats_label = tk.Label(stats_frame, text="Successful: 0 | Failed: 0 | Total: 0", 
                                               font=("Arial", 9), bg="#f0f0f0", fg="#666666")
        self.processing_stats_label.pack(side="left")
        
        # ========== Output Excel Section ==========
        output_section = tk.LabelFrame(main_frame, text="Output Excel File", 
                                      font=("Arial", 11, "bold"), bg="#f0f0f0", fg="#800000")
        output_section.pack(fill="x", pady=(0, 15))
        
        output_content = tk.Frame(output_section, bg="#f0f0f0")
        output_content.pack(fill="both", expand=True, padx=15, pady=10)
        
        # Instructions
        output_instruction = "Select where to save the detailed output Excel file. This file will contain comprehensive tracking data for all processed clients (Refiled, Not Refiled, etc.), including modifier information (93/95), session medium, payer claim control numbers, and more."
        tk.Label(output_content, text=output_instruction, 
                font=("Arial", 9), bg="#f0f0f0", fg="#666666", wraplength=600, 
                justify="left").pack(pady=(0, 10))
        
        # Output Excel selection row
        output_row = tk.Frame(output_content, bg="#f0f0f0")
        output_row.pack(fill="x")
        
        tk.Label(output_row, text="Output Excel File:", font=("Arial", 10), bg="#f0f0f0",
                anchor="w").pack(side="left", padx=(0, 10))
        
        self.output_excel_entry = tk.Entry(output_row, font=("Arial", 9), width=50, state="readonly")
        self.output_excel_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        browse_output_button = tk.Button(output_row, text="Browse...", 
                                        command=self._browse_output_excel,
                                        bg="#800020", fg="white", font=("Arial", 9),
                                        padx=12, pady=4, cursor="hand2", relief="flat", bd=0)
        browse_output_button.pack(side="left", padx=(0, 5))
        
        clear_output_button = tk.Button(output_row, text="Clear", 
                                       command=self._clear_output_excel,
                                       bg="#800020", fg="white", font=("Arial", 9),
                                       padx=12, pady=4, cursor="hand2", relief="flat", bd=0)
        clear_output_button.pack(side="left")
        
        # Output Excel status label
        self.output_excel_status_label = tk.Label(output_content, text="No output file selected (will use default location)", 
                                                  font=("Arial", 9, "italic"), bg="#f0f0f0", fg="#666666")
        self.output_excel_status_label.pack(pady=(5, 0))
        
        # ========== Testing Section ==========
        testing_section = tk.LabelFrame(main_frame, text="Testing", 
                                        font=("Arial", 11, "bold"), bg="#f0f0f0", fg="#800000")
        testing_section.pack(fill="x", pady=(0, 15))
        
        testing_content = tk.Frame(testing_section, bg="#f0f0f0")
        testing_content.pack(fill="both", expand=True, padx=15, pady=10)
        
        # Instructions
        testing_instruction = "Test navigation for clients by entering a row number or row range from your Excel file (1-based, where row 1 is the header). Examples: '5' for single row, '5-10' for range."
        tk.Label(testing_content, text=testing_instruction, 
                font=("Arial", 9), bg="#f0f0f0", fg="#666666", wraplength=600, 
                justify="left").pack(pady=(0, 10))
        
        # Testing controls row
        testing_row = tk.Frame(testing_content, bg="#f0f0f0")
        testing_row.pack(fill="x")
        
        tk.Label(testing_row, text="Row Number/Range:", font=("Arial", 10), bg="#f0f0f0",
                anchor="w").pack(side="left", padx=(0, 10))
        
        self.test_row_entry = tk.Entry(testing_row, font=("Arial", 10), width=20)
        self.test_row_entry.pack(side="left", padx=(0, 10))
        
        # Test button
        self.test_navigation_button = tk.Button(testing_row, text="Test Navigation", 
                                                command=self._test_navigation,
                                                bg="#28a745", fg="white", font=("Arial", 10, "bold"),
                                                padx=18, pady=6, cursor="hand2", state="disabled",
                                                relief="flat", bd=0)
        self.test_navigation_button.pack(side="left", padx=(0, 10))
        
        # Test status label
        self.test_status_label = tk.Label(testing_row, text="Status: Enter row number/range and click 'Test Navigation'", 
                                          font=("Arial", 9), bg="#f0f0f0", fg="#666666")
        self.test_status_label.pack(side="left", padx=(10, 0))
        
        # ========== Complex Testing Section ==========
        complex_testing_section = tk.LabelFrame(main_frame, text="Complex Testing", 
                                                font=("Arial", 11, "bold"), bg="#f0f0f0", fg="#800000")
        complex_testing_section.pack(fill="x", pady=(0, 15))
        
        complex_testing_content = tk.Frame(complex_testing_section, bg="#f0f0f0")
        complex_testing_content.pack(fill="both", expand=True, padx=15, pady=10)
        
        # Instructions
        complex_testing_instruction = "Complex testing runs the FULL refiling flow (clicks Save Changes, navigates to Submit Primary Claims, filters date range, selects checkbox, selects Amended Claim, enters payer claim control) but stops right before clicking the final 'Submit Claims' button. The bot will stay on the page for your review. Enter a row number or row range from your Excel file (1-based, where row 1 is the header). Examples: '5' for single row, '5-10' for range."
        tk.Label(complex_testing_content, text=complex_testing_instruction, 
                font=("Arial", 9), bg="#f0f0f0", fg="#666666", wraplength=600, 
                justify="left").pack(pady=(0, 10))
        
        # Complex Testing controls row
        complex_testing_row = tk.Frame(complex_testing_content, bg="#f0f0f0")
        complex_testing_row.pack(fill="x")
        
        tk.Label(complex_testing_row, text="Row Number/Range:", font=("Arial", 10), bg="#f0f0f0",
                anchor="w").pack(side="left", padx=(0, 10))
        
        self.complex_test_row_entry = tk.Entry(complex_testing_row, font=("Arial", 10), width=20)
        self.complex_test_row_entry.pack(side="left", padx=(0, 10))
        
        # Complex Test button
        self.complex_test_navigation_button = tk.Button(complex_testing_row, text="Complex Test", 
                                                         command=self._complex_test_navigation,
                                                         bg="#ff9500", fg="white", font=("Arial", 10, "bold"),
                                                         padx=18, pady=6, cursor="hand2", state="disabled",
                                                         relief="flat", bd=0)
        self.complex_test_navigation_button.pack(side="left", padx=(0, 10))
        
        # Complex Test status label
        self.complex_test_status_label = tk.Label(complex_testing_row, text="Status: Enter row number/range and click 'Complex Test'", 
                                                   font=("Arial", 9), bg="#f0f0f0", fg="#666666")
        self.complex_test_status_label.pack(side="left", padx=(10, 0))
        
        # ========== Activity Log Section ==========
        log_section = tk.LabelFrame(main_frame, text="Activity Log", 
                                   font=("Arial", 11, "bold"), bg="#f0f0f0", fg="#800000")
        log_section.pack(fill="both", expand=True, pady=(0, 15))
        
        log_content = tk.Frame(log_section, bg="#f0f0f0")
        log_content.pack(fill="both", expand=True, padx=15, pady=10)
        
        # Log text box with scrollbar
        self.log_text = scrolledtext.ScrolledText(log_content, height=15, 
                                                  font=("Consolas", 9), 
                                                  bg="#ffffff", fg="#000000",
                                                  wrap=tk.WORD)
        self.log_text.pack(fill="both", expand=True)
        self._start_gui_log_dispatcher()
        self._drain_gui_log_queue()
        
        # Bottom buttons
        button_frame = tk.Frame(main_frame, bg="#f0f0f0")
        button_frame.pack(fill="x", pady=(0, 5))
        
        # Stop button (left side)
        self.stop_button = tk.Button(button_frame, text="⛔ Stop Bot", 
                                     command=self._request_stop,
                                     bg="#dc3545", fg="white", font=("Arial", 10, "bold"),
                                     padx=18, pady=6, cursor="hand2",
                                     relief="flat", bd=0)
        self.stop_button.pack(side="left", padx=(0, 10))
        
        # Close button on the right
        self.close_button = tk.Button(button_frame, text="Close", 
                                     command=self._close_window,
                                     bg="#800020", fg="white", font=("Arial", 10),
                                     padx=18, pady=6, cursor="hand2",
                                     relief="flat", bd=0)
        self.close_button.pack(side="right")
        
        # Initial log message
        self.gui_log("TN Refiling Bot initialized")
        if self.users:
            self.gui_log(f"Loaded {len(self.users)} saved user(s) - Select a user from the dropdown or add a new one")
        else:
            self.gui_log("No users saved - Click 'Add User' to create your first user profile")
        
        # Bind username/password changes to enable login button
        def check_credentials(event=None):
            self._enable_login()
        
        self.username_entry.bind("<KeyRelease>", check_credentials)
        self.password_entry.bind("<KeyRelease>", check_credentials)
    
    def _enable_processing_button(self):
        """Enable the Start Processing button after login and data load"""
        if hasattr(self, 'start_processing_button'):
            # Enable if logged in and has client data
            if self.is_logged_in and self.excel_client_data:
                self.start_processing_button.config(state="normal")
                self.processing_status_label.config(text="Status: Ready to process", fg="#28a745")
                self.gui_log(f"Processing ready: {len(self.excel_client_data)} client(s) loaded")
                
                # Also enable test button
                if hasattr(self, 'test_navigation_button'):
                    self.test_navigation_button.config(state="normal")
                    self.complex_test_navigation_button.config(state="normal")
                    if hasattr(self, 'test_status_label'):
                        self.test_status_label.config(text=f"Status: Ready to test (Loaded {len(self.excel_client_data)} client(s))", fg="#666666")
                    if hasattr(self, 'complex_test_status_label'):
                        self.complex_test_status_label.config(text=f"Status: Ready to complex test (Loaded {len(self.excel_client_data)} client(s))", fg="#666666")
            elif self.is_logged_in:
                self.start_processing_button.config(state="normal")
                self.processing_status_label.config(text="Status: Ready (load Excel/CSV for batch processing)", fg="#ff9500")
                
                # Disable test button if no data
                if hasattr(self, 'test_navigation_button'):
                    self.test_navigation_button.config(state="disabled")
                    self.complex_test_navigation_button.config(state="disabled")
            else:
                self.start_processing_button.config(state="disabled")
                self.processing_status_label.config(text="Status: Login required", fg="#666666")
                
                # Disable test button if not logged in
                if hasattr(self, 'test_navigation_button'):
                    self.test_navigation_button.config(state="disabled")
                    self.complex_test_navigation_button.config(state="disabled")
    
    def _test_navigation(self):
        """Test navigation for a row number or row range from the Excel file
        
        Supports:
        - Single row: "5" 
        - Range: "5-10" or "2386-2390"
        """
        if not self.is_logged_in:
            messagebox.showwarning("Not Logged In", "Please log in to Therapy Notes first")
            return
        
        if not self.excel_client_data:
            messagebox.showwarning("No Data", "Please load an Excel/CSV file with client data first")
            return
        
        if not self.driver:
            messagebox.showwarning("Browser Not Open", "Browser session not found. Please log in again.")
            return
        
        # Reset tracked clients for new test
        self.tracked_clients = []
        self.skipped_clients = []
        self.correct_modifier_clients = []
        
        # Reset modifier tracking
        self.current_original_modifier = None
        self.current_session_medium = None
        self.current_new_modifier = None
        self.current_modifier_action = None
        self.current_primary_policy_name = None
        self.current_primary_policy_member_id = None
        self.current_primary_payment_description = None
        
        # Get row input from entry
        row_input = self.test_row_entry.get().strip()
        if not row_input:
            messagebox.showwarning("Invalid Input", "Please enter a row number or range (e.g., '5' or '5-10')")
            return
        
        try:
            # Parse row input - support both single number and range
            row_numbers = []
            if '-' in row_input:
                # Range format: "5-10"
                parts = row_input.split('-')
                if len(parts) != 2:
                    raise ValueError("Invalid range format. Use 'start-end' (e.g., '5-10')")
                start_row = int(parts[0].strip())
                end_row = int(parts[1].strip())
                
                # Validate range
                if start_row < 2 or end_row < 2:
                    messagebox.showwarning("Invalid Row", "Row numbers must be 2 or greater (row 1 is the header)")
                    return
                if start_row > end_row:
                    messagebox.showwarning("Invalid Range", "Start row must be less than or equal to end row")
                    return
                
                # Generate list of row numbers
                row_numbers = list(range(start_row, end_row + 1))
            else:
                # Single row format: "5"
                row_number = int(row_input)
                if row_number < 2:
                    messagebox.showwarning("Invalid Row", "Row number must be 2 or greater (row 1 is the header)")
                    return
                row_numbers = [row_number]
            
            # Validate we have rows to process
            if not row_numbers:
                messagebox.showwarning("Invalid Input", "No rows to process")
                return
            
            self.gui_log(f"\n{'='*80}", level="INFO")
            self.gui_log(f"TESTING NAVIGATION FOR ROW(S): {row_input} ({len(row_numbers)} row(s))", level="INFO")
            self.gui_log(f"{'='*80}\n", level="INFO")
            
            # Find all clients for the row numbers
            clients_to_test = []
            missing_rows = []
            
            for row_num in row_numbers:
                client_data = None
                for client in self.excel_client_data:
                    if client.get('excel_row') == row_num:
                        client_data = client
                        break
                
                if client_data:
                    clients_to_test.append((row_num, client_data))
                else:
                    missing_rows.append(row_num)
            
            # Report missing rows
            if missing_rows:
                missing_str = ', '.join(map(str, missing_rows[:10]))  # Show first 10
                if len(missing_rows) > 10:
                    missing_str += f", ... ({len(missing_rows)} total)"
                self.gui_log(f"⚠️ Warning: {len(missing_rows)} row(s) not found in loaded data: {missing_str}", level="WARNING")
                self.gui_log("These rows may be filtered out, hidden, or contain 'New Referrals'", level="INFO")
            
            if not clients_to_test:
                messagebox.showwarning("No Rows Found", 
                    f"None of the requested row(s) were found in the loaded data.\n\n"
                    f"Requested: {row_input}\n"
                    f"Currently loaded: {len(self.excel_client_data)} visible data row(s).\n\n"
                    f"Rows may be filtered out, hidden, or contain 'New Referrals'.")
                return
            
            # Process each client in the range
            total_clients = len(clients_to_test)
            successful = 0
            failed = 0
            
            for idx, (row_number, client_data) in enumerate(clients_to_test, 1):
                # Set testing mode flag for this client
                self.is_testing_mode = True
                self.test_row_number = row_number
                
                self.gui_log(f"\n{'='*80}", level="INFO")
                self.gui_log(f"TESTING ROW {row_number} ({idx}/{total_clients}): {client_data.get('client_name', 'N/A')}", level="INFO")
                self.gui_log(f"{'='*80}", level="INFO")
                self.gui_log(f"Client: {client_data.get('client_name', 'N/A')}", level="INFO")
                self.gui_log(f"DOB: {client_data.get('dob', 'N/A')}", level="INFO")
                self.gui_log(f"Date of Service: {client_data.get('date_of_service', 'N/A')}", level="INFO")
                self.gui_log(f"{'='*80}\n", level="INFO")
                
                # Update test status
                if hasattr(self, 'test_status_label'):
                    self.test_status_label.config(
                        text=f"Status: Testing row {row_number} ({idx}/{total_clients}) - {client_data.get('client_name', 'N/A')}", 
                        fg="#ff9500"
                    )
                
                # Reset stop flag
                self.stop_requested = False
                
                # Store current client info
                client_name = client_data.get('client_name', '')
                dob = client_data.get('dob', '')
                dos = client_data.get('date_of_service', '')
                insurance = client_data.get('insurance', None)
                claim_status = client_data.get('claim_status', None)
                self.current_client_name = client_name
                self.current_client_dob = dob
                self.current_date_of_service = dos
                self.current_insurance = insurance  # Store insurance for this client
                self.current_claim_status = claim_status  # Store claim submission status for this client
                
                # Reset modifier tracking for this client (will be extracted when popup appears)
                self.current_original_modifier = None
                self.current_session_medium = None
                self.current_new_modifier = None
                self.current_modifier_action = None
                self.current_primary_policy_name = None
                self.current_primary_policy_member_id = None
                self.current_primary_payment_description = None
                self.current_primary_policy_name = None
                self.current_primary_policy_member_id = None
                self.current_primary_payment_description = None
                
                # Step 1: Ensure we're on the Patients page
                current_url = self.driver.current_url
                if "patients" not in current_url.lower():
                    self.gui_log("Not on Patients page - Navigating to Patients...", level="INFO")
                    if not self._navigate_to_patients():
                        self.gui_log(f"❌ Row {row_number}: Failed to navigate to Patients page", level="ERROR")
                        failed += 1
                        # Track failed client
                        self._add_tracked_client(
                            status='Error - Navigation Failed',
                            original_modifier='N/A',
                            new_modifier='N/A',
                            modifier_action='N/A',
                            reason=f'Could not navigate to Patients page',
                            error_message=f'Navigation failed for row {row_number}'
                        )
                        continue
                
                # Step 2: Reset search field for subsequent clients
                if idx > 1:
                    if not self._reset_search_field():
                        self.gui_log(f"⚠️ Row {row_number}: Could not reset search field, continuing anyway...", level="WARNING")
                
                # Step 3: Search for the client
                search_result = self._search_for_client(client_name, dob, dos)
                
                if not search_result:
                    self.gui_log(f"❌ Row {row_number}: Failed to search for client after all retry attempts", level="ERROR")
                    failed += 1
                    # Track failed client
                    self._add_tracked_client(
                        status='Error - Client Not Found',
                        original_modifier='N/A',
                        new_modifier='N/A',
                        modifier_action='N/A',
                        reason=f'Could not find client in Therapy Notes after all retry attempts',
                        error_message=f'Client search failed for row {row_number}: {client_name}'
                    )
                    continue
                
                # Note: The navigation will continue automatically through the full workflow
                # When it reaches Save Changes button, it will stop if is_testing_mode is True
                # The client will be tracked in the comprehensive tracking system
                self.gui_log(f"✅ Row {row_number}: Navigation test started - Will stop before Save Changes button", level="INFO")
                
                # Navigation continues automatically - the workflow will handle tracking
                # For successful completion, tracking happens in the workflow itself
                # We'll mark it as successful after the workflow completes
                successful += 1
                
                # Small delay between clients
                time.sleep(1)
            
            # Reset testing mode after all rows processed
            self.is_testing_mode = False
            self.test_row_number = None
            
            # Save output Excel for all tested rows
            if self.tracked_clients:
                self._save_output_excel()
            
            # Update final status
            if hasattr(self, 'test_status_label'):
                self.test_status_label.config(
                    text=f"Status: Test complete - {successful} successful, {failed} failed out of {total_clients} row(s)", 
                    fg="#28a745" if failed == 0 else "#ff9500"
                )
            
            self.gui_log(f"\n{'='*80}", level="INFO")
            self.gui_log(f"TESTING COMPLETE: {successful} successful, {failed} failed out of {total_clients} row(s)", level="INFO")
            if missing_rows:
                self.gui_log(f"Note: {len(missing_rows)} row(s) were not found in loaded data", level="INFO")
            self.gui_log(f"{'='*80}\n", level="INFO")
            
        except ValueError as ve:
            messagebox.showerror("Invalid Input", f"'{row_input}' is not a valid row number or range.\n\n"
                                 f"Use a single number (e.g., '5') or a range (e.g., '5-10').\n\n"
                                 f"Error: {str(ve)}")
        except Exception as e:
            self.log_error(f"Error testing navigation: {e}")
            messagebox.showerror("Error", f"Failed to test navigation:\n{e}")
            if hasattr(self, 'test_status_label'):
                self.test_status_label.config(text="Status: Test failed", fg="#dc3545")
            self.is_testing_mode = False
            self.test_row_number = None
    
    def _complex_test_navigation(self):
        """Complex test navigation - runs full flow but stops before Submit Claims button
        
        This mode:
        - Actually clicks Save Changes (unlike regular test mode)
        - Runs through the entire refiling flow
        - Stops right before clicking "Submit Claims" button
        - Stays on the page for user review (doesn't navigate back)
        """
        try:
            # Check if logged in
            if not self.is_logged_in:
                messagebox.showwarning("Not Logged In", "Please log in first before running complex test.")
                return
            
            # Check if Excel file is loaded
            if not self.excel_client_data:
                messagebox.showwarning("No Data Loaded", "Please load an Excel/CSV file first.")
                return
            
            # Get row input
            row_input = self.complex_test_row_entry.get().strip()
            if not row_input:
                messagebox.showwarning("Invalid Input", "Please enter a row number or range (e.g., '5' or '5-10').")
                return
            
            # Parse row input (supports single number or range)
            try:
                if '-' in row_input:
                    # Range input (e.g., "5-10")
                    parts = row_input.split('-')
                    if len(parts) != 2:
                        raise ValueError("Invalid range format")
                    start_row = int(parts[0].strip())
                    end_row = int(parts[1].strip())
                    if start_row < 2:
                        raise ValueError("Row numbers must be >= 2 (row 1 is header)")
                    if start_row > end_row:
                        raise ValueError("Start row must be <= end row")
                    row_numbers = list(range(start_row, end_row + 1))
                else:
                    # Single number
                    row_num = int(row_input.strip())
                    if row_num < 2:
                        raise ValueError("Row numbers must be >= 2 (row 1 is header)")
                    row_numbers = [row_num]
            except ValueError as ve:
                messagebox.showerror("Invalid Input", f"'{row_input}' is not a valid row number or range.\n\n"
                                    f"Use a single number (e.g., '5') or a range (e.g., '5-10').\n\n"
                                    f"Error: {str(ve)}")
                return
            
            # Validate we have rows to process
            if not row_numbers:
                messagebox.showwarning("Invalid Input", "No rows to process")
                return
            
            self.gui_log(f"\n{'='*80}", level="INFO")
            self.gui_log(f"COMPLEX TESTING FOR ROW(S): {row_input} ({len(row_numbers)} row(s))", level="INFO")
            self.gui_log(f"{'='*80}\n", level="INFO")
            
            # Find all clients for the row numbers
            clients_to_test = []
            missing_rows = []
            
            for row_num in row_numbers:
                client_data = None
                for client in self.excel_client_data:
                    if client.get('excel_row') == row_num:
                        client_data = client
                        break
                
                if client_data:
                    clients_to_test.append((row_num, client_data))
                else:
                    missing_rows.append(row_num)
            
            # Report missing rows
            if missing_rows:
                missing_str = ', '.join(map(str, missing_rows[:10]))  # Show first 10
                if len(missing_rows) > 10:
                    missing_str += f", ... ({len(missing_rows)} total)"
                self.gui_log(f"⚠️ Warning: {len(missing_rows)} row(s) not found in loaded data: {missing_str}", level="WARNING")
                self.gui_log("These rows may be filtered out, hidden, or contain 'New Referrals'", level="INFO")
            
            if not clients_to_test:
                messagebox.showwarning("No Rows Found", 
                    f"None of the requested row(s) were found in the loaded data.\n\n"
                    f"Requested: {row_input}\n"
                    f"Currently loaded: {len(self.excel_client_data)} visible data row(s).\n\n"
                    f"Rows may be filtered out, hidden, or contain 'New Referrals'.")
                return
            
            # Process each client in the range
            total_clients = len(clients_to_test)
            successful = 0
            failed = 0
            
            for idx, (row_number, client_data) in enumerate(clients_to_test, 1):
                # Set complex testing mode flag for this client
                self.is_complex_testing_mode = True
                self.test_row_number = row_number
                
                self.gui_log(f"\n{'='*80}", level="INFO")
                self.gui_log(f"COMPLEX TESTING ROW {row_number} ({idx}/{total_clients}): {client_data.get('client_name', 'N/A')}", level="INFO")
                self.gui_log(f"{'='*80}", level="INFO")
                self.gui_log(f"Client: {client_data.get('client_name', 'N/A')}", level="INFO")
                self.gui_log(f"DOB: {client_data.get('dob', 'N/A')}", level="INFO")
                self.gui_log(f"Date of Service: {client_data.get('date_of_service', 'N/A')}", level="INFO")
                self.gui_log(f"{'='*80}\n", level="INFO")
                
                # Update complex test status
                if hasattr(self, 'complex_test_status_label'):
                    self.complex_test_status_label.config(
                        text=f"Status: Complex testing row {row_number} ({idx}/{total_clients}) - {client_data.get('client_name', 'N/A')}", 
                        fg="#ff9500"
                    )
                
                # Reset stop flag
                self.stop_requested = False
                
                # Store current client info
                client_name = client_data.get('client_name', '')
                dob = client_data.get('dob', '')
                dos = client_data.get('date_of_service', '')
                insurance = client_data.get('insurance', None)
                claim_status = client_data.get('claim_status', None)
                self.current_client_name = client_name
                self.current_client_dob = dob
                self.current_date_of_service = dos
                self.current_insurance = insurance  # Store insurance for this client
                self.current_claim_status = claim_status  # Store claim submission status for this client
                
                # Reset modifier tracking for this client (will be extracted when popup appears)
                self.current_original_modifier = None
                self.current_session_medium = None
                self.current_new_modifier = None
                self.current_modifier_action = None
                
                # Step 1: Ensure we're on the Patients page
                current_url = self.driver.current_url
                if "patients" not in current_url.lower():
                    self.gui_log("Not on Patients page - Navigating to Patients...", level="INFO")
                    if not self._navigate_to_patients():
                        self.gui_log(f"❌ Row {row_number}: Failed to navigate to Patients page", level="ERROR")
                        failed += 1
                        # Track failed client
                        self._add_tracked_client(
                            status='Error - Navigation Failed',
                            original_modifier='N/A',
                            new_modifier='N/A',
                            modifier_action='N/A',
                            reason=f'Could not navigate to Patients page',
                            error_message=f'Navigation failed for row {row_number}'
                        )
                        continue
                
                # Step 2: Reset search field for subsequent clients
                if idx > 1:
                    if not self._reset_search_field():
                        self.gui_log(f"⚠️ Row {row_number}: Could not reset search field, continuing anyway...", level="WARNING")
                
                # Step 3: Search for the client
                search_result = self._search_for_client(client_name, dob, dos)
                
                if not search_result:
                    self.gui_log(f"❌ Row {row_number}: Failed to search for client after all retry attempts", level="ERROR")
                    failed += 1
                    # Track failed client
                    self._add_tracked_client(
                        status='Error - Client Not Found',
                        original_modifier='N/A',
                        new_modifier='N/A',
                        modifier_action='N/A',
                        reason=f'Could not find client in Therapy Notes after all retry attempts',
                        error_message=f'Client search failed for row {row_number}: {client_name}'
                    )
                    continue
                
                # Note: The navigation will continue automatically through the full workflow
                # Complex testing mode will click Save Changes and continue through all steps
                # but stop before clicking Submit Claims button
                # The client will be tracked in the comprehensive tracking system
                self.gui_log(f"✅ Row {row_number}: Complex test started - Will run full flow but stop before Submit Claims", level="INFO")
                
                # Navigation continues automatically - the workflow will handle tracking
                # For successful completion, tracking happens in the workflow itself
                # We'll mark it as successful after the workflow completes
                successful += 1
                
                # Small delay between clients
                time.sleep(1)
            
            # Reset complex testing mode after all rows processed
            self.is_complex_testing_mode = False
            self.test_row_number = None
            
            # Save output Excel for all tested rows
            if self.tracked_clients:
                self.gui_log(f"Saving output Excel with {len(self.tracked_clients)} tracked client(s)...", level="INFO")
                self._save_output_excel()
                self.gui_log(f"✅ Output Excel saved successfully with {len(self.tracked_clients)} client record(s)", level="INFO")
            else:
                self.gui_log("⚠️ No tracked clients to save in output Excel", level="WARNING")
            
            # Update final status
            if hasattr(self, 'complex_test_status_label'):
                self.complex_test_status_label.config(
                    text=f"Status: Complex test complete - {successful} successful, {failed} failed out of {total_clients} row(s)", 
                    fg="#28a745" if failed == 0 else "#ff9500"
                )
            
            self.gui_log(f"\n{'='*80}", level="INFO")
            self.gui_log(f"COMPLEX TESTING COMPLETE: {successful} successful, {failed} failed out of {total_clients} row(s)", level="INFO")
            if missing_rows:
                self.gui_log(f"Note: {len(missing_rows)} row(s) were not found in loaded data", level="INFO")
            if self.tracked_clients:
                self.gui_log(f"✅ Output Excel saved with {len(self.tracked_clients)} client record(s)", level="INFO")
            self.gui_log(f"{'='*80}\n", level="INFO")
            
        except ValueError as ve:
            messagebox.showwarning("Invalid Input", f"'{row_input}' is not a valid row number or range.\n\n"
                                 f"Use a single number (e.g., '5') or a range (e.g., '5-10').\n\n"
                                 f"Error: {str(ve)}")
        except Exception as e:
            self.log_error(f"Error complex testing navigation: {e}")
            messagebox.showerror("Error", f"Failed to run complex test:\n{e}")
            if hasattr(self, 'complex_test_status_label'):
                self.complex_test_status_label.config(text="Status: Complex test failed", fg="#dc3545")
            self.is_complex_testing_mode = False
            self.test_row_number = None
    
    def _start_processing(self):
        """Start the processing workflow"""
        if not self.is_logged_in:
            messagebox.showwarning("Not Logged In", "Please log in to Therapy Notes first")
            return
        
        if not self.excel_client_data:
            messagebox.showwarning("No Data", "Please load an Excel/CSV file with client data first")
            return
        
        if not self.driver:
            messagebox.showwarning("Browser Not Open", "Browser session not found. Please log in again.")
            return
        
        # Determine processing scope
        if self.run_all_rows_var.get() == 1:
            selected_clients = list(self.excel_client_data)
            scope_description = f"all {len(selected_clients)} loaded client(s)"
        else:
            start_text = self.processing_start_row_entry.get().strip()
            end_text = self.processing_end_row_entry.get().strip()
            
            if not start_text:
                messagebox.showwarning("Invalid Range", "Enter a starting row number (Excel rows are 1-based; row 1 is the header).")
                return
            
            try:
                start_row = int(start_text)
            except ValueError:
                messagebox.showwarning("Invalid Range", f"Starting row '{start_text}' is not a valid number.")
                return
            
            if end_text:
                try:
                    end_row = int(end_text)
                except ValueError:
                    messagebox.showwarning("Invalid Range", f"Ending row '{end_text}' is not a valid number.")
                    return
            else:
                end_row = start_row
            
            if start_row < 2 or end_row < 2:
                messagebox.showwarning("Invalid Range", "Row numbers must be 2 or greater (row 1 is the header).")
                return
            if end_row < start_row:
                messagebox.showwarning("Invalid Range", "Ending row must be greater than or equal to starting row.")
                return
            
            selected_clients = []
            for client in self.excel_client_data:
                excel_row = client.get('excel_row')
                try:
                    excel_row_int = int(excel_row)
                except (TypeError, ValueError):
                    continue
                if start_row <= excel_row_int <= end_row:
                    selected_clients.append(client)
            
            if not selected_clients:
                messagebox.showwarning("No Matching Rows", f"No clients found between Excel rows {start_row} and {end_row}.")
                return
            
            scope_description = f"rows {start_row}-{end_row} ({len(selected_clients)} client(s))"
        
        # Store selected clients for processing
        self.current_processing_clients = selected_clients
        
        # Reset testing mode flag - we're doing actual processing, not testing
        self.is_testing_mode = False
        
        # Reset tracked clients for new processing run
        self.tracked_clients = []
        self.skipped_clients = []
        self.correct_modifier_clients = []
        
        # Reset stop flag
        self.stop_requested = False
        
        # Reset statistics
        self.processing_stats = {
            'total': len(self.current_processing_clients),
            'successful': 0,
            'failed': 0,
            'current': 0
        }
        
        self.gui_log(f"Starting processing workflow for {scope_description}...", level="INFO")
        self.update_status("Processing started...", "#ff9500")
        
        # Update processing status
        self.processing_status_label.config(text="Status: Processing...", fg="#ff9500")
        self.start_processing_button.config(state="disabled")
        
        # Run processing in managed background thread
        thread = self._start_worker_thread("process-clients", self._process_clients)
        if thread is None:
            self.processing_status_label.config(text="Status: Shutdown in progress", fg="#dc3545")
            self.start_processing_button.config(state="normal")
    
    def _reset_search_field(self):
        """Reset search field by navigating back to Patients page and clearing search
        
        This is used when moving to the next client after a search failure.
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            self.gui_log("Resetting search field for next client...", level="DEBUG")
            
            # Navigate back to Patients page (click Patients button again)
            if not self._navigate_to_patients():
                self.gui_log("⚠️ Could not navigate back to Patients page", level="WARNING")
                return False
            
            # Wait for search field to be available (use dynamic wait instead of fixed sleep)
            try:
                search_field = WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.ID, "ctl00_BodyContent_TextBoxSearchPatientName"))
                )
                search_field.clear()
                self.gui_log("✅ Search field cleared", level="DEBUG")
                return True
            except:
                try:
                    search_field = self.driver.find_element(By.NAME, "ctl00$BodyContent$TextBoxSearchPatientName")
                    search_field.clear()
                    self.gui_log("✅ Search field cleared", level="DEBUG")
                    return True
                except Exception as e:
                    self.gui_log(f"⚠️ Could not clear search field: {e}", level="WARNING")
                    return False
                    
        except Exception as e:
            self.log_error("Error resetting search field", exception=e, include_traceback=True)
            return False
    
    def _process_clients(self):
        """Process clients from Excel/CSV data - will be implemented step-by-step"""
        try:
            data_to_process = self.current_processing_clients if self.current_processing_clients else list(self.excel_client_data)
            total = len(data_to_process)
            
            if total == 0:
                self.gui_log("No clients available for processing - aborting workflow", level="WARNING")
                self.processing_status_label.config(text="Status: No clients to process", fg="#dc3545")
                self.start_processing_button.config(state="normal")
                return
            
            self.gui_log(f"Processing {total} client(s)...", level="INFO")
            
            for idx, client in enumerate(data_to_process, 1):
                # Check for stop request
                if self.check_stop_requested():
                    self.gui_log("Processing stopped by user", level="WARNING")
                    break
                
                self.processing_stats['current'] = idx
                
                # Update progress in GUI
                if self.root:
                    self.root.after(0, lambda i=idx, t=total: self._update_processing_progress(i, t))
                
                client_name = client.get('client_name', 'Unknown')
                dob = client.get('dob', '')
                dos = client.get('date_of_service', '')
                
                # Store current client info for potential skipping
                self.current_client_name = client_name
                self.current_client_dob = dob
                self.current_date_of_service = dos
                self.current_original_modifier = None
                self.current_session_medium = None
                self.current_new_modifier = None
                self.current_modifier_action = None
                self.current_primary_policy_name = None
                self.current_primary_policy_member_id = None
                self.current_primary_payment_description = None
                
                self.gui_log(f"\n{'='*80}", level="INFO")
                self.gui_log(f"Processing client {idx}/{total}: {client_name}", level="INFO")
                self.gui_log(f"{'='*80}", level="INFO")
                self.gui_log(f"  DOB: {dob}, Date of Service: {dos}", level="INFO")
                
                # Step 1: Ensure we're on Patients page (for first client or after reset)
                if idx == 1:
                    current_url = self.driver.current_url
                    if "patients" not in current_url.lower():
                        self.gui_log("Not on Patients page - Navigating to Patients...", level="INFO")
                        if not self._navigate_to_patients():
                            self.gui_log(f"❌ Failed to navigate to Patients page for client {idx}", level="ERROR")
                            self.processing_stats['failed'] += 1
                            continue
                else:
                    # For subsequent clients, only navigate/reset if needed
                    current_url = self.driver.current_url
                    on_patients = "patients" in current_url.lower()
                    search_ready = False
                    if on_patients:
                        try:
                            search_field = self.driver.find_element(By.ID, "ctl00_BodyContent_TextBoxSearchPatientName")
                            if search_field.is_displayed() and search_field.is_enabled():
                                search_field.clear()
                                self.gui_log("✅ Search field ready - staying on Patients page", level="DEBUG")
                                search_ready = True
                        except Exception:
                            search_ready = False
                    if not search_ready:
                        if not self._reset_search_field():
                            self.gui_log(f"⚠️ Could not reset search field for client {idx}, continuing anyway...", level="WARNING")
                
                # Step 2: Search for the client with retry logic
                search_successful = self._search_for_client(client_name, dob, dos)
                
                if not search_successful:
                    self.gui_log(f"❌ Could not find client {idx}: {client_name} after all retry attempts", level="WARNING")
                    self.gui_log(f"Moving to next client...", level="INFO")
                    self.processing_stats['failed'] += 1
                    
                    # Track failed client
                    self._add_tracked_client(
                        status='Error - Client Not Found',
                        original_modifier='N/A',
                        new_modifier='N/A',
                        modifier_action='N/A',
                        reason=f'Could not find client in Therapy Notes after all retry attempts',
                        error_message=f'Client search failed: {client_name} (DOB: {dob}, DOS: {dos})'
                    )
                    
                    # Reset search field for next client
                    if idx < total:
                        self._reset_search_field()
                    continue
                
                # Navigation continues in _search_for_client -> _navigate_to_billing_tab -> _click_all_items_button
                # Date clicking will happen after "All Items" is clicked
                # After clicking payment amount, check for View ERA button
                # This happens in _click_payment_amount_link -> _check_and_click_view_era_button
                
                # Check if client was skipped during navigation (no View ERA button or no payments made)
                # The skip logic is handled in _check_and_click_view_era_button or _click_payment_amount_link
                # which navigates back to Patients if button doesn't exist or if no payments were made
                # The _navigate_back_to_billing_tab -> _click_all_items_button chain may return "skip_client" or "no_payments"
                
                # If client was skipped, it's already in skipped_clients list and tracked
                # and navigation back to Patients has already occurred - continue to next client
                if search_successful == "skip_client":
                    self.gui_log(f"⏭️ Client {idx} skipped - Continuing to next client...", level="INFO")
                    self.processing_stats['failed'] += 1
                    
                    # Note: Skip tracking is already done in _check_and_click_view_era_button
                    # when View ERA button is not found
                    
                    # Reset search field for next client
                    if idx < total:
                        self._reset_search_field()
                    continue
                
                # Check if "no payments made" scenario was encountered
                if search_successful == "no_payments":
                    self.gui_log(f"⏭️ Client {idx} skipped (no payments made) - Continuing to next client...", level="INFO")
                    self.processing_stats['failed'] += 1
                    
                    # Note: Tracking is already done in _click_payment_amount_link
                    # when "No payments have been made" message is found
                    
                    # Reset search field for next client
                    if idx < total:
                        self._reset_search_field()
                    continue
                
                # Check if "payment mismatch" scenario was encountered
                if search_successful == "payment_mismatch":
                    self.gui_log(f"⏭️ Client {idx} skipped (payment did not match primary policy) - Continuing to next client...", level="INFO")
                    self.processing_stats['failed'] += 1
                    
                    # Reset search field for next client
                    if idx < total:
                        self._reset_search_field()
                    continue
                
                # Check if "modifier already correct" scenario was encountered
                if search_successful == "modifier_correct":
                    self.gui_log(f"⏭️ Client {idx} skipped (modifier already correct) - Continuing to next client...", level="INFO")
                    self.processing_stats['failed'] += 1
                    
                    # Note: Tracking is already done in _click_progress_note_button
                    # when modifier matches expected value
                    
                    # Reset search field for next client
                    if idx < total:
                        self._reset_search_field()
                    continue
                
                if search_successful == "session_medium_missing":
                    self.gui_log(f"⏭️ Client {idx} skipped (session medium not found) - Continuing to next client...", level="INFO")
                    self.processing_stats['failed'] += 1
                    
                    # Reset search field for next client
                    if idx < total:
                        self._reset_search_field()
                    continue
                
                # Note: The navigation chain after clicking date again may return "next_client"
                # if modifier was correct (popup closed, navigated back to Patients)
                # This is handled in _click_date_of_service_again -> _check_and_update_modifier flow
                # If modifier was correct, the flow will have already:
                # - Closed the popup
                # - Added client to correct_modifier_clients list
                # - Navigated back to Patients page
                # So we can just continue to next client
                
                # Check if we need to handle "next_client" signal from the navigation chain
                # The check happens inside _click_date_of_service_again -> _check_and_update_modifier
                # which handles the "correct" modifier case and returns "next_client"
                
                # Continue processing - normal completion
                self.gui_log(f"✅ Client {idx} processing completed", level="INFO")
                self.processing_stats['successful'] += 1
                time.sleep(1)  # Small delay between clients
            
            # Final update and save output Excel
            if self.root:
                self.root.after(0, self._finish_processing)
            
            # Save output Excel with skipped clients and correct modifier clients
            if self.skipped_clients or self.correct_modifier_clients:
                self._save_output_excel()
                
        except Exception as e:
            self.log_error("Error during processing", exception=e, include_traceback=True)
            if self.root:
                self.root.after(0, lambda: self.processing_status_label.config(text="Status: Error occurred", fg="#dc3545"))
                self.root.after(0, lambda: self.start_processing_button.config(state="normal"))
    
    def _update_processing_progress(self, current, total):
        """Update processing progress in GUI (called from main thread)"""
        if hasattr(self, 'processing_progress_label'):
            self.processing_progress_label.config(text=f"{current}/{total} clients processed")
        
        if hasattr(self, 'processing_stats_label'):
            successful = self.processing_stats.get('successful', 0)
            failed = self.processing_stats.get('failed', 0)
            self.processing_stats_label.config(text=f"Successful: {successful} | Failed: {failed} | Total: {total}")
    
    def _finish_processing(self):
        """Finish processing and update GUI (called from main thread)"""
        total = self.processing_stats.get('total', 0)
        successful = self.processing_stats.get('successful', 0)
        failed = self.processing_stats.get('failed', 0)
        
        self.processing_status_label.config(text=f"Status: Complete - {successful} successful, {failed} failed", 
                                          fg="#28a745" if failed == 0 else "#ff9500")
        self.start_processing_button.config(state="normal")
        self.gui_log(f"Processing complete: {successful} successful, {failed} failed out of {total} total", level="INFO")
        self.current_processing_clients = []
    
    def _start_login(self):
        """Start login process in a separate thread"""
        if not self.username_entry.get().strip() or not self.password_entry.get().strip():
            messagebox.showwarning("Missing Credentials", "Please enter username and password")
            return
        
        # Reset stop flag when starting new operation
        self.stop_requested = False
        
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        
        self.gui_log(f"Starting login workflow for user: {username}", level="DEBUG")
        
        # Run login in separate thread to avoid blocking GUI
        thread = self._start_worker_thread("login", self.login, args=(username, password))
        if thread is None:
            self.gui_log("Login aborted because shutdown is in progress", level="WARNING", include_context=False)
    
    def _extract_claim(self):
        """Extract claim number from selected PDF"""
        if not self.selected_pdf_path:
            messagebox.showwarning("No PDF Selected", "Please select a PDF file first")
            return
        
        # Run extraction in separate thread
        thread = self._start_worker_thread("pdf-extraction", self._extract_claim_thread)
        if thread is None:
            self.gui_log("PDF extraction aborted because shutdown is in progress", level="WARNING", include_context=False)
    
    def _extract_claim_thread(self):
        """Extract claim number in separate thread"""
        try:
            self.gui_log(f"Starting PDF claim extraction for: {Path(self.selected_pdf_path).name}", level="DEBUG")
            result = self.extract_claim_number_from_pdf(self.selected_pdf_path)
            
            if self.check_stop_requested():
                return
            
            if result and result.get('claim_number'):
                claim_number = result['claim_number']
                self.gui_log(f"Successfully extracted claim number: {claim_number}", level="INFO")
                self.root.after(0, lambda: messagebox.showinfo("Claim Number Found", 
                                                               f"Claim Number: {claim_number}"))
            else:
                self.gui_log("Could not extract claim number from PDF", level="WARNING")
                self.root.after(0, lambda: messagebox.showwarning("Claim Number Not Found", 
                                                                 "Could not extract claim number from PDF. Please check the document."))
        except Exception as e:
            self.log_error("Error during claim extraction", exception=e, include_traceback=True)
            self.root.after(0, lambda: messagebox.showerror("Extraction Error", 
                                                           f"Error extracting claim number:\n{e}"))
    
    def _init_chrome_driver(self):
        """Initialize Chrome WebDriver"""
        if not SELENIUM_AVAILABLE:
            self.log_error("Selenium not available - cannot open browser", include_traceback=False)
            self.root.after(0, lambda: messagebox.showerror("Selenium Not Installed", 
                                                           "Selenium WebDriver is required for browser automation.\n\n"
                                                           "Please install it with:\npip install selenium"))
            return False
        
        try:
            self.gui_log("Initializing Chrome WebDriver...", level="DEBUG")
            self.update_status("Starting Chrome browser...", "#ff9500")
            
            # Chrome options
            chrome_options = Options()
            # Keep browser open for development
            chrome_options.add_experimental_option("detach", True)
            # Disable automation flags to avoid detection
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            # Initialize Chrome driver with automatic ChromeDriver management if available
            self.gui_log("Launching Chrome browser...", level="INFO")
            if WEBDRIVER_MANAGER_AVAILABLE:
                self.gui_log("Using webdriver-manager for automatic ChromeDriver management", level="DEBUG")
                service = ChromeService(ChromeDriverManager().install())
                self.driver = webdriver.Chrome(service=service, options=chrome_options)
            else:
                self.gui_log("Using system ChromeDriver (make sure ChromeDriver is in PATH)", level="DEBUG")
                self.driver = webdriver.Chrome(options=chrome_options)
            
            # Set up wait for elements
            self.wait = WebDriverWait(self.driver, 10)
            
            self.gui_log("✅ Chrome browser opened successfully", level="INFO")
            self.update_status("Chrome browser opened", "#28a745")
            
            return True
            
        except WebDriverException as e:
            self.log_error("Failed to initialize Chrome WebDriver", exception=e, include_traceback=True)
            self.root.after(0, lambda: messagebox.showerror("Chrome Driver Error", 
                                                           f"Failed to start Chrome browser:\n{e}\n\n"
                                                           "Make sure Chrome is installed and ChromeDriver is available.\n"
                                                           "You may need to install webdriver-manager:\npip install webdriver-manager"))
            return False
        except Exception as e:
            self.log_error("Unexpected error initializing Chrome driver", exception=e, include_traceback=True)
            self.update_status(f"Driver error: {str(e)}", "#dc3545")
            return False
    
    def _close_driver(self):
        """Close Chrome WebDriver"""
        try:
            if self.driver:
                self.gui_log("Closing Chrome browser...", level="DEBUG")
                self.driver.quit()
                self.driver = None
                self.wait = None
                self.gui_log("✅ Chrome browser closed", level="INFO")
        except Exception as e:
            self.log_error("Error closing Chrome driver", exception=e, include_traceback=True)
    
    def _navigate_to_patients(self):
        """Navigate to Patients section by clicking the Patients button on the left sidebar
        
        Element structure:
        <div><svg width="16" height="16" viewBox="0 0 16 16"...>
        SVG contains a path with id="Primary" and specific path data for circle-user icon
        
        Returns:
            bool: True if navigation successful, False otherwise
        """
        try:
            self.gui_log("Step 1: Navigating to Patients section...", level="INFO")
            self.update_status("Clicking Patients button...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Strategy 1: Find by SVG structure (circle-user icon)
            # Look for the SVG with the specific path ID "Primary" or viewBox="0 0 16 16"
            try:
                self.gui_log("Looking for Patients button (searching for circle-user SVG icon)...", level="DEBUG")
                
                # Wait for page to be ready (check document.readyState instead of fixed sleep)
                WebDriverWait(self.driver, 3).until(
                    lambda driver: driver.execute_script("return document.readyState") == "complete"
                )
                time.sleep(0.5)  # Small delay for any dynamic content
                
                # Try multiple strategies to find the Patients button
                patients_button = None
                
                # Strategy 1a: Find SVG with the specific viewBox and path structure
                try:
                    # Find SVG with viewBox="0 0 16 16" and path with id="Primary"
                    svg_xpath = "//svg[@viewBox='0 0 16 16']//path[@id='Primary']"
                    svg_element = self.wait.until(
                        EC.presence_of_element_located((By.XPATH, svg_xpath))
                    )
                    self.gui_log("✅ Found circle-user SVG icon", level="DEBUG")
                    
                    # Navigate to the clickable parent element (usually a link or button)
                    # Try to find the parent anchor tag or clickable div
                    parent = svg_element.find_element(By.XPATH, "./ancestor::a[1] | ./ancestor::div[contains(@class, 'nav') or contains(@class, 'menu')][1] | ./ancestor::button[1]")
                    patients_button = parent
                    self.gui_log("✅ Found clickable parent element for Patients button", level="DEBUG")
                    
                except TimeoutException:
                    self.gui_log("SVG path method failed, trying alternative methods...", level="DEBUG")
                
                # Strategy 1b: Find by link text "Patients"
                if not patients_button:
                    try:
                        self.gui_log("Trying to find 'Patients' link by text...", level="DEBUG")
                        patients_button = self.wait.until(
                            EC.element_to_be_clickable((By.LINK_TEXT, "Patients"))
                        )
                        self.gui_log("✅ Found Patients link by text", level="DEBUG")
                    except TimeoutException:
                        self.gui_log("Link text 'Patients' not found, trying partial text...", level="DEBUG")
                
                # Strategy 1c: Find by partial link text
                if not patients_button:
                    try:
                        patients_button = self.wait.until(
                            EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Patients"))
                        )
                        self.gui_log("✅ Found Patients link by partial text", level="DEBUG")
                    except TimeoutException:
                        self.gui_log("Partial link text 'Patients' not found...", level="DEBUG")
                
                # Strategy 1d: Find by XPath containing "Patients" text
                if not patients_button:
                    try:
                        # Look for any element containing "Patients" text in a navigation context
                        patients_xpath = "//a[contains(text(), 'Patients')] | //button[contains(text(), 'Patients')] | //div[contains(text(), 'Patients') and (contains(@class, 'nav') or contains(@class, 'menu'))]"
                        patients_button = self.wait.until(
                            EC.element_to_be_clickable((By.XPATH, patients_xpath))
                        )
                        self.gui_log("✅ Found Patients element by XPath text search", level="DEBUG")
                    except TimeoutException:
                        self.gui_log("XPath text search for 'Patients' failed...", level="DEBUG")
                
                # Strategy 1e: Find the SVG and click its parent container
                if not patients_button:
                    try:
                        # Find SVG by viewBox attribute
                        svg = self.driver.find_element(By.XPATH, "//svg[@viewBox='0 0 16 16' and .//path[@id='Primary']]")
                        # Find the nearest clickable ancestor
                        patients_button = svg.find_element(By.XPATH, "./ancestor::*[self::a or self::button or (self::div and (@onclick or @role='button'))][1]")
                        self.gui_log("✅ Found Patients button via SVG ancestor", level="DEBUG")
                    except Exception as e:
                        self.gui_log(f"SVG ancestor method failed: {e}", level="DEBUG")
                
                if patients_button:
                    # Scroll into view if needed
                    self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", patients_button)
                    time.sleep(0.5)
                    
                    # Click the button
                    patients_button.click()
                    self.gui_log("✅ Patients button clicked", level="INFO")
                    self.update_status("Patients button clicked - Waiting for page to load...", "#ff9500")
                    
                    # Wait for navigation to complete dynamically
                    try:
                        # Wait for URL to change or contain "patients"
                        WebDriverWait(self.driver, 5).until(
                            lambda driver: "patients" in driver.current_url.lower() or 
                            driver.execute_script("return document.readyState") == "complete"
                        )
                        time.sleep(0.5)  # Small delay for any final dynamic content
                    except TimeoutException:
                        # Fallback: wait a bit and continue
                        time.sleep(1)
                    
                    # Verify navigation by checking URL or page elements
                    current_url = self.driver.current_url
                    self.gui_log(f"Current URL after clicking Patients: {current_url}", level="DEBUG")
                    
                    if "patients" in current_url.lower():
                        self.gui_log("✅ Successfully navigated to Patients section", level="INFO")
                        self.update_status("Navigation successful - On Patients page", "#28a745")
                        return True
                    else:
                        self.gui_log(f"⚠️ URL doesn't contain 'patients' - Current URL: {current_url}", level="WARNING")
                        # Continue anyway - may still be on correct page
                        return True
                else:
                    self.log_error("Could not find Patients button using any method", include_traceback=False)
                    self.update_status("Error: Could not find Patients button", "#dc3545")
                    return False
                    
            except Exception as e:
                self.log_error("Error navigating to Patients section", exception=e, include_traceback=True)
                self.update_status(f"Navigation error: {str(e)}", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Unexpected error during Patients navigation", exception=e, include_traceback=True)
            return False
    
    def _normalize_client_name(self, client_name):
        """Normalize client name for searching in Therapy Notes
        
        Removes periods from middle initials and prepares name variations
        Example: "John H. Doe" -> ("John H Doe", "John Doe")
        
        Args:
            client_name: The client's name (e.g., "John H. Doe" or "Jane Smith")
        
        Returns:
            tuple: (full_name_with_middle, name_without_middle) - both normalized
        """
        if not client_name:
            return ("", "")
        
        # Remove extra spaces
        name = " ".join(client_name.split())
        
        # Remove periods from middle initials (e.g., "John H. Doe" -> "John H Doe")
        # Pattern: letter + period + space (middle initial)
        import re
        name = re.sub(r'([A-Z])\.\s', r'\1 ', name)
        
        # Split into parts
        parts = name.split()
        
        if len(parts) >= 3:
            # Has middle name/initial
            first = parts[0]
            middle = parts[1]
            last = " ".join(parts[2:])  # Handle last names with spaces
            
            full_name = f"{first} {middle} {last}"
            name_without_middle = f"{first} {last}"
            
            return (full_name, name_without_middle)
        elif len(parts) == 2:
            # Just first and last name
            return (name, name)
        else:
            # Single name or empty
            return (name, name)
    
    def _normalize_dob_for_comparison(self, dob_str):
        """Normalize DOB string for comparison across different formats
        
        Handles formats like:
        - MM/DD/YYYY (e.g., "01/22/1963")
        - MM/DD/YY (e.g., "01/22/63")
        - YYYY/MM/DD (e.g., "1963/01/22")
        - YY/MM/DD (e.g., "63/01/22")
        
        Args:
            dob_str: Date of birth as string
        
        Returns:
            tuple: (normalized_dob_str, normalized_dob_date) or (None, None) if invalid
            Format: "YYYY-MM-DD" string and date object for comparison
        """
        if not dob_str or str(dob_str).strip() in ['', 'nan', 'None', 'N/A']:
            return (None, None)
        
        try:
            import datetime
            
            dob_str = str(dob_str).strip()
            
            # Remove time component if present (Excel dates sometimes include time)
            if ' ' in dob_str:
                dob_str = dob_str.split()[0]
            
            # Try different parsing strategies
            parsed_date = None
            
            # Strategy 1: Try parsing with slashes (MM/DD/YYYY or MM/DD/YY)
            if '/' in dob_str:
                parts = dob_str.split('/')
                if len(parts) == 3:
                    month = parts[0].strip()
                    day = parts[1].strip()
                    year = parts[2].strip()
                    
                    # Handle 2-digit year
                    if len(year) == 2:
                        year_int = int(year)
                        # Choose century based on current year (pivot window)
                        current_year = datetime.datetime.now().year
                        pivot = (current_year % 100) + 5
                        if pivot >= 100:
                            pivot -= 100
                        if year_int <= pivot:
                            year = f"20{year_int:02d}"
                        else:
                            year = f"19{year_int:02d}"
                    
                    # Try MM/DD/YYYY format
                    try:
                        parsed_date = datetime.datetime(int(year), int(month), int(day))
                    except Exception:
                        # Try YYYY/MM/DD format
                        try:
                            parsed_date = datetime.datetime(int(month), int(day), int(year))
                            # If this worked but year is obviously wrong, swap back
                            if int(month) > 12:
                                # month was actually year
                                parsed_date = datetime.datetime(int(month), int(day), int(year))
                        except:
                            pass
            
            # Strategy 2: Try parsing with dashes (YYYY-MM-DD)
            if not parsed_date and '-' in dob_str:
                try:
                    parts = dob_str.split('-')
                    if len(parts) == 3:
                        # Assume YYYY-MM-DD format
                        parsed_date = datetime.datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                except:
                    pass
            
            # Strategy 3: Try pandas datetime parsing (handles many Excel formats)
            if not parsed_date:
                try:
                    import pandas as pd
                    parsed_date = pd.to_datetime(dob_str)
                    if isinstance(parsed_date, pd.Timestamp):
                        parsed_date = parsed_date.to_pydatetime()
                except:
                    pass
            
            # Strategy 4: Try datetime.strptime with common formats
            if not parsed_date:
                date_formats = [
                    '%m/%d/%Y',
                    '%m/%d/%y',
                    '%Y/%m/%d',
                    '%y/%m/%d',
                    '%d/%m/%Y',
                    '%d/%m/%y',
                    '%Y-%m-%d',
                    '%m-%d-%Y',
                ]
                
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.datetime.strptime(dob_str, fmt)
                        break
                    except:
                        continue
            
            if parsed_date:
                # Prevent future-year DOBs caused by two-digit parsing defaults
                current_year = datetime.datetime.now().year
                if parsed_date.year > current_year:
                    try:
                        parsed_date = parsed_date.replace(year=parsed_date.year - 100)
                    except ValueError:
                        # Handle leap-day edge cases (e.g., Feb 29 -> Feb 28)
                        parsed_date = parsed_date.replace(year=parsed_date.year - 100, day=28)
                # Normalize to YYYY-MM-DD string for comparison
                normalized_str = parsed_date.strftime('%Y-%m-%d')
                return (normalized_str, parsed_date)
            else:
                self.gui_log(f"⚠️ Could not parse DOB: {dob_str}", level="WARNING")
                return (None, None)
                
        except Exception as e:
            self.gui_log(f"Error normalizing DOB '{dob_str}': {e}", level="WARNING")
            return (None, None)
    
    def _search_for_client(self, client_name, client_dob=None, date_of_service=None):
        """Search for a client by name in the Patients search field with smart retry logic
        
        Handles:
        - Middle names/initials (tries full name, then without middle)
        - Removes periods from middle initials
        - DOB matching from dropdown (will be implemented when dropdown element is provided)
        
        Args:
            client_name: The client's name (e.g., "John H. Doe" or "Jane Smith")
            client_dob: The client's date of birth from Excel (for matching in dropdown)
        
        Returns:
            bool: True if search field found and filled successfully, False otherwise
        """
        try:
            self.gui_log(f"Step 2: Searching for client: {client_name} (DOB: {client_dob or 'N/A'})...", level="INFO")
            self.update_status(f"Searching for client: {client_name}...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Store client info early so it's available throughout the navigation chain
            # This must be set BEFORE calling _match_client_in_dropdown() because _match_client_in_dropdown()
            # calls _navigate_to_billing_tab() which needs this value
            self.current_client_name = client_name
            self.current_client_dob = client_dob if client_dob else None
            if date_of_service:
                self.current_date_of_service = date_of_service
                self.gui_log(f"Date of service stored for navigation: {date_of_service}", level="DEBUG")
            else:
                self.current_date_of_service = None
            
            # Normalize client name - get variations
            full_name, name_without_middle = self._normalize_client_name(client_name)
            self.gui_log(f"Name variations - Full: '{full_name}', Without middle: '{name_without_middle}'", level="DEBUG")
            
            # Normalize DOB if provided
            normalized_dob_str = None
            if client_dob:
                normalized_dob_str, _ = self._normalize_dob_for_comparison(client_dob)
                if normalized_dob_str:
                    self.gui_log(f"Normalized DOB for matching: {normalized_dob_str}", level="DEBUG")
            
            # Find search field
            search_field = None
            try:
                self.gui_log("Looking for patient search field (ID: ctl00_BodyContent_TextBoxSearchPatientName)...", level="DEBUG")
                
                search_field = self.wait.until(
                    EC.presence_of_element_located((By.ID, "ctl00_BodyContent_TextBoxSearchPatientName"))
                )
                self.gui_log("✅ Patient search field found", level="DEBUG")
                
            except TimeoutException:
                self.gui_log("Trying alternative method: searching by name attribute...", level="DEBUG")
                try:
                    search_field = self.wait.until(
                        EC.presence_of_element_located((By.NAME, "ctl00$BodyContent$TextBoxSearchPatientName"))
                    )
                    self.gui_log("✅ Patient search field found by name attribute", level="DEBUG")
                except Exception as e:
                    self.log_error("Failed to find patient search field", exception=e, include_traceback=True)
                    self.update_status("Error: Could not find search field", "#dc3545")
                    return False
            
            # Scroll into view
            self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", search_field)
            time.sleep(0.5)
            
            # Try searching with name variations
            search_names = []
            if full_name != name_without_middle:
                # If name has middle initial/name, try full name first, then without middle
                search_names = [full_name, name_without_middle]
            else:
                # Just try the single name
                search_names = [full_name]
            
            for attempt, name_to_search in enumerate(search_names, 1):
                self.gui_log(f"Search attempt {attempt}/{len(search_names)}: '{name_to_search}'", level="INFO")
                
                try:
                    # Clear search field
                    search_field.clear()
                    time.sleep(0.2)  # Small delay to ensure clear is processed
                    
                    # Enter client name
                    search_field.send_keys(name_to_search)
                    self.gui_log(f"✅ Entered name: '{name_to_search}'", level="INFO")
                    
                    # Wait for dropdown menu to appear dynamically (no fixed delay - wait handles it)
                    self.gui_log("Waiting for dropdown menu to appear...", level="DEBUG")
                    
                    # Wait for dropdown container to appear using multiple strategies
                    dropdown_found = False
                    dropdown_container = None
                    
                    # Strategy 1: Find by ID (try capital C first, then lowercase - like consent bot v2)
                    try:
                        # Try ContentBubbleResultsContainer (capital C) - this is what consent bot v2 uses
                        dropdown_container = self.wait.until(
                            EC.presence_of_element_located((By.ID, "ContentBubbleResultsContainer"))
                        )
                        self.gui_log("✅ Dropdown menu appeared (by ID: ContentBubbleResultsContainer)", level="INFO")
                        dropdown_found = True
                    except TimeoutException:
                        self.gui_log("Dropdown not found by ID (ContentBubbleResultsContainer), trying lowercase...", level="DEBUG")
                        # Try lowercase version
                        try:
                            dropdown_container = self.wait.until(
                                EC.presence_of_element_located((By.ID, "contentbubbleresultscontainer"))
                            )
                            self.gui_log("✅ Dropdown menu appeared (by ID: contentbubbleresultscontainer)", level="INFO")
                            dropdown_found = True
                        except TimeoutException:
                            self.gui_log("Dropdown not found by ID (either case), trying alternative methods...", level="DEBUG")
                    
                    # Strategy 2: Find by partial ID match
                    if not dropdown_container:
                        try:
                            # Try finding by partial ID
                            dropdown_container = self.wait.until(
                                EC.presence_of_element_located((By.XPATH, "//div[contains(@id, 'contentbubble') or contains(@id, 'resultscontainer')]"))
                            )
                            self.gui_log("✅ Dropdown menu appeared (by partial ID)", level="INFO")
                            dropdown_found = True
                        except TimeoutException:
                            self.gui_log("Dropdown not found by partial ID...", level="DEBUG")
                    
                    # Strategy 3: Find by class or attribute patterns
                    if not dropdown_container:
                        try:
                            # Look for div elements that might be the dropdown container
                            # Common patterns: contains "dropdown", "results", "bubble", "autocomplete"
                            dropdown_container = self.wait.until(
                                EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'dropdown') or contains(@class, 'results') or contains(@class, 'autocomplete') or contains(@class, 'bubble')]"))
                            )
                            self.gui_log("✅ Dropdown menu appeared (by class pattern)", level="INFO")
                            dropdown_found = True
                        except TimeoutException:
                            self.gui_log("Dropdown not found by class pattern...", level="DEBUG")
                    
                    # Strategy 4: Try finding any visible dropdown-like element after search
                    if not dropdown_container:
                        try:
                            # Wait a bit more and look for any div that appeared after typing
                            time.sleep(2)
                            # Find elements that contain the client name
                            dropdown_candidates = self.driver.find_elements(By.XPATH, f"//*[contains(text(), '{name_to_search.split()[0]}')]")
                            if dropdown_candidates:
                                # Find the parent container
                                for candidate in dropdown_candidates:
                                    parent = candidate.find_element(By.XPATH, "./ancestor::div[contains(@id, 'contentbubble') or contains(@id, 'results')][1]")
                                    if parent:
                                        dropdown_container = parent
                                        self.gui_log("✅ Dropdown menu appeared (by text search)", level="INFO")
                                        dropdown_found = True
                                        break
                        except:
                            pass
                    
                    if not dropdown_found:
                        self.gui_log("⚠️ Dropdown menu did not appear - client may not exist or dropdown may have different structure", level="WARNING")
                        # Continue to next search attempt or return False
                        if attempt < len(search_names):
                            continue  # Try next name variation
                        else:
                            return False  # All attempts failed
                    
                    if dropdown_found:
                        # Find and match client by DOB
                        client_matched = self._match_client_in_dropdown(dropdown_container, normalized_dob_str, client_name)
                        
                        if client_matched == True:
                            self.gui_log(f"✅ Successfully matched and clicked client: {client_name}", level="INFO")
                            
                            # Date of service is already stored at the beginning of this function
                            # No need to set it again here
                            
                            return True
                        elif client_matched == "no_payments":
                            # No payments made - already handled (popup closed, navigated back, client tracked)
                            return "no_payments"
                        elif client_matched == "modifier_correct":
                            # Modifier already correct - already handled (navigated to Patients, client tracked)
                            return "modifier_correct"
                        elif client_matched == "session_medium_missing":
                            return "session_medium_missing"
                        elif client_matched == "payment_mismatch":
                            return "payment_mismatch"
                        else:
                            self.gui_log(f"⚠️ Client not found in dropdown for: {name_to_search}", level="WARNING")
                            # Try next name variation if available
                            if attempt < len(search_names):
                                self.gui_log(f"Retrying with different name variation...", level="INFO")
                                continue
                            else:
                                return False  # All attempts failed
                    
                except Exception as e:
                    self.gui_log(f"Error during search attempt {attempt}: {e}", level="WARNING")
                    if attempt < len(search_names):
                        self.gui_log(f"Retrying with different name variation...", level="INFO")
                        continue
                    else:
                        self.log_error(f"All search attempts failed for client: {client_name}", exception=e, include_traceback=True)
                        return False
            
            # If we get here, all attempts failed
            self.gui_log(f"❌ Could not find client after all search attempts: {client_name}", level="WARNING")
            return False
                
        except Exception as e:
            self.log_error("Unexpected error during client search", exception=e, include_traceback=True)
            return False
    
    def _match_client_in_dropdown(self, dropdown_container, expected_dob_str, client_name):
        """Match client in dropdown by DOB and click the matching option
        
        Args:
            dropdown_container: The dropdown container element (id: contentbubbleresultscontainer)
            expected_dob_str: Normalized DOB string from Excel (format: "YYYY-MM-DD")
            client_name: The client's name for logging
        
        Returns:
            bool: True if client matched and clicked, False otherwise
        """
        try:
            self.gui_log("Matching client in dropdown by DOB...", level="INFO")
            
            # Find all client options in the dropdown
            # The structure will depend on how Therapy Notes displays the dropdown
            # Try multiple strategies to find clickable client items
            client_options = []
            
            # Strategy 1: Find all links/clickable elements within dropdown
            try:
                # IMPORTANT: Therapy Notes dropdown uses .ui-menu-item class for each result
                # This is the same approach used in consent bot v2
                
                # Wait a bit more for results to populate (like consent bot v2 does)
                time.sleep(0.5)
                
                # Strategy 1: Look for .ui-menu-item elements (primary method from consent bot v2)
                result_selectors = [
                    "div.ui-menu-item",
                    ".ui-menu-item",
                    "li.ui-menu-item",
                    "a.ui-menu-item",
                    "*[class*='menu-item']",
                    "div[class*='result']",
                ]
                
                for sel in result_selectors:
                    try:
                        # Search WITHIN the container, not globally (critical!)
                        results = dropdown_container.find_elements(By.CSS_SELECTOR, sel)
                        if results:
                            client_options.extend(results)
                            self.gui_log(f"Found {len(results)} results using selector: {sel}", level="DEBUG")
                            break  # Use first successful selector
                    except Exception as e:
                        self.gui_log(f"Error with selector {sel}: {e}", level="DEBUG")
                        continue
                
                # Strategy 2: If no .ui-menu-item found, try finding any clickable elements in the container
                if not client_options:
                    try:
                        all_elements = dropdown_container.find_elements(By.TAG_NAME, "*")
                        clickable = [el for el in all_elements if el.is_displayed() and el.text.strip() and len(el.text.strip()) > 3]
                        if clickable:
                            client_options.extend(clickable[:10])  # Limit to first 10 to avoid too many
                            self.gui_log(f"Found {len(clickable)} clickable elements in container (fallback)", level="DEBUG")
                    except Exception as e:
                        self.gui_log(f"Error finding clickable elements: {e}", level="DEBUG")
                
                # Strategy 3: Fallback - look for links and divs (original approach)
                if not client_options:
                    try:
                        all_elements = dropdown_container.find_elements(By.TAG_NAME, "a")
                        client_options.extend(all_elements)
                        
                        list_items = dropdown_container.find_elements(By.TAG_NAME, "li")
                        client_options.extend(list_items)
                        
                        # Look for divs with click handlers
                        divs = dropdown_container.find_elements(By.TAG_NAME, "div")
                        for div in divs:
                            onclick = div.get_attribute("onclick")
                            role = div.get_attribute("role")
                            style = div.get_attribute("style") or ""
                            has_text = div.text.strip()
                            
                            if onclick or role == "button" or "cursor:pointer" in style.lower():
                                client_options.append(div)
                            elif has_text and len(has_text) > 3 and re.search(r'\d{1,2}/\d{1,2}/\d{2,4}', has_text):
                                client_options.append(div)
                    except Exception as e:
                        self.gui_log(f"Error in fallback element finding: {e}", level="DEBUG")
                
                # Remove duplicates while preserving order
                seen = set()
                unique_options = []
                for opt in client_options:
                    opt_id = id(opt)
                    if opt_id not in seen:
                        seen.add(opt_id)
                        unique_options.append(opt)
                client_options = unique_options
                
                self.gui_log(f"Found {len(client_options)} potential client option(s) in dropdown", level="DEBUG")
                
            except Exception as e:
                self.gui_log(f"Error finding dropdown options: {e}", level="WARNING")
                import traceback
                self.gui_log(traceback.format_exc(), level="DEBUG")
            
            if not client_options:
                self.gui_log("⚠️ No client options found in dropdown", level="WARNING")
                return False
            
            # Extract and match DOB from each option
            for option_idx, option in enumerate(client_options, 1):
                try:
                    # Get the text content of this option
                    option_text = option.text.strip()
                    
                    if not option_text:
                        continue
                    
                    self.gui_log(f"Checking option {option_idx}: {option_text[:80]}...", level="DEBUG")
                    
                    # Extract DOB from option text
                    # Therapy Notes dropdown format: "Name - 9/18/1955" or "Name (9/18/1955)" or "Name 9/18/1955"
                    # DOB format in dropdown: MM/DD/YYYY (e.g., "9/18/1955")
                    option_dob_str = None
                    
                    # Try to find DOB in the text using various patterns
                    import re
                    
                    # Pattern 1: MM/DD/YYYY or MM/DD/YY (Therapy Notes format: "9/18/1955" or "9/18/55")
                    # This is the primary format in Therapy Notes dropdown
                    dob_pattern1 = r'(\d{1,2}/\d{1,2}/\d{2,4})'
                    match = re.search(dob_pattern1, option_text)
                    if match:
                        dob_in_text = match.group(1)
                        self.gui_log(f"   Found DOB in dropdown text: {dob_in_text}", level="DEBUG")
                        option_dob_str, _ = self._normalize_dob_for_comparison(dob_in_text)
                        if option_dob_str:
                            self.gui_log(f"   Normalized dropdown DOB: {option_dob_str}", level="DEBUG")
                    
                    # Pattern 2: YYYY-MM-DD
                    if not option_dob_str:
                        dob_pattern2 = r'(\d{4}-\d{2}-\d{2})'
                        match = re.search(dob_pattern2, option_text)
                        if match:
                            option_dob_str = match.group(1)  # Already normalized
                    
                    # Pattern 3: Look for dates in parentheses or after dashes
                    if not option_dob_str:
                        # Try extracting from text after last "-" or within parentheses
                        if '-' in option_text:
                            parts = option_text.split('-')
                            if len(parts) > 1:
                                # Last part might contain DOB
                                potential_dob = parts[-1].strip()
                                option_dob_str, _ = self._normalize_dob_for_comparison(potential_dob)
                        
                        if not option_dob_str and '(' in option_text:
                            # Extract text within parentheses
                            paren_match = re.search(r'\(([^)]+)\)', option_text)
                            if paren_match:
                                potential_dob = paren_match.group(1).strip()
                                option_dob_str, _ = self._normalize_dob_for_comparison(potential_dob)
                    
                    if option_dob_str and expected_dob_str:
                        # Compare DOBs
                        if option_dob_str == expected_dob_str:
                            self.gui_log(f"✅ DOB match found! Option: {option_text[:80]}", level="INFO")
                            self.gui_log(f"   Matched DOB: {option_dob_str}", level="INFO")
                            
                            # Scroll option into view
                            self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", option)
                            time.sleep(0.5)
                            
                            # Click the matching option
                            # First, try to find the actual clickable link within the option element
                            clickable_element = None
                            
                            try:
                                # If option is already a link, use it directly
                                if option.tag_name.lower() == "a":
                                    clickable_element = option
                                else:
                                    # Try to find a link within the option element
                                    try:
                                        inner_link = option.find_element(By.TAG_NAME, "a")
                                        if inner_link:
                                            clickable_element = inner_link
                                    except:
                                        pass
                                    
                                    # If no link found, check if the option element itself is clickable
                                    if not clickable_element:
                                        # Check if element has click handlers or is interactive
                                        onclick = option.get_attribute("onclick")
                                        href = option.get_attribute("href")
                                        role = option.get_attribute("role")
                                        
                                        if onclick or href or role in ["button", "link", "option"]:
                                            clickable_element = option
                                        else:
                                            # As last resort, use the option element itself
                                            clickable_element = option
                                
                                if not clickable_element:
                                    self.gui_log("⚠️ Could not find clickable element within option", level="WARNING")
                                    clickable_element = option  # Fallback to original element
                                
                                # Scroll element into view first
                                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", clickable_element)
                                time.sleep(0.3)
                                
                                # Try multiple click strategies
                                clicked = False
                                
                                # Strategy 1: Try regular click
                                try:
                                    clickable_element.click()
                                    self.gui_log(f"✅ Clicked matching client option (regular click)", level="INFO")
                                    clicked = True
                                except Exception as click_error:
                                    self.gui_log(f"Regular click failed, trying JavaScript click: {click_error}", level="DEBUG")
                                
                                # Strategy 2: Try JavaScript click if regular click failed
                                if not clicked:
                                    try:
                                        self.driver.execute_script("arguments[0].click();", clickable_element)
                                        self.gui_log(f"✅ Clicked matching client option (JavaScript)", level="INFO")
                                        clicked = True
                                    except Exception as js_error:
                                        self.gui_log(f"JavaScript click failed, trying ActionChains: {js_error}", level="DEBUG")
                                
                                # Strategy 3: Try ActionChains move_to_element and click
                                if not clicked:
                                    try:
                                        from selenium.webdriver.common.action_chains import ActionChains
                                        actions = ActionChains(self.driver)
                                        actions.move_to_element(clickable_element).click().perform()
                                        self.gui_log(f"✅ Clicked matching client option (ActionChains)", level="INFO")
                                        clicked = True
                                    except Exception as ac_error:
                                        self.gui_log(f"ActionChains click failed: {ac_error}", level="WARNING")
                                
                                if not clicked:
                                    raise Exception("All click strategies failed")
                                
                                # Wait for navigation to complete
                                time.sleep(2)
                                
                                # Verify we navigated to client page
                                current_url = self.driver.current_url
                                self.gui_log(f"Current URL after clicking client: {current_url}", level="DEBUG")
                                
                                # Verify navigation by checking URL changed or waiting for page elements
                                if "patients" in current_url.lower() and "edit" in current_url.lower():
                                    self.gui_log("✅ Verified navigation to client page", level="INFO")
                                
                                # Navigate to Billing tab
                                billing_result = self._navigate_to_billing_tab()
                                if billing_result == True:
                                    self.gui_log("✅ Successfully navigated to Billing tab", level="INFO")
                                    return True
                                elif billing_result == "no_payments":
                                    # No payments made - already handled (popup closed, navigated back, client tracked)
                                    return "no_payments"
                                elif billing_result == "modifier_correct":
                                    # Modifier already correct - already handled (navigated to Patients, client tracked)
                                    return "modifier_correct"
                                elif billing_result == "session_medium_missing":
                                    return "session_medium_missing"
                                elif billing_result == "payment_mismatch":
                                    return "payment_mismatch"
                                else:
                                    self.gui_log("⚠️ Could not navigate to Billing tab", level="WARNING")
                                    # Still return True since client was clicked successfully
                                    return True
                                
                            except Exception as click_error:
                                self.log_error(f"Failed to click client option after all strategies: {click_error}", exception=click_error, include_traceback=True)
                                
                                # Last resort: Try to find any link in the dropdown that matches and click it
                                try:
                                    self.gui_log("Attempting last resort: searching for matching link by text...", level="WARNING")
                                    all_links = dropdown_container.find_elements(By.TAG_NAME, "a")
                                    for link in all_links:
                                        link_text = link.text.strip()
                                        if option_text in link_text or link_text in option_text:
                                            self.driver.execute_script("arguments[0].click();", link)
                                            self.gui_log(f"✅ Clicked client using text match fallback", level="INFO")
                                            time.sleep(2)
                                            return True
                                except Exception as fallback_error:
                                    self.gui_log(f"Fallback click also failed: {fallback_error}", level="ERROR")
                                
                                return False
                        else:
                            self.gui_log(f"   DOB mismatch - Option DOB: {option_dob_str}, Expected: {expected_dob_str}", level="DEBUG")
                    elif not option_dob_str:
                        self.gui_log(f"   Could not extract DOB from option text", level="DEBUG")
                    elif not expected_dob_str:
                        self.gui_log(f"   ⚠️ No expected DOB provided for matching", level="WARNING")
                        # If no DOB provided, might need to select by name only or return False
                        # For now, continue checking other options
                
                except Exception as e:
                    self.gui_log(f"Error checking option {option_idx}: {e}", level="WARNING")
                    continue
            
            # If we get here, no matching client was found
            self.gui_log(f"❌ No matching client found in dropdown with DOB: {expected_dob_str}", level="WARNING")
            return False
            
        except Exception as e:
            self.log_error("Error matching client in dropdown", exception=e, include_traceback=True)
            return False
    
    def _navigate_to_billing_tab(self):
        """Navigate to the Billing tab on the client page
        
        Element: <a data-tab-id="Patient Billing" href="#tab=Patient+Billing">Billing</a>
        
        Returns:
            bool: True if Billing tab clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 3: Navigating to Billing tab...", level="INFO")
            self.update_status("Clicking Billing tab...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for page to fully load after clicking client
            time.sleep(2)
            
            # Find and click Billing tab using multiple strategies
            billing_tab = None
            
            # Strategy 1: Find by data-tab-id attribute
            try:
                self.gui_log("Looking for Billing tab (data-tab-id='Patient Billing')...", level="DEBUG")
                billing_tab = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-tab-id='Patient Billing']"))
                )
                self.gui_log("✅ Found Billing tab by data-tab-id", level="DEBUG")
            except TimeoutException:
                self.gui_log("Billing tab not found by data-tab-id, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by link text "Billing"
            if not billing_tab:
                try:
                    self.gui_log("Trying to find 'Billing' link by text...", level="DEBUG")
                    billing_tab = self.wait.until(
                        EC.element_to_be_clickable((By.LINK_TEXT, "Billing"))
                    )
                    self.gui_log("✅ Found Billing tab by link text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Billing link not found by text, trying partial text...", level="DEBUG")
            
            # Strategy 3: Find by partial link text
            if not billing_tab:
                try:
                    billing_tab = self.wait.until(
                        EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Billing"))
                    )
                    self.gui_log("✅ Found Billing tab by partial link text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Billing link not found by partial text...", level="DEBUG")
            
            # Strategy 4: Find by href containing "Patient+Billing"
            if not billing_tab:
                try:
                    billing_tab = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[contains(@href, 'Patient+Billing') or contains(@href, 'Patient Billing')]"))
                    )
                    self.gui_log("✅ Found Billing tab by href", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Billing tab not found by href...", level="DEBUG")
            
            # Strategy 5: Find by XPath with text and href
            if not billing_tab:
                try:
                    billing_tab = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Billing') and contains(@href, 'tab=')]"))
                    )
                    self.gui_log("✅ Found Billing tab by XPath with text and href", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Billing tab not found by XPath...", level="DEBUG")
            
            if billing_tab:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", billing_tab)
                time.sleep(0.5)
                
                # Click the Billing tab
                try:
                    billing_tab.click()
                    self.gui_log("✅ Billing tab clicked", level="INFO")
                    self.update_status("Billing tab clicked - Waiting for page to load...", "#ff9500")
                    
                    # Wait for tab to load
                    time.sleep(2)
                    
                    # Verify navigation by checking URL or page elements
                    current_url = self.driver.current_url
                    self.gui_log(f"Current URL after clicking Billing tab: {current_url}", level="DEBUG")
                    
                    # Check if URL contains billing indicator or tab is active
                    if "billing" in current_url.lower() or "#tab=" in current_url.lower():
                        self.gui_log("✅ Successfully navigated to Billing tab", level="INFO")
                        self.update_status("On Billing tab", "#28a745")
                        
                        # Click "All Items" button
                        all_items_result = self._click_all_items_button()
                        if all_items_result == True:
                            # After clicking "All Items", click the date of service
                            if hasattr(self, 'current_date_of_service') and self.current_date_of_service:
                                date_result = self._click_date_of_service(self.current_date_of_service)
                                if date_result == True:
                                    return True
                                elif date_result == "no_payments":
                                    # No payments made - already handled (popup closed, navigated back, client tracked)
                                    return "no_payments"
                                elif date_result == "modifier_correct":
                                    # Modifier already correct - already handled (navigated to Patients, client tracked)
                                    return "modifier_correct"
                                elif date_result == "session_medium_missing":
                                    return "session_medium_missing"
                                elif date_result == "payment_mismatch":
                                    return "payment_mismatch"
                                else:
                                    self.gui_log("⚠️ Could not click date of service, but continuing...", level="WARNING")
                                    return True
                            else:
                                self.gui_log("⚠️ No date of service available to click", level="WARNING")
                                return True
                        elif all_items_result == "no_payments":
                            # No payments made - propagate up
                            return "no_payments"
                        elif all_items_result == "modifier_correct":
                            # Modifier already correct - propagate up
                            return "modifier_correct"
                        else:
                            # Still return True as Billing tab was clicked successfully
                            self.gui_log("⚠️ Could not click All Items button, but continuing...", level="WARNING")
                            return True
                    else:
                        self.gui_log("⚠️ URL doesn't show billing tab, but tab may still be active", level="WARNING")
                        # Still return True as tab was clicked
                        return True
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", billing_tab)
                        self.gui_log("✅ Billing tab clicked (JavaScript)", level="INFO")
                        time.sleep(2)
                        return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click Billing tab: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find Billing tab using any method", include_traceback=False)
                self.update_status("Error: Could not find Billing tab", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error navigating to Billing tab", exception=e, include_traceback=True)
            self.update_status(f"Billing tab error: {str(e)}", "#dc3545")
            return False
    
    def _click_all_items_button(self):
        """Click the "All Items" button on the Billing tab
        
        Element: <a tabindex="0" class="quick-search-link" id="SearchBillingTransactionsFilter__AllItems" data-testid="quicksearchfilter-allitems-link">All Items</a>
        
        Returns:
            bool: True if All Items button clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 4: Clicking 'All Items' button...", level="INFO")
            self.update_status("Clicking All Items button...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for Billing tab content to load
            time.sleep(2)
            
            # Find and click All Items button using multiple strategies
            all_items_button = None
            
            # Strategy 1: Find by ID
            try:
                self.gui_log("Looking for All Items button (ID: SearchBillingTransactionsFilter__AllItems)...", level="DEBUG")
                all_items_button = self.wait.until(
                    EC.element_to_be_clickable((By.ID, "SearchBillingTransactionsFilter__AllItems"))
                )
                self.gui_log("✅ Found All Items button by ID", level="DEBUG")
            except TimeoutException:
                self.gui_log("All Items button not found by ID, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by data-testid attribute
            if not all_items_button:
                try:
                    self.gui_log("Trying to find All Items button by data-testid...", level="DEBUG")
                    all_items_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[@data-testid='quicksearchfilter-allitems-link']"))
                    )
                    self.gui_log("✅ Found All Items button by data-testid", level="DEBUG")
                except TimeoutException:
                    self.gui_log("All Items button not found by data-testid, trying text...", level="DEBUG")
            
            # Strategy 3: Find by link text "All Items"
            if not all_items_button:
                try:
                    self.gui_log("Trying to find 'All Items' link by text...", level="DEBUG")
                    all_items_button = self.wait.until(
                        EC.element_to_be_clickable((By.LINK_TEXT, "All Items"))
                    )
                    self.gui_log("✅ Found All Items button by link text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("All Items link not found by text, trying partial text...", level="DEBUG")
            
            # Strategy 4: Find by partial link text
            if not all_items_button:
                try:
                    all_items_button = self.wait.until(
                        EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "All Items"))
                    )
                    self.gui_log("✅ Found All Items button by partial link text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("All Items link not found by partial text...", level="DEBUG")
            
            # Strategy 5: Find by class and text
            if not all_items_button:
                try:
                    all_items_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[@class='quick-search-link' and contains(text(), 'All Items')]"))
                    )
                    self.gui_log("✅ Found All Items button by class and text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("All Items button not found by class and text...", level="DEBUG")
            
            if all_items_button:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", all_items_button)
                time.sleep(0.5)
                
                # Click the All Items button
                try:
                    all_items_button.click()
                    self.gui_log("✅ All Items button clicked", level="INFO")
                    self.update_status("All Items clicked - Waiting for page to load...", "#ff9500")
                    
                    # Wait for content to load
                    time.sleep(2)
                    
                    # Verify click by checking if button appears active or content loaded
                    current_url = self.driver.current_url
                    self.gui_log(f"Current URL after clicking All Items: {current_url}", level="DEBUG")
                    
                    self.gui_log("✅ Successfully clicked All Items button", level="INFO")
                    self.update_status("All Items selected", "#28a745")
                    return True
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", all_items_button)
                        self.gui_log("✅ All Items button clicked (JavaScript)", level="INFO")
                        time.sleep(2)
                        self.update_status("All Items selected", "#28a745")
                        return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click All Items button: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find All Items button using any method", include_traceback=False)
                self.update_status("Error: Could not find All Items button", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error clicking All Items button", exception=e, include_traceback=True)
            self.update_status(f"All Items button error: {str(e)}", "#dc3545")
            return False
    
    def _normalize_date_of_service(self, dos_str):
        """Normalize date of service string to Therapy Notes format for comparison
        
        Therapy Notes format: M/D/YY (e.g., "8/10/25")
        Excel may have various formats: MM/DD/YYYY, YYYY-MM-DD, etc.
        
        Args:
            dos_str: Date of service string from Excel (various formats)
        
        Returns:
            tuple: (normalized_dos_str, dos_datetime) where normalized_dos_str is in "M/D/YY" format
                   and dos_datetime is a datetime object for comparison, or (None, None) if parsing fails
        """
        if not dos_str:
            return None, None
        
        try:
            import re
            from datetime import datetime
            
            dos_str = str(dos_str).strip()
            
            # Handle different Excel date formats
            # Pattern 1: M/D/YY or MM/DD/YY (e.g., "8/10/25", "08/10/25")
            match = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2})$', dos_str)
            if match:
                month, day, year = match.groups()
                # Convert 2-digit year to 4-digit (assuming 2000s)
                year_int = int(year)
                if year_int < 100:
                    year_int = 2000 + year_int
                
                # Create normalized string in M/D/YY format (remove leading zeros from month/day)
                normalized = f"{int(month)}/{int(day)}/{str(year_int)[-2:]}"
                dt = datetime(year_int, int(month), int(day))
                return normalized, dt
            
            # Pattern 2: M/D/YYYY or MM/DD/YYYY (e.g., "8/10/2025", "08/10/2025")
            match = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})$', dos_str)
            if match:
                month, day, year = match.groups()
                year_int = int(year)
                
                # Create normalized string in M/D/YY format
                normalized = f"{int(month)}/{int(day)}/{str(year_int)[-2:]}"
                dt = datetime(year_int, int(month), int(day))
                return normalized, dt
            
            # Pattern 3: YYYY-MM-DD (e.g., "2025-08-10")
            match = re.match(r'(\d{4})-(\d{2})-(\d{2})$', dos_str)
            if match:
                year, month, day = match.groups()
                year_int = int(year)
                
                # Create normalized string in M/D/YY format
                normalized = f"{int(month)}/{int(day)}/{str(year_int)[-2:]}"
                dt = datetime(year_int, int(month), int(day))
                return normalized, dt
            
            # Pattern 4: Try parsing with datetime parser (handles various formats)
            try:
                # Try common formats
                formats = [
                    "%m/%d/%y", "%m/%d/%Y",
                    "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y",
                    "%m-%d-%Y", "%m-%d-%y",
                    "%Y/%m/%d"
                ]
                
                dt = None
                for fmt in formats:
                    try:
                        dt = datetime.strptime(dos_str, fmt)
                        year_int = dt.year
                        
                        # Create normalized string in M/D/YY format
                        normalized = f"{dt.month}/{dt.day}/{str(year_int)[-2:]}"
                        return normalized, dt
                    except ValueError:
                        continue
                
                # If no format matched, try pandas parser if available
                if not dt and EXCEL_AVAILABLE and pd:
                    try:
                        # pandas sometimes uses different formats
                        # Try to parse as-is
                        dt_pd = pd.to_datetime(dos_str, errors='coerce')
                        if pd.notna(dt_pd) and isinstance(dt_pd, pd.Timestamp):
                            dt = dt_pd.to_pydatetime()
                            year_int = dt.year
                            normalized = f"{dt.month}/{dt.day}/{str(year_int)[-2:]}"
                            return normalized, dt
                    except Exception:
                        pass
                        
            except Exception as parse_error:
                self.gui_log(f"Could not parse date of service: {dos_str}, error: {parse_error}", level="DEBUG")
            
            return None, None
            
        except Exception as e:
            self.log_error(f"Error normalizing date of service: {dos_str}", exception=e, include_traceback=False)
            return None, None
    
    def _click_date_of_service(self, date_of_service):
        """Click the date of service link from the list of dates on the Billing tab
        
        Element: <a href="#" data-testid="billingstatementtable-paymentdate-link" data-preferred_tooltip_position="0">8/10/25</a>
        
        Args:
            date_of_service: Date of service string from Excel (various formats)
        
        Returns:
            bool: True if date clicked successfully, False otherwise
        """
        try:
            self.gui_log(f"Step 5: Looking for date of service: {date_of_service}", level="INFO")
            self.update_status(f"Searching for date: {date_of_service}...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Normalize the date of service to Therapy Notes format
            normalized_dos_str, dos_datetime = self._normalize_date_of_service(date_of_service)
            
            if not normalized_dos_str or not dos_datetime:
                self.log_error(f"Could not normalize date of service: {date_of_service}", include_traceback=False)
                self.update_status(f"Error: Invalid date format: {date_of_service}", "#dc3545")
                return False
            
            self.gui_log(f"Normalized date of service to Therapy Notes format: {normalized_dos_str}", level="DEBUG")
            
            # Wait for date links to load
            time.sleep(2)
            
            # Find all date links using multiple strategies
            date_links = []
            
            # Strategy 1: Find by data-testid
            try:
                self.gui_log("Looking for date links by data-testid...", level="DEBUG")
                date_links = self.wait.until(
                    EC.presence_of_all_elements_located((By.XPATH, "//a[@data-testid='billingstatementtable-paymentdate-link']"))
                )
                self.gui_log(f"✅ Found {len(date_links)} date link(s) by data-testid", level="DEBUG")
            except TimeoutException:
                self.gui_log("No date links found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by link text pattern (contains date)
            if not date_links:
                try:
                    self.gui_log("Trying to find date links by href='#' pattern...", level="DEBUG")
                    date_links = self.driver.find_elements(By.XPATH, "//a[@href='#' and contains(@data-testid, 'paymentdate')]")
                    if date_links:
                        self.gui_log(f"✅ Found {len(date_links)} date link(s) by href pattern", level="DEBUG")
                except Exception:
                    pass
            
            # Strategy 3: Find links containing date-like text
            if not date_links:
                try:
                    # Look for links with M/D/YY pattern
                    import re
                    all_links = self.driver.find_elements(By.TAG_NAME, "a")
                    date_pattern = re.compile(r'^\d{1,2}/\d{1,2}/\d{2}$')
                    for link in all_links:
                        link_text = link.text.strip()
                        if date_pattern.match(link_text):
                            date_links.append(link)
                    if date_links:
                        self.gui_log(f"✅ Found {len(date_links)} date link(s) by text pattern", level="DEBUG")
                except Exception as e:
                    self.gui_log(f"Error finding date links by pattern: {e}", level="DEBUG")
            
            if not date_links:
                self.log_error("Could not find any date links on the page", include_traceback=False)
                self.update_status("Error: No date links found", "#dc3545")
                return False
            
            self.gui_log(f"Found {len(date_links)} potential date link(s) to search through", level="INFO")
            
            # Search for matching date
            matched_link = None
            for link_idx, link in enumerate(date_links, 1):
                try:
                    link_text = link.text.strip()
                    self.gui_log(f"Checking date link {link_idx}: '{link_text}'", level="DEBUG")
                    
                    # Normalize the link text date for comparison
                    normalized_link_date, link_datetime = self._normalize_date_of_service(link_text)
                    
                    if normalized_link_date and link_datetime:
                        # Compare normalized strings (M/D/YY format)
                        if normalized_dos_str == normalized_link_date:
                            matched_link = link
                            self.gui_log(f"✅ Date match found! Link text: '{link_text}', Normalized: {normalized_link_date}", level="INFO")
                            break
                        # Also compare datetime objects for extra safety
                        elif dos_datetime.date() == link_datetime.date():
                            matched_link = link
                            self.gui_log(f"✅ Date match found (datetime comparison)! Link text: '{link_text}'", level="INFO")
                            break
                    else:
                        # If normalization failed, try direct text comparison
                        if link_text == date_of_service or link_text == normalized_dos_str:
                            matched_link = link
                            self.gui_log(f"✅ Date match found (direct text)! Link text: '{link_text}'", level="INFO")
                            break
                        
                except Exception as e:
                    self.gui_log(f"Error checking date link {link_idx}: {e}", level="DEBUG")
                    continue
            
            if matched_link:
                # Scroll into view
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", matched_link)
                time.sleep(0.5)
                
                # Click the date link
                try:
                    matched_link.click()
                    self.gui_log(f"✅ Date of service clicked: {matched_link.text.strip()}", level="INFO")
                    self.update_status(f"Date clicked: {matched_link.text.strip()}", "#28a745")
                    
                    # Wait for popup to appear
                    time.sleep(2)
                    
                    # Extract original modifier immediately after popup appears (before navigating to Notes)
                    # This allows us to check if modifier is already correct later and skip the entire flow if so
                    if not hasattr(self, 'current_session_medium') or not self.current_session_medium:
                        # Only extract on first pass (before Notes tab)
                        self._extract_original_modifier()
                        self._extract_primary_policy_info()
                    
                    # Check if this is the first time clicking date (should go to Notes) or second time (should go to History)
                    # If we've already analyzed the session medium, we're on the second pass and should click History
                    if hasattr(self, 'current_session_medium') and self.current_session_medium:
                        if (not getattr(self, 'current_primary_policy_name', None) or
                                self.current_primary_policy_name == 'N/A'):
                            self._extract_primary_policy_info()
                        # Second pass - click History tab
                        history_result = self._click_history_tab()
                        if history_result == True:
                            return True
                        elif history_result == "no_payments":
                            # No payments made - already handled (popup closed, navigated back, client tracked)
                            return "no_payments"
                        elif history_result == "skip_client":
                            return "skip_client"
                        elif history_result == "payment_mismatch":
                            return "payment_mismatch"
                        else:
                            self.gui_log("⚠️ Could not click History tab, but date was clicked successfully", level="WARNING")
                            return True
                    else:
                        # First pass - click Notes tab
                        notes_result = self._click_notes_tab()
                        if notes_result == True:
                            return True
                        elif notes_result == "modifier_correct":
                            # Modifier already correct - already handled (navigated to Patients, client tracked)
                            return "modifier_correct"
                        elif notes_result == "session_medium_missing":
                            return "session_medium_missing"
                        elif notes_result == "skip_client":
                            return "skip_client"
                        else:
                            self.gui_log("⚠️ Could not click Notes tab, but date was clicked successfully", level="WARNING")
                            return True
                    
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", matched_link)
                        self.gui_log(f"✅ Date of service clicked (JavaScript): {matched_link.text.strip()}", level="INFO")
                        self.update_status(f"Date clicked: {matched_link.text.strip()}", "#28a745")
                        time.sleep(2)
                        
                        if not hasattr(self, 'current_session_medium') or not self.current_session_medium:
                            self._extract_original_modifier()
                            self._extract_primary_policy_info()
                        
                        # Check if this is the first time clicking date (should go to Notes) or second time (should go to History)
                        if hasattr(self, 'current_session_medium') and self.current_session_medium:
                            if (not getattr(self, 'current_primary_policy_name', None) or
                                    self.current_primary_policy_name == 'N/A'):
                                self._extract_primary_policy_info()
                            # Second pass - click History tab
                            history_result = self._click_history_tab()
                            if history_result == True:
                                return True
                            elif history_result == "no_payments":
                                # No payments made - already handled (popup closed, navigated back, client tracked)
                                return "no_payments"
                            elif history_result == "skip_client":
                                return "skip_client"
                            elif history_result == "payment_mismatch":
                                return "payment_mismatch"
                            else:
                                self.gui_log("⚠️ Could not click History tab, but date was clicked successfully", level="WARNING")
                                return True
                        else:
                            # First pass - click Notes tab
                            notes_result = self._click_notes_tab()
                            if notes_result == True:
                                return True
                            elif notes_result == "modifier_correct":
                                return "modifier_correct"
                            elif notes_result == "session_medium_missing":
                                return "session_medium_missing"
                            elif notes_result == "skip_client":
                                return "skip_client"
                            else:
                                self.gui_log("⚠️ Could not click Notes tab, but date was clicked successfully", level="WARNING")
                                return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click date link: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error(f"Could not find matching date of service: {date_of_service} (normalized: {normalized_dos_str})", include_traceback=False)
                self.update_status(f"Error: Date not found: {date_of_service}", "#dc3545")
                
                # Log all available dates for debugging
                available_dates = []
                for link in date_links[:10]:  # Show first 10 dates
                    try:
                        available_dates.append(link.text.strip())
                    except:
                        pass
                if available_dates:
                    self.gui_log(f"Available dates on page: {', '.join(available_dates)}", level="DEBUG")
                
                return False
                
        except Exception as e:
            self.log_error("Error clicking date of service", exception=e, include_traceback=True)
            self.update_status(f"Date click error: {str(e)}", "#dc3545")
            return False
    
    def _click_notes_tab(self):
        """Click the Notes tab in the popup that appears after clicking the date of service
        
        Element: <a data-testid="calendarentryviewer-notes-tab" href="#" tabindex="0" style="text-align: center;">Notes</a>
        
        Returns:
            bool: True if Notes tab clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 6: Waiting for popup and looking for Notes tab...", level="INFO")
            self.update_status("Looking for Notes tab in popup...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for popup to fully load
            time.sleep(2)
            
            # Find Notes tab using multiple strategies
            notes_tab = None
            
            # Strategy 1: Find by data-testid
            try:
                self.gui_log("Looking for Notes tab (data-testid='calendarentryviewer-notes-tab')...", level="DEBUG")
                notes_tab = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-testid='calendarentryviewer-notes-tab']"))
                )
                self.gui_log("✅ Found Notes tab by data-testid", level="DEBUG")
            except TimeoutException:
                self.gui_log("Notes tab not found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by link text "Notes"
            if not notes_tab:
                try:
                    self.gui_log("Trying to find 'Notes' link by text...", level="DEBUG")
                    notes_tab = self.wait.until(
                        EC.element_to_be_clickable((By.LINK_TEXT, "Notes"))
                    )
                    self.gui_log("✅ Found Notes tab by link text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Notes link not found by text, trying partial text...", level="DEBUG")
            
            # Strategy 3: Find by partial link text
            if not notes_tab:
                try:
                    notes_tab = self.wait.until(
                        EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Notes"))
                    )
                    self.gui_log("✅ Found Notes tab by partial link text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Notes link not found by partial text...", level="DEBUG")
            
            # Strategy 4: Find by XPath with href="#" and text "Notes"
            if not notes_tab:
                try:
                    notes_tab = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[@href='#' and contains(text(), 'Notes')]"))
                    )
                    self.gui_log("✅ Found Notes tab by XPath with href and text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Notes tab not found by XPath...", level="DEBUG")
            
            # Strategy 5: Find by data-testid pattern containing "notes-tab"
            if not notes_tab:
                try:
                    notes_tab = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[contains(@data-testid, 'notes-tab')]"))
                    )
                    self.gui_log("✅ Found Notes tab by data-testid pattern", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Notes tab not found by data-testid pattern...", level="DEBUG")
            
            if notes_tab:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", notes_tab)
                time.sleep(0.5)
                
                # Click the Notes tab
                try:
                    notes_tab.click()
                    self.gui_log("✅ Notes tab clicked", level="INFO")
                    self.update_status("Notes tab clicked - Waiting for content to load...", "#ff9500")
                    
                    # Wait for tab content to load
                    time.sleep(2)
                    
                    # Verify tab is active (optional - could check for active class or visible content)
                    current_url = self.driver.current_url
                    self.gui_log(f"Current URL after clicking Notes tab: {current_url}", level="DEBUG")
                    
                    self.gui_log("✅ Successfully clicked Notes tab", level="INFO")
                    self.update_status("On Notes tab", "#28a745")
                    
                    # Click Progress Note button
                    progress_result = self._click_progress_note_button()
                    if progress_result == True:
                        return True
                    elif progress_result == "modifier_correct":
                        # Modifier already correct - already handled (navigated to Patients, client tracked)
                        return "modifier_correct"
                    elif progress_result == "session_medium_missing":
                        return "session_medium_missing"
                    else:
                        self.gui_log("⚠️ Could not click Progress Note button, but Notes tab was clicked", level="WARNING")
                        # Still return True since Notes tab was clicked successfully
                        return True
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", notes_tab)
                        self.gui_log("✅ Notes tab clicked (JavaScript)", level="INFO")
                        time.sleep(2)
                        self.update_status("On Notes tab", "#28a745")
                        
                        # Click Progress Note button
                        progress_result = self._click_progress_note_button()
                        if progress_result == True:
                            return True
                        elif progress_result == "modifier_correct":
                            return "modifier_correct"
                        elif progress_result == "session_medium_missing":
                            return "session_medium_missing"
                        else:
                            self.gui_log("⚠️ Could not click Progress Note button, but Notes tab was clicked", level="WARNING")
                            # Still return True since Notes tab was clicked successfully
                            return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click Notes tab: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find Notes tab using any method", include_traceback=False)
                self.update_status("Error: Could not find Notes tab", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error clicking Notes tab", exception=e, include_traceback=True)
            self.update_status(f"Notes tab error: {str(e)}", "#dc3545")
            return False
    
    def _click_history_tab(self):
        """Click the History tab in the popup that appears after clicking the date of service
        
        Element: <a data-testid="calendarentryviewer-history-tab" href="#" tabindex="0" style="text-align: center;">History</a>
        
        Returns:
            bool: True if History tab clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 9: Waiting for popup and looking for History tab...", level="INFO")
            self.update_status("Looking for History tab in popup...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for popup to fully load
            time.sleep(2)
            
            # Find History tab using multiple strategies
            history_tab = None
            
            # Strategy 1: Find by data-testid
            try:
                self.gui_log("Looking for History tab (data-testid='calendarentryviewer-history-tab')...", level="DEBUG")
                history_tab = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-testid='calendarentryviewer-history-tab']"))
                )
                self.gui_log("✅ Found History tab by data-testid", level="DEBUG")
            except TimeoutException:
                self.gui_log("History tab not found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by link text "History"
            if not history_tab:
                try:
                    self.gui_log("Trying to find 'History' link by text...", level="DEBUG")
                    history_tab = self.wait.until(
                        EC.element_to_be_clickable((By.LINK_TEXT, "History"))
                    )
                    self.gui_log("✅ Found History tab by link text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("History link not found by text, trying partial text...", level="DEBUG")
            
            # Strategy 3: Find by partial link text
            if not history_tab:
                try:
                    history_tab = self.wait.until(
                        EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "History"))
                    )
                    self.gui_log("✅ Found History tab by partial link text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("History link not found by partial text...", level="DEBUG")
            
            # Strategy 4: Find by XPath with href="#" and text "History"
            if not history_tab:
                try:
                    history_tab = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[@href='#' and contains(text(), 'History')]"))
                    )
                    self.gui_log("✅ Found History tab by XPath with href and text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("History tab not found by XPath...", level="DEBUG")
            
            # Strategy 5: Find by data-testid pattern containing "history-tab"
            if not history_tab:
                try:
                    history_tab = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[contains(@data-testid, 'history-tab')]"))
                    )
                    self.gui_log("✅ Found History tab by data-testid pattern", level="DEBUG")
                except TimeoutException:
                    self.gui_log("History tab not found by data-testid pattern...", level="DEBUG")
            
            if history_tab:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", history_tab)
                time.sleep(0.5)
                
                # Click the History tab
                try:
                    history_tab.click()
                    self.gui_log("✅ History tab clicked", level="INFO")
                    self.update_status("History tab clicked - Waiting for content to load...", "#ff9500")
                    
                    # Wait for tab content to load
                    time.sleep(2)
                    
                    # Verify tab is active (optional - could check for active class or visible content)
                    current_url = self.driver.current_url
                    self.gui_log(f"Current URL after clicking History tab: {current_url}", level="DEBUG")
                    
                    self.gui_log("✅ Successfully clicked History tab", level="INFO")
                    self.update_status("On History tab", "#28a745")
                    
                    # Click the payment amount link
                    payment_result = self._click_payment_amount_link()
                    if payment_result == True:
                        return True
                    elif payment_result == "no_payments":
                        # No payments made - already handled (popup closed, navigated back, client tracked)
                        return "no_payments"  # Signal to skip client
                    elif payment_result == "skip_client":
                        return "skip_client"
                    elif payment_result == "payment_mismatch":
                        return "payment_mismatch"
                    else:
                        self.gui_log("⚠️ Could not click payment amount link, but History tab was clicked", level="WARNING")
                        # Still return True since History tab was clicked successfully
                        return True
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", history_tab)
                        self.gui_log("✅ History tab clicked (JavaScript)", level="INFO")
                        time.sleep(2)
                        self.update_status("On History tab", "#28a745")
                        
                        # Click the payment amount link
                        payment_result = self._click_payment_amount_link()
                        if payment_result == True:
                            return True
                        elif payment_result == "no_payments":
                            return "no_payments"
                        elif payment_result == "skip_client":
                            return "skip_client"
                        elif payment_result == "payment_mismatch":
                            return "payment_mismatch"
                        else:
                            self.gui_log("⚠️ Could not click payment amount link, but History tab was clicked", level="WARNING")
                            # Still return True since History tab was clicked successfully
                            return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click History tab: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find History tab using any method", include_traceback=False)
                self.update_status("Error: Could not find History tab", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error clicking History tab", exception=e, include_traceback=True)
            self.update_status(f"History tab error: {str(e)}", "#dc3545")
            return False
    
    def _click_payment_amount_link(self):
        """Click the payment amount link in the History tab under the Payments header
        
        Element: <a href="#" data-testid="paymentstable-paymentamount-link-0">$2,750.31 total</a>
        Note: The number at the end may vary (0, 1, 2, etc.)
        
        Returns:
            bool: True if payment amount link clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 10: Looking for payment amount link in History tab...", level="INFO")
            self.update_status("Looking for payment amount link...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for History tab content to fully load
            time.sleep(2)
            
            # Find payment amount link using multiple strategies
            payment_link = None
            matched_payment_description = None
            self.current_primary_payment_description = 'N/A'
            
            primary_policy_name = getattr(self, 'current_primary_policy_name', None)
            if not primary_policy_name or primary_policy_name == 'N/A':
                self.gui_log("❌ Primary policy information unavailable - cannot safely select payment", level="ERROR")
                self.update_status("Error: Missing primary policy info - Skipping client", "#dc3545")
                self._handle_payment_policy_mismatch(reason="Primary policy information unavailable")
                return "payment_mismatch"
            
            try:
                normalized_primary = primary_policy_name.strip().lower()
                self.gui_log(f"Matching payment entries to primary policy '{primary_policy_name}'...", level="INFO")
                description_containers = self.driver.find_elements(
                    By.XPATH,
                    "//div[contains(@data-testid, 'paymentstable-description-container')]"
                )
                for container in description_containers:
                    description_text = container.text.strip()
                    if not description_text:
                        continue
                    normalized_description = " ".join(description_text.split())
                    if normalized_primary in normalized_description.lower():
                        self.gui_log(
                            f"✅ Found payment description matching primary policy: '{normalized_description}'",
                            level="INFO"
                        )
                        try:
                            candidate_link = container.find_element(
                                By.XPATH,
                                ".//a[contains(@data-testid, 'paymentstable-paymentamount-link')]"
                            )
                            payment_link = candidate_link
                            matched_payment_description = normalized_description
                            break
                        except Exception as link_error:
                            self.gui_log(
                                f"Matching description found but payment link missing: {link_error}",
                                level="DEBUG"
                            )
                if not payment_link:
                    self.gui_log(
                        f"❌ No payment entries matched primary policy '{primary_policy_name}'",
                        level="ERROR"
                    )
                    self.update_status("Error: Payment did not match primary policy - Skipping client", "#dc3545")
                    self._handle_payment_policy_mismatch(
                        reason=f"No payment found matching primary policy '{primary_policy_name}'"
                    )
                    return "payment_mismatch"
            except Exception as match_error:
                self.gui_log(f"Error matching payment to primary policy: {match_error}", level="DEBUG")
                self.update_status("Error: Payment match failed - Skipping client", "#dc3545")
                self._handle_payment_policy_mismatch(
                    reason=f"Error while matching payment to policy '{primary_policy_name}'"
                )
                return "payment_mismatch"
            
            if payment_link:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", payment_link)
                time.sleep(0.5)
                
                # Get the link text for logging
                link_text = payment_link.text.strip()
                self.gui_log(f"Payment amount link found: '{link_text}'", level="INFO")
                payment_description_to_store = matched_payment_description or link_text or 'N/A'
                self.current_primary_payment_description = payment_description_to_store
                
                # Click the payment amount link
                try:
                    payment_link.click()
                    self.gui_log(f"✅ Payment amount link clicked: '{link_text}'", level="INFO")
                    self.update_status(f"Payment amount clicked: {link_text}", "#28a745")
                    
                    # Wait for page to load properly - wait for document ready state and any content to appear
                    try:
                        # Use a short wait to check if page has loaded
                        short_wait = WebDriverWait(self.driver, 5)
                        # Wait for page to be in ready state and for body to be present
                        short_wait.until(lambda driver: driver.execute_script("return document.readyState") == "complete")
                        # Give a brief moment for dynamic content to load
                        time.sleep(1)
                        self.gui_log("Page loaded after clicking payment amount", level="DEBUG")
                    except TimeoutException:
                        # If page doesn't fully load in 5 seconds, proceed anyway (might be slow network)
                        self.gui_log("Page load timeout, proceeding anyway...", level="DEBUG")
                        time.sleep(1)
                    
                    # After clicking payment amount, check for View ERA button
                    view_era_result = self._check_and_click_view_era_button()
                    if view_era_result == True:
                        # View ERA button exists and was clicked - flow continues
                        return True
                    elif view_era_result == "skip_client":
                        return "skip_client"
                    else:
                        # View ERA button doesn't exist - will be handled by caller
                        return "skip_client"
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", payment_link)
                        self.gui_log(f"✅ Payment amount link clicked (JavaScript): '{link_text}'", level="INFO")
                        self.update_status(f"Payment amount clicked: {link_text}", "#28a745")
                        
                        # Wait for page to load properly - wait for document ready state and any content to appear
                        try:
                            # Use a short wait to check if page has loaded
                            short_wait = WebDriverWait(self.driver, 5)
                            # Wait for page to be in ready state and for body to be present
                            short_wait.until(lambda driver: driver.execute_script("return document.readyState") == "complete")
                            # Give a brief moment for dynamic content to load
                            time.sleep(1)
                            self.gui_log("Page loaded after clicking payment amount (JavaScript)", level="DEBUG")
                        except TimeoutException:
                            # If page doesn't fully load in 5 seconds, proceed anyway (might be slow network)
                            self.gui_log("Page load timeout, proceeding anyway...", level="DEBUG")
                            time.sleep(1)
                        
                        # After clicking payment amount, check for View ERA button
                        view_era_result = self._check_and_click_view_era_button()
                        if view_era_result == True:
                            # View ERA button exists and was clicked - flow continues
                            return True
                        elif view_era_result == "skip_client":
                            return "skip_client"
                        else:
                            # View ERA button doesn't exist - will be handled by caller
                            return "skip_client"
                    except Exception as js_error:
                        self.log_error(f"Failed to click payment amount link: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                # Should not be reached because missing link handled earlier, but keep safety
                self.log_error("Could not find payment amount link using any method", include_traceback=False)
                self.update_status("Error: Could not find payment amount link", "#dc3545")
                self._handle_payment_policy_mismatch(reason="Payment link not available after matching attempt")
                return "payment_mismatch"
                
        except Exception as e:
            self.log_error("Error clicking payment amount link", exception=e, include_traceback=True)
            self.update_status(f"Payment amount link error: {str(e)}", "#dc3545")
            self._handle_payment_policy_mismatch(reason=f"Exception occurred clicking payment amount link: {e}")
            return "payment_mismatch"
    
    def _check_and_click_view_era_button(self):
        """Check for and click the View ERA button on the payment page
        
        Element: <span>View ERA</span>
        If button doesn't exist, skip client and return to Patients page
        
        Returns:
            bool or str: True if View ERA button found and clicked,
                         "skip_client" if the flow should move to the next client,
                         False if the button doesn't exist or another recoverable issue occurred.
        """
        try:
            self.gui_log("Step 11: Looking for View ERA button...", level="INFO")
            self.update_status("Looking for View ERA button...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Use a shorter wait timeout (2 seconds) for quick element checks instead of 10 seconds
            # This prevents long waits when View ERA button doesn't exist
            quick_wait = WebDriverWait(self.driver, 2)
            
            # Find View ERA button using multiple strategies
            view_era_button = None
            
            # Strategy 1: Find span element with text "View ERA"
            try:
                self.gui_log("Looking for View ERA button (span with text 'View ERA')...", level="DEBUG")
                view_era_button = quick_wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//span[text()='View ERA']"))
                )
                self.gui_log("✅ Found View ERA button by span text", level="DEBUG")
            except TimeoutException:
                self.gui_log("View ERA button not found by span text, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by partial text match
            if not view_era_button:
                try:
                    self.gui_log("Trying to find View ERA button by partial text...", level="DEBUG")
                    view_era_button = quick_wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'View ERA')]"))
                    )
                    self.gui_log("✅ Found View ERA button by partial text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("View ERA button not found by partial text, trying other methods...", level="DEBUG")
            
            # Strategy 3: Find clickable parent element containing the span
            if not view_era_button:
                try:
                    self.gui_log("Trying to find parent element containing 'View ERA' span...", level="DEBUG")
                    # Look for button, link, or div containing the span
                    view_era_button = quick_wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[.//span[contains(text(), 'View ERA')]] | //a[.//span[contains(text(), 'View ERA')]] | //div[.//span[contains(text(), 'View ERA')]]"))
                    )
                    self.gui_log("✅ Found View ERA button by parent element", level="DEBUG")
                except TimeoutException:
                    self.gui_log("View ERA button not found by parent element...", level="DEBUG")
            
            # Strategy 4: Find by case-insensitive text match
            if not view_era_button:
                try:
                    self.gui_log("Trying to find View ERA button with case-insensitive match...", level="DEBUG")
                    view_era_button = quick_wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//span[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'view era')]"))
                    )
                    self.gui_log("✅ Found View ERA button by case-insensitive text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("View ERA button not found by case-insensitive text...", level="DEBUG")
            
            # Strategy 5: Find span and check if parent is clickable
            if not view_era_button:
                try:
                    self.gui_log("Trying to find span and click parent/ancestor...", level="DEBUG")
                    spans = self.driver.find_elements(By.XPATH, "//span[contains(text(), 'View ERA')]")
                    if spans:
                        # Try clicking the span itself or its parent
                        for span in spans:
                            try:
                                # Check if span is clickable
                                if span.is_displayed() and span.is_enabled():
                                    view_era_button = span
                                    self.gui_log("✅ Found View ERA button by span element", level="DEBUG")
                                    break
                                # Otherwise try parent
                                parent = span.find_element(By.XPATH, "..")
                                if parent.is_displayed() and parent.is_enabled():
                                    view_era_button = parent
                                    self.gui_log("✅ Found View ERA button by parent of span", level="DEBUG")
                                    break
                            except:
                                continue
                except Exception as e:
                    self.gui_log(f"Error finding span and parent: {e}", level="DEBUG")
            
            if view_era_button:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", view_era_button)
                time.sleep(0.5)
                
                # Click the View ERA button
                try:
                    view_era_button.click()
                    self.gui_log("✅ View ERA button clicked", level="INFO")
                    self.update_status("View ERA clicked - Continuing flow...", "#28a745")
                    
                    # Wait for popup to appear
                    time.sleep(3)
                    
                    # Click the Patient Filter dropdown in the popup
                    dropdown_result = self._click_patient_filter_dropdown()
                    if dropdown_result == True:
                        return True
                    elif dropdown_result == "skip_client":
                        return "skip_client"
                    else:
                        self.gui_log("⚠️ Could not click Patient Filter dropdown after View ERA", level="WARNING")
                        return False
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", view_era_button)
                        self.gui_log("✅ View ERA button clicked (JavaScript)", level="INFO")
                        time.sleep(3)
                        self.update_status("View ERA clicked - Waiting for popup...", "#ff9500")
                        
                        # Click the Patient Filter dropdown in the popup
                        dropdown_result = self._click_patient_filter_dropdown()
                        if dropdown_result == True:
                            return True
                        elif dropdown_result == "skip_client":
                            return "skip_client"
                        else:
                            self.gui_log("⚠️ Could not click Patient Filter dropdown after View ERA (JavaScript)", level="WARNING")
                            return False
                    except Exception as js_error:
                        self.log_error(f"Failed to click View ERA button: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                # View ERA button doesn't exist - skip client
                self.gui_log("⚠️ View ERA button not found - Client will be skipped", level="WARNING")
                self.update_status("No View ERA button found - Skipping client...", "#ff9500")
                
                # Get current client info for skipping
                client_name = getattr(self, 'current_client_name', 'Unknown')
                dob = getattr(self, 'current_client_dob', '')
                dos = getattr(self, 'current_date_of_service', '')
                
                # Add to skipped clients list (legacy)
                self.skipped_clients.append({
                    'client_name': client_name,
                    'dob': dob,
                    'date_of_service': dos,
                    'reason': 'No ERA in Therapy Notes'
                })
                
                # Add to comprehensive tracking
                self._add_tracked_client(
                    status='Not Refiled - No View ERA',
                    original_modifier='N/A',
                    new_modifier='N/A',
                    modifier_action='N/A',
                    reason='No ERA in Therapy Notes - View ERA button not found',
                    error_message=''
                )
                
                self.gui_log(f"📝 Client '{client_name}' added to skipped list (reason: No ERA in Therapy Notes)", level="INFO")
                
                # Navigate back to Patients page
                if self._navigate_to_patients_from_payment_page():
                    self.gui_log("✅ Navigated back to Patients page - Continuing with next client", level="INFO")
                    return False  # Return False to indicate button doesn't exist
                else:
                    self.gui_log("⚠️ Could not navigate back to Patients page", level="WARNING")
                    return False
                
        except Exception as e:
            self.log_error("Error checking for View ERA button", exception=e, include_traceback=True)
            self.update_status(f"View ERA check error: {str(e)}", "#dc3545")
            return False
    
    def _click_patient_filter_dropdown(self):
        """Click the Patient Filter dropdown in the View ERA popup
        
        Element: <select id="PatientFilterDropdown" style="padding-right: 30px; max-width: 356px;">
        Clicking this opens a dropdown with patient options
        
        Returns:
            bool: True if Patient Filter dropdown clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 12: Looking for Patient Filter dropdown in View ERA popup...", level="INFO")
            self.update_status("Looking for Patient Filter dropdown...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for popup to fully load
            time.sleep(2)
            
            # Find Patient Filter dropdown using multiple strategies
            patient_filter_dropdown = None
            
            # Strategy 1: Find by ID
            try:
                self.gui_log("Looking for Patient Filter dropdown (ID: PatientFilterDropdown)...", level="DEBUG")
                patient_filter_dropdown = self.wait.until(
                    EC.presence_of_element_located((By.ID, "PatientFilterDropdown"))
                )
                self.gui_log("✅ Found Patient Filter dropdown by ID", level="DEBUG")
            except TimeoutException:
                self.gui_log("Patient Filter dropdown not found by ID, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by name attribute
            if not patient_filter_dropdown:
                try:
                    self.gui_log("Trying to find Patient Filter dropdown by name...", level="DEBUG")
                    patient_filter_dropdown = self.wait.until(
                        EC.presence_of_element_located((By.NAME, "PatientFilterDropdown"))
                    )
                    self.gui_log("✅ Found Patient Filter dropdown by name", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Patient Filter dropdown not found by name, trying XPath...", level="DEBUG")
            
            # Strategy 3: Find by XPath with select tag and id
            if not patient_filter_dropdown:
                try:
                    self.gui_log("Trying to find Patient Filter dropdown by XPath (select with id)...", level="DEBUG")
                    patient_filter_dropdown = self.wait.until(
                        EC.presence_of_element_located((By.XPATH, "//select[@id='PatientFilterDropdown']"))
                    )
                    self.gui_log("✅ Found Patient Filter dropdown by XPath", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Patient Filter dropdown not found by XPath...", level="DEBUG")
            
            # Strategy 4: Find any select element containing "Patient" or "Filter"
            if not patient_filter_dropdown:
                try:
                    self.gui_log("Trying to find Patient Filter dropdown by select element pattern...", level="DEBUG")
                    selects = self.driver.find_elements(By.TAG_NAME, "select")
                    for select in selects:
                        select_id = select.get_attribute("id") or ""
                        select_name = select.get_attribute("name") or ""
                        if "patient" in select_id.lower() or "filter" in select_id.lower() or "patient" in select_name.lower():
                            patient_filter_dropdown = select
                            self.gui_log(f"✅ Found Patient Filter dropdown by pattern matching: {select_id or select_name}", level="DEBUG")
                            break
                except Exception as e:
                    self.gui_log(f"Error finding dropdown by pattern: {e}", level="DEBUG")
            
            if patient_filter_dropdown:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", patient_filter_dropdown)
                time.sleep(0.5)
                
                # Check if dropdown is already visible/enabled
                if not patient_filter_dropdown.is_displayed():
                    self.gui_log("⚠️ Patient Filter dropdown is not visible", level="WARNING")
                
                # Click the Patient Filter dropdown to open it
                try:
                    # For select elements, we can click directly or use Select class
                    from selenium.webdriver.support.ui import Select
                    
                    # First, try clicking the select element itself
                    patient_filter_dropdown.click()
                    self.gui_log("✅ Patient Filter dropdown clicked", level="INFO")
                    self.update_status("Patient Filter dropdown clicked - Waiting for options...", "#ff9500")
                    
                    # Wait for dropdown options to appear
                    time.sleep(1.5)
                    
                    # Verify dropdown is open and select the correct patient
                    try:
                        from selenium.webdriver.support.ui import Select
                        select_obj = Select(patient_filter_dropdown)
                        options = select_obj.options
                        self.gui_log(f"✅ Patient Filter dropdown opened - Found {len(options)} option(s)", level="INFO")
                        
                        # Select the correct patient from dropdown
                        select_result = self._select_patient_from_dropdown(select_obj, options)
                        if select_result == True:
                            self.update_status("Correct patient selected from dropdown", "#28a745")
                            return True
                        elif select_result == "skip_client":
                            return "skip_client"
                        elif getattr(self, 'last_patient_dropdown_not_found', False):
                            return self._handle_patient_not_found_in_era_dropdown()
                        else:
                            self.gui_log("⚠️ Could not select patient from dropdown", level="WARNING")
                            self.update_status("Error: Patient not selected", "#dc3545")
                            return False
                    except Exception as verify_error:
                        self.gui_log(f"⚠️ Could not verify dropdown options, but dropdown may still be open: {verify_error}", level="WARNING")
                        # Still return True as dropdown was clicked
                        return True
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", patient_filter_dropdown)
                        self.gui_log("✅ Patient Filter dropdown clicked (JavaScript)", level="INFO")
                        self.update_status("Patient Filter dropdown clicked - Waiting for options...", "#ff9500")
                        time.sleep(1.5)
                        
                        # Verify dropdown is open and select the correct patient
                        try:
                            from selenium.webdriver.support.ui import Select
                            select_obj = Select(patient_filter_dropdown)
                            options = select_obj.options
                            self.gui_log(f"✅ Patient Filter dropdown opened (JavaScript) - Found {len(options)} option(s)", level="INFO")
                            
                            # Select the correct patient from dropdown
                            select_result = self._select_patient_from_dropdown(select_obj, options)
                            if select_result == True:
                                self.update_status("Correct patient selected from dropdown", "#28a745")
                                return True
                            elif select_result == "skip_client":
                                return "skip_client"
                            elif getattr(self, 'last_patient_dropdown_not_found', False):
                                return self._handle_patient_not_found_in_era_dropdown()
                            else:
                                self.gui_log("⚠️ Could not select patient from dropdown", level="WARNING")
                                self.update_status("Error: Patient not selected", "#dc3545")
                                return False
                        except:
                            self.gui_log("⚠️ Could not verify dropdown options, but dropdown may still be open", level="WARNING")
                            return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click Patient Filter dropdown: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find Patient Filter dropdown using any method", include_traceback=False)
                self.update_status("Error: Could not find Patient Filter dropdown", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error clicking Patient Filter dropdown", exception=e, include_traceback=True)
            self.update_status(f"Patient Filter dropdown error: {str(e)}", "#dc3545")
            return False
    
    def _select_patient_from_dropdown(self, select_obj, options):
        """Select the correct patient from the Patient Filter dropdown
        
        Dropdown shows names in format "LAST, FIRST" (e.g., "Marcone, Lorraine")
        Must match against the current client being processed from Excel
        
        Args:
            select_obj: Selenium Select object for the dropdown
            options: List of option elements from the dropdown
        
        Returns:
            bool: True if patient selected successfully, False otherwise
        """
        try:
            import re
            self.gui_log("Step 13: Selecting correct patient from dropdown...", level="INFO")
            
            # Reset tracking flags for dropdown attempts
            self.last_patient_dropdown_not_found = False
            self.last_patient_dropdown_options = []
            self.last_patient_dropdown_search_variants = {}
            
            # Get current client name
            client_name = getattr(self, 'current_client_name', '')
            if not client_name:
                self.gui_log("⚠️ No current client name available for dropdown selection", level="WARNING")
                return False
            
            self.gui_log(f"Looking for patient in dropdown: '{client_name}'", level="INFO")
            
            # Normalize client name to "LAST, FIRST" format for matching
            # Client name from Excel may be in format "First Last" or "First Middle Last"
            # Need to convert to "LAST, FIRST" format to match dropdown
            
            # Parse client name - could be "First Last" or "First Middle Last"
            name_parts = client_name.strip().split()
            
            if len(name_parts) < 2:
                self.gui_log(f"⚠️ Client name '{client_name}' doesn't have enough parts for matching", level="WARNING")
                return False
            
            # Get last name (last part) and first name (first part)
            first_name = name_parts[0].strip().upper()  # Normalize to uppercase
            last_name = name_parts[-1].strip().upper()  # Normalize to uppercase
            middle_name = name_parts[1:-1] if len(name_parts) > 2 else []
            last_name_parts = [part.strip().upper() for part in name_parts[1:]]  # Everything after first name
            
            # Create matching variations (all uppercase to match Therapy Notes format)
            # Format 1: "LAST, FIRST" (standard dropdown format, all uppercase)
            standard_format = f"{last_name}, {first_name}"
            
            # Build last name variants to support combined or multi-word last names
            last_name_variants = set()
            if last_name_parts:
                # Include final token
                last_name_variants.add(last_name_parts[-1])
                # Include combined last name with spaces
                combined_with_space = " ".join(last_name_parts).strip().upper()
                if combined_with_space:
                    last_name_variants.add(combined_with_space)
                    combined_no_space = re.sub(r'[^A-Z]', '', combined_with_space)
                    if combined_no_space:
                        last_name_variants.add(combined_no_space)
                # Include hyphenated variant
                hyphenated = "-".join(last_name_parts).upper()
                if hyphenated:
                    last_name_variants.add(hyphenated)
            # Always include the standard last name token
            last_name_variants.add(last_name)
            
            # Format 2: "LAST, FIRST M" (with middle initial, all uppercase)
            # Note: Therapy Notes may have middle initial even if Excel doesn't
            # So we'll try matching with any single letter middle initial
            with_middle = None
            if middle_name:
                middle_initial = middle_name[0][0].upper() if middle_name[0] else ""
                with_middle = f"{last_name}, {first_name} {middle_initial}"
            
            # Format 3: "LAST, FIRST M." (with middle initial and period, all uppercase)
            with_middle_period = None
            if middle_name:
                middle_initial_period = f"{middle_name[0][0]}." if middle_name[0] else ""
                with_middle_period = f"{last_name}, {first_name} {middle_initial_period}".upper()
            
            # Format 4: Value format "FIRST|MIDDLE|LAST" (all uppercase, pipe-separated)
            value_format = f"{first_name}|{middle_name[0][0].upper() if middle_name and middle_name[0] else ''}|{last_name}" if middle_name else f"{first_name}|{last_name}"
            
            self.gui_log(f"Searching for patient variants:", level="DEBUG")
            self.gui_log(f"  Standard: '{standard_format}'", level="DEBUG")
            if with_middle:
                self.gui_log(f"  With middle: '{with_middle}'", level="DEBUG")
            if with_middle_period:
                self.gui_log(f"  With middle period: '{with_middle_period}'", level="DEBUG")
            self.gui_log(f"  Value format: '{value_format}'", level="DEBUG")
            if len(last_name_variants) > 1:
                self.gui_log(f"  Last name variants: {', '.join(sorted(last_name_variants))}", level="DEBUG")
            
            self.last_patient_dropdown_search_variants = {
                'standard': standard_format,
                'with_middle': with_middle or '',
                'with_middle_period': with_middle_period or '',
                'value_format': value_format,
                'last_name_variants': sorted(last_name_variants)
            }
            
            # Search through dropdown options
            matched_option = None
            matched_value = None
            
            for option in options:
                option_text = option.text.strip().upper()  # Normalize to uppercase for comparison
                option_value = option.get_attribute("value")
                
                # Skip "All patients" option
                if option_value == "" or "all" in option_text.lower():
                    continue
                
                self.gui_log(f"Checking dropdown option: '{option_text}' (value: '{option_value}')", level="DEBUG")
                
                # Strategy 1: Try matching by value format first (most reliable)
                # Value format: "FIRST|MIDDLE|LAST" or "FIRST|LAST"
                # Even if Excel doesn't have middle, try matching value format
                if option_value and "|" in option_value:
                    value_parts = option_value.split("|")
                    # Check if first and last name match (case-insensitive)
                    if len(value_parts) >= 2:
                        value_first = value_parts[0].strip().upper()
                        value_last = value_parts[-1].strip().upper()
                        normalized_value_last = re.sub(r'[^A-Z]', '', value_last)
                        if (
                            value_first == first_name
                            and any(
                                normalized_value_last == re.sub(r'[^A-Z]', '', variant.upper())
                                for variant in last_name_variants
                            )
                        ):
                            matched_option = option
                            matched_value = option_value
                            self.gui_log(f"✅ Found match by value format: '{option_text}' (value: '{option_value}')", level="INFO")
                            break
                
                # Strategy 2: Try exact match of display text (case-insensitive)
                if option_text == standard_format:
                    matched_option = option
                    matched_value = option_value
                    self.gui_log(f"✅ Found exact match: '{option_text}'", level="INFO")
                    break
                
                # Strategy 3: Try match with middle initial (if Excel has middle)
                if with_middle and option_text == with_middle:
                    matched_option = option
                    matched_value = option_value
                    self.gui_log(f"✅ Found match with middle initial: '{option_text}'", level="INFO")
                    break
                
                # Strategy 4: Try match with middle initial and period (if Excel has middle)
                if with_middle_period and option_text == with_middle_period:
                    matched_option = option
                    matched_value = option_value
                    self.gui_log(f"✅ Found match with middle initial and period: '{option_text}'", level="INFO")
                    break
                
                # Strategy 5: Match by last name and first name, ignoring middle initial
                # This handles cases where Therapy Notes has a middle initial but Excel doesn't
                # Format: "LAST, FIRST" or "LAST, FIRST M"
                option_parts = option_text.split(',')
                if len(option_parts) == 2:
                    option_last = option_parts[0].strip()
                    # Get first name (may have middle initial after it, but we'll just take the first word)
                    option_first_part = option_parts[1].strip().split()[0] if option_parts[1].strip().split() else ""
                    
                    # Match if last name and first name match (case-insensitive)
                    normalized_option_last = re.sub(r'[^A-Z]', '', option_last.upper())
                    if (
                        option_first_part.upper() == first_name
                        and any(
                            normalized_option_last == re.sub(r'[^A-Z]', '', variant.upper())
                            for variant in last_name_variants
                        )
                    ):
                        matched_option = option
                        matched_value = option_value
                        self.gui_log(f"✅ Found match by last and first name (ignoring middle): '{option_text}'", level="INFO")
                        break
            
            if matched_option and matched_value:
                # Select the matched patient
                try:
                    select_obj.select_by_value(matched_value)
                    self.gui_log(f"✅ Selected patient from dropdown: '{matched_option.text.strip()}'", level="INFO")
                    self.update_status(f"Patient selected: {matched_option.text.strip()}", "#28a745")
                    
                    # Wait for selection to take effect
                    time.sleep(1)
                    
                    # Click the Filter button to update the list
                    filter_result = self._click_filter_button()
                    if filter_result == True:
                        return True
                    elif filter_result == "skip_client":
                        return "skip_client"
                    else:
                        self.gui_log("⚠️ Could not click Filter button, but patient was selected", level="WARNING")
                        # Still return True since patient was selected
                        return True
                except Exception as select_error:
                    self.log_error(f"Error selecting patient from dropdown: {select_error}", exception=select_error, include_traceback=False)
                    return False
            else:
                # Log available options for debugging
                available_options = [opt.text.strip() for opt in options if opt.get_attribute("value")]
                self.gui_log(f"⚠️ Could not find matching patient in dropdown", level="WARNING")
                self.gui_log(f"  Searched for: '{standard_format}' (and variants)", level="DEBUG")
                self.gui_log(f"  Available options: {', '.join(available_options[:10])}", level="DEBUG")
                self.last_patient_dropdown_options = available_options
                self.last_patient_dropdown_not_found = True
                return False
                
        except Exception as e:
            self.log_error("Error selecting patient from dropdown", exception=e, include_traceback=True)
            return False
    
    def _click_filter_button(self):
        """Click the Filter button to update the list of patients after selecting from dropdown
        
        Element: <input type="button" value="Filter" tabindex="0">
        
        Returns:
            bool: True if Filter button clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 14: Looking for Filter button...", level="INFO")
            self.update_status("Looking for Filter button...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait a moment for the dropdown selection to register
            time.sleep(1)
            
            # Find Filter button using multiple strategies
            filter_button = None
            
            # Strategy 1: Find by value attribute "Filter"
            try:
                self.gui_log("Looking for Filter button (value='Filter')...", level="DEBUG")
                filter_button = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//input[@type='button' and @value='Filter']"))
                )
                self.gui_log("✅ Found Filter button by value attribute", level="DEBUG")
            except TimeoutException:
                self.gui_log("Filter button not found by value, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by type and value containing "Filter"
            if not filter_button:
                try:
                    self.gui_log("Trying to find Filter button by type and value pattern...", level="DEBUG")
                    filter_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@type='button' and contains(@value, 'Filter')]"))
                    )
                    self.gui_log("✅ Found Filter button by value pattern", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Filter button not found by value pattern, trying button tag...", level="DEBUG")
            
            # Strategy 3: Find button element with text "Filter"
            if not filter_button:
                try:
                    self.gui_log("Trying to find Filter button by button tag and text...", level="DEBUG")
                    filter_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Filter')]"))
                    )
                    self.gui_log("✅ Found Filter button by button tag", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Filter button not found by button tag...", level="DEBUG")
            
            # Strategy 4: Find any input button near the Patient Filter dropdown
            if not filter_button:
                try:
                    self.gui_log("Trying to find Filter button near Patient Filter dropdown...", level="DEBUG")
                    # Try to find input button after the PatientFilterDropdown select element
                    filter_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//select[@id='PatientFilterDropdown']//following::input[@type='button' and contains(@value, 'Filter')]"))
                    )
                    self.gui_log("✅ Found Filter button near dropdown", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Filter button not found near dropdown...", level="DEBUG")
            
            # Strategy 5: Find any clickable element with "Filter" text
            if not filter_button:
                try:
                    self.gui_log("Trying to find Filter button by any element with 'Filter' text...", level="DEBUG")
                    filter_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), 'Filter') and (@type='button' or @role='button')]"))
                    )
                    self.gui_log("✅ Found Filter button by text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Filter button not found by text...", level="DEBUG")
            
            if filter_button:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", filter_button)
                time.sleep(0.5)
                
                # Click the Filter button
                try:
                    filter_button.click()
                    self.gui_log("✅ Filter button clicked", level="INFO")
                    self.update_status("Filter button clicked - Waiting for results...", "#ff9500")
                    
                    # Wait for filtered results to appear
                    time.sleep(3)
                    
                    self.gui_log("✅ Successfully clicked Filter button and waited for results", level="INFO")
                    
                    # Verify filtered patient matches current client
                    if self._verify_filtered_patient():
                        self.update_status("Filter applied - Correct patient verified", "#28a745")
                        # Click "Show Claim Details and Remarks" button
                        if self._click_show_claim_details_button():
                            return True
                        else:
                            self.gui_log("⚠️ Could not click Show Claim Details button", level="WARNING")
                            # Still return True since filter was applied correctly
                            return True
                    else:
                        # Patient doesn't match - need to retry filtering
                        self.gui_log("⚠️ Filtered patient doesn't match current client - Retrying filter...", level="WARNING")
                        self.update_status("Incorrect patient - Retrying filter...", "#ff9500")
                        
                        # Retry: Re-select patient from dropdown and click Filter again
                        retry_result = self._retry_patient_filter()
                        if retry_result == True:
                            # Verify again after retry
                            if self._verify_filtered_patient():
                                # Click "Show Claim Details and Remarks" button
                                if self._click_show_claim_details_button():
                                    return True
                        else:
                            if retry_result == "skip_client":
                                return "skip_client"
                            self.gui_log("⚠️ Patient still doesn't match after retry", level="WARNING")
                        
                        return True  # Still return True to continue (may need manual intervention)
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", filter_button)
                        self.gui_log("✅ Filter button clicked (JavaScript)", level="INFO")
                        self.update_status("Filter button clicked - Waiting for results...", "#ff9500")
                        time.sleep(3)
                        
                        # Verify filtered patient matches current client
                        if self._verify_filtered_patient():
                            self.update_status("Filter applied - Correct patient verified", "#28a745")
                            # Click "Show Claim Details and Remarks" button
                            if self._click_show_claim_details_button():
                                return True
                            else:
                                self.gui_log("⚠️ Could not click Show Claim Details button", level="WARNING")
                                return True
                        else:
                            # Patient doesn't match - need to retry filtering
                            self.gui_log("⚠️ Filtered patient doesn't match current client - Retrying filter...", level="WARNING")
                            self.update_status("Incorrect patient - Retrying filter...", "#ff9500")
                            
                            # Retry: Re-select patient from dropdown and click Filter again
                            retry_result = self._retry_patient_filter()
                            if retry_result == True:
                                # Verify again after retry
                                if self._verify_filtered_patient():
                                    # Click "Show Claim Details and Remarks" button
                                    if self._click_show_claim_details_button():
                                        return True
                            else:
                                if retry_result == "skip_client":
                                    return "skip_client"
                            return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click Filter button: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find Filter button using any method", include_traceback=False)
                self.update_status("Error: Could not find Filter button", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error clicking Filter button", exception=e, include_traceback=True)
            self.update_status(f"Filter button error: {str(e)}", "#dc3545")
            return False
    
    def _verify_filtered_patient(self):
        """Verify that the filtered results show the correct patient
        
        Looks for text like "Patient: CONSUELO R PALENCIA - K4025598101" 
        to confirm the correct patient is displayed after filtering
        
        Returns:
            bool: True if correct patient is found, False otherwise
        """
        try:
            self.gui_log("Verifying filtered patient matches current client...", level="INFO")
            
            # Get current client name for matching
            client_name = getattr(self, 'current_client_name', '')
            if not client_name:
                self.gui_log("⚠️ No current client name available for verification", level="WARNING")
                return False
            
            # Parse client name to get components for matching
            name_parts = client_name.strip().split()
            if len(name_parts) < 2:
                self.gui_log(f"⚠️ Client name '{client_name}' doesn't have enough parts for verification", level="WARNING")
                return False
            
            first_name_upper = name_parts[0].strip().upper()
            last_name_upper = name_parts[-1].strip().upper()
            middle_parts = [part.strip().upper() for part in name_parts[1:-1]]
            last_name_parts = [part.strip().upper() for part in name_parts[1:]]  # includes middle + last parts
            
            import re
            
            # Build last-name variants to handle concatenated or hyphenated cases
            last_name_variants = set()
            if last_name_parts:
                combined_with_space = " ".join(last_name_parts).strip().upper()
                combined_no_space = re.sub(r'[^A-Z]', '', combined_with_space)
                hyphenated = "-".join(last_name_parts).upper()
                
                if combined_with_space:
                    last_name_variants.add(combined_with_space)
                if combined_no_space:
                    last_name_variants.add(combined_no_space)
                if hyphenated:
                    last_name_variants.add(hyphenated)
                last_name_variants.add(last_name_parts[-1])
            last_name_variants.add(last_name_upper)
            
            middle_initial = (middle_parts[0][0] if middle_parts and middle_parts[0] else "").upper()
            combined_middle = " ".join(middle_parts).upper() if middle_parts else ""
            
            def normalize_text(value: str) -> str:
                return re.sub(r'[^A-Z0-9]', '', value.upper())
            
            # Generate normalized name variants to search for
            target_variants = set()
            for last_variant in last_name_variants:
                combos = [
                    f"{last_variant}, {first_name_upper}",
                    f"{first_name_upper} {last_variant}",
                    f"{first_name_upper}{last_variant}",
                    f"{last_variant}{first_name_upper}",
                ]
                if combined_middle:
                    combos.extend([
                        f"{last_variant}, {first_name_upper} {combined_middle}",
                        f"{first_name_upper} {combined_middle} {last_variant}",
                    ])
                if middle_initial:
                    combos.extend([
                        f"{last_variant}, {first_name_upper} {middle_initial}",
                        f"{last_variant}, {first_name_upper} {middle_initial}.",
                        f"{first_name_upper} {middle_initial} {last_variant}",
                        f"{first_name_upper} {middle_initial}.{last_variant}",
                    ])
                for combo in combos:
                    if combo:
                        target_variants.add(normalize_text(combo))
            
            if not target_variants:
                target_variants.add(normalize_text(f"{first_name_upper} {last_name_upper}"))
            
            # Get page text to search for patient info
            try:
                page_text = self.driver.find_element(By.TAG_NAME, "body").text
            except:
                # Try alternative method
                page_text = self.driver.page_source
            
            # Look for "Patient:" section with the patient name
            # Format: "Patient: CONSUELO R PALENCIA - K4025598101"
            # We'll search for variations
            
            # Pattern 1: "Patient: FIRST LAST - ID" or "Patient: FIRST MIDDLE LAST - ID"
            patient_pattern1 = re.compile(r'Patient:\s*([^-]+)', re.IGNORECASE)
            matches1 = patient_pattern1.findall(page_text)
            
            # Pattern 2: Look for patient name near "Patient:" text
            patient_sections = []
            for match in matches1:
                patient_sections.append(match.strip())
            
            self.gui_log(f"Found {len(patient_sections)} patient section(s) on page", level="DEBUG")
            
            # Helper to evaluate whether any variant is present in provided text
            def text_contains_target(raw_text: str) -> bool:
                normalized = normalize_text(raw_text)
                for variant in target_variants:
                    if variant and variant in normalized:
                        return True
                return False
            
            # Check each patient section for a match
            for patient_text in patient_sections:
                self.gui_log(f"Checking patient section: '{patient_text[:80]}...'", level="DEBUG")
                if text_contains_target(patient_text):
                    self.gui_log(f"✅ Verified filtered patient matches current client: '{patient_text.strip()}'", level="INFO")
                    return True
            
            # As fallback, scan individual lines from whole page (handles lists like "GENAOGONZALEZ, NAYERIS")
            for line in page_text.splitlines():
                if text_contains_target(line):
                    self.gui_log(f"✅ Verified filtered patient from page text line: '{line.strip()}'", level="INFO")
                    return True
            
            # If no match found
            if patient_sections:
                self.gui_log(f"⚠️ Patient sections found but none match current client '{client_name}'", level="WARNING")
                self.gui_log(f"  Found patient sections: {patient_sections[:3]}", level="DEBUG")
            else:
                self.gui_log(f"⚠️ No patient sections found on page", level="WARNING")
            
            return False
            
        except Exception as e:
            self.log_error("Error verifying filtered patient", exception=e, include_traceback=True)
            return False
    
    def _retry_patient_filter(self):
        """Retry selecting patient from dropdown and clicking Filter button
        
        Returns:
            bool: True if retry was successful, False otherwise
        """
        try:
            self.gui_log("Retrying patient filter selection...", level="INFO")
            
            # Find the Patient Filter dropdown again
            try:
                patient_filter_dropdown = self.wait.until(
                    EC.presence_of_element_located((By.ID, "PatientFilterDropdown"))
                )
            except:
                patient_filter_dropdown = self.driver.find_element(By.XPATH, "//select[@id='PatientFilterDropdown']")
            
            if not patient_filter_dropdown:
                self.gui_log("⚠️ Could not find Patient Filter dropdown for retry", level="WARNING")
                return False
            
            # Select patient from dropdown again
            from selenium.webdriver.support.ui import Select
            select_obj = Select(patient_filter_dropdown)
            options = select_obj.options
            
            select_result = self._select_patient_from_dropdown(select_obj, options)
            if select_result == True:
                # Wait a moment
                time.sleep(1)
                
                # Click Filter button again
                if self._click_filter_button():
                    return True
                else:
                    self.gui_log("⚠️ Could not click Filter button on retry", level="WARNING")
                    return False
            elif getattr(self, 'last_patient_dropdown_not_found', False):
                self._handle_patient_not_found_in_era_dropdown()
                return "skip_client"
            else:
                self.gui_log("⚠️ Could not select patient from dropdown on retry", level="WARNING")
                return False
                
        except Exception as e:
            self.log_error("Error retrying patient filter", exception=e, include_traceback=True)
            return False
    
    def _click_show_claim_details_button(self):
        """Click the "Show Claim Details and Remarks" button after filtering
        
        First clicks "Show Details and Remarks for All Claims" if available,
        then looks for the individual claim's "Show Claim Details and Remarks" button.
        
        Element for All Claims: <div id="DivShowHideAllClaimDetails" data-testid="showhideallclaimdetails-container"><a href="#" data-link-state="collapsed" data-link-group="{&quot;claim&quot;:null}"><span class="show-hide-link-state-label">Show </span>Details and Remarks for All Claims</a></div>
        
        Element for individual claim: <a href="#" data-link-state="collapsed" data-link-group="{&quot;claim&quot;:102829507}" data-testid="show-hide-link-claim-details-and-remarks" tabindex="0"><span class="show-hide-link-state-label">Show </span>Claim Details and Remarks</a>
        
        Returns:
            bool: True if button clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 15: Looking for 'Show Claim Details and Remarks' button...", level="INFO")
            self.update_status("Looking for Show Claim Details button...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for page to stabilize
            time.sleep(2)
            
            # Try to find and click EITHER "Show Details and Remarks for All Claims" OR individual "Show Claim Details and Remarks"
            # NOT BOTH - clicking both will collapse the information
            # Priority: Try "All Claims" first, if not found, try individual claim button
            clicked_button = False
            
            # First, try "Show Details and Remarks for All Claims"
            all_claims_button = None
            
            # Strategy 1: Find by container ID
            try:
                self.gui_log("Checking for 'Show Details and Remarks for All Claims' button (by ID)...", level="DEBUG")
                all_claims_container = self.driver.find_elements(By.ID, "DivShowHideAllClaimDetails")
                if all_claims_container:
                    # Find the link within the container
                    all_claims_button = all_claims_container[0].find_element(By.TAG_NAME, "a")
                    if all_claims_button and "Details and Remarks for All Claims" in all_claims_button.text:
                        self.gui_log("✅ Found 'Show Details and Remarks for All Claims' button by ID", level="DEBUG")
            except Exception as e:
                self.gui_log(f"Could not find 'All Claims' button by ID: {e}", level="DEBUG")
            
            # Strategy 2: Find by data-testid
            if not all_claims_button:
                try:
                    self.gui_log("Checking for 'All Claims' button (by data-testid)...", level="DEBUG")
                    all_claims_container = self.driver.find_elements(By.XPATH, "//div[@data-testid='showhideallclaimdetails-container']")
                    if all_claims_container:
                        all_claims_button = all_claims_container[0].find_element(By.TAG_NAME, "a")
                        if all_claims_button:
                            self.gui_log("✅ Found 'Show Details and Remarks for All Claims' button by data-testid", level="DEBUG")
                except Exception as e:
                    self.gui_log(f"Could not find 'All Claims' button by data-testid: {e}", level="DEBUG")
            
            # Strategy 3: Find by text "Details and Remarks for All Claims"
            if not all_claims_button:
                try:
                    self.gui_log("Checking for 'All Claims' button (by text)...", level="DEBUG")
                    all_claims_button = self.driver.find_element(By.XPATH, "//a[contains(text(), 'Details and Remarks for All Claims')]")
                    if all_claims_button:
                        self.gui_log("✅ Found 'Show Details and Remarks for All Claims' button by text", level="DEBUG")
                except Exception:
                    self.gui_log("Could not find 'All Claims' button by text", level="DEBUG")
            
            # If "All Claims" button found, click it and proceed with extraction (don't look for individual claim button)
            if all_claims_button:
                try:
                    self.gui_log("Clicking 'Show Details and Remarks for All Claims'...", level="INFO")
                    self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", all_claims_button)
                    time.sleep(0.5)
                    all_claims_button.click()
                    self.gui_log("✅ Clicked 'Show Details and Remarks for All Claims'", level="INFO")
                    # Wait for all claims to expand
                    time.sleep(2)
                    clicked_button = True
                    
                    # Proceed with Payer Claim Control # extraction
                    self.update_status("Claim Details expanded - Searching for Payer Claim Control #...", "#ff9500")
                    payer_claim_control = self._find_and_extract_payer_claim_control()
                    
                    if payer_claim_control:
                        self.gui_log(f"✅ Payer Claim Control # found: {payer_claim_control}", level="INFO")
                        self.current_payer_claim_control = payer_claim_control  # Store for later use
                        self.update_status(f"Payer Claim Control #: {payer_claim_control}", "#28a745")
                        
                        # Close the popup by clicking X button
                        if self._close_era_popup():
                            # Navigate back to previous page using Alt+Left Arrow (with fallback)
                            navigation_success = False
                            try:
                                if self._navigate_back_using_keyboard():
                                    navigation_success = True
                            except Exception as nav_error:
                                self.gui_log(f"⚠️ Keyboard navigation failed: {nav_error} - trying fallback...", level="WARNING")
                            
                            if not navigation_success:
                                # If keyboard shortcut failed, try full navigation fallback
                                self.gui_log("⚠️ Keyboard navigation failed - trying full navigation fallback...", level="WARNING")
                                if not self._navigate_back_via_patients_search():
                                    self.gui_log("⚠️ Both navigation methods failed, but Payer Claim Control # was found", level="WARNING")
                                    return False
                            
                            # Ensure we're on the correct page with "All Items" selected
                            if self._ensure_all_items_selected():
                                # Click the date of service again
                                if hasattr(self, 'current_date_of_service') and self.current_date_of_service:
                                    if self._click_date_of_service_again():
                                        return True
                                    else:
                                        self.gui_log("⚠️ Could not click date of service again, but continuing...", level="WARNING")
                                        return True
                                else:
                                    self.gui_log("⚠️ No date of service available to click", level="WARNING")
                                    return True
                            else:
                                self.gui_log("⚠️ Could not ensure All Items is selected, but navigation completed", level="WARNING")
                                return True
                        else:
                            self.gui_log("⚠️ Could not close popup, but Payer Claim Control # was found", level="WARNING")
                            # Still try to navigate back
                            if self._navigate_back_using_keyboard():
                                if self._ensure_all_items_selected():
                                    if hasattr(self, 'current_date_of_service') and self.current_date_of_service:
                                        self._click_date_of_service_again()
                            else:
                                # Try fallback if keyboard failed
                                self._navigate_back_via_patients_search()
                            return True
                    else:
                        # Payer Claim Control # is REQUIRED - skip client if not found
                        self.gui_log("❌ Payer Claim Control # not found - Client will be skipped (REQUIRED)", level="ERROR")
                        self.update_status("Error: Payer Claim Control # not found - Skipping client...", "#dc3545")
                        
                        # Add to skipped clients list
                        client_name = getattr(self, 'current_client_name', 'Unknown')
                        dob = getattr(self, 'current_client_dob', '')
                        dos = getattr(self, 'current_date_of_service', '')
                        
                        self.skipped_clients.append({
                            'client_name': client_name,
                            'dob': dob,
                            'date_of_service': dos,
                            'reason': 'Unable to locate Payer Claim Control #'
                        })
                        
                        # Add to comprehensive tracking
                        self._add_tracked_client(
                            status='Not Refiled - Missing Payer Claim Control',
                            original_modifier='N/A',
                            new_modifier='N/A',
                            modifier_action='N/A',
                            reason='Unable to locate Payer Claim Control # in ERA - Cannot proceed with modifier check',
                            error_message='Payer Claim Control # not found'
                        )
                        
                        self.gui_log(f"📝 Client '{client_name}' added to skipped list (reason: Unable to locate Payer Claim Control #)", level="INFO")
                        
                        # Close popup (try even if failed)
                        self._close_era_popup()
                        time.sleep(2)
                        
                        # Navigate back to Patients page to continue with next client
                        if self._navigate_to_patients_from_payment_page():
                            self.gui_log("✅ Navigated back to Patients page - Continuing with next client", level="INFO")
                            return "skip_client"
                        else:
                            self.gui_log("⚠️ Could not navigate back to Patients page", level="WARNING")
                            return "skip_client"  # Still return skip_client
                    
                except Exception as e:
                    self.gui_log(f"⚠️ Could not click 'All Claims' button: {e}, trying individual claim button...", level="WARNING")
            
            # Only look for individual "Show Claim Details and Remarks" button if "All Claims" button was NOT found/clicked
            if not clicked_button:
                self.gui_log("'All Claims' button not found - Looking for individual 'Show Claim Details and Remarks' button...", level="INFO")
                
                claim_details_button = None
                
                # Strategy 1: Find by data-testid
                try:
                    self.gui_log("Looking for Show Claim Details button (data-testid='show-hide-link-claim-details-and-remarks')...", level="DEBUG")
                    claim_details_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[@data-testid='show-hide-link-claim-details-and-remarks']"))
                    )
                    self.gui_log("✅ Found Show Claim Details button by data-testid", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Show Claim Details button not found by data-testid, trying alternative methods...", level="DEBUG")
                
                # Strategy 2: Find by text containing "Claim Details and Remarks"
                if not claim_details_button:
                    try:
                        self.gui_log("Trying to find Show Claim Details button by text...", level="DEBUG")
                        claim_details_button = self.wait.until(
                            EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Claim Details and Remarks')]"))
                        )
                        self.gui_log("✅ Found Show Claim Details button by text", level="DEBUG")
                    except TimeoutException:
                        self.gui_log("Show Claim Details button not found by text, trying partial text...", level="DEBUG")
                
                # Strategy 3: Find by partial text "Claim Details"
                if not claim_details_button:
                    try:
                        claim_details_button = self.wait.until(
                            EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Claim Details')]"))
                        )
                        self.gui_log("✅ Found Show Claim Details button by partial text", level="DEBUG")
                    except TimeoutException:
                        self.gui_log("Show Claim Details button not found by partial text...", level="DEBUG")
                
                # Strategy 4: Find by data-link-group containing "claim"
                if not claim_details_button:
                    try:
                        self.gui_log("Trying to find Show Claim Details button by data-link-group...", level="DEBUG")
                        claim_details_button = self.wait.until(
                            EC.element_to_be_clickable((By.XPATH, "//a[contains(@data-link-group, 'claim') and contains(text(), 'Claim Details')]"))
                        )
                        self.gui_log("✅ Found Show Claim Details button by data-link-group", level="DEBUG")
                    except TimeoutException:
                        self.gui_log("Show Claim Details button not found by data-link-group...", level="DEBUG")
                
                # Strategy 5: Find by span with "Show" text followed by "Claim Details"
                if not claim_details_button:
                    try:
                        self.gui_log("Trying to find Show Claim Details button by span structure...", level="DEBUG")
                        claim_details_button = self.wait.until(
                            EC.element_to_be_clickable((By.XPATH, "//a[.//span[contains(text(), 'Show')] and contains(text(), 'Claim Details and Remarks')]"))
                        )
                        self.gui_log("✅ Found Show Claim Details button by span structure", level="DEBUG")
                    except TimeoutException:
                        self.gui_log("Show Claim Details button not found by span structure...", level="DEBUG")
                
                # Click the button if found
                if claim_details_button:
                    # Scroll into view if needed
                    self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", claim_details_button)
                    time.sleep(0.5)
                    
                    # Click the Show Claim Details button
                    try:
                        claim_details_button.click()
                        self.gui_log("✅ Show Claim Details and Remarks button clicked", level="INFO")
                        self.update_status("Show Claim Details clicked - Waiting for details to expand...", "#ff9500")
                        
                        # Wait for details to expand
                        time.sleep(2)
                        
                        self.gui_log("✅ Successfully clicked Show Claim Details and Remarks button", level="INFO")
                        self.update_status("Claim Details expanded - Searching for Payer Claim Control #...", "#ff9500")
                        
                        # Find and extract Payer Claim Control # (with scrolling if needed)
                        # THIS IS REQUIRED - if not found, skip client
                        payer_claim_control = self._find_and_extract_payer_claim_control()
                        clicked_button = True  # Mark that we clicked a button
                    except Exception as click_error:
                        # Try JavaScript click if regular click fails
                        try:
                            self.driver.execute_script("arguments[0].click();", claim_details_button)
                            self.gui_log("✅ Show Claim Details and Remarks button clicked (JavaScript)", level="INFO")
                            self.update_status("Show Claim Details clicked - Waiting for details to expand...", "#ff9500")
                            
                            # Wait for details to expand
                            time.sleep(2)
                            
                            self.gui_log("✅ Successfully clicked Show Claim Details and Remarks button (JavaScript)", level="INFO")
                            self.update_status("Claim Details expanded - Searching for Payer Claim Control #...", "#ff9500")
                            
                            # Find and extract Payer Claim Control # (with scrolling if needed)
                            # THIS IS REQUIRED - if not found, skip client
                            payer_claim_control = self._find_and_extract_payer_claim_control()
                            clicked_button = True  # Mark that we clicked a button
                        except Exception as js_error:
                            self.log_error(f"Failed to click Show Claim Details button (both regular and JavaScript): {click_error}", exception=js_error, include_traceback=False)
                
                # If individual claim button was clicked successfully, proceed with extraction
                if clicked_button:
                    # Find and extract Payer Claim Control # (with scrolling if needed)
                    # THIS IS REQUIRED - if not found, skip client
                    if 'payer_claim_control' not in locals():
                        payer_claim_control = self._find_and_extract_payer_claim_control()
                    
                    if payer_claim_control:
                        self.gui_log(f"✅ Payer Claim Control # found: {payer_claim_control}", level="INFO")
                        self.current_payer_claim_control = payer_claim_control  # Store for later use
                        self.update_status(f"Payer Claim Control #: {payer_claim_control}", "#28a745")
                        
                        # Close the popup by clicking X button
                        if self._close_era_popup():
                            # Navigate back to previous page using Alt+Left Arrow (with fallback)
                            navigation_success = False
                            try:
                                if self._navigate_back_using_keyboard():
                                    navigation_success = True
                            except Exception as nav_error:
                                self.gui_log(f"⚠️ Keyboard navigation failed: {nav_error} - trying fallback...", level="WARNING")
                            
                            if not navigation_success:
                                # If keyboard shortcut failed, try full navigation fallback
                                self.gui_log("⚠️ Keyboard navigation failed - trying full navigation fallback...", level="WARNING")
                                if not self._navigate_back_via_patients_search():
                                    self.gui_log("⚠️ Both navigation methods failed, but Payer Claim Control # was found", level="WARNING")
                                    return False
                            
                            # Ensure we're on the correct page with "All Items" selected
                            if self._ensure_all_items_selected():
                                # Click the date of service again
                                if hasattr(self, 'current_date_of_service') and self.current_date_of_service:
                                    if self._click_date_of_service_again():
                                        return True
                                    else:
                                        self.gui_log("⚠️ Could not click date of service again, but continuing...", level="WARNING")
                                        return True
                                else:
                                    self.gui_log("⚠️ No date of service available to click", level="WARNING")
                                    return True
                            else:
                                self.gui_log("⚠️ Could not ensure All Items is selected, but navigation completed", level="WARNING")
                                return True
                        else:
                            self.gui_log("⚠️ Could not close popup, but Payer Claim Control # was found", level="WARNING")
                            # Still try to navigate back
                            navigation_success = False
                            try:
                                if self._navigate_back_using_keyboard():
                                    navigation_success = True
                            except Exception as nav_error:
                                self.gui_log(f"⚠️ Keyboard navigation failed: {nav_error} - trying fallback...", level="WARNING")
                                # Try full navigation fallback
                                if self._navigate_back_via_patients_search():
                                    navigation_success = True
                            
                            if not navigation_success:
                                # Try fallback if keyboard failed
                                self._navigate_back_via_patients_search()
                            if self._ensure_all_items_selected():
                                if hasattr(self, 'current_date_of_service') and self.current_date_of_service:
                                    self._click_date_of_service_again()
                            return True
                    else:
                        # Payer Claim Control # is REQUIRED - skip client if not found
                        self.gui_log("❌ Payer Claim Control # not found - Client will be skipped (REQUIRED)", level="ERROR")
                        self.update_status("Error: Payer Claim Control # not found - Skipping client...", "#dc3545")
                        
                        # Add to skipped clients list
                        client_name = getattr(self, 'current_client_name', 'Unknown')
                        dob = getattr(self, 'current_client_dob', '')
                        dos = getattr(self, 'current_date_of_service', '')
                        
                        self.skipped_clients.append({
                            'client_name': client_name,
                            'dob': dob,
                            'date_of_service': dos,
                            'reason': 'Unable to locate Payer Claim Control #'
                        })
                        
                        # Add to comprehensive tracking
                        self._add_tracked_client(
                            status='Not Refiled - Missing Payer Claim Control',
                            original_modifier='N/A',
                            new_modifier='N/A',
                            modifier_action='N/A',
                            reason='Unable to locate Payer Claim Control # in ERA - Cannot proceed with modifier check',
                            error_message='Payer Claim Control # not found'
                        )
                        
                        self.gui_log(f"📝 Client '{client_name}' added to skipped list (reason: Unable to locate Payer Claim Control #)", level="INFO")
                        
                        # Close popup (try even if failed)
                        self._close_era_popup()
                        time.sleep(2)
                        
                        # Navigate back to Patients page to continue with next client
                        if self._navigate_to_patients_from_payment_page():
                            self.gui_log("✅ Navigated back to Patients page - Continuing with next client", level="INFO")
                            return "skip_client"
                        else:
                            self.gui_log("⚠️ Could not navigate back to Patients page", level="WARNING")
                            return "skip_client"  # Still return skip_client
                
                else:
                    # No button found at all
                    self.log_error("Could not find Show Claim Details and Remarks button (neither 'All Claims' nor individual)", include_traceback=False)
                    self.update_status("Error: Could not find Show Claim Details button", "#dc3545")
                    return False
                
        except Exception as e:
            self.log_error("Error clicking Show Claim Details and Remarks button", exception=e, include_traceback=True)
            self.update_status(f"Show Claim Details button error: {str(e)}", "#dc3545")
            return False
    
    def _find_and_extract_payer_claim_control(self):
        """Find and extract the Payer Claim Control # from the expanded claim details
        
        Uses element-based search by data-testid attributes:
        - Label element: data-testid="eraviewer-payerclaimcontrolnumber-label"
        - Value element: data-testid="payerclaimcontrolnumber-container-[number]" (number varies)
        
        This approach is faster and more reliable than text scrolling.
        
        Returns:
            str: The Payer Claim Control number if found, None otherwise
        """
        try:
            self.gui_log("Step 16: Searching for Payer Claim Control #...", level="INFO")
            self.update_status("Searching for Payer Claim Control #...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return None
            
            # Wait for claim details to fully load
            time.sleep(2)
            
            # Try element-based search first (faster and more reliable)
            payer_claim_control = self._search_for_payer_claim_control_by_elements()
            
            if payer_claim_control:
                self.gui_log(f"✅ Found Payer Claim Control # via elements: {payer_claim_control}", level="INFO")
                return payer_claim_control
            
            # Fallback to text-based search if element search fails
            self.gui_log("Element-based search failed - Trying text-based search...", level="DEBUG")
            payer_claim_control = self._search_for_payer_claim_control_text()
            
            if payer_claim_control:
                self.gui_log(f"✅ Found Payer Claim Control # via text search: {payer_claim_control}", level="INFO")
                return payer_claim_control
            
            # If still not found, try scrolling as last resort
            self.gui_log("Text search failed - Trying scroll search as last resort...", level="DEBUG")
            payer_claim_control = self._search_for_payer_claim_control_with_scrolling()
            
            if payer_claim_control:
                self.gui_log(f"✅ Found Payer Claim Control # after scrolling: {payer_claim_control}", level="INFO")
                return payer_claim_control
            else:
                self.log_error("Could not find Payer Claim Control # using any method", include_traceback=False)
                self.update_status("Error: Payer Claim Control # not found", "#dc3545")
                return None
                
        except Exception as e:
            self.log_error("Error finding Payer Claim Control #", exception=e, include_traceback=True)
            return None
    
    def _close_era_popup(self):
        """Close the ERA popup by clicking the X button
        
        Element: <i class="fa6 fa-lg fa-2x fa-times"></i>
        
        Returns:
            bool: True if popup closed successfully, False otherwise
        """
        try:
            self.gui_log("Step 17: Closing ERA popup...", level="INFO")
            self.update_status("Closing ERA popup...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait a moment before closing
            time.sleep(1)
            
            # Find X button using multiple strategies
            close_button = None
            
            # Strategy 1: Find by class combination
            try:
                self.gui_log("Looking for X button by class (fa-times)...", level="DEBUG")
                close_button = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//i[contains(@class, 'fa-times')]"))
                )
                self.gui_log("✅ Found X button by fa-times class", level="DEBUG")
            except TimeoutException:
                self.gui_log("X button not found by fa-times, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by exact class combination
            if not close_button:
                try:
                    self.gui_log("Trying to find X button by exact class combination...", level="DEBUG")
                    close_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//i[contains(@class, 'fa6') and contains(@class, 'fa-times')]"))
                    )
                    self.gui_log("✅ Found X button by exact class combination", level="DEBUG")
                except TimeoutException:
                    self.gui_log("X button not found by exact class, trying other methods...", level="DEBUG")
            
            # Strategy 3: Find by class containing "times" or "close"
            if not close_button:
                try:
                    self.gui_log("Trying to find X button by class containing 'times'...", level="DEBUG")
                    close_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//i[contains(@class, 'times') or contains(@class, 'close')]"))
                    )
                    self.gui_log("✅ Found X button by class containing 'times'", level="DEBUG")
                except TimeoutException:
                    self.gui_log("X button not found by class pattern...", level="DEBUG")
            
            # Strategy 4: Find clickable parent element containing the icon
            if not close_button:
                try:
                    self.gui_log("Trying to find parent element containing fa-times icon...", level="DEBUG")
                    close_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[.//i[contains(@class, 'fa-times')]] | //a[.//i[contains(@class, 'fa-times')]] | //div[.//i[contains(@class, 'fa-times')]]"))
                    )
                    self.gui_log("✅ Found X button by parent element", level="DEBUG")
                except TimeoutException:
                    self.gui_log("X button not found by parent element...", level="DEBUG")
            
            # Strategy 5: Find by aria-label or title containing "close"
            if not close_button:
                try:
                    self.gui_log("Trying to find X button by aria-label or title...", level="DEBUG")
                    close_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//*[@aria-label[contains(., 'close') or contains(., 'Close')]] | //*[@title[contains(., 'close') or contains(., 'Close')]]"))
                    )
                    self.gui_log("✅ Found X button by aria-label/title", level="DEBUG")
                except TimeoutException:
                    self.gui_log("X button not found by aria-label/title...", level="DEBUG")
            
            if close_button:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", close_button)
                time.sleep(0.5)
                
                # Click the X button
                try:
                    close_button.click()
                    self.gui_log("✅ X button clicked - Popup closing...", level="INFO")
                    self.update_status("Closing popup...", "#ff9500")
                    
                    # Wait for popup to close
                    time.sleep(2)
                    
                    self.gui_log("✅ Successfully closed ERA popup", level="INFO")
                    self.update_status("Popup closed", "#28a745")
                    return True
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", close_button)
                        self.gui_log("✅ X button clicked (JavaScript) - Popup closing...", level="INFO")
                        time.sleep(2)
                        self.update_status("Popup closed", "#28a745")
                        return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click X button: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find X button using any method", include_traceback=False)
                self.update_status("Error: Could not find X button", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error closing ERA popup", exception=e, include_traceback=True)
            self.update_status(f"Close popup error: {str(e)}", "#dc3545")
            return False
    
    def _navigate_back_using_keyboard(self):
        """Navigate back to the previous page using Alt+Left Arrow keyboard shortcut
        
        This is equivalent to clicking the browser's back button
        
        Returns:
            bool: True if navigation back was successful, False otherwise
        """
        try:
            self.gui_log("Step 18: Navigating back to previous page using Alt+Left Arrow...", level="INFO")
            self.update_status("Navigating back...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait a moment before navigating
            time.sleep(1)
            
            # Get the current URL before navigating back
            try:
                current_url_before = self.driver.current_url
                self.gui_log(f"Current URL before navigation: {current_url_before}", level="DEBUG")
            except:
                current_url_before = None
            
            # Send Alt+Left Arrow keyboard shortcut
            try:
                # Find the body element to send keys to
                body_element = self.driver.find_element(By.TAG_NAME, "body")
                
                # Send Alt+Left Arrow
                # Use ActionChains for key combinations
                actions = ActionChains(self.driver)
                actions.key_down(Keys.ALT).send_keys(Keys.ARROW_LEFT).key_up(Keys.ALT).perform()
                
                self.gui_log("✅ Sent Alt+Left Arrow keyboard shortcut", level="INFO")
                
                # Wait for navigation to complete
                time.sleep(3)
                
                # Verify navigation by checking URL change
                current_url_after = self.driver.current_url
                self.gui_log(f"Current URL after navigation: {current_url_after}", level="DEBUG")
                
                if current_url_before and current_url_after != current_url_before:
                    self.gui_log("✅ Successfully navigated back to previous page", level="INFO")
                    self.update_status("Navigated back to dates page", "#28a745")
                    return True
                else:
                    # URL didn't change - keyboard shortcut didn't work
                    self.gui_log("⚠️ URL didn't change - keyboard shortcut didn't work", level="WARNING")
                    # Don't return True - raise exception to trigger fallback
                    raise Exception("Keyboard shortcut didn't navigate back - URL unchanged")
                    
            except Exception as keys_error:
                # Fallback 1: Try using browser's back() method
                try:
                    self.gui_log("Keyboard shortcut failed, trying browser.back() method...", level="INFO")
                    url_before_back = self.driver.current_url
                    self.driver.back()
                    time.sleep(3)
                    url_after_back = self.driver.current_url
                    if url_before_back != url_after_back:
                        self.gui_log("✅ Used browser.back() method to navigate back", level="INFO")
                        self.update_status("Navigated back to dates page", "#28a745")
                        return True
                    else:
                        self.gui_log("⚠️ browser.back() didn't change URL - trying full navigation fallback...", level="WARNING")
                        # Fallback 2: Navigate back via Patients → search → Billing → All Items
                        return self._navigate_back_via_patients_search()
                except Exception as back_error:
                    self.gui_log(f"⚠️ browser.back() failed: {back_error} - trying full navigation fallback...", level="WARNING")
                    # Fallback 2: Navigate back via Patients → search → Billing → All Items
                    return self._navigate_back_via_patients_search()
                
        except Exception as e:
            self.log_error("Error navigating back using keyboard", exception=e, include_traceback=True)
            # Try full navigation fallback
            return self._navigate_back_via_patients_search()
    
    def _navigate_back_via_patients_search(self):
        """Navigate back to the billing dates page by going through Patients → search → Billing → All Items
        
        This is a fallback method when keyboard shortcuts don't work.
        
        Returns:
            bool: True if navigation successful, False otherwise
        """
        try:
            self.gui_log("Using full navigation fallback: Patients → search → Billing → All Items...", level="INFO")
            self.update_status("Navigating back via Patients page...", "#ff9500")
            
            # Get current client info (should be stored)
            client_name = getattr(self, 'current_client_name', None)
            client_dob = getattr(self, 'current_client_dob', None)
            dos = getattr(self, 'current_date_of_service', None)
            
            if not client_name:
                self.gui_log("⚠️ No current client name available for navigation fallback", level="WARNING")
                return False
            
            # Step 1: Navigate to Patients page
            if not self._navigate_to_patients():
                self.gui_log("⚠️ Could not navigate to Patients page in fallback", level="WARNING")
                return False
            
            # Step 2: Search for the client again
            if not self._search_for_client(client_name, client_dob, dos):
                self.gui_log("⚠️ Could not search for client in navigation fallback", level="WARNING")
                return False
            
            # Step 3: Navigate to Billing tab (should already be done by _search_for_client, but ensure it)
            if not self._navigate_to_billing_tab():
                self.gui_log("⚠️ Could not navigate to Billing tab in fallback", level="WARNING")
                return False
            
            # Step 4: Click All Items
            if not self._click_all_items_button():
                self.gui_log("⚠️ Could not click All Items in navigation fallback", level="WARNING")
                return False
            
            self.gui_log("✅ Successfully navigated back via Patients → search → Billing → All Items", level="INFO")
            self.update_status("Navigated back to dates page (via fallback)", "#28a745")
            return True
            
        except Exception as e:
            self.log_error("Error in navigation fallback", exception=e, include_traceback=True)
            return False
    
    def _ensure_all_items_selected(self):
        """Ensure we're on the Billing tab with "All Items" selected
        
        Checks if "All Items" button has the "selected" class, and clicks it if not selected.
        
        Element: <a tabindex="0" class="quick-search-link selected" id="SearchBillingTransactionsFilter__AllItems" data-testid="quicksearchfilter-allitems-link">All Items</a>
        
        Returns:
            bool: True if "All Items" is selected or was successfully clicked, False otherwise
        """
        try:
            self.gui_log("Verifying 'All Items' is selected...", level="INFO")
            self.update_status("Verifying All Items is selected...", "#ff9500")
            
            # Wait for page to load after navigation
            time.sleep(2)
            
            # Find the "All Items" button
            try:
                all_items_button = self.wait.until(
                    EC.presence_of_element_located((By.ID, "SearchBillingTransactionsFilter__AllItems"))
                )
                
                # Check if it has the "selected" class
                classes = all_items_button.get_attribute("class") or ""
                
                if "selected" in classes:
                    self.gui_log("✅ All Items is already selected", level="INFO")
                    self.update_status("All Items is selected", "#28a745")
                    return True
                else:
                    self.gui_log("All Items is not selected - Clicking it now...", level="INFO")
                    # Click the "All Items" button using the existing method
                    if self._click_all_items_button():
                        self.gui_log("✅ All Items button clicked successfully", level="INFO")
                        return True
                    else:
                        self.gui_log("⚠️ Could not click All Items button", level="WARNING")
                        return False
                        
            except TimeoutException:
                self.gui_log("⚠️ Could not find 'All Items' button - May not be on correct page", level="WARNING")
                # Don't try to navigate to Billing tab here - that would find the wrong Billing button
                # Return False and let the caller handle navigation
                return False
                
        except Exception as e:
            self.log_error("Error ensuring All Items is selected", exception=e, include_traceback=True)
            return False
    
    def _click_date_of_service_again(self):
        """Click the date of service again after navigating back from ERA popup
        
        This is the third time clicking the date (after extracting Payer Claim Control #).
        The date should be clicked to open the popup again, but we don't automatically
        click any tab - we wait for the user's next instruction.
        
        Element: <a href="#" data-testid="billingstatementtable-paymentdate-link" data-preferred_tooltip_position="0">8/10/25</a>
        
        Returns:
            bool: True if date clicked successfully, False otherwise
        """
        try:
            self.gui_log("Clicking date of service again (third pass)...", level="INFO")
            self.update_status(f"Clicking date: {self.current_date_of_service}...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Normalize the date of service to Therapy Notes format
            normalized_dos_str, dos_datetime = self._normalize_date_of_service(self.current_date_of_service)
            
            if not normalized_dos_str or not dos_datetime:
                self.log_error(f"Could not normalize date of service: {self.current_date_of_service}", include_traceback=False)
                self.update_status(f"Error: Invalid date format: {self.current_date_of_service}", "#dc3545")
                return False
            
            self.gui_log(f"Normalized date of service to Therapy Notes format: {normalized_dos_str}", level="DEBUG")
            
            # Wait for date links to load
            time.sleep(2)
            
            # Find all date links using multiple strategies
            date_links = []
            
            # Strategy 1: Find by data-testid
            try:
                self.gui_log("Looking for date links by data-testid...", level="DEBUG")
                date_links = self.wait.until(
                    EC.presence_of_all_elements_located((By.XPATH, "//a[@data-testid='billingstatementtable-paymentdate-link']"))
                )
                self.gui_log(f"✅ Found {len(date_links)} date link(s) by data-testid", level="DEBUG")
            except TimeoutException:
                self.gui_log("No date links found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by link text pattern (contains date)
            if not date_links:
                try:
                    self.gui_log("Trying to find date links by href='#' pattern...", level="DEBUG")
                    date_links = self.driver.find_elements(By.XPATH, "//a[@href='#' and contains(@data-testid, 'paymentdate')]")
                    if date_links:
                        self.gui_log(f"✅ Found {len(date_links)} date link(s) by href pattern", level="DEBUG")
                except Exception:
                    pass
            
            # Strategy 3: Find links containing date-like text
            if not date_links:
                try:
                    # Look for links with M/D/YY pattern
                    import re
                    all_links = self.driver.find_elements(By.TAG_NAME, "a")
                    date_pattern = re.compile(r'^\d{1,2}/\d{1,2}/\d{2}$')
                    for link in all_links:
                        link_text = link.text.strip()
                        if date_pattern.match(link_text):
                            date_links.append(link)
                    if date_links:
                        self.gui_log(f"✅ Found {len(date_links)} date link(s) by text pattern", level="DEBUG")
                except Exception as e:
                    self.gui_log(f"Error finding date links by pattern: {e}", level="DEBUG")
            
            if not date_links:
                self.log_error("Could not find any date links on the page", include_traceback=False)
                self.update_status("Error: No date links found", "#dc3545")
                return False
            
            self.gui_log(f"Found {len(date_links)} potential date link(s) to search through", level="INFO")
            
            # Search for matching date
            matched_link = None
            for link_idx, link in enumerate(date_links, 1):
                try:
                    link_text = link.text.strip()
                    self.gui_log(f"Checking date link {link_idx}: '{link_text}'", level="DEBUG")
                    
                    # Normalize the link text date for comparison
                    normalized_link_date, link_datetime = self._normalize_date_of_service(link_text)
                    
                    if normalized_link_date and link_datetime:
                        # Compare normalized strings (M/D/YY format)
                        if normalized_dos_str == normalized_link_date:
                            matched_link = link
                            self.gui_log(f"✅ Date match found! Link text: '{link_text}', Normalized: {normalized_link_date}", level="INFO")
                            break
                        # Also compare datetime objects for extra safety
                        elif dos_datetime.date() == link_datetime.date():
                            matched_link = link
                            self.gui_log(f"✅ Date match found (datetime comparison)! Link text: '{link_text}'", level="INFO")
                            break
                    else:
                        # If normalization failed, try direct text comparison
                        if link_text == self.current_date_of_service or link_text == normalized_dos_str:
                            matched_link = link
                            self.gui_log(f"✅ Date match found (direct text)! Link text: '{link_text}'", level="INFO")
                            break
                        
                except Exception as e:
                    self.gui_log(f"Error checking date link {link_idx}: {e}", level="DEBUG")
                    continue
            
            if matched_link:
                # Scroll into view
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", matched_link)
                time.sleep(0.5)
                
                # Click the date link
                try:
                    matched_link.click()
                    self.gui_log(f"✅ Date of service clicked: {matched_link.text.strip()}", level="INFO")
                    self.update_status(f"Date clicked: {matched_link.text.strip()} - Popup open", "#28a745")
                    
                    # Wait for popup to appear
                    time.sleep(2)
                    self.gui_log("✅ Popup should be open", level="INFO")
                    
                    # Verify we have both Payer Claim Control # and session medium before checking/updating modifier
                    if self._verify_required_data():
                        # Check and update modifier if needed
                        modifier_status = self._check_and_update_modifier()
                        if modifier_status == "correct":
                            # Modifier was already correct - close popup and proceed to next client
                            if self._close_billing_popup():
                                # Add to correct modifier clients list (legacy)
                                self._add_correct_modifier_client()
                                
                                # Add to comprehensive tracking
                                original_mod = getattr(self, 'current_original_modifier', 'N/A')
                                self._add_tracked_client(
                                    status='Not Refiled - Modifier Correct',
                                    original_modifier=original_mod,
                                    new_modifier='No Change',
                                    modifier_action='Already Correct',
                                    reason='Modifier was already correct - No resubmission needed',
                                    error_message=''
                                )
                                
                                # Navigate back to Patients page for next client
                                if self._navigate_to_patients():
                                    return "next_client"  # Signal to proceed to next client
                                else:
                                    self.gui_log("⚠️ Could not navigate to Patients page, but continuing...", level="WARNING")
                                    return "next_client"
                            else:
                                self.gui_log("⚠️ Could not close popup, but continuing...", level="WARNING")
                                return "next_client"
                        elif modifier_status == "updated":
                            # Modifier was updated - continue with next steps in the workflow
                            # DO NOT close popup or navigate away yet - workflow is still in progress
                            self.gui_log("✅ Modifier updated successfully - Continuing workflow...", level="INFO")
                            self.update_status("Modifier updated - Navigating to Status tab...", "#28a745")
                            
                            # Navigate to Status tab and update status dropdown
                            if self._click_status_tab():
                                # Update status dropdown to "Pending Resubmission"
                                if self._update_status_dropdown("Pending Resubmission"):
                                    # Check if we're in testing mode - stop before saving changes
                                    # Complex testing mode will continue (it clicks Save Changes)
                                    if self.is_testing_mode and not self.is_complex_testing_mode:
                                        self.gui_log("🧪 TEST MODE: Stopping before Save Changes button (no changes will be saved)", level="INFO")
                                        self.update_status("Test complete - Stopped before Save Changes", "#28a745")
                                        
                                        # Add to tracking for test mode
                                        original_mod = getattr(self, 'current_original_modifier', 'N/A')
                                        new_mod = getattr(self, 'current_new_modifier', 'N/A')
                                        mod_action = getattr(self, 'current_modifier_action', 'Updated')
                                        
                                        self._add_tracked_client(
                                            status='Refiled (Test Mode - No Save)',
                                            original_modifier=original_mod,
                                            new_modifier=new_mod,
                                            modifier_action=mod_action,
                                            reason='Modifier updated and status set to Pending Resubmission (Test Mode - Stopped before Save Changes)',
                                            error_message=''
                                        )
                                        
                                        if hasattr(self, 'test_status_label') and self.test_row_number:
                                            self.test_status_label.config(text=f"Status: Test complete for row {self.test_row_number} - Stopped before Save Changes", fg="#28a745")
                                        
                                        # Save output Excel for test mode
                                        if self.tracked_clients:
                                            self._save_output_excel()
                                        
                                        # Check for and close any unrelated windows that may have opened
                                        self._check_and_close_unrelated_windows()
                                        
                                        # Close the billing popup (in test mode, changes weren't saved, so dialog will appear)
                                        self.gui_log("Closing billing popup after test mode...", level="INFO")
                                        if self._close_billing_popup():
                                            self.gui_log("✅ Billing popup closed after test mode", level="INFO")
                                        else:
                                            self.gui_log("⚠️ Could not close billing popup, but continuing...", level="WARNING")
                                        
                                        # Reset testing mode after test completes
                                        self.is_testing_mode = False
                                        self.test_row_number = None
                                        return True  # Return True to indicate success, but stop here
                                    
                                    # Click Save Changes button (only in non-testing mode OR complex testing mode)
                                    # Complex testing mode will click Save Changes and continue
                                    # Wait a moment after closing unrelated windows to ensure popup is ready
                                    time.sleep(1)
                                    if self._click_save_changes_button():
                                        # Client successfully refiled - add to tracking
                                        # Note: At this point, we're not in test mode (test mode stops before Save Changes)
                                        original_mod = getattr(self, 'current_original_modifier', 'N/A')
                                        new_mod = getattr(self, 'current_new_modifier', 'N/A')
                                        mod_action = getattr(self, 'current_modifier_action', 'Updated')
                                        
                                        # Actually refiled
                                        self._add_tracked_client(
                                            status='Refiled',
                                            original_modifier=original_mod,
                                            new_modifier=new_mod,
                                            modifier_action=mod_action,
                                            reason='Modifier updated, status set to Pending Resubmission, and Save Changes clicked successfully',
                                            error_message=''
                                        )
                                        
                                        # Step 24: After Save Changes, popup closes automatically
                                        # Now navigate to Submit Primary Claims page
                                        if self._click_submit_primary_claims():
                                            # Step 25: Filter by date range on the Submit Primary Claims page
                                            if self._filter_date_range_for_refiling():
                                                self.gui_log("✅ Successfully filtered date range for refiling", level="INFO")
                                                
                                                # Step 26: Select the checkbox for the DOS
                                                if self._select_dos_checkbox():
                                                    # Step 27: Select "Amended Claim" from the resubmission type dropdown
                                                    if self._select_amended_claim():
                                                        # Step 28: Enter the payer claim control number
                                                        if self._enter_payer_claim_control():
                                                            # Step 29: Click Submit Claims button
                                                            # Check if we're in complex testing mode - stop before clicking Submit Claims
                                                            if self.is_complex_testing_mode:
                                                                self.gui_log("🧪 COMPLEX TEST MODE: Stopping before 'Submit Claims' button - Page will remain open for review", level="INFO")
                                                                self.update_status("Complex test complete - Stopped before Submit Claims (page open for review)", "#28a745")
                                                                
                                                                # Add to tracking for complex test mode
                                                                original_mod = getattr(self, 'current_original_modifier', 'N/A')
                                                                new_mod = getattr(self, 'current_new_modifier', 'N/A')
                                                                mod_action = getattr(self, 'current_modifier_action', 'Updated')
                                                                
                                                                self._add_tracked_client(
                                                                    status='Refiled (Complex Test - No Submit)',
                                                                    original_modifier=original_mod,
                                                                    new_modifier=new_mod,
                                                                    modifier_action=mod_action,
                                                                    reason='Modifier updated, status set to Pending Resubmission, Save Changes clicked, and all steps completed up to Submit Claims (Complex Test Mode - Stopped before Submit Claims)',
                                                                    error_message=''
                                                                )
                                                                
                                                                # Save output Excel for complex test mode
                                                                if self.tracked_clients:
                                                                    self.gui_log(f"Saving output Excel with {len(self.tracked_clients)} tracked client(s)...", level="INFO")
                                                                    self._save_output_excel()
                                                                    self.gui_log(f"✅ Output Excel saved successfully with {len(self.tracked_clients)} client record(s)", level="INFO")
                                                                else:
                                                                    self.gui_log("⚠️ No tracked clients to save in output Excel", level="WARNING")
                                                                
                                                                # Don't navigate away - stay on page for review
                                                                return True
                                                            elif self._click_submit_claims_button():
                                                                self.gui_log("✅ Successfully submitted claim for refiling", level="INFO")
                                                                self._handle_claims_submitted_popup()
                                                                # Flow complete - navigate back to Patients page to continue with next client
                                                                if self._navigate_to_patients():
                                                                    self.gui_log("✅ Navigated back to Patients page - Ready for next client", level="INFO")
                                                                    return True
                                                                else:
                                                                    self.gui_log("⚠️ Could not navigate back to Patients page, but continuing...", level="WARNING")
                                                                    return True
                                                            else:
                                                                self.gui_log("⚠️ Could not click Submit Claims button, but continuing...", level="WARNING")
                                                                return True
                                                        else:
                                                            self.gui_log("⚠️ Could not enter payer claim control number, but continuing...", level="WARNING")
                                                            return True
                                                    else:
                                                        self.gui_log("⚠️ Could not select Amended Claim, but continuing...", level="WARNING")
                                                        return True
                                                else:
                                                    self.gui_log("⚠️ Could not select DOS checkbox, but continuing...", level="WARNING")
                                                    return True
                                            else:
                                                self.gui_log("⚠️ Could not filter date range, but continuing...", level="WARNING")
                                                return True
                                        else:
                                            self.gui_log("⚠️ Could not navigate to Submit Primary Claims, but continuing...", level="WARNING")
                                            return True
                                    else:
                                        self.gui_log("⚠️ Could not click Save Changes button, but continuing...", level="WARNING")
                                        return True
                                else:
                                    self.gui_log("⚠️ Could not update status dropdown, but continuing...", level="WARNING")
                                    return True
                            else:
                                self.gui_log("⚠️ Could not navigate to Status tab, but continuing...", level="WARNING")
                                return True
                        else:
                            self.gui_log("⚠️ Could not check/update modifier, but continuing...", level="WARNING")
                            return True
                    else:
                        self.gui_log("⚠️ Missing required data (Payer Claim Control # or session medium) - Cannot proceed with modifier check", level="WARNING")
                        return True
                    
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", matched_link)
                        self.gui_log(f"✅ Date of service clicked (JavaScript): {matched_link.text.strip()}", level="INFO")
                        self.update_status(f"Date clicked: {matched_link.text.strip()} - Popup open", "#28a745")
                        time.sleep(2)
                        
                        # Verify we have both Payer Claim Control # and session medium before checking/updating modifier
                        if self._verify_required_data():
                            # Check and update modifier if needed
                            modifier_status = self._check_and_update_modifier()
                            if modifier_status == "correct":
                                # Modifier was already correct - close popup and proceed to next client
                                if self._close_billing_popup():
                                    # Add to correct modifier clients list
                                    self._add_correct_modifier_client()
                                    # Navigate back to Patients page for next client
                                    if self._navigate_to_patients():
                                        return "next_client"  # Signal to proceed to next client
                                    else:
                                        self.gui_log("⚠️ Could not navigate to Patients page, but continuing...", level="WARNING")
                                        return "next_client"
                                else:
                                    self.gui_log("⚠️ Could not close popup, but continuing...", level="WARNING")
                                    return "next_client"
                            elif modifier_status == "updated":
                                # Modifier was updated - continue with next steps in the workflow
                                # DO NOT close popup or navigate away yet - workflow is still in progress
                                self.gui_log("✅ Modifier updated successfully - Continuing workflow...", level="INFO")
                                self.update_status("Modifier updated - Navigating to Status tab...", "#28a745")
                                
                                # Navigate to Status tab and update status dropdown
                                if self._click_status_tab():
                                    # Update status dropdown to "Pending Resubmission"
                                    if self._update_status_dropdown("Pending Resubmission"):
                                        # Check if we're in testing mode - stop before saving changes
                                        # Complex testing mode will continue (it clicks Save Changes)
                                        if self.is_testing_mode and not self.is_complex_testing_mode:
                                            self.gui_log("🧪 TEST MODE: Stopping before Save Changes button (no changes will be saved)", level="INFO")
                                            self.update_status("Test complete - Stopped before Save Changes", "#28a745")
                                            if hasattr(self, 'test_status_label') and self.test_row_number:
                                                self.test_status_label.config(text=f"Status: Test complete for row {self.test_row_number} - Stopped before Save Changes", fg="#28a745")
                                            
                                            # Close the billing popup (in test mode, changes weren't saved, so dialog will appear)
                                            self.gui_log("Closing billing popup after test mode...", level="INFO")
                                            if self._close_billing_popup():
                                                self.gui_log("✅ Billing popup closed after test mode", level="INFO")
                                            else:
                                                self.gui_log("⚠️ Could not close billing popup, but continuing...", level="WARNING")
                                            
                                            # Reset testing mode after test completes
                                            self.is_testing_mode = False
                                            self.test_row_number = None
                                            return True  # Return True to indicate success, but stop here
                                        
                                        # Click Save Changes button (only in non-testing mode OR complex testing mode)
                                        # Complex testing mode will click Save Changes and continue
                                        if self._click_save_changes_button():
                                            # Client successfully refiled - add to tracking
                                            original_mod = getattr(self, 'current_original_modifier', 'N/A')
                                            new_mod = getattr(self, 'current_new_modifier', 'N/A')
                                            mod_action = getattr(self, 'current_modifier_action', 'Updated')
                                            
                                            # Actually refiled
                                            self._add_tracked_client(
                                                status='Refiled',
                                                original_modifier=original_mod,
                                                new_modifier=new_mod,
                                                modifier_action=mod_action,
                                                reason='Modifier updated, status set to Pending Resubmission, and Save Changes clicked successfully',
                                                error_message=''
                                            )
                                            
                                            # Step 24: After Save Changes, popup closes automatically
                                            # Now navigate to Submit Primary Claims page
                                            if self._click_submit_primary_claims():
                                                # Step 25: Filter by date range on the Submit Primary Claims page
                                                if self._filter_date_range_for_refiling():
                                                    self.gui_log("✅ Successfully filtered date range for refiling", level="INFO")
                                                    
                                                    # Step 26: Select the checkbox for the DOS
                                                    if self._select_dos_checkbox():
                                                        # Step 27: Select "Amended Claim" from the resubmission type dropdown
                                                        if self._select_amended_claim():
                                                            # Step 28: Enter the payer claim control number
                                                            if self._enter_payer_claim_control():
                                                                # Step 29: Click Submit Claims button
                                                                # Check if we're in complex testing mode - stop before clicking Submit Claims
                                                                if self.is_complex_testing_mode:
                                                                    self.gui_log("🧪 COMPLEX TEST MODE: Stopping before 'Submit Claims' button - Page will remain open for review", level="INFO")
                                                                    self.update_status("Complex test complete - Stopped before Submit Claims (page open for review)", "#28a745")
                                                                    
                                                                    # Add to tracking for complex test mode
                                                                    original_mod = getattr(self, 'current_original_modifier', 'N/A')
                                                                    new_mod = getattr(self, 'current_new_modifier', 'N/A')
                                                                    mod_action = getattr(self, 'current_modifier_action', 'Updated')
                                                                    
                                                                    self._add_tracked_client(
                                                                        status='Refiled (Complex Test - No Submit)',
                                                                        original_modifier=original_mod,
                                                                        new_modifier=new_mod,
                                                                        modifier_action=mod_action,
                                                                        reason='Modifier updated, status set to Pending Resubmission, Save Changes clicked, and all steps completed up to Submit Claims (Complex Test Mode - Stopped before Submit Claims)',
                                                                        error_message=''
                                                                    )
                                                                    
                                                                    # Save output Excel for complex test mode
                                                                    if self.tracked_clients:
                                                                        self._save_output_excel()
                                                                    
                                                                    # Don't navigate away - stay on page for review
                                                                    return True
                                                                elif self._click_submit_claims_button():
                                                                    self.gui_log("✅ Successfully submitted claim for refiling", level="INFO")
                                                                    self._handle_claims_submitted_popup()
                                                                    # Flow complete - navigate back to Patients page to continue with next client
                                                                    if self._navigate_to_patients():
                                                                        self.gui_log("✅ Navigated back to Patients page - Ready for next client", level="INFO")
                                                                        return True
                                                                    else:
                                                                        self.gui_log("⚠️ Could not navigate back to Patients page, but continuing...", level="WARNING")
                                                                        return True
                                                                else:
                                                                    self.gui_log("⚠️ Could not click Submit Claims button, but continuing...", level="WARNING")
                                                                    return True
                                                            else:
                                                                self.gui_log("⚠️ Could not enter payer claim control number, but continuing...", level="WARNING")
                                                                return True
                                                        else:
                                                            self.gui_log("⚠️ Could not select Amended Claim, but continuing...", level="WARNING")
                                                            return True
                                                    else:
                                                        self.gui_log("⚠️ Could not select DOS checkbox, but continuing...", level="WARNING")
                                                        return True
                                                else:
                                                    self.gui_log("⚠️ Could not filter date range, but continuing...", level="WARNING")
                                                    return True
                                            else:
                                                self.gui_log("⚠️ Could not navigate to Submit Primary Claims, but continuing...", level="WARNING")
                                                return True
                                        else:
                                            self.gui_log("⚠️ Could not click Save Changes button, but continuing...", level="WARNING")
                                            return True
                                    else:
                                        self.gui_log("⚠️ Could not update status dropdown, but continuing...", level="WARNING")
                                        return True
                                else:
                                    self.gui_log("⚠️ Could not navigate to Status tab, but continuing...", level="WARNING")
                                    return True
                            else:
                                # modifier_status is neither "correct" nor "updated" (shouldn't happen, but handle it)
                                self.gui_log("⚠️ Unexpected modifier status, but continuing...", level="WARNING")
                                return True
                        else:
                            self.gui_log("⚠️ Missing required data (Payer Claim Control # or session medium) - Cannot proceed with modifier check", level="WARNING")
                            return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click date link: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error(f"Could not find matching date of service: {self.current_date_of_service} (normalized: {normalized_dos_str})", include_traceback=False)
                self.update_status(f"Error: Date not found: {self.current_date_of_service}", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error clicking date of service again", exception=e, include_traceback=True)
            self.update_status(f"Date click error: {str(e)}", "#dc3545")
            return False
    
    def _verify_required_data(self):
        """Verify that we have both Payer Claim Control # and session medium before proceeding
        
        Returns:
            bool: True if both required data are available, False otherwise
        """
        try:
            self.gui_log("Verifying required data: Payer Claim Control # and session medium...", level="INFO")
            
            has_payer_claim_control = hasattr(self, 'current_payer_claim_control') and self.current_payer_claim_control
            has_session_medium = hasattr(self, 'current_session_medium') and self.current_session_medium
            
            if has_payer_claim_control:
                self.gui_log(f"✅ Payer Claim Control #: {self.current_payer_claim_control}", level="INFO")
            else:
                self.gui_log("❌ Payer Claim Control # not found", level="ERROR")
            
            if has_session_medium:
                self.gui_log(f"✅ Session medium: {self.current_session_medium}", level="INFO")
            else:
                self.gui_log("❌ Session medium not identified", level="ERROR")
            
            if has_payer_claim_control and has_session_medium:
                self.gui_log("✅ Both required data verified - Ready to check modifier", level="INFO")
                return True
            else:
                self.log_error("Missing required data - Cannot proceed with modifier check", include_traceback=False)
                return False
                
        except Exception as e:
            self.log_error("Error verifying required data", exception=e, include_traceback=True)
            return False
    
    def _extract_primary_policy_info(self):
        """Extract the primary policy name and member ID from the billing popup."""
        try:
            self.gui_log("Extracting primary policy information from popup...", level="INFO")
            
            # Allow popup elements to stabilize
            time.sleep(1)
            
            primary_name = None
            member_id = None
            
            # Strategy 1: Find by primary payer data-testid
            try:
                self.gui_log("Looking for primary payer link by data-testid...", level="DEBUG")
                quick_wait = WebDriverWait(self.driver, 5)
                primary_link = quick_wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//a[contains(@data-testid, 'primarypayer')]")
                    )
                )
                link_text = primary_link.text.strip()
                if link_text:
                    primary_name = " ".join(link_text.split())
                else:
                    # Try to gather from child spans if primary link text is empty
                    spans = primary_link.find_elements(By.TAG_NAME, "span")
                    for span in spans:
                        span_text = span.text.strip()
                        if span_text:
                            primary_name = " ".join(span_text.split())
                            break
                if primary_name:
                    self.gui_log(f"✅ Primary policy name captured: '{primary_name}'", level="INFO")
            except TimeoutException:
                self.gui_log("Primary payer link not found by data-testid", level="DEBUG")
            except Exception as primary_link_error:
                self.gui_log(f"Error reading primary payer link: {primary_link_error}", level="DEBUG")
            
            # Strategy 2: Find by label "Primary Policy"
            if not primary_name:
                try:
                    self.gui_log("Trying to locate primary policy via label fallback...", level="DEBUG")
                    label_element = self.driver.find_element(
                        By.XPATH,
                        "//label[contains(normalize-space(text()), 'Primary Policy')]"
                    )
                    primary_container = label_element.find_element(
                        By.XPATH,
                        "./following::div[contains(@class, 'DynamicInputViewModeContent')][1]"
                    )
                    primary_text = primary_container.text.strip()
                    if primary_text:
                        primary_name = " ".join(primary_text.split())
                        self.gui_log(f"✅ Primary policy name captured via label: '{primary_name}'", level="INFO")
                except Exception as label_error:
                    self.gui_log(f"Primary policy label fallback failed: {label_error}", level="DEBUG")
            
            # Strategy 3: Find member ID by data-testid
            try:
                self.gui_log("Looking for primary policy member ID...", level="DEBUG")
                member_span = self.driver.find_element(
                    By.XPATH, "//span[contains(@data-testid, 'primarypayermemberid')]"
                )
                span_text = member_span.text.strip()
                if span_text:
                    member_id = span_text
                    self.gui_log(f"✅ Primary policy member ID captured: '{member_id}'", level="INFO")
            except Exception as member_error:
                self.gui_log(f"Primary policy member ID not found: {member_error}", level="DEBUG")
            
            stored_name = primary_name or 'N/A'
            stored_member_id = member_id or 'N/A'
            
            if stored_name == 'N/A':
                self.gui_log("⚠️ Primary policy name not found in popup", level="WARNING")
            if stored_member_id == 'N/A':
                self.gui_log("⚠️ Primary policy member ID not found in popup", level="WARNING")
            
            self.current_primary_policy_name = stored_name
            self.current_primary_policy_member_id = stored_member_id
            
            return stored_name, stored_member_id
        
        except Exception as e:
            self.gui_log(f"Error extracting primary policy information: {e}", level="WARNING")
            self.current_primary_policy_name = 'N/A'
            self.current_primary_policy_member_id = 'N/A'
            return None, None
    
    def _handle_payment_policy_mismatch(self, reason):
        """Handle scenarios where the payment could not be matched to the captured primary policy."""
        try:
            client_name = getattr(self, 'current_client_name', 'Unknown')
            dob = getattr(self, 'current_client_dob', '')
            dos = getattr(self, 'current_date_of_service', '')
            
            mismatch_reason = reason or 'Payment could not be matched to captured primary policy'
            
            # Record in legacy skipped list
            try:
                self.skipped_clients.append({
                    'client_name': client_name,
                    'dob': dob,
                    'date_of_service': dos,
                    'reason': mismatch_reason
                })
            except Exception as skip_error:
                self.gui_log(f"Error recording skipped client for mismatch: {skip_error}", level="DEBUG")
            
            # Record in comprehensive tracking
            try:
                original_mod = getattr(self, 'current_original_modifier', None) or 'N/A'
                self._add_tracked_client(
                    status='Not Refiled - Payment/Policy Mismatch',
                    original_modifier=original_mod,
                    new_modifier='N/A',
                    modifier_action='N/A',
                    reason=mismatch_reason,
                    error_message='Payment selection aborted due to policy mismatch'
                )
            except Exception as track_error:
                self.gui_log(f"Error tracking payment mismatch: {track_error}", level="DEBUG")
            
            # Close popup if possible
            try:
                if self._close_billing_popup():
                    self.gui_log("✅ Billing popup closed after payment mismatch", level="INFO")
                else:
                    self.gui_log("⚠️ Could not close billing popup after payment mismatch", level="WARNING")
            except Exception as close_error:
                self.gui_log(f"Error closing popup after payment mismatch: {close_error}", level="DEBUG")
            
            # Navigate back to Patients page to continue workflow
            try:
                if self._navigate_to_patients():
                    self.gui_log("✅ Navigated back to Patients page after payment mismatch", level="INFO")
                else:
                    self.gui_log("⚠️ Could not navigate back to Patients page after payment mismatch", level="WARNING")
            except Exception as nav_error:
                self.gui_log(f"Error navigating back to Patients page after mismatch: {nav_error}", level="DEBUG")
        
        except Exception as e:
            self.gui_log(f"Unhandled error during payment mismatch handling: {e}", level="DEBUG")
    
    def _handle_patient_not_found_in_era_dropdown(self):
        """Handle scenarios where the View ERA patient dropdown does not contain the current client."""
        try:
            client_name = getattr(self, 'current_client_name', 'Unknown')
            dob = getattr(self, 'current_client_dob', '')
            dos = getattr(self, 'current_date_of_service', '')
            original_mod = getattr(self, 'current_original_modifier', None) or 'N/A'
            
            search_variants = getattr(self, 'last_patient_dropdown_search_variants', {}) or {}
            available_options = getattr(self, 'last_patient_dropdown_options', []) or []
            
            searched_for = search_variants.get('standard') or client_name
            variants_summary = []
            if search_variants.get('with_middle'):
                variants_summary.append(search_variants['with_middle'])
            if search_variants.get('with_middle_period'):
                variants_summary.append(search_variants['with_middle_period'])
            if search_variants.get('last_name_variants'):
                variants_summary.append(f"Last name variants: {', '.join(search_variants['last_name_variants'])}")
            variants_text = "; ".join(variants_summary)
            
            option_preview = ", ".join(available_options[:10])
            reason = f"Patient not found in ERA dropdown (searched for '{searched_for}')"
            if option_preview:
                reason += f"; available options include: {option_preview}"
            
            self.gui_log(f"❌ {reason}", level="ERROR")
            if variants_text:
                self.gui_log(f"  Search variants used: {variants_text}", level="DEBUG")
            self.update_status("ERA dropdown missing patient - Skipping client...", "#dc3545")
            
            # Record in skipped clients list
            try:
                self.skipped_clients.append({
                    'client_name': client_name,
                    'dob': dob,
                    'date_of_service': dos,
                    'reason': 'Patient not listed in ERA dropdown'
                })
            except Exception as skip_error:
                self.gui_log(f"Error recording skipped client for ERA dropdown: {skip_error}", level="DEBUG")
            
            # Record in comprehensive tracking
            try:
                self._add_tracked_client(
                    status='Not Refiled - ERA Dropdown Missing Patient',
                    original_modifier=original_mod,
                    new_modifier='N/A',
                    modifier_action='N/A',
                    reason='Patient not listed in ERA dropdown (View ERA popup)',
                    error_message=reason
                )
            except Exception as track_error:
                self.gui_log(f"Error tracking ERA dropdown miss: {track_error}", level="DEBUG")
            
            # Close the ERA popup before navigating away
            try:
                if self._close_era_popup():
                    self.gui_log("✅ ERA popup closed after missing patient in dropdown", level="INFO")
                else:
                    self.gui_log("⚠️ Could not close ERA popup after missing patient in dropdown", level="WARNING")
            except Exception as close_error:
                self.gui_log(f"Error closing ERA popup after missing patient: {close_error}", level="DEBUG")
            
            # Navigate back to Patients page to continue workflow
            try:
                if self._navigate_to_patients_from_payment_page():
                    self.gui_log("✅ Navigated back to Patients page after missing ERA dropdown patient", level="INFO")
                else:
                    self.gui_log("⚠️ Could not navigate back to Patients page after missing ERA dropdown patient", level="WARNING")
            except Exception as nav_error:
                self.gui_log(f"Error navigating back to Patients page after missing dropdown patient: {nav_error}", level="DEBUG")
            
        finally:
            # Reset tracking flags to avoid affecting future dropdown attempts
            self.last_patient_dropdown_not_found = False
            self.last_patient_dropdown_options = []
            self.last_patient_dropdown_search_variants = {}
        
        return "skip_client"
    
    def _extract_original_modifier(self):
        """Extract the original modifier from the popup immediately after it appears
        
        This should be called right after clicking the date of service, before navigating to Notes tab.
        The modifier field is available immediately when the popup opens.
        
        Element: <input type="text" maxlength="2" data-testid="modifierseditor-code-input-1" style="text-transform: uppercase; width: 25px;">
        
        Returns:
            str: The modifier value ('93', '95', or 'N/A' if not found), or None on error
        """
        try:
            self.gui_log("Extracting original modifier from popup...", level="INFO")
            
            # Wait for popup to fully load - give it more time to ensure modifier field is available
            time.sleep(2)
            
            # Find the modifier input field using multiple strategies
            modifier_input = None
            
            # Strategy 1: Find by data-testid (try visible first, then just presence)
            try:
                self.gui_log("Looking for modifier input by data-testid...", level="DEBUG")
                quick_wait = WebDriverWait(self.driver, 5)
                # First try to find it as visible
                try:
                    modifier_input = quick_wait.until(
                        EC.visibility_of_element_located((By.XPATH, "//input[@data-testid='modifierseditor-code-input-1']"))
                    )
                    self.gui_log("✅ Found modifier input by data-testid (visible)", level="DEBUG")
                except TimeoutException:
                    # If not visible, try just presence
                    modifier_input = quick_wait.until(
                        EC.presence_of_element_located((By.XPATH, "//input[@data-testid='modifierseditor-code-input-1']"))
                    )
                    self.gui_log("✅ Found modifier input by data-testid (present)", level="DEBUG")
            except TimeoutException:
                self.gui_log("Modifier input not found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by type="text" with maxlength="2"
            if not modifier_input:
                try:
                    self.gui_log("Trying to find modifier input by type and maxlength...", level="DEBUG")
                    quick_wait = WebDriverWait(self.driver, 5)
                    try:
                        modifier_input = quick_wait.until(
                            EC.visibility_of_element_located((By.XPATH, "//input[@type='text' and @maxlength='2']"))
                        )
                        self.gui_log("✅ Found modifier input by type and maxlength (visible)", level="DEBUG")
                    except TimeoutException:
                        modifier_input = quick_wait.until(
                            EC.presence_of_element_located((By.XPATH, "//input[@type='text' and @maxlength='2']"))
                        )
                        self.gui_log("✅ Found modifier input by type and maxlength (present)", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Modifier input not found by type/maxlength...", level="DEBUG")
            
            # Strategy 3: Find by modifier-related attributes
            if not modifier_input:
                try:
                    self.gui_log("Trying to find modifier input by modifier-related attributes...", level="DEBUG")
                    quick_wait = WebDriverWait(self.driver, 5)
                    try:
                        modifier_input = quick_wait.until(
                            EC.visibility_of_element_located((By.XPATH, "//input[contains(@data-testid, 'modifier') or contains(@id, 'modifier') or contains(@name, 'modifier')]"))
                        )
                        self.gui_log("✅ Found modifier input by modifier-related attributes (visible)", level="DEBUG")
                    except TimeoutException:
                        modifier_input = quick_wait.until(
                            EC.presence_of_element_located((By.XPATH, "//input[contains(@data-testid, 'modifier') or contains(@id, 'modifier') or contains(@name, 'modifier')]"))
                        )
                        self.gui_log("✅ Found modifier input by modifier-related attributes (present)", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Modifier input not found by modifier attributes...", level="DEBUG")
            
            if not modifier_input:
                self.gui_log("⚠️ Could not find modifier input field - Will extract later", level="WARNING")
                # Don't return False - this is not critical, we can extract later
                self.current_original_modifier = 'N/A'
                return None
            
            # Read current modifier value
            try:
                current_modifier = modifier_input.get_attribute("value") or ""
                current_modifier = current_modifier.strip().upper()
                
                if current_modifier:
                    self.gui_log(f"✅ Original modifier extracted: '{current_modifier}'", level="INFO")
                    # Store original modifier for tracking
                    self.current_original_modifier = current_modifier
                    return current_modifier
                else:
                    self.gui_log("⚠️ Modifier input field found but is empty", level="WARNING")
                    self.current_original_modifier = 'N/A'
                    return 'N/A'
                    
            except Exception as read_error:
                self.gui_log(f"Error reading modifier value: {read_error}", level="WARNING")
                self.current_original_modifier = 'N/A'
                return 'N/A'
                
        except Exception as e:
            self.gui_log(f"Error extracting original modifier: {e}", level="WARNING")
            self.current_original_modifier = 'N/A'
            return None
    
    def _check_and_update_modifier(self):
        """Check the current modifier in the popup and update it if it doesn't match the session medium
        
        Modifier logic:
        - "95" = video session
        - "93" = phone session
        
        Element: <input type="text" maxlength="2" data-testid="modifierseditor-code-input-1" style="text-transform: uppercase; width: 25px;">
        
        Returns:
            bool: True if modifier check/update completed successfully, False otherwise
        """
        try:
            self.gui_log("Step 19: Checking current modifier in popup...", level="INFO")
            self.update_status("Checking modifier...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Determine expected modifier based on session medium
            expected_modifier = None
            if self.current_session_medium == "video":
                expected_modifier = "95"
            elif self.current_session_medium == "phone":
                expected_modifier = "93"
            else:
                self.log_error(f"Unknown session medium: {self.current_session_medium} - Cannot determine expected modifier", include_traceback=False)
                self.update_status(f"Error: Unknown session medium: {self.current_session_medium}", "#dc3545")
                return False
            
            self.gui_log(f"Expected modifier for {self.current_session_medium} session: {expected_modifier}", level="INFO")
            
            # Wait for popup to fully load
            time.sleep(2)
            
            # Find the modifier input field using multiple strategies
            modifier_input = None
            
            # Strategy 1: Find by data-testid
            try:
                self.gui_log("Looking for modifier input by data-testid...", level="DEBUG")
                modifier_input = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[@data-testid='modifierseditor-code-input-1']"))
                )
                self.gui_log("✅ Found modifier input by data-testid", level="DEBUG")
            except TimeoutException:
                self.gui_log("Modifier input not found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by name or type="text" with maxlength="2"
            if not modifier_input:
                try:
                    self.gui_log("Trying to find modifier input by type and maxlength...", level="DEBUG")
                    modifier_input = self.wait.until(
                        EC.presence_of_element_located((By.XPATH, "//input[@type='text' and @maxlength='2']"))
                    )
                    self.gui_log("✅ Found modifier input by type and maxlength", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Modifier input not found by type/maxlength...", level="DEBUG")
            
            # Strategy 3: Find by placeholder or nearby label text
            if not modifier_input:
                try:
                    self.gui_log("Trying to find modifier input by nearby text...", level="DEBUG")
                    # Look for input near "Modifier" text
                    modifier_input = self.wait.until(
                        EC.presence_of_element_located((By.XPATH, "//input[contains(@data-testid, 'modifier') or contains(@id, 'modifier') or contains(@name, 'modifier')]"))
                    )
                    self.gui_log("✅ Found modifier input by modifier-related attributes", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Modifier input not found by modifier attributes...", level="DEBUG")
            
            if not modifier_input:
                self.log_error("Could not find modifier input field", include_traceback=False)
                self.update_status("Error: Modifier input field not found", "#dc3545")
                return False
            
            # Scroll into view (use gentler scroll to avoid triggering events)
            # Use 'auto' instead of 'smooth' and 'nearest' instead of 'center' to be less intrusive
            try:
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'auto', block: 'nearest'});", modifier_input)
                time.sleep(0.3)
            except:
                pass  # If scroll fails, continue anyway
            
            # Read current modifier value
            try:
                current_modifier = modifier_input.get_attribute("value") or ""
                current_modifier = current_modifier.strip().upper()
                self.gui_log(f"Current modifier in popup: '{current_modifier}'", level="INFO")
                
                # Store original modifier for tracking
                self.current_original_modifier = current_modifier if current_modifier else 'N/A'
                self.current_expected_modifier = expected_modifier
                
                if current_modifier == expected_modifier:
                    self.gui_log(f"✅ Modifier is correct! Current: '{current_modifier}', Expected: '{expected_modifier}' - Claim was submitted correctly", level="INFO")
                    self.update_status(f"Modifier correct: {current_modifier} ({self.current_session_medium})", "#28a745")
                    
                    # Store tracking data
                    self.current_new_modifier = 'No Change'
                    self.current_modifier_action = 'Already Correct'
                    
                    return "correct"  # Return "correct" status
                else:
                    self.gui_log(f"⚠️ Modifier mismatch! Current: '{current_modifier}', Expected: '{expected_modifier}' - Updating modifier...", level="WARNING")
                    self.update_status(f"Updating modifier: {current_modifier} → {expected_modifier}", "#ff9500")
                    
                    # Clear the current value
                    modifier_input.clear()
                    time.sleep(0.3)
                    
                    # Enter the expected modifier
                    modifier_input.send_keys(expected_modifier)
                    time.sleep(0.5)
                    
                    # Verify the value was set correctly
                    updated_value = modifier_input.get_attribute("value") or ""
                    updated_value = updated_value.strip().upper()
                    
                    if updated_value == expected_modifier:
                        self.gui_log(f"✅ Modifier updated successfully! Changed from '{current_modifier}' to '{expected_modifier}'", level="INFO")
                        self.update_status(f"Modifier updated: {expected_modifier} ({self.current_session_medium})", "#28a745")
                        
                        # Store tracking data
                        self.current_new_modifier = updated_value
                        self.current_modifier_action = 'Updated'
                        
                        # Check for and close any unrelated windows that may have opened
                        self._check_and_close_unrelated_windows()
                        
                        return "updated"  # Return "updated" status
                    else:
                        self.log_error(f"Failed to update modifier - Expected: '{expected_modifier}', Got: '{updated_value}'", include_traceback=False)
                        self.update_status(f"Error: Could not update modifier", "#dc3545")
                        
                        # Store tracking data even on error
                        self.current_new_modifier = updated_value if updated_value else 'Update Failed'
                        self.current_modifier_action = 'Update Failed'
                        
                        return False
                        
            except Exception as read_error:
                self.log_error(f"Error reading/updating modifier: {read_error}", exception=read_error, include_traceback=True)
                self.update_status(f"Error reading modifier: {str(read_error)}", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error checking/updating modifier", exception=e, include_traceback=True)
            self.update_status(f"Modifier check error: {str(e)}", "#dc3545")
            return False
    
    def _close_billing_popup(self):
        """Close the billing popup by clicking the X button
        
        Element: <i class="fa6 fa-lg fa-2x fa-times"></i>
        
        Returns:
            bool: True if popup closed successfully, False otherwise
        """
        try:
            self.gui_log("Step 20: Closing billing popup...", level="INFO")
            self.update_status("Closing popup...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait a moment before closing
            time.sleep(1)
            
            # Find X button using multiple strategies (same as ERA popup)
            close_button = None
            
            # Strategy 1: Find by class combination
            try:
                self.gui_log("Looking for X button by class (fa-times)...", level="DEBUG")
                close_button = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//i[contains(@class, 'fa-times')]"))
                )
                self.gui_log("✅ Found X button by fa-times class", level="DEBUG")
            except TimeoutException:
                self.gui_log("X button not found by fa-times, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by exact class combination
            if not close_button:
                try:
                    self.gui_log("Trying to find X button by exact class combination...", level="DEBUG")
                    close_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//i[contains(@class, 'fa6') and contains(@class, 'fa-lg') and contains(@class, 'fa-2x') and contains(@class, 'fa-times')]"))
                    )
                    self.gui_log("✅ Found X button by exact class combination", level="DEBUG")
                except TimeoutException:
                    self.gui_log("X button not found by exact class, trying other methods...", level="DEBUG")
            
            # Strategy 3: Find by class containing "times" or "close"
            if not close_button:
                try:
                    self.gui_log("Trying to find X button by class containing 'times'...", level="DEBUG")
                    close_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//i[contains(@class, 'times') or contains(@class, 'close')]"))
                    )
                    self.gui_log("✅ Found X button by class containing 'times'", level="DEBUG")
                except TimeoutException:
                    self.gui_log("X button not found by class pattern...", level="DEBUG")
            
            # Strategy 4: Find clickable parent element containing the icon
            if not close_button:
                try:
                    self.gui_log("Trying to find parent element containing fa-times icon...", level="DEBUG")
                    close_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[.//i[contains(@class, 'fa-times')]] | //a[.//i[contains(@class, 'fa-times')]] | //div[.//i[contains(@class, 'fa-times')]]"))
                    )
                    self.gui_log("✅ Found X button by parent element", level="DEBUG")
                except TimeoutException:
                    self.gui_log("X button not found by parent element...", level="DEBUG")
            
            if close_button:
                # Scroll into view
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", close_button)
                time.sleep(0.5)
                
                # Click the X button
                try:
                    close_button.click()
                    self.gui_log("✅ Billing popup closed (X button clicked)", level="INFO")
                    self.update_status("Popup closed", "#28a745")
                    
                    # In test mode, check if confirmation dialog appears (unsaved changes warning)
                    # This only happens in test mode because in actual runs, changes are saved before closing
                    if self.is_testing_mode:
                        time.sleep(1)  # Wait for dialog to appear if it will
                        if self._handle_leave_and_discard_dialog():
                            self.gui_log("✅ Confirmation dialog handled (Leave and Discard Changes)", level="INFO")
                        
                    time.sleep(2)  # Wait for popup to close
                    return True
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", close_button)
                        self.gui_log("✅ Billing popup closed (X button clicked via JavaScript)", level="INFO")
                        self.update_status("Popup closed", "#28a745")
                        
                        # In test mode, check if confirmation dialog appears (unsaved changes warning)
                        if self.is_testing_mode:
                            time.sleep(1)  # Wait for dialog to appear if it will
                            if self._handle_leave_and_discard_dialog():
                                self.gui_log("✅ Confirmation dialog handled (Leave and Discard Changes)", level="INFO")
                        
                        time.sleep(2)  # Wait for popup to close
                        return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click X button: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find X button using any method", include_traceback=False)
                self.update_status("Error: Could not find X button", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error closing billing popup", exception=e, include_traceback=True)
            self.update_status(f"Close popup error: {str(e)}", "#dc3545")
            return False
    
    def _handle_leave_and_discard_dialog(self):
        """Handle the "Leave and Discard Changes" confirmation dialog that appears in test mode
        
        This dialog appears only in test/training mode when the bot closes the popup without saving changes.
        In actual runs, changes are saved before closing, so this dialog won't appear.
        
        Element: <psy-button class="button-save hydrated" data-testid="prompt-dialog-button1">Leave and Discard Changes</psy-button>
        
        Returns:
            bool: True if dialog was found and handled, False otherwise
        """
        try:
            self.gui_log("Checking for 'Leave and Discard Changes' confirmation dialog...", level="DEBUG")
            
            # Use a shorter wait time since this might not always appear
            quick_wait = WebDriverWait(self.driver, 3)
            discard_button = None
            
            # Strategy 1: Find by data-testid
            try:
                self.gui_log("Looking for 'Leave and Discard Changes' button by data-testid...", level="DEBUG")
                discard_button = quick_wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//psy-button[@data-testid='prompt-dialog-button1']"))
                )
                self.gui_log("✅ Found 'Leave and Discard Changes' button by data-testid", level="DEBUG")
            except TimeoutException:
                self.gui_log("'Leave and Discard Changes' button not found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by text content if data-testid fails
            if not discard_button:
                try:
                    self.gui_log("Trying to find 'Leave and Discard Changes' button by text...", level="DEBUG")
                    discard_button = quick_wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//psy-button[contains(text(), 'Leave and Discard Changes')]"))
                    )
                    self.gui_log("✅ Found 'Leave and Discard Changes' button by text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("'Leave and Discard Changes' button not found by text, trying other methods...", level="DEBUG")
            
            # Strategy 3: Find by class and text combination
            if not discard_button:
                try:
                    self.gui_log("Trying to find 'Leave and Discard Changes' button by class and text...", level="DEBUG")
                    discard_button = quick_wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//psy-button[contains(@class, 'button-save') and contains(text(), 'Leave and Discard Changes')]"))
                    )
                    self.gui_log("✅ Found 'Leave and Discard Changes' button by class and text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("'Leave and Discard Changes' button not found by class and text, trying other methods...", level="DEBUG")
            
            # Strategy 4: Find button element within shadow DOM (if needed)
            if not discard_button:
                try:
                    self.gui_log("Trying to find button within shadow DOM...", level="DEBUG")
                    # Try to find the button element inside psy-button
                    discard_button = quick_wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//psy-button[@data-testid='prompt-dialog-button1']//button"))
                    )
                    self.gui_log("✅ Found 'Leave and Discard Changes' button within shadow DOM", level="DEBUG")
                except TimeoutException:
                    self.gui_log("'Leave and Discard Changes' button not found within shadow DOM...", level="DEBUG")
            
            if discard_button:
                # Scroll into view
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", discard_button)
                time.sleep(0.5)
                
                # Click the button
                try:
                    discard_button.click()
                    self.gui_log("✅ Clicked 'Leave and Discard Changes' button", level="INFO")
                    self.update_status("Discarded changes", "#28a745")
                    time.sleep(1)  # Wait for dialog to close
                    return True
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", discard_button)
                        self.gui_log("✅ Clicked 'Leave and Discard Changes' button (JavaScript)", level="INFO")
                        self.update_status("Discarded changes", "#28a745")
                        time.sleep(1)  # Wait for dialog to close
                        return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click 'Leave and Discard Changes' button: {click_error}", exception=js_error, include_traceback=False)
                        return False
            else:
                # Button not found - dialog might not have appeared, which is fine
                self.gui_log("No confirmation dialog to handle", level="DEBUG")
                return False
                
        except TimeoutException:
            # Dialog didn't appear - that's fine, just return False
            self.gui_log("No confirmation dialog appeared (or already dismissed)", level="DEBUG")
            return False
        except Exception as e:
            # Log error but don't fail - dialog might not appear
            self.log_error("Error handling 'Leave and Discard Changes' dialog", exception=e, include_traceback=False)
            return False
    
    def _check_and_close_unrelated_windows(self):
        """Check if we're on an unrelated page and close any unrelated tabs/windows
        
        This function detects if the bot has navigated to an unexpected page (like blog.therapynotes.com)
        and closes that tab/window, switching back to the main TherapyNotes application window.
        
        Returns:
            bool: True if unrelated windows were found and closed, False otherwise
        """
        try:
            # Check current URL to see if we're on a non-app page
            current_url = self.driver.current_url
            self.gui_log(f"Checking current URL: {current_url}", level="DEBUG")
            
            # TherapyNotes app pages should be on www.therapynotes.com/app/...
            # Pages like blog.therapynotes.com/status are unrelated
            is_app_page = 'www.therapynotes.com/app' in current_url or 'therapynotes.com/app' in current_url
            
            if not is_app_page:
                self.gui_log(f"⚠️ Detected unrelated page: {current_url}", level="WARNING")
                self.update_status("Detected unrelated page - Closing tab...", "#ff9500")
                
                # Get all window handles
                try:
                    all_windows = self.driver.window_handles
                    self.gui_log(f"Found {len(all_windows)} window(s)/tab(s)", level="DEBUG")
                    
                    if len(all_windows) > 1:
                        # Multiple windows/tabs exist - find and close the unrelated one
                        main_window = None
                        unrelated_window = None
                        
                        # Check each window to find which one is the main app window
                        for window_handle in all_windows:
                            self.driver.switch_to.window(window_handle)
                            window_url = self.driver.current_url
                            self.gui_log(f"Window URL: {window_url}", level="DEBUG")
                            
                            if 'www.therapynotes.com/app' in window_url or 'therapynotes.com/app' in window_url:
                                # This is the main app window
                                if main_window is None:
                                    main_window = window_handle
                            else:
                                # This is an unrelated window (like blog.therapynotes.com)
                                unrelated_window = window_handle
                        
                        # Close the unrelated window if found
                        if unrelated_window:
                            self.driver.switch_to.window(unrelated_window)
                            self.driver.close()
                            self.gui_log(f"✅ Closed unrelated window/tab: {current_url}", level="INFO")
                            self.update_status("Closed unrelated tab", "#28a745")
                            
                            # Switch back to main window
                            if main_window:
                                self.driver.switch_to.window(main_window)
                                time.sleep(1)  # Wait for switch
                                self.gui_log("✅ Switched back to main TherapyNotes window", level="INFO")
                                self.update_status("Back on main page", "#28a745")
                                return True
                            else:
                                # No main window found, switch to first remaining window
                                remaining_windows = self.driver.window_handles
                                if remaining_windows:
                                    self.driver.switch_to.window(remaining_windows[0])
                                    self.gui_log("✅ Switched to remaining window", level="INFO")
                                    return True
                        else:
                            # Current window is unrelated but it's the only window
                            # Navigate back to main app using browser navigation
                            if main_window:
                                self.driver.switch_to.window(main_window)
                                self.gui_log("✅ Switched back to main window", level="INFO")
                                return True
                            else:
                                # Navigate directly to the Patients page
                                self.gui_log("Navigating back to main TherapyNotes app...", level="INFO")
                                self.driver.get("https://www.therapynotes.com/app/patients/")
                                time.sleep(2)  # Wait for navigation
                                self.gui_log("✅ Navigated back to main TherapyNotes app", level="INFO")
                                return True
                    else:
                        # Only one window exists - just navigate back
                        self.gui_log("Only one window exists - Navigating back to main app...", level="INFO")
                        self.driver.get("https://www.therapynotes.com/app/patients/")
                        time.sleep(2)  # Wait for navigation
                        self.gui_log("✅ Navigated back to main TherapyNotes app", level="INFO")
                        self.update_status("Back on main page", "#28a745")
                        return True
                        
                except Exception as window_error:
                    self.log_error("Error managing windows", exception=window_error, include_traceback=False)
                    # Fallback: try to navigate back
                    try:
                        self.driver.get("https://www.therapynotes.com/app/patients/")
                        time.sleep(2)
                        self.gui_log("✅ Navigated back to main TherapyNotes app (fallback)", level="INFO")
                        return True
                    except Exception as nav_error:
                        self.log_error("Error navigating back", exception=nav_error, include_traceback=False)
                        return False
            else:
                # We're on the correct page, but check if there are multiple windows
                # (unrelated tab might have opened but we're still on the main window)
                try:
                    all_windows = self.driver.window_handles
                    if len(all_windows) > 1:
                        # Check other windows for unrelated pages
                        main_window = self.driver.current_window_handle
                        unrelated_windows = []
                        
                        for window_handle in all_windows:
                            if window_handle != main_window:
                                self.driver.switch_to.window(window_handle)
                                window_url = self.driver.current_url
                                
                                # Check if this is an unrelated page
                                if 'www.therapynotes.com/app' not in window_url and 'therapynotes.com/app' not in window_url:
                                    unrelated_windows.append(window_handle)
                                
                                # Switch back to main window
                                self.driver.switch_to.window(main_window)
                        
                        # Close unrelated windows
                        if unrelated_windows:
                            for window_handle in unrelated_windows:
                                self.driver.switch_to.window(window_handle)
                                window_url = self.driver.current_url
                                self.gui_log(f"Closing unrelated tab: {window_url}", level="INFO")
                                self.driver.close()
                                self.update_status("Closed unrelated tab", "#28a745")
                            
                            # Switch back to main window
                            self.driver.switch_to.window(main_window)
                            self.gui_log(f"✅ Closed {len(unrelated_windows)} unrelated tab(s)", level="INFO")
                            return True
                except Exception as check_error:
                    # If there's an error checking, just continue - we're on the right page
                    self.gui_log(f"Error checking other windows: {check_error}", level="DEBUG")
                    return False
                
                # We're on the correct page and no unrelated windows
                return False
                
        except Exception as e:
            self.log_error("Error checking for unrelated windows", exception=e, include_traceback=False)
            # Don't fail - just continue
            return False
    
    def _add_correct_modifier_client(self):
        """Add current client to correct_modifier_clients list for output Excel"""
        try:
            client_name = getattr(self, 'current_client_name', 'Unknown')
            dob = getattr(self, 'current_client_dob', '')
            dos = getattr(self, 'current_date_of_service', '')
            
            self.correct_modifier_clients.append({
                'client_name': client_name,
                'dob': dob,
                'date_of_service': dos,
                'reason': 'Modifier was already correct, Not Resubmitted'
            })
            
            self.gui_log(f"📝 Client '{client_name}' added to correct modifier list (reason: Modifier was already correct, Not Resubmitted)", level="INFO")
            
        except Exception as e:
            self.log_error("Error adding client to correct modifier list", exception=e, include_traceback=True)
    
    def _add_tracked_client(self, status, original_modifier=None, new_modifier=None, modifier_action=None, reason=None, error_message=None):
        """Add client to comprehensive tracking list with all details
        
        Args:
            status: Status string (e.g., 'Refiled', 'Not Refiled - No View ERA', 'Not Refiled - Modifier Correct', 'Error')
            original_modifier: Original modifier value ('93', '95', or None)
            new_modifier: New modifier value (what it was changed to, or 'No Change')
            modifier_action: Action taken ('Updated', 'Already Correct', 'Not Checked', 'N/A')
            reason: Detailed reason for status
            error_message: Error details if any
        """
        try:
            client_name = getattr(self, 'current_client_name', 'Unknown')
            dob = getattr(self, 'current_client_dob', '')
            dos = getattr(self, 'current_date_of_service', '')
            excel_row = None
            insurance = None
            claim_status = None
            
            # Find Excel row number, insurance, and claim status if available
            if self.excel_client_data:
                for client in self.excel_client_data:
                    if (client.get('client_name') == client_name and 
                        client.get('dob') == dob and 
                        client.get('date_of_service') == dos):
                        excel_row = client.get('excel_row')
                        insurance = client.get('insurance')  # Extract insurance from Excel data
                        claim_status = client.get('claim_status')  # Extract claim status from Excel data
                        break
            
            # Get session medium and expected modifier
            session_medium = getattr(self, 'current_session_medium', None) or 'N/A'
            expected_modifier = 'N/A'
            if session_medium == 'video':
                expected_modifier = '95'
            elif session_medium == 'phone':
                expected_modifier = '93'
            
            # Get payer claim control number
            payer_claim_control = getattr(self, 'current_payer_claim_control', None) or 'N/A'
            
            # Get primary policy details captured from popup
            primary_policy = getattr(self, 'current_primary_policy_name', None) or 'N/A'
            primary_policy_member_id = getattr(self, 'current_primary_policy_member_id', None) or 'N/A'
            primary_policy_payment = getattr(self, 'current_primary_payment_description', None) or 'N/A'
            
            # Determine modifier values
            if original_modifier is None:
                original_modifier = 'N/A'
            if new_modifier is None:
                if modifier_action == 'Already Correct':
                    new_modifier = 'No Change'
                else:
                    new_modifier = 'N/A'
            
            # Create comprehensive tracking record
            tracked_client = {
                'client_name': client_name,
                'dob': dob,
                'date_of_service': dos,
                'insurance': insurance or 'N/A',  # Add insurance from Excel
                'claim_status': claim_status or 'N/A',  # Add claim submission status from Excel
                'excel_row': excel_row or 'N/A',
                'status': status,
                'original_modifier': str(original_modifier),
                'new_modifier': str(new_modifier),
                'session_medium': session_medium,
                'expected_modifier': expected_modifier,
                'payer_claim_control': str(payer_claim_control),
                'primary_policy': str(primary_policy),
                'primary_policy_member_id': str(primary_policy_member_id),
                'primary_policy_payment_match': str(primary_policy_payment),
                'modifier_action': modifier_action or 'N/A',
                'reason': reason or '',
                'processing_date': time.strftime("%Y-%m-%d %H:%M:%S"),
                'error_message': error_message or ''
            }
            
            self.tracked_clients.append(tracked_client)
            self.gui_log(f"📊 Client '{client_name}' tracked: {status}", level="DEBUG")
            
        except Exception as e:
            self.log_error("Error adding client to tracked list", exception=e, include_traceback=True)
    
    def _click_status_tab(self):
        """Click the Status tab in the popup (similar to Notes and History tabs)
        
        The Status tab should be in the same popup where Notes and History tabs are located.
        
        Returns:
            bool: True if Status tab clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 21: Looking for Status tab in popup...", level="INFO")
            self.update_status("Looking for Status tab...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for popup content to fully load
            time.sleep(2)
            
            # Find Status tab using multiple strategies
            status_tab = None
            
            # Strategy 1: Find by link text "Status"
            try:
                self.gui_log("Looking for Status tab (link text 'Status')...", level="DEBUG")
                status_tab = self.wait.until(
                    EC.element_to_be_clickable((By.LINK_TEXT, "Status"))
                )
                self.gui_log("✅ Found Status tab by link text", level="DEBUG")
            except TimeoutException:
                self.gui_log("Status tab not found by link text, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by partial link text
            if not status_tab:
                try:
                    self.gui_log("Trying to find 'Status' link by partial text...", level="DEBUG")
                    status_tab = self.wait.until(
                        EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Status"))
                    )
                    self.gui_log("✅ Found Status tab by partial link text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Status link not found by partial text, trying other methods...", level="DEBUG")
            
            # Strategy 3: Find by XPath with href="#" and text "Status"
            if not status_tab:
                try:
                    self.gui_log("Trying to find Status tab by XPath with href and text...", level="DEBUG")
                    status_tab = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[@href='#' and contains(text(), 'Status')]"))
                    )
                    self.gui_log("✅ Found Status tab by XPath with href and text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Status tab not found by XPath...", level="DEBUG")
            
            # Strategy 4: Find by data-testid pattern containing "status-tab"
            if not status_tab:
                try:
                    status_tab = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[contains(@data-testid, 'status-tab')]"))
                    )
                    self.gui_log("✅ Found Status tab by data-testid pattern", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Status tab not found by data-testid pattern...", level="DEBUG")
            
            # Strategy 5: Find by tabindex and text "Status"
            if not status_tab:
                try:
                    status_tab = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[@tabindex='0' and contains(text(), 'Status')]"))
                    )
                    self.gui_log("✅ Found Status tab by tabindex and text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Status tab not found by tabindex...", level="DEBUG")
            
            if status_tab:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", status_tab)
                time.sleep(0.5)
                
                # Click the Status tab
                try:
                    status_tab.click()
                    self.gui_log("✅ Status tab clicked", level="INFO")
                    self.update_status("Status tab clicked - Waiting for content to load...", "#ff9500")
                    
                    # Wait for Status tab content to load
                    time.sleep(2)
                    
                    # Verify navigation by checking if status dropdown is visible
                    try:
                        status_dropdown = self.wait.until(
                            EC.presence_of_element_located((By.XPATH, "//select[@data-testid='calendarentrybillingtabservicerow-insurancestatus-input-0']"))
                        )
                        self.gui_log("✅ Status tab content loaded - Status dropdown found", level="INFO")
                        self.update_status("Status tab loaded", "#28a745")
                        return True
                    except TimeoutException:
                        self.gui_log("⚠️ Status dropdown not found, but tab may have loaded", level="WARNING")
                        return True  # Still return True as tab was clicked
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", status_tab)
                        self.gui_log("✅ Status tab clicked (JavaScript)", level="INFO")
                        time.sleep(2)
                        return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click Status tab: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find Status tab using any method", include_traceback=False)
                self.update_status("Error: Could not find Status tab", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error clicking Status tab", exception=e, include_traceback=True)
            self.update_status(f"Status tab error: {str(e)}", "#dc3545")
            return False
    
    def _update_status_dropdown(self, status_option_text):
        """Update the status dropdown in the Status tab by selecting an option
        
        Element: <select class="CalendarEntryBillingTab_StatusDropdown" data-testid="calendarentrybillingtabservicerow-insurancestatus-input-0">
        Option element: <option value="2">Pending Resubmission</option>
        
        Args:
            status_option_text: The text of the option to select (e.g., "Pending Resubmission")
        
        Returns:
            bool: True if status dropdown updated successfully, False otherwise
        """
        try:
            self.gui_log(f"Step 22: Updating status dropdown to '{status_option_text}'...", level="INFO")
            self.update_status(f"Updating status dropdown to '{status_option_text}'...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for Status tab content to fully load
            time.sleep(2)
            
            # Find the status dropdown using multiple strategies
            status_dropdown = None
            
            # Strategy 1: Find by data-testid
            try:
                self.gui_log("Looking for status dropdown by data-testid...", level="DEBUG")
                status_dropdown = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, "//select[@data-testid='calendarentrybillingtabservicerow-insurancestatus-input-0']"))
                )
                self.gui_log("✅ Found status dropdown by data-testid", level="DEBUG")
            except TimeoutException:
                self.gui_log("Status dropdown not found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by class
            if not status_dropdown:
                try:
                    self.gui_log("Trying to find status dropdown by class...", level="DEBUG")
                    status_dropdown = self.wait.until(
                        EC.presence_of_element_located((By.XPATH, "//select[contains(@class, 'CalendarEntryBillingTab_StatusDropdown')]"))
                    )
                    self.gui_log("✅ Found status dropdown by class", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Status dropdown not found by class...", level="DEBUG")
            
            # Strategy 3: Find by containing the status options
            if not status_dropdown:
                try:
                    self.gui_log("Trying to find status dropdown by option text...", level="DEBUG")
                    status_dropdown = self.wait.until(
                        EC.presence_of_element_located((By.XPATH, "//select[.//option[text()='Pending Resubmission']]"))
                    )
                    self.gui_log("✅ Found status dropdown by option text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Status dropdown not found by option text...", level="DEBUG")
            
            if not status_dropdown:
                self.log_error("Could not find status dropdown using any method", include_traceback=False)
                self.update_status("Error: Could not find status dropdown", "#dc3545")
                return False
            
            # Scroll into view (but don't trigger any events - just position it)
            # Use a more gentle scroll that won't trigger events
            try:
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'auto', block: 'nearest'});", status_dropdown)
                time.sleep(0.3)
            except:
                pass  # If scroll fails, continue anyway
            
            # Try human-like interaction: click the dropdown to open it, then click the option
            # This is more similar to what a human does and should avoid triggering unwanted events
            try:
                self.gui_log(f"Selecting status option: '{status_option_text}'...", level="INFO")
                
                # First, click the dropdown to open it (like a human would)
                try:
                    status_dropdown.click()
                    time.sleep(0.5)  # Wait for dropdown options to appear
                except Exception as click_error:
                    self.gui_log(f"Could not click dropdown to open: {click_error}", level="DEBUG")
                
                # Find and click the option with the matching text
                option_xpath = f"//select[@data-testid='calendarentrybillingtabservicerow-insurancestatus-input-0']//option[text()='{status_option_text}']"
                try:
                    option = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, option_xpath))
                    )
                    option.click()
                    self.gui_log(f"✅ Status dropdown updated to '{status_option_text}' (clicked option directly)", level="INFO")
                except TimeoutException:
                    # Fallback: try finding option in a select element
                    try:
                        option = self.driver.find_element(By.XPATH, f"//option[text()='{status_option_text}']")
                        option.click()
                        self.gui_log(f"✅ Status dropdown updated to '{status_option_text}' (fallback click)", level="INFO")
                    except:
                        # Final fallback: use Select class (original method)
                        from selenium.webdriver.support.ui import Select
                        select = Select(status_dropdown)
                        select.select_by_visible_text(status_option_text)
                        self.gui_log(f"✅ Status dropdown updated to '{status_option_text}' (using Select class fallback)", level="INFO")
                
                self.update_status(f"Status updated: {status_option_text}", "#28a745")
                
                # Verify selection by checking the selected value
                time.sleep(0.5)
                try:
                    from selenium.webdriver.support.ui import Select
                    select = Select(status_dropdown)
                    selected_option = select.first_selected_option
                    if selected_option.text.strip() == status_option_text:
                        self.gui_log(f"✅ Verified status dropdown is set to '{status_option_text}'", level="INFO")
                    else:
                        self.gui_log(f"⚠️ Status dropdown selected '{selected_option.text}', expected '{status_option_text}'", level="WARNING")
                except:
                    # If verification fails, still continue
                    self.gui_log("⚠️ Could not verify selection, but continuing...", level="DEBUG")
                
                # Check for and close any unrelated windows that may have opened
                self._check_and_close_unrelated_windows()
                
                return True
                    
            except Exception as select_error:
                # Fallback: Try using Select class if direct clicking failed
                try:
                    self.gui_log(f"Direct click method failed, trying Select class fallback...", level="DEBUG")
                    from selenium.webdriver.support.ui import Select
                    select = Select(status_dropdown)
                    
                    # Try selecting by visible text first
                    try:
                        select.select_by_visible_text(status_option_text)
                        self.gui_log(f"✅ Status dropdown updated to '{status_option_text}' (Select class fallback)", level="INFO")
                        self.update_status(f"Status updated: {status_option_text}", "#28a745")
                        
                        # Check for and close any unrelated windows that may have opened
                        self._check_and_close_unrelated_windows()
                        
                        return True
                    except:
                        # Try selecting by value if visible text fails
                        status_value_map = {
                            "Pending Initial Submission": "1",
                            "Pending Resubmission": "2",
                            "Submitted External": "5-0",
                            "CMS Printed": "5-4",
                            "Superbill Printed": "5-1",
                            "CMS & Superbill Printed": "5-5",
                            "Submitted Electronically": "3",
                            "Partial Payment": "6",
                            "Paid": "4"
                        }
                        
                        if status_option_text in status_value_map:
                            value = status_value_map[status_option_text]
                            self.gui_log(f"Trying to select by value '{value}'...", level="DEBUG")
                            select.select_by_value(value)
                            self.gui_log(f"✅ Status dropdown updated to '{status_option_text}' (by value)", level="INFO")
                            self.update_status(f"Status updated: {status_option_text}", "#28a745")
                            
                            # Check for and close any unrelated windows that may have opened
                            self._check_and_close_unrelated_windows()
                            
                            return True
                        else:
                            self.log_error(f"Unknown status option: {status_option_text}", exception=select_error, include_traceback=True)
                            return False
                except Exception as value_error:
                    self.log_error(f"Error selecting status option: {select_error}", exception=value_error, include_traceback=True)
                    return False
                
        except Exception as e:
            self.log_error("Error updating status dropdown", exception=e, include_traceback=True)
            self.update_status(f"Status dropdown error: {str(e)}", "#dc3545")
            return False
    
    def _click_save_changes_button(self):
        """Click the Save Changes button after updating the modifier and status
        
        Element: <psy-button data-testid="calendarentrybillingtab-save-button" class="hydrated">
        The button text is "Save Changes" and it's inside a Shadow DOM custom element.
        
        Returns:
            bool: True if Save Changes button clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 23: Looking for Save Changes button...", level="INFO")
            self.update_status("Looking for Save Changes button...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait a moment for changes to be ready to save
            time.sleep(2)
            
            # Find Save Changes button using multiple strategies
            save_button = None
            
            # Strategy 1: Find by data-testid on psy-button
            try:
                self.gui_log("Looking for Save Changes button by data-testid...", level="DEBUG")
                save_button = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//psy-button[@data-testid='calendarentrybillingtab-save-button']"))
                )
                self.gui_log("✅ Found Save Changes button by data-testid", level="DEBUG")
            except TimeoutException:
                self.gui_log("Save Changes button not found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find button by text "Save Changes" (may be in Shadow DOM)
            if not save_button:
                try:
                    self.gui_log("Trying to find Save Changes button by text...", level="DEBUG")
                    # Look for button containing "Save Changes" text
                    save_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Save Changes')] | //*[contains(text(), 'Save Changes') and (local-name()='button' or local-name()='psy-button')]"))
                    )
                    self.gui_log("✅ Found Save Changes button by text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Save Changes button not found by text...", level="DEBUG")
            
            # Strategy 3: Find psy-button and access Shadow DOM button
            if not save_button:
                try:
                    self.gui_log("Trying to find Save Changes button via Shadow DOM...", level="DEBUG")
                    # Find the psy-button element
                    psy_button = self.wait.until(
                        EC.presence_of_element_located((By.XPATH, "//psy-button[@data-testid='calendarentrybillingtab-save-button']"))
                    )
                    # Access the button inside Shadow DOM via JavaScript
                    save_button = self.driver.execute_script("""
                        var psyButton = arguments[0];
                        var shadowRoot = psyButton.shadowRoot;
                        if (shadowRoot) {
                            return shadowRoot.querySelector('button');
                        }
                        return null;
                    """, psy_button)
                    if save_button:
                        self.gui_log("✅ Found Save Changes button via Shadow DOM", level="DEBUG")
                except Exception as shadow_error:
                    self.gui_log(f"Shadow DOM access failed: {shadow_error}", level="DEBUG")
            
            # Strategy 4: Find by button with data-variant="accent"
            if not save_button:
                try:
                    self.gui_log("Trying to find Save Changes button by data-variant...", level="DEBUG")
                    # Try to find button via XPath (may be inside Shadow DOM)
                    save_button = self.driver.find_element(By.XPATH, "//button[@data-variant='accent' and contains(text(), 'Save Changes')]")
                    self.gui_log("✅ Found Save Changes button by data-variant", level="DEBUG")
                except:
                    pass
            
            # Strategy 5: Find psy-button and click it directly (it should be clickable)
            if not save_button:
                try:
                    self.gui_log("Trying to click psy-button directly...", level="DEBUG")
                    psy_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//psy-button[@data-testid='calendarentrybillingtab-save-button']"))
                    )
                    # Click the psy-button element directly - it should handle the click
                    psy_button.click()
                    self.gui_log("✅ Save Changes button clicked (via psy-button)", level="INFO")
                    self.update_status("Save Changes clicked", "#28a745")
                    time.sleep(3)  # Wait for save to complete
                    return True
                except Exception as direct_error:
                    self.gui_log(f"Direct psy-button click failed: {direct_error}", level="DEBUG")
            
            if save_button:
                # Scroll into view
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", save_button)
                time.sleep(0.5)
                
                # Click the Save Changes button
                try:
                    save_button.click()
                    self.gui_log("✅ Save Changes button clicked", level="INFO")
                    self.update_status("Save Changes clicked - Waiting for save to complete...", "#28a745")
                    
                    # Wait for save to complete
                    time.sleep(3)
                    
                    self.gui_log("✅ Changes saved successfully", level="INFO")
                    self.update_status("Changes saved successfully", "#28a745")
                    return True
                    
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", save_button)
                        self.gui_log("✅ Save Changes button clicked (JavaScript)", level="INFO")
                        self.update_status("Save Changes clicked - Waiting for save to complete...", "#28a745")
                        time.sleep(3)
                        return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click Save Changes button: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find Save Changes button using any method", include_traceback=False)
                self.update_status("Error: Could not find Save Changes button", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error clicking Save Changes button", exception=e, include_traceback=True)
            self.update_status(f"Save Changes error: {str(e)}", "#dc3545")
            return False
    
    def _click_submit_primary_claims(self):
        """Click the "Submit Primary Claims" button after saving changes
        
        Element: <a href="#" id="BillingLinksPanel__SubmitPrimaryClaimsLink" data-testid="billing-links-panel-link-SubmitPrimaryClaims">Submit Primary Claims<span class="ClaimsCount countBubble__span " id="BillingLinksPanel__SubmitPrimaryClaimsLinkCountBubble" data-testid="count-bubble">3</span></a>
        
        This button navigates to a new page where we can filter and submit claims for refiling.
        
        Returns:
            bool: True if button clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 24: Looking for 'Submit Primary Claims' button...", level="INFO")
            self.update_status("Looking for Submit Primary Claims button...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait a moment for popup to close and page to be ready after Save Changes
            time.sleep(2)
            
            # Find Submit Primary Claims button using multiple strategies
            submit_button = None
            
            # Strategy 1: Find by data-testid
            try:
                self.gui_log("Looking for 'Submit Primary Claims' button by data-testid...", level="DEBUG")
                submit_button = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-testid='billing-links-panel-link-SubmitPrimaryClaims']"))
                )
                self.gui_log("✅ Found 'Submit Primary Claims' button by data-testid", level="DEBUG")
            except TimeoutException:
                self.gui_log("'Submit Primary Claims' button not found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by ID
            if not submit_button:
                try:
                    self.gui_log("Trying to find 'Submit Primary Claims' button by ID...", level="DEBUG")
                    submit_button = self.wait.until(
                        EC.element_to_be_clickable((By.ID, "BillingLinksPanel__SubmitPrimaryClaimsLink"))
                    )
                    self.gui_log("✅ Found 'Submit Primary Claims' button by ID", level="DEBUG")
                except TimeoutException:
                    self.gui_log("'Submit Primary Claims' button not found by ID, trying other methods...", level="DEBUG")
            
            # Strategy 3: Find by text content
            if not submit_button:
                try:
                    self.gui_log("Trying to find 'Submit Primary Claims' button by text...", level="DEBUG")
                    submit_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Submit Primary Claims')]"))
                    )
                    self.gui_log("✅ Found 'Submit Primary Claims' button by text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("'Submit Primary Claims' button not found by text...", level="DEBUG")
            
            # Strategy 4: Find by partial text match
            if not submit_button:
                try:
                    self.gui_log("Trying to find 'Submit Primary Claims' button by partial text...", level="DEBUG")
                    submit_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[contains(., 'Submit Primary Claims')]"))
                    )
                    self.gui_log("✅ Found 'Submit Primary Claims' button by partial text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("'Submit Primary Claims' button not found by partial text...", level="DEBUG")
            
            if submit_button:
                # Scroll into view
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", submit_button)
                time.sleep(0.5)
                
                # Click the button
                try:
                    submit_button.click()
                    self.gui_log("✅ 'Submit Primary Claims' button clicked", level="INFO")
                    self.update_status("Navigating to Submit Primary Claims page...", "#ff9500")
                    
                    # Wait for navigation to complete
                    try:
                        WebDriverWait(self.driver, 5).until(
                            lambda driver: driver.execute_script("return document.readyState") == "complete"
                        )
                        time.sleep(2)  # Wait for page to fully load
                        
                        # Verify navigation by checking URL
                        current_url = self.driver.current_url
                        self.gui_log(f"Current URL after clicking Submit Primary Claims: {current_url}", level="DEBUG")
                        
                        self.gui_log("✅ Successfully navigated to Submit Primary Claims page", level="INFO")
                        self.update_status("On Submit Primary Claims page", "#28a745")
                        return True
                    except TimeoutException:
                        # Still return True if navigation seems to have occurred
                        self.gui_log("⚠️ Timeout waiting for page load, but navigation may have occurred", level="WARNING")
                        return True
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", submit_button)
                        self.gui_log("✅ 'Submit Primary Claims' button clicked (JavaScript)", level="INFO")
                        self.update_status("Navigating to Submit Primary Claims page...", "#ff9500")
                        time.sleep(2)
                        
                        current_url = self.driver.current_url
                        self.gui_log(f"Current URL after clicking Submit Primary Claims (JavaScript): {current_url}", level="DEBUG")
                        
                        self.gui_log("✅ Successfully navigated to Submit Primary Claims page", level="INFO")
                        self.update_status("On Submit Primary Claims page", "#28a745")
                        return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click 'Submit Primary Claims' button: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find 'Submit Primary Claims' button using any method", include_traceback=False)
                self.update_status("Error: Could not find Submit Primary Claims button", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error clicking 'Submit Primary Claims' button", exception=e, include_traceback=True)
            self.update_status(f"Submit Primary Claims error: {str(e)}", "#dc3545")
            return False
    
    def _filter_date_range_for_refiling(self):
        """Filter the date range on the Submit Primary Claims page to match the date of service being refiled
        
        This function:
        1. Finds the Start Date input field and enters the date of service
        2. Finds the End Date input field and enters the same date of service
        3. Clicks the Search button to filter the results
        
        Elements:
        - Start Date: <input type="text" id="DateRangeSelector_StartDateInput" aria-label="start date" data-testid="sec-daterangestartdate-input" maxlength="20" class="date-range-selector-input" autocomplete="off" placeholder="m/d/yyyy">
        - End Date: <input type="text" aria-label="end date" data-testid="sec-daterangeenddate-input" id="DateRangeSelector_EndDateInput" maxlength="20" class="date-range-selector-input" autocomplete="off" placeholder="m/d/yyyy">
        - Search Button: <input name="BatchClaimCreator__Search-Button" id="BatchClaimCreator__Search-Button" type="button" value="Search" tabindex="0">
        
        Returns:
            bool: True if date range filtered successfully, False otherwise
        """
        try:
            self.gui_log("Step 25: Filtering date range for refiling...", level="INFO")
            self.update_status("Filtering date range...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for page to be ready
            time.sleep(2)
            
            # Get the date of service and format it for the date range inputs
            # The date should be in format m/d/yyyy (e.g., 8/1/2025)
            dos = getattr(self, 'current_date_of_service', None)
            if not dos:
                self.log_error("No date of service available for filtering", include_traceback=False)
                return False
            
            # Format the date to Therapy Notes format (m/d/yyyy) using existing normalization function
            dos_str = str(dos)
            normalized_dos_str, dos_datetime = self._normalize_date_of_service(dos_str)
            
            if not normalized_dos_str or not dos_datetime:
                self.log_error(f"Could not normalize date of service for filtering: {dos_str}", include_traceback=False)
                return False
            
            # Convert normalized date (which is in m/d/yy format) to m/d/yyyy format for date range inputs
            # The normalized format might be like "8/1/25", we need "8/1/2025"
            try:
                from datetime import datetime
                # Try parsing the normalized date and reformat to full year
                parsed_date = datetime.strptime(normalized_dos_str, '%m/%d/%y')
                formatted_date = parsed_date.strftime('%m/%d/%Y')
            except:
                # Fallback: try to use the datetime object directly if available
                if dos_datetime:
                    formatted_date = dos_datetime.strftime('%m/%d/%Y')
                else:
                    # Last resort: try to parse the original date string
                    try:
                        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S', '%m/%d/%y']:
                            try:
                                dt = datetime.strptime(dos_str.split()[0] if ' ' in dos_str else dos_str, fmt)
                                formatted_date = dt.strftime('%m/%d/%Y')
                                break
                            except ValueError:
                                continue
                        else:
                            formatted_date = normalized_dos_str  # Use as-is if we can't parse
                    except:
                        formatted_date = normalized_dos_str  # Use as-is if we can't parse
            
            self.gui_log(f"Filtering for date: {formatted_date}", level="INFO")
            
            # Find and fill Start Date input
            start_date_input = None
            try:
                self.gui_log("Looking for Start Date input field...", level="DEBUG")
                # Strategy 1: Find by data-testid
                try:
                    start_date_input = self.wait.until(
                        EC.presence_of_element_located((By.XPATH, "//input[@data-testid='sec-daterangestartdate-input']"))
                    )
                    self.gui_log("✅ Found Start Date input by data-testid", level="DEBUG")
                except TimeoutException:
                    # Strategy 2: Find by ID
                    try:
                        start_date_input = self.wait.until(
                            EC.presence_of_element_located((By.ID, "DateRangeSelector_StartDateInput"))
                        )
                        self.gui_log("✅ Found Start Date input by ID", level="DEBUG")
                    except TimeoutException:
                        # Strategy 3: Find by aria-label
                        start_date_input = self.wait.until(
                            EC.presence_of_element_located((By.XPATH, "//input[@aria-label='start date']"))
                        )
                        self.gui_log("✅ Found Start Date input by aria-label", level="DEBUG")
            except TimeoutException:
                self.log_error("Could not find Start Date input field", include_traceback=False)
                return False
            
            # Find and fill End Date input
            end_date_input = None
            try:
                self.gui_log("Looking for End Date input field...", level="DEBUG")
                # Strategy 1: Find by data-testid
                try:
                    end_date_input = self.wait.until(
                        EC.presence_of_element_located((By.XPATH, "//input[@data-testid='sec-daterangeenddate-input']"))
                    )
                    self.gui_log("✅ Found End Date input by data-testid", level="DEBUG")
                except TimeoutException:
                    # Strategy 2: Find by ID
                    try:
                        end_date_input = self.wait.until(
                            EC.presence_of_element_located((By.ID, "DateRangeSelector_EndDateInput"))
                        )
                        self.gui_log("✅ Found End Date input by ID", level="DEBUG")
                    except TimeoutException:
                        # Strategy 3: Find by aria-label
                        end_date_input = self.wait.until(
                            EC.presence_of_element_located((By.XPATH, "//input[@aria-label='end date']"))
                        )
                        self.gui_log("✅ Found End Date input by aria-label", level="DEBUG")
            except TimeoutException:
                self.log_error("Could not find End Date input field", include_traceback=False)
                return False
            
            # Fill both date inputs
            try:
                # Clear and fill Start Date
                start_date_input.clear()
                time.sleep(0.3)
                start_date_input.send_keys(formatted_date)
                self.gui_log(f"✅ Start Date entered: {formatted_date}", level="INFO")
                
                # Clear and fill End Date (same date)
                end_date_input.clear()
                time.sleep(0.3)
                end_date_input.send_keys(formatted_date)
                self.gui_log(f"✅ End Date entered: {formatted_date}", level="INFO")
                
                # Find and click Search button
                search_button = None
                try:
                    self.gui_log("Looking for Search button...", level="DEBUG")
                    # Strategy 1: Find by ID
                    try:
                        search_button = self.wait.until(
                            EC.element_to_be_clickable((By.ID, "BatchClaimCreator__Search-Button"))
                        )
                        self.gui_log("✅ Found Search button by ID", level="DEBUG")
                    except TimeoutException:
                        # Strategy 2: Find by name
                        try:
                            search_button = self.wait.until(
                                EC.element_to_be_clickable((By.NAME, "BatchClaimCreator__Search-Button"))
                            )
                            self.gui_log("✅ Found Search button by name", level="DEBUG")
                        except TimeoutException:
                            # Strategy 3: Find by value="Search"
                            search_button = self.wait.until(
                                EC.element_to_be_clickable((By.XPATH, "//input[@type='button' and @value='Search']"))
                            )
                            self.gui_log("✅ Found Search button by value", level="DEBUG")
                except TimeoutException:
                    self.log_error("Could not find Search button", include_traceback=False)
                    return False
                
                # Click Search button
                if search_button:
                    search_button.click()
                    self.gui_log("✅ Search button clicked", level="INFO")
                    self.update_status("Searching for claims...", "#ff9500")
                    
                    # Wait for search results to load
                    time.sleep(3)
                    
                    self.gui_log("✅ Date range filtered successfully", level="INFO")
                    self.update_status("Date range filtered", "#28a745")
                    return True
                else:
                    self.log_error("Search button not found", include_traceback=False)
                    return False
                    
            except Exception as fill_error:
                self.log_error("Error filling date range inputs", exception=fill_error, include_traceback=True)
                return False
                
        except Exception as e:
            self.log_error("Error filtering date range for refiling", exception=e, include_traceback=True)
            self.update_status(f"Date range filter error: {str(e)}", "#dc3545")
            return False
    
    def _select_dos_checkbox(self):
        """Select the checkbox for the date of service on the Submit Primary Claims page
        
        Element: <input type="checkbox" data-testid="pendingclaimappointment-selectitem-checkbox-0">
        
        Returns:
            bool: True if checkbox selected successfully, False otherwise
        """
        try:
            self.gui_log("Step 26: Looking for DOS checkbox...", level="INFO")
            self.update_status("Selecting DOS checkbox...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for checkbox to be available
            time.sleep(1)
            
            # Find the checkbox using multiple strategies
            checkbox = None
            
            # Strategy 1: Find by data-testid
            try:
                self.gui_log("Looking for DOS checkbox by data-testid...", level="DEBUG")
                checkbox = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//input[@type='checkbox' and @data-testid='pendingclaimappointment-selectitem-checkbox-0']"))
                )
                self.gui_log("✅ Found DOS checkbox by data-testid", level="DEBUG")
            except TimeoutException:
                self.gui_log("DOS checkbox not found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by data-testid pattern (try index 0, 1, 2, etc.)
            if not checkbox:
                for index in range(5):  # Try first 5 checkboxes
                    try:
                        self.gui_log(f"Trying to find DOS checkbox by data-testid with index {index}...", level="DEBUG")
                        checkbox = self.wait.until(
                            EC.element_to_be_clickable((By.XPATH, f"//input[@type='checkbox' and @data-testid='pendingclaimappointment-selectitem-checkbox-{index}']"))
                        )
                        self.gui_log(f"✅ Found DOS checkbox by data-testid with index {index}", level="DEBUG")
                        break
                    except TimeoutException:
                        continue
            
            # Strategy 3: Find first unchecked checkbox
            if not checkbox:
                try:
                    self.gui_log("Trying to find first unchecked checkbox...", level="DEBUG")
                    checkboxes = self.driver.find_elements(By.XPATH, "//input[@type='checkbox' and contains(@data-testid, 'pendingclaimappointment-selectitem-checkbox')]")
                    for cb in checkboxes:
                        if not cb.is_selected():
                            checkbox = cb
                            self.gui_log("✅ Found unchecked DOS checkbox", level="DEBUG")
                            break
                except Exception:
                    self.gui_log("Could not find checkbox by unchecked status...", level="DEBUG")
            
            if checkbox:
                # Check if already selected
                if checkbox.is_selected():
                    self.gui_log("✅ DOS checkbox is already selected", level="INFO")
                    self.update_status("DOS checkbox selected", "#28a745")
                    return True
                
                # Scroll into view
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", checkbox)
                time.sleep(0.5)
                
                # Click the checkbox
                try:
                    checkbox.click()
                    self.gui_log("✅ DOS checkbox selected", level="INFO")
                    self.update_status("DOS checkbox selected", "#28a745")
                    time.sleep(0.5)  # Wait for selection to register
                    return True
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", checkbox)
                        self.gui_log("✅ DOS checkbox selected (JavaScript)", level="INFO")
                        self.update_status("DOS checkbox selected", "#28a745")
                        time.sleep(0.5)
                        return True
                    except Exception as js_error:
                        self.log_error(f"Failed to select DOS checkbox: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find DOS checkbox using any method", include_traceback=False)
                self.update_status("Error: Could not find DOS checkbox", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error selecting DOS checkbox", exception=e, include_traceback=True)
            self.update_status(f"DOS checkbox error: {str(e)}", "#dc3545")
            return False
    
    def _select_amended_claim(self):
        """Select "Amended Claim" from the resubmission type dropdown
        
        Element: <select data-testid="pendingclaim-resubmissionclaimfrequency-select-0" id="resubmission-claim-frequency-select-id-0">
        Option: <option value="2">Amended Claim</option>
        
        Returns:
            bool: True if "Amended Claim" selected successfully, False otherwise
        """
        try:
            self.gui_log("Step 27: Selecting 'Amended Claim' from resubmission type dropdown...", level="INFO")
            self.update_status("Selecting Amended Claim...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for dropdown to be available
            time.sleep(1)
            
            # Find the dropdown using multiple strategies
            dropdown = None
            
            # Strategy 1: Find by data-testid
            try:
                self.gui_log("Looking for resubmission type dropdown by data-testid...", level="DEBUG")
                dropdown = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, "//select[@data-testid='pendingclaim-resubmissionclaimfrequency-select-0']"))
                )
                self.gui_log("✅ Found resubmission type dropdown by data-testid", level="DEBUG")
            except TimeoutException:
                self.gui_log("Resubmission type dropdown not found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by ID
            if not dropdown:
                try:
                    self.gui_log("Trying to find resubmission type dropdown by ID...", level="DEBUG")
                    dropdown = self.wait.until(
                        EC.presence_of_element_located((By.ID, "resubmission-claim-frequency-select-id-0"))
                    )
                    self.gui_log("✅ Found resubmission type dropdown by ID", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Resubmission type dropdown not found by ID, trying other methods...", level="DEBUG")
            
            # Strategy 3: Find by data-testid pattern (try index 0, 1, 2, etc.)
            if not dropdown:
                for index in range(5):  # Try first 5 dropdowns
                    try:
                        self.gui_log(f"Trying to find resubmission type dropdown by data-testid with index {index}...", level="DEBUG")
                        dropdown = self.wait.until(
                            EC.presence_of_element_located((By.XPATH, f"//select[@data-testid='pendingclaim-resubmissionclaimfrequency-select-{index}']"))
                        )
                        self.gui_log(f"✅ Found resubmission type dropdown by data-testid with index {index}", level="DEBUG")
                        break
                    except TimeoutException:
                        continue
            
            # Strategy 4: Find select element containing "Amended Claim" option
            if not dropdown:
                try:
                    self.gui_log("Trying to find resubmission type dropdown by option content...", level="DEBUG")
                    dropdown = self.wait.until(
                        EC.presence_of_element_located((By.XPATH, "//select[.//option[@value='2' and text()='Amended Claim']]"))
                    )
                    self.gui_log("✅ Found resubmission type dropdown by option content", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Resubmission type dropdown not found by option content...", level="DEBUG")
            
            if dropdown:
                from selenium.webdriver.support.ui import Select
                
                # Scroll into view
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", dropdown)
                time.sleep(0.5)
                
                # Create Select object
                select = Select(dropdown)
                
                # Select "Amended Claim" by value (value="2")
                try:
                    select.select_by_value("2")
                    self.gui_log("✅ Selected 'Amended Claim' from dropdown", level="INFO")
                    self.update_status("Amended Claim selected", "#28a745")
                    time.sleep(0.5)  # Wait for selection to register
                    return True
                except Exception as select_error:
                    # Try selecting by visible text
                    try:
                        select.select_by_visible_text("Amended Claim")
                        self.gui_log("✅ Selected 'Amended Claim' from dropdown (by text)", level="INFO")
                        self.update_status("Amended Claim selected", "#28a745")
                        time.sleep(0.5)
                        return True
                    except Exception as text_error:
                        self.log_error(f"Failed to select 'Amended Claim': {select_error}", exception=text_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find resubmission type dropdown using any method", include_traceback=False)
                self.update_status("Error: Could not find resubmission type dropdown", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error selecting 'Amended Claim' from dropdown", exception=e, include_traceback=True)
            self.update_status(f"Amended Claim selection error: {str(e)}", "#dc3545")
            return False
    
    def _enter_payer_claim_control(self):
        """Enter the payer claim control number in the input field
        
        Element: <input data-testid="pendingclaim-payerclaimcontrol-input-0" id="resubmission-payer-claim-control-number-text-input-id-0" type="text" maxlength="50">
        
        Uses the payer claim control number that was extracted earlier (stored in self.current_payer_claim_control)
        
        Returns:
            bool: True if payer claim control number entered successfully, False otherwise
        """
        try:
            self.gui_log("Step 28: Entering payer claim control number...", level="INFO")
            self.update_status("Entering payer claim control number...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Get the payer claim control number that was extracted earlier
            payer_claim_control = getattr(self, 'current_payer_claim_control', None)
            if not payer_claim_control:
                self.log_error("No payer claim control number available - was not extracted earlier", include_traceback=False)
                self.update_status("Error: No payer claim control number available", "#dc3545")
                return False
            
            self.gui_log(f"Entering payer claim control number: {payer_claim_control}", level="INFO")
            
            # Wait for input field to be available
            time.sleep(1)
            
            # Find the input field using multiple strategies
            input_field = None
            
            # Strategy 1: Find by data-testid
            try:
                self.gui_log("Looking for payer claim control input field by data-testid...", level="DEBUG")
                input_field = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[@data-testid='pendingclaim-payerclaimcontrol-input-0']"))
                )
                self.gui_log("✅ Found payer claim control input field by data-testid", level="DEBUG")
            except TimeoutException:
                self.gui_log("Payer claim control input field not found by data-testid, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by ID
            if not input_field:
                try:
                    self.gui_log("Trying to find payer claim control input field by ID...", level="DEBUG")
                    input_field = self.wait.until(
                        EC.presence_of_element_located((By.ID, "resubmission-payer-claim-control-number-text-input-id-0"))
                    )
                    self.gui_log("✅ Found payer claim control input field by ID", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Payer claim control input field not found by ID, trying other methods...", level="DEBUG")
            
            # Strategy 3: Find by data-testid pattern (try index 0, 1, 2, etc.)
            if not input_field:
                for index in range(5):  # Try first 5 input fields
                    try:
                        self.gui_log(f"Trying to find payer claim control input field by data-testid with index {index}...", level="DEBUG")
                        input_field = self.wait.until(
                            EC.presence_of_element_located((By.XPATH, f"//input[@data-testid='pendingclaim-payerclaimcontrol-input-{index}']"))
                        )
                        self.gui_log(f"✅ Found payer claim control input field by data-testid with index {index}", level="DEBUG")
                        break
                    except TimeoutException:
                        continue
            
            if input_field:
                # Scroll into view
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", input_field)
                time.sleep(0.5)
                
                # Clear and enter the payer claim control number
                try:
                    input_field.clear()
                    time.sleep(0.3)
                    input_field.send_keys(payer_claim_control)
                    self.gui_log(f"✅ Entered payer claim control number: {payer_claim_control}", level="INFO")
                    self.update_status(f"Payer claim control number entered", "#28a745")
                    
                    # TherapyNotes quirk: ensure input loses focus before clicking Submit
                    label_clicked = False
                    try:
                        from selenium.webdriver.common.action_chains import ActionChains
                        actions = ActionChains(self.driver)
                        actions.move_to_element(input_field).move_by_offset(0, -35).click().perform()
                        label_clicked = True
                        self.gui_log("🖱️ Clicked above payer claim control input to commit value before submission", level="DEBUG")
                    except Exception as offset_error:
                        self.gui_log(f"⚠️ Offset click above input failed: {offset_error}", level="DEBUG")
                        try:
                            # Try clicking on form container as fallback
                            container = self.wait.until(
                                EC.presence_of_element_located(
                                    (By.XPATH, "//div[contains(@class, 'responsive-form-grid')]")
                                )
                            )
                            actions = ActionChains(self.driver)
                            actions.move_to_element(container).move_by_offset(20, 20).click().perform()
                            label_clicked = True
                            self.gui_log("🖱️ Clicked form container near payer claim control input", level="DEBUG")
                        except Exception as container_error:
                            self.gui_log(f"⚠️ Form container click failed: {container_error}", level="DEBUG")
                    
                    if not label_clicked:
                        try:
                            input_field.send_keys(Keys.TAB)
                            self.gui_log("🖱️ Sent TAB to move focus away from payer claim control input", level="DEBUG")
                        except Exception as tab_error:
                            self.gui_log(f"⚠️ Could not send TAB key after payer claim control entry: {tab_error}", level="WARNING")
                    
                    time.sleep(0.5)  # Wait for input to register
                    return True
                except Exception as fill_error:
                    self.log_error(f"Failed to enter payer claim control number: {fill_error}", exception=fill_error, include_traceback=True)
                    return False
            else:
                self.log_error("Could not find payer claim control input field using any method", include_traceback=False)
                self.update_status("Error: Could not find payer claim control input field", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error entering payer claim control number", exception=e, include_traceback=True)
            self.update_status(f"Payer claim control number error: {str(e)}", "#dc3545")
            return False
    
    def _click_submit_claims_button(self):
        """Click the "Submit Claims" button to submit the refiled claim
        
        Element: <input name="BatchClaimCreator__SubmitClaims-Button" id="BatchClaimCreator__SubmitClaims-Button" type="button" value="Submit Claims" tabindex="0">
        
        Returns:
            bool: True if Submit Claims button clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 29: Looking for 'Submit Claims' button...", level="INFO")
            self.update_status("Submitting claims...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for button to be available
            time.sleep(1)
            
            # Find the Submit Claims button using multiple strategies
            submit_button = None
            
            # Strategy 1: Find by ID
            try:
                self.gui_log("Looking for 'Submit Claims' button by ID...", level="DEBUG")
                submit_button = self.wait.until(
                    EC.element_to_be_clickable((By.ID, "BatchClaimCreator__SubmitClaims-Button"))
                )
                self.gui_log("✅ Found 'Submit Claims' button by ID", level="DEBUG")
            except TimeoutException:
                self.gui_log("'Submit Claims' button not found by ID, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by name
            if not submit_button:
                try:
                    self.gui_log("Trying to find 'Submit Claims' button by name...", level="DEBUG")
                    submit_button = self.wait.until(
                        EC.element_to_be_clickable((By.NAME, "BatchClaimCreator__SubmitClaims-Button"))
                    )
                    self.gui_log("✅ Found 'Submit Claims' button by name", level="DEBUG")
                except TimeoutException:
                    self.gui_log("'Submit Claims' button not found by name, trying other methods...", level="DEBUG")
            
            # Strategy 3: Find by value="Submit Claims"
            if not submit_button:
                try:
                    self.gui_log("Trying to find 'Submit Claims' button by value...", level="DEBUG")
                    submit_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@type='button' and @value='Submit Claims']"))
                    )
                    self.gui_log("✅ Found 'Submit Claims' button by value", level="DEBUG")
                except TimeoutException:
                    self.gui_log("'Submit Claims' button not found by value...", level="DEBUG")
            
            # Strategy 4: Find by partial text match
            if not submit_button:
                try:
                    self.gui_log("Trying to find 'Submit Claims' button by partial text...", level="DEBUG")
                    submit_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//input[contains(@value, 'Submit Claims')]"))
                    )
                    self.gui_log("✅ Found 'Submit Claims' button by partial text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("'Submit Claims' button not found by partial text...", level="DEBUG")
            
            if submit_button:
                # Scroll into view
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", submit_button)
                time.sleep(0.5)
                
                # Click the button
                try:
                    submit_button.click()
                    self.gui_log("✅ 'Submit Claims' button clicked", level="INFO")
                    self.update_status("Claims submitted", "#28a745")
                    
                    # Wait for submission to complete
                    time.sleep(3)
                    
                    self.gui_log("✅ Successfully submitted claims", level="INFO")
                    return True
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", submit_button)
                        self.gui_log("✅ 'Submit Claims' button clicked (JavaScript)", level="INFO")
                        self.update_status("Claims submitted", "#28a745")
                        time.sleep(3)
                        self.gui_log("✅ Successfully submitted claims", level="INFO")
                        return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click 'Submit Claims' button: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find 'Submit Claims' button using any method", include_traceback=False)
                self.update_status("Error: Could not find Submit Claims button", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error clicking 'Submit Claims' button", exception=e, include_traceback=True)
            self.update_status(f"Submit Claims error: {str(e)}", "#dc3545")
            return False
    
    def _handle_claims_submitted_popup(self):
        """Close the 'Claims Submitted' confirmation popup (with refresh fallback)."""
        try:
            self.gui_log("Checking for 'Claims Submitted' confirmation popup...", level="INFO")
            
            popup_locator = (By.XPATH, "//div[contains(@class, 'Dialog') and .//*[contains(text(), 'Claims Submitted')]]")
            try:
                popup = WebDriverWait(self.driver, 3).until(EC.presence_of_element_located(popup_locator))
                self.gui_log("✅ 'Claims Submitted' popup detected", level="INFO")
            except TimeoutException:
                self.gui_log("No 'Claims Submitted' popup detected (nothing to close)", level="DEBUG")
                return True
            
            close_strategies = [
                ".//button[contains(@class, 'close')]",
                ".//button[@aria-label='Close']",
                ".//button[contains(translate(text(), 'close', 'CLOSE'), 'CLOSE')]",
                ".//button[contains(translate(text(), 'ok', 'OK'), 'OK')]",
                ".//span[contains(@class, 'close') or contains(text(), '×')]",
            ]
            
            for strategy in close_strategies:
                try:
                    close_element = popup.find_element(By.XPATH, strategy)
                    self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'auto', block: 'center'});", close_element)
                    time.sleep(0.2)
                    try:
                        close_element.click()
                    except Exception:
                        self.driver.execute_script("arguments[0].click();", close_element)
                    WebDriverWait(self.driver, 5).until(EC.invisibility_of_element_located(popup_locator))
                    self.gui_log(f"✅ Closed 'Claims Submitted' popup using selector: {strategy}", level="DEBUG")
                    return True
                except Exception as close_error:
                    self.gui_log(f"⚠️ Close attempt with selector {strategy} failed: {close_error}", level="DEBUG")
            
            # Try pressing ESC as another fallback
            try:
                active_element = self.driver.switch_to.active_element
                active_element.send_keys(Keys.ESCAPE)
                WebDriverWait(self.driver, 3).until(EC.invisibility_of_element_located(popup_locator))
                self.gui_log("✅ Closed 'Claims Submitted' popup via ESC key", level="INFO")
                return True
            except Exception as esc_error:
                self.gui_log(f"⚠️ ESC key did not dismiss popup: {esc_error}", level="DEBUG")
            
            # Final fallback: refresh page to clear popup
            self.gui_log("⚠️ Unable to close popup directly - refreshing page as fallback", level="WARNING")
            self.driver.refresh()
            WebDriverWait(self.driver, 10).until(lambda drv: drv.execute_script("return document.readyState") == "complete")
            time.sleep(1)
            self.gui_log("✅ Page refreshed after submission popup to clear overlay", level="INFO")
            return True
        
        except Exception as e:
            self.log_error("Error handling 'Claims Submitted' popup", exception=e, include_traceback=True)
            return False
    
    def _search_for_payer_claim_control_by_elements(self):
        """Search for Payer Claim Control # using element-based approach with data-testid
        
        Label element: <td data-testid="eraviewer-payerclaimcontrolnumber-label">Payer Claim Control #:</td>
        Value element: <td data-testid="payerclaimcontrolnumber-container-7">01   082925 09574 00005</td>
        
        Returns:
            str: The Payer Claim Control number if found, None otherwise
        """
        try:
            # Strategy 1: Find label element first, then find corresponding value element
            try:
                self.gui_log("Looking for Payer Claim Control label element...", level="DEBUG")
                label_element = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, "//td[@data-testid='eraviewer-payerclaimcontrolnumber-label']"))
                )
                self.gui_log("✅ Found Payer Claim Control label element", level="DEBUG")
                
                # Find the corresponding value element
                # The value element is in the same row or nearby
                # Try to find it by data-testid pattern
                try:
                    # Find value element by pattern (payerclaimcontrolnumber-container-[number])
                    value_elements = self.driver.find_elements(By.XPATH, "//td[starts-with(@data-testid, 'payerclaimcontrolnumber-container-')]")
                    
                    if value_elements:
                        # Get the first matching value element
                        value_element = value_elements[0]
                        payer_claim_control = value_element.text.strip()
                        
                        # Clean up whitespace (handle multiple spaces)
                        payer_claim_control = ' '.join(payer_claim_control.split())
                        
                        if payer_claim_control:
                            self.gui_log(f"✅ Extracted Payer Claim Control # from element: {payer_claim_control}", level="INFO")
                            return payer_claim_control
                        else:
                            self.gui_log("Value element found but text is empty", level="DEBUG")
                except Exception as find_error:
                    self.gui_log(f"Could not find value element: {find_error}", level="DEBUG")
                
                # Strategy 2: Find value element in the same row as label
                try:
                    # Find parent row (tr) of label element
                    parent_row = label_element.find_element(By.XPATH, "./..")  # Parent is tr
                    
                    # Find the value td in the same row
                    value_td = parent_row.find_element(By.XPATH, ".//td[@data-testid[starts-with(., 'payerclaimcontrolnumber-container-')]]")
                    payer_claim_control = value_td.text.strip()
                    payer_claim_control = ' '.join(payer_claim_control.split())
                    
                    if payer_claim_control:
                        self.gui_log(f"✅ Extracted Payer Claim Control # from same row: {payer_claim_control}", level="INFO")
                        return payer_claim_control
                except Exception as row_error:
                    self.gui_log(f"Could not find value in same row: {row_error}", level="DEBUG")
                
                # Strategy 3: Find value element near label (next sibling or following td)
                try:
                    value_td = label_element.find_element(By.XPATH, "./following-sibling::td[@data-testid[starts-with(., 'payerclaimcontrolnumber-container-')]]")
                    payer_claim_control = value_td.text.strip()
                    payer_claim_control = ' '.join(payer_claim_control.split())
                    
                    if payer_claim_control:
                        self.gui_log(f"✅ Extracted Payer Claim Control # from following sibling: {payer_claim_control}", level="INFO")
                        return payer_claim_control
                except Exception as sibling_error:
                    self.gui_log(f"Could not find value in following sibling: {sibling_error}", level="DEBUG")
                    
            except TimeoutException:
                self.gui_log("Label element not found by data-testid", level="DEBUG")
            
            # Strategy 4: Find value element directly by pattern (without finding label first)
            try:
                self.gui_log("Trying to find value element directly by pattern...", level="DEBUG")
                value_elements = self.driver.find_elements(By.XPATH, "//td[starts-with(@data-testid, 'payerclaimcontrolnumber-container-')]")
                
                if value_elements:
                    # Get the first matching value element
                    value_element = value_elements[0]
                    payer_claim_control = value_element.text.strip()
                    payer_claim_control = ' '.join(payer_claim_control.split())
                    
                    if payer_claim_control:
                        self.gui_log(f"✅ Extracted Payer Claim Control # directly from element: {payer_claim_control}", level="INFO")
                        return payer_claim_control
            except Exception as direct_error:
                self.gui_log(f"Could not find value element directly: {direct_error}", level="DEBUG")
            
            return None
            
        except Exception as e:
            self.log_error("Error searching for Payer Claim Control # by elements", exception=e, include_traceback=False)
            return None
    
    def _search_for_payer_claim_control_with_scrolling(self):
        """Search for Payer Claim Control # with scrolling as fallback
        
        Returns:
            str: The Payer Claim Control number if found, None otherwise
        """
        try:
            # Find the popup/modal container to scroll within
            popup_container = None
            try:
                popup_containers = [
                    "//div[contains(@class, 'modal')]",
                    "//div[contains(@class, 'popup')]",
                    "//div[contains(@class, 'dialog')]",
                    "//div[@role='dialog']",
                    "//div[contains(@id, 'era') or contains(@id, 'claim')]",
                    "//body"
                ]
                
                for selector in popup_containers:
                    try:
                        containers = self.driver.find_elements(By.XPATH, selector)
                        if containers:
                            for container in containers:
                                if container.is_displayed():
                                    popup_container = container
                                    break
                            if popup_container:
                                break
                    except:
                        continue
            except:
                popup_container = None
            
            # Scroll search strategy
            max_scroll_attempts = 10
            scroll_distance = 300
            
            for scroll_attempt in range(max_scroll_attempts):
                self.gui_log(f"Scroll attempt {scroll_attempt + 1}/{max_scroll_attempts}...", level="DEBUG")
                
                # Try element-based search after each scroll
                payer_claim_control = self._search_for_payer_claim_control_by_elements()
                
                if payer_claim_control:
                    return payer_claim_control
                
                # Scroll down
                try:
                    if popup_container:
                        self.driver.execute_script("arguments[0].scrollTop += arguments[1];", popup_container, scroll_distance)
                    else:
                        self.driver.execute_script(f"window.scrollBy(0, {scroll_distance});")
                    time.sleep(0.5)
                except:
                    try:
                        self.driver.execute_script(f"window.scrollBy(0, {scroll_distance});")
                        time.sleep(0.5)
                    except:
                        pass
                
                # Check if reached bottom
                try:
                    current_scroll = self.driver.execute_script("return window.pageYOffset;")
                    total_height = self.driver.execute_script("return document.body.scrollHeight;")
                    viewport_height = self.driver.execute_script("return window.innerHeight;")
                    
                    if current_scroll + viewport_height >= total_height - 10:
                        break
                except:
                    pass
            
            return None
            
        except Exception as e:
            self.log_error("Error in scroll search for Payer Claim Control #", exception=e, include_traceback=False)
            return None
    
    def _search_for_payer_claim_control_text(self):
        """Search for "Payer Claim Control #:" text in the current page view and extract the number
        
        Returns:
            str: The Payer Claim Control number if found, None otherwise
        """
        try:
            import re
            
            # Get page text
            try:
                page_text = self.driver.find_element(By.TAG_NAME, "body").text
            except:
                page_text = self.driver.page_source
            
            # Search for "Payer Claim Control #:" pattern
            # Pattern variations:
            # - "Payer Claim Control #: 01 091525 09917 00020"
            # - "Payer Claim Control #:01 091525 09917 00020"
            # - "Payer Claim Control #:	01 091525 09917 00020" (with tab)
            
            patterns = [
                r'Payer Claim Control\s*#\s*:\s*([0-9]{2}\s+[0-9]{6}\s+[0-9]{5}\s+[0-9]{5})',  # Specific format: "01 082925 09574 00005" (preferred)
                r'Payer Claim Control\s*#\s*:\s*([\d\s]{10,30}?)(?=\s*(?:Member ID|Patient|Name|Insured|Rendering|\n|\r|$))',  # Matches digits/spaces until stop words (10-30 chars)
                r'Payer Claim Control\s*#\s*:\s*([\d\s]+?)(?=\s*Member\s+ID|\s*Patient|\s*Name|\s*Insured|\s*Rendering|\n|\r|$)',  # Matches until stop words
            ]
            
            for pattern in patterns:
                match = re.search(pattern, page_text, re.IGNORECASE | re.MULTILINE)
                if match:
                    payer_claim_control = match.group(1).strip()
                    self.gui_log(f"Found Payer Claim Control # pattern: '{payer_claim_control}'", level="DEBUG")
                    
                    # Clean up the number (remove extra whitespace, but preserve single spaces between groups)
                    payer_claim_control = ' '.join(payer_claim_control.split())
                    
                    # Validate it looks like a claim control number (has digits and reasonable length)
                    if payer_claim_control and len(payer_claim_control) <= 30 and re.search(r'\d', payer_claim_control):
                        self.gui_log(f"✅ Extracted Payer Claim Control #: {payer_claim_control}", level="INFO")
                        return payer_claim_control
            
            # Also try searching by XPath for elements containing this text
            try:
                elements = self.driver.find_elements(By.XPATH, "//*[contains(text(), 'Payer Claim Control')]")
                for element in elements:
                    element_text = element.text
                    # Look for the pattern in element text
                    for pattern in patterns:
                        match = re.search(pattern, element_text, re.IGNORECASE | re.MULTILINE)
                        if match:
                            payer_claim_control = match.group(1).strip()
                            # Clean up the number (remove extra whitespace)
                            payer_claim_control = ' '.join(payer_claim_control.split())
                            # Validate it looks like a claim control number
                            if payer_claim_control and len(payer_claim_control) <= 30 and re.search(r'\d', payer_claim_control):
                                self.gui_log(f"✅ Extracted Payer Claim Control # from element: {payer_claim_control}", level="INFO")
                                return payer_claim_control
            except:
                pass
            
            return None
            
        except Exception as e:
            self.log_error("Error searching for Payer Claim Control # text", exception=e, include_traceback=False)
            return None
    
    def _navigate_to_patients_from_payment_page(self):
        """Navigate to Patients page from the payment page
        
        Element: <span>Patients</span>
        
        Returns:
            bool: True if Patients button clicked successfully, False otherwise
        """
        try:
            self.gui_log("Navigating back to Patients page...", level="INFO")
            self.update_status("Returning to Patients page...", "#ff9500")
            
            # Wait for page to stabilize
            time.sleep(2)
            
            # Find Patients button using multiple strategies
            patients_button = None
            
            # Strategy 1: Find span element with text "Patients"
            try:
                self.gui_log("Looking for Patients button (span with text 'Patients')...", level="DEBUG")
                patients_button = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//span[text()='Patients']"))
                )
                self.gui_log("✅ Found Patients button by span text", level="DEBUG")
            except TimeoutException:
                self.gui_log("Patients button not found by span text, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by partial text match
            if not patients_button:
                try:
                    self.gui_log("Trying to find Patients button by partial text...", level="DEBUG")
                    patients_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Patients')]"))
                    )
                    self.gui_log("✅ Found Patients button by partial text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Patients button not found by partial text, trying other methods...", level="DEBUG")
            
            # Strategy 3: Find clickable parent element containing the span
            if not patients_button:
                try:
                    self.gui_log("Trying to find parent element containing 'Patients' span...", level="DEBUG")
                    patients_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[.//span[contains(text(), 'Patients')]] | //a[.//span[contains(text(), 'Patients')]] | //div[.//span[contains(text(), 'Patients')]]"))
                    )
                    self.gui_log("✅ Found Patients button by parent element", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Patients button not found by parent element...", level="DEBUG")
            
            # Strategy 4: Find by SVG with circle-user-v6 (as provided by user)
            if not patients_button:
                try:
                    self.gui_log("Trying to find Patients button by SVG (circle-user-v6)...", level="DEBUG")
                    # Look for div containing SVG with id="circle-user-v6"
                    patients_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//div[.//g[@id='circle-user-v6']] | //button[.//g[@id='circle-user-v6']] | //a[.//g[@id='circle-user-v6']]"))
                    )
                    self.gui_log("✅ Found Patients button by SVG (circle-user-v6)", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Patients button not found by SVG...", level="DEBUG")
            
            # Strategy 5: Try using the existing _navigate_to_patients method if available
            if not patients_button:
                try:
                    self.gui_log("Trying existing _navigate_to_patients method...", level="DEBUG")
                    if self._navigate_to_patients():
                        self.gui_log("✅ Navigated to Patients using existing method", level="DEBUG")
                        return True
                except Exception as e:
                    self.gui_log(f"Error using existing method: {e}", level="DEBUG")
            
            if patients_button:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", patients_button)
                time.sleep(0.5)
                
                # Click the Patients button
                try:
                    patients_button.click()
                    self.gui_log("✅ Patients button clicked", level="INFO")
                    self.update_status("Returned to Patients page", "#28a745")
                    
                    # Wait for navigation to complete
                    time.sleep(3)
                    
                    # Verify navigation
                    current_url = self.driver.current_url
                    self.gui_log(f"Current URL after clicking Patients: {current_url}", level="DEBUG")
                    
                    if "patients" in current_url.lower():
                        self.gui_log("✅ Successfully navigated to Patients page", level="INFO")
                        return True
                    else:
                        self.gui_log("⚠️ URL doesn't show patients page, but navigation may have succeeded", level="WARNING")
                        return True
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", patients_button)
                        self.gui_log("✅ Patients button clicked (JavaScript)", level="INFO")
                        time.sleep(3)
                        self.update_status("Returned to Patients page", "#28a745")
                        return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click Patients button: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find Patients button using any method", include_traceback=False)
                self.update_status("Error: Could not find Patients button", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error navigating to Patients page", exception=e, include_traceback=True)
            self.update_status(f"Patients navigation error: {str(e)}", "#dc3545")
            return False
    
    def _save_output_excel(self):
        """Save comprehensive output Excel file with all tracked client information
        
        Includes detailed tracking data:
        - Status (Refiled, Not Refiled - No View ERA, Not Refiled - Modifier Correct, etc.)
        - Original modifier (93 or 95)
        - New modifier (what it was changed to, or 'No Change')
        - Session medium (video or phone)
        - Expected modifier (based on session medium)
        - Payer Claim Control #
        - Modifier action (Updated, Already Correct, etc.)
        - Processing date/time
        - Error messages (if any)
        """
        try:
            # Use comprehensive tracking if available, otherwise fall back to legacy tracking
            if self.tracked_clients:
                all_clients = self.tracked_clients
                self.gui_log(f"Saving output Excel with {len(all_clients)} tracked client(s)...", level="INFO")
            elif self.skipped_clients or self.correct_modifier_clients:
                # Legacy mode - convert to comprehensive format
                all_clients = []
                for client in self.skipped_clients:
                    # Try to find insurance and claim status from excel_client_data
                    insurance = None
                    claim_status = None
                    client_name = client.get('client_name', 'Unknown')
                    dob = client.get('dob', '')
                    dos = client.get('date_of_service', '')
                    if self.excel_client_data:
                        for excel_client in self.excel_client_data:
                            if (excel_client.get('client_name') == client_name and 
                                excel_client.get('dob') == dob and 
                                excel_client.get('date_of_service') == dos):
                                insurance = excel_client.get('insurance')
                                claim_status = excel_client.get('claim_status')
                                break
                    
                    all_clients.append({
                        'client_name': client_name,
                        'dob': dob,
                        'date_of_service': dos,
                        'insurance': insurance or 'N/A',
                        'claim_status': claim_status or 'N/A',
                        'excel_row': 'N/A',
                        'status': 'Not Refiled - ' + client.get('reason', 'Unknown'),
                        'original_modifier': 'N/A',
                        'new_modifier': 'N/A',
                        'session_medium': 'N/A',
                        'expected_modifier': 'N/A',
                        'payer_claim_control': 'N/A',
                        'modifier_action': 'N/A',
                        'reason': client.get('reason', ''),
                        'processing_date': time.strftime("%Y-%m-%d %H:%M:%S"),
                        'error_message': ''
                    })
                for client in self.correct_modifier_clients:
                    # Try to find insurance and claim status from excel_client_data
                    insurance = None
                    claim_status = None
                    client_name = client.get('client_name', 'Unknown')
                    dob = client.get('dob', '')
                    dos = client.get('date_of_service', '')
                    if self.excel_client_data:
                        for excel_client in self.excel_client_data:
                            if (excel_client.get('client_name') == client_name and 
                                excel_client.get('dob') == dob and 
                                excel_client.get('date_of_service') == dos):
                                insurance = excel_client.get('insurance')
                                claim_status = excel_client.get('claim_status')
                                break
                    
                    all_clients.append({
                        'client_name': client_name,
                        'dob': dob,
                        'date_of_service': dos,
                        'insurance': insurance or 'N/A',
                        'claim_status': claim_status or 'N/A',
                        'excel_row': 'N/A',
                        'status': 'Not Refiled - Modifier Correct',
                        'original_modifier': 'N/A',
                        'new_modifier': 'N/A',
                        'session_medium': 'N/A',
                        'expected_modifier': 'N/A',
                        'payer_claim_control': 'N/A',
                        'modifier_action': 'Already Correct',
                        'reason': client.get('reason', 'Modifier was already correct, Not Resubmitted'),
                        'processing_date': time.strftime("%Y-%m-%d %H:%M:%S"),
                        'error_message': ''
                    })
                self.gui_log(f"Saving output Excel with {len(all_clients)} client(s) (legacy mode)...", level="INFO")
            else:
                self.gui_log("No tracked clients to save", level="INFO")
                return
            
            # Determine output file path
            if self.output_excel_path:
                # Use user-selected path
                output_file = Path(self.output_excel_path)
                # Ensure directory exists
                output_file.parent.mkdir(parents=True, exist_ok=True)
            else:
                # Generate default filename based on input filename and timestamp
                if self.selected_excel_path:
                    input_file = Path(self.selected_excel_path)
                    timestamp = time.strftime("%Y%m%d_%H%M%S")
                    output_file = input_file.parent / f"{input_file.stem}_output_{timestamp}.xlsx"
                else:
                    timestamp = time.strftime("%Y%m%d_%H%M%S")
                    output_file = Path(__file__).parent / f"tn_refiling_output_{timestamp}.xlsx"
            
            # Create DataFrame with comprehensive columns in a logical order
            df = pd.DataFrame(all_clients)
            
            # Define column order for better readability
            column_order = [
                'excel_row',
                'client_name',
                'dob',
                'date_of_service',
                'insurance',
                'claim_status',
                'status',
                'session_medium',
                'expected_modifier',
                'original_modifier',
                'new_modifier',
                'modifier_action',
                'payer_claim_control',
                'primary_policy',
                'primary_policy_member_id',
                'primary_policy_payment_match',
                'processing_date',
                'reason',
                'error_message'
            ]
            
            # Reorder columns (only include columns that exist in the DataFrame)
            existing_columns = [col for col in column_order if col in df.columns]
            remaining_columns = [col for col in df.columns if col not in existing_columns]
            df = df[existing_columns + remaining_columns]
            
            # Save to Excel with formatting
            try:
                # Use openpyxl for better formatting options
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Save initial DataFrame
                df.to_excel(output_file, index=False, engine='openpyxl')
                
                # Open workbook for formatting
                wb = load_workbook(output_file)
                ws = wb.active
                
                # Format header row
                header_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save formatted workbook
                wb.save(output_file)
                wb.close()
                
            except ImportError:
                # Fallback if openpyxl not available for formatting
                df.to_excel(output_file, index=False, engine='openpyxl')
                self.gui_log("Excel saved (formatting skipped - openpyxl not fully available)", level="DEBUG")
            
            self.gui_log(f"✅ Output Excel saved: {output_file}", level="INFO")
            self.update_status(f"Output saved: {output_file.name}", "#28a745")
            
            # Show summary
            status_counts = df['status'].value_counts() if 'status' in df.columns else {}
            summary = "Output Summary:\n"
            for status, count in status_counts.items():
                summary += f"  {status}: {count}\n"
            self.gui_log(summary, level="INFO")
            
            return True
            
        except Exception as e:
            self.log_error("Error saving output Excel", exception=e, include_traceback=True)
            return False
    
    def _click_progress_note_button(self):
        """Click the Progress Note, Consultation Note, or Intake Note button within the Notes tab popup
        
        Element: <span>Progress Note</span>, <span>Consultation Note</span>, or <span>Intake Note</span>
        After clicking, navigates to a new page
        The bot will try Progress Note first, then Consultation Note, then Intake Note if earlier options are not found.
        
        Returns:
            bool: True if Progress Note, Consultation Note, or Intake Note button clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 7: Looking for Progress/Consultation/Intake Note button...", level="INFO")
            self.update_status("Looking for note button...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for Notes tab content to fully load
            time.sleep(1.5)
            
            note_button = None
            note_type = None  # Track which type of note was found
            
            # Rapid lookup for any of the note types
            try:
                note_button = WebDriverWait(self.driver, 6).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//span[text()='Progress Note' or text()='Consultation Note' or text()='Intake Note']")
                    )
                )
                button_text = note_button.text.strip()
                if "Progress Note" in button_text:
                    note_type = "Progress Note"
                elif "Consultation Note" in button_text:
                    note_type = "Consultation Note"
                elif "Intake Note" in button_text:
                    note_type = "Intake Note"
                else:
                    note_type = "Note"
                self.gui_log(f"✅ Found {note_type} button quickly", level="DEBUG")
            except TimeoutException:
                # Fallback to previous thorough search when combined locator fails
                self.gui_log("Quick note button search failed - falling back to full search", level="DEBUG")
            
            # Fall back to detailed strategies only if needed
            if not note_button:
                # First, try to find Progress Note
                try:
                    self.gui_log("Looking for Progress Note button (span with text 'Progress Note')...", level="DEBUG")
                    note_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//span[text()='Progress Note']"))
                    )
                    note_type = "Progress Note"
                    self.gui_log("✅ Found Progress Note button by span text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Progress Note button not found by span text, trying Consultation Note...", level="DEBUG")
            
            # First, try to find Progress Note
            # Strategy 1: Find span element with text "Progress Note"
            try:
                self.gui_log("Looking for Progress Note button (span with text 'Progress Note')...", level="DEBUG")
                note_button = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//span[text()='Progress Note']"))
                )
                note_type = "Progress Note"
                self.gui_log("✅ Found Progress Note button by span text", level="DEBUG")
            except TimeoutException:
                self.gui_log("Progress Note button not found by span text, trying Consultation Note...", level="DEBUG")
            
            # Strategy 2: Find by partial text match for Progress Note
            if not note_button:
                try:
                    self.gui_log("Trying to find Progress Note button by partial text...", level="DEBUG")
                    note_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Progress Note')]"))
                    )
                    note_type = "Progress Note"
                    self.gui_log("✅ Found Progress Note button by partial text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Progress Note button not found by partial text, trying Consultation Note...", level="DEBUG")
            
            # If Progress Note not found, try Consultation Note
            if not note_button:
                try:
                    self.gui_log("Looking for Consultation Note button (span with text 'Consultation Note')...", level="DEBUG")
                    note_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//span[text()='Consultation Note']"))
                    )
                    note_type = "Consultation Note"
                    self.gui_log("✅ Found Consultation Note button by span text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Consultation Note button not found by span text, trying alternative methods...", level="DEBUG")
            
            # Strategy 3: Find Consultation Note by partial text
            if not note_button:
                try:
                    self.gui_log("Trying to find Consultation Note button by partial text...", level="DEBUG")
                    note_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Consultation Note')]"))
                    )
                    note_type = "Consultation Note"
                    self.gui_log("✅ Found Consultation Note button by partial text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Consultation Note button not found by partial text, trying Intake Note...", level="DEBUG")

            # Strategy 3: Find Intake Note by exact text
            if not note_button:
                try:
                    self.gui_log("Looking for Intake Note button (span with text 'Intake Note')...", level="DEBUG")
                    note_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//span[text()='Intake Note']"))
                    )
                    note_type = "Intake Note"
                    self.gui_log("✅ Found Intake Note button by span text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Intake Note button not found by span text, trying partial text...", level="DEBUG")

            # Strategy 4: Find Intake Note by partial text
            if not note_button:
                try:
                    self.gui_log("Trying to find Intake Note button by partial text...", level="DEBUG")
                    note_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Intake Note')]"))
                    )
                    note_type = "Intake Note"
                    self.gui_log("✅ Found Intake Note button by partial text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Intake Note button not found by partial text, trying other methods...", level="DEBUG")
            
            # Strategy 5: Find clickable parent element containing any note type
            if not note_button:
                try:
                    self.gui_log("Trying to find parent element containing note span (Progress/Consultation/Intake)...", level="DEBUG")
                    # Look for button, link, or div containing either span
                    note_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[.//span[contains(text(), 'Progress Note') or contains(text(), 'Consultation Note') or contains(text(), 'Intake Note')]] | //a[.//span[contains(text(), 'Progress Note') or contains(text(), 'Consultation Note') or contains(text(), 'Intake Note')]] | //div[.//span[contains(text(), 'Progress Note') or contains(text(), 'Consultation Note') or contains(text(), 'Intake Note')]]"))
                    )
                    # Determine which type was found
                    button_text = note_button.text
                    if 'Progress Note' in button_text:
                        note_type = "Progress Note"
                    elif 'Consultation Note' in button_text:
                        note_type = "Consultation Note"
                    elif 'Intake Note' in button_text:
                        note_type = "Intake Note"
                    self.gui_log(f"✅ Found {note_type} button by parent element", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Note button not found by parent element...", level="DEBUG")
            
            # Strategy 6: Find by case-insensitive text match for all types
            if not note_button:
                try:
                    self.gui_log("Trying to find note button with case-insensitive match...", level="DEBUG")
                    note_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//span[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'progress note') or contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'consultation note') or contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'intake note')]"))
                    )
                    # Determine which type was found
                    button_text = note_button.text
                    if 'progress note' in button_text.lower():
                        note_type = "Progress Note"
                    elif 'consultation note' in button_text.lower():
                        note_type = "Consultation Note"
                    elif 'intake note' in button_text.lower():
                        note_type = "Intake Note"
                    self.gui_log(f"✅ Found {note_type} button by case-insensitive text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Note button not found by case-insensitive text...", level="DEBUG")
            
            # Strategy 7: Find span and check if parent is clickable (for all types)
            if not note_button:
                try:
                    self.gui_log("Trying to find span and click parent/ancestor...", level="DEBUG")
                    spans = self.driver.find_elements(By.XPATH, "//span[contains(text(), 'Progress Note') or contains(text(), 'Consultation Note') or contains(text(), 'Intake Note')]")
                    if spans:
                        # Try clicking the span itself or its parent
                        for span in spans:
                            try:
                                span_text = span.text
                                # Check if span is clickable
                                if span.is_displayed() and span.is_enabled():
                                    note_button = span
                                    if 'Progress Note' in span_text:
                                        note_type = "Progress Note"
                                    elif 'Consultation Note' in span_text:
                                        note_type = "Consultation Note"
                                    elif 'Intake Note' in span_text:
                                        note_type = "Intake Note"
                                    self.gui_log(f"✅ Found {note_type} button by span element", level="DEBUG")
                                    break
                                # Otherwise try parent
                                parent = span.find_element(By.XPATH, "..")
                                if parent.is_displayed() and parent.is_enabled():
                                    note_button = parent
                                    if 'Progress Note' in span_text:
                                        note_type = "Progress Note"
                                    elif 'Consultation Note' in span_text:
                                        note_type = "Consultation Note"
                                    elif 'Intake Note' in span_text:
                                        note_type = "Intake Note"
                                    self.gui_log(f"✅ Found {note_type} button by parent of span", level="DEBUG")
                                    break
                            except:
                                continue
                except Exception as e:
                    self.gui_log(f"Error finding span and parent: {e}", level="DEBUG")
            
            if note_button:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", note_button)
                time.sleep(0.5)
                
                # Click the note button (Progress/Consultation/Intake)
                try:
                    note_button.click()
                    self.gui_log(f"✅ {note_type} button clicked", level="INFO")
                    self.update_status(f"{note_type} clicked - Navigating to new page...", "#ff9500")
                    
                    # Wait for navigation to new page
                    time.sleep(3)
                    
                    # Verify navigation by checking URL change
                    current_url = self.driver.current_url
                    self.gui_log(f"Current URL after clicking {note_type}: {current_url}", level="DEBUG")
                    
                    self.gui_log(f"✅ Successfully clicked {note_type} button and navigated to new page", level="INFO")
                    self.update_status(f"On {note_type} page", "#28a745")
                    
                    # Analyze the note to determine session medium (phone vs video)
                    # All note types (Progress/Consultation/Intake) contain session medium information
                    session_medium = self._analyze_session_medium()
                    if session_medium:
                        self.gui_log(f"📊 Session medium detected: {session_medium.upper()}", level="INFO")
                        self.current_session_medium = session_medium  # Store for later use
                        
                        # Check if original modifier (extracted from popup earlier) matches expected modifier
                        # If it matches, we can skip the entire History/Payment/ERA flow
                        original_mod = getattr(self, 'current_original_modifier', None)
                        if original_mod and original_mod != 'N/A':
                            expected_modifier = "95" if session_medium == "video" else "93"
                            if original_mod == expected_modifier:
                                self.gui_log(f"✅ Modifier already correct! Original: '{original_mod}', Expected: '{expected_modifier}' ({session_medium}) - Skipping refile process", level="INFO")
                                self.update_status(f"Modifier correct: {original_mod} ({session_medium}) - Skipping", "#28a745")
                                
                                # Store tracking data
                                self.current_new_modifier = 'No Change'
                                self.current_modifier_action = 'Already Correct'
                                self.current_expected_modifier = expected_modifier
                                
                                # Track this client (modifier already correct)
                                self._add_tracked_client(
                                    status='Not Refiled - Modifier Correct',
                                    original_modifier=original_mod,
                                    new_modifier='No Change',
                                    modifier_action='Already Correct',
                                    reason=f'Modifier was already correct ({original_mod} matches expected {expected_modifier} for {session_medium} session) - No refile needed',
                                    error_message=''
                                )
                                
                                # Navigate to Patients page from this Progress Note page (Patients button is available here)
                                if self._navigate_to_patients():
                                    self.gui_log("✅ Navigated to Patients - Continuing with next client", level="INFO")
                                    return "modifier_correct"  # Signal to skip rest of flow
                                else:
                                    self.gui_log("⚠️ Could not navigate to Patients, but modifier was correct", level="WARNING")
                                    return "modifier_correct"
                        
                        # Navigate back to Billing tab (modifier was incorrect)
                        if self._navigate_back_to_billing_tab():
                            return True
                        else:
                            self.gui_log("⚠️ Could not navigate back to Billing tab, but session medium was analyzed", level="WARNING")
                            # Still return True since analysis was completed
                            return True
                    else:
                        return self._handle_missing_session_medium(note_type)
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", note_button)
                        self.gui_log(f"✅ {note_type} button clicked (JavaScript)", level="INFO")
                        self.update_status(f"{note_type} clicked - Navigating to new page...", "#ff9500")
                        time.sleep(3)
                        current_url = self.driver.current_url
                        self.gui_log(f"Current URL after clicking {note_type} (JavaScript): {current_url}", level="DEBUG")
                        self.update_status(f"On {note_type} page", "#28a745")
                        
                        # Analyze the note to determine session medium (phone vs video)
                        # All note types (Progress/Consultation/Intake) contain session medium information
                        session_medium = self._analyze_session_medium()
                        if session_medium:
                            self.gui_log(f"📊 Session medium detected: {session_medium.upper()}", level="INFO")
                            self.current_session_medium = session_medium  # Store for later use
                            
                            # Navigate back to Billing tab
                            if self._navigate_back_to_billing_tab():
                                return True
                            else:
                                self.gui_log("⚠️ Could not navigate back to Billing tab, but session medium was analyzed", level="WARNING")
                                # Still return True since analysis was completed
                                return True
                        else:
                            return self._handle_missing_session_medium(note_type)
                    except Exception as js_error:
                        self.log_error(f"Failed to click {note_type} button: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find Progress Note, Consultation Note, or Intake Note button using any method", include_traceback=False)
                self.update_status("Error: Could not find note button", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error clicking note button (Progress/Consultation/Intake Note)", exception=e, include_traceback=True)
            self.update_status(f"Note button error: {str(e)}", "#dc3545")
            return False
    
    def _handle_missing_session_medium(self, note_type):
        """Record and handle scenarios where the session medium cannot be determined."""
        note_label = note_type or "Note"
        self.gui_log(f"⚠️ No session medium keywords found in {note_label} - Skipping client", level="WARNING")
        self.update_status("Session medium not found - Skipping client...", "#dc3545")
        
        # Ensure state reflects that modifier check was not performed
        self.current_session_medium = None
        self.current_modifier_action = 'Not Checked'
        
        client_name = getattr(self, 'current_client_name', 'Unknown')
        dob = getattr(self, 'current_client_dob', '')
        dos = getattr(self, 'current_date_of_service', '')
        
        # Record in skipped clients list
        try:
            skipped_entry = {
                'client_name': client_name,
                'dob': dob,
                'date_of_service': dos,
                'reason': f'Session medium keywords not found in {note_label}'
            }
            skipped_list = getattr(self, 'skipped_clients', None)
            if skipped_list is None:
                self.skipped_clients = []
                skipped_list = self.skipped_clients
            skipped_list.append(skipped_entry)
            self.gui_log(f"📝 Client '{client_name}' added to skipped list (reason: Session medium not found)", level="INFO")
        except Exception as skip_error:
            self.gui_log(f"⚠️ Could not add client to skipped list: {skip_error}", level="DEBUG")
        
        # Track detailed output
        original_mod = getattr(self, 'current_original_modifier', None) or 'N/A'
        try:
            self._add_tracked_client(
                status='Not Refiled - Session Medium Missing',
                original_modifier=original_mod,
                new_modifier='N/A',
                modifier_action='Not Checked',
                reason=f'Session medium keywords not found in {note_label} - Unable to determine expected modifier',
                error_message=f'Session medium detection failed in {note_label}'
            )
        except Exception as track_error:
            self.gui_log(f"⚠️ Could not track client for missing session medium: {track_error}", level="DEBUG")
        
        # Attempt to navigate back to Patients page to continue processing
        navigated_to_patients = False
        try:
            navigated_to_patients = self._navigate_to_patients()
        except Exception as nav_error:
            self.gui_log(f"⚠️ Error while navigating to Patients after missing session medium: {nav_error}", level="DEBUG")
        
        if navigated_to_patients:
            self.gui_log("✅ Navigated to Patients - Continuing with next client", level="INFO")
        else:
            self.gui_log("⚠️ Could not navigate to Patients after missing session medium", level="WARNING")
        
        return "session_medium_missing"
    
    def _analyze_session_medium(self):
        """Analyze the Progress/Consultation/Intake Note page to determine session medium (phone vs video)
        
        Extracts all text from the page and searches for keywords indicating:
        - Phone: phone, telephone, telephonic, phonic, audio, cell, cellphone
        - Video: video, zoom, videoconference, etc.
        
        Works with Progress Note, Consultation Note, and Intake Note pages.
        
        Returns:
            str: "phone", "video", or None if unable to determine
        """
        try:
            self.gui_log("Analyzing note page (Progress/Consultation/Intake) to determine session medium...", level="INFO")
            self.update_status("Analyzing session medium...", "#ff9500")
            
            # Wait for page to fully load
            time.sleep(2)
            
            # Extract all visible text from the page
            # Try to get text from the body element
            try:
                page_text = self.driver.find_element(By.TAG_NAME, "body").text
                self.gui_log(f"Extracted {len(page_text)} characters of text from note page", level="DEBUG")
            except Exception as e:
                self.gui_log(f"Could not extract text from body element: {e}", level="WARNING")
                # Try alternative method - get all text from main content area
                try:
                    # Look for common content containers
                    page_text = ""
                    content_selectors = [
                        "main", "article", "div[role='main']", 
                        ".content", "#content", ".main-content",
                        "div.progress-note", "div.note-content"
                    ]
                    for selector in content_selectors:
                        try:
                            elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                            if elements:
                                page_text = "\n".join([elem.text for elem in elements if elem.text])
                                if page_text:
                                    break
                        except:
                            continue
                    
                    if not page_text:
                        # Last resort: get all text from the page source
                        page_source = self.driver.page_source
                        # This won't be ideal but better than nothing
                        from html.parser import HTMLParser
                        # For now, just log that we're using page source
                        self.gui_log("Using page source as fallback for text extraction", level="WARNING")
                        page_text = self.driver.find_element(By.XPATH, "//body").text if self.driver.find_elements(By.XPATH, "//body") else ""
                        
                except Exception as e2:
                    self.log_error("Could not extract text from note page (Progress/Consultation/Intake Note)", exception=e2, include_traceback=False)
                    return None
            
            if not page_text or len(page_text.strip()) < 10:
                self.gui_log("⚠️ No meaningful text found on note page", level="WARNING")
                return None
            
            # Normalize text to lowercase for case-insensitive matching
            text_lower = page_text.lower()
            self.gui_log(f"Analyzing {len(text_lower)} characters of text for session medium keywords...", level="DEBUG")
            
            # Define keyword lists
            phone_keywords = [
                "phone", "telephone", "telephonic", "phonic", "audio", 
                "cell", "cellphone", "phone call", "telephone call",
                "telephonic session", "audio session", "via phone",
                "by phone", "on the phone", "over the phone"
            ]
            
            video_keywords = [
                "video", "zoom", "videoconference", "video conference",
                "video call", "video session", "via video", "by video",
                "over video", "zoom session", "televideo", "telehealth video"
            ]
            
            # Search for phone keywords
            phone_matches = []
            for keyword in phone_keywords:
                if keyword.lower() in text_lower:
                    phone_matches.append(keyword)
                    self.gui_log(f"  Found phone keyword: '{keyword}'", level="DEBUG")
            
            # Search for video keywords
            video_matches = []
            for keyword in video_keywords:
                if keyword.lower() in text_lower:
                    video_matches.append(keyword)
                    self.gui_log(f"  Found video keyword: '{keyword}'", level="DEBUG")
            
            # Determine categorization
            has_phone_keywords = len(phone_matches) > 0
            has_video_keywords = len(video_matches) > 0
            
            if has_video_keywords and not has_phone_keywords:
                self.gui_log(f"✅ Session categorized as VIDEO (found keywords: {', '.join(video_matches)})", level="INFO")
                return "video"
            elif has_phone_keywords and not has_video_keywords:
                self.gui_log(f"✅ Session categorized as PHONE (found keywords: {', '.join(phone_matches)})", level="INFO")
                return "phone"
            elif has_video_keywords and has_phone_keywords:
                # Both found - need to determine which is more prominent
                # Check for context around keywords
                self.gui_log(f"⚠️ Both phone and video keywords found - analyzing context...", level="WARNING")
                self.gui_log(f"  Phone keywords: {', '.join(phone_matches)}", level="DEBUG")
                self.gui_log(f"  Video keywords: {', '.join(video_matches)}", level="DEBUG")
                
                # Look for phrases that indicate the primary medium
                # Check for "met over video" or "met via video" which strongly indicates video
                video_phrases = ["met over video", "met via video", "session over video", "via zoom", "zoom session"]
                phone_phrases = ["met over phone", "met via phone", "session over phone", "telephone session", "phone session"]
                
                has_video_phrase = any(phrase in text_lower for phrase in video_phrases)
                has_phone_phrase = any(phrase in text_lower for phrase in phone_phrases)
                
                if has_video_phrase and not has_phone_phrase:
                    self.gui_log(f"✅ Session categorized as VIDEO (found video phrase)", level="INFO")
                    return "video"
                elif has_phone_phrase and not has_video_phrase:
                    self.gui_log(f"✅ Session categorized as PHONE (found phone phrase)", level="INFO")
                    return "phone"
                else:
                    # If still ambiguous, check count of keyword occurrences
                    video_count = sum(text_lower.count(kw) for kw in video_keywords)
                    phone_count = sum(text_lower.count(kw) for kw in phone_keywords)
                    
                    if video_count > phone_count:
                        self.gui_log(f"✅ Session categorized as VIDEO (video keywords appear more frequently)", level="INFO")
                        return "video"
                    elif phone_count > video_count:
                        self.gui_log(f"✅ Session categorized as PHONE (phone keywords appear more frequently)", level="INFO")
                        return "phone"
                    else:
                        # Default to video if equally ambiguous (video is more common in telehealth)
                        self.gui_log(f"⚠️ Ambiguous - defaulting to VIDEO (video/phone keywords equally present)", level="WARNING")
                        return "video"
            else:
                self.gui_log(f"⚠️ No session medium keywords found in Progress Note", level="WARNING")
                return None
                
        except Exception as e:
            self.log_error("Error analyzing session medium", exception=e, include_traceback=True)
            return None
    
    def _navigate_back_to_billing_tab(self):
        """Navigate back to the Billing tab from the Progress Note page
        
        Element: <a data-tab-id="Patient Billing" href="/app/patients/edit/CQl8z0eAq9nR4wlaj0aZig/#tab=Patient+Billing">Billing</a>
        Note: The patient ID in the href will vary, but the pattern is consistent
        
        Returns:
            bool: True if Billing tab clicked successfully, False otherwise
        """
        try:
            self.gui_log("Step 8: Navigating back to Billing tab...", level="INFO")
            self.update_status("Navigating back to Billing tab...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Wait for page to stabilize
            time.sleep(2)
            
            # Find Billing tab using multiple strategies
            billing_tab = None
            
            # Strategy 1: Find by data-tab-id="Patient Billing"
            try:
                self.gui_log("Looking for Billing tab (data-tab-id='Patient Billing')...", level="DEBUG")
                billing_tab = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-tab-id='Patient Billing']"))
                )
                self.gui_log("✅ Found Billing tab by data-tab-id", level="DEBUG")
            except TimeoutException:
                self.gui_log("Billing tab not found by data-tab-id, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by href pattern containing "tab=Patient+Billing"
            if not billing_tab:
                try:
                    self.gui_log("Trying to find Billing tab by href pattern...", level="DEBUG")
                    billing_tab = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[contains(@href, 'tab=Patient+Billing') or contains(@href, 'tab=Patient Billing')]"))
                    )
                    self.gui_log("✅ Found Billing tab by href pattern", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Billing tab not found by href pattern, trying link text...", level="DEBUG")
            
            # Strategy 3: Find by link text "Billing"
            if not billing_tab:
                try:
                    self.gui_log("Trying to find 'Billing' link by text...", level="DEBUG")
                    billing_tab = self.wait.until(
                        EC.element_to_be_clickable((By.LINK_TEXT, "Billing"))
                    )
                    self.gui_log("✅ Found Billing tab by link text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Billing link not found by text, trying partial text...", level="DEBUG")
            
            # Strategy 4: Find by partial link text
            if not billing_tab:
                try:
                    billing_tab = self.wait.until(
                        EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Billing"))
                    )
                    self.gui_log("✅ Found Billing tab by partial link text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Billing link not found by partial text...", level="DEBUG")
            
            # Strategy 5: Find by XPath with href containing "/app/patients" and "Billing" text
            if not billing_tab:
                try:
                    billing_tab = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[contains(@href, '/app/patients') and contains(text(), 'Billing')]"))
                    )
                    self.gui_log("✅ Found Billing tab by XPath with href and text", level="DEBUG")
                except TimeoutException:
                    self.gui_log("Billing tab not found by XPath...", level="DEBUG")
            
            if billing_tab:
                # Scroll into view if needed
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", billing_tab)
                time.sleep(0.5)
                
                # Click the Billing tab
                try:
                    billing_tab.click()
                    self.gui_log("✅ Billing tab clicked", level="INFO")
                    self.update_status("Billing tab clicked - Waiting for page to load...", "#ff9500")
                    
                    # Wait for navigation to complete
                    time.sleep(3)
                    
                    # Verify navigation by checking URL
                    current_url = self.driver.current_url
                    self.gui_log(f"Current URL after clicking Billing tab: {current_url}", level="DEBUG")
                    
                    # Check if URL contains billing indicator or tab is active
                    if "billing" in current_url.lower() or "#tab=" in current_url.lower() or "/app/patients" in current_url.lower():
                        self.gui_log("✅ Successfully navigated back to Billing tab", level="INFO")
                        self.update_status("Back on Billing tab", "#28a745")
                        
                        # Log the retained session medium category
                        if hasattr(self, 'current_session_medium') and self.current_session_medium:
                            self.gui_log(f"📊 Retained session medium category: {self.current_session_medium.upper()}", level="INFO")
                        
                        # Click "All Items" button again
                        if self._click_all_items_button():
                            # After clicking "All Items", click the same date of service again
                            if hasattr(self, 'current_date_of_service') and self.current_date_of_service:
                                if self._click_date_of_service(self.current_date_of_service):
                                    return True
                                else:
                                    self.gui_log("⚠️ Could not click date of service, but All Items was clicked", level="WARNING")
                                    return True
                            else:
                                self.gui_log("⚠️ No date of service available to click", level="WARNING")
                                return True
                        else:
                            self.gui_log("⚠️ Could not click All Items button, but Billing tab was navigated to", level="WARNING")
                            return True
                    else:
                        self.gui_log("⚠️ URL doesn't show billing tab, but tab may still be active", level="WARNING")
                        # Still return True as tab was clicked
                        if hasattr(self, 'current_session_medium') and self.current_session_medium:
                            self.gui_log(f"📊 Retained session medium category: {self.current_session_medium.upper()}", level="INFO")
                        
                        # Click "All Items" button again
                        if self._click_all_items_button():
                            # After clicking "All Items", click the same date of service again
                            if hasattr(self, 'current_date_of_service') and self.current_date_of_service:
                                if self._click_date_of_service(self.current_date_of_service):
                                    return True
                                else:
                                    self.gui_log("⚠️ Could not click date of service, but All Items was clicked", level="WARNING")
                                    return True
                            else:
                                self.gui_log("⚠️ No date of service available to click", level="WARNING")
                                return True
                        else:
                            self.gui_log("⚠️ Could not click All Items button, but Billing tab was navigated to", level="WARNING")
                            return True
                        
                except Exception as click_error:
                    # Try JavaScript click if regular click fails
                    try:
                        self.driver.execute_script("arguments[0].click();", billing_tab)
                        self.gui_log("✅ Billing tab clicked (JavaScript)", level="INFO")
                        time.sleep(3)
                        current_url = self.driver.current_url
                        self.gui_log(f"Current URL after clicking Billing tab (JavaScript): {current_url}", level="DEBUG")
                        self.update_status("Back on Billing tab", "#28a745")
                        
                        # Log the retained session medium category
                        if hasattr(self, 'current_session_medium') and self.current_session_medium:
                            self.gui_log(f"📊 Retained session medium category: {self.current_session_medium.upper()}", level="INFO")
                        
                        # Click "All Items" button again
                        if self._click_all_items_button():
                            # After clicking "All Items", click the same date of service again
                            if hasattr(self, 'current_date_of_service') and self.current_date_of_service:
                                if self._click_date_of_service(self.current_date_of_service):
                                    return True
                                else:
                                    self.gui_log("⚠️ Could not click date of service, but All Items was clicked", level="WARNING")
                                    return True
                            else:
                                self.gui_log("⚠️ No date of service available to click", level="WARNING")
                                return True
                        else:
                            self.gui_log("⚠️ Could not click All Items button, but Billing tab was navigated to", level="WARNING")
                            return True
                    except Exception as js_error:
                        self.log_error(f"Failed to click Billing tab: {click_error}", exception=js_error, include_traceback=True)
                        return False
            else:
                self.log_error("Could not find Billing tab using any method", include_traceback=False)
                self.update_status("Error: Could not find Billing tab", "#dc3545")
                return False
                
        except Exception as e:
            self.log_error("Error navigating back to Billing tab", exception=e, include_traceback=True)
            self.update_status(f"Billing tab navigation error: {str(e)}", "#dc3545")
            return False
    
    def login(self, username, password):
        """Perform login to Therapy Notes
        
        This method will be implemented step-by-step with the user
        to handle Therapy Notes login and navigation.
        
        URL: https://www.therapynotes.com/app/login/IntegritySWS/?r=%2fapp%2fpatients%2f
        """
        try:
            self.gui_log("🔑 Starting login process...", level="INFO")
            self.update_status("Logging in...", "#ff9500")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Log debug information
            self.gui_log(f"Therapy Notes URL: {self.therapy_notes_url}", level="DEBUG")
            self.gui_log(f"Username: {username}", level="DEBUG")
            self.gui_log(f"Password length: {len(password)} characters", level="DEBUG")
            
            # Initialize Chrome driver if not already initialized
            if not self.driver:
                if not self._init_chrome_driver():
                    return False
            
            # Check for stop request after driver initialization
            if self.check_stop_requested():
                return False
            
            # Navigate to Therapy Notes login page
            self.gui_log(f"Navigating to Therapy Notes login page...", level="INFO")
            self.update_status("Loading Therapy Notes...", "#ff9500")
            
            self.driver.get(self.therapy_notes_url)
            
            # Wait a moment for page to load
            time.sleep(2)
            
            # Check current page title/URL for debugging
            current_url = self.driver.current_url
            page_title = self.driver.title
            self.gui_log(f"Page loaded - URL: {current_url}", level="DEBUG")
            self.gui_log(f"Page title: {page_title}", level="DEBUG")
            
            self.gui_log("✅ Browser opened and navigated to Therapy Notes login page", level="INFO")
            self.gui_log("PROCEEDING TO FILL LOGIN FORM...", level="INFO")
            
            # Check for stop request
            if self.check_stop_requested():
                self.gui_log("Stop requested - aborting login", level="WARNING")
                return False
            
            # Find and fill username field
            self.gui_log("Step 1: Looking for username field...", level="INFO")
            self.update_status("Filling username...", "#ff9500")
            
            try:
                self.gui_log("Waiting for username field (ID: Login__UsernameField)...", level="DEBUG")
                # Wait for username field to be present and visible
                username_field = self.wait.until(
                    EC.presence_of_element_located((By.ID, "Login__UsernameField"))
                )
                
                self.gui_log("✅ Username field found", level="INFO")
                
                # Clear any existing text and fill username
                username_field.clear()
                time.sleep(0.3)  # Small delay after clear
                username_field.send_keys(username)
                
                self.gui_log(f"✅ Username entered: {username}", level="INFO")
                self.update_status("Username entered - Looking for password field...", "#ff9500")
                
            except TimeoutException:
                self.log_error("Username field not found - page may not have loaded", include_traceback=False)
                self.gui_log("Trying alternative method: searching by name attribute...", level="DEBUG")
                try:
                    username_field = self.driver.find_element(By.NAME, "username")
                    username_field.clear()
                    time.sleep(0.3)
                    username_field.send_keys(username)
                    self.gui_log(f"✅ Username entered using name attribute: {username}", level="INFO")
                except Exception as e:
                    self.log_error("Failed to find username field using multiple methods", exception=e, include_traceback=True)
                    self.update_status("Error: Could not find username field", "#dc3545")
                    return False
            except Exception as e:
                self.log_error("Error filling username field", exception=e, include_traceback=True)
                self.update_status(f"Error: {str(e)}", "#dc3545")
                return False
            
            self.gui_log("Username field filled successfully", level="DEBUG")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Find and fill password field
            self.gui_log("Looking for password field...", level="DEBUG")
            self.update_status("Filling password...", "#ff9500")
            
            try:
                # Wait for password field to be present and visible
                password_field = self.wait.until(
                    EC.presence_of_element_located((By.ID, "Login__Password"))
                )
                
                self.gui_log("✅ Password field found", level="DEBUG")
                
                # Clear any existing text and fill password
                password_field.clear()
                time.sleep(0.3)  # Small delay after clear
                password_field.send_keys(password)
                
                self.gui_log("✅ Password entered", level="INFO")
                self.update_status("Password entered - Looking for login button...", "#ff9500")
                
            except TimeoutException:
                self.log_error("Password field not found - page may not have loaded", include_traceback=False)
                self.gui_log("Trying alternative method: searching by name attribute...", level="DEBUG")
                try:
                    password_field = self.driver.find_element(By.NAME, "Password")
                    password_field.clear()
                    time.sleep(0.3)
                    password_field.send_keys(password)
                    self.gui_log("✅ Password entered using name attribute", level="INFO")
                except Exception as e:
                    self.log_error("Failed to find password field using multiple methods", exception=e, include_traceback=True)
                    self.update_status("Error: Could not find password field", "#dc3545")
                    return False
            except Exception as e:
                self.log_error("Error filling password field", exception=e, include_traceback=True)
                self.update_status(f"Error: {str(e)}", "#dc3545")
                return False
            
            self.gui_log("Password field filled successfully", level="DEBUG")
            
            # Check for stop request
            if self.check_stop_requested():
                return False
            
            # Find and click login button
            self.gui_log("Looking for login button...", level="DEBUG")
            self.update_status("Clicking login button...", "#ff9500")
            
            try:
                # Wait for login button to be present
                login_button = self.wait.until(
                    EC.presence_of_element_located((By.ID, "Login__LogInButton"))
                )
                
                self.gui_log("✅ Login button found", level="DEBUG")
                
                # Wait for button to be enabled (it may start disabled until form is valid)
                # Check if button is disabled
                is_disabled = login_button.get_attribute("aria-disabled")
                if is_disabled == "true":
                    self.gui_log("Login button is disabled, waiting for it to be enabled...", level="DEBUG")
                    # Wait for button to become enabled
                    self.wait.until(
                        lambda driver: login_button.get_attribute("aria-disabled") != "true"
                    )
                    self.gui_log("✅ Login button is now enabled", level="DEBUG")
                
                # Small delay before clicking
                time.sleep(0.5)
                
                # Click the login button
                login_button.click()
                
                self.gui_log("✅ Login button clicked", level="INFO")
                self.update_status("Login submitted - Waiting for page to load...", "#ff9500")
                
            except TimeoutException:
                self.log_error("Login button not found or did not become enabled", include_traceback=False)
                self.gui_log("Trying alternative method: searching by button text...", level="DEBUG")
                try:
                    # Try finding by button text
                    login_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Log In')]"))
                    )
                    login_button.click()
                    self.gui_log("✅ Login button clicked using text search", level="INFO")
                except Exception as e:
                    self.log_error("Failed to find/login button using multiple methods", exception=e, include_traceback=True)
                    self.update_status("Error: Could not find login button", "#dc3545")
                    return False
            except Exception as e:
                self.log_error("Error clicking login button", exception=e, include_traceback=True)
                self.update_status(f"Error: {str(e)}", "#dc3545")
                return False
            
            # Wait for navigation after login (check for URL change or new page elements)
            self.gui_log("Waiting for login to complete...", level="DEBUG")
            time.sleep(3)  # Wait for page to start loading
            
            # Check if we've navigated away from login page
            try:
                current_url_after = self.driver.current_url
                self.gui_log(f"Current URL after login: {current_url_after}", level="DEBUG")
                
                if "login" not in current_url_after.lower():
                    self.gui_log("✅ Successfully logged in - Navigated away from login page", level="INFO")
                    self.update_status("Login successful!", "#28a745")
                    self.is_logged_in = True
                    
                    # Navigate to Patients section
                    if self._navigate_to_patients():
                        self.gui_log("✅ Navigation to Patients section completed", level="INFO")
                    else:
                        self.gui_log("⚠️ Could not navigate to Patients section", level="WARNING")
                else:
                    self.gui_log("⚠️ Still on login page - may need to check for errors", level="WARNING")
                    self.update_status("Login may have failed - Check browser", "#ff9500")
                    
            except Exception as e:
                self.log_error("Error checking login status", exception=e, include_traceback=True)
                # Continue anyway - login may have succeeded
            
            self.gui_log("Login process completed. Ready for processing.")
            
            # Enable Start Processing button after successful login
            if self.root:
                self.root.after(0, self._enable_processing_button)
            
            return True
            
        except Exception as e:
            self.log_error("Error during login process", exception=e, include_traceback=True)
            self.update_status(f"Login error: {str(e)}", "#dc3545")
            # Try to close driver if there was an error
            try:
                if self.driver:
                    self._close_driver()
            except:
                pass
            return False


def main():
    """Main entry point"""
    print("\n=== TN Refiling Bot ===")
    print("Starting in 3 seconds...")
    print("Move mouse to corner to abort\n")
    
    time.sleep(3)
    
    # Create bot instance
    bot = TNRefilingBot()
    
    # Create the main window
    bot.create_main_window()
    
    # Start the main loop
    bot.root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nBot execution cancelled by user")
    except Exception as e:
        print(f"\n\nFatal error: {e}")
        logging.exception("Fatal error occurred")
        messagebox.showerror("Fatal Error", f"A fatal error occurred:\n{e}")



