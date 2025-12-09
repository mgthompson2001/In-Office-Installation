#!/usr/bin/env python3
"""
Medicare Modifier Comparison Bot

This bot compares two Excel files to identify modifier mismatches that need refiling.
It matches records based on Name, DOB, and Date of Service, then compares Session Medium
(File 1, Column F) with Modifier (File 2, Column G) to determine which claims need refiling.
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import logging
from pathlib import Path
import threading
import queue
import time
from datetime import datetime
from typing import Optional, List, Dict, Any
import re

# Optional dependencies
try:
    import pandas as pd
    EXCEL_AVAILABLE = True
except ImportError:
    pd = None
    EXCEL_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    openpyxl = None
    OPENPYXL_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    pdfplumber = None
    PDFPLUMBER_AVAILABLE = False

# Configure logging
SCRIPT_DIR = Path(__file__).parent
LOG_FILE_PATH = SCRIPT_DIR / "medicare_modifier_comparison_bot.log"

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE_PATH),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class MedicareModifierComparisonBot:
    """Main bot class for comparing Medicare modifier data"""
    
    def __init__(self):
        self.root: tk.Tk | None = None
        self.log_text: scrolledtext.ScrolledText | None = None
        self.gui_log_queue: queue.Queue[tuple[str, str]] = queue.Queue()
        self._gui_log_dispatcher_active = False
        
        # File paths
        self.log_file_path: Path | None = None  # File 1: Refile Medicare Log
        self.scrape_file_path: Path | None = None  # File 2: Scrape Excel
        self.output_path: Path | None = None
        
        # Column mappings for File 1 (Refile Medicare Log)
        self.file1_mappings: Dict[str, str] = {
            "Name": "",
            "First Name": "",
            "Last Name": "",
            "DOB": "",
            "Date of Service": "",
            "Session Medium": ""
        }
        
        # Column mappings for File 2 (Scrape Excel)
        self.file2_mappings: Dict[str, str] = {
            "Name": "",
            "First Name": "",
            "Last Name": "",
            "DOB": "",
            "Date of Service": "",
            "Modifier": ""
        }
        
        # Available columns (will be populated when files are loaded)
        self.file1_columns: List[str] = []
        self.file2_columns: List[str] = []
        
        # Data storage
        self.log_data: List[Dict[str, Any]] = []
        self.scrape_data: List[Dict[str, Any]] = []
        self.comparison_results: List[Dict[str, Any]] = []
        
        # PDF/Excel Synthesis tab variables
        self.pdf_file_path: Path | None = None  # PDF file for synthesis
        self.excel_synthesis_file_path: Path | None = None  # Excel file for synthesis
        self.synthesis_output_path: Path | None = None
        self.pdf_synthesis_data: List[Dict[str, Any]] = []
        self.excel_synthesis_data: List[Dict[str, Any]] = []
        self.synthesis_results: List[Dict[str, Any]] = []
        
        # Synthesis column mapping for Excel file
        self.synthesis_excel_columns: List[str] = []  # Available columns in synthesis Excel
        self.synthesis_session_medium_mapping: str = ""  # User-provided mapping for Session Medium column
        
        # Processing control
        self.processing = False
        self.stop_requested = False
        
    def create_gui(self):
        """Create the GUI interface"""
        self.root = tk.Tk()
        self.root.title("Medicare Modifier Comparison Bot - Version 3.1.0, Last Updated 12/04/2025")
        self.root.geometry("1000x900")
        
        # Header
        header = tk.Frame(self.root, bg="#660000")
        header.pack(fill="x")
        tk.Label(header, text="Medicare Modifier Comparison Bot", bg="#660000", fg="white",
                 font=("Segoe UI", 14, "bold"), pady=8).pack(side="left", padx=12)
        
        # Create notebook (tabbed interface)
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill="both", expand=True, padx=25, pady=25)
        
        # Tab 1: Excel Comparison
        tab1_frame = tk.Frame(notebook, bg="#f0f0f0")
        notebook.add(tab1_frame, text="Excel Comparison")
        
        main_frame = tk.Frame(tab1_frame, bg="#f0f0f0")
        main_frame.pack(fill="both", expand=True, padx=25, pady=25)
        
        # Instructions
        instructions = tk.Label(main_frame,
                               text="This bot compares two Excel files to identify modifier mismatches.\n"
                                    "It matches records by Name, DOB, and Date of Service, then compares\n"
                                    "Session Medium (File 1, Column F) with Modifier (File 2, Column G).",
                               font=("Segoe UI", 10),
                               justify="left",
                               wraplength=800,
                               bg="#f0f0f0")
        instructions.pack(pady=(0, 20))
        
        # File 1: Refile Medicare Log
        file1_frame = tk.LabelFrame(main_frame, text="File 1: Refile Medicare Log (Session Medium)", 
                                    font=("Segoe UI", 11, "bold"), bg="#f0f0f0")
        file1_frame.pack(fill="x", pady=(0, 15))
        
        file1_content = tk.Frame(file1_frame, bg="#f0f0f0")
        file1_content.pack(fill="x", padx=15, pady=10)
        
        tk.Label(file1_content, text="Excel File:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.log_file_entry = tk.Entry(file1_content, font=("Segoe UI", 9), width=60, state="readonly")
        self.log_file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        tk.Button(file1_content, text="Browse...", command=self._browse_log_file,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
        
        # File 2: Scrape Excel
        file2_frame = tk.LabelFrame(main_frame, text="File 2: Scrape Excel (Modifiers)", 
                                    font=("Segoe UI", 11, "bold"), bg="#f0f0f0")
        file2_frame.pack(fill="x", pady=(0, 15))
        
        file2_content = tk.Frame(file2_frame, bg="#f0f0f0")
        file2_content.pack(fill="x", padx=15, pady=10)
        
        tk.Label(file2_content, text="Excel File:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.scrape_file_entry = tk.Entry(file2_content, font=("Segoe UI", 9), width=60, state="readonly")
        self.scrape_file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        tk.Button(file2_content, text="Browse...", command=self._browse_scrape_file,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
        
        # Column Mapping Section
        mapping_frame = tk.LabelFrame(main_frame, text="Column Mapping (Optional - Leave empty for auto-detect)", 
                                      font=("Segoe UI", 11, "bold"), bg="#f0f0f0")
        mapping_frame.pack(fill="x", pady=(0, 15))
        
        mapping_content = tk.Frame(mapping_frame, bg="#f0f0f0")
        mapping_content.pack(fill="both", expand=True, padx=15, pady=10)
        
        instruction_text = ("Map columns for matching. Type column names (e.g., 'Excel_First_Name') or column letters (e.g., 'B', 'C', 'AB').\n"
                           "Leave empty to use automatic detection. Click 'Show Columns' after selecting files to see available columns.")
        tk.Label(mapping_content, text=instruction_text,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=900, justify="left").pack(anchor="w", pady=(0, 15))
        
        # Two-column layout for File 1 and File 2 mappings
        mapping_columns = tk.Frame(mapping_content, bg="#f0f0f0")
        mapping_columns.pack(fill="both", expand=True)
        
        # File 1 mappings (left column)
        file1_mapping_frame = tk.LabelFrame(mapping_columns, text="File 1: Refile Medicare Log", 
                                            font=("Segoe UI", 9, "bold"), bg="#f0f0f0")
        file1_mapping_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        file1_mapping_content = tk.Frame(file1_mapping_frame, bg="#f0f0f0")
        file1_mapping_content.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.file1_mapping_entries: Dict[str, tk.Entry] = {}
        file1_fields = [
            ("Name", "Full name or separate first/last"),
            ("First Name", "Optional - if name is split"),
            ("Last Name", "Optional - if name is split"),
            ("DOB", "Date of birth"),
            ("Date of Service", "Date of service"),
            ("Session Medium", "Session medium (Column F)")
        ]
        
        for field_name, description in file1_fields:
            row = tk.Frame(file1_mapping_content, bg="#f0f0f0")
            row.pack(fill="x", pady=(0, 8))
            
            tk.Label(row, text=f"{field_name}:", font=("Segoe UI", 9), bg="#f0f0f0", width=18, anchor="w").pack(side="left", padx=(0, 8))
            
            entry = tk.Entry(row, font=("Segoe UI", 9), width=20)
            entry.pack(side="left", padx=(0, 8))
            entry.bind("<KeyRelease>", lambda e, field=field_name: self._on_file1_mapping_changed(field))
            self.file1_mapping_entries[field_name] = entry
            
            tk.Label(row, text=description, font=("Segoe UI", 8, "italic"), 
                    bg="#f0f0f0", fg="#666666").pack(side="left")
        
        # File 2 mappings (right column)
        file2_mapping_frame = tk.LabelFrame(mapping_columns, text="File 2: Scrape Excel", 
                                            font=("Segoe UI", 9, "bold"), bg="#f0f0f0")
        file2_mapping_frame.pack(side="left", fill="both", expand=True, padx=(10, 0))
        
        file2_mapping_content = tk.Frame(file2_mapping_frame, bg="#f0f0f0")
        file2_mapping_content.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.file2_mapping_entries: Dict[str, tk.Entry] = {}
        file2_fields = [
            ("Name", "Full name or separate first/last"),
            ("First Name", "Optional - if name is split"),
            ("Last Name", "Optional - if name is split"),
            ("DOB", "Date of birth"),
            ("Date of Service", "Date of service"),
            ("Modifier", "Modifier (Column G)")
        ]
        
        for field_name, description in file2_fields:
            row = tk.Frame(file2_mapping_content, bg="#f0f0f0")
            row.pack(fill="x", pady=(0, 8))
            
            tk.Label(row, text=f"{field_name}:", font=("Segoe UI", 9), bg="#f0f0f0", width=18, anchor="w").pack(side="left", padx=(0, 8))
            
            entry = tk.Entry(row, font=("Segoe UI", 9), width=20)
            entry.pack(side="left", padx=(0, 8))
            entry.bind("<KeyRelease>", lambda e, field=field_name: self._on_file2_mapping_changed(field))
            self.file2_mapping_entries[field_name] = entry
            
            tk.Label(row, text=description, font=("Segoe UI", 8, "italic"), 
                    bg="#f0f0f0", fg="#666666").pack(side="left")
        
        # Show columns buttons
        button_row = tk.Frame(mapping_content, bg="#f0f0f0")
        button_row.pack(fill="x", pady=(10, 0))
        
        tk.Button(button_row, text="Show File 1 Columns", command=self._show_file1_columns,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 10))
        
        tk.Button(button_row, text="Show File 2 Columns", command=self._show_file2_columns,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
        
        # Output file
        output_frame = tk.LabelFrame(main_frame, text="Output Excel File", 
                                     font=("Segoe UI", 11, "bold"), bg="#f0f0f0")
        output_frame.pack(fill="x", pady=(0, 15))
        
        output_content = tk.Frame(output_frame, bg="#f0f0f0")
        output_content.pack(fill="x", padx=15, pady=10)
        
        tk.Label(output_content, text="Output File:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.output_file_entry = tk.Entry(output_content, font=("Segoe UI", 9), width=60, state="readonly")
        self.output_file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        tk.Button(output_content, text="Browse...", command=self._browse_output_file,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
        
        # Process button
        button_frame = tk.Frame(main_frame, bg="#f0f0f0")
        button_frame.pack(pady=20)
        
        self.process_button = tk.Button(button_frame, text="Compare and Generate Output",
                                        command=self._start_processing,
                                        bg="#006644", fg="white", font=("Segoe UI", 11, "bold"),
                                        padx=20, pady=10, cursor="hand2", relief="flat")
        self.process_button.pack()
        
        # Log area
        log_frame = tk.LabelFrame(main_frame, text="Log", font=("Segoe UI", 11, "bold"), bg="#f0f0f0")
        log_frame.pack(fill="both", expand=True, pady=(0, 0))
        
        log_content = tk.Frame(log_frame, bg="#f0f0f0")
        log_content.pack(fill="both", expand=True, padx=15, pady=10)
        
        self.log_text = scrolledtext.ScrolledText(log_content, height=15, font=("Courier New", 9),
                                                  bg="#ffffff", fg="#000000", wrap=tk.WORD)
        self.log_text.pack(fill="both", expand=True)
        
        # Tab 2: PDF/Excel Synthesis
        tab2_frame = tk.Frame(notebook, bg="#f0f0f0")
        notebook.add(tab2_frame, text="PDF/Excel Synthesis")
        
        self._create_synthesis_tab(tab2_frame)
        
        # Start GUI log dispatcher
        self._start_gui_log_dispatcher()
        
        # Initial log message
        self.gui_log("Ready. Please select both Excel files and an output location.")
        
    def _browse_log_file(self):
        """Browse for File 1: Refile Medicare Log"""
        filetypes = [("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")]
        path = filedialog.askopenfilename(title="Select Refile Medicare Log Excel File", filetypes=filetypes)
        if path:
            self.log_file_path = Path(path)
            self.log_file_entry.config(state="normal")
            self.log_file_entry.delete(0, tk.END)
            self.log_file_entry.insert(0, str(self.log_file_path))
            self.log_file_entry.config(state="readonly")
            self.gui_log(f"Selected File 1: {self.log_file_path.name}")
            # Load columns for mapping
            self._load_file1_columns()
    
    def _browse_scrape_file(self):
        """Browse for File 2: Scrape Excel"""
        filetypes = [("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")]
        path = filedialog.askopenfilename(title="Select Scrape Excel File", filetypes=filetypes)
        if path:
            self.scrape_file_path = Path(path)
            self.scrape_file_entry.config(state="normal")
            self.scrape_file_entry.delete(0, tk.END)
            self.scrape_file_entry.insert(0, str(self.scrape_file_path))
            self.scrape_file_entry.config(state="readonly")
            self.gui_log(f"Selected File 2: {self.scrape_file_path.name}")
            # Load columns for mapping
            self._load_file2_columns()
    
    def _browse_output_file(self):
        """Browse for output Excel file"""
        filetypes = [("Excel files", "*.xlsx"), ("All files", "*.*")]
        path = filedialog.asksaveasfilename(title="Save Output Excel File As", 
                                           defaultextension=".xlsx", filetypes=filetypes)
        if path:
            self.output_path = Path(path)
            self.output_file_entry.config(state="normal")
            self.output_file_entry.delete(0, tk.END)
            self.output_file_entry.insert(0, str(self.output_path))
            self.output_file_entry.config(state="readonly")
            self.gui_log(f"Output file: {self.output_path.name}")
    
    def _start_processing(self):
        """Start the comparison process in a separate thread"""
        if self.processing:
            messagebox.showwarning("Already Processing", "Comparison is already in progress.")
            return
        
        if not self.log_file_path or not self.log_file_path.exists():
            messagebox.showerror("Missing File", "Please select File 1 (Refile Medicare Log).")
            return
        
        if not self.scrape_file_path or not self.scrape_file_path.exists():
            messagebox.showerror("Missing File", "Please select File 2 (Scrape Excel).")
            return
        
        if not self.output_path:
            messagebox.showerror("Missing Output", "Please select an output file location.")
            return
        
        # Disable process button
        self.process_button.config(state="disabled", text="Processing...")
        
        # Start processing in thread
        thread = threading.Thread(target=self._process_comparison, daemon=True)
        thread.start()
    
    def _process_comparison(self):
        """Process the comparison between the two Excel files"""
        try:
            self.processing = True
            self.gui_log("\n" + "="*70)
            self.gui_log("Starting Medicare Modifier Comparison")
            self.gui_log("="*70)
            
            # Load File 1: Refile Medicare Log
            self.gui_log(f"\n📄 Loading File 1: {self.log_file_path.name}")
            self.log_data = self._load_excel_file(self.log_file_path)
            self.gui_log(f"   ✅ Loaded {len(self.log_data)} record(s) from File 1")
            
            # Load File 2: Scrape Excel
            self.gui_log(f"\n📄 Loading File 2: {self.scrape_file_path.name}")
            self.scrape_data = self._load_excel_file(self.scrape_file_path)
            self.gui_log(f"   ✅ Loaded {len(self.scrape_data)} record(s) from File 2")
            
            # Match records and compare
            self.gui_log(f"\n🔍 Matching records and comparing modifiers...")
            self._match_and_compare()
            
            # Generate output
            self.gui_log(f"\n📊 Generating output Excel...")
            self._generate_output()
            
            self.gui_log(f"\n✅ Comparison complete! Output saved to: {self.output_path}")
            self.root.after(0, lambda: messagebox.showinfo("Success", 
                f"Comparison complete!\n\n"
                f"Results saved to:\n{self.output_path}\n\n"
                f"Total records processed: {len(self.comparison_results)}"))
            
        except Exception as e:
            error_msg = f"Error during comparison: {str(e)}"
            self.gui_log(f"\n❌ {error_msg}", level="ERROR")
            logger.exception("Error in _process_comparison")
            self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
        finally:
            self.processing = False
            self.root.after(0, lambda: self.process_button.config(state="normal", text="Compare and Generate Output"))
    
    def _load_excel_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """Load Excel file and return list of dictionaries"""
        if not EXCEL_AVAILABLE:
            raise Exception("pandas is required. Install with: pip install pandas openpyxl")
        
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            # Convert DataFrame to list of dictionaries
            records = df.to_dict('records')
            return records
        except Exception as e:
            raise Exception(f"Failed to load Excel file {file_path.name}: {str(e)}")
    
    def _normalize_name(self, name: str) -> str:
        """Normalize name for matching (remove extra spaces, convert to uppercase)"""
        if not name or pd.isna(name):
            return ""
        return " ".join(str(name).strip().upper().split())
    
    def _normalize_dob(self, dob) -> str:
        """Normalize DOB for matching - handles MM/DD/YY, MM/DD/YYYY, and YYYY-MM-DD formats"""
        if not dob or pd.isna(dob):
            return ""
        
        # Convert to string and remove timestamp if present
        dob_str = str(dob).strip()
        if ' ' in dob_str:
            dob_str = dob_str.split(' ')[0]
        
        # Handle YYYY-MM-DD format (e.g., "1935-02-07")
        match = re.match(r'(\d{4})-(\d{2})-(\d{2})', dob_str)
        if match:
            return match.group(0)  # Return as-is: "1935-02-07"
        
        # Handle MM/DD/YYYY format (e.g., "2/7/1935" or "02/07/1935")
        match = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', dob_str)
        if match:
            month, day, year = match.groups()
            # Format as YYYY-MM-DD for consistent comparison
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        
        # Handle MM/DD/YY format (2-digit year, e.g., "8/21/54")
        match = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2})', dob_str)
        if match:
            month, day, year_2digit = match.groups()
            # Convert 2-digit year to 4-digit
            # Assume years < 50 are 20xx, years >= 50 are 19xx
            year_int = int(year_2digit)
            if year_int < 50:
                year = f"20{year_2digit}"
            else:
                year = f"19{year_2digit}"
            # Format as YYYY-MM-DD
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        
        # If no match, return cleaned string (remove non-digits except / and -)
        cleaned = re.sub(r'[^\d/-]', '', dob_str)
        return cleaned if cleaned else dob_str
    
    def _normalize_date(self, date) -> str:
        """Normalize date for matching - handles multiple date formats robustly"""
        if not date:
            return ""
        
        # Handle pandas/Excel date objects (datetime64, Timestamp)
        try:
            if hasattr(date, 'strftime'):
                # It's a datetime object
                return date.strftime('%Y-%m-%d')
            if hasattr(date, 'date'):
                # It has a date() method (like pandas Timestamp)
                return date.date().strftime('%Y-%m-%d')
        except (AttributeError, ValueError):
            pass
        
        # Handle NaN values
        try:
            if pd.notna is not None and not pd.notna(date):
                return ""
        except (TypeError, ValueError):
            pass
        
        # Convert to string and clean
        date_str = str(date).strip()
        if not date_str or date_str.lower() in ['nan', 'none', 'nat', '']:
            return ""
        
        # Remove time component if present (e.g., "2025-07-04 00:00:00")
        if ' ' in date_str:
            date_str = date_str.split(' ')[0]
        
        # Remove timestamp milliseconds if present
        if '.' in date_str and 'T' in date_str:
            date_str = date_str.split('.')[0]
        
        # Handle YYYY-MM-DD format (e.g., "2025-07-04" or "1935-02-07")
        match = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', date_str)
        if match:
            year, month, day = match.groups()
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        
        # Handle MM/DD/YYYY format (e.g., "7/4/2025" or "07/04/2025" or "7/04/2025")
        match = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', date_str)
        if match:
            month, day, year = match.groups()
            # Format as YYYY-MM-DD for consistent comparison
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        
        # Handle MM-DD-YYYY format (e.g., "7-4-2025" or "07-04-2025")
        match = re.match(r'(\d{1,2})-(\d{1,2})-(\d{4})', date_str)
        if match:
            month, day, year = match.groups()
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        
        # Handle MM.DD.YYYY format (e.g., "7.4.2025")
        match = re.match(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', date_str)
        if match:
            month, day, year = match.groups()
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        
        # Handle MM/DD/YY format (2-digit year, e.g., "7/4/25" or "8/21/54")
        match = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2})$', date_str)
        if match:
            month, day, year_2digit = match.groups()
            # Convert 2-digit year to 4-digit
            # Assume years < 50 are 20xx, years >= 50 are 19xx
            try:
                year_int = int(year_2digit)
                if year_int < 50:
                    year = f"20{year_2digit}"
                else:
                    year = f"19{year_2digit}"
                return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            except ValueError:
                pass
        
        # Handle YYYYMMDD format (e.g., "20250704")
        match = re.match(r'(\d{4})(\d{2})(\d{2})$', date_str)
        if match:
            year, month, day = match.groups()
            return f"{year}-{month}-{day}"
        
        # Try to extract date components from any remaining format
        # Look for 4-digit year
        year_match = re.search(r'(\d{4})', date_str)
        if year_match:
            year = year_match.group(1)
            # Look for month and day before/after year
            parts = re.findall(r'\d{1,2}', date_str)
            if len(parts) >= 3:
                # Try different orders
                for i, part in enumerate(parts):
                    if part == year[-2:]:  # Found 2-digit year
                        if i > 0 and i < len(parts) - 1:
                            month = parts[i-1]
                            day = parts[i+1]
                            try:
                                if 1 <= int(month) <= 12 and 1 <= int(day) <= 31:
                                    return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                            except ValueError:
                                pass
        
        # If no match, return cleaned string (remove non-digits except /, -, and .)
        cleaned = re.sub(r'[^\d/.-]', '', date_str)
        # Try one more time to parse cleaned string
        if cleaned and len(cleaned) >= 6:
            # Try common patterns on cleaned string
            for pattern in [
                r'(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})',  # MM/DD/YYYY
                r'(\d{4})[/.-](\d{1,2})[/.-](\d{1,2})',  # YYYY/MM/DD
            ]:
                match = re.match(pattern, cleaned)
                if match:
                    groups = match.groups()
                    if len(groups) == 3:
                        # Determine if first group is year (4 digits) or month (1-2 digits)
                        if len(groups[0]) == 4:
                            year, month, day = groups
                            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        else:
                            month, day, year = groups
                            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        
        return cleaned if cleaned else date_str
    
    def _dates_match(self, date1_normalized: str, date2_normalized: str) -> bool:
        """Check if two normalized dates match (allows for format variations).
        
        Args:
            date1_normalized: First normalized date (YYYY-MM-DD format)
            date2_normalized: Second normalized date (YYYY-MM-DD format)
        
        Returns:
            True if dates match, False otherwise
        """
        if not date1_normalized or not date2_normalized:
            return False
        
        # Exact match
        if date1_normalized == date2_normalized:
            return True
        
        # Try to normalize both again to catch any edge cases
        try:
            d1 = self._normalize_date(date1_normalized)
            d2 = self._normalize_date(date2_normalized)
            return d1 == d2 and d1 != "" and d2 != ""
        except Exception:
            return False
    
    def _extract_column_value(self, record: Dict, column_letter: str, default_col_name: str = None) -> str:
        """Extract value from record by column letter or column name"""
        # Try column letter first (convert to index)
        col_index = self._column_letter_to_index(column_letter)
        if col_index is not None:
            columns = list(record.keys())
            if 0 <= col_index < len(columns):
                value = record.get(columns[col_index], "")
                return str(value).strip() if value and not pd.isna(value) else ""
        
        # Try default column name
        if default_col_name:
            value = record.get(default_col_name, "")
            return str(value).strip() if value and not pd.isna(value) else ""
        
        return ""
    
    def _column_letter_to_index(self, column_letter: str) -> int | None:
        """Convert Excel column letter to 0-based index"""
        if not column_letter:
            return None
        
        column_letter = column_letter.strip().upper()
        if not column_letter.isalpha():
            return None
        
        result = 0
        for char in column_letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1
    
    def _find_column_by_name(self, record: Dict, search_terms: List[str]) -> str:
        """Find column value by searching for column names containing search terms"""
        for key in record.keys():
            key_lower = str(key).lower()
            for term in search_terms:
                if term.lower() in key_lower:
                    value = record.get(key, "")
                    if value and not pd.isna(value):
                        return str(value).strip()
        return ""
    
    def _resolve_column_reference(self, user_input: str, record: Dict) -> str:
        """Resolve user input (column name or letter) to actual column value"""
        if not user_input:
            return ""
        
        user_input = user_input.strip()
        
        # Try as column letter first
        col_index = self._column_letter_to_index(user_input)
        if col_index is not None:
            columns = list(record.keys())
            if 0 <= col_index < len(columns):
                value = record.get(columns[col_index], "")
                return str(value).strip() if value and not pd.isna(value) else ""
        
        # Try exact column name match
        if user_input in record:
            value = record.get(user_input, "")
            return str(value).strip() if value and not pd.isna(value) else ""
        
        # Try case-insensitive match
        for key in record.keys():
            if str(key).strip().lower() == user_input.lower():
                value = record.get(key, "")
                return str(value).strip() if value and not pd.isna(value) else ""
        
        return ""
    
    def _extract_name_from_record(self, record: Dict, is_file1: bool = True) -> str:
        """Extract full name from record using mappings or auto-detect"""
        mappings = self.file1_mappings if is_file1 else self.file2_mappings
        entries = self.file1_mapping_entries if is_file1 else self.file2_mapping_entries
        
        # Try mapped columns first
        if entries and "Name" in entries:
            name_mapping = entries["Name"].get().strip()
            if name_mapping:
                name = self._resolve_column_reference(name_mapping, record)
                if name:
                    return name
        
        # Try separate first/last name if mapped
        first_name = ""
        last_name = ""
        if entries and "First Name" in entries:
            first_mapping = entries["First Name"].get().strip()
            if first_mapping:
                first_name = self._resolve_column_reference(first_mapping, record)
        
        if entries and "Last Name" in entries:
            last_mapping = entries["Last Name"].get().strip()
            if last_mapping:
                last_name = self._resolve_column_reference(last_mapping, record)
        
        if first_name or last_name:
            return f"{first_name} {last_name}".strip()
        
        # Fallback to auto-detect
        full_name = self._find_column_by_name(record, ["name", "client name", "patient name", "full name"])
        if full_name:
            return full_name
        
        # Try separate first and last name
        first_name = self._find_column_by_name(record, ["first name", "fname", "firstname"])
        last_name = self._find_column_by_name(record, ["last name", "lname", "lastname"])
        
        if first_name or last_name:
            return f"{first_name} {last_name}".strip()
        
        return ""
    
    def _match_and_compare(self):
        """Match records between the two files and compare modifiers"""
        self.comparison_results = []
        match_count = 0
        no_match_count = 0
        
        # Expected modifier mapping based on session medium
        # Session Medium -> Expected Modifier
        session_medium_to_modifier = {
            "video": "95",
            "telehealth": "95",
            "telemedicine": "95",
            "phone": "93",
            "telephone": "93",
            "in-person": "",
            "in person": "",
            "office": ""
        }
        
        # Build lookup dictionary for File 2 to speed up matching
        # Key: (normalized_name, normalized_dob, normalized_dos)
        self.gui_log("   Building lookup index for File 2...")
        scrape_lookup: Dict[tuple, List[Dict]] = {}
        file2_samples = []  # Store samples for debugging
        
        for scrape_record in self.scrape_data:
            # Extract matching fields from File 2
            name2 = self._normalize_name(self._extract_name_from_record(scrape_record, is_file1=False))
            dob2 = ""
            if "DOB" in self.file2_mapping_entries:
                dob_mapping = self.file2_mapping_entries["DOB"].get().strip()
                if dob_mapping:
                    dob2 = self._normalize_dob(self._resolve_column_reference(dob_mapping, scrape_record))
            if not dob2:
                dob2 = self._normalize_dob(self._find_column_by_name(scrape_record, ["dob", "date of birth", "birthdate"]))
            
            dos2 = ""
            if "Date of Service" in self.file2_mapping_entries:
                dos_mapping = self.file2_mapping_entries["Date of Service"].get().strip()
                if dos_mapping:
                    dos2 = self._normalize_date(self._resolve_column_reference(dos_mapping, scrape_record))
            if not dos2:
                dos2 = self._normalize_date(self._find_column_by_name(scrape_record, ["date of service", "dos", "service date"]))
            
            # Store first few samples for debugging (before normalization)
            if len(file2_samples) < 5 and name2 and dob2 and dos2:
                # Get raw values before normalization for debugging
                name2_raw = self._extract_name_from_record(scrape_record, is_file1=False)
                dob2_raw = ""
                if "DOB" in self.file2_mapping_entries:
                    dob_mapping = self.file2_mapping_entries["DOB"].get().strip()
                    if dob_mapping:
                        dob2_raw = self._resolve_column_reference(dob_mapping, scrape_record)
                if not dob2_raw:
                    dob2_raw = self._find_column_by_name(scrape_record, ["dob", "date of birth", "birthdate"])
                
                dos2_raw = ""
                if "Date of Service" in self.file2_mapping_entries:
                    dos_mapping = self.file2_mapping_entries["Date of Service"].get().strip()
                    if dos_mapping:
                        dos2_raw = self._resolve_column_reference(dos_mapping, scrape_record)
                if not dos2_raw:
                    dos2_raw = self._find_column_by_name(scrape_record, ["date of service", "dos", "service date"])
                
                file2_samples.append({
                    "raw_name": name2_raw,
                    "raw_dob": dob2_raw,
                    "raw_dos": dos2_raw,
                    "normalized_name": name2,
                    "normalized_dob": dob2,
                    "normalized_dos": dos2
                })
            
            # Create lookup keys with multiple variations
            if name2 and dob2 and dos2:
                name_parts = name2.split()
                
                # Clean dates for flexible matching
                dob2_clean = re.sub(r'[^\d/]', '', dob2)
                dos2_clean = re.sub(r'[^\d/]', '', dos2)
                
                # Add multiple key variations for flexible matching
                keys_to_add = [
                    (name2, dob2, dos2),  # Exact match
                    (name2, dob2_clean, dos2_clean),  # Exact name, cleaned dates
                ]
                
                # Add first + last name variations
                if len(name_parts) >= 2:
                    first_last_name = f"{name_parts[0]} {name_parts[-1]}"
                    keys_to_add.extend([
                        (first_last_name, dob2, dos2),  # First+Last, original dates
                        (first_last_name, dob2_clean, dos2_clean),  # First+Last, cleaned dates
                    ])
                
                # Add all key variations to lookup
                for key in keys_to_add:
                    if key not in scrape_lookup:
                        scrape_lookup[key] = []
                    if scrape_record not in scrape_lookup[key]:
                        scrape_lookup[key].append(scrape_record)
        
        self.gui_log(f"   Built lookup index with {len(scrape_lookup)} unique key(s)")
        
        # Debug: Show sample File 2 values
        if file2_samples:
            self.gui_log(f"\n   📋 Sample File 2 values (raw → normalized):")
            for i, sample in enumerate(file2_samples[:5], 1):
                self.gui_log(f"      {i}. Name: '{sample['raw_name']}' → '{sample['normalized_name']}'")
                self.gui_log(f"         DOB:  '{sample['raw_dob']}' → '{sample['normalized_dob']}'")
                self.gui_log(f"         DOS:  '{sample['raw_dos']}' → '{sample['normalized_dos']}'")
        
        # Debug: Show sample lookup keys
        sample_keys = list(scrape_lookup.keys())[:5]
        self.gui_log(f"\n   🔑 Sample File 2 lookup keys (first 5):")
        for key in sample_keys:
            self.gui_log(f"      {key}")
        
        total_records = len(self.log_data)
        no_match_samples = []  # Store first few no-match samples for debugging
        
        for idx, log_record in enumerate(self.log_data):
            # Progress logging every 50 records
            if (idx + 1) % 50 == 0 or idx == 0:
                self.gui_log(f"   Processing record {idx + 1}/{total_records}...")
            # Extract matching fields from File 1 using mappings or auto-detect
            name1 = self._normalize_name(self._extract_name_from_record(log_record, is_file1=True))
            
            # Extract DOB using mapping or auto-detect
            dob1 = ""
            if "DOB" in self.file1_mapping_entries:
                dob_mapping = self.file1_mapping_entries["DOB"].get().strip()
                if dob_mapping:
                    dob1_raw = self._resolve_column_reference(dob_mapping, log_record)
                    dob1 = self._normalize_dob(dob1_raw)
            if not dob1:
                dob1_raw = self._find_column_by_name(log_record, ["dob", "date of birth", "birthdate"])
                dob1 = self._normalize_dob(dob1_raw)
            
            # Extract Date of Service using mapping or auto-detect
            dos1 = ""
            if "Date of Service" in self.file1_mapping_entries:
                dos_mapping = self.file1_mapping_entries["Date of Service"].get().strip()
                if dos_mapping:
                    dos1_raw = self._resolve_column_reference(dos_mapping, log_record)
                    dos1 = self._normalize_date(dos1_raw)
            if not dos1:
                dos1_raw = self._find_column_by_name(log_record, ["date of service", "dos", "service date"])
                dos1 = self._normalize_date(dos1_raw)
            
            # Debug: Show first few File 1 values (raw and normalized)
            if idx < 5:
                name1_raw = self._extract_name_from_record(log_record, is_file1=True)
                dob1_raw = ""
                if "DOB" in self.file1_mapping_entries:
                    dob_mapping = self.file1_mapping_entries["DOB"].get().strip()
                    if dob_mapping:
                        dob1_raw = self._resolve_column_reference(dob_mapping, log_record)
                if not dob1_raw:
                    dob1_raw = self._find_column_by_name(log_record, ["dob", "date of birth", "birthdate"])
                
                dos1_raw = ""
                if "Date of Service" in self.file1_mapping_entries:
                    dos_mapping = self.file1_mapping_entries["Date of Service"].get().strip()
                    if dos_mapping:
                        dos1_raw = self._resolve_column_reference(dos_mapping, log_record)
                if not dos1_raw:
                    dos1_raw = self._find_column_by_name(log_record, ["date of service", "dos", "service date"])
                
                if idx == 0:
                    self.gui_log(f"\n   📋 Sample File 1 values (raw → normalized):")
                self.gui_log(f"      {idx + 1}. Name: '{name1_raw}' → '{name1}'")
                self.gui_log(f"         DOB:  '{dob1_raw}' → '{dob1}'")
                self.gui_log(f"         DOS:  '{dos1_raw}' → '{dos1}'")
            
            # Extract Session Medium using mapping or Column F
            session_medium = ""
            if "Session Medium" in self.file1_mapping_entries:
                session_mapping = self.file1_mapping_entries["Session Medium"].get().strip()
                if session_mapping:
                    session_medium = self._resolve_column_reference(session_mapping, log_record)
            if not session_medium:
                session_medium = self._extract_column_value(log_record, "F", "Session Medium")
            if not session_medium:
                session_medium = self._find_column_by_name(log_record, ["session medium", "medium", "session type"])
            session_medium_lower = session_medium.lower().strip() if session_medium else ""
            
            # Determine expected modifier based on session medium
            expected_modifier = ""
            for key, mod in session_medium_to_modifier.items():
                if key in session_medium_lower:
                    expected_modifier = mod
                    break
            
            # Try to find matching record in File 2 using lookup dictionary
            matched_scrape_record = None
            
            if name1 and dob1 and dos1:
                name1_parts = name1.split() if name1 else []
                
                # Build list of all keys to try (in order of preference)
                keys_to_try = []
                
                # 1. Exact match
                keys_to_try.append((name1, dob1, dos1))
                
                # 2. First + last name only (ignore middle names)
                if len(name1_parts) >= 2:
                    first_last_name = f"{name1_parts[0]} {name1_parts[-1]}"
                    keys_to_try.append((first_last_name, dob1, dos1))
                
                # 3. Try with cleaned dates (normalize should already do this, but try anyway)
                dob1_clean = re.sub(r'[^\d/]', '', dob1)
                dos1_clean = re.sub(r'[^\d/]', '', dos1)
                if dob1_clean and dos1_clean and (dob1_clean != dob1 or dos1_clean != dos1):
                    keys_to_try.append((name1, dob1_clean, dos1_clean))
                    if len(name1_parts) >= 2:
                        first_last_name = f"{name1_parts[0]} {name1_parts[-1]}"
                        keys_to_try.append((first_last_name, dob1_clean, dos1_clean))
                
                # 4. Try just DOB + DOS match (name might be slightly different)
                # This is a fallback - only if exact matches fail
                if len(name1_parts) >= 2:
                    # Try matching just first name + DOB + DOS
                    keys_to_try.append((name1_parts[0], dob1, dos1))
                
                # Try all keys
                for lookup_key in keys_to_try:
                    if lookup_key in scrape_lookup:
                        matched_scrape_record = scrape_lookup[lookup_key][0]  # Take first match
                        break
                
                # Debug: Log first few no-match attempts
                if not matched_scrape_record and len(no_match_samples) < 5:
                    dob1_clean = re.sub(r'[^\d/]', '', dob1) if dob1 else ""
                    dos1_clean = re.sub(r'[^\d/]', '', dos1) if dos1 else ""
                    name1_parts = name1.split() if name1 else []
                    first_last_name = f"{name1_parts[0]} {name1_parts[-1]}" if len(name1_parts) >= 2 else ""
                    
                    attempted_keys = keys_to_try.copy()  # Use the same keys we tried
                    
                    no_match_samples.append({
                        "name": name1,
                        "dob": dob1,
                        "dos": dos1,
                        "keys_attempted": attempted_keys
                    })
            elif len(no_match_samples) < 5:
                # Missing required fields
                no_match_samples.append({
                    "name": name1 or "(empty)",
                    "dob": dob1 or "(empty)",
                    "dos": dos1 or "(empty)",
                    "keys_attempted": []
                })
            
            # Extract modifier from File 2 using mapping or Column G
            actual_modifier = ""
            if matched_scrape_record:
                if "Modifier" in self.file2_mapping_entries:
                    mod_mapping = self.file2_mapping_entries["Modifier"].get().strip()
                    if mod_mapping:
                        actual_modifier = self._resolve_column_reference(mod_mapping, matched_scrape_record)
                if not actual_modifier:
                    actual_modifier = self._extract_column_value(matched_scrape_record, "G", "Modifier")
                match_count += 1
            else:
                no_match_count += 1
            
            # Determine if refiling is needed
            needs_refile = False
            refile_reason = ""
            
            if matched_scrape_record:
                if expected_modifier and actual_modifier:
                    if expected_modifier != actual_modifier:
                        needs_refile = True
                        refile_reason = f"Modifier mismatch: Expected {expected_modifier} (for {session_medium}), but found {actual_modifier}"
                elif expected_modifier and not actual_modifier:
                    needs_refile = True
                    refile_reason = f"Missing modifier: Expected {expected_modifier} for {session_medium} session"
                elif not expected_modifier and actual_modifier:
                    needs_refile = True
                    refile_reason = f"Unexpected modifier: Found {actual_modifier} but session medium is {session_medium} (in-person should have no modifier)"
            else:
                refile_reason = "No matching record found in File 2"
            
            # Create comparison result
            # Get original values for display (before normalization)
            orig_name1 = self._extract_name_from_record(log_record, is_file1=True)
            orig_dob1 = ""
            if "DOB" in self.file1_mapping_entries:
                dob_mapping = self.file1_mapping_entries["DOB"].get().strip()
                if dob_mapping:
                    orig_dob1 = self._resolve_column_reference(dob_mapping, log_record)
            if not orig_dob1:
                orig_dob1 = self._find_column_by_name(log_record, ["dob", "date of birth"])
            
            orig_dos1 = ""
            if "Date of Service" in self.file1_mapping_entries:
                dos_mapping = self.file1_mapping_entries["Date of Service"].get().strip()
                if dos_mapping:
                    orig_dos1 = self._resolve_column_reference(dos_mapping, log_record)
            if not orig_dos1:
                orig_dos1 = self._find_column_by_name(log_record, ["date of service", "dos"])
            
            result = {
                "Name": name1 or orig_name1,
                "DOB": dob1 or orig_dob1,
                "Date_of_Service": dos1 or orig_dos1,
                "Session_Medium": session_medium,
                "Expected_Modifier": expected_modifier,
                "Actual_Modifier": actual_modifier,
                "Needs_Refile": "Yes" if needs_refile else "No",
                "Refile_Reason": refile_reason,
                "Match_Status": "Matched" if matched_scrape_record else "No Match"
            }
            
            # Add all other fields from both files
            for key, value in log_record.items():
                if key not in result:
                    result[f"File1_{key}"] = value
            
            if matched_scrape_record:
                for key, value in matched_scrape_record.items():
                    if key not in result:
                        result[f"File2_{key}"] = value
            
            self.comparison_results.append(result)
        
        self.gui_log(f"\n📊 Matching Summary:")
        self.gui_log(f"   ✅ Matched: {match_count} record(s)")
        self.gui_log(f"   ⚠️  No Match: {no_match_count} record(s)")
        self.gui_log(f"   📋 Total: {len(self.comparison_results)} record(s)")
        
        # Debug: Show why matches failed
        if match_count == 0 and no_match_samples:
            self.gui_log(f"\n🔍 Debug: Why matches failed (first {len(no_match_samples)} samples):")
            for i, sample in enumerate(no_match_samples[:5], 1):
                self.gui_log(f"\n   Sample {i} from File 1:")
                self.gui_log(f"      Name='{sample['name']}'")
                self.gui_log(f"      DOB='{sample['dob']}'")
                self.gui_log(f"      DOS='{sample['dos']}'")
                if sample['keys_attempted']:
                    self.gui_log(f"      Attempted lookup keys ({len(sample['keys_attempted'])}):")
                    for key_idx, key in enumerate(sample['keys_attempted'], 1):
                        found = "✅ FOUND" if key in scrape_lookup else "❌ NOT FOUND"
                        self.gui_log(f"         {key_idx}. {key} - {found}")
                    
                    # Check if any similar keys exist
                    if sample['keys_attempted']:
                        first_attempted_key = sample['keys_attempted'][0]
                        if first_attempted_key and len(first_attempted_key) == 3:
                            attempted_name, attempted_dob, attempted_dos = first_attempted_key
                            
                            # Find keys with same name but different dates
                            similar_name_keys = [k for k in scrape_lookup.keys() 
                                              if len(k) == 3 and k[0] == attempted_name][:5]
                            if similar_name_keys:
                                self.gui_log(f"      Found keys with same name (different dates): {similar_name_keys}")
                            
                            # Find keys with same DOB but different name
                            similar_dob_keys = [k for k in scrape_lookup.keys() 
                                              if len(k) == 3 and k[1] == attempted_dob][:5]
                            if similar_dob_keys:
                                self.gui_log(f"      Found keys with same DOB (different name): {similar_dob_keys}")
                            
                            # Find keys with same DOS but different name/DOB
                            similar_dos_keys = [k for k in scrape_lookup.keys() 
                                              if len(k) == 3 and k[2] == attempted_dos][:5]
                            if similar_dos_keys:
                                self.gui_log(f"      Found keys with same DOS (different name/DOB): {similar_dos_keys}")
                        
                        # Try fuzzy name matching
                        if attempted_name:
                            name_parts = attempted_name.split()
                            if len(name_parts) >= 2:
                                first_name = name_parts[0]
                                last_name = name_parts[-1]
                                # Find keys that start with same first name or end with same last name
                                fuzzy_keys = [k for k in scrape_lookup.keys() 
                                            if len(k) == 3 and (
                                                (isinstance(k[0], str) and k[0].startswith(first_name)) or
                                                (isinstance(k[0], str) and k[0].endswith(last_name))
                                            )][:5]
                                if fuzzy_keys:
                                    self.gui_log(f"      Found keys with similar name (first/last match): {fuzzy_keys}")
    
    def _generate_output(self):
        """Generate output Excel file with comparison results"""
        if not self.comparison_results:
            raise Exception("No comparison results to output")
        
        if not EXCEL_AVAILABLE:
            raise Exception("pandas is required. Install with: pip install pandas openpyxl")
        
        # Create DataFrame
        df = pd.DataFrame(self.comparison_results)
        
        # Reorder columns to put important ones first
        priority_columns = [
            "Name", "DOB", "Date_of_Service", "Session_Medium",
            "Expected_Modifier", "Actual_Modifier", "Needs_Refile",
            "Refile_Reason", "Match_Status"
        ]
        
        # Get remaining columns
        other_columns = [col for col in df.columns if col not in priority_columns]
        
        # Reorder
        ordered_columns = [col for col in priority_columns if col in df.columns] + sorted(other_columns)
        df = df[ordered_columns]
        
        # Save to Excel
        df.to_excel(self.output_path, index=False, engine='openpyxl')
        self.gui_log(f"   ✅ Generated output with {len(df)} row(s) and {len(df.columns)} column(s)")
    
    def gui_log(self, message: str, level: str = "INFO"):
        """Add message to GUI log queue"""
        timestamp = datetime.now().strftime("[%H:%M:%S]")
        log_entry = f"{timestamp} {message}"
        self.gui_log_queue.put((log_entry, level))
        logger.log(getattr(logging, level, logging.INFO), message)
    
    def _start_gui_log_dispatcher(self):
        """Start the GUI log message dispatcher"""
        if self._gui_log_dispatcher_active:
            return
        
        self._gui_log_dispatcher_active = True
        
        def dispatch_logs():
            while self._gui_log_dispatcher_active:
                try:
                    log_entry, level = self.gui_log_queue.get(timeout=0.1)
                    
                    # Determine color based on level
                    color_map = {
                        "INFO": "#000000",
                        "WARNING": "#ff6600",
                        "ERROR": "#cc0000",
                        "DEBUG": "#666666"
                    }
                    color = color_map.get(level, "#000000")
                    
                    self.root.after(0, lambda entry=log_entry, c=color: self._append_log(entry, c))
                except queue.Empty:
                    continue
                except Exception:
                    break
        
        thread = threading.Thread(target=dispatch_logs, daemon=True)
        thread.start()
    
    def _append_log(self, message: str, color: str = "#000000"):
        """Append message to log text widget"""
        if self.log_text:
            self.log_text.config(state="normal")
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
            self.log_text.config(state="disabled")
    
    def _load_file1_columns(self):
        """Load column names from File 1"""
        if not self.log_file_path or not self.log_file_path.exists():
            return
        
        try:
            df = pd.read_excel(self.log_file_path, engine='openpyxl', nrows=0)  # Read only headers
            self.file1_columns = list(df.columns)
            self.gui_log(f"Loaded {len(self.file1_columns)} columns from File 1")
        except Exception as e:
            self.gui_log(f"Warning: Could not load columns from File 1: {str(e)}", level="WARNING")
    
    def _load_file2_columns(self):
        """Load column names from File 2"""
        if not self.scrape_file_path or not self.scrape_file_path.exists():
            return
        
        try:
            df = pd.read_excel(self.scrape_file_path, engine='openpyxl', nrows=0)  # Read only headers
            self.file2_columns = list(df.columns)
            self.gui_log(f"Loaded {len(self.file2_columns)} columns from File 2")
        except Exception as e:
            self.gui_log(f"Warning: Could not load columns from File 2: {str(e)}", level="WARNING")
    
    def _show_file1_columns(self):
        """Show available columns from File 1"""
        if not self.file1_columns:
            if not self.log_file_path:
                messagebox.showinfo("No File", "Please select File 1 first.")
                return
            self._load_file1_columns()
        
        if not self.file1_columns:
            messagebox.showinfo("No Columns", "No columns found in File 1.")
            return
        
        def get_column_letter(index):
            """Convert 0-based index to Excel column letter"""
            result = ""
            index += 1
            while index > 0:
                index -= 1
                result = chr(65 + (index % 26)) + result
                index //= 26
            return result
        
        columns_text = "\n".join([f"{get_column_letter(i)}. {col}" for i, col in enumerate(self.file1_columns)])
        messagebox.showinfo(
            "File 1 Columns",
            f"Found {len(self.file1_columns)} column(s) in File 1:\n\n{columns_text}\n\n"
            "Type column names or letters (e.g., 'B', 'C', 'AB') in the mapping fields above."
        )
    
    def _show_file2_columns(self):
        """Show available columns from File 2"""
        if not self.file2_columns:
            if not self.scrape_file_path:
                messagebox.showinfo("No File", "Please select File 2 first.")
                return
            self._load_file2_columns()
        
        if not self.file2_columns:
            messagebox.showinfo("No Columns", "No columns found in File 2.")
            return
        
        def get_column_letter(index):
            """Convert 0-based index to Excel column letter"""
            result = ""
            index += 1
            while index > 0:
                index -= 1
                result = chr(65 + (index % 26)) + result
                index //= 26
            return result
        
        columns_text = "\n".join([f"{get_column_letter(i)}. {col}" for i, col in enumerate(self.file2_columns)])
        messagebox.showinfo(
            "File 2 Columns",
            f"Found {len(self.file2_columns)} column(s) in File 2:\n\n{columns_text}\n\n"
            "Type column names or letters (e.g., 'B', 'C', 'AB') in the mapping fields above."
        )
    
    def _on_file1_mapping_changed(self, field_name: str):
        """Handle File 1 column mapping change"""
        if field_name in self.file1_mapping_entries:
            entry = self.file1_mapping_entries[field_name]
            value = entry.get().strip()
            self.file1_mappings[field_name] = value
            if value:
                self.gui_log(f"File 1 mapping updated: {field_name} -> '{value}'", level="DEBUG")
    
    def _on_file2_mapping_changed(self, field_name: str):
        """Handle File 2 column mapping change"""
        if field_name in self.file2_mapping_entries:
            entry = self.file2_mapping_entries[field_name]
            value = entry.get().strip()
            self.file2_mappings[field_name] = value
            if value:
                self.gui_log(f"File 2 mapping updated: {field_name} -> '{value}'", level="DEBUG")
    
    def _create_synthesis_tab(self, parent: tk.Frame):
        """Create the PDF/Excel Synthesis tab"""
        main_frame = tk.Frame(parent, bg="#f0f0f0")
        main_frame.pack(fill="both", expand=True, padx=25, pady=25)
        
        # Instructions
        instructions = tk.Label(main_frame,
                               text="This feature matches records between a PDF and Excel file based on abbreviated name codes.\n"
                                    "Name code format: First 3 letters of last name + First 2 letters of first name + 3 numbers\n"
                                    "Example: 'Barbara Lukasiewicz' → 'LUKBA000' (from PDF Column 2)",
                               font=("Segoe UI", 10),
                               justify="left",
                               wraplength=800,
                               bg="#f0f0f0")
        instructions.pack(pady=(0, 20))
        
        # PDF File
        pdf_frame = tk.LabelFrame(main_frame, text="PDF File (Name Codes in Column 2)", 
                                  font=("Segoe UI", 11, "bold"), bg="#f0f0f0")
        pdf_frame.pack(fill="x", pady=(0, 15))
        
        pdf_content = tk.Frame(pdf_frame, bg="#f0f0f0")
        pdf_content.pack(fill="x", padx=15, pady=10)
        
        tk.Label(pdf_content, text="PDF File:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.pdf_synthesis_entry = tk.Entry(pdf_content, font=("Segoe UI", 9), width=60, state="readonly")
        self.pdf_synthesis_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        tk.Button(pdf_content, text="Browse...", command=self._browse_pdf_synthesis,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
        
        # Excel File
        excel_frame = tk.LabelFrame(main_frame, text="Excel File (Full Names)", 
                                    font=("Segoe UI", 11, "bold"), bg="#f0f0f0")
        excel_frame.pack(fill="x", pady=(0, 15))
        
        excel_content = tk.Frame(excel_frame, bg="#f0f0f0")
        excel_content.pack(fill="x", padx=15, pady=10)
        
        tk.Label(excel_content, text="Excel File:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.excel_synthesis_entry = tk.Entry(excel_content, font=("Segoe UI", 9), width=60, state="readonly")
        self.excel_synthesis_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        tk.Button(excel_content, text="Browse...", command=self._browse_excel_synthesis,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
        
        # Column Mapping Section for Synthesis Excel
        synthesis_mapping_frame = tk.LabelFrame(main_frame, text="Excel Column Mapping (Optional - Leave empty for auto-detect)", 
                                                font=("Segoe UI", 11, "bold"), bg="#f0f0f0")
        synthesis_mapping_frame.pack(fill="x", pady=(0, 15))
        
        synthesis_mapping_content = tk.Frame(synthesis_mapping_frame, bg="#f0f0f0")
        synthesis_mapping_content.pack(fill="both", expand=True, padx=15, pady=10)
        
        mapping_instruction = ("Map the Session Medium column. Type column name (e.g., 'Session Medium') or column letter (e.g., 'F').\n"
                              "Leave empty to use automatic detection (will check Column F or search for 'Session Medium' in column names).\n"
                              "Click 'Show Excel Columns' after selecting Excel file to see available columns.")
        tk.Label(synthesis_mapping_content, text=mapping_instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=900, justify="left").pack(anchor="w", pady=(0, 10))
        
        # Session Medium mapping row
        session_medium_row = tk.Frame(synthesis_mapping_content, bg="#f0f0f0")
        session_medium_row.pack(fill="x", pady=(0, 10))
        
        tk.Label(session_medium_row, text="Session Medium Column:", font=("Segoe UI", 9), bg="#f0f0f0", width=25, anchor="w").pack(side="left", padx=(0, 10))
        
        self.synthesis_session_medium_entry = tk.Entry(session_medium_row, font=("Segoe UI", 9), width=25)
        self.synthesis_session_medium_entry.pack(side="left", padx=(0, 10))
        self.synthesis_session_medium_entry.bind("<KeyRelease>", lambda e: self._on_synthesis_mapping_changed())
        
        tk.Label(session_medium_row, text="(e.g., 'F' or 'Session Medium')", font=("Segoe UI", 8, "italic"), 
                bg="#f0f0f0", fg="#666666").pack(side="left")
        
        # Show columns button
        button_row = tk.Frame(synthesis_mapping_content, bg="#f0f0f0")
        button_row.pack(fill="x", pady=(5, 0))
        
        tk.Button(button_row, text="Show Excel Columns", command=self._show_synthesis_excel_columns,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
        
        # Output file
        output_frame = tk.LabelFrame(main_frame, text="Output Excel File", 
                                     font=("Segoe UI", 11, "bold"), bg="#f0f0f0")
        output_frame.pack(fill="x", pady=(0, 15))
        
        output_content = tk.Frame(output_frame, bg="#f0f0f0")
        output_content.pack(fill="x", padx=15, pady=10)
        
        tk.Label(output_content, text="Output File:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.synthesis_output_entry = tk.Entry(output_content, font=("Segoe UI", 9), width=60, state="readonly")
        self.synthesis_output_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        tk.Button(output_content, text="Browse...", command=self._browse_synthesis_output,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
        
        # Process button
        button_frame = tk.Frame(main_frame, bg="#f0f0f0")
        button_frame.pack(pady=20)
        
        self.synthesis_process_button = tk.Button(button_frame, text="Synthesize and Generate Output",
                                                 command=self._start_synthesis_processing,
                                                 bg="#006644", fg="white", font=("Segoe UI", 11, "bold"),
                                                 padx=20, pady=10, cursor="hand2", relief="flat")
        self.synthesis_process_button.pack()
        
        # Log area for synthesis tab (shared with main log)
        log_frame = tk.LabelFrame(main_frame, text="Log", font=("Segoe UI", 11, "bold"), bg="#f0f0f0")
        log_frame.pack(fill="both", expand=True, pady=(0, 0))
        
        log_content = tk.Frame(log_frame, bg="#f0f0f0")
        log_content.pack(fill="both", expand=True, padx=15, pady=10)
        
        self.synthesis_log_text = scrolledtext.ScrolledText(log_content, height=15, font=("Courier New", 9),
                                                            bg="#ffffff", fg="#000000", wrap=tk.WORD)
        self.synthesis_log_text.pack(fill="both", expand=True)
    
    def _generate_name_code(self, full_name: str) -> str:
        """Generate name code from full name: First 3 of last name + First 2 of first name + 000
        
        Example: 'Barbara Lukasiewicz' → 'LUKBA000'
        """
        if not full_name:
            return ""
        
        name_parts = str(full_name).strip().upper().split()
        if not name_parts:
            return ""
        
        # Get first name (first part) and last name (last part)
        first_name = name_parts[0] if len(name_parts) > 0 else ""
        last_name = name_parts[-1] if len(name_parts) > 1 else (name_parts[0] if len(name_parts) > 0 else "")
        
        # If only one part, treat it as last name
        if len(name_parts) == 1:
            last_name = name_parts[0]
            first_name = ""
        
        # Extract first 3 letters of last name
        last_code = last_name[:3] if len(last_name) >= 3 else last_name.ljust(3, 'X')
        
        # Extract first 2 letters of first name (or use 'XX' if no first name)
        first_code = first_name[:2] if len(first_name) >= 2 else (first_name.ljust(2, 'X') if first_name else "XX")
        
        # Combine: Last3 + First2 + 000
        name_code = f"{last_code}{first_code}000"
        
        return name_code
    
    def _extract_name_code_from_pdf_value(self, value: str) -> tuple:
        """Extract name code from PDF cell value (handles format like 'LUKBA000')
        
        Returns: (full_code, letter_prefix) where letter_prefix is used for matching
        Example: ('LUKBA000', 'LUKBA')
        """
        if not value:
            return ("", "")
        
        value_str = str(value).strip().upper()
        # Match pattern: 3-7 letters followed by 3 numbers
        match = re.match(r'^([A-Z]{3,7})(\d{3})$', value_str)
        if match:
            letter_part = match.group(1)
            number_part = match.group(2)
            return (value_str, letter_part)  # Return full code and letter prefix
        return ("", "")
    
    def _get_letter_prefix_from_name_code(self, name_code: str) -> str:
        """Extract letter prefix from name code (e.g., 'LUKBA000' → 'LUKBA')"""
        if not name_code:
            return ""
        
        value_str = str(name_code).strip().upper()
        match = re.match(r'^([A-Z]{3,7})', value_str)
        if match:
            return match.group(1)
        return ""
    
    def _name_code_prefix_similarity(self, prefix1: str, prefix2: str) -> float:
        """Calculate similarity between two name code prefixes.
        
        Returns a value between 0.0 (no similarity) and 1.0 (identical).
        Uses simple character overlap and position matching.
        """
        if not prefix1 or not prefix2:
            return 0.0
        
        prefix1 = str(prefix1).strip().upper()
        prefix2 = str(prefix2).strip().upper()
        
        # Exact match
        if prefix1 == prefix2:
            return 1.0
        
        # One is substring of the other (e.g., "LUK" vs "LUKBA")
        if prefix1 in prefix2 or prefix2 in prefix1:
            shorter = min(len(prefix1), len(prefix2))
            longer = max(len(prefix1), len(prefix2))
            return shorter / longer if longer > 0 else 0.0
        
        # Calculate character overlap similarity
        # Check how many characters match at same positions
        min_len = min(len(prefix1), len(prefix2))
        max_len = max(len(prefix1), len(prefix2))
        
        if min_len == 0:
            return 0.0
        
        # Count matching characters at same positions
        matches = sum(1 for i in range(min_len) if prefix1[i] == prefix2[i])
        position_similarity = matches / max_len
        
        # Count character overlap (same characters, any position)
        chars1 = set(prefix1)
        chars2 = set(prefix2)
        overlap = len(chars1 & chars2)
        total_chars = len(chars1 | chars2)
        overlap_similarity = overlap / total_chars if total_chars > 0 else 0.0
        
        # Weighted combination (position matching is more important)
        similarity = (position_similarity * 0.7) + (overlap_similarity * 0.3)
        
        return similarity
    
    def _find_best_name_code_match(self, pdf_prefix: str, excel_prefixes: List[str], threshold: float = 0.6) -> Optional[str]:
        """Find the best matching Excel name code prefix for a PDF prefix.
        
        Args:
            pdf_prefix: The name code prefix from PDF
            excel_prefixes: List of Excel name code prefixes to search
            threshold: Minimum similarity threshold (0.0 to 1.0)
        
        Returns:
            The best matching Excel prefix, or None if no match above threshold
        """
        if not pdf_prefix or not excel_prefixes:
            return None
        
        best_match = None
        best_similarity = 0.0
        
        for excel_prefix in excel_prefixes:
            similarity = self._name_code_prefix_similarity(pdf_prefix, excel_prefix)
            if similarity > best_similarity and similarity >= threshold:
                best_similarity = similarity
                best_match = excel_prefix
        
        return best_match if best_similarity >= threshold else None
    
    def _browse_pdf_synthesis(self):
        """Browse for PDF file for synthesis"""
        filetypes = [("PDF files", "*.pdf"), ("All files", "*.*")]
        path = filedialog.askopenfilename(title="Select PDF File with Name Codes", filetypes=filetypes)
        if path:
            self.pdf_file_path = Path(path)
            self.pdf_synthesis_entry.config(state="normal")
            self.pdf_synthesis_entry.delete(0, tk.END)
            self.pdf_synthesis_entry.insert(0, str(self.pdf_file_path))
            self.pdf_synthesis_entry.config(state="readonly")
            self.gui_log(f"Selected PDF: {self.pdf_file_path.name}")
            # Also log to synthesis tab
            if hasattr(self, 'synthesis_log_text'):
                self.synthesis_log_text.insert(tk.END, f"Selected PDF: {self.pdf_file_path.name}\n")
                self.synthesis_log_text.see(tk.END)
    
    def _browse_excel_synthesis(self):
        """Browse for Excel file for synthesis"""
        filetypes = [("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")]
        path = filedialog.askopenfilename(title="Select Excel File with Full Names", filetypes=filetypes)
        if path:
            self.excel_synthesis_file_path = Path(path)
            self.excel_synthesis_entry.config(state="normal")
            self.excel_synthesis_entry.delete(0, tk.END)
            self.excel_synthesis_entry.insert(0, str(self.excel_synthesis_file_path))
            self.excel_synthesis_entry.config(state="readonly")
            
            # Load columns from Excel file
            self.synthesis_excel_columns = []
            self._load_synthesis_excel_columns()
            
            self.gui_log(f"Selected Excel: {self.excel_synthesis_file_path.name}")
            # Also log to synthesis tab
            if hasattr(self, 'synthesis_log_text'):
                self.synthesis_log_text.insert(tk.END, f"Selected Excel: {self.excel_synthesis_file_path.name}\n")
                if self.synthesis_excel_columns:
                    self.synthesis_log_text.insert(tk.END, f"   Found {len(self.synthesis_excel_columns)} column(s)\n")
                self.synthesis_log_text.see(tk.END)
    
    def _browse_synthesis_output(self):
        """Browse for synthesis output file"""
        filetypes = [("Excel files", "*.xlsx"), ("All files", "*.*")]
        path = filedialog.asksaveasfilename(title="Save Synthesis Output As", 
                                           defaultextension=".xlsx", filetypes=filetypes)
        if path:
            self.synthesis_output_path = Path(path)
            self.synthesis_output_entry.config(state="normal")
            self.synthesis_output_entry.delete(0, tk.END)
            self.synthesis_output_entry.insert(0, str(self.synthesis_output_path))
            self.synthesis_output_entry.config(state="readonly")
            self.gui_log(f"Output file: {self.synthesis_output_path.name}")
            # Also log to synthesis tab
            if hasattr(self, 'synthesis_log_text'):
                self.synthesis_log_text.insert(tk.END, f"Output file: {self.synthesis_output_path.name}\n")
                self.synthesis_log_text.see(tk.END)
    
    def _load_synthesis_excel_columns(self):
        """Load column names from synthesis Excel file"""
        if not self.excel_synthesis_file_path or not self.excel_synthesis_file_path.exists():
            self.synthesis_excel_columns = []
            return
        
        if not EXCEL_AVAILABLE:
            self.synthesis_excel_columns = []
            return
        
        try:
            df = pd.read_excel(self.excel_synthesis_file_path, engine='openpyxl', nrows=0)  # Read only headers
            self.synthesis_excel_columns = list(df.columns)
        except Exception as e:
            self.synthesis_excel_columns = []
            logger.warning(f"Could not load columns from synthesis Excel: {e}")
    
    def _show_synthesis_excel_columns(self):
        """Show available columns from synthesis Excel file"""
        if not self.synthesis_excel_columns:
            if not self.excel_synthesis_file_path:
                messagebox.showinfo("No File", "Please select Excel file first.")
                return
            self._load_synthesis_excel_columns()
        
        if not self.synthesis_excel_columns:
            messagebox.showinfo("No Columns", "No columns found in Excel file.")
            return
        
        def get_column_letter(index):
            """Convert 0-based index to Excel column letter"""
            result = ""
            index += 1
            while index > 0:
                index -= 1
                result = chr(65 + (index % 26)) + result
                index //= 26
            return result
        
        columns_text = "\n".join([f"{get_column_letter(i)}. {col}" for i, col in enumerate(self.synthesis_excel_columns)])
        messagebox.showinfo(
            "Excel Columns",
            f"Found {len(self.synthesis_excel_columns)} column(s) in Excel file:\n\n{columns_text}\n\n"
            "Type column name or letter (e.g., 'F' or 'Session Medium') in the Session Medium mapping field above."
        )
    
    def _on_synthesis_mapping_changed(self):
        """Handle synthesis column mapping change"""
        if hasattr(self, 'synthesis_session_medium_entry'):
            self.synthesis_session_medium_mapping = self.synthesis_session_medium_entry.get().strip()
        else:
            self.synthesis_session_medium_mapping = ""
    
    def _start_synthesis_processing(self):
        """Start the synthesis process in a separate thread"""
        if self.processing:
            messagebox.showwarning("Already Processing", "Synthesis is already in progress.")
            return
        
        if not self.pdf_file_path or not self.pdf_file_path.exists():
            messagebox.showerror("Missing File", "Please select PDF file.")
            return
        
        if not self.excel_synthesis_file_path or not self.excel_synthesis_file_path.exists():
            messagebox.showerror("Missing File", "Please select Excel file.")
            return
        
        if not self.synthesis_output_path:
            messagebox.showerror("Missing Output", "Please select an output file location.")
            return
        
        # Disable process button
        self.synthesis_process_button.config(state="disabled", text="Processing...")
        
        # Start processing in thread
        thread = threading.Thread(target=self._process_synthesis, daemon=True)
        thread.start()
    
    def _synthesis_log(self, message: str, level: str = "INFO"):
        """Log message to both main log and synthesis tab log"""
        self.gui_log(message, level)
        if hasattr(self, 'synthesis_log_text'):
            timestamp = datetime.now().strftime("[%H:%M:%S]")
            log_entry = f"{timestamp} {message}"
            self.root.after(0, lambda: self._append_synthesis_log(log_entry))
    
    def _append_synthesis_log(self, message: str):
        """Append message to synthesis log text widget"""
        if hasattr(self, 'synthesis_log_text'):
            self.synthesis_log_text.config(state="normal")
            self.synthesis_log_text.insert(tk.END, message + "\n")
            self.synthesis_log_text.see(tk.END)
            self.synthesis_log_text.config(state="disabled")
    
    def _process_synthesis(self):
        """Process the PDF/Excel synthesis"""
        try:
            self.processing = True
            self._synthesis_log("\n" + "="*70)
            self._synthesis_log("Starting PDF/Excel Synthesis")
            self._synthesis_log("="*70)
            
            # Parse PDF
            self._synthesis_log(f"\n📄 Parsing PDF: {self.pdf_file_path.name}")
            self.pdf_synthesis_data = self._parse_pdf_synthesis(self.pdf_file_path)
            self._synthesis_log(f"   ✅ Parsed {len(self.pdf_synthesis_data)} record(s) from PDF")
            
            # Parse Excel
            self._synthesis_log(f"\n📄 Parsing Excel: {self.excel_synthesis_file_path.name}")
            self.excel_synthesis_data = self._parse_excel_synthesis(self.excel_synthesis_file_path)
            self._synthesis_log(f"   ✅ Parsed {len(self.excel_synthesis_data)} record(s) from Excel")
            
            # Match and combine
            self._synthesis_log(f"\n🔍 Matching records based on name codes...")
            self._match_and_combine_synthesis()
            
            # Generate output
            self._synthesis_log(f"\n📊 Generating output Excel...")
            self._generate_synthesis_output()
            
            self._synthesis_log(f"\n✅ Synthesis complete! Output saved to: {self.synthesis_output_path}")
            self.root.after(0, lambda: messagebox.showinfo("Success", 
                f"Synthesis complete!\n\n"
                f"Results saved to:\n{self.synthesis_output_path}\n\n"
                f"Total records: {len(self.synthesis_results)}"))
            
        except Exception as e:
            error_msg = f"Error during synthesis: {str(e)}"
            self._synthesis_log(f"\n❌ {error_msg}", level="ERROR")
            logger.exception("Error in _process_synthesis")
            self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
        finally:
            self.processing = False
            self.root.after(0, lambda: self.synthesis_process_button.config(state="normal", text="Synthesize and Generate Output"))
    
    def _parse_pdf_synthesis(self, pdf_path: Path) -> List[Dict[str, Any]]:
        """Parse PDF and extract records with name codes from column 2
        Uses the same robust parsing logic as Medisoft Penelope Data Synthesizer
        """
        if not PDFPLUMBER_AVAILABLE:
            raise Exception("pdfplumber is required. Install with: pip install pdfplumber")
        
        records = []
        seen_record_keys = set()  # Track unique record keys to prevent duplicates
        duplicate_count = 0  # Track how many duplicates we skip
        
        with pdfplumber.open(str(pdf_path)) as pdf:
            total_pages = len(pdf.pages)
            self._synthesis_log(f"   Found {total_pages} page(s)")
            
            for page_num, page in enumerate(pdf.pages, 1):
                if self.stop_requested:
                    return records
                
                # Try to extract tables
                tables = page.extract_tables()
                
                if tables and len(tables) > 0:
                    if page_num <= 3 or page_num % 20 == 0:
                        self._synthesis_log(f"   Page {page_num}: Found {len(tables)} table(s)")
                    
                    for table_idx, table in enumerate(tables):
                        if not table or len(table) == 0:
                            continue
                        
                        # First row should be headers
                        if len(table) < 2:
                            continue
                        
                        # Process data rows - get ALL rows, don't skip any
                        # IMPORTANT: Process ALL rows including row 1 (which might have concatenated data)
                        # Start from index 1 (which is row 2 in Excel, since index 0 is header)
                        for row_idx, row in enumerate(table[1:], start=2):
                            if self.stop_requested:
                                return records
                            
                            # Skip completely empty rows (no cells at all)
                            if not row:
                                continue
                            
                            # Check if row has any data at all
                            # Be very lenient - include rows with ANY non-empty cell
                            has_data = False
                            for cell in row:
                                if cell and str(cell).strip():
                                    has_data = True
                                    break
                            
                            if not has_data:
                                continue
                            
                            # Create row dictionary - store by column index for reliable access
                            row_dict = {}
                            
                            # Store all columns by index for reliable access
                            for col_idx in range(len(row)):
                                value = row[col_idx]
                                col_key = f"Col_{col_idx}"
                                if value is not None:
                                    val_str = str(value).strip()
                                    row_dict[col_key] = val_str
                                else:
                                    row_dict[col_key] = ""
                            
                            # Extract name code from Column 2 (index 2, 0-indexed)
                            # This is the abbreviated name code format: LUKBA000
                            name_code_full = ""
                            name_code_prefix = ""
                            
                            # Method 1: Check Column 2 first (index 2) - this is where name codes are
                            if 2 < len(row) and row[2]:
                                name_code_full, name_code_prefix = self._extract_name_code_from_pdf_value(str(row[2]))
                            
                            # Method 2: Check row 1 (index 1) - it might have concatenated data in column 0
                            if not name_code_full and row_idx == 2 and len(row) > 0 and row[0]:
                                concatenated = str(row[0]).strip()
                                if len(concatenated) > 50:  # Likely concatenated
                                    # Look for name code pattern in concatenated text
                                    name_code_pattern = re.compile(r'\b([A-Z]{3,7}\d{3})\b')
                                    matches = name_code_pattern.findall(concatenated.upper())
                                    if matches:
                                        name_code_full, name_code_prefix = self._extract_name_code_from_pdf_value(matches[0])
                            
                            # Method 3: Search all columns for name code pattern (fallback)
                            if not name_code_full:
                                for col_idx, cell in enumerate(row[:15]):  # Check first 15 columns
                                    if cell:
                                        potential = str(cell).strip()
                                        name_code_full, name_code_prefix = self._extract_name_code_from_pdf_value(potential)
                                        if name_code_full:
                                            break
                            
                            # If no name code found, skip this row
                            if not name_code_full or not name_code_prefix:
                                continue
                            
                            # Store name code information
                            row_dict['Name_Code'] = name_code_full  # Full code with numbers: LUKBA000
                            row_dict['Name_Code_Prefix'] = name_code_prefix  # Letter prefix only: LUKBA
                            row_dict['_page_num'] = page_num
                            row_dict['_row_num'] = row_idx
                            row_dict['_table_idx'] = table_idx
                            
                            # Also extract other fields using the same logic as Medisoft Penelope Data Synthesizer
                            # Column 1: Date of Service (MM/DD/YYYY format)
                            date_of_service = ""
                            if 1 < len(row) and row[1]:
                                # Check if column 1 has date (might be in column 0 or 1)
                                for check_col in [0, 1]:
                                    if check_col < len(row) and row[check_col]:
                                        date_val = str(row[check_col]).strip()
                                        if re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{4}$', date_val):
                                            date_of_service = date_val
                                            break
                            
                            # If not found and this is row 1 (concatenated), extract from column 0
                            if not date_of_service and row_idx == 2 and row[0] and len(str(row[0]).strip()) > 50:
                                concatenated = str(row[0]).strip()
                                date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', concatenated)
                                if date_match:
                                    date_of_service = date_match.group(1)
                            
                            # Store extracted fields
                            row_dict['Date of Service'] = date_of_service
                            row_dict['Date_of_Service'] = date_of_service
                            row_dict['DOS'] = date_of_service
                            
                            # Extract other fields from known column positions (same as Medisoft Penelope Data Synthesizer)
                            # Column 4: Chart/PT Code (if present)
                            chart_value = ""
                            if 4 < len(row) and row[4]:
                                potential = str(row[4]).strip().upper()
                                if re.match(r'^[A-Z]{2,7}\d{3,}$', potential):
                                    chart_value = potential
                            row_dict['Chart'] = chart_value
                            row_dict['PT_Code'] = chart_value
                            
                            # Column 6: Case Number
                            case_num = ""
                            if 6 < len(row) and row[6]:
                                case_val = str(row[6]).strip()
                                if case_val.isdigit() and len(case_val) == 5:
                                    case_num = case_val
                            row_dict['Case'] = case_num
                            row_dict['Case Number'] = case_num
                            
                            # Column 8: Diagnosis Code
                            diagnosis_code = ""
                            if 8 < len(row) and row[8]:
                                diag_val = str(row[8]).strip()
                                if diag_val.isdigit() and len(diag_val) == 1:
                                    diagnosis_code = diag_val
                            row_dict['Diagnosis Code'] = diagnosis_code
                            
                            # Procedure Code extraction - can be in multiple columns
                            # IMPORTANT: Some cells may contain both procedure code and modifier (e.g., "90834 95")
                            # We need to extract them separately
                            procedure_code = ""
                            used_procedure_columns = set()  # Track which columns we used for procedure codes
                            
                            for col_idx in [10, 11, 12]:
                                if col_idx < len(row) and row[col_idx]:
                                    cell_val = str(row[col_idx]).strip()
                                    if not cell_val:
                                        continue
                                    
                                    # Check if cell contains a 5-digit procedure code
                                    # Split by space in case it contains both procedure and modifier
                                    parts = cell_val.split()
                                    for part in parts:
                                        part_clean = part.strip()
                                        if part_clean.isdigit() and len(part_clean) == 5:
                                            procedure_code = part_clean
                                            used_procedure_columns.add(col_idx)
                                            break
                                    
                                    if procedure_code:
                                        break
                            
                            row_dict['Procedure Code'] = procedure_code
                            
                            # Modifier extraction - can be in multiple columns
                            # IMPORTANT: Must NOT extract procedure codes as modifiers
                            # Check columns 12, 13, 14, but skip columns we already used for procedure codes
                            modifier = ""
                            
                            # Priority: Check 13, 14 first (less likely to have procedure codes)
                            # Then check 12 only if we didn't find a procedure code there
                            modifier_check_order = [13, 14]
                            if 12 not in used_procedure_columns:
                                modifier_check_order.insert(0, 12)  # Check 12 first if it's not used for procedure
                            else:
                                modifier_check_order.append(12)  # Check 12 last if it had a procedure code
                            
                            for col_idx in modifier_check_order:
                                if col_idx < len(row) and row[col_idx]:
                                    cell_val = str(row[col_idx]).strip()
                                    if not cell_val:
                                        continue
                                    
                                    # Split by space in case cell contains multiple values
                                    parts = cell_val.split()
                                    for part in parts:
                                        part_clean = part.strip().upper()
                                        
                                        # Skip if it's a 5-digit procedure code
                                        if part_clean.isdigit() and len(part_clean) == 5:
                                            continue
                                        
                                        # Check if it's a valid modifier (1-2 digits or alphanumeric)
                                        if len(part_clean) <= 3:
                                            if re.match(r'^[0-9]{1,2}$|^[A-Z][0-9]$', part_clean):
                                                modifier = part_clean
                                                break
                                    
                                    if modifier:
                                        break
                            
                            # Final validation: Ensure modifier is NOT a procedure code
                            if modifier and modifier.isdigit() and len(modifier) == 5:
                                # This is actually a procedure code, not a modifier
                                if not procedure_code:
                                    procedure_code = modifier
                                modifier = ""
                            
                            row_dict['Modifier'] = modifier
                            row_dict['Mod'] = modifier
                            
                            # Final validation: Ensure procedure code field doesn't contain modifier
                            if procedure_code and procedure_code != row_dict.get('Procedure Code', ''):
                                row_dict['Procedure Code'] = procedure_code
                            
                            # Column 14: Provider/Counselor
                            provider = ""
                            if 14 < len(row) and row[14]:
                                prov_val = str(row[14]).strip()
                                if re.match(r'^[A-Z]{1,4}$', prov_val.upper()) and len(prov_val) <= 4:
                                    provider = prov_val
                            row_dict['Provider'] = provider
                            row_dict['Counselor'] = provider
                            row_dict['Counselor Name'] = provider
                            
                            # Column 18: Amount
                            amount = ""
                            if 18 < len(row) and row[18]:
                                amount_val = str(row[18]).strip()
                                if re.match(r'^\d+\.\d{2}$', amount_val):
                                    amount = amount_val
                            row_dict['Amount'] = amount
                            
                            # DEDUPLICATION: Check if we've already seen this exact record
                            # Use a unique key: name_code + date + procedure_code + modifier (to distinguish same client/date/proc with different modifiers)
                            record_key = f"{name_code_full}|{date_of_service}|{procedure_code}|{modifier}"
                            
                            # Only deduplicate if we have enough info to create a meaningful key
                            if record_key and (name_code_full or date_of_service or procedure_code):
                                # Check if this record key already exists using set for O(1) lookup
                                if record_key in seen_record_keys:
                                    duplicate_count += 1
                                    # Skip duplicate records
                                    continue
                                
                                # Mark this key as seen
                                seen_record_keys.add(record_key)
                            
                            # Add the record
                            records.append(row_dict)
                            
                            # Log progress every 500 rows
                            if len(records) % 500 == 0:
                                self._synthesis_log(f"   Processed {len(records)} unique PDF rows so far...")
                
                # Log total progress every 10 pages
                if page_num % 10 == 0:
                    self._synthesis_log(f"   📊 Total PDF rows extracted so far: {len(records)} (through page {page_num})")
        
        self._synthesis_log(f"   ✅ Extracted {len(records)} unique record(s) from PDF")
        if duplicate_count > 0:
            self._synthesis_log(f"   ⚠️  Skipped {duplicate_count} duplicate record(s) (same name code, date, procedure, and modifier)")
        
        # Debug: Show sample data
        if records:
            self._synthesis_log(f"   📋 Sample extracted data (first 3 rows):")
            for idx, row in enumerate(records[:3], 1):
                name_code = row.get('Name_Code', '')
                date = row.get('Date_of_Service', '')
                modifier = row.get('Modifier', '')
                provider = row.get('Provider', '')
                self._synthesis_log(f"      Row {idx}: NameCode={name_code}, Date={date}, Modifier={modifier}, Provider={provider}")
        
        return records
    
    def _parse_excel_synthesis(self, excel_path: Path) -> List[Dict[str, Any]]:
        """Parse Excel and generate name codes from full names"""
        if not EXCEL_AVAILABLE:
            raise Exception("pandas is required. Install with: pip install pandas openpyxl")
        
        df = pd.read_excel(excel_path, engine='openpyxl')
        
        records = []
        for idx, row in df.iterrows():
            # Try to find name column (check common names)
            name_value = ""
            for col in df.columns:
                col_lower = str(col).lower()
                if any(term in col_lower for term in ['name', 'patient', 'client']):
                    name_value = str(row[col]).strip() if pd.notna(row[col]) else ""
                    if name_value:
                        break
            
            # If no name found, skip this row
            if not name_value:
                continue
            
            # Generate name code
            name_code = self._generate_name_code(name_value)
            name_code_prefix = self._get_letter_prefix_from_name_code(name_code)
            
            # Extract Date of Service from Excel
            date_of_service = ""
            date_of_service_normalized = ""
            for col in df.columns:
                col_lower = str(col).lower()
                if any(term in col_lower for term in ['date of service', 'dos', 'service date', 'date']):
                    date_value = row[col]
                    if pd.notna(date_value):
                        date_str = str(date_value).strip()
                        # Normalize date for matching
                        date_of_service = date_str
                        date_of_service_normalized = self._normalize_date(date_str)
                        break
            
            # Extract Procedure Code/Service Code from Excel (Column E = index 4)
            # Priority: Check column name first, then fallback to Column E (index 4)
            procedure_code = ""
            procedure_code_col = None
            
            def normalize_procedure_code(val):
                """Normalize procedure code - handle floats like 90834.0 -> 90834"""
                if pd.isna(val):
                    return ""
                # Convert to string and strip
                proc_str = str(val).strip()
                # Remove .0 suffix if it's a float representation (e.g., "90834.0" -> "90834")
                if proc_str.endswith('.0'):
                    # Check if it's a whole number float (e.g., "90834.0")
                    try:
                        float_val = float(proc_str)
                        if float_val == int(float_val):
                            proc_str = str(int(float_val))
                    except (ValueError, TypeError):
                        # If conversion fails, just remove .0 if present
                        if proc_str.endswith('.0'):
                            proc_str = proc_str[:-2]
                return proc_str
            
            # Method 1: Check column names for procedure/service code
            for col_idx, col in enumerate(df.columns):
                col_lower = str(col).lower()
                if any(term in col_lower for term in ['procedure', 'service code', 'code', 'cpt']):
                    if pd.notna(row[col]):
                        proc_val = normalize_procedure_code(row[col])
                        # Check if it's a 5-digit procedure code (e.g., "90834")
                        if proc_val.isdigit() and len(proc_val) == 5:
                            procedure_code = proc_val
                            procedure_code_col = col
                            break
                        # Also check if it contains a 5-digit code (e.g., "90834 95" or "90834-95")
                        parts = re.split(r'[\s\-]+', proc_val)
                        for part in parts:
                            part_norm = normalize_procedure_code(part)
                            if part_norm.isdigit() and len(part_norm) == 5:
                                procedure_code = part_norm
                                procedure_code_col = col
                                break
                        if procedure_code:
                            break
            
            # Method 2: If not found, check Column E (index 4) directly
            if not procedure_code and len(df.columns) > 4:
                col_e = df.columns[4]  # Column E (0-indexed = 4)
                if pd.notna(row[col_e]):
                    proc_val = normalize_procedure_code(row[col_e])
                    if proc_val.isdigit() and len(proc_val) == 5:
                        procedure_code = proc_val
                        procedure_code_col = col_e
                    else:
                        # Check if it contains a 5-digit code
                        parts = re.split(r'[\s\-]+', proc_val)
                        for part in parts:
                            part_norm = normalize_procedure_code(part)
                            if part_norm.isdigit() and len(part_norm) == 5:
                                procedure_code = part_norm
                                procedure_code_col = col_e
                                break
            
            # Normalize procedure code before storing (handle floats)
            if procedure_code:
                procedure_code = normalize_procedure_code(procedure_code)
            
            # Extract Session Medium from Excel
            # Priority: User-provided mapping > Column name search > Column F (index 5)
            session_medium = ""
            
            # Method 1: Use user-provided mapping if available
            if self.synthesis_session_medium_mapping:
                mapping = self.synthesis_session_medium_mapping.strip()
                if mapping:
                    # Try to resolve the mapping (could be column name or letter)
                    resolved_col = None
                    
                    # Check if it's a column letter (e.g., "F", "AB")
                    if re.match(r'^[A-Z]+$', mapping.upper()):
                        # Convert column letter to index (e.g., "F" = 5, "A" = 0)
                        col_letter = mapping.upper()
                        col_index = 0
                        for char in col_letter:
                            col_index = col_index * 26 + (ord(char) - ord('A') + 1)
                        col_index -= 1  # Convert to 0-based
                        
                        if 0 <= col_index < len(df.columns):
                            resolved_col = df.columns[col_index]
                    else:
                        # Check if it's a column name (exact match or partial)
                        for col in df.columns:
                            col_str = str(col).strip()
                            if mapping.lower() == col_str.lower() or mapping.lower() in col_str.lower():
                                resolved_col = col
                                break
                    
                    # Extract value from resolved column
                    if resolved_col is not None and pd.notna(row[resolved_col]):
                        session_medium = str(row[resolved_col]).strip()
            
            # Method 2: Auto-detect if no user mapping provided or mapping didn't work
            if not session_medium:
                # Check column names for session medium
                for col in df.columns:
                    col_lower = str(col).lower()
                    if any(term in col_lower for term in ['session medium', 'medium', 'session type', 'session medium type']):
                        if pd.notna(row[col]):
                            session_medium = str(row[col]).strip()
                            if session_medium:
                                break
                
                # Method 3: If still not found, check Column F (index 5) directly
                if not session_medium and len(df.columns) > 5:
                    col_f = df.columns[5]  # Column F (0-indexed = 5)
                    if pd.notna(row[col_f]):
                        session_medium = str(row[col_f]).strip()
            
            # Create record
            record = {
                'Excel_Name': name_value,
                'Name_Code': name_code,  # Full code with 000: LUKBA000
                'Name_Code_Prefix': name_code_prefix,  # Letter prefix only: LUKBA
                'Excel_Date_of_Service': date_of_service,  # Original date
                'Excel_Date_of_Service_Normalized': date_of_service_normalized,  # Normalized for matching
                'Excel_Procedure_Code': procedure_code,  # Procedure code from Excel (Column E)
                'Excel_Session_Medium': session_medium,  # Session medium from Excel (Column F or by name)
                'Excel_Row': idx + 2  # +2 because Excel rows start at 1 and we skip header
            }
            
            # Store all Excel columns
            for col in df.columns:
                value = row[col]
                if pd.notna(value):
                    record[f'Excel_{col}'] = value
                else:
                    record[f'Excel_{col}'] = ""
            
            records.append(record)
        
        return records
    
    def _match_and_combine_synthesis(self):
        """Match PDF and Excel records based on name codes, date of service, AND procedure code"""
        self.synthesis_results = []
        match_count = 0
        no_match_count = 0
        
        def normalize_proc_code(proc_code):
            """Normalize procedure code for comparison - handle floats and strings"""
            if not proc_code or pd.isna(proc_code):
                return ""
            proc_str = str(proc_code).strip()
            # Remove .0 suffix if it's a float representation (e.g., "90834.0" -> "90834")
            if proc_str.endswith('.0'):
                try:
                    float_val = float(proc_str)
                    if float_val == int(float_val):
                        proc_str = str(int(float_val))
                except (ValueError, TypeError):
                    # If conversion fails, just remove .0 if present
                    if proc_str.endswith('.0'):
                        proc_str = proc_str[:-2]
            # Also handle cases where it might be a float that pandas converted
            elif '.' in proc_str:
                try:
                    float_val = float(proc_str)
                    if float_val == int(float_val):
                        proc_str = str(int(float_val))
                except (ValueError, TypeError):
                    pass
            return proc_str
        
        # Build lookup dictionary for Excel records by (name_code_prefix, normalized_date, procedure_code)
        # Key: (name_code_prefix, normalized_date, procedure_code) -> List[Excel records]
        # Also create fallback lookups for partial matches
        excel_lookup_full: Dict[tuple, List[Dict]] = {}  # Full match: name + date + procedure
        excel_lookup_date: Dict[tuple, List[Dict]] = {}  # Partial match: name + date
        excel_lookup_name: Dict[str, List[Dict]] = {}  # Fallback: name only
        
        for excel_record in self.excel_synthesis_data:
            name_code_prefix = excel_record.get('Name_Code_Prefix', '')
            excel_date_normalized = excel_record.get('Excel_Date_of_Service_Normalized', '')
            excel_procedure_code = normalize_proc_code(excel_record.get('Excel_Procedure_Code', ''))
            
            if not name_code_prefix:
                continue
            
            # Full match: name + date + procedure code
            if excel_date_normalized and excel_procedure_code:
                full_key = (name_code_prefix, excel_date_normalized, excel_procedure_code)
                if full_key not in excel_lookup_full:
                    excel_lookup_full[full_key] = []
                excel_lookup_full[full_key].append(excel_record)
            
            # Partial match: name + date (no procedure code)
            if excel_date_normalized:
                date_key = (name_code_prefix, excel_date_normalized)
                if date_key not in excel_lookup_date:
                    excel_lookup_date[date_key] = []
                excel_lookup_date[date_key].append(excel_record)
            
            # Fallback: name only
            if name_code_prefix not in excel_lookup_name:
                excel_lookup_name[name_code_prefix] = []
            excel_lookup_name[name_code_prefix].append(excel_record)
        
        self._synthesis_log(f"   Built Excel lookup:")
        self._synthesis_log(f"      Full matches (name+date+procedure): {len(excel_lookup_full)}")
        self._synthesis_log(f"      Partial matches (name+date): {len(excel_lookup_date)}")
        self._synthesis_log(f"      Name-only fallback: {len(excel_lookup_name)}")
        
        # Match PDF records with Excel records (match on name code prefix, date of service, AND procedure code)
        # IMPORTANT: Multiple PDF name codes with same prefix but different numbers (e.g., WILD1000, WILD1001)
        # should NOT all match to the same Excel record. We need to ensure procedure code matches to distinguish them.
        
        # Track which Excel records have already been matched to avoid double-matching
        # Key: (name_prefix, date, procedure_code) -> list of matched PDF name codes
        excel_match_history: Dict[tuple, List[str]] = {}
        
        # Track which Excel records got matched (by their unique key)
        matched_excel_keys: set = set()  # Track (name_code_prefix, date_normalized, proc_normalized) keys that were matched
        excel_records_output: set = set()  # Track (excel_name, excel_date, excel_proc) that have been output - ONE row per Excel record
        
        for pdf_record in self.pdf_synthesis_data:
            pdf_name_code = pdf_record.get('Name_Code', '')  # Full code: WILD1000 or WILD1001
            pdf_name_code_prefix = pdf_record.get('Name_Code_Prefix', '')  # Prefix: WILDI
            pdf_date = pdf_record.get('Date_of_Service', '')  # Original date from PDF
            pdf_date_normalized = self._normalize_date(pdf_date) if pdf_date else ""  # Normalized for matching
            # Normalize PDF procedure code (might be string, float, or int)
            pdf_proc_raw = pdf_record.get('Procedure Code', '')
            pdf_procedure_code = normalize_proc_code(pdf_proc_raw) if pdf_proc_raw else ""
            
            # Try to match on name code + date + procedure code first (most accurate)
            matched_excel = None
            match_method = "none"
            
            if pdf_name_code_prefix:
                # CRITICAL: We must match on name prefix + date + procedure code to avoid matching
                # different PDF name codes (WILD1000, WILD1001) to the same Excel record
                
                # Method 1: Try exact match on (name_code_prefix, normalized_date, procedure_code) - MOST ACCURATE
                if pdf_date_normalized and pdf_procedure_code:
                    full_lookup_key = (pdf_name_code_prefix, pdf_date_normalized, pdf_procedure_code)
                    if full_lookup_key in excel_lookup_full:
                        # Check if this Excel record was already matched to a different PDF name code
                        if full_lookup_key in excel_match_history:
                            matched_codes = excel_match_history[full_lookup_key]
                            if pdf_name_code in matched_codes:
                                # Same PDF code, already matched - skip
                                matched_excel = excel_lookup_full[full_lookup_key][0]
                                match_method = "name+date+procedure (already matched)"
                            else:
                                # Different PDF code trying to match same Excel - this is suspicious
                                # Only allow if procedure code matches exactly
                                matched_excel = excel_lookup_full[full_lookup_key][0]
                                match_method = f"name+date+procedure (different PDF code: {pdf_name_code} vs {matched_codes[0]})"
                                self._synthesis_log(f"   ⚠️  Warning: PDF code {pdf_name_code} matches same Excel as {matched_codes[0]} - verifying procedure code")
                        else:
                            matched_excel = excel_lookup_full[full_lookup_key][0]
                            match_method = "name+date+procedure"
                            match_count += 1
                            # Track this match
                            if full_lookup_key not in excel_match_history:
                                excel_match_history[full_lookup_key] = []
                            excel_match_history[full_lookup_key].append(pdf_name_code)
                            # Mark this Excel record as matched
                            matched_excel_keys.add(full_lookup_key)
                
                # Method 2: Try match on (name_code_prefix, normalized_date) - MUST verify procedure code
                if not matched_excel and pdf_date_normalized:
                    # Try exact date match first
                    date_lookup_key = (pdf_name_code_prefix, pdf_date_normalized)
                    if date_lookup_key in excel_lookup_date:
                        # Check if any of the matching Excel records have a matching procedure code
                        potential_matches = excel_lookup_date[date_lookup_key]
                        
                        if pdf_procedure_code:
                            # CRITICAL: Only match if procedure code matches - this distinguishes different clients
                            # with the same name prefix (e.g., WILD1000 vs WILD1001)
                            # Additionally, prefer matches where PDF name code suffix matches Excel-generated suffix (000)
                            best_match = None
                            best_match_score = -1
                            
                            for candidate in potential_matches:
                                candidate_proc = normalize_proc_code(candidate.get('Excel_Procedure_Code', ''))
                                # Normalize both for comparison
                                if candidate_proc and pdf_procedure_code and candidate_proc == pdf_procedure_code:
                                    # Calculate match score:
                                    # - Score 2: Excel generates "000" and PDF ends in "000" (e.g., WILD1000) - PERFECT MATCH
                                    # - Score 1: Excel generates "000" but PDF ends differently (e.g., WILD1001) - VERIFY
                                    excel_generated_code = candidate.get('Name_Code', '')
                                    pdf_name_code_suffix = pdf_name_code[-3:] if len(pdf_name_code) >= 3 else ""
                                    excel_name_code_suffix = excel_generated_code[-3:] if len(excel_generated_code) >= 3 else ""
                                    
                                    match_score = 0
                                    if pdf_name_code_suffix == "000" or pdf_name_code_suffix == excel_name_code_suffix:
                                        match_score = 2  # Suffix matches - preferred
                                    elif excel_name_code_suffix == "000" and pdf_name_code_suffix != "000":
                                        match_score = 1  # Different suffix - still valid if procedure matches
                                    
                                    # Only consider this match if it's better than previous best
                                    if match_score > best_match_score:
                                        # Verify this isn't already matched to a different PDF code
                                        candidate_key = (pdf_name_code_prefix, pdf_date_normalized, candidate_proc)
                                        if candidate_key not in excel_match_history or pdf_name_code in excel_match_history[candidate_key]:
                                            best_match = candidate
                                            best_match_score = match_score
                            
                            if best_match:
                                matched_excel = best_match
                                if best_match_score == 2:
                                    match_method = "name+date+procedure (found in partial, suffix matches)"
                                else:
                                    match_method = f"name+date+procedure (found in partial, suffix differs: PDF={pdf_name_code[-3:]}, Excel=000)"
                                    self._synthesis_log(f"   ⚠️  Warning: PDF {pdf_name_code} has different suffix but procedure code matches - VERIFY this is same client")
                                
                                match_count += 1
                                # Track this match
                                best_match_proc = normalize_proc_code(best_match.get('Excel_Procedure_Code', ''))
                                candidate_key = (pdf_name_code_prefix, pdf_date_normalized, best_match_proc)
                                if candidate_key not in excel_match_history:
                                    excel_match_history[candidate_key] = []
                                excel_match_history[candidate_key].append(pdf_name_code)
                                # Mark this Excel record as matched
                                matched_excel_keys.add(candidate_key)
                        
                        # REQUIRE exact procedure code match - no flexible matching
                        # If no procedure code match found, do not match
                        if not matched_excel:
                            if pdf_procedure_code:
                                match_method = f"name+date (procedure code mismatch: PDF={pdf_procedure_code}, Excel candidates don't match)"
                                self._synthesis_log(f"   ⚠️  No match for PDF {pdf_name_code}: procedure code {pdf_procedure_code} doesn't match any Excel records")
                            else:
                                match_method = f"name+date (procedure code missing in PDF - cannot verify match)"
                                self._synthesis_log(f"   ⚠️  No match for PDF {pdf_name_code}: procedure code missing in PDF")
                
                # Method 3: STRICT - No fuzzy matching, no fallback without exact matches
                # Only match if we have exact name code prefix, date, AND procedure code
                # No matching on name only or name+date without procedure code
            
            if not matched_excel:
                no_match_count += 1
                # Log detailed reason why no match was found
                reasons = []
                if not pdf_name_code_prefix:
                    reasons.append("missing name code prefix")
                else:
                    reasons.append(f"name code prefix: '{pdf_name_code_prefix}'")
                
                if not pdf_date_normalized:
                    reasons.append("missing or invalid date")
                else:
                    reasons.append(f"date: '{pdf_date_normalized}'")
                
                if not pdf_procedure_code:
                    reasons.append("missing procedure code")
                else:
                    reasons.append(f"procedure code: '{pdf_procedure_code}'")
                
                # Check if name code exists in Excel but date/procedure don't match
                name_exists = pdf_name_code_prefix in excel_lookup_name
                if name_exists and no_match_count <= 10:  # Log first 10 no-matches with details
                    excel_candidates = excel_lookup_name[pdf_name_code_prefix]
                    sample_excel = excel_candidates[0] if excel_candidates else None
                    if sample_excel:
                        excel_date = sample_excel.get('Excel_Date_of_Service_Normalized', '')
                        excel_proc = sample_excel.get('Excel_Procedure_Code', '')
                        self._synthesis_log(f"   ❌ No match for PDF {pdf_name_code}: Found Excel with same name, but date/procedure differ")
                        self._synthesis_log(f"      PDF: date={pdf_date_normalized}, proc={pdf_procedure_code}")
                        self._synthesis_log(f"      Excel sample: date={excel_date}, proc={excel_proc}")
                
                match_method = f"no match ({', '.join(reasons)})"
            
            # Combine records
            combined = {}
            
            # Add PDF fields (this will include Date_of_Service from pdf_record)
            for key, value in pdf_record.items():
                combined[f'PDF_{key}'] = value
            
            # Ensure PDF date and procedure code info is explicitly set
            combined['PDF_Date_of_Service'] = pdf_date
            combined['PDF_Date_of_Service_Normalized'] = pdf_date_normalized
            combined['PDF_Procedure_Code'] = pdf_procedure_code
            combined['PDF_Name_Code_Prefix'] = pdf_name_code_prefix  # For debugging
            
            # Add Excel fields if matched
            if matched_excel:
                # Track that this Excel record has been output (one row per Excel record)
                excel_name = matched_excel.get('Excel_Name', '')
                excel_date = matched_excel.get('Excel_Date_of_Service', '')
                excel_proc_raw = matched_excel.get('Excel_Procedure_Code', '')
                excel_proc = normalize_proc_code(excel_proc_raw)
                
                # Check if this Excel record has already been output (only one row per Excel record)
                if excel_name and excel_date and excel_proc:
                    excel_output_key = (str(excel_name).strip(), str(excel_date).strip(), excel_proc)
                    if excel_output_key in excel_records_output:
                        # This Excel record was already matched to a different PDF record - skip this duplicate
                        # Don't create a duplicate row, skip to next PDF record
                        # But still count this as a match for statistics
                        continue
                    else:
                        # First time this Excel record is being matched - track it
                        excel_records_output.add(excel_output_key)
                
                for key, value in matched_excel.items():
                    if key not in combined:
                        combined[f'Excel_{key}'] = value
                combined['Match_Status'] = 'Matched'
                combined['Name_Code'] = pdf_name_code
                combined['Match_Method'] = match_method
                
                # Add match key showing what was matched on
                excel_date_normalized = matched_excel.get('Excel_Date_of_Service_Normalized', '')
                excel_procedure_code = normalize_proc_code(matched_excel.get('Excel_Procedure_Code', ''))
                
                if pdf_date_normalized and pdf_procedure_code:
                    if excel_date_normalized == pdf_date_normalized and excel_procedure_code == pdf_procedure_code:
                        combined['Match_Key'] = f"{pdf_name_code_prefix}|{pdf_date_normalized}|{pdf_procedure_code} (Name+Date+Procedure)"
                    elif excel_date_normalized == pdf_date_normalized:
                        combined['Match_Key'] = f"{pdf_name_code_prefix}|{pdf_date_normalized} (Name+Date, Proc: PDF={pdf_procedure_code}, Excel={excel_procedure_code})"
                    else:
                        combined['Match_Key'] = f"{pdf_name_code_prefix} (Name only, Date/Proc may differ)"
                elif pdf_date_normalized:
                    if excel_date_normalized == pdf_date_normalized:
                        combined['Match_Key'] = f"{pdf_name_code_prefix}|{pdf_date_normalized} (Name+Date, no procedure)"
                    else:
                        combined['Match_Key'] = f"{pdf_name_code_prefix} (Name only - date mismatch)"
                else:
                    combined['Match_Key'] = f"{pdf_name_code_prefix} (Name only - no PDF date)"
                
                # MODIFIER COMPARISON LOGIC: Compare Session Medium from Excel with Modifier from PDF
                # Expected modifier mapping based on session medium
                session_medium_to_modifier = {
                    "video": "95",
                    "telehealth": "95",
                    "telemedicine": "95",
                    "phone": "93",
                    "telephone": "93",
                    "in-person": "",
                    "in person": "",
                    "office": "",
                    "in office": ""
                }
                
                # Get session medium from Excel (disregard any modifiers in Excel - those are incorrect)
                excel_session_medium = matched_excel.get('Excel_Session_Medium', '').strip()
                excel_session_medium_lower = excel_session_medium.lower() if excel_session_medium else ""
                
                # Get modifier from PDF (this is the actual modifier that was billed)
                pdf_modifier = pdf_record.get('Modifier', '').strip() or pdf_record.get('Mod', '').strip()
                
                # Determine expected modifier based on session medium
                expected_modifier = ""
                for key, mod in session_medium_to_modifier.items():
                    if key in excel_session_medium_lower:
                        expected_modifier = mod
                        break
                
                # Compare and determine if refiling is needed
                needs_refile = False
                refile_reason = ""
                
                if excel_session_medium:
                    if expected_modifier and pdf_modifier:
                        # Both expected and actual modifier exist - check if they match
                        if expected_modifier != pdf_modifier:
                            needs_refile = True
                            refile_reason = f"Modifier mismatch: Session medium is '{excel_session_medium}' (expects {expected_modifier}), but PDF shows {pdf_modifier}"
                    elif expected_modifier and not pdf_modifier:
                        # Expected modifier but none found in PDF
                        needs_refile = True
                        refile_reason = f"Missing modifier: Session medium is '{excel_session_medium}' (expects {expected_modifier}), but no modifier found in PDF"
                    elif not expected_modifier and pdf_modifier:
                        # In-person should have no modifier, but PDF has one
                        needs_refile = True
                        refile_reason = f"Unexpected modifier: Session medium is '{excel_session_medium}' (in-person should have no modifier), but PDF shows {pdf_modifier}"
                    elif not expected_modifier and not pdf_modifier:
                        # In-person with no modifier - correct
                        refile_reason = "Correct: In-person session with no modifier"
                    else:
                        # Session medium found but no clear mapping
                        refile_reason = f"Session medium '{excel_session_medium}' found but unable to determine expected modifier"
                else:
                    # No session medium in Excel
                    refile_reason = "No session medium found in Excel - cannot verify modifier"
                
                # Add comparison results to combined record
                combined['Session_Medium'] = excel_session_medium
                combined['PDF_Modifier'] = pdf_modifier
                combined['Expected_Modifier'] = expected_modifier
                combined['Needs_Refile'] = 'Yes' if needs_refile else 'No'
                combined['Refile_Reason'] = refile_reason
            else:
                combined['Match_Status'] = 'No Match'
                combined['Name_Code'] = pdf_name_code
                combined['Match_Method'] = match_method
                combined['Match_Key'] = pdf_name_code_prefix
                combined['No_Excel_Match'] = f'No matching Excel record found for name code "{pdf_name_code_prefix}"'
                if pdf_date_normalized:
                    combined['No_Excel_Match'] += f' with date "{pdf_date_normalized}"'
                if pdf_procedure_code:
                    combined['No_Excel_Match'] += f' and procedure code "{pdf_procedure_code}"'
                
                # For no-match cases, still show PDF modifier but can't compare
                pdf_modifier = pdf_record.get('Modifier', '').strip() or pdf_record.get('Mod', '').strip()
                combined['Session_Medium'] = ""  # No Excel match, so no session medium
                combined['PDF_Modifier'] = pdf_modifier
                combined['Expected_Modifier'] = ""
                combined['Needs_Refile'] = 'Cannot Determine'
                combined['Refile_Reason'] = "No matching Excel record found - cannot verify modifier without session medium"
            
            self.synthesis_results.append(combined)
        
        # SECOND PASS: Ensure all Excel records are included in output (ONE row per Excel record)
        # Track which Excel records have already been output (using normalized keys)
        # Each Excel record should appear ONCE in the output
        self._synthesis_log(f"\n📋 Second Pass: Ensuring all Excel records are included...")
        
        # Track Excel records that have been output using normalized keys
        excel_records_in_output = set()
        for result in self.synthesis_results:
            # Check both Matched and No PDF Match status
            if result.get('Match_Status') in ['Matched', 'No PDF Match']:
                # Get Excel data - try different column names
                excel_name = result.get('Excel_Name', '') or result.get('Excel_Excel_Name', '')
                excel_date = result.get('Excel_Date_of_Service', '') or result.get('Excel_Excel_Date of Service', '') or result.get('Excel_Excel_Date_of_Service', '')
                excel_proc_raw = result.get('Excel_Procedure_Code', '') or result.get('Excel_Excel_Procedure_Code', '')
                
                # Normalize all values for consistent tracking
                if excel_name and excel_date and excel_proc_raw:
                    excel_name_norm = str(excel_name).strip()
                    excel_date_norm = str(excel_date).strip()
                    excel_proc_norm = normalize_proc_code(excel_proc_raw)
                    
                    if excel_name_norm and excel_date_norm and excel_proc_norm:
                        excel_key = (excel_name_norm, excel_date_norm, excel_proc_norm)
                        excel_records_in_output.add(excel_key)
        
        added_excel_count = 0
        matched_in_second_pass = 0
        
        for excel_record in self.excel_synthesis_data:
            excel_name = excel_record.get('Excel_Name', '')
            excel_date = excel_record.get('Excel_Date_of_Service', '')
            excel_proc_raw = excel_record.get('Excel_Procedure_Code', '')
            excel_proc = normalize_proc_code(excel_proc_raw)
            
            if excel_name and excel_date and excel_proc:
                # Create normalized key for tracking (must match format from first pass)
                excel_name_norm = str(excel_name).strip()
                excel_date_norm = str(excel_date).strip()
                excel_proc_norm = excel_proc  # Already normalized
                excel_key = (excel_name_norm, excel_date_norm, excel_proc_norm)
                
                if excel_key not in excel_records_output:
                    # This Excel record wasn't matched - try to find a PDF match
                    excel_name_code_prefix = excel_record.get('Name_Code_Prefix', '')
                    excel_date_normalized = excel_record.get('Excel_Date_of_Service_Normalized', '')
                    
                    # Search for matching PDF record
                    matching_pdf = None
                    for pdf_record in self.pdf_synthesis_data:
                        pdf_name_code_prefix = pdf_record.get('Name_Code_Prefix', '')
                        pdf_date = pdf_record.get('Date_of_Service', '')
                        pdf_date_normalized = self._normalize_date(pdf_date) if pdf_date else ""
                        pdf_proc_raw = pdf_record.get('Procedure Code', '')
                        pdf_proc = normalize_proc_code(pdf_proc_raw) if pdf_proc_raw else ""
                        
                        # Check if this PDF record matches the Excel record
                        if (pdf_name_code_prefix == excel_name_code_prefix and 
                            pdf_date_normalized == excel_date_normalized and 
                            pdf_proc == excel_proc):
                            matching_pdf = pdf_record
                            break
                    
                    combined = {}
                    
                    # Add all Excel fields
                    for key, value in excel_record.items():
                        combined[f'Excel_{key}'] = value
                    
                    if matching_pdf:
                        # Found a match! Use the same logic as first pass
                        matched_in_second_pass += 1
                        combined['Match_Status'] = 'Matched'
                        combined['Name_Code'] = matching_pdf.get('Name_Code', '')
                        combined['Match_Method'] = 'Matched in second pass (Excel -> PDF)'
                        combined['Match_Key'] = f"{excel_name_code_prefix}|{excel_date_normalized}|{excel_proc}"
                        
                        # Add PDF fields
                        combined['PDF_Date_of_Service'] = matching_pdf.get('Date_of_Service', '')
                        combined['PDF_Date_of_Service_Normalized'] = pdf_date_normalized
                        combined['PDF_Procedure_Code'] = pdf_proc
                        pdf_modifier = matching_pdf.get('Modifier', '').strip() or matching_pdf.get('Mod', '').strip()
                        combined['PDF_Modifier'] = pdf_modifier
                        combined['PDF_Name_Code'] = matching_pdf.get('Name_Code', '')
                        combined['PDF_Name_Code_Prefix'] = pdf_name_code_prefix
                        
                        # Add all other PDF fields
                        for key, value in matching_pdf.items():
                            if key not in ['Date_of_Service', 'Procedure Code', 'Modifier', 'Mod', 'Name_Code', 'Name_Code_Prefix']:
                                combined[f'PDF_{key}'] = value
                        
                        # Determine expected modifier and refile status
                        excel_session_medium = excel_record.get('Excel_Session_Medium', '').strip().upper()
                        expected_modifier = ""
                        if 'VIDEO' in excel_session_medium:
                            expected_modifier = "95"
                        elif 'PHONE' in excel_session_medium:
                            expected_modifier = "93"
                        
                        needs_refile = False
                        refile_reason = ""
                        
                        if excel_session_medium:
                            if expected_modifier and pdf_modifier:
                                if expected_modifier != pdf_modifier:
                                    needs_refile = True
                                    refile_reason = f"Modifier mismatch: Session medium is '{excel_session_medium}' (expects {expected_modifier}), but PDF shows {pdf_modifier}"
                                else:
                                    refile_reason = f"Correct: Session medium '{excel_session_medium}' matches modifier {pdf_modifier}"
                            elif expected_modifier and not pdf_modifier:
                                needs_refile = True
                                refile_reason = f"Missing modifier: Session medium is '{excel_session_medium}' (expects {expected_modifier}), but no modifier found in PDF"
                            elif not expected_modifier and pdf_modifier:
                                needs_refile = True
                                refile_reason = f"Unexpected modifier: Session medium is '{excel_session_medium}' (in-person should have no modifier), but PDF shows {pdf_modifier}"
                            else:
                                refile_reason = f"Session medium '{excel_session_medium}' found but unable to determine expected modifier"
                        else:
                            refile_reason = "No session medium found in Excel - cannot verify modifier"
                        
                        combined['Session_Medium'] = excel_record.get('Excel_Session_Medium', '').strip()
                        combined['Expected_Modifier'] = expected_modifier
                        combined['Needs_Refile'] = 'Yes' if needs_refile else 'No'
                        combined['Refile_Reason'] = refile_reason
                    else:
                        # No PDF match found
                        combined['Match_Status'] = 'No PDF Match'
                        combined['Name_Code'] = excel_record.get('Name_Code', '')
                        combined['Match_Method'] = 'Excel record has no matching PDF record'
                        combined['Match_Key'] = f"{excel_name_code_prefix}|{excel_date_normalized}|{excel_proc}"
                        
                        # Add PDF fields as empty
                        combined['PDF_Date_of_Service'] = ""
                        combined['PDF_Procedure_Code'] = ""
                        combined['PDF_Modifier'] = ""
                        
                        # Add comparison fields
                        excel_session_medium = excel_record.get('Excel_Session_Medium', '').strip()
                        combined['Session_Medium'] = excel_session_medium
                        combined['PDF_Modifier'] = ""
                        combined['Expected_Modifier'] = ""
                        combined['Needs_Refile'] = 'Cannot Determine'
                        combined['Refile_Reason'] = "No matching PDF record found - cannot verify modifier"
                    
                    self.synthesis_results.append(combined)
                    excel_records_output.add(excel_key)  # Track that this Excel record is now output
                    added_excel_count += 1
        
        if matched_in_second_pass > 0:
            self._synthesis_log(f"   ✅ Matched {matched_in_second_pass} Excel record(s) to PDF in second pass")
        
        if added_excel_count > 0:
            self._synthesis_log(f"   ✅ Added {added_excel_count} unmatched Excel record(s) to output")
        
        # Check which Excel records didn't get matched
        unmatched_excel_records = []
        for excel_record in self.excel_synthesis_data:
            name_code_prefix = excel_record.get('Name_Code_Prefix', '')
            excel_date_normalized = excel_record.get('Excel_Date_of_Service_Normalized', '')
            excel_procedure_code = normalize_proc_code(excel_record.get('Excel_Procedure_Code', ''))
            
            if name_code_prefix and excel_date_normalized and excel_procedure_code:
                excel_key = (name_code_prefix, excel_date_normalized, excel_procedure_code)
                if excel_key not in matched_excel_keys:
                    unmatched_excel_records.append(excel_record)
        
        self._synthesis_log(f"\n📊 Matching Summary:")
        self._synthesis_log(f"   ✅ Matched: {match_count} PDF record(s) to Excel")
        self._synthesis_log(f"   ⚠️  No Match: {no_match_count} PDF record(s) couldn't find Excel match")
        self._synthesis_log(f"   📋 Total PDF records: {len(self.synthesis_results)}")
        self._synthesis_log(f"   📋 Total Excel records: {len(self.excel_synthesis_data)}")
        self._synthesis_log(f"   ⚠️  Unmatched Excel records: {len(unmatched_excel_records)} (out of {len(self.excel_synthesis_data)})")
        
        # Show details about unmatched Excel records
        if unmatched_excel_records:
            self._synthesis_log(f"\n📊 Unmatched Excel Records Analysis (first 10):")
            for idx, excel_record in enumerate(unmatched_excel_records[:10], 1):
                excel_name = excel_record.get('Excel_Name', 'N/A')
                excel_name_code = excel_record.get('Name_Code_Prefix', 'N/A')
                excel_date = excel_record.get('Excel_Date_of_Service', 'N/A')
                excel_date_norm = excel_record.get('Excel_Date_of_Service_Normalized', 'N/A')
                excel_proc = excel_record.get('Excel_Procedure_Code', 'N/A')
                self._synthesis_log(f"   {idx}. Excel: {excel_name} (Code: {excel_name_code}, Date: {excel_date} [{excel_date_norm}], Proc: {excel_proc})")
            
            if len(unmatched_excel_records) > 10:
                self._synthesis_log(f"   ... and {len(unmatched_excel_records) - 10} more unmatched Excel records")
            
            # Check if there are PDF records that could match but didn't due to strict matching
            self._synthesis_log(f"\n💡 Checking if unmatched Excel records have potential PDF matches...")
            potential_matches_found = 0
            for excel_record in unmatched_excel_records[:5]:  # Check first 5
                excel_name_code = excel_record.get('Name_Code_Prefix', '')
                excel_date_norm = excel_record.get('Excel_Date_of_Service_Normalized', '')
                excel_proc = normalize_proc_code(excel_record.get('Excel_Procedure_Code', ''))
                
                # Look for PDF records with same name and date but different procedure code
                matching_pdfs = []
                for pdf_record in self.pdf_synthesis_data:
                    pdf_name_code = pdf_record.get('Name_Code_Prefix', '')
                    pdf_date_norm = self._normalize_date(pdf_record.get('Date_of_Service', ''))
                    pdf_proc = normalize_proc_code(pdf_record.get('Procedure Code', ''))
                    
                    if pdf_name_code == excel_name_code and pdf_date_norm == excel_date_norm:
                        if pdf_proc != excel_proc:
                            matching_pdfs.append(f"Proc: {pdf_proc}")
                
                if matching_pdfs:
                    unique_procs = list(set(matching_pdfs))[:3]  # Show first 3 unique
                    self._synthesis_log(f"   Excel {excel_name_code} ({excel_date_norm}): Found PDF with same name+date but different procedure codes: {', '.join(unique_procs)}")
                    potential_matches_found += 1
        
        # Analyze unmatched records to help diagnose issues
        if no_match_count > 0:
            self._synthesis_log(f"\n📊 Unmatched Records Analysis (first 10):")
            unmatched_count = 0
            for result in self.synthesis_results:
                if result.get('Match_Status') == 'No Match':
                    unmatched_count += 1
                    if unmatched_count <= 10:
                        pdf_name = result.get('PDF_Name_Code_Prefix', 'N/A')
                        pdf_date = result.get('PDF_Date_of_Service', 'N/A')
                        pdf_proc = result.get('PDF_Procedure_Code', 'N/A')
                        match_reason = result.get('Match_Method', 'N/A')
                        self._synthesis_log(f"   {unmatched_count}. Name: {pdf_name}, Date: {pdf_date}, Proc: {pdf_proc}")
                        self._synthesis_log(f"      Reason: {match_reason}")
            
            if no_match_count > 10:
                self._synthesis_log(f"   ... and {no_match_count - 10} more unmatched records")
            
            self._synthesis_log(f"\n💡 Troubleshooting Tips:")
            self._synthesis_log(f"   - Check if dates are in the correct format (MM/DD/YYYY)")
            self._synthesis_log(f"   - Verify procedure codes match between PDF and Excel")
            self._synthesis_log(f"   - Ensure name codes are being extracted correctly from both files")
            self._synthesis_log(f"   - Check the output Excel 'Match_Method' column for details on each match")
        
        # Debug: Show why multiple rows might exist
        if self.synthesis_results:
            # Count unique (name, date) combinations in results
            unique_combinations = set()
            for result in self.synthesis_results:
                name_code_prefix = result.get('PDF_Name_Code_Prefix', result.get('Name_Code', ''))
                pdf_date = result.get('PDF_Date_of_Service', result.get('PDF_Date_of_Service', ''))
                excel_name = result.get('Excel_Name', '')
                excel_date = result.get('Excel_Date_of_Service', '')
                if name_code_prefix and (pdf_date or excel_date):
                    key = f"{name_code_prefix}|{pdf_date or excel_date}|{excel_name}"
                    unique_combinations.add(key)
            
            self._synthesis_log(f"\n📊 Row Analysis:")
            self._synthesis_log(f"   Unique (Name Code, Date, Excel Name) combinations: {len(unique_combinations)}")
            self._synthesis_log(f"   Total PDF rows: {len(self.pdf_synthesis_data)}")
            self._synthesis_log(f"   Total Excel rows: {len(self.excel_synthesis_data)}")
            self._synthesis_log(f"   Total output rows: {len(self.synthesis_results)}")
            
            if len(self.synthesis_results) > len(unique_combinations):
                self._synthesis_log(f"\n   ℹ️  Note: Multiple PDF rows match the same Excel row.")
                self._synthesis_log(f"      This is expected if the PDF has multiple transactions/services")
                self._synthesis_log(f"      for the same client on the same date.")
                
                # Show examples of multiple matches
                example_counts = {}
                pdf_date_counts = {}  # Track PDF dates per Excel row
                for result in self.synthesis_results:
                    excel_name = result.get('Excel_Name', '')
                    excel_date = result.get('Excel_Date_of_Service', '')
                    pdf_date = result.get('PDF_Date_of_Service', '')
                    if excel_name and excel_date:
                        key = f"{excel_name}|{excel_date}"
                        example_counts[key] = example_counts.get(key, 0) + 1
                        if key not in pdf_date_counts:
                            pdf_date_counts[key] = []
                        if pdf_date:
                            pdf_date_counts[key].append(pdf_date)
                
                multi_match_examples = {k: v for k, v in example_counts.items() if v > 1}
                if multi_match_examples:
                    self._synthesis_log(f"\n   📋 Examples of Excel rows with multiple PDF matches:")
                    for i, (key, count) in enumerate(list(multi_match_examples.items())[:5], 1):
                        name, date = key.split('|', 1)
                        pdf_dates = set(pdf_date_counts.get(key, []))
                        self._synthesis_log(f"      {i}. Excel: {name} ({date})")
                        self._synthesis_log(f"         Matched {count} PDF row(s)")
                        if len(pdf_dates) == 1:
                            self._synthesis_log(f"         All PDF rows have date: {list(pdf_dates)[0]}")
                        elif len(pdf_dates) > 1:
                            self._synthesis_log(f"         PDF dates vary: {sorted(pdf_dates)[:3]}")
                        else:
                            self._synthesis_log(f"         PDF dates: (not extracted)")
                    
                    self._synthesis_log(f"\n   💡 Explanation:")
                    self._synthesis_log(f"      The PDF contains multiple rows (transactions/services)")
                    self._synthesis_log(f"      for the same client on the same date. Each PDF row")
                    self._synthesis_log(f"      represents a different transaction, but they all")
                    self._synthesis_log(f"      match to the same Excel row (one client session).")
                    self._synthesis_log(f"      This is expected behavior and preserves all PDF data.")
    
    def _generate_synthesis_output(self):
        """Generate output Excel file with synthesis results"""
        if not self.synthesis_results:
            raise Exception("No synthesis results to output")
        
        if not EXCEL_AVAILABLE:
            raise Exception("pandas is required. Install with: pip install pandas openpyxl")
        
        # Create DataFrame
        df = pd.DataFrame(self.synthesis_results)
        
        # Reorder columns to put important ones first - especially comparison columns
        priority_columns = [
            'Name_Code', 
            'Match_Status',
            'Session_Medium',
            'PDF_Modifier',
            'Expected_Modifier',
            'Needs_Refile',
            'Refile_Reason',
            'Match_Method',
            'Excel_Name',
            'PDF_Date_of_Service',
            'Excel_Date_of_Service',
            'PDF_Procedure_Code',
            'Excel_Procedure_Code',
            'Match_Key'
        ]
        
        # Get remaining columns
        other_columns = [col for col in df.columns if col not in priority_columns]
        
        # Reorder
        ordered_columns = [col for col in priority_columns if col in df.columns] + sorted(other_columns)
        df = df[ordered_columns]
        
        # Save to Excel
        df.to_excel(self.synthesis_output_path, index=False, engine='openpyxl')
        self._synthesis_log(f"   ✅ Generated output with {len(df)} row(s) and {len(df.columns)} column(s)")
    
    def run(self):
        """Run the bot GUI"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Missing Dependency", 
                "pandas is required. Install with: pip install pandas openpyxl")
            return
        
        if not PDFPLUMBER_AVAILABLE:
            # Show warning but allow GUI to open (PDF tab will show error if used)
            messagebox.showwarning("Missing Dependency", 
                "pdfplumber is not installed. The PDF/Excel Synthesis tab will not work.\n\n"
                "Install with: pip install pdfplumber")
        
        self.create_gui()
        self.root.mainloop()


def main():
    """Main entry point"""
    bot = MedicareModifierComparisonBot()
    bot.run()


if __name__ == "__main__":
    main()

