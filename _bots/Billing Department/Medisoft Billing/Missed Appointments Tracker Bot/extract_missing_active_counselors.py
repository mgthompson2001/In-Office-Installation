import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Set encoding to handle special characters
sys.stdout.reconfigure(encoding='utf-8')

master_file = r"G:\Company\Master Counselor\Master Counselor List.xlsx"
output_file1 = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\November Missed Appointment tracker part 1.xlsx"
output_file2 = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\Nov Missed Part 2.xlsx"
output_file3 = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\Log for remaining Run\Part 3 (Rows 109 through End)\part 3.xlsx"
output_excel = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\Missing Active Counselors - For Re-Analysis.xlsx"

def normalize_counselor_name(name):
    """Normalize counselor name to 'first last' format for comparison"""
    if pd.isna(name) or name == '' or str(name).lower() == 'nan':
        return None
    
    name_str = str(name).strip()
    
    # Remove "COUNSELOR:" prefix if present
    import re
    name_str = re.sub(r'^counselor:\s*', '', name_str, flags=re.IGNORECASE)
    name_str = name_str.strip()
    
    # Handle "Last, First" format
    if ',' in name_str:
        parts = [p.strip() for p in name_str.split(',')]
        if len(parts) >= 2:
            # Reverse to "First Last"
            name_str = ' '.join(reversed(parts))
    
    # Normalize whitespace and convert to lowercase
    name_str = ' '.join(name_str.split()).lower()
    return name_str

print("=" * 80)
print("EXTRACTING MISSING ACTIVE COUNSELORS")
print("=" * 80)

# Read Master Counselor List
print("\n1. Reading Master Counselor List...")
master_df = pd.read_excel(master_file)
print(f"   ✓ Master file loaded: {len(master_df)} rows, {len(master_df.columns)} columns")
print(f"   ✓ Columns: {list(master_df.columns)}")

# Filter out invalid rows (notes, instructions, etc.)
invalid_keywords = ['key', 'blue', 'green', 'orange', 'purple', 'red', 'white', 'yellow', 
                    'gray', 'bold', 'counselors must', 'supervision', 'mail', 'check']
master_df_clean = master_df[
    master_df['First Name'].notna() & 
    master_df['Last Name'].notna() &
    (master_df['First Name'].astype(str).str.strip() != 'nan') &
    (master_df['Last Name'].astype(str).str.strip() != 'nan') &
    (master_df['First Name'].astype(str).str.strip() != '') &
    (master_df['Last Name'].astype(str).str.strip() != '')
].copy()

name_combined = (master_df_clean['First Name'].astype(str).str.lower() + ' ' + 
                 master_df_clean['Last Name'].astype(str).str.lower())
is_valid = ~name_combined.str.contains('|'.join(invalid_keywords), case=False, na=False)
master_df_clean = master_df_clean[is_valid].copy()

print(f"   ✓ Valid counselors after filtering: {len(master_df_clean)}")

# Identify resigned (red) counselors
print("\n2. Identifying resigned (red) counselors...")
wb = load_workbook(master_file, data_only=True)
ws = wb.active

red_counselors = []
last_name_col = 1
first_name_col = 2

for row_idx in range(2, ws.max_row + 1):
    is_red = False
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.fill and cell.fill.start_color:
            fill_color = str(cell.fill.start_color.rgb).upper()
            if fill_color and ('FF0000' in fill_color or 'FFFF0000' in fill_color):
                is_red = True
                break
    
    if is_red:
        last_name = ws.cell(row=row_idx, column=last_name_col).value
        first_name = ws.cell(row=row_idx, column=first_name_col).value
        if pd.notna(last_name) and pd.notna(first_name) and str(last_name).strip() != 'nan' and str(first_name).strip() != 'nan':
            red_counselors.append((str(first_name).strip().lower(), str(last_name).strip().lower()))

red_set = set(red_counselors)
print(f"   ✓ Found {len(red_set)} resigned counselors")

# Identify LOA counselors
print("\n3. Identifying LOA counselors...")
loa_counselors = []
if 'Date of Term' in master_df_clean.columns:
    for idx, row in master_df_clean.iterrows():
        date_of_term = row.get('Date of Term')
        if pd.notna(date_of_term):
            date_str = str(date_of_term).upper()
            if 'LOA' in date_str or 'LEAVE' in date_str:
                first_name = str(row.get('First Name', '')).strip().lower()
                last_name = str(row.get('Last Name', '')).strip().lower()
                loa_counselors.append((first_name, last_name))

loa_set = set(loa_counselors)
print(f"   ✓ Found {len(loa_set)} LOA counselors")

all_resigned_or_loa = red_set | loa_set
print(f"   ✓ Total resigned OR LOA: {len(all_resigned_or_loa)}")

# Read all output files and collect counselor names
print("\n4. Reading output files and collecting counselor names...")
all_output_counselors = set()

# File 1
df1 = pd.read_excel(output_file1)
if 'Counselor Name' in df1.columns:
    counselors1_raw = df1['Counselor Name'].apply(normalize_counselor_name)
    counselors1 = {c for c in counselors1_raw if c is not None}
    all_output_counselors.update(counselors1)
    print(f"   ✓ Part 1: {len(counselors1)} unique counselors")

# File 2
df2 = pd.read_excel(output_file2)
if 'Counselor Name' in df2.columns:
    counselors2_raw = df2['Counselor Name'].apply(normalize_counselor_name)
    counselors2 = {c for c in counselors2_raw if c is not None}
    all_output_counselors.update(counselors2)
    print(f"   ✓ Part 2: {len(counselors2)} unique counselors")

# File 3
df3 = pd.read_excel(output_file3)
if 'Counselor Name' in df3.columns:
    counselors3_raw = df3['Counselor Name'].apply(normalize_counselor_name)
    counselors3 = {c for c in counselors3_raw if c is not None}
    all_output_counselors.update(counselors3)
    print(f"   ✓ Part 3: {len(counselors3)} unique counselors")

print(f"   ✓ Total unique counselors in output files: {len(all_output_counselors)}")

# Find active counselors who are missing
print("\n5. Identifying active counselors who are missing...")

# Create normalized names for master list
master_df_clean['Normalized_Name'] = (
    master_df_clean['First Name'].astype(str).str.strip().str.lower() + ' ' + 
    master_df_clean['Last Name'].astype(str).str.strip().str.lower()
).str.strip()

# Create normalized names for comparison
master_normalized_set = set(master_df_clean['Normalized_Name'].str.lower().str.strip())
missing_normalized = master_normalized_set - all_output_counselors

# Filter out resigned/LOA
missing_active = []
for idx, row in master_df_clean.iterrows():
    normalized_name = row['Normalized_Name'].lower().strip()
    first_name = str(row['First Name']).strip().lower()
    last_name = str(row['Last Name']).strip().lower()
    name_tuple = (first_name, last_name)
    
    # Check if missing from output
    if normalized_name in missing_normalized:
        # Check if NOT resigned/LOA
        if name_tuple not in all_resigned_or_loa:
            missing_active.append(idx)

missing_active_df = master_df_clean.loc[missing_active].copy()

print(f"   ✓ Found {len(missing_active_df)} active counselors who are missing")

# Verify no overlap with output files
print("\n6. Verifying no overlap with existing output files...")
missing_active_normalized = set(missing_active_df['Normalized_Name'].str.lower().str.strip())

overlap_with_output = missing_active_normalized & all_output_counselors
if overlap_with_output:
    print(f"   ⚠ WARNING: Found {len(overlap_with_output)} counselors in missing list who ARE in output files!")
    print(f"   Overlapping names: {sorted(list(overlap_with_output))[:10]}")
else:
    print(f"   ✓ CONFIRMED: No overlap with existing output files")

# Sort by Last Name, then First Name
missing_active_df = missing_active_df.sort_values(['Last Name', 'First Name'])

# Create Excel file with exact same format as master list
print("\n7. Creating Excel file with exact same format as master list...")

# Use all columns from master list
output_columns = list(master_df.columns)
missing_active_excel = missing_active_df[output_columns].copy()

# Save to Excel with same formatting
with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    missing_active_excel.to_excel(writer, sheet_name='ISWS Counselors', index=False)
    
    # Get the worksheet
    worksheet = writer.sheets['ISWS Counselors']
    
    # Auto-adjust column widths
    for idx, col in enumerate(missing_active_excel.columns, 1):
        max_length = max(
            missing_active_excel[col].astype(str).map(len).max(),
            len(str(col))
        )
        col_letter = get_column_letter(idx)
        worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

print(f"   ✓ Excel file saved: {output_excel}")
print(f"   ✓ Contains {len(missing_active_excel)} active counselors")
print(f"   ✓ Same format as master list with {len(output_columns)} columns")

# Print summary
print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"\n📊 COUNTS:")
print(f"   • Active counselors missing from output: {len(missing_active_excel)}")
print(f"   • Verified NO overlap with existing 3 output files")
print(f"   • Excel file ready for re-analysis: {output_excel}")

print(f"\n📋 FIRST 10 COUNSELORS IN THE LIST:")
for idx, row in missing_active_excel.head(10).iterrows():
    print(f"   {idx + 1}. {row['First Name']} {row['Last Name']} - {row.get('Email', 'N/A')}")

print("\n" + "=" * 80)
print("EXTRACTION COMPLETE")
print("=" * 80)

