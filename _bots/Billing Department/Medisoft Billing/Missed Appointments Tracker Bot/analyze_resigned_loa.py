import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Set encoding to handle special characters
sys.stdout.reconfigure(encoding='utf-8')

master_file = r"G:\Company\Master Counselor\Master Counselor List.xlsx"

print("=" * 80)
print("ANALYZING RESIGNED AND LOA COUNSELORS")
print("=" * 80)

# Read the Excel file with openpyxl to check cell colors
print("\n1. Reading Master Counselor List with formatting information...")
wb = load_workbook(master_file, data_only=True)
ws = wb.active

# Read with pandas for data
master_df = pd.read_excel(master_file)

print(f"   ✓ Master file loaded: {len(master_df)} rows")
print(f"   ✓ Sheet name: {ws.title}")
print(f"   ✓ Columns: {list(master_df.columns)[:5]}...")

# Find the row and column indices
# Assuming First Name is in column B (index 2) and Last Name is in column A (index 1)
# We need to find which columns contain First Name and Last Name
first_name_col = None
last_name_col = None
date_of_term_col = None

for idx, col_name in enumerate(master_df.columns, 1):
    if 'first name' in col_name.lower():
        first_name_col = idx
    if 'last name' in col_name.lower():
        last_name_col = idx
    if 'date of term' in col_name.lower():
        date_of_term_col = idx

print(f"\n2. Column positions:")
print(f"   • First Name column: {first_name_col}")
print(f"   • Last Name column: {last_name_col}")
print(f"   • Date of Term column: {date_of_term_col}")

# Check for red highlighting (resigned counselors)
print("\n3. Checking for red highlighting (resigned counselors)...")
red_counselors = []
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # Red
red_fill_alt = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Alternative red

# Check each row (skip header row, which is row 1 in Excel, but pandas uses 0-indexed)
for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
    # Check if the row has red fill in any cell
    is_red = False
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.fill and cell.fill.start_color:
            # Check if it's red (RGB: FF0000 or similar)
            fill_color = str(cell.fill.start_color.rgb).upper()
            if fill_color and ('FF0000' in fill_color or 'FFFF0000' in fill_color):
                is_red = True
                break
    
    if is_red:
        # Get counselor name
        if last_name_col and first_name_col:
            last_name = ws.cell(row=row_idx, column=last_name_col).value
            first_name = ws.cell(row=row_idx, column=first_name_col).value
            if pd.notna(last_name) and pd.notna(first_name) and str(last_name).strip() != 'nan' and str(first_name).strip() != 'nan':
                red_counselors.append({
                    'First Name': str(first_name).strip(),
                    'Last Name': str(last_name).strip(),
                    'Row': row_idx
                })

print(f"   ✓ Found {len(red_counselors)} counselors with red highlighting")

# Check for LOA in Date of Term column
print("\n4. Checking for Leave of Absence (LOA) indicators...")
if 'Date of Term' in master_df.columns:
    loa_counselors = []
    for idx, row in master_df.iterrows():
        date_of_term = row.get('Date of Term')
        if pd.notna(date_of_term):
            date_str = str(date_of_term).upper()
            if 'LOA' in date_str or 'LEAVE' in date_str:
                loa_counselors.append({
                    'First Name': str(row.get('First Name', '')).strip(),
                    'Last Name': str(row.get('Last Name', '')).strip(),
                    'Date of Term': date_str,
                    'Row': idx + 2  # +2 because Excel is 1-indexed and has header
                })
    
    print(f"   ✓ Found {len(loa_counselors)} counselors with LOA indicators")
else:
    loa_counselors = []
    print("   ⚠ Date of Term column not found")

# Combine red and LOA counselors (remove duplicates)
print("\n5. Combining resigned (red) and LOA counselors...")

# Create sets for comparison
red_set = {(c['First Name'].lower().strip(), c['Last Name'].lower().strip()) for c in red_counselors}
loa_set = {(c['First Name'].lower().strip(), c['Last Name'].lower().strip()) for c in loa_counselors}

# Combine unique counselors
all_resigned_or_loa = red_set | loa_set
resigned_and_loa = red_set & loa_set

print(f"   • Red (resigned): {len(red_set)} unique counselors")
print(f"   • LOA: {len(loa_set)} unique counselors")
print(f"   • Both red AND LOA: {len(resigned_and_loa)} counselors")
print(f"   • TOTAL unique (red OR LOA): {len(all_resigned_or_loa)} counselors")

# Now compare with missing counselors
print("\n6. Comparing with missing counselors from output files...")

# Read the missing counselors report we generated
missing_file = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\Missing Counselors Report.xlsx"
try:
    missing_df = pd.read_excel(missing_file)
    missing_set = {(str(row['First Name']).lower().strip(), str(row['Last Name']).lower().strip()) 
                    for _, row in missing_df.iterrows() 
                    if pd.notna(row['First Name']) and pd.notna(row['Last Name'])}
    print(f"   ✓ Loaded {len(missing_set)} missing counselors from report")
except Exception as e:
    print(f"   ⚠ Could not load missing counselors file: {e}")
    missing_set = set()

# Find discrepancies
if missing_set:
    missing_and_resigned_loa = missing_set & all_resigned_or_loa
    missing_but_not_resigned_loa = missing_set - all_resigned_or_loa
    resigned_loa_but_not_missing = all_resigned_or_loa - missing_set
    
    print(f"\n7. DISCREPANCY ANALYSIS:")
    print(f"   • Missing counselors who ARE resigned/LOA: {len(missing_and_resigned_loa)}")
    print(f"   • Missing counselors who are NOT resigned/LOA: {len(missing_but_not_resigned_loa)}")
    print(f"   • Resigned/LOA counselors who are NOT missing: {len(resigned_loa_but_not_missing)}")
    
    if missing_but_not_resigned_loa:
        print(f"\n   ⚠ PROBLEM: {len(missing_but_not_resigned_loa)} missing counselors are NOT resigned/LOA:")
        for name in sorted(list(missing_but_not_resigned_loa))[:20]:
            print(f"     - {name[0].title()} {name[1].title()}")
        if len(missing_but_not_resigned_loa) > 20:
            print(f"     ... and {len(missing_but_not_resigned_loa) - 20} more")
    
    if resigned_loa_but_not_missing:
        print(f"\n   ✓ GOOD: {len(resigned_loa_but_not_missing)} resigned/LOA counselors correctly excluded from output")

# Print detailed breakdown
print("\n" + "=" * 80)
print("DETAILED BREAKDOWN")
print("=" * 80)

if red_counselors:
    print(f"\nRED (RESIGNED) COUNSELORS ({len(red_counselors)}):")
    for i, c in enumerate(red_counselors[:10], 1):
        print(f"   {i}. {c['First Name']} {c['Last Name']}")
    if len(red_counselors) > 10:
        print(f"   ... and {len(red_counselors) - 10} more")

if loa_counselors:
    print(f"\nLOA COUNSELORS ({len(loa_counselors)}):")
    for i, c in enumerate(loa_counselors[:10], 1):
        print(f"   {i}. {c['First Name']} {c['Last Name']} - {c['Date of Term']}")
    if len(loa_counselors) > 10:
        print(f"   ... and {len(loa_counselors) - 10} more")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)

