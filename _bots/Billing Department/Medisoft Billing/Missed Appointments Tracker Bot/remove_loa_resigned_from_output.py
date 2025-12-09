#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to remove LOA and Resigned counselors from the merged output Excel file.
Preserves formatting and structure.
"""

import pandas as pd
import os
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

# File paths
master_file = r"G:\Company\Master Counselor\Master Counselor List.xlsx"
input_file = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\The stuff\Final Merged - Complete.xlsx"
output_file = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\The stuff\Final Merged - Complete (Active Only).xlsx"

print("="*80)
print("REMOVING LOA AND RESIGNED COUNSELORS FROM OUTPUT")
print("="*80)

# Read Master Counselor List to identify LOA/Resigned
print("\n1. Reading Master Counselor List...")
master_df = pd.read_excel(master_file)

# Build set of LOA and Resigned counselor names
loa_counselors = set()
resigned_counselors = set()

if 'Last Name' in master_df.columns and 'First Name' in master_df.columns:
    for idx, row in master_df.iterrows():
        last_name = str(row.get('Last Name', '')).strip() if pd.notna(row.get('Last Name')) else ""
        first_name = str(row.get('First Name', '')).strip() if pd.notna(row.get('First Name')) else ""
        
        if not last_name or not first_name or last_name.lower() == 'nan' or first_name.lower() == 'nan':
            continue
        
        # Skip KEY rows
        if any(keyword in last_name.lower() for keyword in ['key', 'all counselors', 'minimum', 'cases']):
            continue
        if any(keyword in first_name.lower() for keyword in ['key', 'all counselors', 'minimum', 'cases']):
            continue
        
        counselor_name = f"{last_name}, {first_name}"
        
        # Check Notes for LOA/resigned
        notes_text = str(row.get('Notes', '')).strip() if pd.notna(row.get('Notes')) else ""
        notes_lower = notes_text.lower() if notes_text else ""
        
        # Check for LOA
        is_loa = False
        if notes_lower:
            loa_keywords = ['loa', 'leave of absence', 'on leave']
            if any(keyword in notes_lower for keyword in loa_keywords):
                if 'returned from' not in notes_lower and 'has returned' not in notes_lower and 'returning from' not in notes_lower:
                    is_loa = True
        
        # Check for Resigned
        is_resigned = False
        if notes_lower:
            resigned_keywords = ['resigned', 'resigning', 'resignation', 'will be resigning']
            if any(keyword in notes_lower for keyword in resigned_keywords):
                if 'rescinded' not in notes_lower:
                    is_resigned = True
        
        if is_loa:
            loa_counselors.add(counselor_name)
        if is_resigned:
            resigned_counselors.add(counselor_name)

print(f"   [OK] Found {len(loa_counselors)} LOA counselors in Master List")
print(f"   [OK] Found {len(resigned_counselors)} Resigned counselors in Master List")

# Normalize names for comparison
def normalize_name(name):
    name = str(name).strip().lower()
    name = ' '.join(name.split())
    if ',' in name:
        parts = [p.strip() for p in name.split(',')]
        if len(parts) >= 2:
            return f"{parts[0]}, {parts[1]}"
    return name

# Create normalized sets
loa_normalized = {normalize_name(name): name for name in loa_counselors}
resigned_normalized = {normalize_name(name): name for name in resigned_counselors}

# Read input Excel with formatting
print("\n2. Reading merged output Excel file...")
wb = load_workbook(input_file, data_only=False)
ws = wb.active

# Get header row
headers = []
for cell in ws[1]:
    headers.append(cell.value)

print(f"   [OK] Found {len(headers)} columns")
print(f"   [OK] Total rows in input: {ws.max_row}")

# Process rows and identify which to keep
print("\n3. Identifying rows to remove...")
rows_to_remove = set()
counselor_header_rows = {}  # row_num -> counselor_name
counselors_removed = set()  # Track which counselors we're removing

# Find Counselor Name column index
counselor_name_col = None
for col_idx, header in enumerate(headers, 1):
    if header and str(header).strip().lower() == 'counselor name':
        counselor_name_col = col_idx
        break

if not counselor_name_col:
    print("   [ERROR] Could not find 'Counselor Name' column!")
    sys.exit(1)

print(f"   [OK] Counselor Name column: {counselor_name_col}")

# First pass: check ALL rows in Counselor Name column for LOA/Resigned counselors
print("   Checking all rows in Counselor Name column...")
for row_num in range(2, ws.max_row + 1):
    # Check Counselor Name column for this row
    if counselor_name_col:
        counselor_name_cell = ws.cell(row=row_num, column=counselor_name_col).value
        if counselor_name_cell:
            counselor_name_str = str(counselor_name_cell).strip()
            if counselor_name_str and counselor_name_str.lower() != 'nan':
                # Check if this counselor should be removed
                counselor_norm = normalize_name(counselor_name_str)
                
                if counselor_norm in loa_normalized or counselor_norm in resigned_normalized:
                    # This is a row for a LOA/Resigned counselor - remove it
                    rows_to_remove.add(row_num)
                    if counselor_name_str not in counselors_removed:
                        counselors_removed.add(counselor_name_str)
                        status = "LOA" if counselor_norm in loa_normalized else "Resigned"
                        print(f"   [REMOVE] {status}: {counselor_name_str} (row {row_num})")
    
    # Also check if this is a counselor header row (starts with "COUNSELOR:")
    first_cell_value = ws.cell(row=row_num, column=1).value
    first_cell_str = str(first_cell_value).strip() if first_cell_value else ""
    
    if first_cell_str.upper().startswith("COUNSELOR:"):
        # Extract counselor name
        counselor_name = first_cell_str.replace("COUNSELOR:", "").strip()
        counselor_header_rows[row_num] = counselor_name
        
        # Check if this counselor should be removed (if not already marked)
        counselor_norm = normalize_name(counselor_name)
        should_remove = False
        status = ""
        
        if counselor_norm in loa_normalized:
            should_remove = True
            status = "LOA"
        elif counselor_norm in resigned_normalized:
            should_remove = True
            status = "Resigned"
        
        if should_remove:
            rows_to_remove.add(row_num)
            if counselor_name not in counselors_removed:
                counselors_removed.add(counselor_name)
                print(f"   [REMOVE] {status}: {counselor_name} (header row {row_num})")

print(f"\n   [OK] Identified {len(rows_to_remove)} rows to remove")

# Create new workbook with filtered data
print("\n4. Creating cleaned Excel file...")
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = "Missed Appointments"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
counselor_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
counselor_font = Font(bold=True)

# Copy header row
for col_idx, header in enumerate(headers, 1):
    cell = new_ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='left', vertical='center')

# Copy rows that should be kept
new_row = 2
rows_kept = 0
rows_removed = 0

for row_num in range(2, ws.max_row + 1):
    if row_num not in rows_to_remove:
        # Copy this row
        is_counselor_header = row_num in counselor_header_rows
        
        for col_idx in range(1, len(headers) + 1):
            source_cell = ws.cell(row=row_num, column=col_idx)
            target_cell = new_ws.cell(row=new_row, column=col_idx, value=source_cell.value)
            
            # Apply counselor header formatting if this is a counselor header row
            if is_counselor_header:
                target_cell.fill = counselor_fill
                target_cell.font = counselor_font
                target_cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                # For client rows: ensure white/transparent background (no fill)
                # Don't copy fill from source - client rows should have no background color
                target_cell.fill = PatternFill()  # No fill (transparent/white)
                
                # Copy font (but ensure text is black/visible)
                try:
                    if source_cell.font:
                        # Get font color, but default to black if it's black or None
                        font_color = source_cell.font.color
                        if font_color and hasattr(font_color, 'rgb'):
                            color_rgb = str(font_color.rgb)
                            # If font is black or very dark, keep it; otherwise use black for visibility
                            if color_rgb and (color_rgb.upper() in ['FF000000', '00000000', '000000', None]):
                                font_color_val = '000000'  # Black
                            else:
                                font_color_val = color_rgb if color_rgb else '000000'
                        else:
                            font_color_val = '000000'  # Default to black
                        
                        target_cell.font = Font(
                            name=source_cell.font.name if source_cell.font.name else 'Calibri',
                            size=source_cell.font.size if source_cell.font.size else 11,
                            bold=source_cell.font.bold if source_cell.font.bold else False,
                            italic=source_cell.font.italic if source_cell.font.italic else False,
                            color=font_color_val
                        )
                    else:
                        # Default font if source has no font
                        target_cell.font = Font(name='Calibri', size=11, color='000000')
                except:
                    # Default font on error
                    target_cell.font = Font(name='Calibri', size=11, color='000000')
                
                # Set alignment
                target_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        new_row += 1
        rows_kept += 1
    else:
        rows_removed += 1

# Auto-adjust column widths
print("\n5. Adjusting column widths...")
for col_idx, header in enumerate(headers, 1):
    max_length = len(str(header))
    col_letter = get_column_letter(col_idx)
    
    for row in new_ws[col_letter]:
        try:
            if row.value:
                max_length = max(max_length, len(str(row.value)))
        except:
            pass
    
    new_ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

# Save file
print("\n6. Saving cleaned Excel file...")
try:
    new_wb.save(output_file)
    print(f"   [OK] Saved to: {output_file}")
except Exception as e:
    print(f"   [ERROR] Error saving file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# Summary
print("\n" + "="*80)
print("CLEANUP SUMMARY")
print("="*80)
print(f"\nInput file: {os.path.basename(input_file)}")
print(f"Output file: {os.path.basename(output_file)}")
print(f"\nRows in input: {ws.max_row}")
print(f"Rows kept: {rows_kept}")
print(f"Rows removed: {rows_removed}")
print(f"Rows in output: {new_row - 1}")

# Count counselors removed
loa_removed = sum(1 for name in counselor_header_rows.values() if normalize_name(name) in loa_normalized)
resigned_removed = sum(1 for name in counselor_header_rows.values() if normalize_name(name) in resigned_normalized)

print(f"\nCounselors removed:")
print(f"  - LOA: {loa_removed}")
print(f"  - Resigned: {resigned_removed}")
print(f"  - Total: {loa_removed + resigned_removed}")

print(f"\n✅ CLEANUP COMPLETE - File is ready for email sending!")
print(f"✅ All LOA and Resigned counselors have been removed")
print("="*80)

