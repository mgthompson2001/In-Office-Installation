import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Set encoding to handle special characters
sys.stdout.reconfigure(encoding='utf-8')

master_file = r"G:\Company\Master Counselor\Master Counselor List.xlsx"
missing_file = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\Missing Counselors Report.xlsx"
output_report = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\Discrepancy Report - Active Counselors Missing.xlsx"

print("=" * 80)
print("GENERATING DISCREPANCY REPORT")
print("=" * 80)

# Read Master Counselor List
print("\n1. Reading Master Counselor List...")
wb = load_workbook(master_file, data_only=True)
ws = wb.active
master_df = pd.read_excel(master_file)

# Filter out invalid rows
invalid_keywords = ['key', 'blue', 'green', 'orange', 'purple', 'red', 'white', 'yellow', 
                    'gray', 'bold', 'counselors must', 'supervision', 'mail', 'check']
master_df = master_df[
    master_df['First Name'].notna() & 
    master_df['Last Name'].notna() &
    (master_df['First Name'].astype(str).str.strip() != 'nan') &
    (master_df['Last Name'].astype(str).str.strip() != 'nan') &
    (master_df['First Name'].astype(str).str.strip() != '') &
    (master_df['Last Name'].astype(str).str.strip() != '')
].copy()

name_combined = (master_df['First Name'].astype(str).str.lower() + ' ' + 
                 master_df['Last Name'].astype(str).str.lower())
is_valid = ~name_combined.str.contains('|'.join(invalid_keywords), case=False, na=False)
master_df = master_df[is_valid].copy()

print(f"   ✓ Valid counselors: {len(master_df)}")

# Find red (resigned) counselors
print("\n2. Identifying resigned (red) counselors...")
red_counselors = []
last_name_col = 1
first_name_col = 2

for row_idx in range(2, ws.max_row + 1):
    is_red = False
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.fill and cell.fill.start_color:
            fill_color = str(cell.fill.start_color.rgb).upper()
            if fill_color and ('FF0000' in fill_color or 'FFFF0000' in fill_color):
                is_red = True
                break
    
    if is_red:
        last_name = ws.cell(row=row_idx, column=last_name_col).value
        first_name = ws.cell(row=row_idx, column=first_name_col).value
        if pd.notna(last_name) and pd.notna(first_name) and str(last_name).strip() != 'nan' and str(first_name).strip() != 'nan':
            red_counselors.append((str(first_name).strip().lower(), str(last_name).strip().lower()))

red_set = set(red_counselors)
print(f"   ✓ Found {len(red_set)} resigned counselors")

# Find LOA counselors
print("\n3. Identifying LOA counselors...")
loa_counselors = []
if 'Date of Term' in master_df.columns:
    for idx, row in master_df.iterrows():
        date_of_term = row.get('Date of Term')
        if pd.notna(date_of_term):
            date_str = str(date_of_term).upper()
            if 'LOA' in date_str or 'LEAVE' in date_str:
                first_name = str(row.get('First Name', '')).strip().lower()
                last_name = str(row.get('Last Name', '')).strip().lower()
                loa_counselors.append((first_name, last_name))

loa_set = set(loa_counselors)
print(f"   ✓ Found {len(loa_set)} LOA counselors")

# Combine resigned and LOA
all_resigned_or_loa = red_set | loa_set
print(f"   ✓ Total resigned OR LOA: {len(all_resigned_or_loa)}")

# Read missing counselors
print("\n4. Reading missing counselors report...")
missing_df = pd.read_excel(missing_file)
missing_set = {(str(row['First Name']).lower().strip(), str(row['Last Name']).lower().strip()) 
                for _, row in missing_df.iterrows() 
                if pd.notna(row['First Name']) and pd.notna(row['Last Name'])}
print(f"   ✓ Found {len(missing_set)} missing counselors")

# Categorize missing counselors
print("\n5. Categorizing missing counselors...")
missing_and_resigned_loa = missing_set & all_resigned_or_loa
missing_but_active = missing_set - all_resigned_or_loa

print(f"   • Missing counselors who ARE resigned/LOA: {len(missing_and_resigned_loa)}")
print(f"   • Missing counselors who are ACTIVE (PROBLEM): {len(missing_but_active)}")

# Create detailed report for active counselors who are missing
print("\n6. Creating Excel report for active counselors who are missing...")

# Get full records for active missing counselors
active_missing_records = []
for name_tuple in missing_but_active:
    first_name, last_name = name_tuple
    matching_rows = master_df[
        (master_df['First Name'].astype(str).str.strip().str.lower() == first_name) &
        (master_df['Last Name'].astype(str).str.strip().str.lower() == last_name)
    ]
    if not matching_rows.empty:
        active_missing_records.append(matching_rows.iloc[0])

if active_missing_records:
    active_missing_df = pd.DataFrame(active_missing_records)
    
    # Select relevant columns
    excel_columns = [
        'First Name', 
        'Last Name', 
        'Email',
        'Program? (In-Home, Telehealth, FUH, Sup, IA Only)',
        'date of hire',
        'Date of Term',
        'Phone #',
        'Sup',
        'Lic #',
        'NPI'
    ]
    
    available_columns = [col for col in excel_columns if col in active_missing_df.columns]
    active_missing_excel = active_missing_df[available_columns].copy()
    
    # Sort by Last Name, then First Name
    active_missing_excel = active_missing_excel.sort_values(['Last Name', 'First Name'])
    
    # Add status column
    active_missing_excel.insert(0, 'Status', 'ACTIVE - MISSING FROM OUTPUT')
    
    # Save to Excel
    with pd.ExcelWriter(output_report, engine='openpyxl') as writer:
        active_missing_excel.to_excel(writer, sheet_name='Active Missing Counselors', index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Active Missing Counselors']
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(active_missing_excel.columns, 1):
            max_length = max(
                active_missing_excel[col].astype(str).map(len).max(),
                len(str(col))
            )
            col_letter = get_column_letter(idx)
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    print(f"   ✓ Excel report saved: {output_report}")
    print(f"   ✓ Contains {len(active_missing_excel)} active counselors who are missing")

# Print summary
print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"\n📊 COUNTS:")
print(f"   • Total counselors in Master List: {len(master_df)}")
print(f"   • Resigned (red): {len(red_set)}")
print(f"   • On LOA: {len(loa_set)}")
print(f"   • Total resigned OR LOA: {len(all_resigned_or_loa)}")
print(f"   • Missing from output files: {len(missing_set)}")
print(f"\n🔍 DISCREPANCY:")
print(f"   • Missing counselors who ARE resigned/LOA (acceptable): {len(missing_and_resigned_loa)}")
print(f"   • Missing counselors who are ACTIVE (PROBLEM): {len(missing_but_active)}")
print(f"\n⚠️  CRITICAL ISSUE:")
print(f"   {len(missing_but_active)} active counselors are missing from the output files!")
print(f"   These counselors should have been included in the missed appointments analysis.")
print(f"\n📄 Report saved to: {output_report}")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)

