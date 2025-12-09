#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fix black client rows in the merged Excel file.
"""

import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding='utf-8')

input_file = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\The stuff\Final Merged - Complete (Active Only).xlsx"

print("="*80)
print("FIXING BLACK CLIENT ROWS IN MERGED FILE")
print("="*80)

print("\n1. Loading Excel file...")
wb = load_workbook(input_file, data_only=False)
ws = wb.active

print(f"   [OK] Total rows: {ws.max_row}")

# Define styles
counselor_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
counselor_font = Font(bold=True)
client_fill = PatternFill()  # No fill (white background)
client_font = Font(name='Calibri', size=11, color='000000')  # Black text

print("\n2. Fixing cell formatting...")
rows_fixed = 0

for row_num in range(2, ws.max_row + 1):
    first_cell_value = ws.cell(row=row_num, column=1).value
    first_cell_str = str(first_cell_value).strip() if first_cell_value else ""
    
    # Check if this is a counselor header row
    is_counselor_header = first_cell_str.upper().startswith("COUNSELOR:")
    
    # Fix all cells in this row
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        
        if is_counselor_header:
            # Counselor header row - light blue background, bold
            cell.fill = counselor_fill
            cell.font = counselor_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            # Client row - white background, black text
            cell.fill = client_fill
            cell.font = client_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            rows_fixed += 1

print(f"   [OK] Fixed {rows_fixed} client row cells")

print("\n3. Saving fixed file...")
try:
    wb.save(input_file)
    print(f"   [OK] File saved: {input_file}")
except Exception as e:
    print(f"   [ERROR] Error saving file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print("\n" + "="*80)
print("✅ FIX COMPLETE - All client rows now have white background and black text!")
print("="*80)

