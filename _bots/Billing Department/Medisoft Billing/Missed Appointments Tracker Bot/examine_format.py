import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

file1 = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\November Missed Appointment tracker part 1.xlsx"

wb = load_workbook(file1)
ws = wb.active

print("Examining format...")
print(f"Sheet: {ws.title}")
print(f"Total rows: {ws.max_row}")
print("\nFirst 30 rows:")

for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=False), 1):
    first_cell = row[0]
    value = first_cell.value
    is_bold = first_cell.font.bold if first_cell.font else False
    fill_color = None
    if first_cell.fill and first_cell.fill.start_color:
        fill_color = str(first_cell.fill.start_color.rgb).upper()
    
    print(f"Row {i}: Value='{value}', Bold={is_bold}, Fill={fill_color}")

