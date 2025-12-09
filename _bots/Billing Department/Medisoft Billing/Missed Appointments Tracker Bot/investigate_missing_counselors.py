import pandas as pd
import sys
from openpyxl import load_workbook
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

master_file = r"G:\Company\Master Counselor\Master Counselor List.xlsx"
missing_active_file = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\Missing Active Counselors - For Re-Analysis.xlsx"

print("=" * 80)
print("INVESTIGATING WHY COUNSELORS WERE SKIPPED")
print("=" * 80)

# Read Master Counselor List
print("\n1. Reading Master Counselor List...")
master_df = pd.read_excel(master_file, sheet_name='ISWS Counselors')

# Read missing active counselors
print("\n2. Reading missing active counselors...")
missing_df = pd.read_excel(missing_active_file)

print(f"   ✓ Found {len(missing_df)} missing active counselors")

# Create a comprehensive analysis
print("\n3. Analyzing each missing counselor...")

analysis_results = []

for idx, missing_row in missing_df.iterrows():
    first_name = str(missing_row['First Name']).strip()
    last_name = str(missing_row['Last Name']).strip()
    
    # Find matching row in master list
    master_match = master_df[
        (master_df['First Name'].astype(str).str.strip() == first_name) &
        (master_df['Last Name'].astype(str).str.strip() == last_name)
    ]
    
    if not master_match.empty:
        master_row = master_match.iloc[0]
        
        # Check various potential reasons for skipping
        reasons = []
        
        # Check Date of Term
        date_of_term = master_row.get('Date of Term')
        if pd.notna(date_of_term):
            date_str = str(date_of_term).upper()
            if 'LOA' in date_str or 'LEAVE' in date_str:
                reasons.append(f"LOA in Date of Term: {date_str}")
            elif any(char.isdigit() for char in date_str):
                # Has digits - might be a date
                reasons.append(f"Date in Date of Term: {date_str}")
        
        # Check Notes
        notes = master_row.get('Notes', '')
        if pd.notna(notes):
            notes_str = str(notes).upper()
            if 'LOA' in notes_str or 'LEAVE' in notes_str:
                reasons.append(f"LOA mentioned in Notes")
            if 'RESIGN' in notes_str:
                reasons.append(f"Resign mentioned in Notes")
        
        # Check Email
        email = master_row.get('Email', '')
        if pd.isna(email) or str(email).strip() == '':
            reasons.append("NO EMAIL ADDRESS")
        
        # Check Program
        program = master_row.get('Program? (In-Home, Telehealth, FUH, Sup, IA Only)', '')
        if pd.notna(program):
            program_str = str(program).upper()
            if 'SUP' in program_str and 'IA' not in program_str:
                reasons.append(f"Program is Supervisor only: {program}")
            if 'IA ONLY' in program_str:
                reasons.append(f"Program is IA Only: {program}")
        
        analysis_results.append({
            'Counselor': f"{first_name} {last_name}",
            'Email': email if pd.notna(email) else 'MISSING',
            'Program': program if pd.notna(program) else 'N/A',
            'Date of Term': date_of_term if pd.notna(date_of_term) else 'N/A',
            'Notes': str(notes)[:100] if pd.notna(notes) else 'N/A',
            'Potential Reasons': '; '.join(reasons) if reasons else 'UNKNOWN - Need to check Therapy Notes lookup',
            'Row Number': master_match.index[0] + 2  # Excel row number
        })
    else:
        analysis_results.append({
            'Counselor': f"{first_name} {last_name}",
            'Email': 'NOT FOUND IN MASTER',
            'Program': 'N/A',
            'Date of Term': 'N/A',
            'Notes': 'N/A',
            'Potential Reasons': 'NOT FOUND IN MASTER LIST',
            'Row Number': 'N/A'
        })

# Create analysis DataFrame
analysis_df = pd.DataFrame(analysis_results)

# Count reasons
print("\n4. Summary of potential reasons:")
no_email_count = len(analysis_df[analysis_df['Email'] == 'MISSING'])
loa_in_date_count = len(analysis_df[analysis_df['Potential Reasons'].str.contains('LOA in Date of Term', na=False)])
loa_in_notes_count = len(analysis_df[analysis_df['Potential Reasons'].str.contains('LOA mentioned in Notes', na=False)])
resign_in_notes_count = len(analysis_df[analysis_df['Potential Reasons'].str.contains('Resign mentioned in Notes', na=False)])
sup_only_count = len(analysis_df[analysis_df['Potential Reasons'].str.contains('Supervisor only', na=False)])
ia_only_count = len(analysis_df[analysis_df['Potential Reasons'].str.contains('IA Only', na=False)])
unknown_count = len(analysis_df[analysis_df['Potential Reasons'].str.contains('UNKNOWN', na=False)])

print(f"   • No email address: {no_email_count}")
print(f"   • LOA in Date of Term: {loa_in_date_count}")
print(f"   • LOA in Notes: {loa_in_notes_count}")
print(f"   • Resign in Notes: {resign_in_notes_count}")
print(f"   • Supervisor only program: {sup_only_count}")
print(f"   • IA Only program: {ia_only_count}")
print(f"   • Unknown (need to check Therapy Notes lookup): {unknown_count}")

# Save detailed analysis
output_file = r"C:\Users\mthompson\OneDrive - Integrity Senior Services\Desktop\Missed Appt. Tracker Output\Investigation - Why Counselors Were Skipped.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    analysis_df.to_excel(writer, sheet_name='Investigation Results', index=False)
    
    # Auto-adjust column widths
    worksheet = writer.sheets['Investigation Results']
    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(analysis_df.columns, 1):
        max_length = max(
            analysis_df[col].astype(str).map(len).max(),
            len(str(col))
        )
        col_letter = get_column_letter(idx)
        worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

print(f"\n5. Detailed analysis saved to: {output_file}")

# Show sample of counselors with no email
print("\n6. Sample counselors with NO EMAIL (likely reason for skipping):")
no_email_counselors = analysis_df[analysis_df['Email'] == 'MISSING']
if len(no_email_counselors) > 0:
    for idx, row in no_email_counselors.head(10).iterrows():
        print(f"   • {row['Counselor']} - Program: {row['Program']}")
else:
    print("   ✓ All counselors have email addresses")

# Show sample of unknown reasons
print("\n7. Sample counselors with UNKNOWN reasons (need Therapy Notes investigation):")
unknown_counselors = analysis_df[analysis_df['Potential Reasons'].str.contains('UNKNOWN', na=False)]
if len(unknown_counselors) > 0:
    for idx, row in unknown_counselors.head(10).iterrows():
        print(f"   • {row['Counselor']} - Email: {row['Email']} - Program: {row['Program']}")
else:
    print("   ✓ All counselors have identifiable reasons")

print("\n" + "=" * 80)
print("INVESTIGATION COMPLETE")
print("=" * 80)

