#!/usr/bin/env python3
"""
Medicare Refiling Bot - Session Medium Audit

This standalone bot audits TherapyNotes sessions to determine whether the recorded
modifier matches the documented session medium (video vs phone). It was designed
to mirror the look and feel of the TN Refiling Bot while remaining completely
independent so future changes to either bot do not impact the other.
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog, simpledialog
import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path
import threading
import queue
import time
import json
import os
import subprocess
from datetime import datetime
from typing import Optional, Tuple, List, Dict, Any
import base64
import re

import pyautogui  # type: ignore

# Optional dependencies
try:
    import keyboard  # type: ignore
    KEYBOARD_AVAILABLE = True
except ImportError:
    KEYBOARD_AVAILABLE = False
    keyboard = None  # type: ignore

try:
    import pandas as pd  # type: ignore
    EXCEL_AVAILABLE = True
except ImportError:
    pd = None  # type: ignore
    EXCEL_AVAILABLE = False

try:
    import openpyxl  # type: ignore
    OPENPYXL_AVAILABLE = True
except ImportError:
    openpyxl = None  # type: ignore
    OPENPYXL_AVAILABLE = False

# PDF form filling libraries
try:
    import pdfrw  # type: ignore
    PDFRW_AVAILABLE = True
except ImportError:
    pdfrw = None  # type: ignore
    PDFRW_AVAILABLE = False

try:
    import PyPDF2  # type: ignore
    PYPDF2_AVAILABLE = True
except ImportError:
    PyPDF2 = None  # type: ignore
    PYPDF2_AVAILABLE = False

# PDF to image conversion for visual preview
try:
    import fitz  # PyMuPDF  # type: ignore
    PYMUPDF_AVAILABLE = True
except ImportError:
    fitz = None  # type: ignore
    PYMUPDF_AVAILABLE = False

# Image handling for PDF preview
try:
    from PIL import Image, ImageTk  # type: ignore
    PIL_AVAILABLE = True
except ImportError:
    Image = None  # type: ignore
    ImageTk = None  # type: ignore
    PIL_AVAILABLE = False

try:
    from selenium import webdriver  # type: ignore
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.common.exceptions import WebDriverException, TimeoutException
    SELENIUM_AVAILABLE = True
except ImportError:
    webdriver = None  # type: ignore
    Service = None  # type: ignore
    Options = None  # type: ignore
    By = None  # type: ignore
    Keys = None  # type: ignore
    WebDriverWait = None  # type: ignore
    EC = None  # type: ignore
    TimeoutException = Exception  # type: ignore
    WebDriverException = Exception  # type: ignore
    SELENIUM_AVAILABLE = False

# Optional webdriver-manager
try:
    from webdriver_manager.chrome import ChromeDriverManager  # type: ignore
    WEBDRIVER_MANAGER_AVAILABLE = True
except ImportError:
    ChromeDriverManager = None  # type: ignore
    WEBDRIVER_MANAGER_AVAILABLE = False

pyautogui.PAUSE = 1
pyautogui.FAILSAFE = True
try:
    import keyring  # type: ignore
    from keyring.errors import KeyringError  # type: ignore
    KEYRING_AVAILABLE = True
except ImportError:
    keyring = None  # type: ignore
    KeyringError = Exception
    KEYRING_AVAILABLE = False


# -----------------------------------------------------------------------------
# Logging configuration
# -----------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).parent
LOG_FILE_PATH = SCRIPT_DIR / "medicare_refiling_bot.log"

file_handler = RotatingFileHandler(
    filename=str(LOG_FILE_PATH),
    maxBytes=5 * 1024 * 1024,
    backupCount=5,
    encoding="utf-8"
)

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        file_handler,
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("MedicareRefilingBot")


# -----------------------------------------------------------------------------
# Main Bot Class
# -----------------------------------------------------------------------------
class MedicareRefilingBot:
    """Main bot container with GUI, user management, and audit workflow."""

    WINDOW_TITLE = "Medicare Refiling Bot - Session Medium Audit"
    KEYRING_SERVICE = "Medicare Refiling Bot"
    THERAPY_NOTES_URL = "https://www.therapynotes.com/app/login/IntegritySWS/?r=%2fapp%2fpatients%2f"

    def __init__(self) -> None:
        self.root: tk.Tk | None = None
        self.log_text: scrolledtext.ScrolledText | None = None
        self.gui_log_queue: queue.Queue[tuple[str, str]] = queue.Queue()
        self._gui_log_dispatcher_active = False

        # Credentials
        self.users_file = SCRIPT_DIR / "medicare_users.json"
        self.users: dict[str, dict[str, str]] = {}
        self.selected_user: str | None = None

        # Input/output selections
        self.selected_excel_path: Path | None = None
        self.output_excel_path: Path | None = None
        self.row_range: tuple[int | None, int | None] = (None, None)
        
        # Column mapping for flexible Excel column selection
        self.column_mappings: dict[str, str] = {
            "First Name": "",
            "Last Name": "",
            "Date of Service": "",
            "Original Modifier": "",
            "Expected Modifier": ""
        }
        self.available_columns: list[str] = []  # Will be populated when Excel is loaded
        
        # PDF form filling selections
        self.pdf_template_path: Path | None = None
        self.pdf_output_folder: Path | None = None
        self.pdf_audit_excel_path: Path | None = None
        self.pdf_filling_stop_requested: bool = False
        
        # Medicare Excel filling selections
        self.medicare_excel_template_path: Path | None = None
        self.medicare_excel_audit_excel_path: Path | None = None
        self.medicare_excel_output_path: Path | None = None
        self.medicare_excel_filling_stop_requested: bool = False
        
        # Medicare Excel column mappings for input Excel (using column letters)
        # These map to Medicare Excel output columns:
        # A: Patient's name, B: Patient's HICN/MBI, C: Date of service, E: Procedure Code
        self.medicare_excel_column_mappings: dict[str, str] = {
            "Patient's Name (Column A)": "",  # Column letter from input Excel
            "Patient's HICN/MBI (Column B)": "",  # Column letter from input Excel
            "Date of Service (Column C)": "",  # Column letter from input Excel
            "Procedure Code (Column E)": ""  # Column letter from input Excel
        }
        
        # Custom note to populate in Medicare Excel
        self.medicare_excel_custom_note: str = ""
        
        # PDF field mapping configuration file
        self.pdf_config_file = SCRIPT_DIR / "pdf_field_mapping_config.json"
        
        # PDF field mapping configuration
        # This will be loaded from the config file, with defaults below
        # Format: {"PDF Field Name": {"type": "excel"|"static"|"function", "value": "...", "enabled": true|false}}
        self.pdf_field_mapping_config: Dict[str, str | callable] = {}
        self.pdf_field_mapping_config_raw: Dict[str, Dict[str, Any]] = {}
        
        # Load PDF field mapping configuration from file
        self._load_pdf_field_mapping_config()

        # Threading control
        self._thread_lock = threading.Lock()
        self._active_threads: set[threading.Thread] = set()
        self._shutdown_requested = False
        self.stop_requested = False

        # Cached data
        self.excel_client_data: list[dict[str, str]] = []
        self.audit_results: list[dict[str, Any]] = []

        # Selenium
        self.driver = None
        self.wait = None
        self.is_logged_in = False
        self.current_client_name: str | None = None
        self.current_client_dob: str | None = None
        self.current_date_of_service: str | None = None
        self.current_patient_member_id: str | None = None
        self.current_service_code: str | None = None  # Service Code from input Excel (e.g., 90837, 90834)
        self.current_session_medium: Optional[str] = None
        self.current_original_modifier: Optional[str] = None

        # Load saved users on init
        self.load_users()

    # ------------------------------------------------------------------
    # Logging helpers
    # ------------------------------------------------------------------
    def log_error(self, message: str, exception: Exception | None = None, include_traceback: bool = True) -> None:
        if exception:
            if include_traceback:
                logger.exception(message, exc_info=exception)
            else:
                logger.error(f"{message}: {exception}")
        else:
            logger.error(message)
        self.gui_log(message, level="ERROR")

    def check_stop_requested(self) -> bool:
        return self.stop_requested or self._shutdown_requested

    # ------------------------------------------------------------------
    # Utility helpers
    # ------------------------------------------------------------------
    def gui_log(self, message: str, level: str = "INFO", include_context: bool = True) -> None:
        """Push log messages to the GUI log window."""
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        prefix = f"[{level}] " if include_context else ""
        text = f"{timestamp} {prefix}{message}\n"
        logger.log(getattr(logging, level.upper(), logging.INFO), message)
        self.gui_log_queue.put((text, level))
        if self.root and not self._gui_log_dispatcher_active:
            self._start_gui_log_dispatcher()

    def update_status(self, message: str, color: str = "#0066cc") -> None:
        """Update the GUI status label if available."""
        if hasattr(self, "status_label") and self.status_label:
            try:
                self.status_label.config(text=message, fg=color)
            except tk.TclError:
                pass

    def _start_gui_log_dispatcher(self) -> None:
        if not self.root or self._gui_log_dispatcher_active:
            return
        self._gui_log_dispatcher_active = True
        self.root.after(50, self._process_gui_log_queue)

    def _process_gui_log_queue(self) -> None:
        if not self.root:
            self._gui_log_dispatcher_active = False
            return
        try:
            while True:
                text, level = self.gui_log_queue.get_nowait()
                if self.log_text:
                    self.log_text.insert(tk.END, text)
                    self.log_text.see(tk.END)
        except queue.Empty:
            pass
        finally:
            if self.root and not self._shutdown_requested:
                self.root.after(50, self._process_gui_log_queue)
            else:
                self._gui_log_dispatcher_active = False

    # ------------------------------------------------------------------
    # Keyring helpers
    # ------------------------------------------------------------------
    def _ensure_keyring_available(self) -> bool:
        if KEYRING_AVAILABLE:
            return True
        if not getattr(self, "_keyring_warning_shown", False):
            self.gui_log(
                "Keyring module not available. Passwords will not be stored securely.",
                level="WARNING"
            )
            self._keyring_warning_shown = True
        return False

    def _store_password(self, username: str, password: str, display_name: Optional[str] = None) -> bool:
        if not username or not password:
            return False
        if not KEYRING_AVAILABLE:
            if display_name:
                encoded = base64.b64encode(password.encode("utf-8")).decode("utf-8")
                self.users.setdefault(display_name, {})["password_b64"] = encoded
                self.save_users()
                return True
            self._ensure_keyring_available()
            return False
        try:
            keyring.set_password(self.KEYRING_SERVICE, username, password)
            if display_name:
                entry = self.users.get(display_name)
                if entry and entry.get("password_b64"):
                    entry.pop("password_b64", None)
                    self.save_users()
            return True
        except KeyringError as e:  # type: ignore
            logger.warning("Failed to store password: %s", e)
            return False

    def _retrieve_password(self, username: str, display_name: Optional[str] = None) -> str | None:
        if not username:
            return None
        if KEYRING_AVAILABLE:
            try:
                stored = keyring.get_password(self.KEYRING_SERVICE, username)  # type: ignore
                if stored:
                    return stored
            except KeyringError as e:  # type: ignore
                logger.warning("Failed to retrieve password: %s", e)
        if display_name:
            encoded = self.users.get(display_name, {}).get("password_b64")
            if encoded:
                try:
                    return base64.b64decode(encoded.encode("utf-8")).decode("utf-8")
                except Exception:
                    return None
        self._ensure_keyring_available()
        return None

    def _delete_password(self, username: str, display_name: Optional[str] = None) -> None:
        if username and KEYRING_AVAILABLE:
            try:
                keyring.delete_password(self.KEYRING_SERVICE, username)  # type: ignore
            except KeyringError:
                pass
        if display_name and display_name in self.users:
            if self.users[display_name].get("password_b64"):
                self.users[display_name].pop("password_b64", None)
                self.save_users()

    # ------------------------------------------------------------------
    # User management
    # ------------------------------------------------------------------
    def load_users(self) -> None:
        if not self.users_file.exists():
            self.users = {}
            return
        try:
            with open(self.users_file, "r", encoding="utf-8") as f:
                stored = json.load(f)
            if isinstance(stored, dict):
                self.users = stored
            else:
                self.users = {}
        except Exception as exc:
            logger.error("Failed to load users JSON: %s", exc)
            self.users = {}

        # Hydrate credentials from keyring if available
        for display_name, creds in list(self.users.items()):
            username = creds.get("username", "")
            if not username:
                continue
            password = self._retrieve_password(username, display_name=display_name)
            creds["password_cached"] = bool(password)

    def save_users(self) -> None:
        try:
            with open(self.users_file, "w", encoding="utf-8") as f:
                json.dump(self.users, f, indent=2)
        except Exception as exc:
            logger.error("Failed to save users JSON: %s", exc)

    def _update_user_dropdown(self) -> None:
        if not hasattr(self, "user_dropdown"):
            return
        usernames = sorted(self.users.keys())
        current = self.user_dropdown.get() if self.user_dropdown else ""
        self.user_dropdown["values"] = usernames
        if usernames and current in usernames:
            self.user_dropdown.set(current)
        elif usernames:
            self.user_dropdown.set(usernames[0])
            self._on_user_selected()
        else:
            self.user_dropdown.set("")

    def _on_user_selected(self, _event=None) -> None:
        display_name = self.user_dropdown.get().strip() if hasattr(self, "user_dropdown") else ""
        if not display_name:
            return
        creds = self.users.get(display_name, {})
        username = creds.get("username", "")
        password = self._retrieve_password(username, display_name=display_name) if username else ""

        if hasattr(self, "username_entry"):
            self.username_entry.delete(0, tk.END)
            self.username_entry.insert(0, username)
        if hasattr(self, "password_entry"):
            self.password_entry.delete(0, tk.END)
            if password:
                self.password_entry.insert(0, password)

        self.selected_user = display_name
        self.gui_log(f"Selected user profile: {display_name}", level="DEBUG")

    def _add_user(self) -> None:
        display_name = simpledialog.askstring(
            "Add User",
            "Enter display name for this profile:",
            parent=self.root
        )
        if not display_name:
            return
        display_name = display_name.strip()
        if not display_name:
            messagebox.showerror("Invalid Name", "Display name cannot be empty.")
            return
        if display_name in self.users:
            messagebox.showwarning("Duplicate", "A user with this display name already exists.")
            return

        username = simpledialog.askstring(
            "Add User",
            "Enter TherapyNotes username:",
            parent=self.root
        )
        if not username:
            return
        username = username.strip()

        password = simpledialog.askstring(
            "Add User",
            "Enter TherapyNotes password:",
            show="*",
            parent=self.root
        )
        if password is None:
            return

        self.users[display_name] = {"username": username}
        self.save_users()
        stored = self._store_password(username, password, display_name=display_name)
        self.users[display_name]["password_cached"] = stored
        self.save_users()

        self.gui_log(f"Added user profile: {display_name}")
        self._update_user_dropdown()
        self.user_dropdown.set(display_name)
        self._on_user_selected()

    def _update_credentials(self) -> None:
        display_name = self.user_dropdown.get().strip() if hasattr(self, "user_dropdown") else ""
        if not display_name or display_name not in self.users:
            messagebox.showwarning("Select User", "Please select a user profile first.")
            return

        creds = self.users[display_name]
        username = simpledialog.askstring(
            "Update Credentials",
            "Enter TherapyNotes username:",
            initialvalue=creds.get("username", ""),
            parent=self.root
        )
        if username is None:
            return
        username = username.strip()

        password = simpledialog.askstring(
            "Update Credentials",
            "Enter TherapyNotes password:",
            show="*",
            parent=self.root
        )
        if password is None:
            # Keep existing password if user cancels
            password = self._retrieve_password(username, display_name=display_name)
            if not password:
                return

        old_username = creds.get("username")
        if old_username and old_username != username:
            self._delete_password(old_username, display_name=display_name)

        self.users[display_name] = {"username": username}
        self.save_users()
        stored = self._store_password(username, password, display_name=display_name)
        self.users[display_name]["password_cached"] = stored
        self.save_users()
        self.gui_log(f"Updated credentials for {display_name}")
        self._update_user_dropdown()
        self.user_dropdown.set(display_name)

    # ------------------------------------------------------------------
    # GUI construction
    # ------------------------------------------------------------------
    def create_main_window(self) -> None:
        self.root = tk.Tk()
        self.root.title(self.WINDOW_TITLE)
        self.root.geometry("900x820")
        self.root.configure(bg="#f0f0f0")
        self.root.protocol("WM_DELETE_WINDOW", self._close_window)

        # Center window
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (self.root.winfo_width() // 2)
        y = (self.root.winfo_screenheight() // 2) - (self.root.winfo_height() // 2)
        self.root.geometry(f"+{x}+{y}")

        # Header
        header_frame = tk.Frame(self.root, bg="#003366", height=70)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)

        title_label = tk.Label(header_frame, text="Medicare Refiling Bot",
                               font=("Segoe UI", 18, "bold"), bg="#003366", fg="white")
        title_label.pack(expand=True)

        subtitle_label = tk.Label(header_frame,
                                  text="Audit TherapyNotes session mediums vs modifiers",
                                  font=("Segoe UI", 10),
                                  bg="#003366",
                                  fg="#c6daf7")
        subtitle_label.pack()

        # Create notebook (tabbed interface)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Tab 1: Audit
        audit_tab = tk.Frame(self.notebook, bg="#f0f0f0")
        self.notebook.add(audit_tab, text="Audit")
        self._build_audit_tab(audit_tab)
        
        # Tab 2: PDF Forms
        pdf_forms_tab = tk.Frame(self.notebook, bg="#f0f0f0")
        self.notebook.add(pdf_forms_tab, text="PDF Forms")
        self._build_pdf_forms_tab(pdf_forms_tab)
        
        # Tab 3: Medicare Excel
        medicare_excel_tab = tk.Frame(self.notebook, bg="#f0f0f0")
        self.notebook.add(medicare_excel_tab, text="Medicare Excel")
        self._build_medicare_excel_tab(medicare_excel_tab)
        
        # Log section (shared across tabs)
        log_frame = tk.Frame(self.root, bg="#f0f0f0")
        log_frame.pack(fill="both", expand=False, padx=10, pady=(0, 10))
        self._build_log_section(log_frame)

        self._start_gui_log_dispatcher()
        self.gui_log("Medicare Refiling Bot initialized. Ready.")

    def _build_audit_tab(self, parent: tk.Widget) -> None:
        """Build the Audit tab with scrollable content."""
        # Scrollable canvas for audit tab
        canvas_container = tk.Frame(parent, bg="#f0f0f0")
        canvas_container.pack(fill="both", expand=True)

        canvas = tk.Canvas(canvas_container, bg="#f0f0f0", highlightthickness=0)
        scrollbar = tk.Scrollbar(canvas_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#f0f0f0")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)

        canvas.bind("<Configure>", on_canvas_configure)

        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", on_mousewheel)

        main_frame = tk.Frame(scrollable_frame, bg="#f0f0f0")
        main_frame.pack(fill="both", expand=True, padx=20, pady=15)

        self._build_login_section(main_frame)
        self._build_excel_section(main_frame)
        self._build_column_mapping_section(main_frame)
        self._build_row_selection_section(main_frame)
        self._build_output_section(main_frame)
        self._build_processing_section(main_frame)

    def _build_pdf_forms_tab(self, parent: tk.Widget) -> None:
        """Build the PDF Forms tab for filling Medicare forms."""
        # Scrollable canvas for PDF forms tab
        canvas_container = tk.Frame(parent, bg="#f0f0f0")
        canvas_container.pack(fill="both", expand=True)

        canvas = tk.Canvas(canvas_container, bg="#f0f0f0", highlightthickness=0)
        scrollbar = tk.Scrollbar(canvas_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#f0f0f0")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)

        canvas.bind("<Configure>", on_canvas_configure)

        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", on_mousewheel)

        main_frame = tk.Frame(scrollable_frame, bg="#f0f0f0")
        main_frame.pack(fill="both", expand=True, padx=20, pady=15)

        self._build_pdf_template_section(main_frame)
        self._build_pdf_field_config_section(main_frame)
        self._build_pdf_audit_excel_section(main_frame)
        self._build_pdf_output_section(main_frame)
        self._build_pdf_processing_section(main_frame)
    
    def _build_medicare_excel_tab(self, parent: tk.Widget) -> None:
        """Build the Medicare Excel tab for filling Medicare Excel forms."""
        # Scrollable canvas for Medicare Excel tab
        canvas_container = tk.Frame(parent, bg="#f0f0f0")
        canvas_container.pack(fill="both", expand=True)
        
        canvas = tk.Canvas(canvas_container, bg="#f0f0f0", highlightthickness=0)
        scrollbar = tk.Scrollbar(canvas_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#f0f0f0")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        def on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)
        
        canvas.bind("<Configure>", on_canvas_configure)
        
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        main_frame = tk.Frame(scrollable_frame, bg="#f0f0f0")
        main_frame.pack(fill="both", expand=True, padx=20, pady=15)
        
        self._build_medicare_excel_template_section(main_frame)
        self._build_medicare_excel_audit_excel_section(main_frame)
        self._build_medicare_excel_column_mapping_section(main_frame)
        self._build_medicare_excel_note_section(main_frame)
        self._build_medicare_excel_output_section(main_frame)
        self._build_medicare_excel_processing_section(main_frame)
    
    def _build_medicare_excel_template_section(self, parent: tk.Widget) -> None:
        """Build the Medicare Excel template selection section."""
        section = tk.LabelFrame(parent, text="Medicare Excel Template",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))
        
        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)
        
        instruction = "Select the Medicare Excel template file (LVAM Request Form)."
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10), anchor="w")
        
        row = tk.Frame(content, bg="#f0f0f0")
        row.pack(fill="x")
        
        tk.Label(row, text="Template:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.medicare_excel_template_entry = tk.Entry(row, font=("Segoe UI", 9), width=50, state="readonly")
        self.medicare_excel_template_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        tk.Button(row, text="Browse...", command=self._browse_medicare_excel_template,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 5))
        
        tk.Button(row, text="Clear", command=self._clear_medicare_excel_template,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
        
        self.medicare_excel_template_status_label = tk.Label(content,
                                                              text="No template selected.",
                                                              font=("Segoe UI", 9, "italic"),
                                                              bg="#f0f0f0", fg="#666666")
        self.medicare_excel_template_status_label.pack(pady=(6, 0))
    
    def _build_medicare_excel_audit_excel_section(self, parent: tk.Widget) -> None:
        """Build the audit Excel file selection section for Medicare Excel."""
        section = tk.LabelFrame(parent, text="Audit Results Excel",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))
        
        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)
        
        instruction = "Select the audit results Excel file (output from Audit tab). Only clients with 'Needs Refile' status will be processed."
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10), anchor="w")
        
        row = tk.Frame(content, bg="#f0f0f0")
        row.pack(fill="x")
        
        tk.Label(row, text="Audit Excel:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.medicare_excel_audit_excel_entry = tk.Entry(row, font=("Segoe UI", 9), width=50, state="readonly")
        self.medicare_excel_audit_excel_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        tk.Button(row, text="Browse...", command=self._browse_medicare_excel_audit_excel,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 5))
        
        tk.Button(row, text="Clear", command=self._clear_medicare_excel_audit_excel,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
        
        self.medicare_excel_audit_excel_status_label = tk.Label(content,
                                                                 text="No audit Excel selected.",
                                                                 font=("Segoe UI", 9, "italic"),
                                                                 bg="#f0f0f0", fg="#666666")
        self.medicare_excel_audit_excel_status_label.pack(pady=(6, 0))
    
    def _build_medicare_excel_column_mapping_section(self, parent: tk.Widget) -> None:
        """Build the column mapping section for Medicare Excel input data."""
        section = tk.LabelFrame(parent, text="Column Mapping (Input Excel)",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))
        
        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)
        
        instruction = "Map the column letters from your input Excel file to the data fields needed for the Medicare Excel form. Enter column letters (e.g., A, B, C, D) that correspond to each field in your input Excel file."
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10), anchor="w")
        
        # Create a frame for the column mappings
        mapping_frame = tk.Frame(content, bg="#f0f0f0")
        mapping_frame.pack(fill="x")
        
        # Row 1: Patient's Name (Column A) and Patient's HICN/MBI (Column B)
        row1 = tk.Frame(mapping_frame, bg="#f0f0f0")
        row1.pack(fill="x", pady=(0, 8))
        
        tk.Label(row1, text="Patient's Name (Column A):", font=("Segoe UI", 9), bg="#f0f0f0", width=25, anchor="w").pack(side="left", padx=(0, 10))
        self.medicare_excel_col_patient_name = tk.Entry(row1, font=("Segoe UI", 9), width=8)
        self.medicare_excel_col_patient_name.pack(side="left", padx=(0, 20))
        self.medicare_excel_col_patient_name.insert(0, self.medicare_excel_column_mappings.get("Patient's Name (Column A)", ""))
        
        tk.Label(row1, text="Patient's HICN/MBI (Column B):", font=("Segoe UI", 9), bg="#f0f0f0", width=28, anchor="w").pack(side="left", padx=(0, 10))
        self.medicare_excel_col_hicn = tk.Entry(row1, font=("Segoe UI", 9), width=8)
        self.medicare_excel_col_hicn.pack(side="left")
        self.medicare_excel_col_hicn.insert(0, self.medicare_excel_column_mappings.get("Patient's HICN/MBI (Column B)", ""))
        
        # Row 2: Date of Service (Column C) and Procedure Code (Column E)
        row2 = tk.Frame(mapping_frame, bg="#f0f0f0")
        row2.pack(fill="x", pady=(0, 8))
        
        tk.Label(row2, text="Date of Service (Column C):", font=("Segoe UI", 9), bg="#f0f0f0", width=25, anchor="w").pack(side="left", padx=(0, 10))
        self.medicare_excel_col_dos = tk.Entry(row2, font=("Segoe UI", 9), width=8)
        self.medicare_excel_col_dos.pack(side="left", padx=(0, 20))
        self.medicare_excel_col_dos.insert(0, self.medicare_excel_column_mappings.get("Date of Service (Column C)", ""))
        
        tk.Label(row2, text="Procedure Code (Column E):", font=("Segoe UI", 9), bg="#f0f0f0", width=28, anchor="w").pack(side="left", padx=(0, 10))
        self.medicare_excel_col_service_code = tk.Entry(row2, font=("Segoe UI", 9), width=8)
        self.medicare_excel_col_service_code.pack(side="left")
        self.medicare_excel_col_service_code.insert(0, self.medicare_excel_column_mappings.get("Procedure Code (Column E)", ""))
        
        # Apply button
        apply_row = tk.Frame(mapping_frame, bg="#f0f0f0")
        apply_row.pack(fill="x", pady=(5, 0))
        
        def _apply_column_mappings():
            """Apply the column mappings from the GUI fields."""
            self.medicare_excel_column_mappings["Patient's Name (Column A)"] = self.medicare_excel_col_patient_name.get().strip().upper()
            self.medicare_excel_column_mappings["Patient's HICN/MBI (Column B)"] = self.medicare_excel_col_hicn.get().strip().upper()
            self.medicare_excel_column_mappings["Date of Service (Column C)"] = self.medicare_excel_col_dos.get().strip().upper()
            self.medicare_excel_column_mappings["Procedure Code (Column E)"] = self.medicare_excel_col_service_code.get().strip().upper()
            self.gui_log(f"Applied column mappings: {self.medicare_excel_column_mappings}", level="INFO")
        
        tk.Button(apply_row, text="Apply Column Mappings", command=_apply_column_mappings,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
    
    def _build_medicare_excel_note_section(self, parent: tk.Widget) -> None:
        """Build the custom note section for Medicare Excel."""
        section = tk.LabelFrame(parent, text="Custom Note",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))
        
        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)
        
        instruction = "Enter a custom note that will be populated in the 'Explain Correction Needed' column (Column F) of the Medicare Excel form for all clients. Leave empty to use auto-generated notes based on modifier changes."
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10), anchor="w")
        
        row = tk.Frame(content, bg="#f0f0f0")
        row.pack(fill="x")
        
        tk.Label(row, text="Custom Note:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10), anchor="n")
        
        # Use a Text widget for multi-line note input
        note_frame = tk.Frame(row, bg="#f0f0f0")
        note_frame.pack(side="left", fill="both", expand=True)
        
        self.medicare_excel_custom_note_text = scrolledtext.ScrolledText(note_frame, height=3, 
                                                                          font=("Segoe UI", 9),
                                                                          wrap=tk.WORD)
        self.medicare_excel_custom_note_text.pack(fill="both", expand=True)
        if self.medicare_excel_custom_note:
            self.medicare_excel_custom_note_text.insert("1.0", self.medicare_excel_custom_note)
        
        # Apply button for note
        apply_note_row = tk.Frame(content, bg="#f0f0f0")
        apply_note_row.pack(fill="x", pady=(10, 0))
        
        def _apply_custom_note():
            """Apply the custom note from the GUI field."""
            self.medicare_excel_custom_note = self.medicare_excel_custom_note_text.get("1.0", tk.END).strip()
            if self.medicare_excel_custom_note:
                self.gui_log(f"Applied custom note: {self.medicare_excel_custom_note[:50]}...", level="INFO")
            else:
                self.gui_log("Cleared custom note - will use auto-generated notes", level="INFO")
        
        tk.Button(apply_note_row, text="Apply Custom Note", command=_apply_custom_note,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
    
    def _build_medicare_excel_output_section(self, parent: tk.Widget) -> None:
        """Build the output file selection section for Medicare Excel."""
        section = tk.LabelFrame(parent, text="Output File",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))
        
        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)
        
        instruction = "Select where to save the filled Medicare Excel file. If not provided, a timestamped file will be created."
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10), anchor="w")
        
        row = tk.Frame(content, bg="#f0f0f0")
        row.pack(fill="x")
        
        tk.Label(row, text="Output File:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.medicare_excel_output_entry = tk.Entry(row, font=("Segoe UI", 9), width=50, state="readonly")
        self.medicare_excel_output_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        tk.Button(row, text="Browse...", command=self._browse_medicare_excel_output,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 5))
        
        tk.Button(row, text="Clear", command=self._clear_medicare_excel_output,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
        
        self.medicare_excel_output_status_label = tk.Label(content,
                                                            text="No output file selected.",
                                                            font=("Segoe UI", 9, "italic"),
                                                            bg="#f0f0f0", fg="#666666")
        self.medicare_excel_output_status_label.pack(pady=(6, 0))
    
    def _build_medicare_excel_processing_section(self, parent: tk.Widget) -> None:
        """Build the Medicare Excel processing section with start/stop buttons."""
        section = tk.LabelFrame(parent, text="Process Medicare Excel",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))
        
        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)
        
        instruction = "Click 'Start Filling Medicare Excel' to populate the Medicare Excel form for all clients with 'Needs Refile' status."
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10), anchor="w")
        
        button_row = tk.Frame(content, bg="#f0f0f0")
        button_row.pack(fill="x")
        
        self.medicare_excel_start_button = tk.Button(button_row, text="Start Filling Medicare Excel",
                                                      command=self._start_medicare_excel_filling,
                                                      bg="#28a745", fg="white", font=("Segoe UI", 10, "bold"),
                                                      padx=20, pady=8, cursor="hand2", relief="flat")
        self.medicare_excel_start_button.pack(side="left", padx=(0, 10))
        
        self.medicare_excel_stop_button = tk.Button(button_row, text="Stop",
                                                     command=self._stop_medicare_excel_filling,
                                                     bg="#dc3545", fg="white", font=("Segoe UI", 10),
                                                     padx=20, pady=8, cursor="hand2", relief="flat",
                                                     state="disabled")
        self.medicare_excel_stop_button.pack(side="left")
        
        self.medicare_excel_status_label = tk.Label(content,
                                                     text="Ready to fill Medicare Excel form.",
                                                     font=("Segoe UI", 9, "bold"),
                                                     bg="#f0f0f0", fg="#006644")
        self.medicare_excel_status_label.pack(pady=(10, 0), anchor="w")
        
        self.medicare_excel_stats_label = tk.Label(content,
                                                    text="",
                                                    font=("Segoe UI", 9),
                                                    bg="#f0f0f0", fg="#666666")
        self.medicare_excel_stats_label.pack(pady=(5, 0), anchor="w")

    def _build_login_section(self, parent: tk.Widget) -> None:
        section = tk.LabelFrame(parent, text="TherapyNotes Login Credentials",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))

        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)

        # User profile row
        user_row = tk.Frame(content, bg="#f0f0f0")
        user_row.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))

        tk.Label(user_row, text="User:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.user_dropdown = ttk.Combobox(user_row, font=("Segoe UI", 10), width=25, state="readonly")
        self.user_dropdown.pack(side="left", padx=(0, 10))
        self.user_dropdown.bind("<<ComboboxSelected>>", self._on_user_selected)
        self._update_user_dropdown()

        tk.Button(user_row, text="Add User", command=self._add_user,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 5))

        tk.Button(user_row, text="Update Credentials", command=self._update_credentials,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 5))

        # Username / password fields
        tk.Label(content, text="Username:", font=("Segoe UI", 10), bg="#f0f0f0").grid(row=1, column=0, sticky="w")
        self.username_entry = tk.Entry(content, font=("Segoe UI", 10), width=40)
        self.username_entry.grid(row=2, column=0, sticky="ew", pady=(0, 10))

        tk.Label(content, text="Password:", font=("Segoe UI", 10), bg="#f0f0f0").grid(row=3, column=0, sticky="w")
        self.password_entry = tk.Entry(content, font=("Segoe UI", 10), width=40, show="*")
        self.password_entry.grid(row=4, column=0, sticky="ew", pady=(0, 10))

        self.status_label = tk.Label(content,
                                     text="Status: Awaiting login.",
                                     font=("Segoe UI", 9, "bold"),
                                     bg="#f0f0f0", fg="#0066cc")
        self.status_label.grid(row=5, column=0, sticky="w")
        content.grid_columnconfigure(0, weight=1)

    def _build_excel_section(self, parent: tk.Widget) -> None:
        section = tk.LabelFrame(parent, text="Client Data Input (Excel/CSV)",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))

        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)

        instruction = ("Select an Excel (.xlsx) or CSV file containing client data "
                       "(Name, DOB, Date of Service, modifiers, etc.).")
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10))

        row = tk.Frame(content, bg="#f0f0f0")
        row.pack(fill="x")

        tk.Label(row, text="Excel/CSV File:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.excel_entry = tk.Entry(row, font=("Segoe UI", 9), width=50, state="readonly")
        self.excel_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        tk.Button(row, text="Browse...", command=self._browse_excel,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 5))

        tk.Button(row, text="Clear", command=self._clear_excel,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")

        self.excel_status_label = tk.Label(content,
                                           text="No file selected.",
                                           font=("Segoe UI", 9, "italic"),
                                           bg="#f0f0f0", fg="#666666")
        self.excel_status_label.pack(pady=(6, 0))
        
        # Preview section
        preview_frame = tk.Frame(content, bg="#f0f0f0")
        preview_frame.pack(fill="both", expand=True, pady=(10, 0))
        
        tk.Label(preview_frame, text="Preview (First 20 clients to be processed):",
                 font=("Segoe UI", 9, "bold"), bg="#f0f0f0", fg="#003366").pack(anchor="w", pady=(0, 5))
        
        # Create preview text widget with scrollbar
        preview_container = tk.Frame(preview_frame, bg="#f0f0f0")
        preview_container.pack(fill="both", expand=True)
        
        preview_scrollbar = tk.Scrollbar(preview_container)
        preview_scrollbar.pack(side="right", fill="y")
        
        self.preview_text = tk.Text(preview_container,
                                    height=8,
                                    font=("Courier New", 9),
                                    bg="#ffffff",
                                    fg="#000000",
                                    yscrollcommand=preview_scrollbar.set,
                                    wrap=tk.NONE,
                                    state="disabled")
        self.preview_text.pack(side="left", fill="both", expand=True)
        preview_scrollbar.config(command=self.preview_text.yview)
        
        self.preview_count_label = tk.Label(preview_frame,
                                           text="No preview available.",
                                           font=("Segoe UI", 8, "italic"),
                                           bg="#f0f0f0", fg="#666666")
        self.preview_count_label.pack(pady=(5, 0))
        
        # Initialize preview with default message
        self.preview_text.config(state="normal")
        self.preview_text.insert(tk.END, "No preview available. Select an Excel/CSV file to see preview.")
        self.preview_text.config(state="disabled")

    def _build_column_mapping_section(self, parent: tk.Widget) -> None:
        """Build the column mapping section for flexible Excel column selection."""
        section = tk.LabelFrame(parent, text="Column Mapping (Optional)",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))

        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)

        instruction = ("Map Excel columns to required fields. Leave empty to use automatic detection.\n"
                       "If your Excel has non-standard column names, specify them here.\n\n"
                       "Note: 'Expected Modifier' is optional for initial runs - the bot will deduce it during the audit workflow.")
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(anchor="w", pady=(0, 10))

        # Store entry fields for column mappings
        self.column_mapping_entries: dict[str, tk.Entry] = {}

        mapping_fields = [
            ("First Name", "First Name or separate First Name column"),
            ("Last Name", "Last Name or separate Last Name column"),
            ("Date of Service", "Date of Service"),
            ("Original Modifier", "Original Modifier (e.g., map 'Modifier' column to this)"),
            ("Expected Modifier", "Expected Modifier (optional - bot will deduce during audit if empty)")
        ]

        for idx, (field_name, description) in enumerate(mapping_fields):
            row = tk.Frame(content, bg="#f0f0f0")
            row.pack(fill="x", pady=(0, 8))

            tk.Label(row, text=f"{field_name}:", font=("Segoe UI", 9), bg="#f0f0f0", width=20, anchor="w").pack(side="left", padx=(0, 10))
            
            entry = tk.Entry(row, font=("Segoe UI", 9), width=30)
            entry.pack(side="left", padx=(0, 10))
            entry.bind("<KeyRelease>", lambda e, field=field_name: self._on_column_mapping_changed(field))
            self.column_mapping_entries[field_name] = entry

            tk.Label(row, text=description, font=("Segoe UI", 8, "italic"), 
                    bg="#f0f0f0", fg="#666666").pack(side="left")

        # Info label showing available columns
        info_row = tk.Frame(content, bg="#f0f0f0")
        info_row.pack(fill="x", pady=(10, 0))

        tk.Button(info_row, text="Show Available Columns",
                  command=self._show_available_columns,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")

        self.column_mapping_status = tk.Label(info_row,
                                              text="Type column names (e.g., 'Excel_First_Name') or column letters (e.g., 'B', 'C', 'AB') or leave empty for auto-detect.",
                                              font=("Segoe UI", 8, "italic"),
                                              bg="#f0f0f0", fg="#666666")
        self.column_mapping_status.pack(side="left", padx=(15, 0))

    def _show_available_columns(self) -> None:
        """Show available columns from the loaded Excel file in a message box."""
        if not self.selected_excel_path or not self.selected_excel_path.exists():
            messagebox.showinfo("No File", "Please select an Excel file first.")
            return

        try:
            # Load Excel to get column names (without applying mappings to avoid recursion)
            if not OPENPYXL_AVAILABLE:
                messagebox.showerror("Missing Dependency", "openpyxl is required to read Excel files.")
                return

            wb = openpyxl.load_workbook(self.selected_excel_path, data_only=True, read_only=True)
            ws = wb.active
            header_row = 1
            headers = []
            for cell in ws[header_row]:
                headers.append(str(cell.value) if cell.value else "")
            wb.close()

            if not headers:
                messagebox.showinfo("No Columns", "No columns found in Excel file.")
                return

            # Show columns in a message box with column letters
            import string
            def get_column_letter(index):
                """Convert 0-based index to Excel column letter (A, B, ..., Z, AA, AB, etc.)"""
                result = ""
                index += 1  # Convert to 1-based
                while index > 0:
                    index -= 1
                    result = chr(65 + (index % 26)) + result
                    index //= 26
                return result
            
            columns_text = "\n".join([f"{get_column_letter(i)}. {col}" for i, col in enumerate(headers)])
            messagebox.showinfo(
                "Available Columns",
                f"Found {len(headers)} column(s) in the Excel file:\n\n{columns_text}\n\n"
                "You can type either:\n"
                "- Column name (e.g., 'Excel_First_Name')\n"
                "- Column letter (e.g., 'B', 'C', 'AB')\n\n"
                "Type in the mapping fields above."
            )
            self.available_columns = headers
            self.column_mapping_status.config(
                text=f"Found {len(headers)} column(s). Type column names or letters (e.g., 'B', 'C', 'AB') above or leave empty for auto-detect.",
                fg="#006644"
            )

        except Exception as e:
            messagebox.showerror("Error", f"Error loading columns: {str(e)}")
            self.gui_log(f"Error showing available columns: {str(e)}", level="ERROR")

    def _on_column_mapping_changed(self, field_name: str) -> None:
        """Handle column mapping text change."""
        entry = self.column_mapping_entries.get(field_name)
        if entry:
            column_name = entry.get().strip()
            self.column_mappings[field_name] = column_name
            if column_name:
                self.gui_log(f"Column mapping updated: {field_name} -> '{column_name}'", level="DEBUG")

    def _build_row_selection_section(self, parent: tk.Widget) -> None:
        section = tk.LabelFrame(parent, text="Row Selection",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))

        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)

        tk.Label(content, text="Specify optional row range to audit (1-indexed, inclusive).",
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555").pack(anchor="w")

        row = tk.Frame(content, bg="#f0f0f0")
        row.pack(fill="x", pady=(10, 0))

        tk.Label(row, text="Start Row:", font=("Segoe UI", 9), bg="#f0f0f0").pack(side="left")
        self.start_row_entry = tk.Entry(row, font=("Segoe UI", 9), width=8)
        self.start_row_entry.pack(side="left", padx=(5, 20))

        tk.Label(row, text="End Row:", font=("Segoe UI", 9), bg="#f0f0f0").pack(side="left")
        self.end_row_entry = tk.Entry(row, font=("Segoe UI", 9), width=8)
        self.end_row_entry.pack(side="left", padx=(5, 10))

        tk.Button(row, text="Apply",
                  command=self._apply_row_range,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=10, pady=4, relief="flat", cursor="hand2").pack(side="left")

        self.row_status_label = tk.Label(content,
                                         text="Row range: All rows",
                                         font=("Segoe UI", 9, "italic"),
                                         bg="#f0f0f0", fg="#666666")
        self.row_status_label.pack(pady=(6, 0), anchor="w")

    def _build_output_section(self, parent: tk.Widget) -> None:
        section = tk.LabelFrame(parent, text="Audit Output",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))

        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)

        instruction = "Select where to save the audit Excel file. If not provided, a timestamped file will be created."
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10), anchor="w")

        row = tk.Frame(content, bg="#f0f0f0")
        row.pack(fill="x")

        tk.Label(row, text="Output Excel:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.output_entry = tk.Entry(row, font=("Segoe UI", 9), width=50, state="readonly")
        self.output_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        tk.Button(row, text="Browse...", command=self._browse_output_excel,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 5))

        tk.Button(row, text="Clear", command=self._clear_output_excel,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")

        self.output_status_label = tk.Label(content,
                                            text="No output file selected.",
                                            font=("Segoe UI", 9, "italic"),
                                            bg="#f0f0f0", fg="#666666")
        self.output_status_label.pack(pady=(6, 0), anchor="w")

    def _build_processing_section(self, parent: tk.Widget) -> None:
        section = tk.LabelFrame(parent, text="Audit Processing",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))

        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)

        tk.Label(content,
                 text="After logging in and loading client data, start the audit to review modifiers vs session medium.",
                 font=("Segoe UI", 9),
                 bg="#f0f0f0", fg="#555555", wraplength=600,
                 justify="left").pack(pady=(0, 10), anchor="w")

        row = tk.Frame(content, bg="#f0f0f0")
        row.pack(fill="x")

        self.start_button = tk.Button(row, text="Start Audit",
                                      command=self._start_audit,
                                      bg="#006644", fg="white", font=("Segoe UI", 10, "bold"),
                                      padx=18, pady=6, cursor="hand2", relief="flat", state="normal")
        self.start_button.pack(side="left", padx=(0, 10))

        self.stop_button = tk.Button(row, text="Stop",
                                     command=self.request_stop,
                                     bg="#a60000", fg="white", font=("Segoe UI", 10, "bold"),
                                     padx=18, pady=6, cursor="hand2", relief="flat", state="disabled")
        self.stop_button.pack(side="left")

        self.processing_stats_label = tk.Label(content,
                                               text="Clients audited: 0 | Pending: 0",
                                               font=("Segoe UI", 9),
                                               bg="#f0f0f0", fg="#555555")
        self.processing_stats_label.pack(pady=(8, 0), anchor="w")

    def _build_log_section(self, parent: tk.Widget) -> None:
        section = tk.LabelFrame(parent, text="Activity Log",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="both", expand=True)

        self.log_text = scrolledtext.ScrolledText(section, wrap=tk.WORD, font=("Consolas", 9),
                                                  bg="#1e1e1e", fg="#dcdcdc", height=10)
        self.log_text.pack(fill="both", expand=True, padx=10, pady=10)
        self.log_text.insert(tk.END, "Medicare Refiling Bot ready.\n")

        footer = tk.Frame(parent, bg="#f0f0f0")
        footer.pack(fill="x", pady=(5, 0))

        tk.Button(footer, text="Open Logs Folder",
                  command=self._open_logs_folder,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")

        tk.Button(footer, text="Close",
                  command=self._close_window,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="right")

    # ------------------------------------------------------------------
    # PDF Forms Tab Sections
    # ------------------------------------------------------------------
    def _build_pdf_template_section(self, parent: tk.Widget) -> None:
        """Build the PDF template selection section."""
        section = tk.LabelFrame(parent, text="PDF Form Template",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))

        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)

        instruction = "Select the Medicare reopening request form template (PDF)."
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10), anchor="w")

        row = tk.Frame(content, bg="#f0f0f0")
        row.pack(fill="x")

        tk.Label(row, text="PDF Template:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.pdf_template_entry = tk.Entry(row, font=("Segoe UI", 9), width=50, state="readonly")
        self.pdf_template_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        tk.Button(row, text="Browse...", command=self._browse_pdf_template,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 5))

        tk.Button(row, text="Clear", command=self._clear_pdf_template,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")

        # Add button to list PDF fields
        button_row = tk.Frame(content, bg="#f0f0f0")
        button_row.pack(fill="x", pady=(10, 0))

        tk.Button(button_row, text="List PDF Form Fields",
                  command=self._list_and_display_pdf_fields,
                  bg="#28a745", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")

        tk.Label(button_row, text="(Click to see all fields in the PDF form)",
                 font=("Segoe UI", 8, "italic"), bg="#f0f0f0", fg="#666666").pack(side="left", padx=(10, 0))

        self.pdf_template_status_label = tk.Label(content,
                                                   text="No template selected.",
                                                   font=("Segoe UI", 9, "italic"),
                                                   bg="#f0f0f0", fg="#666666")
        self.pdf_template_status_label.pack(pady=(6, 0))
    
    def _build_pdf_field_config_section(self, parent: tk.Widget) -> None:
        """Build the PDF field configuration section."""
        section = tk.LabelFrame(parent, text="PDF Field Configuration",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="both", expand=True, pady=(0, 15))
        
        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)
        
        instruction = (
            "Configure which PDF fields to fill and what values to use. "
            "Fields not configured will be left blank."
        )
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10), anchor="w")
        
        # Buttons row
        button_row = tk.Frame(content, bg="#f0f0f0")
        button_row.pack(fill="x", pady=(0, 10))
        
        tk.Button(button_row, text="Visual Field Mapper",
                  command=self._open_visual_field_mapper,
                  bg="#28a745", fg="white", font=("Segoe UI", 10, "bold"),
                  padx=15, pady=6, cursor="hand2", relief="flat").pack(side="left", padx=(0, 10))
        
        tk.Button(button_row, text="List Fields (Text)",
                  command=self._open_field_config_window,
                  bg="#6c757d", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 10))
        
        tk.Button(button_row, text="Load Defaults",
                  command=self._load_default_field_mappings,
                  bg="#17a2b8", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 10))
        
        tk.Button(button_row, text="Clear All",
                  command=self._clear_field_mappings,
                  bg="#dc3545", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")
        
        # Status label
        self.pdf_config_status_label = tk.Label(content,
                                                 text="No fields configured. Click 'Visual Field Mapper' to set up field mappings.",
                                                 font=("Segoe UI", 9, "italic"),
                                                 bg="#f0f0f0", fg="#666666",
                                                 wraplength=600, justify="left")
        self.pdf_config_status_label.pack(pady=(5, 0), anchor="w")
        
        # Update status label
        self._update_pdf_config_status()

    def _build_pdf_audit_excel_section(self, parent: tk.Widget) -> None:
        """Build the audit Excel file selection section."""
        section = tk.LabelFrame(parent, text="Audit Results Excel",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))

        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)

        instruction = "Select the audit results Excel file (output from Audit tab). Only clients with 'Needs Refile' status will be processed."
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10), anchor="w")

        row = tk.Frame(content, bg="#f0f0f0")
        row.pack(fill="x")

        tk.Label(row, text="Audit Excel:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.pdf_audit_excel_entry = tk.Entry(row, font=("Segoe UI", 9), width=50, state="readonly")
        self.pdf_audit_excel_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        tk.Button(row, text="Browse...", command=self._browse_pdf_audit_excel,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 5))

        tk.Button(row, text="Clear", command=self._clear_pdf_audit_excel,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")

        self.pdf_audit_excel_status_label = tk.Label(content,
                                                      text="No audit Excel selected.",
                                                      font=("Segoe UI", 9, "italic"),
                                                      bg="#f0f0f0", fg="#666666")
        self.pdf_audit_excel_status_label.pack(pady=(6, 0))

    def _build_pdf_output_section(self, parent: tk.Widget) -> None:
        """Build the PDF output folder selection section."""
        section = tk.LabelFrame(parent, text="Output Folder",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))

        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)

        instruction = "Select the master folder where filled PDF forms will be saved. Each client will get their own subfolder."
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10), anchor="w")

        row = tk.Frame(content, bg="#f0f0f0")
        row.pack(fill="x")

        tk.Label(row, text="Output Folder:", font=("Segoe UI", 10), bg="#f0f0f0").pack(side="left", padx=(0, 10))
        self.pdf_output_folder_entry = tk.Entry(row, font=("Segoe UI", 9), width=50, state="readonly")
        self.pdf_output_folder_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        tk.Button(row, text="Browse...", command=self._browse_pdf_output_folder,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left", padx=(0, 5))

        tk.Button(row, text="Clear", command=self._clear_pdf_output_folder,
                  bg="#003366", fg="white", font=("Segoe UI", 9),
                  padx=12, pady=4, cursor="hand2", relief="flat").pack(side="left")

        self.pdf_output_folder_status_label = tk.Label(content,
                                                        text="No output folder selected.",
                                                        font=("Segoe UI", 9, "italic"),
                                                        bg="#f0f0f0", fg="#666666")
        self.pdf_output_folder_status_label.pack(pady=(6, 0))

    def _build_pdf_processing_section(self, parent: tk.Widget) -> None:
        """Build the PDF processing section with start/stop buttons."""
        section = tk.LabelFrame(parent, text="Process PDF Forms",
                                font=("Segoe UI", 11, "bold"),
                                bg="#f0f0f0", fg="#003366")
        section.pack(fill="x", pady=(0, 15))

        content = tk.Frame(section, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=15, pady=10)

        instruction = "Click 'Start Filling PDFs' to fill out PDF forms for all clients with 'Needs Refile' status."
        tk.Label(content, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=600, justify="left").pack(pady=(0, 10), anchor="w")

        button_row = tk.Frame(content, bg="#f0f0f0")
        button_row.pack(fill="x")

        self.pdf_start_button = tk.Button(button_row, text="Start Filling PDFs",
                                           command=self._start_pdf_filling,
                                           bg="#28a745", fg="white", font=("Segoe UI", 10, "bold"),
                                           padx=20, pady=8, cursor="hand2", relief="flat")
        self.pdf_start_button.pack(side="left", padx=(0, 10))

        self.pdf_stop_button = tk.Button(button_row, text="Stop",
                                          command=self._stop_pdf_filling,
                                          bg="#dc3545", fg="white", font=("Segoe UI", 10, "bold"),
                                          padx=20, pady=8, cursor="hand2", relief="flat", state="disabled")
        self.pdf_stop_button.pack(side="left")

        self.pdf_status_label = tk.Label(content,
                                          text="Ready to fill PDF forms.",
                                          font=("Segoe UI", 9, "bold"),
                                          bg="#f0f0f0", fg="#0066cc")
        self.pdf_status_label.pack(pady=(10, 0), anchor="w")

        self.pdf_stats_label = tk.Label(content,
                                         text="PDFs filled: 0 | Pending: 0",
                                         font=("Segoe UI", 9),
                                         bg="#f0f0f0", fg="#555555")
        self.pdf_stats_label.pack(pady=(8, 0), anchor="w")

    # ------------------------------------------------------------------
    # File selection helpers
    # ------------------------------------------------------------------
    def _browse_excel(self) -> None:
        filetypes = [
            ("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"),
            ("CSV files", "*.csv"),
            ("All files", "*.*"),
        ]
        path = filedialog.askopenfilename(title="Select Excel/CSV file", filetypes=filetypes)
        if not path:
            return
        self.selected_excel_path = Path(path)
        self.excel_entry.config(state="normal")
        self.excel_entry.delete(0, tk.END)
        self.excel_entry.insert(0, str(self.selected_excel_path))
        self.excel_entry.config(state="readonly")
        self.excel_status_label.config(text=f"Selected: {self.selected_excel_path.name}", fg="#006644")
        self.gui_log(f"Selected input file: {self.selected_excel_path}")
        
        # Update column mapping status
        self.column_mapping_status.config(
            text="Type column names (e.g., 'Excel_First_Name') or column letters (e.g., 'B', 'C', 'AB') or click 'Show Available Columns'.",
            fg="#666666"
        )
        
        # Update preview
        self._update_preview()

    def _clear_excel(self) -> None:
        self.selected_excel_path = None
        self.excel_entry.config(state="normal")
        self.excel_entry.delete(0, tk.END)
        self.excel_entry.config(state="readonly")
        self.excel_status_label.config(text="No file selected.", fg="#666666")
        self.gui_log("Cleared Excel selection.")
        
        # Clear preview
        self._clear_preview()
    
    def _update_preview(self) -> None:
        """Update the preview with clients that will be processed."""
        if not self.selected_excel_path or not self.selected_excel_path.exists():
            self._clear_preview()
            return
        
        if not EXCEL_AVAILABLE:
            self._clear_preview()
            return
        
        try:
            # Load data (respecting filters)
            records = self._load_excel_data(self.selected_excel_path)
            
            if not records:
                self.preview_text.config(state="normal")
                self.preview_text.delete(1.0, tk.END)
                self.preview_text.insert(tk.END, "No data found in file.")
                self.preview_text.config(state="disabled")
                self.preview_count_label.config(text="No data found.", fg="#666666")
                return
            
            # Extract preview data (first 20 clients)
            preview_limit = 20
            preview_rows = []
            
            for idx, row in enumerate(records[:preview_limit], 1):
                normalized_row = self._normalize_row_dict(row)
                client_name = self._extract_client_name_fields(normalized_row).strip()
                dob = self._extract_dob_field(normalized_row).strip()
                dos = self._extract_dos_field(normalized_row).strip()
                
                if not dos:
                    # Try alternate format if DOS stored as Excel serial number
                    raw_dos = next((normalized_row[key] for key in normalized_row if "dos" in key or "date" in key), "")
                    dos = self._value_to_string(raw_dos)
                
                # Format preview row
                name_display = client_name if client_name else "[Missing Name]"
                dob_display = dob if dob else "[Missing DOB]"
                dos_display = dos if dos else "[Missing DOS]"
                
                # Truncate long names for display
                if len(name_display) > 35:
                    name_display = name_display[:32] + "..."
                
                preview_rows.append((idx, name_display, dob_display, dos_display))
            
            # Update preview text
            self.preview_text.config(state="normal")
            self.preview_text.delete(1.0, tk.END)
            
            # Header with better formatting
            self.preview_text.insert(tk.END, f"{'#':<4} {'Client Name':<38} {'DOB':<15} {'Date of Service':<20}\n")
            self.preview_text.insert(tk.END, "-" * 77 + "\n")
            
            # Preview rows
            for idx, name, dob, dos in preview_rows:
                self.preview_text.insert(tk.END, f"{idx:3d}. {name:<38} {dob:<15} {dos:<20}\n")
            
            if len(records) > preview_limit:
                self.preview_text.insert(tk.END, f"\n... and {len(records) - preview_limit} more client(s)\n")
            
            self.preview_text.config(state="disabled")
            
            # Update count label
            total_count = len(records)
            self.preview_count_label.config(
                text=f"Total clients to process: {total_count} | Showing first {min(preview_limit, total_count)}",
                fg="#006644"
            )
            
        except Exception as e:
            self.preview_text.config(state="normal")
            self.preview_text.delete(1.0, tk.END)
            self.preview_text.insert(tk.END, f"Error loading preview: {str(e)}")
            self.preview_text.config(state="disabled")
            self.preview_count_label.config(text="Error loading preview.", fg="#dc3545")
            self.gui_log(f"Error updating preview: {e}", level="ERROR")
    
    def _clear_preview(self) -> None:
        """Clear the preview display."""
        self.preview_text.config(state="normal")
        self.preview_text.delete(1.0, tk.END)
        self.preview_text.insert(tk.END, "No preview available. Select an Excel/CSV file to see preview.")
        self.preview_text.config(state="disabled")
        self.preview_count_label.config(text="No preview available.", fg="#666666")

    def _browse_output_excel(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Select output Excel file",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            return
        self.output_excel_path = Path(path)
        self.output_entry.config(state="normal")
        self.output_entry.delete(0, tk.END)
        self.output_entry.insert(0, str(self.output_excel_path))
        self.output_entry.config(state="readonly")
        self.output_status_label.config(text=f"Saving to: {self.output_excel_path}", fg="#006644")
        self.gui_log(f"Output path set: {self.output_excel_path}")

    def _clear_output_excel(self) -> None:
        self.output_excel_path = None
        self.output_entry.config(state="normal")
        self.output_entry.delete(0, tk.END)
        self.output_entry.config(state="readonly")
        self.output_status_label.config(text="No output file selected.", fg="#666666")
        self.gui_log("Cleared output Excel selection.")

    def _browse_pdf_template(self) -> None:
        """Browse for PDF template file."""
        filetypes = [("PDF files", "*.pdf"), ("All files", "*.*")]
        path = filedialog.askopenfilename(title="Select PDF Form Template", filetypes=filetypes)
        if not path:
            return
        self.pdf_template_path = Path(path)
        self.pdf_template_entry.config(state="normal")
        self.pdf_template_entry.delete(0, tk.END)
        self.pdf_template_entry.insert(0, str(self.pdf_template_path))
        self.pdf_template_entry.config(state="readonly")
        self.pdf_template_status_label.config(text=f"Selected: {self.pdf_template_path.name}", fg="#006644")
        self.gui_log(f"Selected PDF template: {self.pdf_template_path}")
        
        # Automatically list PDF fields when template is selected
        if PDFRW_AVAILABLE or PYPDF2_AVAILABLE:
            # List fields but don't show popup (just log to console)
            # User can click "List PDF Form Fields" button if they want to see the popup
            try:
                fields = self._list_pdf_form_fields()
                if fields:
                    self.gui_log(f"PDF template loaded. Found {len(fields)} form field(s).", level="INFO")
                    # Update configuration status if config window exists
                    self._update_pdf_config_status()
            except Exception as e:
                self.gui_log(f"Error listing PDF fields: {e}", level="WARNING")
    
    def _list_and_display_pdf_fields(self) -> None:
        """List and display all PDF form fields in a message box and log."""
        if not self.pdf_template_path or not self.pdf_template_path.exists():
            messagebox.showwarning("No Template", "Please select a PDF template first.")
            return
        
        if not PDFRW_AVAILABLE and not PYPDF2_AVAILABLE:
            messagebox.showerror("Missing Library", "PDF library is required to list PDF fields.\nInstall with: pip install pdfrw OR pip install PyPDF2")
            return
        
        try:
            fields = self._list_pdf_form_fields()
            
            # Also try to get more detailed information about the PDF structure
            pdf_info = self._analyze_pdf_structure()
            if pdf_info:
                self.gui_log(f"PDF Structure Analysis: {pdf_info}", level="INFO")
            
            if not fields:
                # Provide more helpful error message with diagnostic info
                error_msg = (
                    f"No fillable form fields detected in this PDF.\n\n"
                    f"PDF Structure: {pdf_info if pdf_info else 'Unable to analyze'}\n\n"
                    f"Possible reasons:\n"
                    f"1. The PDF uses XFA (XML Forms Architecture) which may require different handling\n"
                    f"2. The PDF fields are embedded in a non-standard format\n"
                    f"3. The PDF may need to be converted to a standard fillable PDF format\n"
                    f"4. The PDF might use JavaScript or dynamic form generation\n\n"
                    f"Try:\n"
                    f"1. Opening the PDF in Adobe Acrobat and saving it as a new PDF\n"
                    f"2. Using Adobe Acrobat to check if the form fields are properly named\n"
                    f"3. Checking if the PDF needs to be flattened and re-created as a fillable form\n\n"
                    f"Note: Even if fields are not detected, the bot may still be able to fill the PDF\n"
                    f"using field name patterns you provide. Check the log for more diagnostic information."
                )
                messagebox.showwarning("No Fields Detected", error_msg)
                self.gui_log("WARNING: No PDF form fields detected. PDF may still be fillable if fields are properly named.", level="WARNING")
                self.gui_log("TIP: You can still configure field mappings manually if you know the field names.", level="INFO")
                return
            
            # Display fields in message box
            fields_text = f"Found {len(fields)} PDF form field(s):\n\n"
            for i, field in enumerate(fields, 1):
                fields_text += f"{i}. {field}\n"
            
            # Show in message box (truncated if too many)
            if len(fields) > 50:
                fields_text = fields_text[:3000] + f"\n... and {len(fields) - 50} more fields"
                messagebox.showinfo("PDF Form Fields", fields_text)
            else:
                messagebox.showinfo("PDF Form Fields", fields_text)
            
            # Also log to GUI
            self.gui_log(f"Found {len(fields)} PDF form field(s):", level="INFO")
            for field in fields[:20]:  # Log first 20
                self.gui_log(f"  - {field}", level="INFO")
            if len(fields) > 20:
                self.gui_log(f"  ... and {len(fields) - 20} more fields", level="INFO")
            
            # Log instruction for configuration
            self.gui_log("\n" + "="*70, level="INFO")
            self.gui_log("TO CONFIGURE FIELD MAPPINGS:", level="INFO")
            self.gui_log("="*70, level="INFO")
            self.gui_log("1. Copy the PDF field names listed above", level="INFO")
            self.gui_log("2. Update the 'pdf_field_mapping_config' dictionary in the code", level="INFO")
            self.gui_log("3. Map each PDF field name to a client_data key:", level="INFO")
            self.gui_log("   Example: 'Patient Name': 'client_name'", level="INFO")
            self.gui_log("Available client_data keys: client_name, dob, dos, patient_member_id, etc.", level="INFO")
            self.gui_log("="*70, level="INFO")
            
        except Exception as e:
            self.log_error("Error listing PDF fields", exception=e, include_traceback=True)
            messagebox.showerror("Error", f"Error listing PDF fields:\n{e}")

    def _analyze_pdf_structure(self) -> str:
        """Analyze PDF structure to provide diagnostic information.
        
        Returns:
            str: Diagnostic information about PDF structure
        """
        info_parts = []
        
        try:
            if PDFRW_AVAILABLE and self.pdf_template_path and self.pdf_template_path.exists():
                template_pdf = pdfrw.PdfReader(str(self.pdf_template_path))  # type: ignore
                
                # Check for Root
                if template_pdf.Root:
                    info_parts.append("Has Root")
                    
                    # Check for AcroForm
                    if template_pdf.Root.AcroForm:
                        acro_form = template_pdf.Root.AcroForm
                        info_parts.append("Has AcroForm")
                        
                        # Check for Fields array
                        if acro_form.get('/Fields'):
                            fields_array = acro_form['/Fields']
                            info_parts.append(f"{len(fields_array)} Fields in AcroForm")
                        else:
                            info_parts.append("No Fields array in AcroForm")
                        
                        # Check for XFA
                        if acro_form.get('/XFA'):
                            info_parts.append("Uses XFA (XML Forms Architecture) - may not be fully supported")
                    else:
                        info_parts.append("No AcroForm found")
                else:
                    info_parts.append("No Root structure")
                
                # Check pages
                if template_pdf.pages:
                    info_parts.append(f"{len(template_pdf.pages)} page(s)")
                    
                    # Check annotations on first page
                    if template_pdf.pages[0].get('/Annots'):
                        annots = template_pdf.pages[0]['/Annots']
                        info_parts.append(f"{len(annots)} annotation(s) on first page")
                        
                        # Count widget annotations
                        widget_count = 0
                        for annot in annots:
                            if annot:
                                if isinstance(annot, pdfrw.PdfDict):  # type: ignore
                                    annot_obj = annot
                                else:
                                    annot_obj = pdfrw.PdfDict(annot)  # type: ignore
                                
                                subtype = annot_obj.get('/Subtype')
                                if subtype:
                                    if isinstance(subtype, str) and subtype == '/Widget':
                                        widget_count += 1
                                    else:
                                        # Check if it's a PdfName object (compare string representation)
                                        subtype_str = str(subtype)
                                        if '/Widget' in subtype_str or 'Widget' in subtype_str:
                                            widget_count += 1
                                        # Also try direct comparison
                                        try:
                                            if hasattr(pdfrw, 'PdfName') and subtype == pdfrw.PdfName('Widget'):  # type: ignore
                                                widget_count += 1
                                        except Exception:
                                            pass
                        
                        if widget_count > 0:
                            info_parts.append(f"{widget_count} Widget annotation(s) on first page")
                    else:
                        info_parts.append("No annotations on first page")
            
        except Exception as e:
            info_parts.append(f"Error analyzing: {e}")
        
        return ", ".join(info_parts) if info_parts else "Unable to analyze PDF structure"
    
    def _clear_pdf_template(self) -> None:
        """Clear PDF template selection."""
        self.pdf_template_path = None
        self.pdf_template_entry.config(state="normal")
        self.pdf_template_entry.delete(0, tk.END)
        self.pdf_template_entry.config(state="readonly")
        self.pdf_template_status_label.config(text="No template selected.", fg="#666666")
        self.gui_log("Cleared PDF template selection.")

    def _browse_pdf_audit_excel(self) -> None:
        """Browse for audit results Excel file."""
        filetypes = [("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")]
        path = filedialog.askopenfilename(title="Select Audit Results Excel", filetypes=filetypes)
        if not path:
            return
        self.pdf_audit_excel_path = Path(path)
        self.pdf_audit_excel_entry.config(state="normal")
        self.pdf_audit_excel_entry.delete(0, tk.END)
        self.pdf_audit_excel_entry.insert(0, str(self.pdf_audit_excel_path))
        self.pdf_audit_excel_entry.config(state="readonly")
        self.pdf_audit_excel_status_label.config(text=f"Selected: {self.pdf_audit_excel_path.name}", fg="#006644")
        self.gui_log(f"Selected audit Excel: {self.pdf_audit_excel_path}")
        
        # Update preview of clients that need refiling
        self._update_pdf_preview()

    def _clear_pdf_audit_excel(self) -> None:
        """Clear audit Excel selection."""
        self.pdf_audit_excel_path = None
        self.pdf_audit_excel_entry.config(state="normal")
        self.pdf_audit_excel_entry.delete(0, tk.END)
        self.pdf_audit_excel_entry.config(state="readonly")
        self.pdf_audit_excel_status_label.config(text="No audit Excel selected.", fg="#666666")
        self.gui_log("Cleared audit Excel selection.")

    def _browse_pdf_output_folder(self) -> None:
        """Browse for PDF output folder."""
        path = filedialog.askdirectory(title="Select Output Folder for Filled PDFs")
        if not path:
            return
        self.pdf_output_folder = Path(path)
        self.pdf_output_folder_entry.config(state="normal")
        self.pdf_output_folder_entry.delete(0, tk.END)
        self.pdf_output_folder_entry.insert(0, str(self.pdf_output_folder))
        self.pdf_output_folder_entry.config(state="readonly")
        self.pdf_output_folder_status_label.config(text=f"Selected: {self.pdf_output_folder.name}", fg="#006644")
        self.gui_log(f"Selected PDF output folder: {self.pdf_output_folder}")

    def _clear_pdf_output_folder(self) -> None:
        """Clear PDF output folder selection."""
        self.pdf_output_folder = None
        self.pdf_output_folder_entry.config(state="normal")
        self.pdf_output_folder_entry.delete(0, tk.END)
        self.pdf_output_folder_entry.config(state="readonly")
        self.pdf_output_folder_status_label.config(text="No output folder selected.", fg="#666666")
        self.gui_log("Cleared PDF output folder selection.")
    
    def _browse_medicare_excel_template(self) -> None:
        """Browse for Medicare Excel template file."""
        filetypes = [("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")]
        path = filedialog.askopenfilename(title="Select Medicare Excel Template", filetypes=filetypes)
        if not path:
            return
        self.medicare_excel_template_path = Path(path)
        self.medicare_excel_template_entry.config(state="normal")
        self.medicare_excel_template_entry.delete(0, tk.END)
        self.medicare_excel_template_entry.insert(0, str(self.medicare_excel_template_path))
        self.medicare_excel_template_entry.config(state="readonly")
        self.medicare_excel_template_status_label.config(text=f"Selected: {self.medicare_excel_template_path.name}", fg="#006644")
        self.gui_log(f"Selected Medicare Excel template: {self.medicare_excel_template_path}")
    
    def _clear_medicare_excel_template(self) -> None:
        """Clear Medicare Excel template selection."""
        self.medicare_excel_template_path = None
        self.medicare_excel_template_entry.config(state="normal")
        self.medicare_excel_template_entry.delete(0, tk.END)
        self.medicare_excel_template_entry.config(state="readonly")
        self.medicare_excel_template_status_label.config(text="No template selected.", fg="#666666")
        self.gui_log("Cleared Medicare Excel template selection.")
    
    def _browse_medicare_excel_audit_excel(self) -> None:
        """Browse for audit results Excel file for Medicare Excel filling."""
        filetypes = [("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")]
        path = filedialog.askopenfilename(title="Select Audit Results Excel", filetypes=filetypes)
        if not path:
            return
        self.medicare_excel_audit_excel_path = Path(path)
        self.medicare_excel_audit_excel_entry.config(state="normal")
        self.medicare_excel_audit_excel_entry.delete(0, tk.END)
        self.medicare_excel_audit_excel_entry.insert(0, str(self.medicare_excel_audit_excel_path))
        self.medicare_excel_audit_excel_entry.config(state="readonly")
        self.medicare_excel_audit_excel_status_label.config(text=f"Selected: {self.medicare_excel_audit_excel_path.name}", fg="#006644")
        self.gui_log(f"Selected audit Excel for Medicare Excel: {self.medicare_excel_audit_excel_path}")
        
        # Try to load and show available columns
        if EXCEL_AVAILABLE:
            try:
                df = pd.read_excel(self.medicare_excel_audit_excel_path)  # type: ignore
                columns = df.columns.tolist()
                self.gui_log(f"Available columns in Excel: {', '.join(columns)}", level="INFO")
            except Exception as exc:
                self.gui_log(f"Could not read Excel columns: {exc}", level="WARNING")
    
    def _clear_medicare_excel_audit_excel(self) -> None:
        """Clear Medicare Excel audit Excel selection."""
        self.medicare_excel_audit_excel_path = None
        self.medicare_excel_audit_excel_entry.config(state="normal")
        self.medicare_excel_audit_excel_entry.delete(0, tk.END)
        self.medicare_excel_audit_excel_entry.config(state="readonly")
        self.medicare_excel_audit_excel_status_label.config(text="No audit Excel selected.", fg="#666666")
        self.gui_log("Cleared Medicare Excel audit Excel selection.")
    
    def _browse_medicare_excel_output(self) -> None:
        """Browse for Medicare Excel output file."""
        filetypes = [("Excel files", "*.xlsx"), ("All files", "*.*")]
        path = filedialog.asksaveasfilename(title="Select Output File for Filled Medicare Excel",
                                           filetypes=filetypes, defaultextension=".xlsx")
        if not path:
            return
        self.medicare_excel_output_path = Path(path)
        self.medicare_excel_output_entry.config(state="normal")
        self.medicare_excel_output_entry.delete(0, tk.END)
        self.medicare_excel_output_entry.insert(0, str(self.medicare_excel_output_path))
        self.medicare_excel_output_entry.config(state="readonly")
        self.medicare_excel_output_status_label.config(text=f"Saving to: {self.medicare_excel_output_path.name}", fg="#006644")
        self.gui_log(f"Medicare Excel output path set: {self.medicare_excel_output_path}")
    
    def _clear_medicare_excel_output(self) -> None:
        """Clear Medicare Excel output selection."""
        self.medicare_excel_output_path = None
        self.medicare_excel_output_entry.config(state="normal")
        self.medicare_excel_output_entry.delete(0, tk.END)
        self.medicare_excel_output_entry.config(state="readonly")
        self.medicare_excel_output_status_label.config(text="No output file selected.", fg="#666666")
        self.gui_log("Cleared Medicare Excel output selection.")

    def _update_pdf_preview(self) -> None:
        """Update preview of clients that need refiling (if preview widget exists)."""
        # This can be added later if needed
        pass

    def _apply_row_range(self) -> None:
        start_val = self.start_row_entry.get().strip()
        end_val = self.end_row_entry.get().strip()
        try:
            start_row = int(start_val) if start_val else None
            end_row = int(end_val) if end_val else None
            if start_row is not None and start_row <= 0:
                raise ValueError("Start row must be positive.")
            if end_row is not None and end_row <= 0:
                raise ValueError("End row must be positive.")
            if start_row is not None and end_row is not None and end_row < start_row:
                raise ValueError("End row must be greater than or equal to start row.")
            self.row_range = (start_row, end_row)
            if start_row or end_row:
                text = f"Row range: {start_row or 1} - {end_row or 'end'}"
            else:
                text = "Row range: All rows"
            self.row_status_label.config(text=text, fg="#006644")
            self.gui_log(f"Row range set to: {text}")
            
            # Refresh preview if Excel file is loaded
            if self.selected_excel_path:
                self._update_preview()
        except ValueError as exc:
            messagebox.showerror("Invalid Row Range", str(exc))

    # ------------------------------------------------------------------
    # Thread management
    # ------------------------------------------------------------------
    def _track_thread(self, thread: threading.Thread) -> None:
        with self._thread_lock:
            self._active_threads.add(thread)

    def _release_thread(self, thread: threading.Thread) -> None:
        with self._thread_lock:
            self._active_threads.discard(thread)

    # ------------------------------------------------------------------
    # Audit workflow
    # ------------------------------------------------------------------
    def _start_audit(self) -> None:
        if self._shutdown_requested:
            return
        if not self.username_entry.get().strip() or not self.password_entry.get().strip():
            messagebox.showwarning("Missing Credentials", "Please provide username and password.")
            return
        if not self.selected_excel_path:
            messagebox.showwarning("No Data", "Please select an Excel/CSV file before starting the audit.")
            return
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Missing Dependency", "Pandas and openpyxl are required to load Excel/CSV files and respect filters. Install with: pip install pandas openpyxl")
            return
        if not SELENIUM_AVAILABLE:
            messagebox.showerror(
                "Missing Dependency",
                "Selenium is required for web automation.\nInstall with: pip install selenium webdriver-manager"
            )
            return

        self.gui_log("Starting audit workflow...", level="INFO")
        self.update_status("Audit started...", "#ff9500")
        self.stop_requested = False
        self.start_button.config(state="disabled")
        self.stop_button.config(state="normal")

        audit_thread = threading.Thread(target=self._run_audit_workflow, name="MedicareAudit", daemon=False)
        self._track_thread(audit_thread)
        audit_thread.start()

    def _run_audit_workflow(self) -> None:
        try:
            if not self._ensure_logged_in():
                self.gui_log("Cannot start audit without a valid TherapyNotes session.", level="ERROR")
                return

            self.gui_log("Loading client data...", level="INFO")
            try:
                data = self._load_excel_data(self.selected_excel_path)
            except Exception as exc:
                self.gui_log(f"Failed to load Excel data: {exc}", level="ERROR")
                if self.root:
                    self.root.after(0, lambda: messagebox.showerror("Excel Error", f"Failed to load data:\n{exc}"))
                return

            total_rows = len(data)
            if total_rows == 0:
                self.gui_log("No client rows found in the selected file.", level="WARNING")
                if self.root:
                    self.root.after(0, lambda: messagebox.showwarning("No Data", "The selected file contains no rows to audit."))
                return

            self.gui_log(f"Loaded {total_rows} rows from input file.", level="INFO")
            self.update_status(f"Loaded {total_rows} clients. Beginning audit...", "#ff9500")

            self.audit_results = []
            audited = 0

            for idx, row in enumerate(data, 1):
                if self.check_stop_requested():
                    self.gui_log("Audit stopped by user.", level="WARNING")
                    break

                try:
                    self._process_client_row(idx, row, total_rows)
                    audited += 1
                except Exception as exc:
                    self.log_error(f"Unexpected error auditing row {idx}: {exc}", exception=exc)

                self._update_processing_stats(audited, total_rows - idx)

            if self.audit_results:
                self._save_audit_results()

            self.gui_log(f"Audit complete. Total rows reviewed: {audited}", level="INFO")
            if not self.check_stop_requested():
                self.update_status(f"Audit complete. Reviewed {audited} row(s).", "#006644")
        finally:
            current_thread = threading.current_thread()
            self._release_thread(current_thread)
            if self.root:
                self.root.after(0, self._on_audit_finished)

    def _update_processing_stats(self, audited: int, pending: int) -> None:
        if self.processing_stats_label:
            self.processing_stats_label.config(text=f"Clients audited: {audited} | Pending: {pending}")

    def _on_audit_finished(self) -> None:
        self.start_button.config(state="normal")
        self.stop_button.config(state="disabled")
        if self.stop_requested:
            self.update_status("Audit stopped by user.", "#a60000")
        self.stop_requested = False

    def request_stop(self) -> None:
        if not self.stop_requested:
            self.stop_requested = True
            self.gui_log("Stop requested. Finishing current client...", level="WARNING")
            self.update_status("Stop requested...", "#a60000")
            self.stop_button.config(state="disabled")
    
    def _start_pdf_filling(self) -> None:
        """Start the PDF form filling workflow."""
        if self._shutdown_requested:
            return
        
        # Validate inputs
        if not self.pdf_template_path or not self.pdf_template_path.exists():
            messagebox.showwarning("Missing Template", "Please select a PDF form template.")
            return
        
        if not self.pdf_audit_excel_path or not self.pdf_audit_excel_path.exists():
            messagebox.showwarning("Missing Audit Excel", "Please select the audit results Excel file.")
            return
        
        if not self.pdf_output_folder:
            messagebox.showwarning("Missing Output Folder", "Please select an output folder for filled PDFs.")
            return
        
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Missing Dependency", "Pandas is required to read Excel files. Install with: pip install pandas openpyxl")
            return
        
        if not PDFRW_AVAILABLE and not PYPDF2_AVAILABLE:
            messagebox.showerror("Missing Dependency", "PDF library required. Install with: pip install pdfrw OR pip install PyPDF2")
            return
        
        self.gui_log("Starting PDF form filling workflow...", level="INFO")
        self.pdf_status_label.config(text="PDF filling started...", fg="#ff9500")
        self.pdf_filling_stop_requested = False
        self.pdf_start_button.config(state="disabled")
        self.pdf_stop_button.config(state="normal")
        
        pdf_filling_thread = threading.Thread(target=self._run_pdf_filling_workflow, name="PDFFilling", daemon=False)
        self._track_thread(pdf_filling_thread)
        pdf_filling_thread.start()
    
    def _stop_pdf_filling(self) -> None:
        """Stop the PDF form filling workflow."""
        if not self.pdf_filling_stop_requested:
            self.pdf_filling_stop_requested = True
            self.gui_log("Stop requested. Finishing current PDF...", level="WARNING")
            self.pdf_status_label.config(text="Stop requested...", fg="#a60000")
            self.pdf_stop_button.config(state="disabled")
    
    def _start_medicare_excel_filling(self) -> None:
        """Start the Medicare Excel filling workflow."""
        if self._shutdown_requested:
            return
        
        # Validate inputs
        if not self.medicare_excel_template_path or not self.medicare_excel_template_path.exists():
            messagebox.showwarning("Missing Template", "Please select the Medicare Excel template file.")
            return
        
        if not self.medicare_excel_audit_excel_path or not self.medicare_excel_audit_excel_path.exists():
            messagebox.showwarning("Missing Audit Excel", "Please select the audit results Excel file.")
            return
        
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Missing Dependency", "Pandas and openpyxl are required to read/write Excel files. Install with: pip install pandas openpyxl")
            return
        
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Missing Dependency", "openpyxl is required to write Excel files. Install with: pip install openpyxl")
            return
        
        self.gui_log("Starting Medicare Excel filling workflow...", level="INFO")
        self.medicare_excel_status_label.config(text="Medicare Excel filling started...", fg="#ff9500")
        self.medicare_excel_filling_stop_requested = False
        self.medicare_excel_start_button.config(state="disabled")
        self.medicare_excel_stop_button.config(state="normal")
        
        medicare_excel_thread = threading.Thread(target=self._run_medicare_excel_filling_workflow, name="MedicareExcelFilling", daemon=False)
        self._track_thread(medicare_excel_thread)
        medicare_excel_thread.start()
    
    def _stop_medicare_excel_filling(self) -> None:
        """Stop the Medicare Excel filling workflow."""
        if not self.medicare_excel_filling_stop_requested:
            self.medicare_excel_filling_stop_requested = True
            self.gui_log("Stop requested. Finishing current row...", level="WARNING")
            self.medicare_excel_status_label.config(text="Stop requested...", fg="#a60000")
            self.medicare_excel_stop_button.config(state="disabled")
    
    def _run_medicare_excel_filling_workflow(self) -> None:
        """Main Medicare Excel filling workflow."""
        try:
            # Apply column mappings from GUI (in case user didn't click Apply)
            if hasattr(self, 'medicare_excel_col_patient_name'):
                self.medicare_excel_column_mappings["Patient's Name (Column A)"] = self.medicare_excel_col_patient_name.get().strip().upper()
                self.medicare_excel_column_mappings["Patient's HICN/MBI (Column B)"] = self.medicare_excel_col_hicn.get().strip().upper()
                self.medicare_excel_column_mappings["Date of Service (Column C)"] = self.medicare_excel_col_dos.get().strip().upper()
                self.medicare_excel_column_mappings["Procedure Code (Column E)"] = self.medicare_excel_col_service_code.get().strip().upper()
            
            # Apply custom note from GUI (in case user didn't click Apply)
            if hasattr(self, 'medicare_excel_custom_note_text'):
                self.medicare_excel_custom_note = self.medicare_excel_custom_note_text.get("1.0", tk.END).strip()
            
            # Load audit results Excel
            self.gui_log("Loading audit results Excel...", level="INFO")
            try:
                df = pd.read_excel(self.medicare_excel_audit_excel_path)  # type: ignore
            except Exception as exc:
                self.gui_log(f"Failed to load audit Excel: {exc}", level="ERROR")
                if self.root:
                    self.root.after(0, lambda: messagebox.showerror("Excel Error", f"Failed to load audit Excel:\n{exc}"))
                return
            
            # Validate column letter mappings
            missing_columns = []
            invalid_columns = []
            for field_name, col_letter in self.medicare_excel_column_mappings.items():
                if not col_letter:
                    missing_columns.append(f"{field_name} (no column letter specified)")
                else:
                    col_index = self._column_letter_to_index(col_letter)
                    if col_index is None:
                        invalid_columns.append(f"{field_name} (invalid column letter: '{col_letter}')")
                    elif col_index >= len(df.columns):
                        invalid_columns.append(f"{field_name} (column '{col_letter}' is beyond available columns)")
            
            if missing_columns or invalid_columns:
                error_msg = ""
                if missing_columns:
                    error_msg += "The following fields have no column letter specified:\n" + "\n".join(f"  - {col}" for col in missing_columns) + "\n\n"
                if invalid_columns:
                    error_msg += "The following column letters are invalid:\n" + "\n".join(f"  - {col}" for col in invalid_columns) + "\n\n"
                error_msg += f"Available columns in Excel:\n"
                for i, col_name in enumerate(df.columns[:10]):  # Show first 10 columns
                    col_letter = self._index_to_column_letter(i)
                    error_msg += f"  - {col_letter}: {col_name}\n"
                if len(df.columns) > 10:
                    error_msg += f"  ... and {len(df.columns) - 10} more columns\n"
                self.gui_log(f"ERROR: Column mapping issues: {', '.join(missing_columns + invalid_columns)}", level="ERROR")
                if self.root:
                    self.root.after(0, lambda: messagebox.showerror("Column Mapping Error", error_msg))
                return
            
            # Filter for "Needs Refile" clients
            if "Status" not in df.columns:
                self.gui_log("ERROR: 'Status' column not found in audit Excel.", level="ERROR")
                if self.root:
                    self.root.after(0, lambda: messagebox.showerror("Invalid Excel", "The audit Excel file must have a 'Status' column."))
                return
            
            needs_refile = df[df["Status"] == "Needs Refile"].copy()
            
            if len(needs_refile) == 0:
                self.gui_log("No clients with 'Needs Refile' status found in audit Excel.", level="WARNING")
                if self.root:
                    self.root.after(0, lambda: messagebox.showinfo("No Clients", "No clients with 'Needs Refile' status found."))
                return
            
            self.gui_log(f"Found {len(needs_refile)} client(s) with 'Needs Refile' status.", level="INFO")
            self.medicare_excel_status_label.config(text=f"Found {len(needs_refile)} client(s) to process...", fg="#ff9500")
            
            # Load Medicare Excel template using openpyxl
            self.gui_log("Loading Medicare Excel template...", level="INFO")
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(self.medicare_excel_template_path))
                ws = wb.active
            except Exception as exc:
                self.gui_log(f"Failed to load Medicare Excel template: {exc}", level="ERROR")
                if self.root:
                    self.root.after(0, lambda: messagebox.showerror("Template Error", f"Failed to load Medicare Excel template:\n{exc}"))
                return
            
            # Find starting row (should be row 3, but verify by checking header row)
            start_row = 3
            header_row = 2  # Row 2 contains headers
            # Verify headers are in row 2
            if ws.cell(row=header_row, column=1).value and "Patient's name" in str(ws.cell(row=header_row, column=1).value):
                start_row = 3
            else:
                # Try to find header row
                for row_num in range(1, min(10, ws.max_row + 1)):
                    cell_value = ws.cell(row=row_num, column=1).value
                    if cell_value and "Patient's name" in str(cell_value):
                        start_row = row_num + 1
                        break
            
            self.gui_log(f"Starting to fill Medicare Excel from row {start_row}...", level="INFO")
            
            # Process each client
            filled_count = 0
            current_row = start_row
            
            # Get column indices (0-based) from column letters
            patient_name_letter = self.medicare_excel_column_mappings.get("Patient's Name (Column A)", "")
            hicn_letter = self.medicare_excel_column_mappings.get("Patient's HICN/MBI (Column B)", "")
            dos_letter = self.medicare_excel_column_mappings.get("Date of Service (Column C)", "")
            service_code_letter = self.medicare_excel_column_mappings.get("Procedure Code (Column E)", "")
            
            patient_name_idx = self._column_letter_to_index(patient_name_letter) if patient_name_letter else None
            hicn_idx = self._column_letter_to_index(hicn_letter) if hicn_letter else None
            dos_idx = self._column_letter_to_index(dos_letter) if dos_letter else None
            service_code_idx = self._column_letter_to_index(service_code_letter) if service_code_letter else None
            
            # Iterate through DataFrame using position index
            for pos_idx in range(len(needs_refile)):
                if self.medicare_excel_filling_stop_requested:
                    self.gui_log("Medicare Excel filling stopped by user.", level="WARNING")
                    break
                
                try:
                    # Extract data using column indices from the DataFrame
                    client_name = ""
                    if patient_name_idx is not None and patient_name_idx < len(needs_refile.columns):
                        client_name = str(needs_refile.iloc[pos_idx, patient_name_idx]).strip()
                    
                    patient_member_id = ""
                    if hicn_idx is not None and hicn_idx < len(needs_refile.columns):
                        patient_member_id = str(needs_refile.iloc[pos_idx, hicn_idx]).strip()
                    
                    date_of_service = ""
                    if dos_idx is not None and dos_idx < len(needs_refile.columns):
                        date_of_service = str(needs_refile.iloc[pos_idx, dos_idx]).strip()
                    
                    service_code = ""
                    if service_code_idx is not None and service_code_idx < len(needs_refile.columns):
                        service_code = str(needs_refile.iloc[pos_idx, service_code_idx]).strip()
                    
                    if not client_name:
                        self.gui_log(f"Skipping row {pos_idx + 1}: Missing client name (column: {patient_name_letter}).", level="WARNING")
                        continue
                    
                    # Generate reasoning note - use custom note if provided
                    if self.medicare_excel_custom_note:
                        reasoning_note = self.medicare_excel_custom_note
                    else:
                        # Default note if no custom note provided
                        reasoning_note = "Modifier correction needed as per correct session medium"
                    
                    # Fill Medicare Excel row
                    # Column A: Patient's name
                    ws.cell(row=current_row, column=1, value=client_name)
                    # Column B: Patient's HICN/MBI
                    ws.cell(row=current_row, column=2, value=patient_member_id)
                    # Column C: Date of service
                    ws.cell(row=current_row, column=3, value=date_of_service)
                    # Column D: ICN (leave blank)
                    ws.cell(row=current_row, column=4, value="")
                    # Column E: Procedure Code
                    ws.cell(row=current_row, column=5, value=service_code)
                    # Column F: Explain Correction Needed
                    ws.cell(row=current_row, column=6, value=reasoning_note)
                    
                    filled_count += 1
                    current_row += 1
                    
                    self.gui_log(f"Filled row for {client_name} (row {current_row - 1})", level="INFO")
                    
                    # Update stats
                    pending = len(needs_refile) - (pos_idx + 1)
                    if self.root:
                        self.root.after(0, lambda c=filled_count, p=pending: 
                                       self._update_medicare_excel_stats(c, p))
                    
                except Exception as exc:
                    self.log_error(f"Error filling Medicare Excel for row {pos_idx + 1}: {exc}", exception=exc)
                    continue
            
            # Save the filled Medicare Excel
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            if self.medicare_excel_output_path:
                output_path = self.medicare_excel_output_path
            else:
                output_dir = SCRIPT_DIR / "output"
                output_dir.mkdir(parents=True, exist_ok=True)
                output_path = output_dir / f"medicare_refiling_excel_{timestamp}.xlsx"
            
            try:
                wb.save(str(output_path))
                self.gui_log(f"✅ Medicare Excel saved to {output_path}", level="INFO")
                if self.root:
                    self.root.after(0, lambda: self.medicare_excel_status_label.config(
                        text=f"Medicare Excel saved: {output_path.name}", fg="#28a745"))
            except Exception as exc:
                self.log_error(f"Failed to save Medicare Excel: {exc}", exception=exc)
                if self.root:
                    self.root.after(0, lambda: messagebox.showerror("Save Error", f"Failed to save Medicare Excel:\n{exc}"))
                return
            
            self.gui_log(f"Medicare Excel filling complete. Filled {filled_count} row(s).", level="INFO")
            if not self.medicare_excel_filling_stop_requested:
                if self.root:
                    self.root.after(0, lambda: self.medicare_excel_status_label.config(
                        text=f"Medicare Excel filling complete. Filled {filled_count} row(s).", fg="#28a745"))
        except Exception as exc:
            self.log_error("Unexpected error in Medicare Excel filling workflow", exception=exc)
        finally:
            current_thread = threading.current_thread()
            self._release_thread(current_thread)
            if self.root:
                self.root.after(0, self._on_medicare_excel_filling_finished)
    
    def _generate_reasoning_note(self, original_modifier: str, expected_modifier: str, notes: str) -> str:
        """Generate a reasoning note for the Medicare Excel form.
        
        Args:
            original_modifier: Original modifier from audit (e.g., "95", "93")
            expected_modifier: Expected modifier from audit (e.g., "95", "93")
            notes: Notes from audit results
            
        Returns:
            A brief reasoning note, e.g., "Modifier change from 95 to 93, as per correct session medium"
        """
        original_mod = original_modifier.strip().upper()
        expected_mod = expected_modifier.strip().upper()
        
        if original_mod and expected_mod and original_mod != expected_mod:
            return f"Modifier change from {original_mod} to {expected_mod}, as per correct session medium"
        elif notes:
            # Use notes if available, but make it brief
            if len(notes) > 100:
                return notes[:97] + "..."
            return notes
        else:
            return "Modifier correction needed as per correct session medium"
    
    def _update_medicare_excel_stats(self, filled: int, pending: int) -> None:
        """Update Medicare Excel filling statistics in GUI."""
        if self.medicare_excel_stats_label:
            self.medicare_excel_stats_label.config(text=f"Filled: {filled} | Pending: {pending}")
    
    def _on_medicare_excel_filling_finished(self) -> None:
        """Reset GUI elements when Medicare Excel filling is complete."""
        self.medicare_excel_start_button.config(state="normal")
        self.medicare_excel_stop_button.config(state="disabled")
        if self.medicare_excel_filling_stop_requested:
            self.medicare_excel_status_label.config(text="Medicare Excel filling stopped by user.", fg="#a60000")
        self.medicare_excel_filling_stop_requested = False
    
    def _run_pdf_filling_workflow(self) -> None:
        """Main PDF form filling workflow."""
        try:
            # Load audit results Excel
            self.gui_log("Loading audit results Excel...", level="INFO")
            try:
                df = pd.read_excel(self.pdf_audit_excel_path)  # type: ignore
            except Exception as exc:
                self.gui_log(f"Failed to load audit Excel: {exc}", level="ERROR")
                if self.root:
                    self.root.after(0, lambda: messagebox.showerror("Excel Error", f"Failed to load audit Excel:\n{exc}"))
                return
            
            # Filter for "Needs Refile" clients
            if "Status" not in df.columns:
                self.gui_log("ERROR: 'Status' column not found in audit Excel.", level="ERROR")
                if self.root:
                    self.root.after(0, lambda: messagebox.showerror("Invalid Excel", "The audit Excel file must have a 'Status' column."))
                return
            
            needs_refile = df[df["Status"] == "Needs Refile"].copy()
            
            if len(needs_refile) == 0:
                self.gui_log("No clients with 'Needs Refile' status found in audit Excel.", level="WARNING")
                if self.root:
                    self.root.after(0, lambda: messagebox.showinfo("No Clients", "No clients with 'Needs Refile' status found."))
                return
            
            self.gui_log(f"Found {len(needs_refile)} client(s) that need refiling.", level="INFO")
            self.pdf_status_label.config(text=f"Found {len(needs_refile)} client(s) to process...", fg="#ff9500")
            
            # Process each client
            filled_count = 0
            total_count = len(needs_refile)
            
            for idx, row in needs_refile.iterrows():
                if self.pdf_filling_stop_requested:
                    self.gui_log("PDF filling stopped by user.", level="WARNING")
                    break
                
                try:
                    client_name = str(row.get("Client Name", "")).strip()
                    if not client_name:
                        self.gui_log(f"Row {idx + 1} skipped - missing client name.", level="WARNING")
                        continue
                    
                    self.gui_log(f"\n{'=' * 70}\nProcessing {filled_count + 1}/{total_count}: {client_name}\n{'=' * 70}", level="INFO")
                    
                    # Fill PDF form for this client
                    success = self._fill_pdf_form_for_client(row, client_name)
                    
                    if success:
                        filled_count += 1
                        self.gui_log(f"✅ Successfully filled PDF for {client_name}", level="INFO")
                    else:
                        self.gui_log(f"❌ Failed to fill PDF for {client_name}", level="ERROR")
                    
                    # Update stats
                    self._update_pdf_stats(filled_count, total_count - (filled_count + 1))
                    
                except Exception as exc:
                    self.log_error(f"Unexpected error filling PDF for row {idx + 1}: {exc}", exception=exc)
            
            self.gui_log(f"\nPDF filling complete. Filled {filled_count} PDF(s) out of {total_count} client(s).", level="INFO")
            if not self.pdf_filling_stop_requested:
                self.pdf_status_label.config(text=f"PDF filling complete. Filled {filled_count} PDF(s).", fg="#28a745")
        finally:
            current_thread = threading.current_thread()
            self._release_thread(current_thread)
            if self.root:
                self.root.after(0, self._on_pdf_filling_finished)
    
    def _update_pdf_stats(self, filled: int, pending: int) -> None:
        """Update PDF filling statistics."""
        if self.pdf_stats_label:
            self.pdf_stats_label.config(text=f"PDFs filled: {filled} | Pending: {pending}")
    
    def _on_pdf_filling_finished(self) -> None:
        """Called when PDF filling workflow finishes."""
        self.pdf_start_button.config(state="normal")
        self.pdf_stop_button.config(state="disabled")
        if self.pdf_filling_stop_requested:
            self.pdf_status_label.config(text="PDF filling stopped by user.", fg="#a60000")
        self.pdf_filling_stop_requested = False
    
    def _fill_pdf_form_for_client(self, client_row: "pd.Series", client_name: str) -> bool:
        """Fill PDF form for a single client.
        
        Args:
            client_row: Pandas Series containing client data from audit Excel
            client_name: Client name (used for folder creation)
        
        Returns:
            bool: True if PDF was filled and saved successfully, False otherwise
        """
        try:
            if not self.pdf_template_path or not self.pdf_template_path.exists():
                self.gui_log(f"ERROR: PDF template not found: {self.pdf_template_path}", level="ERROR")
                return False
            
            if not self.pdf_output_folder:
                self.gui_log("ERROR: Output folder not selected.", level="ERROR")
                return False
            
            # Create client folder
            client_folder = self._create_client_folder(client_name)
            if not client_folder:
                self.gui_log(f"ERROR: Could not create folder for {client_name}", level="ERROR")
                return False
            
            # Extract client data from Excel row
            client_data = self._extract_client_data_for_pdf(client_row)
            
            # Fill PDF form
            filled_pdf_path = self._fill_pdf_form(client_data, client_folder, client_name)
            
            if not filled_pdf_path:
                self.gui_log(f"ERROR: Could not fill PDF for {client_name}", level="ERROR")
                return False
            
            # Flatten PDF (make non-editable)
            flattened_pdf_path = self._flatten_pdf(filled_pdf_path, client_folder, client_name)
            
            if not flattened_pdf_path:
                self.gui_log(f"WARNING: Could not flatten PDF for {client_name}, but filled PDF saved.", level="WARNING")
                # Use filled PDF if flattening failed
                flattened_pdf_path = filled_pdf_path
            
            self.gui_log(f"✅ PDF filled and saved: {flattened_pdf_path}", level="INFO")
            return True
            
        except Exception as e:
            self.log_error(f"Error filling PDF for {client_name}", exception=e, include_traceback=True)
            return False
    
    def _extract_client_data_for_pdf(self, client_row: "pd.Series") -> Dict[str, Any]:
        """Extract client data from Excel row for PDF form filling.
        
        This method uses robust column name matching to extract data from the Excel file,
        including handling variations in column names (e.g., "Patient Member #", "MBI", etc.).
        
        Args:
            client_row: Pandas Series containing client data (from audit results Excel)
        
        Returns:
            dict: Dictionary containing client data for PDF form
        """
        # Normalize the row to handle column name variations
        normalized_row = self._normalize_row_dict(client_row.to_dict())
        
        # Extract data using robust synonym-based extraction (same as audit workflow)
        client_name = self._extract_client_name_fields(normalized_row).strip()
        dob = self._extract_dob_field(normalized_row).strip()
        dos = self._extract_dos_field(normalized_row).strip()
        patient_member_id = self._extract_patient_member_id_field(normalized_row).strip()
        
        # Also try to get data from common column names in audit results Excel
        # (These are the column names used in the audit output Excel file)
        client_data = {
            "client_name": client_name or str(client_row.get("Client Name", "")).strip(),
            "dob": dob or str(client_row.get("DOB", "")).strip(),
            "dos": dos or str(client_row.get("Date of Service", "")).strip(),
            "patient_member_id": patient_member_id or str(client_row.get("Patient Member #", "")).strip(),
            "session_medium": str(client_row.get("Session Medium", "")).strip(),
            "original_modifier": str(client_row.get("Original Modifier", "")).strip(),
            "expected_modifier": str(client_row.get("Expected Modifier", "")).strip(),
            "status": str(client_row.get("Status", "")).strip(),
            "notes": str(client_row.get("Notes", "")).strip(),
        }
        
        # Note: "Expected Modifier" is optional for initial runs
        # The bot's audit workflow will deduce it by checking TherapyNotes for session medium keywords
        # and comparing against the documented session medium to determine what the modifier should have been
        # So if "Expected Modifier" is empty here, that's expected - it will be filled during audit
        
        # Parse name into first and last (if needed)
        name_parts = client_data["client_name"].split(maxsplit=1)
        client_data["first_name"] = name_parts[0] if len(name_parts) > 0 else ""
        client_data["last_name"] = name_parts[1] if len(name_parts) > 1 else ""
        
        # Log extraction for debugging
        if patient_member_id:
            self.gui_log(f"Extracted Patient Member ID (MBI): {patient_member_id}", level="DEBUG")
        else:
            self.gui_log(f"WARNING: Patient Member ID (MBI) not found in Excel row for {client_name}", level="WARNING")
        
        return client_data
    
    def _create_client_folder(self, client_name: str) -> Path | None:
        """Create a folder for the client in the output folder.
        
        Args:
            client_name: Client name (used for folder name)
        
        Returns:
            Path: Path to created folder, or None if creation failed
        """
        try:
            if not self.pdf_output_folder:
                return None
            
            # Sanitize client name for folder name (remove invalid characters)
            safe_folder_name = re.sub(r'[<>:"/\\|?*]', '_', client_name)
            safe_folder_name = safe_folder_name.strip()
            
            if not safe_folder_name:
                safe_folder_name = "Unknown_Client"
            
            client_folder = self.pdf_output_folder / safe_folder_name
            client_folder.mkdir(parents=True, exist_ok=True)
            
            self.gui_log(f"Created client folder: {client_folder}", level="DEBUG")
            return client_folder
            
        except Exception as e:
            self.gui_log(f"Error creating client folder: {e}", level="ERROR")
            return None
    
    def _fill_pdf_form(self, client_data: Dict[str, Any], output_folder: Path, client_name: str) -> Path | None:
        """Fill PDF form with client data.
        
        Args:
            client_data: Dictionary containing client data
            output_folder: Folder where filled PDF should be saved
            client_name: Client name (used for filename)
        
        Returns:
            Path: Path to filled PDF, or None if filling failed
        """
        try:
            if not self.pdf_template_path or not self.pdf_template_path.exists():
                return None
            
            # Try pdfrw first (best for form filling)
            if PDFRW_AVAILABLE:
                return self._fill_pdf_with_pdfrw(client_data, output_folder, client_name)
            elif PYPDF2_AVAILABLE:
                return self._fill_pdf_with_pypdf2(client_data, output_folder, client_name)
            else:
                self.gui_log("ERROR: No PDF library available for form filling.", level="ERROR")
                return None
                
        except Exception as e:
            self.log_error("Error filling PDF form", exception=e, include_traceback=True)
            return None
    
    def _fill_pdf_with_pdfrw(self, client_data: Dict[str, Any], output_folder: Path, client_name: str) -> Path | None:
        """Fill PDF form using pdfrw library.
        
        This method tries multiple approaches to fill PDF forms:
        1. Fill AcroForm fields directly
        2. Fill page annotation widgets
        3. Handle nested form fields
        
        Args:
            client_data: Dictionary containing client data
            output_folder: Folder where filled PDF should be saved
            client_name: Client name (used for filename)
        
        Returns:
            Path: Path to filled PDF, or None if filling failed
        """
        try:
            if not PDFRW_AVAILABLE:
                return None
            
            self.gui_log(f"Filling PDF form using pdfrw for {client_name}...", level="INFO")
            
            # Read template PDF
            template_pdf = pdfrw.PdfReader(str(self.pdf_template_path))  # type: ignore
            
            # Map client data to PDF form fields
            field_mapping = self._get_pdf_field_mapping(client_data)
            
            if not field_mapping:
                self.gui_log("WARNING: No field mappings configured. PDF will be saved without filling.", level="WARNING")
                # Still save the PDF even without filling
            else:
                # Log field mappings for debugging
                self.gui_log(f"Field mappings: {list(field_mapping.keys())}", level="DEBUG")
            
            # Fill form fields using multiple methods
            fields_filled = 0
            
            # Method 1: Fill AcroForm fields directly
            if template_pdf.Root and template_pdf.Root.AcroForm:
                acro_form = template_pdf.Root.AcroForm
                if acro_form.get('/Fields'):
                    fields_array = acro_form['/Fields']
                    self.gui_log(f"Attempting to fill {len(fields_array)} AcroForm field(s)...", level="DEBUG")
                    
                    for field_obj in fields_array:
                        if field_obj is None:
                            continue
                        
                        field_name_obj = field_obj.get('/T')
                        if field_name_obj:
                            field_name_str = self._extract_field_name(field_name_obj)
                            if field_name_str:
                                # Find matching value
                                matched_key = self._match_field_name(field_name_str, field_mapping)
                                if matched_key and matched_key in field_mapping:
                                    field_value = str(field_mapping[matched_key])
                                    if field_value:
                                        try:
                                            # Set field value
                                            field_obj[pdfrw.PdfName('V')] = pdfrw.PdfString(field_value)  # type: ignore
                                            # Update appearance stream
                                            field_obj[pdfrw.PdfName('AP')] = pdfrw.PdfDict()  # type: ignore
                                            # Remove read-only flag if present
                                            if '/Ff' in field_obj:
                                                # Clear read-only flag (bit 1)
                                                current_flags = field_obj['/Ff']
                                                if isinstance(current_flags, int):
                                                    field_obj['/Ff'] = pdfrw.PdfObject(current_flags & ~1)  # type: ignore
                                            fields_filled += 1
                                            self.gui_log(f"Filled AcroForm field '{field_name_str}': '{field_value}'", level="DEBUG")
                                        except Exception as e:
                                            self.gui_log(f"Error filling AcroForm field '{field_name_str}': {e}", level="WARNING")
            
            # Method 2: Fill page annotation widgets
            for page_num, page in enumerate(template_pdf.pages, 1):
                annotations = page.get('/Annots')
                if not annotations:
                    continue
                
                self.gui_log(f"Processing page {page_num} annotations...", level="DEBUG")
                
                for annotation in annotations:
                    if annotation is None:
                        continue
                    
                    # Get annotation object
                    if isinstance(annotation, pdfrw.PdfDict):  # type: ignore
                        annotation_obj = annotation
                    else:
                        annotation_obj = pdfrw.PdfDict(annotation)  # type: ignore
                    
                    # Check if it's a form field widget
                    if annotation_obj.get('/Subtype') == '/Widget':
                        # Get field name
                        field_name_obj = annotation_obj.get('/T')
                        field_name_str = ""
                        
                        if field_name_obj:
                            field_name_str = self._extract_field_name(field_name_obj)
                        else:
                            # Check parent for field name
                            if '/Parent' in annotation_obj:
                                parent = annotation_obj['/Parent']
                                if parent:
                                    parent_name_obj = parent.get('/T')
                                    if parent_name_obj:
                                        field_name_str = self._extract_field_name(parent_name_obj)
                        
                        if field_name_str:
                            # Find matching value
                            matched_key = self._match_field_name(field_name_str, field_mapping)
                            if matched_key and matched_key in field_mapping:
                                field_value = str(field_mapping[matched_key])
                                # Only fill if value is not empty (empty means leave blank)
                                if field_value:
                                    try:
                                        # Set field value
                                        annotation_obj[pdfrw.PdfName('V')] = pdfrw.PdfString(field_value)  # type: ignore
                                        # Update appearance (needed for field to display)
                                        annotation_obj[pdfrw.PdfName('AP')] = pdfrw.PdfDict()  # type: ignore
                                        # Remove read-only flag
                                        if '/Ff' in annotation_obj:
                                            current_flags = annotation_obj['/Ff']
                                            if isinstance(current_flags, int):
                                                annotation_obj['/Ff'] = pdfrw.PdfObject(current_flags & ~1)  # type: ignore
                                        fields_filled += 1
                                        self.gui_log(f"✅ Filled widget field '{field_name_str}': '{field_value}'", level="INFO")
                                    except Exception as e:
                                        self.gui_log(f"❌ Error filling widget field '{field_name_str}': {e}", level="WARNING")
                                else:
                                    self.gui_log(f"⏭️  Skipping field '{field_name_str}' (mapped to empty string - leaving blank)", level="DEBUG")
                            else:
                                # Field not in mapping - leave blank
                                self.gui_log(f"⏭️  Skipping field '{field_name_str}' (not in mapping - leaving blank)", level="DEBUG")
            
            self.gui_log(f"Filled {fields_filled} field(s) in PDF form.", level="INFO")
            
            # Save filled PDF
            safe_filename = re.sub(r'[<>:"/\\|?*]', '_', client_name)
            output_pdf_path = output_folder / f"{safe_filename}_filled.pdf"
            
            pdfrw.PdfWriter().write(str(output_pdf_path), template_pdf)  # type: ignore
            
            self.gui_log(f"✅ PDF form filled: {output_pdf_path}", level="INFO")
            return output_pdf_path
            
        except Exception as e:
            self.log_error("Error filling PDF with pdfrw", exception=e, include_traceback=True)
            return None
    
    def _match_field_name(self, field_name_str: str, field_mapping: Dict[str, str]) -> str | None:
        """Match a field name to a key in the field mapping.
        
        This method handles field name matching with special characters, parentheses, and variations.
        PDF field names often have parentheses and escaped characters that need to be normalized.
        
        Args:
            field_name_str: Field name from PDF (may have parentheses, escaped chars, etc.)
            field_mapping: Dictionary of field mappings
        
        Returns:
            str: Matched key from field_mapping, or None if no match
        """
        if not field_name_str or not field_mapping:
            return None
        
        # Clean up field name (remove parentheses, escape characters, etc.)
        cleaned_field_name = field_name_str.strip('()[]')
        cleaned_field_name = cleaned_field_name.replace('\\', '')  # Remove escape characters
        cleaned_field_name = cleaned_field_name.strip()
        
        # Try exact match first (with cleaned name)
        if cleaned_field_name in field_mapping:
            return cleaned_field_name
        
        # Try exact match with original name
        if field_name_str in field_mapping:
            return field_name_str
        
        # Try case-insensitive match (with cleaned name)
        cleaned_field_name_lower = cleaned_field_name.lower().strip()
        for key in field_mapping.keys():
            key_cleaned = key.strip('()[]').replace('\\', '').strip()
            # Exact match after cleaning
            if key_cleaned.lower().strip() == cleaned_field_name_lower:
                return key
            # Original key match (case-insensitive)
            if key.lower().strip() == cleaned_field_name_lower:
                return key
        
        # Try partial match (field name contains key or vice versa)
        for key in field_mapping.keys():
            key_cleaned = key.strip('()[]').replace('\\', '').strip().lower()
            if cleaned_field_name_lower in key_cleaned or key_cleaned in cleaned_field_name_lower:
                return key
        
        # Try matching without special characters (parentheses, etc.)
        field_name_no_special = re.sub(r'[^\w\s]', '', cleaned_field_name_lower)
        for key in field_mapping.keys():
            key_no_special = re.sub(r'[^\w\s]', '', key.lower().strip())
            if field_name_no_special == key_no_special:
                return key
        
        return None
    
    def _fill_pdf_with_pypdf2(self, client_data: Dict[str, Any], output_folder: Path, client_name: str) -> Path | None:
        """Fill PDF form using PyPDF2 library (fallback).
        
        Args:
            client_data: Dictionary containing client data
            output_folder: Folder where filled PDF should be saved
            client_name: Client name (used for filename)
        
        Returns:
            Path: Path to filled PDF, or None if filling failed
        """
        try:
            if not PYPDF2_AVAILABLE:
                return None
            
            self.gui_log(f"Filling PDF form using PyPDF2 for {client_name}...", level="INFO")
            
            # Read template PDF
            with open(self.pdf_template_path, 'rb') as template_file:
                pdf_reader = PyPDF2.PdfReader(template_file)  # type: ignore
                pdf_writer = PyPDF2.PdfWriter()  # type: ignore
                
                # Copy all pages
                for page in pdf_reader.pages:
                    pdf_writer.add_page(page)
                
                # Get form fields (if any)
                if pdf_reader.metadata:
                    # PyPDF2 doesn't directly support form filling, so we'd need to use pdfrw
                    # or another library for actual form filling
                    self.gui_log("WARNING: PyPDF2 doesn't support form filling directly. Using pdfrw is recommended.", level="WARNING")
                    return None
            
            # Save filled PDF
            safe_filename = re.sub(r'[<>:"/\\|?*]', '_', client_name)
            output_pdf_path = output_folder / f"{safe_filename}_filled.pdf"
            
            with open(output_pdf_path, 'wb') as output_file:
                pdf_writer.write(output_file)
            
            self.gui_log(f"✅ PDF saved: {output_pdf_path}", level="INFO")
            return output_pdf_path
            
        except Exception as e:
            self.log_error("Error filling PDF with PyPDF2", exception=e, include_traceback=True)
            return None
    
    def _get_pdf_field_mapping(self, client_data: Dict[str, Any]) -> Dict[str, str]:
        """Map client data to PDF form field names.
        
        This method uses the configured field mapping (self.pdf_field_mapping_config) to map
        client data from the Excel file to PDF form fields. The user must configure this mapping
        based on the actual PDF form structure.
        
        Args:
            client_data: Dictionary containing client data from Excel
        
        Returns:
            dict: Dictionary mapping PDF field names to values (strings)
        """
        # First, list all PDF form fields for reference
        pdf_fields = self._list_pdf_form_fields()
        
        if pdf_fields:
            self.gui_log(f"Found {len(pdf_fields)} PDF form field(s) in template.", level="INFO")
            # Log first 20 fields for reference
            fields_preview = ', '.join(pdf_fields[:20])
            if len(pdf_fields) > 20:
                fields_preview += f", ... and {len(pdf_fields) - 20} more"
            self.gui_log(f"PDF fields: {fields_preview}", level="DEBUG")
        else:
            self.gui_log("WARNING: No PDF form fields found in template. PDF may not be fillable.", level="WARNING")
        
        # Create field mapping based on user configuration
        field_mapping = {}
        
        # Check if user has configured field mappings
        if self.pdf_field_mapping_config:
            # Use configured mappings
            for pdf_field_name, config_item in self.pdf_field_mapping_config.items():
                # Handle both old format (string) and new format (dict)
                if isinstance(config_item, dict):
                    config_type = config_item.get("type", "static")
                    config_value = config_item.get("value", "")
                    field_type = config_item.get("field_type", "text")
                else:
                    # Old format: config_item is a string (backwards compatibility)
                    config_type = "excel" if config_item in client_data else "static"
                    config_value = config_item
                    field_type = "text"
                
                # Find matching PDF field (handle variations in field names)
                matched_field = None
                
                # Clean up configured field name (remove parentheses, escape chars, etc.)
                cleaned_config_name = pdf_field_name.strip('()[]').replace('\\', '').strip()
                cleaned_config_name_lower = cleaned_config_name.lower().strip()
                
                # Try to find matching field in PDF
                if pdf_fields:
                    for field in pdf_fields:
                        # Clean up PDF field name
                        cleaned_field = field.strip('()[]').replace('\\', '').strip()
                        cleaned_field_lower = cleaned_field.lower().strip()
                        
                        # Try exact match (after cleaning)
                        if cleaned_field == cleaned_config_name or cleaned_field_lower == cleaned_config_name_lower:
                            matched_field = field
                            break
                        
                        # Try partial match (field name contains config name or vice versa)
                        if cleaned_config_name_lower in cleaned_field_lower or cleaned_field_lower in cleaned_config_name_lower:
                            matched_field = field
                            break
                    
                    # If found a match, use the actual PDF field name
                    if matched_field:
                        pdf_field_name = matched_field
                    else:
                        # Try one more time with original field name
                        if pdf_field_name not in pdf_fields:
                            # Try case-insensitive match with original names
                            for field in pdf_fields:
                                if field.lower() == pdf_field_name.lower():
                                    matched_field = field
                                    break
                            
                            if matched_field:
                                pdf_field_name = matched_field
                            else:
                                self.gui_log(f"WARNING: Configured PDF field '{pdf_field_name}' not found in form. Trying to match anyway...", level="WARNING")
                                # Continue anyway - the _match_field_name method will try to match it
                else:
                    self.gui_log(f"WARNING: No PDF fields found, but attempting to use configured field '{pdf_field_name}'", level="WARNING")
                
                # Get value from client_data based on config_type
                value = ""
                
                if config_type == "function":
                    # If config_value is a function, call it with client_data
                    if callable(config_value):
                        try:
                            value = config_value(client_data)
                            value = str(value) if value else ""
                            if value:
                                self.gui_log(f"Mapped '{pdf_field_name}' -> function result: '{value[:50]}...'", level="DEBUG")
                        except Exception as e:
                            self.gui_log(f"Error getting value for field '{pdf_field_name}': {e}", level="WARNING")
                            value = ""
                elif config_type == "excel":
                    # If config_value is an Excel column name (client_data key)
                    if config_value in client_data:
                        value = client_data.get(config_value, "")
                        value = str(value) if value else ""
                        if value:
                            self.gui_log(f"Mapped '{pdf_field_name}' -> '{config_value}': '{value}'", level="DEBUG")
                    else:
                        self.gui_log(f"WARNING: Excel column '{config_value}' not found in client_data for field '{pdf_field_name}'", level="WARNING")
                        value = ""
                elif config_type == "static":
                    # Static value
                    value = str(config_value) if config_value else ""
                    if value:
                        self.gui_log(f"Mapped '{pdf_field_name}' -> static value: '{value}'", level="DEBUG")
                else:
                    # Fallback: treat as direct value
                    value = str(config_value) if config_value else ""
                
                # Handle checkbox fields: convert value to "Yes" or "Off"
                if field_type == "checkbox":
                    if value:
                        # Convert checkbox value to PDF format
                        value_lower = value.lower().strip()
                        if value_lower in ["checked", "yes", "true", "1", "on"]:
                            value = "Yes"
                        elif value_lower in ["unchecked", "off", "no", "false", "0", ""]:
                            value = "Off"
                        else:
                            # If value is truthy (non-empty), assume checked
                            value = "Yes"
                    else:
                        # Empty value means unchecked
                        value = "Off"
                    self.gui_log(f"Checkbox field '{pdf_field_name}' mapped to: '{value}'", level="DEBUG")
                
                # Add to field mapping
                field_mapping[pdf_field_name] = value
            
            if field_mapping:
                self.gui_log(f"Mapped {len(field_mapping)} field(s) using configuration.", level="INFO")
                # Log mapped fields for debugging
                mapped_fields_preview = ', '.join(list(field_mapping.keys())[:10])
                if len(field_mapping) > 10:
                    mapped_fields_preview += f", ... and {len(field_mapping) - 10} more"
                self.gui_log(f"Mapped fields: {mapped_fields_preview}", level="DEBUG")
        else:
            # No configuration found - try automatic pattern matching as fallback
            self.gui_log("WARNING: No field mapping configuration found. Attempting automatic pattern matching...", level="WARNING")
            field_mapping = self._auto_map_pdf_fields(client_data, pdf_fields)
        
        # If still no mappings, provide helpful error message
        if not field_mapping:
            self.gui_log("ERROR: No field mappings configured. Please configure field mappings in the code.", level="ERROR")
            if pdf_fields:
                self.gui_log(f"Available PDF fields: {', '.join(pdf_fields)}", level="INFO")
                self.gui_log("Available client_data keys: " + ', '.join(client_data.keys()), level="INFO")
        
        return field_mapping
    
    def _auto_map_pdf_fields(self, client_data: Dict[str, Any], pdf_fields: List[str]) -> Dict[str, str]:
        """Attempt to automatically map PDF fields based on common patterns.
        
        This is a fallback method that tries to match PDF field names to client data
        using common naming patterns. This may not work for all PDF forms.
        
        Args:
            client_data: Dictionary containing client data
            pdf_fields: List of PDF form field names
        
        Returns:
            dict: Dictionary mapping PDF field names to values
        """
        field_mapping = {}
        
        if not pdf_fields:
            return field_mapping
        
        # Try to match client data to PDF fields using common patterns
        for field_name in pdf_fields:
            field_name_lower = field_name.lower()
            
            # Patient name fields
            if any(term in field_name_lower for term in ['name', 'patient', 'beneficiary']):
                if 'first' in field_name_lower:
                    field_mapping[field_name] = client_data.get("first_name", "")
                elif 'last' in field_name_lower:
                    field_mapping[field_name] = client_data.get("last_name", "")
                else:
                    field_mapping[field_name] = client_data.get("client_name", "")
            
            # DOB fields
            elif any(term in field_name_lower for term in ['dob', 'date of birth', 'birth', 'birthdate']):
                field_mapping[field_name] = client_data.get("dob", "")
            
            # Date of Service fields
            elif any(term in field_name_lower for term in ['dos', 'date of service', 'service date']):
                if 'service' in field_name_lower or 'dos' in field_name_lower:
                    field_mapping[field_name] = client_data.get("dos", "")
            
            # Patient Member ID / MBI fields
            # MBI = Medicare Beneficiary Identifier (same as Patient Member #)
            elif any(term in field_name_lower for term in ['mbi', 'member', 'medicare id', 'patient id', 
                                                            'medicare number', 'medicare beneficiary', 
                                                            'beneficiary identifier', 'beneficiary id']):
                # Specifically handle MBI field
                if 'mbi' in field_name_lower or 'medicare beneficiary' in field_name_lower:
                    field_mapping[field_name] = client_data.get("patient_member_id", "")
                # Also handle generic member/ID fields
                elif 'member' in field_name_lower or 'id' in field_name_lower or 'number' in field_name_lower:
                    field_mapping[field_name] = client_data.get("patient_member_id", "")
            
            # Modifier fields
            elif any(term in field_name_lower for term in ['modifier', 'code', 'cpt']):
                if 'original' in field_name_lower or 'current' in field_name_lower:
                    field_mapping[field_name] = client_data.get("original_modifier", "")
                elif 'expected' in field_name_lower or 'correct' in field_name_lower:
                    field_mapping[field_name] = client_data.get("expected_modifier", "")
        
        if field_mapping:
            self.gui_log(f"Auto-mapped {len(field_mapping)} field(s) using pattern matching.", level="INFO")
        
        return field_mapping
    
    def _list_pdf_form_fields(self) -> List[str]:
        """List all form field names in the PDF template.
        
        This method tries multiple approaches to detect PDF form fields:
        1. pdfrw - checks AcroForm fields
        2. pdfrw - checks page annotations (Widget subtypes)
        3. PyPDF2 - checks form fields
        4. Direct PDF structure inspection
        
        Returns:
            list: List of PDF form field names
        """
        fields = []
        
        # Method 1: Try pdfrw (AcroForm fields)
        if PDFRW_AVAILABLE and self.pdf_template_path and self.pdf_template_path.exists():
            try:
                template_pdf = pdfrw.PdfReader(str(self.pdf_template_path))  # type: ignore
                
                # Check if PDF has Root
                if not template_pdf.Root:
                    self.gui_log("PDF has no Root structure.", level="WARNING")
                else:
                    # Check if PDF has AcroForm
                    if template_pdf.Root.AcroForm:
                        acro_form = template_pdf.Root.AcroForm
                        self.gui_log("PDF has AcroForm structure.", level="INFO")
                        
                        # Check for Fields array in AcroForm
                        if acro_form.get('/Fields'):
                            fields_array = acro_form['/Fields']
                            self.gui_log(f"Found Fields array with {len(fields_array)} field(s).", level="INFO")
                            
                            # Recursively extract fields (handles nested fields)
                            self._extract_fields_from_array(fields_array, fields)
                        else:
                            self.gui_log("AcroForm exists but no Fields array found. Checking page annotations...", level="INFO")
                    else:
                        self.gui_log("PDF Root exists but no AcroForm found. Checking page annotations...", level="INFO")
                    
                    # Always check page annotations (Widget subtypes) - this is often more reliable
                    for page_num, page in enumerate(template_pdf.pages, 1):
                        annotations = page.get('/Annots')
                        if not annotations:
                            continue
                        
                        self.gui_log(f"Page {page_num} has {len(annotations)} annotation(s).", level="DEBUG")
                        
                        for annotation in annotations:
                            if annotation is None:
                                continue
                            
                            if isinstance(annotation, pdfrw.PdfDict):  # type: ignore
                                annotation_obj = annotation
                            else:
                                annotation_obj = pdfrw.PdfDict(annotation)  # type: ignore
                            
                            # Check subtype - handle both string and PdfName
                            subtype = annotation_obj.get('/Subtype')
                            is_widget = False
                            if subtype:
                                if isinstance(subtype, str) and subtype == '/Widget':
                                    is_widget = True
                                else:
                                    # Check if it's a PdfName object (compare string representation)
                                    subtype_str = str(subtype)
                                    if '/Widget' in subtype_str or 'Widget' in subtype_str:
                                        is_widget = True
                                    # Also try direct comparison
                                    try:
                                        if hasattr(pdfrw, 'PdfName') and subtype == pdfrw.PdfName('Widget'):  # type: ignore
                                            is_widget = True
                                    except Exception:
                                        pass
                            
                            # Check for Widget subtype (form fields)
                            if is_widget:
                                # Get field name from /T
                                field_name_obj = annotation_obj.get('/T')
                                if field_name_obj:
                                    field_name_str = self._extract_field_name(field_name_obj)
                                    if field_name_str and field_name_str not in fields:
                                        fields.append(field_name_str)
                                        self.gui_log(f"Found Widget field: {field_name_str}", level="DEBUG")
                                
                                # Also check /Parent for field name
                                if '/Parent' in annotation_obj:
                                    parent = annotation_obj['/Parent']
                                    if parent:
                                        parent_name_obj = parent.get('/T')
                                        if parent_name_obj:
                                            field_name_str = self._extract_field_name(parent_name_obj)
                                            if field_name_str and field_name_str not in fields:
                                                fields.append(field_name_str)
                                                self.gui_log(f"Found Parent field: {field_name_str}", level="DEBUG")
                            
                            # Also check for any annotation with /FT (Field Type) - indicates form field
                            if '/FT' in annotation_obj:
                                field_name_obj = annotation_obj.get('/T')
                                if field_name_obj:
                                    field_name_str = self._extract_field_name(field_name_obj)
                                    if field_name_str and field_name_str not in fields:
                                        fields.append(field_name_str)
                                        self.gui_log(f"Found form field (has /FT): {field_name_str}", level="DEBUG")
                            
                            # Check for /Ff (Field Flags) - also indicates form field
                            if '/Ff' in annotation_obj:
                                field_name_obj = annotation_obj.get('/T')
                                if field_name_obj:
                                    field_name_str = self._extract_field_name(field_name_obj)
                                    if field_name_str and field_name_str not in fields:
                                        fields.append(field_name_str)
                                        self.gui_log(f"Found form field (has /Ff): {field_name_str}", level="DEBUG")
                
            except Exception as e:
                self.gui_log(f"Error listing PDF form fields with pdfrw: {e}", level="WARNING")
                import traceback
                self.gui_log(f"Traceback: {traceback.format_exc()}", level="DEBUG")
        
        # Method 2: Try PyPDF2 (alternative method) - even if pdfrw found fields, try PyPDF2 for comparison
        if PYPDF2_AVAILABLE and self.pdf_template_path and self.pdf_template_path.exists():
            try:
                with open(self.pdf_template_path, 'rb') as pdf_file:
                    pdf_reader = PyPDF2.PdfReader(pdf_file)  # type: ignore
                    
                    self.gui_log("Trying PyPDF2 for field detection...", level="DEBUG")
                    
                    # Check if PDF has form fields (PyPDF2 3.x method)
                    if hasattr(pdf_reader, 'get_form_text_fields'):
                        try:
                            form_fields = pdf_reader.get_form_text_fields()
                            if form_fields:
                                self.gui_log(f"PyPDF2 found {len(form_fields)} form field(s) via get_form_text_fields().", level="INFO")
                                for field_name in form_fields.keys():
                                    if field_name and field_name not in fields:
                                        fields.append(str(field_name))
                                        self.gui_log(f"PyPDF2 found field: {field_name}", level="DEBUG")
                        except Exception as e:
                            self.gui_log(f"PyPDF2 get_form_text_fields() error: {e}", level="DEBUG")
                    
                    # Try PyPDF2 2.x method (getFields)
                    if hasattr(pdf_reader, 'getFields'):
                        try:
                            pdf_fields = pdf_reader.getFields()  # type: ignore
                            if pdf_fields:
                                self.gui_log(f"PyPDF2 found {len(pdf_fields)} field(s) via getFields().", level="INFO")
                                for field_name in pdf_fields.keys():
                                    if field_name and field_name not in fields:
                                        fields.append(str(field_name))
                                        self.gui_log(f"PyPDF2 found field: {field_name}", level="DEBUG")
                        except Exception as e:
                            self.gui_log(f"PyPDF2 getFields() error: {e}", level="DEBUG")
                    
                    # Try accessing fields via pages (PyPDF2 3.x)
                    try:
                        for page_num, page in enumerate(pdf_reader.pages, 1):
                            # Check for annotations
                            if '/Annots' in page:
                                annotations = page['/Annots']
                                if annotations:
                                    self.gui_log(f"PyPDF2: Page {page_num} has {len(annotations)} annotation(s).", level="DEBUG")
                    except Exception as e:
                        self.gui_log(f"PyPDF2 page annotation check error: {e}", level="DEBUG")
                    
            except Exception as e:
                self.gui_log(f"Error listing PDF form fields with PyPDF2: {e}", level="WARNING")
                import traceback
                self.gui_log(f"PyPDF2 traceback: {traceback.format_exc()}", level="DEBUG")
        
        # Method 3: Try to read PDF structure directly (fallback)
        if not fields and self.pdf_template_path and self.pdf_template_path.exists():
            try:
                # Read PDF as binary and search for common form field patterns
                with open(self.pdf_template_path, 'rb') as f:
                    pdf_content = f.read()
                    
                    # Search for common form field markers
                    # This is a last resort method
                    import re
                    # Look for /T (field name) patterns
                    name_patterns = re.findall(rb'/T\s*\(([^)]+)\)', pdf_content)
                    for pattern in name_patterns:
                        try:
                            field_name = pattern.decode('utf-8', errors='ignore').strip()
                            if field_name and field_name not in fields:
                                fields.append(field_name)
                        except Exception:
                            pass
                    
                    if fields:
                        self.gui_log(f"Found {len(fields)} field(s) via pattern matching.", level="DEBUG")
            except Exception as e:
                self.gui_log(f"Error in pattern matching: {e}", level="DEBUG")
        
        return fields
    
    def _extract_fields_from_array(self, fields_array, fields_list: List[str]) -> None:
        """Recursively extract field names from a PDF Fields array.
        
        Args:
            fields_array: PDF Fields array (can contain nested fields)
            fields_list: List to append field names to
        """
        try:
            if not fields_array:
                return
            
            for field_obj in fields_array:
                if field_obj is None:
                    continue
                
                # Get field name
                field_name_obj = field_obj.get('/T')
                field_name_str = ""
                if field_name_obj:
                    field_name_str = self._extract_field_name(field_name_obj)
                    if field_name_str and field_name_str not in fields_list:
                        fields_list.append(field_name_str)
                        self.gui_log(f"Extracted field from array: {field_name_str}", level="DEBUG")
                
                # Check for nested fields (Kids array)
                if '/Kids' in field_obj:
                    kids_array = field_obj['/Kids']
                    if kids_array:
                        self.gui_log(f"Found nested fields (Kids array) for field: {field_name_str if field_name_str else 'Unknown'}", level="DEBUG")
                        self._extract_fields_from_array(kids_array, fields_list)
        except Exception as e:
            self.gui_log(f"Error extracting fields from array: {e}", level="WARNING")
    
    def _extract_field_name(self, field_name_obj) -> str:
        """Extract field name from various PDF object types.
        
        Args:
            field_name_obj: PDF object containing field name
        
        Returns:
            str: Extracted field name, or empty string if extraction fails
        """
        try:
            if field_name_obj is None:
                return ""
            
            # Handle different PDF object types
            # Check if it's a PdfString (without using isinstance on the class directly)
            field_name_str = ""
            try:
                # Try to get the value directly if it's a PdfString
                if hasattr(field_name_obj, 'decode'):
                    try:
                        field_name_str = field_name_obj.decode('utf-8')
                    except Exception:
                        try:
                            field_name_str = field_name_obj.decode('latin-1')
                        except Exception:
                            field_name_str = str(field_name_obj)
                elif isinstance(field_name_obj, bytes):
                    try:
                        field_name_str = field_name_obj.decode('utf-8', errors='ignore')
                    except Exception:
                        try:
                            field_name_str = field_name_obj.decode('latin-1', errors='ignore')
                        except Exception:
                            field_name_str = str(field_name_obj)
                elif isinstance(field_name_obj, str):
                    field_name_str = field_name_obj
                else:
                    # Try to convert to string and clean up
                    field_name_str = str(field_name_obj)
                    # Remove common PDF object markers (parentheses, brackets)
                    field_name_str = field_name_str.strip('()[]')
                    # Remove quotes if present
                    field_name_str = field_name_str.strip('"\'')
                
                # Clean up escaped characters (e.g., "Date\(s\)" -> "Date(s)")
                field_name_str = field_name_str.replace('\\', '')
                
                return field_name_str
            except Exception:
                # Fallback: just convert to string
                return str(field_name_obj).strip('()[]"\'')
        except Exception as e:
            self.gui_log(f"Error extracting field name: {e}", level="DEBUG")
            return ""
    
    def _flatten_pdf(self, pdf_path: Path, output_folder: Path, client_name: str) -> Path | None:
        """Flatten PDF (make it non-editable).
        
        Args:
            pdf_path: Path to filled PDF
            output_folder: Folder where flattened PDF should be saved
            client_name: Client name (used for filename)
        
        Returns:
            Path: Path to flattened PDF, or None if flattening failed
        """
        try:
            if not pdf_path or not pdf_path.exists():
                return None
            
            self.gui_log(f"Flattening PDF for {client_name}...", level="DEBUG")
            
            # Try pdfrw first (better for flattening)
            if PDFRW_AVAILABLE:
                return self._flatten_pdf_with_pdfrw(pdf_path, output_folder, client_name)
            elif PYPDF2_AVAILABLE:
                return self._flatten_pdf_with_pypdf2(pdf_path, output_folder, client_name)
            else:
                self.gui_log("WARNING: No PDF library available for flattening. PDF will remain editable.", level="WARNING")
                return pdf_path
                
        except Exception as e:
            self.log_error("Error flattening PDF", exception=e, include_traceback=True)
            return pdf_path  # Return original PDF if flattening fails
    
    def _flatten_pdf_with_pdfrw(self, pdf_path: Path, output_folder: Path, client_name: str) -> Path | None:
        """Flatten PDF using pdfrw library.
        
        Args:
            pdf_path: Path to filled PDF
            output_folder: Folder where flattened PDF should be saved
            client_name: Client name (used for filename)
        
        Returns:
            Path: Path to flattened PDF, or None if flattening failed
        """
        try:
            if not PDFRW_AVAILABLE:
                return None
            
            # Read filled PDF
            pdf = pdfrw.PdfReader(str(pdf_path))  # type: ignore
            
            # Remove form fields (flatten) by removing AcroForm
            if pdf.Root and pdf.Root.AcroForm:
                # Remove AcroForm to flatten
                del pdf.Root.AcroForm
            
            # Remove annotations from pages (flatten form fields)
            for page in pdf.pages:
                if '/Annots' in page:
                    # Keep only non-form annotations (like signatures)
                    # For full flattening, remove all annotations
                    # For now, we'll remove form field annotations
                    annotations = page['/Annots']
                    if annotations:
                        # Filter out form field annotations
                        filtered_annotations = []
                        for annot in annotations:
                            if annot and annot.get('/Subtype') != '/Widget':
                                filtered_annotations.append(annot)
                        if filtered_annotations:
                            page['/Annots'] = pdfrw.PdfArray(filtered_annotations)  # type: ignore
                        else:
                            del page['/Annots']
            
            # Save flattened PDF
            safe_filename = re.sub(r'[<>:"/\\|?*]', '_', client_name)
            flattened_pdf_path = output_folder / f"{safe_filename}_flattened.pdf"
            
            pdfrw.PdfWriter().write(str(flattened_pdf_path), pdf)  # type: ignore
            
            # Delete the original filled PDF (keep only flattened version)
            try:
                if pdf_path.exists() and pdf_path != flattened_pdf_path:
                    pdf_path.unlink()
            except Exception:
                pass  # If deletion fails, keep both files
            
            self.gui_log(f"✅ PDF flattened: {flattened_pdf_path}", level="INFO")
            return flattened_pdf_path
            
        except Exception as e:
            self.log_error("Error flattening PDF with pdfrw", exception=e, include_traceback=True)
            return pdf_path  # Return original PDF if flattening fails
    
    def _flatten_pdf_with_pypdf2(self, pdf_path: Path, output_folder: Path, client_name: str) -> Path | None:
        """Flatten PDF using PyPDF2 library (fallback).
        
        Args:
            pdf_path: Path to filled PDF
            output_folder: Folder where flattened PDF should be saved
            client_name: Client name (used for filename)
        
        Returns:
            Path: Path to flattened PDF, or None if flattening failed
        """
        try:
            if not PYPDF2_AVAILABLE:
                return None
            
            # Read filled PDF
            with open(pdf_path, 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)  # type: ignore
                pdf_writer = PyPDF2.PdfWriter()  # type: ignore
                
                # Copy all pages (this effectively flattens form fields in PyPDF2)
                for page in pdf_reader.pages:
                    pdf_writer.add_page(page)
            
            # Save flattened PDF
            safe_filename = re.sub(r'[<>:"/\\|?*]', '_', client_name)
            flattened_pdf_path = output_folder / f"{safe_filename}_flattened.pdf"
            
            with open(flattened_pdf_path, 'wb') as output_file:
                pdf_writer.write(output_file)
            
            # Delete the original filled PDF (keep only flattened version)
            try:
                if pdf_path.exists() and pdf_path != flattened_pdf_path:
                    pdf_path.unlink()
            except Exception:
                pass  # If deletion fails, keep both files
            
            self.gui_log(f"✅ PDF flattened: {flattened_pdf_path}", level="INFO")
            return flattened_pdf_path
            
        except Exception as e:
            self.log_error("Error flattening PDF with PyPDF2", exception=e, include_traceback=True)
            return pdf_path  # Return original PDF if flattening fails

    def _load_excel_data(self, path: Path | None) -> list[dict[str, str]]:
        """Load Excel/CSV data, respecting Excel filters (only visible rows)."""
        if not path or not path.exists():
            raise FileNotFoundError(f"Input file not found: {path}")
        
        records: list[dict[str, str]] = []
        
        if path.suffix.lower() == ".csv":
            # CSV files don't have filters, so read normally
            df = pd.read_csv(path)  # type: ignore
        else:
            # Excel files: respect filters and only read visible rows
            if OPENPYXL_AVAILABLE:
                try:
                    self.gui_log("Loading Excel file and respecting filters (only visible rows)...", level="INFO")
                    df = self._load_excel_with_filters(path)
                except Exception as e:
                    self.gui_log(f"Warning: Could not respect Excel filters: {e}. Reading all rows.", level="WARNING")
                    df = pd.read_excel(path)  # type: ignore
            else:
                self.gui_log("Warning: openpyxl not available. Reading all rows (filters not respected).", level="WARNING")
                df = pd.read_excel(path)  # type: ignore
        
        # Apply row range filter if specified
        # Note: Row range refers to ORIGINAL Excel row numbers, not DataFrame indices
        start_row, end_row = self.row_range
        if start_row or end_row:
            # Filter based on original Excel row numbers (stored in _original_row_number column)
            if '_original_row_number' in df.columns:
                # Filter by original row numbers
                mask = True
                if start_row:
                    mask = mask & (df['_original_row_number'] >= start_row)
                if end_row:
                    mask = mask & (df['_original_row_number'] <= end_row)
                df = df[mask].copy()
                self.gui_log(f"Applied row range filter: original Excel rows {start_row} to {end_row}", level="INFO")
                self.gui_log(f"  Found {len(df)} visible row(s) within the specified range.", level="INFO")
            else:
                # Fallback: if original row numbers not available, use DataFrame indices
                start_idx = (start_row - 1) if start_row else 0
                end_idx = end_row if end_row else len(df)
                df = df.iloc[start_idx:end_idx]
                self.gui_log(f"Applied row range filter: rows {start_row} to {end_row} (using DataFrame indices)", level="INFO")
            
            # Remove the helper column before returning
            if '_original_row_number' in df.columns:
                df = df.drop(columns=['_original_row_number'])
        
        records = df.fillna("").to_dict(orient="records")
        self.gui_log(f"Loaded {len(records)} row(s) from input file.", level="INFO")
        return records
    
    def _load_excel_with_filters(self, path: Path) -> "pd.DataFrame":
        """Load Excel file respecting filters - only visible rows are included.
        
        Note: This method respects Excel filters by checking which rows are marked
        as hidden. For filters to work correctly, the Excel file must be saved
        with filters applied (rows should be marked as hidden).
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required to respect Excel filters")
        
        # Load workbook with openpyxl (read-only mode for better performance)
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)  # type: ignore
        ws = wb.active  # Get active worksheet
        
        # Check if auto_filter exists (indicates filters might be applied)
        has_auto_filter = ws.auto_filter is not None and ws.auto_filter.ref is not None
        if has_auto_filter:
            self.gui_log(f"Detected Excel auto-filter: {ws.auto_filter.ref}", level="DEBUG")
        
        # Find header row (usually row 1)
        header_row = 1
        headers = []
        for cell in ws[header_row]:
            headers.append(str(cell.value) if cell.value else "")
        
        if not headers:
            wb.close()
            raise ValueError("No headers found in Excel file")
        
        # Collect visible rows (skip hidden rows) and track original row numbers
        visible_rows = []
        original_row_numbers = []  # Track original Excel row number for each visible row
        total_rows = ws.max_row
        hidden_count = 0
        visible_count = 0
        
        self.gui_log(f"Scanning {total_rows} total row(s) for visible (non-hidden) rows...", level="DEBUG")
        
        for row_num in range(header_row + 1, total_rows + 1):
            # Check if row is hidden (filtered rows are marked as hidden)
            row_dim = ws.row_dimensions.get(row_num)
            is_hidden = row_dim and row_dim.hidden
            
            if is_hidden:
                hidden_count += 1
                continue  # Skip hidden rows
            
            # Extract row data
            row_data = []
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                value = cell.value
                # Convert value to string, handling None, dates, etc.
                if value is None:
                    row_data.append("")
                elif isinstance(value, datetime):
                    row_data.append(value.strftime("%m/%d/%Y"))
                elif hasattr(value, "strftime"):
                    try:
                        row_data.append(value.strftime("%m/%d/%Y"))
                    except Exception:
                        row_data.append(str(value))
                else:
                    row_data.append(str(value))
            
            visible_rows.append(row_data)
            original_row_numbers.append(row_num)  # Store original Excel row number
            visible_count += 1
        
        wb.close()
        
        # Log results
        self.gui_log(f"Found {visible_count} visible row(s) out of {total_rows - header_row} data row(s).", level="INFO")
        if hidden_count > 0:
            self.gui_log(f"Skipped {hidden_count} hidden/filtered row(s).", level="INFO")
        elif has_auto_filter:
            self.gui_log("⚠️ WARNING: Auto-filter detected in Excel file, but no hidden rows found.", level="WARNING")
            self.gui_log("⚠️ This may mean filters are not active, or the file needs to be saved with filters applied.", level="WARNING")
            self.gui_log("⚠️ To ensure filters are respected: Apply filters in Excel, then save the file.", level="WARNING")
            self.gui_log("⚠️ All visible rows will be processed. Consider using the row range filter if needed.", level="WARNING")
        
        # Convert to pandas DataFrame
        if visible_rows:
            df = pd.DataFrame(visible_rows, columns=headers)  # type: ignore
            # Add original row numbers as a column for filtering
            df['_original_row_number'] = original_row_numbers
        else:
            # No visible rows found, create empty DataFrame with headers
            self.gui_log("Warning: No visible rows found. Creating empty DataFrame.", level="WARNING")
            df = pd.DataFrame(columns=headers)  # type: ignore
            df['_original_row_number'] = []
        
        # Apply column mappings if specified
        df = self._apply_column_mappings(df)
        
        return df

    # ------------------------------------------------------------------
    # Selenium setup and authentication
    # ------------------------------------------------------------------
    def _init_chrome_driver(self) -> bool:
        if not SELENIUM_AVAILABLE:
            return False

        if self.driver:
            return True

        try:
            chrome_options = Options()
            chrome_options.add_argument("--disable-notifications")
            chrome_options.add_argument("--disable-infobars")
            chrome_options.add_argument("--start-maximized")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")

            if WEBDRIVER_MANAGER_AVAILABLE and ChromeDriverManager:
                driver_path = ChromeDriverManager().install()
                service = Service(driver_path)
                self.driver = webdriver.Chrome(service=service, options=chrome_options)
            else:
                self.driver = webdriver.Chrome(options=chrome_options)

            self.wait = WebDriverWait(self.driver, 10)
            self.gui_log("✅ Chrome browser launched successfully.", level="INFO")
            return True
        except WebDriverException as exc:
            self.log_error("Failed to initialize Chrome WebDriver", exception=exc)
            if self.root:
                self.root.after(0, lambda: messagebox.showerror(
                    "Chrome Driver Error",
                    f"Failed to start Chrome browser:\n{exc}\n\n"
                    "Ensure Chrome is installed and chromedriver is accessible."
                ))
            return False

    def _close_driver(self) -> None:
        try:
            if self.driver:
                self.gui_log("Closing browser...", level="DEBUG")
                self.driver.quit()
        except Exception as exc:
            self.log_error("Error closing Chrome driver", exception=exc, include_traceback=False)
        finally:
            self.driver = None
            self.wait = None
            self.is_logged_in = False

    def _ensure_logged_in(self) -> bool:
        if self.driver and self.is_logged_in:
            try:
                current_url = self.driver.current_url.lower()
                if "therapynotes" in current_url:
                    return True
            except Exception:
                pass

        return self._perform_login()

    def _perform_login(self) -> bool:
        if not self._init_chrome_driver():
            return False

        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()

        if not username or not password:
            self.gui_log("Username and password required for login.", level="ERROR")
            return False

        try:
            assert self.driver is not None
            assert self.wait is not None

            self.gui_log("Navigating to TherapyNotes login page...", level="INFO")
            self.driver.get(self.THERAPY_NOTES_URL)
            time.sleep(2)

            if self.check_stop_requested():
                return False

            username_field = self.wait.until(
                EC.presence_of_element_located((By.ID, "Login__UsernameField"))
            )
            username_field.clear()
            username_field.send_keys(username)

            password_field = self.wait.until(
                EC.presence_of_element_located((By.ID, "Login__Password"))
            )
            password_field.clear()
            password_field.send_keys(password)

            login_button = self.wait.until(
                EC.element_to_be_clickable((By.ID, "Login__LogInButton"))
            )
            login_button.click()
            self.gui_log("Login submitted, waiting for redirect...", level="INFO")

            time.sleep(3)
            if self.check_stop_requested():
                return False

            current_url = self.driver.current_url.lower()
            if "login" in current_url:
                self.gui_log("Login may have failed. Please verify credentials.", level="WARNING")
                return False

            self.is_logged_in = True
            self.update_status("Login successful.", "#28a745")
            self.gui_log("✅ Logged into TherapyNotes.", level="INFO")

            if not self._navigate_to_patients():
                self.gui_log("Could not navigate to Patients page automatically.", level="WARNING")
            return True

        except Exception as exc:
            self.log_error("Error during TherapyNotes login", exception=exc)
            self._close_driver()
            return False

    # ------------------------------------------------------------------
    # Core audit workflow helpers
    # ------------------------------------------------------------------
    def _column_letter_to_index(self, column_letter: str) -> int | None:
        """Convert Excel column letter (e.g., 'A', 'B', 'AA', 'AB') to 0-based index.
        
        Args:
            column_letter: Excel column letter(s) like 'A', 'B', 'Z', 'AA', 'AB', etc.
        
        Returns:
            0-based column index, or None if invalid
        """
        if not column_letter:
            return None
        
        column_letter = column_letter.strip().upper()
        
        # Check if it's a valid column letter (only A-Z)
        if not column_letter.isalpha():
            return None
        
        # Convert column letter to index (A=0, B=1, ..., Z=25, AA=26, AB=27, etc.)
        result = 0
        for char in column_letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1  # Convert to 0-based index
    
    def _index_to_column_letter(self, index: int) -> str:
        """Convert 0-based column index to Excel column letter (e.g., 0='A', 1='B', 26='AA').
        
        Args:
            index: 0-based column index
            
        Returns:
            Excel column letter(s) like 'A', 'B', 'AA', 'AB', etc.
        """
        result = ""
        index += 1  # Convert to 1-based for calculation
        while index > 0:
            index -= 1
            result = chr(65 + (index % 26)) + result
            index //= 26
        return result
    
    def _resolve_column_reference(self, user_input: str, df: pd.DataFrame) -> str | None:
        """Resolve user input to actual column name.
        
        Handles both column names and column letters (e.g., 'B', 'C', 'AB').
        
        Args:
            user_input: User input (column name or column letter)
            df: DataFrame to get column names from
        
        Returns:
            Column name if found, None otherwise
        """
        if not user_input or df.empty:
            return None
        
        user_input = user_input.strip()
        
        # First, try exact column name match
        if user_input in df.columns:
            return user_input
        
        # Try case-insensitive match
        for col in df.columns:
            if str(col).strip().lower() == user_input.lower():
                return col
        
        # Try as column letter (e.g., 'B', 'C', 'AB')
        col_index = self._column_letter_to_index(user_input)
        if col_index is not None and 0 <= col_index < len(df.columns):
            return df.columns[col_index]
        
        return None
    
    def _apply_column_mappings(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply user-specified column mappings to DataFrame.
        
        Renames columns according to user mappings so that extraction logic can find them.
        Supports both column names (e.g., "Excel_First_Name") and column letters (e.g., "B", "C", "AB").
        
        Example: If user maps "B" or "Excel_First_Name" -> "First Name", then that column 
        will be renamed to "First Name" so the extraction logic can find it.
        """
        if df.empty:
            return df
        
        # Create a copy to avoid modifying original
        df_mapped = df.copy()
        
        # Mapping from internal field names to expected column names
        field_to_standard_name = {
            "First Name": "First Name",
            "Last Name": "Last Name",
            "Date of Service": "Date of Service",
            "Original Modifier": "Original Modifier",
            "Expected Modifier": "Expected Modifier"
        }
        
        rename_dict = {}
        
        # Apply mappings: rename user-selected columns to standard names
        for field_name, user_input in self.column_mappings.items():
            if not user_input:
                continue
            
            # Resolve user input (column name or column letter) to actual column name
            actual_column = self._resolve_column_reference(user_input, df_mapped)
            
            if actual_column:
                standard_name = field_to_standard_name.get(field_name, field_name)
                if standard_name != actual_column:  # Only rename if different
                    rename_dict[actual_column] = standard_name
                    self.gui_log(f"Column mapping: '{user_input}' ({actual_column}) -> '{standard_name}'", level="DEBUG")
            else:
                self.gui_log(f"Warning: Column reference '{user_input}' not found for field '{field_name}'", level="WARNING")
        
        # Apply renames
        if rename_dict:
            df_mapped = df_mapped.rename(columns=rename_dict)
            self.gui_log(f"Applied {len(rename_dict)} column mapping(s)", level="INFO")
        
        return df_mapped
    
    def _normalize_column_key(self, key: str) -> str:
        key = (key or "").strip().lower()
        normalized = "".join(ch for ch in key if ch.isalnum())
        return normalized

    def _normalize_row_dict(self, row: Dict[str, Any]) -> Dict[str, Any]:
        normalized: Dict[str, Any] = {}
        for key, value in row.items():
            normalized[self._normalize_column_key(str(key))] = value
        return normalized

    def _value_to_string(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, )):
            return value.strftime("%m/%d/%Y")
        if hasattr(value, "strftime"):
            try:
                return value.strftime("%m/%d/%Y")
            except Exception:
                pass
        return str(value).strip()

    def _extract_from_synonyms(self, normalized_row: Dict[str, Any], synonyms: List[str]) -> str:
        for key in synonyms:
            if key in normalized_row:
                value = self._value_to_string(normalized_row[key])
                if value:
                    return value
        return ""

    def _extract_client_name_fields(self, normalized_row: Dict[str, Any]) -> str:
        # Check if user mapped "First Name" or "Last Name" columns
        # These will be normalized to "firstname" and "lastname"
        if "firstname" in normalized_row or "lastname" in normalized_row:
            first = self._extract_from_synonyms(normalized_row, ["firstname"])
            last = self._extract_from_synonyms(normalized_row, ["lastname"])
            if first or last:
                return f"{first} {last}".strip()
        
        # Try full name synonyms
        name_synonyms = [
            "clientname", "patientname", "client", "patient", "fullname", "name",
            "clientfullname", "patientfullname", "membername"
        ]
        name = self._extract_from_synonyms(normalized_row, name_synonyms)
        if name:
            return name

        # Try separate first/last name columns with various synonyms
        first_synonyms = ["firstname", "clientfirstname", "patientfirstname", "fname"]
        last_synonyms = ["lastname", "clientlastname", "patientlastname", "lname"]
        first = self._extract_from_synonyms(normalized_row, first_synonyms)
        last = self._extract_from_synonyms(normalized_row, last_synonyms)
        if first or last:
            return f"{first} {last}".strip()
        return ""

    def _extract_dob_field(self, normalized_row: Dict[str, Any]) -> str:
        dob_synonyms = [
            "dob", "dateofbirth", "birthdate", "clientdob", "patientdob", "memberdob"
        ]
        return self._extract_from_synonyms(normalized_row, dob_synonyms)

    def _extract_dos_field(self, normalized_row: Dict[str, Any]) -> str:
        dos_synonyms = [
            "dateofservice", "dos", "servicedate", "clientdos", "patientdos",
            "sessiondate", "visitdate", "billingdos", "date", "appointmentdate"
        ]
        return self._extract_from_synonyms(normalized_row, dos_synonyms)
    
    def _extract_patient_member_id_field(self, normalized_row: Dict[str, Any]) -> str:
        """Extract Patient Member ID (MBI) from normalized row.
        
        This extracts the Medicare Beneficiary Identifier (MBI) or Patient Member Number
        from the Excel file. This is the same value that should be used for the MBI field
        in the Medicare PDF form.
        """
        member_id_synonyms = [
            "patientmemberid", "memberid", "patientmember", "member",
            "clientmemberid", "patientid", "membernumber", "membernum",
            "mbi", "medicarebeneficiaryidentifier", "medicareid", "medicarenumber",
            "patientmember#", "patient member#", "patient member #", "member#",
            "medicarebeneficiaryid", "beneficiaryid", "beneficiarynumber"
        ]
        return self._extract_from_synonyms(normalized_row, member_id_synonyms)
    
    def _extract_service_code_field(self, normalized_row: Dict[str, Any]) -> str:
        """Extract Service Code (procedure code) from normalized row.
        
        This extracts the service/procedure code (e.g., 90837, 90834, 90835) from the input Excel.
        The user mentioned it's in column F "Service Code".
        Note: Column names are normalized to lowercase with no spaces/special chars (e.g., "Service Code" -> "servicecode")
        """
        # Try exact match first (normalized column name "Service Code" becomes "servicecode")
        if "servicecode" in normalized_row:
            value = self._value_to_string(normalized_row["servicecode"])
            if value:
                return value
        
        # Try other synonyms
        service_code_synonyms = [
            "servicecode",  # Primary match for "Service Code" column
            "service_code", "procedurecode", "procedure_code",
            "cptcode", "cpt_code", "cpt", "code", "service", "procedure",
            "servcode", "proc_code", "billingcode", "billing_code",
            "procedure", "servicedesc", "servicedescription"
        ]
        return self._extract_from_synonyms(normalized_row, service_code_synonyms)

    def _process_client_row(self, index: int, row: Dict[str, Any], total_rows: int) -> None:
        """Process a single client row with comprehensive error handling and cleanup."""
        normalized_row = self._normalize_row_dict(row)
        client_name = self._extract_client_name_fields(normalized_row)
        client_name = client_name.strip()
        dob = self._extract_dob_field(normalized_row).strip()
        dos = self._extract_dos_field(normalized_row).strip()
        patient_member_id = self._extract_patient_member_id_field(normalized_row).strip()
        service_code = self._extract_service_code_field(normalized_row).strip()
        if not dos:
            # Try alternate format if DOS stored as Excel serial number
            raw_dos = next((normalized_row[key] for key in normalized_row if "dos" in key or "date" in key), "")
            dos = self._value_to_string(raw_dos)

        if not client_name:
            self.gui_log(f"Row {index} skipped - missing client name.", level="WARNING")
            self.audit_results.append({
                "Client Name": "",
                "DOB": dob,
                "Date of Service": dos,
                "Patient Member #": patient_member_id,
                "Service Code": service_code or "",
                "Session Medium": "N/A",
                "Original Modifier": "N/A",
                "Expected Modifier": "N/A",
                "Status": "Skipped - Missing Name",
                "Notes": "Client name not provided in input file."
            })
            return

        if not dos:
            self.gui_log(f"Row {index} skipped - missing date of service for {client_name}.", level="WARNING")
            self.audit_results.append({
                "Client Name": client_name,
                "DOB": dob,
                "Date of Service": "",
                "Patient Member #": patient_member_id,
                "Service Code": service_code or "",
                "Session Medium": "N/A",
                "Original Modifier": "N/A",
                "Expected Modifier": "N/A",
                "Status": "Skipped - Missing DOS",
                "Notes": "Date of service not provided in input file."
            })
            return

        self.current_client_name = client_name
        self.current_client_dob = dob
        self.current_date_of_service = dos
        self.current_patient_member_id = patient_member_id
        self.current_service_code = service_code
        self.current_session_medium = None
        self.current_original_modifier = None

        self.gui_log(f"\n{'=' * 70}\nProcessing {index}/{total_rows}: {client_name} | DOB: {dob} | DOS: {dos}\n{'=' * 70}", level="INFO")

        # Wrap the entire processing in try-except to ensure cleanup
        try:
            if self.check_stop_requested():
                return

            if index == 1:
                if not self._navigate_to_patients():
                    self._record_failure("Navigation Failure", f"Unable to reach Patients page before processing {client_name}.")
                    return
            else:
                # Always ensure we're on the patients page before searching
                if not self._ensure_on_patients_page():
                    self._record_failure("Navigation Failure", f"Unable to navigate to Patients page for {client_name}.")
                    return
                self._reset_search_field()

            if not self._search_for_client(client_name, dob, dos):
                self._record_failure("Client Not Found", f"Could not locate client {client_name} in TherapyNotes.")
                return

            if self.check_stop_requested():
                return

            if not self._navigate_to_billing_tab():
                self._record_failure("Billing Tab Missing", "Unable to open Billing tab.")
                return

            if not self._click_all_items_button():
                self._record_failure("All Items Button Missing", "Could not click 'All Items' button.")
                return

            if not self._click_date_of_service(dos):
                self._record_failure("Date Not Found", f"Could not open date of service {dos}.")
                return

            time.sleep(1)
            if not self._click_notes_tab():
                self._record_failure("Notes Tab Missing", "Unable to open Notes tab in DOS popup.")
                return

            if not self._click_progress_note_button():
                self._record_failure("Note Page Missing", "Unable to open Progress/Consultation/Intake note.")
                return

            session_medium = (self.current_session_medium or "").lower()
            original_mod = (self.current_original_modifier or "N/A").upper()
            expected_mod = "N/A"
            status = "Medium Not Detected"
            notes = ""

            if session_medium in {"video", "phone"}:
                expected_mod = "95" if session_medium == "video" else "93"
                if original_mod == expected_mod:
                    status = "Modifier Correct"
                    notes = "Modifier matches session medium."
                else:
                    status = "Needs Refile"
                    notes = f"Expected modifier {expected_mod} for {session_medium} session."
            else:
                notes = "Could not determine session medium from note."

            self.audit_results.append({
                "Client Name": client_name,
                "DOB": dob,
                "Date of Service": dos,
                "Patient Member #": self.current_patient_member_id or "",
                "Service Code": self.current_service_code or "",
                "Session Medium": session_medium.upper() if session_medium else "UNKNOWN",
                "Original Modifier": original_mod,
                "Expected Modifier": expected_mod,
                "Status": status,
                "Notes": notes
            })

            self.gui_log(f"Result: {status} | Session: {session_medium or 'Unknown'} | Original: {original_mod} | Expected: {expected_mod}", level="INFO")

        except Exception as e:
            # Catch any unexpected exceptions and ensure we record the failure
            self.log_error(f"Unexpected error auditing row {index}", exception=e, include_traceback=True)
            self._record_failure("Unexpected Error", f"An unexpected error occurred: {str(e)}")
        finally:
            # Always navigate back to patients page, regardless of success or failure
            try:
                self._navigate_back_to_patients()
            except Exception as nav_error:
                self.gui_log(f"Error navigating back to patients page: {nav_error}", level="ERROR")
                # Try to force navigation
                try:
                    if self.driver:
                        self.driver.get("https://www.therapynotes.com/app/patients/")
                        time.sleep(2)
                except Exception:
                    pass

    def _record_failure(self, status: str, reason: str) -> None:
        self.gui_log(f"{status}: {reason}", level="WARNING")
        self.audit_results.append({
            "Client Name": self.current_client_name or "",
            "DOB": self.current_client_dob or "",
            "Date of Service": self.current_date_of_service or "",
            "Patient Member #": self.current_patient_member_id or "",
            "Service Code": self.current_service_code or "",
            "Session Medium": "N/A",
            "Original Modifier": "N/A",
            "Expected Modifier": "N/A",
            "Status": status,
            "Notes": reason
        })

    def _save_audit_results(self) -> None:
        if not self.audit_results:
            return

        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        if self.output_excel_path:
            output_path = self.output_excel_path
        else:
            output_dir = SCRIPT_DIR / "output"
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = output_dir / f"medicare_refiling_audit_{timestamp}.xlsx"

        try:
            df = pd.DataFrame(self.audit_results)  # type: ignore
            df.to_excel(output_path, index=False)
            self.gui_log(f"✅ Audit results saved to {output_path}", level="INFO")
            self.update_status(f"Output saved: {output_path.name}", "#28a745")
        except Exception as exc:
            self.log_error("Failed to save audit Excel", exception=exc)

    # ------------------------------------------------------------------
    # Navigation helpers (adapted from TN Refiling Bot)
    # ------------------------------------------------------------------
    def _navigate_to_patients(self) -> bool:
        if not self.driver or not self.wait:
            return False
        try:
            self.gui_log("Navigating to Patients page...", level="DEBUG")
            self.driver.get("https://www.therapynotes.com/app/patients/")
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, "ctl00_BodyContent_TextBoxSearchPatientName"))
            )
            time.sleep(1)
            return True
        except Exception as exc:
            self.log_error("Failed to navigate to Patients page", exception=exc, include_traceback=False)
            return False

    def _reset_search_field(self) -> None:
        if not self.driver:
            return
        try:
            search_field = self.driver.find_element(By.ID, "ctl00_BodyContent_TextBoxSearchPatientName")
            search_field.clear()
        except Exception:
            try:
                search_field = self.driver.find_element(By.NAME, "ctl00$BodyContent$TextBoxSearchPatientName")
                search_field.clear()
            except Exception:
                pass

    def _normalize_client_name(self, client_name: str) -> Tuple[str, str]:
        """Normalize client name for searching in Therapy Notes
        
        Removes periods from middle initials and prepares name variations
        Example: "John H. Doe" -> ("John H Doe", "John Doe")
        
        Args:
            client_name: The client's name (e.g., "John H. Doe" or "Jane Smith")
        
        Returns:
            tuple: (full_name_with_middle, name_without_middle) - both normalized
        """
        if not client_name:
            return ("", "")
        
        # Remove extra spaces
        name = " ".join(client_name.split())
        
        # Remove periods from middle initials (e.g., "John H. Doe" -> "John H Doe")
        # Pattern: letter + period + space (middle initial)
        name = re.sub(r'([A-Z])\.\s', r'\1 ', name)
        
        # Split into parts
        parts = name.split()
        
        if len(parts) >= 3:
            # Has middle name/initial
            first = parts[0]
            middle = parts[1]
            last = " ".join(parts[2:])  # Handle last names with spaces
            
            full_name = f"{first} {middle} {last}"
            name_without_middle = f"{first} {last}"
            
            return (full_name, name_without_middle)
        elif len(parts) == 2:
            # Just first and last name
            return (name, name)
        else:
            # Single name or empty
            return (name, name)

    def _normalize_dob_for_comparison(self, dob_str: str) -> Tuple[Optional[str], Optional[datetime]]:
        """Normalize DOB string for comparison across different formats
        
        Handles formats like:
        - MM/DD/YYYY (e.g., "01/22/1963")
        - MM/DD/YY (e.g., "01/22/63")
        - YYYY/MM/DD (e.g., "1963/01/22")
        - YY/MM/DD (e.g., "63/01/22")
        
        Args:
            dob_str: Date of birth as string
        
        Returns:
            tuple: (normalized_dob_str, normalized_dob_date) or (None, None) if invalid
            Format: "YYYY-MM-DD" string and date object for comparison
        """
        if not dob_str or str(dob_str).strip() in ['', 'nan', 'None', 'N/A']:
            return (None, None)
        
        try:
            dob_str = str(dob_str).strip()
            
            # Remove time component if present (Excel dates sometimes include time)
            if ' ' in dob_str:
                dob_str = dob_str.split()[0]
            
            # Try different parsing strategies
            parsed_date = None
            
            # Strategy 1: Try parsing with slashes (MM/DD/YYYY or MM/DD/YY)
            if '/' in dob_str:
                parts = dob_str.split('/')
                if len(parts) == 3:
                    month = parts[0].strip()
                    day = parts[1].strip()
                    year = parts[2].strip()
                    
                    # Handle 2-digit year
                    # For DOBs, always default to 1900s since this is Medicare refiling
                    # Medicare patients are 65+, so they were born in 1960 or earlier (1900s)
                    if len(year) == 2:
                        year_int = int(year)
                        # Always use 1900s for DOBs (Medicare context: patients are 65+)
                        year = f"19{year_int:02d}"
                    
                    # Try MM/DD/YYYY format
                    try:
                        parsed_date = datetime(int(year), int(month), int(day))
                    except:
                        # Try YYYY/MM/DD format
                        try:
                            parsed_date = datetime(int(month), int(day), int(year))
                            # If this worked but year is obviously wrong, swap back
                            if int(month) > 12:
                                # month was actually year
                                parsed_date = datetime(int(month), int(day), int(year))
                        except:
                            pass
            
            # Strategy 2: Try parsing with dashes (YYYY-MM-DD)
            if not parsed_date and '-' in dob_str:
                try:
                    parts = dob_str.split('-')
                    if len(parts) == 3:
                        # Assume YYYY-MM-DD format
                        parsed_date = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                except:
                    pass
            
            # Strategy 3: Try pandas datetime parsing (handles many Excel formats)
            if not parsed_date:
                try:
                    import pandas as pd
                    parsed_date = pd.to_datetime(dob_str)
                    if isinstance(parsed_date, pd.Timestamp):
                        parsed_date = parsed_date.to_pydatetime()
                except:
                    pass
            
            # Strategy 4: Try datetime.strptime with common formats
            # Note: strptime with %y interprets 00-68 as 2000-2068, 69-99 as 1969-1999
            # For Medicare DOBs, we want all 2-digit years to be 1900s
            # So we preprocess the date string to convert 2-digit years to 1900s before strptime
            if not parsed_date:
                # Preprocess: Convert 2-digit years to 1900s for Medicare context
                preprocessed_dob = dob_str
                if '/' in preprocessed_dob:
                    parts = preprocessed_dob.split('/')
                    if len(parts) == 3:
                        # Check if year is 2 digits
                        year_part = parts[2].strip()
                        if len(year_part) == 2 and year_part.isdigit():
                            # Convert to 1900s
                            year_int = int(year_part)
                            parts[2] = f"19{year_int:02d}"
                            preprocessed_dob = '/'.join(parts)
                elif '-' in preprocessed_dob:
                    parts = preprocessed_dob.split('-')
                    if len(parts) == 3:
                        # Check if first part is 4-digit year (YYYY-MM-DD format)
                        first_part = parts[0].strip()
                        last_part = parts[2].strip()
                        
                        if len(first_part) == 4 and first_part.isdigit():
                            # YYYY-MM-DD format - year is already 4 digits
                            pass
                        elif len(last_part) == 2 and last_part.isdigit():
                            # MM-DD-YY format - convert last part (year) to 1900s
                            year_int = int(last_part)
                            parts[2] = f"19{year_int:02d}"
                            preprocessed_dob = '-'.join(parts)
                
                date_formats = [
                    '%m/%d/%Y',
                    '%Y/%m/%d',
                    '%d/%m/%Y',
                    '%Y-%m-%d',
                    '%m-%d-%Y',
                ]
                
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(preprocessed_dob, fmt)
                        break
                    except:
                        continue
            
            if parsed_date:
                # Normalize to YYYY-MM-DD string for comparison
                normalized_str = parsed_date.strftime('%Y-%m-%d')
                return (normalized_str, parsed_date)
            else:
                self.gui_log(f"⚠️ Could not parse DOB: {dob_str}", level="WARNING")
                return (None, None)
                
        except Exception as e:
            self.gui_log(f"Error normalizing DOB '{dob_str}': {e}", level="WARNING")
            return (None, None)

    def _normalize_date_of_service(self, dos_str: str) -> Tuple[Optional[str], Optional[datetime]]:
        if not dos_str:
            return None, None
        dos_str = str(dos_str).strip()
        try:
            dt = datetime.strptime(dos_str, "%m/%d/%Y")
        except ValueError:
            try:
                dt = datetime.strptime(dos_str, "%Y-%m-%d")
            except ValueError:
                try:
                    dt = datetime.strptime(dos_str, "%m/%d/%y")
                except ValueError:
                    return None, None
        normalized = f"{dt.month}/{dt.day}/{str(dt.year)[-2:]}"
        return normalized, dt

    def _search_for_client(self, client_name: str, dob: str, dos: str) -> bool:
        if not self.driver or not self.wait:
            return False

        normalized_dob, _ = self._normalize_dob_for_comparison(dob)
        full_name, name_no_middle = self._normalize_client_name(client_name)

        name_variants = []
        if full_name:
            name_variants.append(full_name)
        if name_no_middle and name_no_middle != full_name:
            name_variants.append(name_no_middle)

        if not name_variants:
            name_variants.append(client_name)

        try:
            search_field = self.wait.until(
                EC.presence_of_element_located((By.ID, "ctl00_BodyContent_TextBoxSearchPatientName"))
            )
        except TimeoutException:
            try:
                search_field = self.wait.until(
                    EC.presence_of_element_located((By.NAME, "ctl00$BodyContent$TextBoxSearchPatientName"))
                )
            except TimeoutException:
                self.gui_log("Unable to locate TherapyNotes patient search field.", level="ERROR")
                return False

        # Scroll into view
        self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", search_field)
        time.sleep(0.5)

        for attempt, name_variant in enumerate(name_variants, 1):
            if self.check_stop_requested():
                return False

            try:
                self.gui_log(f"Search attempt {attempt}/{len(name_variants)}: '{name_variant}'", level="INFO")
                
                # Clear search field
                search_field.clear()
                time.sleep(0.2)  # Small delay to ensure clear is processed
                
                # Enter client name
                search_field.send_keys(name_variant)
                self.gui_log(f"✅ Entered name: '{name_variant}'", level="INFO")

                # Wait for dropdown menu to appear dynamically (no fixed delay - wait handles it)
                dropdown = self._wait_for_dropdown()
                if not dropdown:
                    self.gui_log("⚠️ Dropdown menu did not appear - client may not exist or dropdown may have different structure", level="WARNING")
                    # Continue to next search attempt or return False
                    if attempt < len(name_variants):
                        continue  # Try next name variation
                    else:
                        return False  # All attempts failed

                # Find and match client by DOB
                match_result = self._match_client_in_dropdown(dropdown, normalized_dob, client_name)
                
                if match_result == True:
                    self.gui_log(f"✅ Successfully matched and clicked client: {client_name}", level="INFO")
                    time.sleep(1.5)
                    return True
                else:
                    self.gui_log(f"⚠️ Client not found in dropdown for: {name_variant}", level="WARNING")
                    # Try next name variation if available
                    if attempt < len(name_variants):
                        self.gui_log(f"Retrying with different name variation...", level="INFO")
                        continue
                    else:
                        return False  # All attempts failed
                        
            except Exception as exc:
                self.gui_log(f"Error during search attempt {attempt}: {exc}", level="WARNING")
                if attempt < len(name_variants):
                    self.gui_log(f"Retrying with different name variation...", level="INFO")
                    continue
                else:
                    self.log_error(f"All search attempts failed for client: {client_name}", exception=exc, include_traceback=True)
                    return False

        return False

    def _wait_for_dropdown(self):
        """Wait for dropdown menu to appear after typing in search field"""
        # Wait a moment for dropdown to appear dynamically (no fixed delay - wait handles it)
        self.gui_log("Waiting for dropdown menu to appear...", level="DEBUG")
        
        # Wait for dropdown container to appear using multiple strategies
        dropdown_found = False
        dropdown_container = None
        
        # Strategy 1: Find by ID (try capital C first, then lowercase - like consent bot v2)
        try:
            # Try ContentBubbleResultsContainer (capital C) - this is what consent bot v2 uses
            dropdown_container = self.wait.until(
                EC.presence_of_element_located((By.ID, "ContentBubbleResultsContainer"))
            )
            self.gui_log("✅ Dropdown menu appeared (by ID: ContentBubbleResultsContainer)", level="INFO")
            dropdown_found = True
        except TimeoutException:
            self.gui_log("Dropdown not found by ID (ContentBubbleResultsContainer), trying lowercase...", level="DEBUG")
            # Try lowercase version
            try:
                dropdown_container = self.wait.until(
                    EC.presence_of_element_located((By.ID, "contentbubbleresultscontainer"))
                )
                self.gui_log("✅ Dropdown menu appeared (by ID: contentbubbleresultscontainer)", level="INFO")
                dropdown_found = True
            except TimeoutException:
                self.gui_log("Dropdown not found by ID (either case), trying alternative methods...", level="DEBUG")
        
        # Strategy 2: Find by partial ID match
        if not dropdown_container:
            try:
                # Try finding by partial ID
                dropdown_container = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, "//div[contains(@id, 'contentbubble') or contains(@id, 'resultscontainer')]"))
                )
                self.gui_log("✅ Dropdown menu appeared (by partial ID)", level="INFO")
                dropdown_found = True
            except TimeoutException:
                self.gui_log("Dropdown not found by partial ID...", level="DEBUG")
        
        # Strategy 3: Find by class or attribute patterns
        if not dropdown_container:
            try:
                # Look for div elements that might be the dropdown container
                # Common patterns: contains "dropdown", "results", "bubble", "autocomplete"
                dropdown_container = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'dropdown') or contains(@class, 'results') or contains(@class, 'autocomplete') or contains(@class, 'bubble')]"))
                )
                self.gui_log("✅ Dropdown menu appeared (by class pattern)", level="INFO")
                dropdown_found = True
            except TimeoutException:
                self.gui_log("Dropdown not found by class pattern...", level="DEBUG")
        
        if dropdown_found and dropdown_container:
            return dropdown_container
        else:
            self.gui_log("⚠️ Dropdown menu did not appear - client may not exist or dropdown may have different structure", level="WARNING")
            return None

    def _match_client_in_dropdown(self, dropdown_container, expected_dob_str: Optional[str], client_name: str) -> bool:
        """Match client in dropdown by DOB and click the matching option
        
        Args:
            dropdown_container: The dropdown container element (id: contentbubbleresultscontainer)
            expected_dob_str: Normalized DOB string from Excel (format: "YYYY-MM-DD")
            client_name: The client's name for logging
        
        Returns:
            bool: True if client matched and clicked, False otherwise
        """
        try:
            self.gui_log("Matching client in dropdown by DOB...", level="INFO")
            
            # Find all client options in the dropdown
            # The structure will depend on how Therapy Notes displays the dropdown
            # Try multiple strategies to find clickable client items
            client_options = []
            
            # Strategy 1: Find all links/clickable elements within dropdown
            try:
                # IMPORTANT: Therapy Notes dropdown uses .ui-menu-item class for each result
                # This is the same approach used in consent bot v2
                
                # Wait a bit more for results to populate (like consent bot v2 does)
                time.sleep(0.5)
                
                # Strategy 1: Look for .ui-menu-item elements (primary method from consent bot v2)
                result_selectors = [
                    "div.ui-menu-item",
                    ".ui-menu-item",
                    "li.ui-menu-item",
                    "a.ui-menu-item",
                    "*[class*='menu-item']",
                    "div[class*='result']",
                ]
                
                for sel in result_selectors:
                    try:
                        # Search WITHIN the container, not globally (critical!)
                        results = dropdown_container.find_elements(By.CSS_SELECTOR, sel)
                        if results:
                            client_options.extend(results)
                            self.gui_log(f"Found {len(results)} results using selector: {sel}", level="DEBUG")
                            break  # Use first successful selector
                    except Exception as e:
                        self.gui_log(f"Error with selector {sel}: {e}", level="DEBUG")
                        continue
                
                # Strategy 2: If no .ui-menu-item found, try finding any clickable elements in the container
                if not client_options:
                    try:
                        all_elements = dropdown_container.find_elements(By.TAG_NAME, "*")
                        clickable = [el for el in all_elements if el.is_displayed() and el.text.strip() and len(el.text.strip()) > 3]
                        if clickable:
                            client_options.extend(clickable[:10])  # Limit to first 10 to avoid too many
                            self.gui_log(f"Found {len(clickable)} clickable elements in container (fallback)", level="DEBUG")
                    except Exception as e:
                        self.gui_log(f"Error finding clickable elements: {e}", level="DEBUG")
                
                # Strategy 3: Fallback - look for links and divs (original approach)
                if not client_options:
                    try:
                        all_elements = dropdown_container.find_elements(By.TAG_NAME, "a")
                        client_options.extend(all_elements)
                        
                        list_items = dropdown_container.find_elements(By.TAG_NAME, "li")
                        client_options.extend(list_items)
                        
                        # Look for divs with click handlers
                        divs = dropdown_container.find_elements(By.TAG_NAME, "div")
                        for div in divs:
                            onclick = div.get_attribute("onclick")
                            role = div.get_attribute("role")
                            style = div.get_attribute("style") or ""
                            has_text = div.text.strip()
                            
                            if onclick or role == "button" or "cursor:pointer" in style.lower():
                                client_options.append(div)
                            elif has_text and len(has_text) > 3 and re.search(r'\d{1,2}/\d{1,2}/\d{2,4}', has_text):
                                client_options.append(div)
                    except Exception as e:
                        self.gui_log(f"Error in fallback element finding: {e}", level="DEBUG")
                
                # Remove duplicates while preserving order
                seen = set()
                unique_options = []
                for opt in client_options:
                    opt_id = id(opt)
                    if opt_id not in seen:
                        seen.add(opt_id)
                        unique_options.append(opt)
                client_options = unique_options
                
                self.gui_log(f"Found {len(client_options)} potential client option(s) in dropdown", level="DEBUG")
                
            except Exception as e:
                self.gui_log(f"Error finding dropdown options: {e}", level="WARNING")
                import traceback
                self.gui_log(traceback.format_exc(), level="DEBUG")
            
            if not client_options:
                self.gui_log("⚠️ No client options found in dropdown", level="WARNING")
                return False
            
            # Extract and match DOB from each option
            for option_idx, option in enumerate(client_options, 1):
                try:
                    # Get the text content of this option
                    option_text = option.text.strip()
                    
                    if not option_text:
                        continue
                    
                    self.gui_log(f"Checking option {option_idx}: {option_text[:80]}...", level="DEBUG")
                    
                    # Extract DOB from option text
                    # Therapy Notes dropdown format: "Name - 9/18/1955" or "Name (9/18/1955)" or "Name 9/18/1955"
                    # DOB format in dropdown: MM/DD/YYYY (e.g., "9/18/1955")
                    option_dob_str = None
                    
                    # Try to find DOB in the text using various patterns
                    # Pattern 1: MM/DD/YYYY or MM/DD/YY (Therapy Notes format: "9/18/1955" or "9/18/55")
                    # This is the primary format in Therapy Notes dropdown
                    dob_pattern1 = r'(\d{1,2}/\d{1,2}/\d{2,4})'
                    match = re.search(dob_pattern1, option_text)
                    if match:
                        dob_in_text = match.group(1)
                        self.gui_log(f"   Found DOB in dropdown text: {dob_in_text}", level="DEBUG")
                        option_dob_str, _ = self._normalize_dob_for_comparison(dob_in_text)
                        if option_dob_str:
                            self.gui_log(f"   Normalized dropdown DOB: {option_dob_str}", level="DEBUG")
                    
                    # Pattern 2: YYYY-MM-DD
                    if not option_dob_str:
                        dob_pattern2 = r'(\d{4}-\d{2}-\d{2})'
                        match = re.search(dob_pattern2, option_text)
                        if match:
                            option_dob_str = match.group(1)  # Already normalized
                    
                    # Pattern 3: Look for dates in parentheses or after dashes
                    if not option_dob_str:
                        # Try extracting from text after last "-" or within parentheses
                        if '-' in option_text:
                            parts = option_text.split('-')
                            if len(parts) > 1:
                                # Last part might contain DOB
                                potential_dob = parts[-1].strip()
                                option_dob_str, _ = self._normalize_dob_for_comparison(potential_dob)
                        
                        if not option_dob_str and '(' in option_text:
                            # Extract text within parentheses
                            paren_match = re.search(r'\(([^)]+)\)', option_text)
                            if paren_match:
                                potential_dob = paren_match.group(1).strip()
                                option_dob_str, _ = self._normalize_dob_for_comparison(potential_dob)
                    
                    if option_dob_str and expected_dob_str:
                        # Compare DOBs
                        if option_dob_str == expected_dob_str:
                            self.gui_log(f"✅ DOB match found! Option: {option_text[:80]}", level="INFO")
                            self.gui_log(f"   Matched DOB: {option_dob_str}", level="INFO")
                            
                            # Scroll option into view
                            self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", option)
                            time.sleep(0.5)
                            
                            # Click the matching option
                            # First, try to find the actual clickable link within the option element
                            clickable_element = None
                            
                            try:
                                # If option is already a link, use it directly
                                if option.tag_name.lower() == "a":
                                    clickable_element = option
                                else:
                                    # Try to find a link within the option element
                                    try:
                                        inner_link = option.find_element(By.TAG_NAME, "a")
                                        if inner_link:
                                            clickable_element = inner_link
                                    except:
                                        pass
                                    
                                    # If no link found, check if the option element itself is clickable
                                    if not clickable_element:
                                        # Check if element has click handlers or is interactive
                                        onclick = option.get_attribute("onclick")
                                        href = option.get_attribute("href")
                                        role = option.get_attribute("role")
                                        
                                        if onclick or href or role in ["button", "link", "option"]:
                                            clickable_element = option
                                        else:
                                            # As last resort, use the option element itself
                                            clickable_element = option
                                
                                if not clickable_element:
                                    self.gui_log("⚠️ Could not find clickable element within option", level="WARNING")
                                    clickable_element = option  # Fallback to original element
                                
                                # Scroll element into view first
                                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", clickable_element)
                                time.sleep(0.3)
                                
                                # Try multiple click strategies
                                clicked = False
                                
                                # Strategy 1: Try regular click
                                try:
                                    clickable_element.click()
                                    self.gui_log(f"✅ Clicked matching client option (regular click)", level="INFO")
                                    clicked = True
                                except Exception as click_error:
                                    self.gui_log(f"Regular click failed, trying JavaScript click: {click_error}", level="DEBUG")
                                
                                # Strategy 2: Try JavaScript click if regular click failed
                                if not clicked:
                                    try:
                                        self.driver.execute_script("arguments[0].click();", clickable_element)
                                        self.gui_log(f"✅ Clicked matching client option (JavaScript)", level="INFO")
                                        clicked = True
                                    except Exception as js_error:
                                        self.gui_log(f"JavaScript click failed, trying ActionChains: {js_error}", level="DEBUG")
                                
                                # Strategy 3: Try ActionChains move_to_element and click
                                if not clicked:
                                    try:
                                        actions = ActionChains(self.driver)
                                        actions.move_to_element(clickable_element).click().perform()
                                        self.gui_log(f"✅ Clicked matching client option (ActionChains)", level="INFO")
                                        clicked = True
                                    except Exception as ac_error:
                                        self.gui_log(f"ActionChains click failed: {ac_error}", level="WARNING")
                                
                                if not clicked:
                                    raise Exception("All click strategies failed")
                                
                                # Wait for navigation to complete
                                time.sleep(2)
                                
                                # Verify we navigated to client page
                                current_url = self.driver.current_url
                                self.gui_log(f"Current URL after clicking client: {current_url}", level="DEBUG")
                                
                                # Verify navigation by checking URL changed or waiting for page elements
                                if "patients" in current_url.lower() and "edit" in current_url.lower():
                                    self.gui_log("✅ Verified navigation to client page", level="INFO")
                                
                                return True
                                
                            except Exception as click_error:
                                self.log_error(f"Failed to click client option after all strategies: {click_error}", exception=click_error, include_traceback=True)
                                
                                # Last resort: Try to find any link in the dropdown that matches and click it
                                try:
                                    self.gui_log("Attempting last resort: searching for matching link by text...", level="WARNING")
                                    all_links = dropdown_container.find_elements(By.TAG_NAME, "a")
                                    for link in all_links:
                                        link_text = link.text.strip()
                                        if option_text in link_text or link_text in option_text:
                                            self.driver.execute_script("arguments[0].click();", link)
                                            self.gui_log(f"✅ Clicked client using text match fallback", level="INFO")
                                            time.sleep(2)
                                            return True
                                except Exception as fallback_error:
                                    self.gui_log(f"Fallback click also failed: {fallback_error}", level="ERROR")
                                
                                return False
                        else:
                            self.gui_log(f"   DOB mismatch - Option DOB: {option_dob_str}, Expected: {expected_dob_str}", level="DEBUG")
                    elif not option_dob_str:
                        self.gui_log(f"   Could not extract DOB from option text", level="DEBUG")
                    elif not expected_dob_str:
                        self.gui_log(f"   ⚠️ No expected DOB provided for matching", level="WARNING")
                        # If no DOB provided, might need to select by name only or return False
                        # For now, continue checking other options
                
                except Exception as e:
                    self.gui_log(f"Error checking option {option_idx}: {e}", level="WARNING")
                    continue
            
            # If we get here, no matching client was found
            self.gui_log(f"❌ No matching client found in dropdown with DOB: {expected_dob_str}", level="WARNING")
            return False
            
        except Exception as e:
            self.log_error("Error matching client in dropdown", exception=e, include_traceback=True)
            return False

    def _navigate_to_billing_tab(self) -> bool:
        """Navigate to Patient Billing tab on client page
        
        Looks for the billing tab with data-tab-id="Patient Billing" or href="#tab=Patient+Billing"
        """
        if not self.driver or not self.wait:
            return False
        
        self.gui_log("Navigating to Patient Billing tab...", level="DEBUG")
        
        try:
            # Strategy 1: Find by data-tab-id attribute (most specific)
            try:
                billing_tab = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-tab-id='Patient Billing']"))
                )
                self.gui_log("✅ Found Patient Billing tab (by data-tab-id)", level="INFO")
                billing_tab.click()
                time.sleep(2)
                return True
            except TimeoutException:
                self.gui_log("Billing tab not found by data-tab-id, trying href...", level="DEBUG")
            
            # Strategy 2: Find by href attribute
            try:
                billing_tab = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@href='#tab=Patient+Billing']"))
                )
                self.gui_log("✅ Found Patient Billing tab (by href)", level="INFO")
                billing_tab.click()
                time.sleep(2)
                return True
            except TimeoutException:
                self.gui_log("Billing tab not found by href, trying partial href match...", level="DEBUG")
            
            # Strategy 3: Find by partial href match (case-insensitive)
            try:
                billing_tab = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[contains(@href, 'Patient+Billing') or contains(@href, 'Patient%20Billing') or contains(@href, 'tab=Patient')]"))
                )
                self.gui_log("✅ Found Patient Billing tab (by partial href)", level="INFO")
                billing_tab.click()
                time.sleep(2)
                return True
            except TimeoutException:
                self.gui_log("Billing tab not found by partial href, trying link text...", level="DEBUG")
            
            # Strategy 4: Fallback to link text (but verify it's the Patient Billing tab)
            try:
                # Find all "Billing" links and look for the one with Patient Billing context
                billing_links = self.driver.find_elements(By.LINK_TEXT, "Billing")
                for link in billing_links:
                    # Check if this link has Patient Billing attributes or is in Patient Billing context
                    href = link.get_attribute("href") or ""
                    data_tab_id = link.get_attribute("data-tab-id") or ""
                    if "Patient+Billing" in href or "Patient%20Billing" in href or data_tab_id == "Patient Billing":
                        self.gui_log("✅ Found Patient Billing tab (by link text with verification)", level="INFO")
                        link.click()
                        time.sleep(2)
                        return True
                
                # If no Patient Billing link found, try the first Billing link (last resort)
                if billing_links:
                    self.gui_log("⚠️ Using first 'Billing' link found (may not be Patient Billing)", level="WARNING")
                    billing_links[0].click()
                    time.sleep(2)
                    return True
            except Exception as e:
                self.gui_log(f"Error finding billing tab by link text: {e}", level="DEBUG")
            
            self.gui_log("❌ Could not find Patient Billing tab", level="ERROR")
            return False
            
        except Exception as e:
            self.log_error("Error navigating to billing tab", exception=e, include_traceback=False)
            return False

    def _click_all_items_button(self) -> bool:
        """Click the 'All Items' button in the billing transactions filter
        
        This button shows all billing transactions, not just filtered ones.
        """
        if not self.driver or not self.wait:
            return False
        
        self.gui_log("Clicking 'All Items' button...", level="DEBUG")
        
        try:
            # Strategy 1: Find by ID (most specific)
            try:
                button = self.wait.until(
                    EC.element_to_be_clickable((By.ID, "SearchBillingTransactionsFilter__AllItems"))
                )
                self.gui_log("✅ Found 'All Items' button (by ID)", level="INFO")
                button.click()
                time.sleep(1.5)  # Wait for transactions to load
                return True
            except TimeoutException:
                self.gui_log("'All Items' button not found by ID, trying alternative methods...", level="DEBUG")
            
            # Strategy 2: Find by partial ID match
            try:
                button = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@id, 'AllItems') or contains(@id, 'all-items')]"))
                )
                self.gui_log("✅ Found 'All Items' button (by partial ID)", level="INFO")
                button.click()
                time.sleep(1.5)
                return True
            except TimeoutException:
                self.gui_log("'All Items' button not found by partial ID...", level="DEBUG")
            
            # Strategy 3: Find by button text
            try:
                button = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'All Items') or contains(text(), 'All')]"))
                )
                self.gui_log("✅ Found 'All Items' button (by text)", level="INFO")
                button.click()
                time.sleep(1.5)
                return True
            except TimeoutException:
                self.gui_log("'All Items' button not found by text...", level="DEBUG")
            
            # Strategy 4: Find by link text (if it's a link instead of button)
            try:
                link = self.wait.until(
                    EC.element_to_be_clickable((By.LINK_TEXT, "All Items"))
                )
                self.gui_log("✅ Found 'All Items' link (by link text)", level="INFO")
                link.click()
                time.sleep(1.5)
                return True
            except TimeoutException:
                self.gui_log("'All Items' link not found by link text...", level="DEBUG")
            
            self.gui_log("❌ Could not find 'All Items' button", level="ERROR")
            return False
            
        except Exception as e:
            self.log_error("Error clicking 'All Items' button", exception=e, include_traceback=False)
            return False

    def _click_date_of_service(self, dos: str) -> bool:
        """Click on the date of service link in the billing transactions table
        
        Args:
            dos: Date of service string (e.g., "08/01/2025")
        
        Returns:
            bool: True if date of service was found and clicked, False otherwise
        """
        if not self.driver or not self.wait:
            return False
        
        self.gui_log(f"Looking for date of service: {dos}", level="INFO")
        
        # Normalize the date of service for comparison
        normalized_dos, dos_dt = self._normalize_date_of_service(dos)
        if not normalized_dos or not dos_dt:
            self.gui_log(f"⚠️ Could not normalize date of service: {dos}", level="WARNING")
            return False
        
        self.gui_log(f"Normalized DOS: {normalized_dos}", level="DEBUG")
        
        try:
            # Wait for billing transaction links to appear
            self.gui_log("Waiting for billing transaction links to appear...", level="DEBUG")
            links = self.wait.until(
                EC.presence_of_all_elements_located((By.XPATH, "//a[@data-testid='billingstatementtable-paymentdate-link']"))
            )
            self.gui_log(f"Found {len(links)} billing transaction link(s)", level="INFO")
        except TimeoutException:
            self.gui_log("⚠️ No billing transaction links found - table may be empty or still loading", level="WARNING")
            return False

        # Find the matching date of service link
        # Use XPath to find by text to avoid stale element issues
        found_dates = []
        matched_text = None
        target_link_idx = None
        
        # First, collect all dates by re-finding links each time to avoid stale elements
        max_retries = 3
        for retry in range(max_retries):
            try:
                # Re-find links each iteration to avoid stale element reference
                current_links = self.driver.find_elements(By.XPATH, "//a[@data-testid='billingstatementtable-paymentdate-link']")
                if not current_links:
                    if retry < max_retries - 1:
                        time.sleep(1)
                        continue
                    break
                
                for idx, link in enumerate(current_links):
                    try:
                        text = link.text.strip()
                        normalized_link, link_dt = self._normalize_date_of_service(text)
                        if normalized_link:
                            found_dates.append(normalized_link)
                            if normalized_link == normalized_dos:
                                matched_text = text
                                target_link_idx = idx
                                self.gui_log(f"✅ Found matching DOS link: {text} -> {normalized_link}", level="INFO")
                                break
                    except Exception as e:
                        # If we get a stale element or other error, continue to next link
                        self.gui_log(f"Error checking link {idx}: {e}", level="DEBUG")
                        continue
                
                if matched_text:
                    break
                    
            except Exception as e:
                self.gui_log(f"Error finding links (retry {retry + 1}/{max_retries}): {e}", level="DEBUG")
                if retry < max_retries - 1:
                    time.sleep(1)
                    continue

        if not matched_text:
            self.gui_log(f"❌ Date of service {dos} (normalized: {normalized_dos}) not found in billing transactions", level="ERROR")
            if found_dates:
                self.gui_log(f"Available dates in table: {', '.join(found_dates[:10])}{'...' if len(found_dates) > 10 else ''}", level="DEBUG")
            return False

        # Click the date of service link using XPath to avoid stale element issues
        try:
            self.gui_log(f"Clicking date of service link: {matched_text}", level="INFO")
            
            # Strategy 1: Use XPath to find the link by text (most reliable)
            try:
                # Use multiple XPath strategies to handle text matching reliably
                # Escape single quotes by using concat() or contains() with proper quoting
                # Try exact match first, then fall back to contains()
                xpaths = [
                    f"//a[@data-testid='billingstatementtable-paymentdate-link' and normalize-space(text())='{matched_text}']",
                    f"//a[@data-testid='billingstatementtable-paymentdate-link' and contains(normalize-space(text()), '{matched_text}')]"
                ]
                clickable_link = None
                for xpath in xpaths:
                    try:
                        # Use shorter timeout for each attempt
                        short_wait = WebDriverWait(self.driver, 3)
                        clickable_link = short_wait.until(
                            EC.element_to_be_clickable((By.XPATH, xpath))
                        )
                        # Verify the text matches exactly
                        if clickable_link.text.strip() == matched_text:
                            break
                        clickable_link = None
                    except (TimeoutException, Exception):
                        continue
                
                if not clickable_link:
                    raise TimeoutException("Could not find clickable link with any XPath strategy")
                
                # Scroll into view
                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", clickable_link)
                time.sleep(0.5)
                
                # Try regular click first
                try:
                    clickable_link.click()
                except Exception:
                    # Try JavaScript click as fallback
                    self.driver.execute_script("arguments[0].click();", clickable_link)
                
                time.sleep(2)  # Wait for DOS popup to open
                
                # Extract original modifier after clicking
                self._extract_original_modifier()
                return True
                
            except (TimeoutException, Exception) as e1:
                # Strategy 2: If XPath fails, try finding by index
                if target_link_idx is not None:
                    try:
                        self.gui_log(f"XPath method failed, trying index method (index: {target_link_idx})...", level="DEBUG")
                        current_links = self.driver.find_elements(By.XPATH, "//a[@data-testid='billingstatementtable-paymentdate-link']")
                        if target_link_idx < len(current_links):
                            clickable_link = current_links[target_link_idx]
                            # Verify it's the right link by text
                            if clickable_link.text.strip() == matched_text:
                                self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", clickable_link)
                                time.sleep(0.5)
                                try:
                                    clickable_link.click()
                                except Exception:
                                    self.driver.execute_script("arguments[0].click();", clickable_link)
                                
                                time.sleep(2)  # Wait for DOS popup to open
                                
                                # Extract original modifier after clicking
                                self._extract_original_modifier()
                                return True
                    except Exception as e2:
                        self.gui_log(f"Index method also failed: {e2}", level="DEBUG")
                
                # If all else fails, log error and return False
                self.gui_log(f"❌ Error clicking date of service link: {e1}", level="ERROR")
                return False
            
        except Exception as e:
            self.gui_log(f"❌ Unexpected error clicking date of service link: {e}", level="ERROR")
            return False

    def _extract_original_modifier(self) -> None:
        if not self.wait:
            return
        try:
            modifier_input = self.wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[@data-testid='modifierseditor-code-input-1']"))
            )
            value = (modifier_input.get_attribute("value") or "").strip().upper()
            self.current_original_modifier = value if value else "N/A"
            self.gui_log(f"Original modifier detected: {self.current_original_modifier}", level="DEBUG")
        except TimeoutException:
            self.current_original_modifier = "N/A"

    def _click_notes_tab(self) -> bool:
        if not self.wait:
            return False
        try:
            notes_tab = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//a[@data-testid='calendarentryviewer-notes-tab']"))
            )
            notes_tab.click()
            time.sleep(1)
            return True
        except TimeoutException:
            return False

    def _click_progress_note_button(self) -> bool:
        if not self.wait:
            return False
        note_selectors = [
            "//span[text()='Progress Note']",
            "//span[contains(text(), 'Progress Note')]",
            "//span[text()='Consultation Note']",
            "//span[contains(text(), 'Consultation Note')]",
            "//span[text()='Intake Note']",
            "//span[contains(text(), 'Intake Note')]",
        ]
        note_button = None
        note_type = ""
        for selector in note_selectors:
            try:
                note_button = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, selector))
                )
                note_type = note_button.text.strip()
                if note_button:
                    break
            except TimeoutException:
                continue

        if not note_button:
            return False

        try:
            note_button.click()
            time.sleep(2)
        except Exception:
            return False

        session_medium = self._analyze_session_medium()
        self.current_session_medium = session_medium
        return True

    def _analyze_session_medium(self) -> Optional[str]:
        """Analyze the Progress/Consultation/Intake Note page to determine session medium (phone vs video)
        
        Extracts all text from the page and searches for keywords indicating:
        - Phone: phone, telephone, telephonic, phonic, audio, cell, cellphone
        - Video: video, zoom, videoconference, televideo, etc.
        
        Works with Progress Note, Consultation Note, and Intake Note pages.
        
        Returns:
            str: "phone", "video", or None if unable to determine
        """
        try:
            self.gui_log("Analyzing note page (Progress/Consultation/Intake) to determine session medium...", level="INFO")
            self.update_status("Analyzing session medium...", "#ff9500")
            
            # Wait for page to fully load
            time.sleep(2)
            
            # Extract all visible text from the page
            # Try to get text from the body element
            try:
                page_text = self.driver.find_element(By.TAG_NAME, "body").text
                self.gui_log(f"Extracted {len(page_text)} characters of text from note page", level="DEBUG")
            except Exception as e:
                self.gui_log(f"Could not extract text from body element: {e}", level="WARNING")
                # Try alternative method - get all text from main content area
                try:
                    # Look for common content containers
                    page_text = ""
                    content_selectors = [
                        "main", "article", "div[role='main']", 
                        ".content", "#content", ".main-content",
                        "div.progress-note", "div.note-content"
                    ]
                    for selector in content_selectors:
                        try:
                            elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                            if elements:
                                page_text = "\n".join([elem.text for elem in elements if elem.text])
                                if page_text:
                                    break
                        except:
                            continue
                    
                    if not page_text:
                        # Last resort: get all text from the page source
                        page_source = self.driver.page_source
                        # This won't be ideal but better than nothing
                        from html.parser import HTMLParser
                        # For now, just log that we're using page source
                        self.gui_log("Using page source as fallback for text extraction", level="WARNING")
                        page_text = self.driver.find_element(By.XPATH, "//body").text if self.driver.find_elements(By.XPATH, "//body") else ""
                        
                except Exception as e2:
                    self.log_error("Could not extract text from note page (Progress/Consultation/Intake Note)", exception=e2, include_traceback=False)
                    return None
            
            if not page_text or len(page_text.strip()) < 10:
                self.gui_log("⚠️ No meaningful text found on note page", level="WARNING")
                return None
            
            # Normalize text to lowercase for case-insensitive matching
            text_lower = page_text.lower()
            self.gui_log(f"Analyzing {len(text_lower)} characters of text for session medium keywords...", level="DEBUG")
            
            # Define keyword lists
            phone_keywords = [
                "phone", "telephone", "telephonic", "phonic", "audio", 
                "cell", "cellphone", "phone call", "telephone call",
                "telephonic session", "audio session", "via phone",
                "by phone", "on the phone", "over the phone"
            ]
            
            video_keywords = [
                "video", "zoom", "videoconference", "video conference",
                "video call", "video session", "via video", "by video",
                "over video", "zoom session", "televideo", "telehealth video"
            ]
            
            # Search for phone keywords
            phone_matches = []
            for keyword in phone_keywords:
                if keyword.lower() in text_lower:
                    phone_matches.append(keyword)
                    self.gui_log(f"  Found phone keyword: '{keyword}'", level="DEBUG")
            
            # Search for video keywords
            video_matches = []
            for keyword in video_keywords:
                if keyword.lower() in text_lower:
                    video_matches.append(keyword)
                    self.gui_log(f"  Found video keyword: '{keyword}'", level="DEBUG")
            
            # Determine categorization
            has_phone_keywords = len(phone_matches) > 0
            has_video_keywords = len(video_matches) > 0
            
            if has_video_keywords and not has_phone_keywords:
                self.gui_log(f"✅ Session categorized as VIDEO (found keywords: {', '.join(video_matches)})", level="INFO")
                return "video"
            elif has_phone_keywords and not has_video_keywords:
                self.gui_log(f"✅ Session categorized as PHONE (found keywords: {', '.join(phone_matches)})", level="INFO")
                return "phone"
            elif has_video_keywords and has_phone_keywords:
                # Both found - need to determine which is more prominent
                # Check for context around keywords
                self.gui_log(f"⚠️ Both phone and video keywords found - analyzing context...", level="WARNING")
                self.gui_log(f"  Phone keywords: {', '.join(phone_matches)}", level="DEBUG")
                self.gui_log(f"  Video keywords: {', '.join(video_matches)}", level="DEBUG")
                
                # Look for phrases that indicate the primary medium
                # Check for "met over video" or "met via video" which strongly indicates video
                video_phrases = ["met over video", "met via video", "session over video", "via zoom", "zoom session"]
                phone_phrases = ["met over phone", "met via phone", "session over phone", "telephone session", "phone session"]
                
                has_video_phrase = any(phrase in text_lower for phrase in video_phrases)
                has_phone_phrase = any(phrase in text_lower for phrase in phone_phrases)
                
                if has_video_phrase and not has_phone_phrase:
                    self.gui_log(f"✅ Session categorized as VIDEO (found video phrase)", level="INFO")
                    return "video"
                elif has_phone_phrase and not has_video_phrase:
                    self.gui_log(f"✅ Session categorized as PHONE (found phone phrase)", level="INFO")
                    return "phone"
                else:
                    # If still ambiguous, check count of keyword occurrences
                    video_count = sum(text_lower.count(kw) for kw in video_keywords)
                    phone_count = sum(text_lower.count(kw) for kw in phone_keywords)
                    
                    if video_count > phone_count:
                        self.gui_log(f"✅ Session categorized as VIDEO (video keywords appear more frequently)", level="INFO")
                        return "video"
                    elif phone_count > video_count:
                        self.gui_log(f"✅ Session categorized as PHONE (phone keywords appear more frequently)", level="INFO")
                        return "phone"
                    else:
                        # Default to video if equally ambiguous (video is more common in telehealth)
                        self.gui_log(f"⚠️ Ambiguous - defaulting to VIDEO (video/phone keywords equally present)", level="WARNING")
                        return "video"
            else:
                self.gui_log(f"⚠️ No session medium keywords found in Progress Note", level="WARNING")
                return None
                
        except Exception as e:
            self.log_error("Error analyzing session medium", exception=e, include_traceback=True)
            return None

    def _ensure_on_patients_page(self) -> bool:
        """Ensure we're on the Patients page before processing next client."""
        if not self.driver or not self.wait:
            return False
        
        try:
            # Check if we're already on the patients page
            current_url = self.driver.current_url.lower()
            if "patients" in current_url:
                try:
                    # Verify search field is present
                    self.driver.find_element(By.ID, "ctl00_BodyContent_TextBoxSearchPatientName")
                    return True
                except Exception:
                    pass
            
            # Navigate to patients page
            return self._navigate_to_patients()
        except Exception as e:
            self.gui_log(f"Error ensuring on patients page: {e}", level="DEBUG")
            return False
    
    def _navigate_back_to_patients(self) -> None:
        """Navigate back to Patients page with multiple fallback strategies."""
        if not self.driver or not self.wait:
            return
        
        self.gui_log("Navigating back to Patients page...", level="DEBUG")
        
        # Strategy 1: Close any open popups/modals first
        try:
            # Try to close popup by pressing Escape or finding close button
            from selenium.webdriver.common.keys import Keys
            try:
                close_buttons = self.driver.find_elements(By.XPATH, "//button[contains(@class, 'close') or contains(@aria-label, 'Close')]")
                if close_buttons:
                    close_buttons[0].click()
                    time.sleep(0.5)
            except Exception:
                try:
                    # Try pressing Escape
                    body = self.driver.find_element(By.TAG_NAME, "body")
                    body.send_keys(Keys.ESCAPE)
                    time.sleep(0.5)
                except Exception:
                    pass
        except Exception:
            pass
        
        # Strategy 2: Try clicking Patients link
        try:
            patients_link = self.wait.until(
                EC.element_to_be_clickable((By.LINK_TEXT, "Patients"))
            )
            patients_link.click()
            time.sleep(1)
            # Verify we're on patients page
            self.wait.until(
                EC.presence_of_element_located((By.ID, "ctl00_BodyContent_TextBoxSearchPatientName"))
            )
            self.gui_log("✅ Navigated back to Patients page (via link)", level="DEBUG")
            return
        except Exception:
            pass
        
        # Strategy 3: Try browser back button
        try:
            self.driver.back()
            time.sleep(1)
            # Check if we're on patients page now
            current_url = self.driver.current_url.lower()
            if "patients" in current_url:
                try:
                    self.wait.until(
                        EC.presence_of_element_located((By.ID, "ctl00_BodyContent_TextBoxSearchPatientName"))
                    )
                    self.gui_log("✅ Navigated back to Patients page (via back button)", level="DEBUG")
                    return
                except Exception:
                    pass
        except Exception:
            pass
        
        # Strategy 4: Force navigation to patients page URL
        try:
            self.driver.get("https://www.therapynotes.com/app/patients/")
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, "ctl00_BodyContent_TextBoxSearchPatientName"))
            )
            self.gui_log("✅ Navigated back to Patients page (via URL)", level="DEBUG")
            time.sleep(1)  # Additional wait for page to stabilize
        except Exception as e:
            self.gui_log(f"⚠️ Warning: Could not verify navigation to Patients page: {e}", level="WARNING")

    # ------------------------------------------------------------------
    # Window management
    # ------------------------------------------------------------------
    def _open_logs_folder(self) -> None:
        try:
            if os.name == "nt":
                os.startfile(str(SCRIPT_DIR))
            else:
                subprocess.Popen(["open", str(SCRIPT_DIR)])
        except Exception as exc:
            messagebox.showerror("Error", f"Unable to open logs folder:\n{exc}")

    # ------------------------------------------------------------------
    # PDF Field Configuration Management
    # ------------------------------------------------------------------
    def _load_pdf_field_mapping_config(self) -> None:
        """Load PDF field mapping configuration from JSON file."""
        try:
            if self.pdf_config_file.exists():
                with open(self.pdf_config_file, 'r', encoding='utf-8') as f:
                    self.pdf_field_mapping_config_raw = json.load(f)
                
                # Convert raw config to working format
                self._convert_config_to_mapping()
                self.gui_log(f"Loaded PDF field mapping configuration from {self.pdf_config_file.name}", level="INFO")
            else:
                # Create default configuration
                self._create_default_field_mapping_config()
                self._save_pdf_field_mapping_config()
                self.gui_log("Created default PDF field mapping configuration.", level="INFO")
        except Exception as e:
            self.log_error("Error loading PDF field mapping configuration", exception=e)
            # Use default configuration
            self._create_default_field_mapping_config()
    
    def _create_default_field_mapping_config(self) -> None:
        """Create default PDF field mapping configuration."""
        self.pdf_field_mapping_config_raw = {
            "Beneficiary's Name": {
                "type": "excel",
                "value": "client_name",
                "enabled": True
            },
            "Date of Birth": {
                "type": "excel",
                "value": "dob",
                "enabled": True
            },
            "MBI": {
                "type": "excel",
                "value": "patient_member_id",
                "enabled": True
            },
            "(MBI)": {
                "type": "excel",
                "value": "patient_member_id",
                "enabled": True
            },
            "Medicare Beneficiary Identifier": {
                "type": "excel",
                "value": "patient_member_id",
                "enabled": True
            },
            "Medicare ID": {
                "type": "excel",
                "value": "patient_member_id",
                "enabled": True
            },
            "Date(s) of Service": {
                "type": "excel",
                "value": "dos",
                "enabled": True
            },
            "Date of service": {
                "type": "excel",
                "value": "dos",
                "enabled": True
            },
            "Modifier": {
                "type": "excel",
                "value": "expected_modifier",
                "enabled": True
            },
            "Explain the needed correction below": {
                "type": "function",
                "value": "auto_explanation",
                "enabled": True
            },
            # Provider fields - disabled by default (can be enabled and configured)
            "Provider Name": {
                "type": "static",
                "value": "",
                "enabled": False
            },
            "Provider Address": {
                "type": "static",
                "value": "",
                "enabled": False
            },
            "NPI": {
                "type": "static",
                "value": "",
                "enabled": False
            },
            "PTAN": {
                "type": "static",
                "value": "",
                "enabled": False
            },
            "Tax ID": {
                "type": "static",
                "value": "",
                "enabled": False
            },
            "Internal Control Number": {
                "type": "static",
                "value": "",
                "enabled": False
            },
            "Rendering practitioner NPI": {
                "type": "static",
                "value": "",
                "enabled": False
            },
        }
        self._convert_config_to_mapping()
    
    def _convert_config_to_mapping(self) -> None:
        """Convert raw config format to working mapping format."""
        self.pdf_field_mapping_config = {}
        
        for field_name, config in self.pdf_field_mapping_config_raw.items():
            if not config.get("enabled", False):
                continue
            
            config_type = config.get("type", "static")
            value = config.get("value", "")
            field_type = config.get("field_type", "text")  # Get field type (text, checkbox, radio)
            
            # Store config with field_type for later use
            if config_type == "excel":
                # Map to Excel column (client_data key), but store as dict with field_type
                self.pdf_field_mapping_config[field_name] = {
                    "type": "excel",
                    "value": value,
                    "field_type": field_type
                }
            elif config_type == "static":
                # Map to static value
                if value:  # Only add if value is not empty
                    self.pdf_field_mapping_config[field_name] = {
                        "type": "static",
                        "value": value,
                        "field_type": field_type
                    }
            elif config_type == "function":
                # Map to function
                if value == "auto_explanation":
                    self.pdf_field_mapping_config[field_name] = {
                        "type": "function",
                        "value": lambda client_data: (
                            f"Modifier correction needed: Original modifier {client_data.get('original_modifier', 'N/A')} "
                            f"should be {client_data.get('expected_modifier', 'N/A')} based on session medium "
                            f"({client_data.get('session_medium', 'Unknown')})."
                        ),
                        "field_type": field_type
                    }
    
    def _save_pdf_field_mapping_config(self) -> bool:
        """Save PDF field mapping configuration to JSON file.
        
        Returns:
            bool: True if saved successfully, False otherwise
        """
        try:
            with open(self.pdf_config_file, 'w', encoding='utf-8') as f:
                json.dump(self.pdf_field_mapping_config_raw, f, indent=2, ensure_ascii=False)
            self.gui_log(f"Saved PDF field mapping configuration to {self.pdf_config_file.name}", level="INFO")
            return True
        except Exception as e:
            self.log_error("Error saving PDF field mapping configuration", exception=e)
            return False
    
    def _update_pdf_config_status(self) -> None:
        """Update the PDF configuration status label."""
        if not hasattr(self, 'pdf_config_status_label'):
            return
        
        enabled_count = sum(1 for config in self.pdf_field_mapping_config_raw.values() if config.get("enabled", False))
        total_count = len(self.pdf_field_mapping_config_raw)
        
        if enabled_count == 0:
            status_text = "No fields configured. Click 'Visual Field Mapper' to set up field mappings."
        else:
            status_text = f"{enabled_count} field(s) configured. Click 'Visual Field Mapper' to edit."
        
        self.pdf_config_status_label.config(text=status_text)
    
    def _open_visual_field_mapper(self) -> None:
        """Open the visual PDF field mapper window."""
        if not self.pdf_template_path or not self.pdf_template_path.exists():
            messagebox.showwarning("No Template", "Please select a PDF template first.")
            return
        
        # Check for required libraries
        if not PYMUPDF_AVAILABLE:
            messagebox.showerror("Missing Library", 
                               "PyMuPDF (fitz) is required for visual PDF mapping.\n"
                               "Install with: pip install PyMuPDF")
            return
        
        if not PIL_AVAILABLE:
            messagebox.showerror("Missing Library", 
                               "PIL/Pillow is required for visual PDF mapping.\n"
                               "Install with: pip install Pillow")
            return
        
        # Create visual mapper window
        VisualFieldMapper(self.root, self.pdf_template_path, self.pdf_field_mapping_config_raw, 
                         self._on_field_config_changed, self.gui_log)
    
    def _on_field_config_changed(self, updated_config: Dict[str, Dict[str, Any]]) -> None:
        """Callback when field configuration is changed in visual mapper.
        
        Args:
            updated_config: Updated field mapping configuration
        """
        self.pdf_field_mapping_config_raw = updated_config
        self._convert_config_to_mapping()
        self._save_pdf_field_mapping_config()
        self._update_pdf_config_status()
        self.gui_log("Field configuration updated from visual mapper.", level="INFO")
    
    def _open_field_config_window(self) -> None:
        """Open the PDF field configuration window (text-based)."""
        if not self.pdf_template_path or not self.pdf_template_path.exists():
            messagebox.showwarning("No Template", "Please select a PDF template first.")
            return
        
        # Get PDF fields
        pdf_fields = self._list_pdf_form_fields()
        if not pdf_fields:
            messagebox.showwarning("No Fields", "No PDF form fields found. Please check your PDF template.")
            return
        
        # Create configuration window
        config_window = tk.Toplevel(self.root)
        config_window.title("PDF Field Configuration - Version 3.1.0, Last Updated 12/04/2025")
        config_window.geometry("900x700")
        config_window.transient(self.root)
        config_window.grab_set()
        config_window.configure(bg="#f0f0f0")
        
        # Main container
        main_frame = tk.Frame(config_window, bg="#f0f0f0")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Instructions
        instruction = (
            "Configure which PDF fields to fill and what values to use.\n"
            "Select a field, choose data source (Excel column or static value), and enter the value.\n"
            "Fields not configured will be left blank."
        )
        tk.Label(main_frame, text=instruction,
                 font=("Segoe UI", 9), bg="#f0f0f0", fg="#555555",
                 wraplength=850, justify="left").pack(pady=(0, 15), anchor="w")
        
        # Scrollable frame for field configuration
        canvas = tk.Canvas(main_frame, bg="#f0f0f0", highlightthickness=0, height=450)
        scrollbar = tk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#f0f0f0")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        def on_canvas_configure(event):
            canvas.itemconfig(canvas.create_window((0, 0), window=scrollable_frame, anchor="nw"), width=event.width)
        
        canvas.bind("<Configure>", on_canvas_configure)
        
        # Available Excel columns (client_data keys)
        # Note: patient_member_id is extracted from "Patient Member #" column in Excel
        # This is the same as MBI (Medicare Beneficiary Identifier) in the PDF form
        excel_columns = [
            "client_name", "first_name", "last_name", "dob", "dos",
            "patient_member_id",  # Patient Member # (MBI) - extracted from Excel column M or "Patient Member #"
            "session_medium", "original_modifier",
            "expected_modifier", "status", "notes"
        ]
        
        # Field configuration widgets
        field_widgets: Dict[str, Dict[str, tk.Widget]] = {}
        
        # Header row
        header_frame = tk.Frame(scrollable_frame, bg="#003366", relief="raised", bd=2)
        header_frame.pack(fill="x", pady=(0, 5))
        
        tk.Label(header_frame, text="Field Name", font=("Segoe UI", 10, "bold"),
                 bg="#003366", fg="white", width=25).grid(row=0, column=0, padx=5, pady=5, sticky="w")
        tk.Label(header_frame, text="Enable", font=("Segoe UI", 10, "bold"),
                 bg="#003366", fg="white", width=8).grid(row=0, column=1, padx=5, pady=5)
        tk.Label(header_frame, text="Data Source", font=("Segoe UI", 10, "bold"),
                 bg="#003366", fg="white", width=15).grid(row=0, column=2, padx=5, pady=5)
        tk.Label(header_frame, text="Value", font=("Segoe UI", 10, "bold"),
                 bg="#003366", fg="white", width=25).grid(row=0, column=3, padx=5, pady=5, sticky="ew")
        
        header_frame.grid_columnconfigure(3, weight=1)
        
        # Create configuration rows for each PDF field
        for field_name in sorted(pdf_fields):
            # Get or create config for this field
            # Check both exact match and cleaned name match
            clean_name = field_name.strip('()[]').replace('\\', '').strip()
            field_config = None
            
            if field_name in self.pdf_field_mapping_config_raw:
                field_config = self.pdf_field_mapping_config_raw[field_name].copy()
            elif clean_name in self.pdf_field_mapping_config_raw:
                # Use config from cleaned name
                field_config = self.pdf_field_mapping_config_raw[clean_name].copy()
                # Also store it under the actual field name for future use
                self.pdf_field_mapping_config_raw[field_name] = field_config.copy()
            else:
                # Create new config (disabled by default)
                field_config = {
                    "type": "static",
                    "value": "",
                    "enabled": False
                }
                self.pdf_field_mapping_config_raw[field_name] = field_config
            
            # Create row frame
            row_frame = tk.Frame(scrollable_frame, bg="#ffffff", relief="raised", bd=1)
            row_frame.pack(fill="x", pady=2, padx=2)
            row_frame.grid_columnconfigure(3, weight=1)
            
            # Field name label
            field_label = tk.Label(row_frame, text=field_name, font=("Segoe UI", 9),
                                   bg="#ffffff", fg="#000000", width=30, anchor="w", wraplength=250)
            field_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
            
            # Enable checkbox
            enable_var = tk.BooleanVar(value=field_config.get("enabled", False))
            enable_check = tk.Checkbutton(row_frame, variable=enable_var, bg="#ffffff")
            enable_check.grid(row=0, column=1, padx=5, pady=5)
            
            # Data source dropdown (Excel column or Static value)
            source_var = tk.StringVar(value=field_config.get("type", "static"))
            source_dropdown = ttk.Combobox(row_frame, textvariable=source_var,
                                          values=["excel", "static"], state="readonly", width=12)
            source_dropdown.grid(row=0, column=2, padx=5, pady=5)
            
            # Value entry/dropdown
            value_var = tk.StringVar(value=field_config.get("value", ""))
            
            # Determine if value should be dropdown (for Excel) or entry (for static)
            value_widget = None
            if source_var.get() == "excel":
                value_dropdown = ttk.Combobox(row_frame, textvariable=value_var,
                                             values=excel_columns, width=25, state="readonly")
                value_dropdown.grid(row=0, column=3, padx=5, pady=5, sticky="ew")
                value_widget = value_dropdown
            else:
                value_entry = tk.Entry(row_frame, textvariable=value_var, width=30)
                value_entry.grid(row=0, column=3, padx=5, pady=5, sticky="ew")
                value_widget = value_entry
            
            # Create handler for source change using proper closure
            def create_source_handler(field: str, row: tk.Frame, col: int, widgets_ref: Dict):
                def handler(event=None):
                    widgets = widgets_ref
                    if field not in widgets:
                        return
                    
                    current_value = widgets[field]["value_var"].get()
                    source = widgets[field]["source_var"].get()
                    
                    # Remove old widget
                    old_widget = widgets[field]["value_widget"]
                    old_widget.destroy()
                    
                    # Create new widget based on source
                    if source == "excel":
                        new_widget = ttk.Combobox(row, textvariable=widgets[field]["value_var"],
                                                 values=excel_columns, width=25, state="readonly")
                        new_widget.grid(row=0, column=col, padx=5, pady=5, sticky="ew")
                        # Set value if it's a valid Excel column
                        if current_value in excel_columns:
                            widgets[field]["value_var"].set(current_value)
                        else:
                            widgets[field]["value_var"].set("")
                    else:
                        new_widget = tk.Entry(row, textvariable=widgets[field]["value_var"], width=30)
                        new_widget.grid(row=0, column=col, padx=5, pady=5, sticky="ew")
                        # Preserve current value for static fields
                        if current_value:
                            widgets[field]["value_var"].set(current_value)
                    
                    # Update reference
                    widgets[field]["value_widget"] = new_widget
                
                return handler
            
            # Store widget references first
            field_widgets[field_name] = {
                "enable_var": enable_var,
                "source_var": source_var,
                "value_var": value_var,
                "value_widget": value_widget,
            }
            
            # Now bind the handler
            source_handler = create_source_handler(field_name, row_frame, 3, field_widgets)
            source_dropdown.bind("<<ComboboxSelected>>", source_handler)
        
        # Buttons frame
        buttons_frame = tk.Frame(main_frame, bg="#f0f0f0")
        buttons_frame.pack(fill="x", pady=(15, 0))
        
        def save_configuration():
            try:
                # Update configuration from widgets
                for field_name, widgets in field_widgets.items():
                    enabled = widgets["enable_var"].get()
                    source_type = widgets["source_var"].get()
                    value = widgets["value_var"].get().strip()
                    
                    # Validate configuration
                    if enabled:
                        if source_type == "excel" and not value:
                            messagebox.showerror("Validation Error", 
                                               f"Field '{field_name}' is enabled but no Excel column selected.\n"
                                               "Please select an Excel column or disable the field.")
                            return
                        elif source_type == "static" and not value:
                            # Static fields can be empty if disabled, but if enabled should have a value
                            if messagebox.askyesno("Empty Static Value",
                                                 f"Field '{field_name}' is enabled but has no static value.\n"
                                                 "Disable this field instead?"):
                                                enabled = False
                    
                    # Update or create config
                    if field_name not in self.pdf_field_mapping_config_raw:
                        self.pdf_field_mapping_config_raw[field_name] = {}
                    
                    self.pdf_field_mapping_config_raw[field_name]["enabled"] = enabled
                    self.pdf_field_mapping_config_raw[field_name]["type"] = source_type
                    self.pdf_field_mapping_config_raw[field_name]["value"] = value
                
                # Convert to working format
                self._convert_config_to_mapping()
                
                # Save to file
                if self._save_pdf_field_mapping_config():
                    # Update status label
                    self._update_pdf_config_status()
                    messagebox.showinfo("Success", "PDF field configuration saved successfully!")
                    config_window.destroy()
                else:
                    messagebox.showerror("Error", "Failed to save configuration. Check the log for details.")
            except Exception as e:
                self.log_error("Error saving configuration", exception=e, include_traceback=True)
                messagebox.showerror("Error", f"Error saving configuration:\n{e}")
        
        def cancel_configuration():
            if messagebox.askyesno("Cancel", "Discard changes and close configuration window?"):
                config_window.destroy()
        
        # Add help label
        help_label = tk.Label(buttons_frame,
                             text="💡 Tip: Enable fields and select 'excel' to fill from Excel data, or 'static' to fill with a fixed value.",
                             font=("Segoe UI", 8, "italic"),
                             bg="#f0f0f0", fg="#666666",
                             wraplength=600, justify="left")
        help_label.pack(side="top", pady=(0, 10), anchor="w")
        
        tk.Button(buttons_frame, text="Save Configuration",
                  command=save_configuration,
                  bg="#28a745", fg="white", font=("Segoe UI", 10, "bold"),
                  padx=20, pady=8, cursor="hand2", relief="flat").pack(side="left", padx=(0, 10))
        
        tk.Button(buttons_frame, text="Cancel",
                  command=cancel_configuration,
                  bg="#6c757d", fg="white", font=("Segoe UI", 10),
                  padx=20, pady=8, cursor="hand2", relief="flat").pack(side="left")
    
    def _load_default_field_mappings(self) -> None:
        """Load default field mappings."""
        if messagebox.askyesno("Load Defaults", 
                               "This will replace your current field configuration with defaults.\n"
                               "Continue?"):
            self._create_default_field_mapping_config()
            self._save_pdf_field_mapping_config()
            self._update_pdf_config_status()
            self.gui_log("Loaded default field mappings.", level="INFO")
            messagebox.showinfo("Success", "Default field mappings loaded successfully!")
    
    def _clear_field_mappings(self) -> None:
        """Clear all field mappings."""
        if messagebox.askyesno("Clear All", 
                               "This will clear all field mappings.\n"
                               "All fields will be left blank.\n"
                               "Continue?"):
            self.pdf_field_mapping_config_raw = {}
            self.pdf_field_mapping_config = {}
            self._save_pdf_field_mapping_config()
            self._update_pdf_config_status()
            self.gui_log("Cleared all field mappings.", level="INFO")
            messagebox.showinfo("Success", "All field mappings cleared.")
    
    def _close_window(self) -> None:
        if self._shutdown_requested:
            return
        self._shutdown_requested = True
        self.gui_log("Shutdown requested. Waiting for background tasks...")
        self.start_button.config(state="disabled")
        self.stop_button.config(state="disabled")
        self.request_stop()
        self.root.after(200, self._wait_for_threads_and_close)

    def _wait_for_threads_and_close(self) -> None:
        with self._thread_lock:
            active = [t for t in self._active_threads if t.is_alive()]
        if not active:
            if self.root:
                self.root.quit()
                self.root.destroy()
            self._close_driver()
            self.gui_log("Bot shutdown complete.")
            return
        self.root.after(200, self._wait_for_threads_and_close)


# -----------------------------------------------------------------------------
# Visual Field Mapper
# -----------------------------------------------------------------------------
class VisualFieldMapper:
    """Visual PDF field mapper that displays PDF and allows clicking fields to configure them."""
    
    def __init__(self, parent: tk.Tk, pdf_path: Path, 
                 field_config: Dict[str, Dict[str, Any]], 
                 on_config_changed: callable,
                 log_func: callable):
        """Initialize visual field mapper.
        
        Args:
            parent: Parent window
            pdf_path: Path to PDF template
            field_config: Current field configuration
            on_config_changed: Callback when configuration changes
            log_func: Logging function
        """
        self.parent = parent
        self.pdf_path = pdf_path
        self.field_config = field_config.copy()
        self.on_config_changed = on_config_changed
        self.log_func = log_func
        
        # Field data: {field_name: {"rect": (x0, y0, x1, y1), "page": int, ...}}
        self.field_data: Dict[str, Dict[str, Any]] = {}
        self.pdf_images: List[Image.Image] = []
        self.pdf_image_tk: List[ImageTk.PhotoImage] = []
        self.pdf_doc = None  # Keep PDF document open for coordinate calculations
        self.pdf_pages: List[Any] = []  # Store PDF pages
        self.pdf_page_dimensions: List[Tuple[float, float]] = []  # Store (width, height) for each page
        self.zoom_factor = 1.0
        self.current_page = 0
        self._waiting_for_manual_position = None  # Field name waiting for manual positioning
        self._waiting_field_type = "text"  # Field type waiting for manual positioning
        self._waiting_for_delete = False  # Waiting for field to delete
        self._selected_field = None  # Currently selected field
        self._resizing_field = None  # Field being resized
        self._resize_handle = None  # Which resize handle (nw, ne, sw, se, n, s, e, w)
        self._resize_start_pos = None  # Starting position for resize
        self._resize_start_rect = None  # Starting rectangle for resize
        self._resize_start_pdf_rect = None  # Starting PDF rectangle for resize
        self._dragging_field = None  # Field being dragged
        self._drag_start_pos = None  # Starting position for drag
        self._drag_offset = None  # Offset from click point to field corner
        self._drag_start_pdf_rect = None  # Starting PDF rectangle for drag
        self._select_mode = False  # Selection mode for dragging/resizing (vs edit mode)
        
        # Available Excel columns
        self.excel_columns = [
            "client_name", "first_name", "last_name", "dob", "dos",
            "patient_member_id", "session_medium", "original_modifier",
            "expected_modifier", "status", "notes"
        ]
        
        # Create window
        self.window = tk.Toplevel(parent)
        self.window.title("Visual PDF Field Mapper - Version 3.1.0, Last Updated 12/04/2025")
        self.window.geometry("1200x800")
        self.window.transient(parent)
        self.window.configure(bg="#f0f0f0")
        
        # Build UI
        self._build_ui()
        
        # Load PDF and extract fields
        self._load_pdf()
        
        # Display first page
        if self.pdf_images:
            self._display_page(0)
    
    def _build_ui(self) -> None:
        """Build the UI for the visual mapper."""
        # Top toolbar
        toolbar = tk.Frame(self.window, bg="#003366", height=60)
        toolbar.pack(fill="x", padx=0, pady=0)
        
        # Instructions
        instruction = tk.Label(toolbar, 
                              text="Click on a PDF form field to configure it. Configured fields are highlighted in green.",
                              font=("Segoe UI", 10), bg="#003366", fg="white")
        instruction.pack(side="left", padx=15, pady=10)
        
        # Page navigation
        nav_frame = tk.Frame(toolbar, bg="#003366")
        nav_frame.pack(side="right", padx=15, pady=10)
        
        self.prev_btn = tk.Button(nav_frame, text="◄ Prev", command=self._prev_page,
                                  bg="#006644", fg="white", font=("Segoe UI", 9),
                                  padx=10, pady=5, cursor="hand2", relief="flat")
        self.prev_btn.pack(side="left", padx=5)
        
        self.page_label = tk.Label(nav_frame, text="Page 1 / 1",
                                   font=("Segoe UI", 10), bg="#003366", fg="white")
        self.page_label.pack(side="left", padx=10)
        
        self.next_btn = tk.Button(nav_frame, text="Next ►", command=self._next_page,
                                  bg="#006644", fg="white", font=("Segoe UI", 9),
                                  padx=10, pady=5, cursor="hand2", relief="flat")
        self.next_btn.pack(side="left", padx=5)
        
        # Zoom controls
        zoom_frame = tk.Frame(toolbar, bg="#003366")
        zoom_frame.pack(side="right", padx=15, pady=10)
        
        tk.Button(zoom_frame, text="Zoom -", command=lambda: self._set_zoom(0.8),
                 bg="#6c757d", fg="white", font=("Segoe UI", 9),
                 padx=8, pady=5, cursor="hand2", relief="flat").pack(side="left", padx=2)
        
        tk.Button(zoom_frame, text="Zoom +", command=lambda: self._set_zoom(1.25),
                 bg="#6c757d", fg="white", font=("Segoe UI", 9),
                 padx=8, pady=5, cursor="hand2", relief="flat").pack(side="left", padx=2)
        
        # Main container with scrollable canvas
        main_container = tk.Frame(self.window, bg="#f0f0f0")
        main_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Canvas for PDF display with scrollbars
        canvas_frame = tk.Frame(main_container, bg="#ffffff", relief="sunken", bd=2)
        canvas_frame.pack(fill="both", expand=True)
        
        self.canvas = tk.Canvas(canvas_frame, bg="#ffffff", highlightthickness=0)
        v_scrollbar = tk.Scrollbar(canvas_frame, orient="vertical", command=self.canvas.yview)
        h_scrollbar = tk.Scrollbar(canvas_frame, orient="horizontal", command=self.canvas.xview)
        
        self.canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Bind events for interaction
        self.canvas.bind("<Button-1>", self._on_canvas_click)
        self.canvas.bind("<B1-Motion>", self._on_canvas_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_canvas_release)
        self.canvas.bind("<Motion>", self._on_canvas_motion)
        self.canvas.bind("<Double-Button-1>", self._on_canvas_double_click)
        self.canvas.bind("<Button-3>", self._on_canvas_right_click)  # Right-click for context menu
        
        # Keyboard shortcuts
        self.window.bind("<Delete>", lambda e: self._delete_selected_field())
        self.window.bind("<Escape>", lambda e: self._clear_selection())
        self.window.focus_set()
        
        # Bottom buttons
        bottom_frame = tk.Frame(self.window, bg="#f0f0f0")
        bottom_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Button(bottom_frame, text="Save Selectors",
                 command=self._save_selectors,
                 bg="#28a745", fg="white", font=("Segoe UI", 10, "bold"),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="left", padx=5)
        
        tk.Button(bottom_frame, text="Add Field Manually",
                 command=self._add_field_manually,
                 bg="#007bff", fg="white", font=("Segoe UI", 10),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="left", padx=5)
        
        # Select Mode toggle button
        self.select_mode_btn = tk.Button(bottom_frame, text="Select Mode (Off)",
                 command=self._toggle_select_mode,
                 bg="#6c757d", fg="white", font=("Segoe UI", 10),
                 padx=20, pady=8, cursor="hand2", relief="flat")
        self.select_mode_btn.pack(side="left", padx=5)
        
        tk.Button(bottom_frame, text="Delete Selector",
                 command=self._delete_field_selector,
                 bg="#ff6b35", fg="white", font=("Segoe UI", 10),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="left", padx=5)
        
        tk.Button(bottom_frame, text="Delete All Selectors",
                 command=self._delete_all_selectors,
                 bg="#c82333", fg="white", font=("Segoe UI", 10, "bold"),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="left", padx=5)
        
        tk.Button(bottom_frame, text="Clear All",
                 command=self._clear_all_fields,
                 bg="#dc3545", fg="white", font=("Segoe UI", 10),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="left", padx=5)
        
        tk.Button(bottom_frame, text="Close",
                 command=self._close_window,
                 bg="#6c757d", fg="white", font=("Segoe UI", 10),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="right", padx=5)
        
        # Status label
        self.status_label = tk.Label(bottom_frame,
                                     text="Ready. Click on a field to configure it.",
                                     font=("Segoe UI", 9, "italic"),
                                     bg="#f0f0f0", fg="#666666")
        self.status_label.pack(side="left", padx=20)
    
    def _load_pdf(self) -> None:
        """Load PDF and extract field information using comprehensive methods."""
        try:
            if not PYMUPDF_AVAILABLE:
                messagebox.showerror("Error", "PyMuPDF (fitz) is required but not available.")
                return
            
            # Open PDF (keep it open for coordinate calculations)
            self.pdf_doc = fitz.open(str(self.pdf_path))  # type: ignore
            
            # First, extract all field names using comprehensive method (pdfrw)
            all_field_names = set()
            field_coords_map = {}  # {field_name: {page: int, rect: [x0, y0, x1, y1], field_type: str}}
            
            # Use pdfrw to extract ALL fields with coordinates
            if PDFRW_AVAILABLE:
                try:
                    template_pdf = pdfrw.PdfReader(str(self.pdf_path))  # type: ignore
                    
                    # Extract fields from AcroForm
                    if template_pdf.Root and template_pdf.Root.AcroForm:
                        acro_form = template_pdf.Root.AcroForm
                        if acro_form.get('/Fields'):
                            fields_array = acro_form['/Fields']
                            # Recursively extract field names
                            field_names_list = []
                            self._extract_field_names_from_array(fields_array, field_names_list)
                            all_field_names.update(field_names_list)
                    
                    # Extract field coordinates from page annotations
                    for page_num, page in enumerate(template_pdf.pages):
                        annotations = page.get('/Annots')
                        if not annotations:
                            continue
                        
                        for annotation in annotations:
                            if annotation is None:
                                continue
                            
                            if isinstance(annotation, pdfrw.PdfDict):  # type: ignore
                                annotation_obj = annotation
                            else:
                                annotation_obj = pdfrw.PdfDict(annotation)  # type: ignore
                            
                            # Check if this is a Widget (form field)
                            subtype = annotation_obj.get('/Subtype')
                            is_widget = False
                            if subtype:
                                subtype_str = str(subtype)
                                if '/Widget' in subtype_str or 'Widget' in subtype_str:
                                    is_widget = True
                            
                            # Also check for /FT (Field Type) or /Ff (Field Flags)
                            has_field_type = '/FT' in annotation_obj or '/Ff' in annotation_obj
                            
                            if is_widget or has_field_type:
                                # Get field name
                                field_name_obj = annotation_obj.get('/T')
                                if not field_name_obj:
                                    # Try parent
                                    if '/Parent' in annotation_obj:
                                        parent = annotation_obj['/Parent']
                                        if parent:
                                            field_name_obj = parent.get('/T')
                                
                                if field_name_obj:
                                    field_name = self._extract_field_name_from_obj(field_name_obj)
                                    if field_name:
                                        all_field_names.add(field_name)
                                        
                                        # Get field rectangle from /Rect
                                        rect_obj = annotation_obj.get('/Rect')
                                        if rect_obj:
                                            # Rect is [x0, y0, x1, y1] in PDF coordinates
                                            # PDF coordinates: (0,0) is bottom-left, y increases upward
                                            try:
                                                if hasattr(rect_obj, '__iter__') and not isinstance(rect_obj, (str, bytes)):
                                                    rect_list = list(rect_obj)
                                                    if len(rect_list) >= 4:
                                                        x0, y0, x1, y1 = float(rect_list[0]), float(rect_list[1]), float(rect_list[2]), float(rect_list[3])
                                                        
                                                        # Get field type
                                                        field_type = "text"
                                                        ft_obj = annotation_obj.get('/FT')
                                                        if ft_obj:
                                                            ft_str = str(ft_obj)
                                                            if '/Btn' in ft_str or 'Btn' in ft_str:
                                                                field_type = "checkbox"
                                                            elif '/Tx' in ft_str or 'Tx' in ft_str:
                                                                field_type = "text"
                                                        
                                                        # Store coordinates (PDF coordinate system: bottom-left origin)
                                                        if field_name not in field_coords_map:
                                                            field_coords_map[field_name] = {
                                                                "page": page_num,
                                                                "rect_pdf": [x0, y0, x1, y1],
                                                                "field_type": field_type
                                                            }
                                            except Exception as e:
                                                self.log_func(f"Error extracting rect for {field_name}: {e}", level="DEBUG")
                
                except Exception as e:
                    self.log_func(f"Error extracting fields with pdfrw: {e}", level="WARNING")
                    import traceback
                    self.log_func(f"Traceback: {traceback.format_exc()}", level="DEBUG")
            
            # Also try PyMuPDF widgets as a supplement (might catch some fields pdfrw misses)
            for page_num in range(len(self.pdf_doc)):
                page = self.pdf_doc[page_num]
                
                # Store page for later use
                self.pdf_pages.append(page)
                
                # Get page dimensions (in PDF points)
                page_rect = page.rect
                page_width = page_rect.width
                page_height = page_rect.height
                self.pdf_page_dimensions.append((page_width, page_height))
                
                # Render page to image (200 DPI for good quality)
                mat = fitz.Matrix(2.0, 2.0)  # type: ignore  # 2x zoom = ~200 DPI
                pix = page.get_pixmap(matrix=mat)
                
                # Convert to PIL Image
                img_data = pix.tobytes("ppm")
                from io import BytesIO
                img = Image.open(BytesIO(img_data))
                self.pdf_images.append(img)
                
                # Extract form fields from PyMuPDF widgets (supplement)
                widgets = page.widgets()
                for widget in widgets:
                    field_name = widget.field_name
                    if not field_name:
                        continue
                    
                    all_field_names.add(field_name)
                    
                    # If we don't have coordinates from pdfrw, use PyMuPDF coordinates
                    if field_name not in field_coords_map:
                        rect = widget.rect
                        field_coords_map[field_name] = {
                            "page": page_num,
                            "rect_pdf": [rect.x0, rect.y0, rect.x1, rect.y1],
                            "field_type": widget.field_type_string or "text"
                        }
            
            # Convert field_coords_map to self.field_data format (using fitz.Rect for compatibility)
            for field_name, coords_info in field_coords_map.items():
                x0, y0, x1, y1 = coords_info["rect_pdf"]
                # Create a fitz.Rect-like object
                class RectLike:
                    def __init__(self, x0, y0, x1, y1):
                        self.x0 = x0
                        self.y0 = y0
                        self.x1 = x1
                        self.y1 = y1
                
                self.field_data[field_name] = {
                    "rect": RectLike(x0, y0, x1, y1),
                    "page": coords_info["page"],
                    "field_type": coords_info.get("field_type", "text"),
                    "field_value": None
                }
            
            # Also add any field names we found but don't have coordinates for (user can manually position)
            for field_name in all_field_names:
                if field_name not in self.field_data:
                    # Create a dummy rect (will need manual positioning)
                    class RectLike:
                        def __init__(self):
                            self.x0 = 0
                            self.y0 = 0
                            self.x1 = 0
                            self.y1 = 0
                    
                    self.field_data[field_name] = {
                        "rect": RectLike(),
                        "page": 0,
                        "field_type": "unknown",
                        "field_value": None,
                        "needs_positioning": True
                    }
            
            self.log_func(f"Loaded PDF: {len(self.pdf_images)} page(s), {len(self.field_data)} field(s) detected", level="INFO")
            
        except Exception as e:
            self.log_func(f"Error loading PDF: {e}", level="ERROR")
            import traceback
            self.log_func(f"Traceback: {traceback.format_exc()}", level="DEBUG")
            messagebox.showerror("Error", f"Error loading PDF:\n{e}")
    
    def _extract_field_names_from_array(self, fields_array, field_names_list: List[str]) -> None:
        """Recursively extract field names from a PDF Fields array (helper for VisualFieldMapper)."""
        try:
            if not fields_array:
                return
            
            for field_obj in fields_array:
                if field_obj is None:
                    continue
                
                # Get field name
                field_name_obj = field_obj.get('/T')
                if field_name_obj:
                    field_name = self._extract_field_name_from_obj(field_name_obj)
                    if field_name and field_name not in field_names_list:
                        field_names_list.append(field_name)
                
                # Check for nested fields (Kids array)
                if '/Kids' in field_obj:
                    kids_array = field_obj['/Kids']
                    if kids_array:
                        self._extract_field_names_from_array(kids_array, field_names_list)
        except Exception as e:
            self.log_func(f"Error extracting field names: {e}", level="DEBUG")
    
    def _extract_field_name_from_obj(self, field_name_obj) -> str:
        """Extract field name from PDF object (helper for VisualFieldMapper)."""
        try:
            if field_name_obj is None:
                return ""
            
            # Handle different PDF object types
            field_name_str = ""
            if hasattr(field_name_obj, 'decode'):
                try:
                    field_name_str = field_name_obj.decode('utf-8')
                except Exception:
                    try:
                        field_name_str = field_name_obj.decode('latin-1')
                    except Exception:
                        field_name_str = str(field_name_obj)
            elif isinstance(field_name_obj, bytes):
                try:
                    field_name_str = field_name_obj.decode('utf-8', errors='ignore')
                except Exception:
                    try:
                        field_name_str = field_name_obj.decode('latin-1', errors='ignore')
                    except Exception:
                        field_name_str = str(field_name_obj)
            elif isinstance(field_name_obj, str):
                field_name_str = field_name_obj
            else:
                field_name_str = str(field_name_obj)
                field_name_str = field_name_str.strip('()[]"\'')
            
            # Clean up escaped characters
            field_name_str = field_name_str.replace('\\', '')
            
            return field_name_str
        except Exception:
            return str(field_name_obj).strip('()[]"\'')
    
    def __del__(self) -> None:
        """Cleanup: close PDF document when mapper is destroyed."""
        if hasattr(self, 'pdf_doc') and self.pdf_doc:
            try:
                self.pdf_doc.close()
            except Exception:
                pass
    
    def _display_page(self, page_num: int) -> None:
        """Display a specific page of the PDF."""
        if page_num < 0 or page_num >= len(self.pdf_images):
            return
        
        self.current_page = page_num
        self.canvas.delete("all")
        
        # Get original image for current page (before zoom)
        original_img = self.pdf_images[page_num]
        
        # Apply zoom to create displayed image
        if abs(self.zoom_factor - 1.0) > 0.01:  # Only resize if zoom changed significantly
            new_width = int(original_img.width * self.zoom_factor)
            new_height = int(original_img.height * self.zoom_factor)
            displayed_img = original_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
        else:
            displayed_img = original_img
        
        # Convert to PhotoImage
        img_tk = ImageTk.PhotoImage(displayed_img)
        self.pdf_image_tk = [img_tk]  # Keep reference to prevent garbage collection
        
        # Display image on canvas
        self.canvas.create_image(0, 0, anchor="nw", image=img_tk)
        self.canvas.config(scrollregion=self.canvas.bbox("all"))
        
        # Draw field rectangles for current page (after image is displayed)
        self._draw_field_rectangles(page_num)
        
        # Update page label
        self.page_label.config(text=f"Page {page_num + 1} / {len(self.pdf_images)}")
        
        # Update navigation buttons
        self.prev_btn.config(state="normal" if page_num > 0 else "disabled")
        self.next_btn.config(state="normal" if page_num < len(self.pdf_images) - 1 else "disabled")
    
    def _draw_field_rectangles(self, page_num: int) -> None:
        """Draw rectangles for fields on the current page."""
        if not self.pdf_images or page_num >= len(self.pdf_images):
            return
        
        # Get original image (before zoom) and displayed image (after zoom)
        original_img = self.pdf_images[page_num]
        displayed_img = self.pdf_image_tk[0] if self.pdf_image_tk else None
        if not displayed_img:
            return
        
        # Get PDF page dimensions (from cached data)
        if page_num < len(self.pdf_page_dimensions):
            pdf_page_width, pdf_page_height = self.pdf_page_dimensions[page_num]
        else:
            # Fallback: estimate from image dimensions
            # Image was rendered at 2x (200 DPI), so PDF dimensions are approximately img_dim / 2
            pdf_page_width = original_img.width / 2.0
            pdf_page_height = original_img.height / 2.0
        
        # Get displayed image dimensions (after zoom)
        displayed_width = displayed_img.width()
        displayed_height = displayed_img.height()
        
        # Get the original image dimensions (before zoom)
        original_width = original_img.width
        original_height = original_img.height
        
        # Calculate zoom scale: how much we've zoomed the displayed image
        zoom_scale_x = displayed_width / original_width if original_width > 0 else 1.0
        zoom_scale_y = displayed_height / original_height if original_height > 0 else 1.0
        
        # Calculate conversion factor: original image pixels per PDF point
        # Image was rendered at 2x (200 DPI) from PDF, so:
        # original_image_pixels = pdf_points * 2 (approximately)
        pdf_to_img_x = original_width / pdf_page_width if pdf_page_width > 0 else 2.0
        pdf_to_img_y = original_height / pdf_page_height if pdf_page_height > 0 else 2.0
        
        # Draw rectangles for fields on this page
        for field_name, field_info in self.field_data.items():
            if field_info["page"] != page_num:
                continue
            
            # Skip fields that need positioning (zero-size rectangles)
            if field_info.get("needs_positioning", False):
                continue
            
            rect = field_info["rect"]  # This is a RectLike or fitz.Rect object
            
            # Convert PDF coordinates to original image coordinates
            # PDF rect: (x0, y0, x1, y1) in PDF points, where (0,0) is bottom-left
            # We need to convert to image coordinates where (0,0) is top-left
            
            x0_pdf = rect.x0
            y0_pdf = rect.y0  # In PDF, y0 is bottom
            x1_pdf = rect.x1
            y1_pdf = rect.y1  # In PDF, y1 is top
            
            # Skip if rectangle is invalid (zero size)
            if abs(x1_pdf - x0_pdf) < 0.1 or abs(y1_pdf - y0_pdf) < 0.1:
                continue
            
            # Convert to image coordinates (flip Y axis)
            # PDF Y increases upward, Image Y increases downward
            x0_img = x0_pdf * pdf_to_img_x
            y0_img = (pdf_page_height - y1_pdf) * pdf_to_img_y  # Flip Y: top becomes y0
            x1_img = x1_pdf * pdf_to_img_x
            y1_img = (pdf_page_height - y0_pdf) * pdf_to_img_y  # Flip Y: bottom becomes y1
            
            # Apply zoom scaling
            x0_displayed = x0_img * zoom_scale_x
            y0_displayed = y0_img * zoom_scale_y
            x1_displayed = x1_img * zoom_scale_x
            y1_displayed = y1_img * zoom_scale_y
            
            # Check if field is configured
            is_configured = (field_name in self.field_config and 
                           self.field_config[field_name].get("enabled", False))
            
            # Check if field is selected (only show selection in select mode)
            is_selected = (field_name == self._selected_field and self._select_mode)
            
            # Draw rectangle (green if configured, blue if selected, yellow if not)
            if is_selected:
                color = "#4488ff"
                outline = "#0066cc"
                width = 3
            elif is_configured:
                color = "#00ff00"
                outline = "#00aa00"
                width = 2
            else:
                color = "#ffff00"
                outline = "#aaaa00"
                width = 1
            
            # Draw rectangle on canvas
            # Note: Tkinter canvas doesn't support alpha directly, so we use stipple
            rect_id = self.canvas.create_rectangle(x0_displayed, y0_displayed, x1_displayed, y1_displayed,
                                        outline=outline, fill=color, 
                                        stipple="gray25", width=width,
                                        tags=(f"field_{field_name}", "field_rect"))
            
            # Draw resize handles if field is selected
            if is_selected:
                handle_size = 8
                handles = [
                    (x0_displayed, y0_displayed, "nw"),  # Top-left
                    ((x0_displayed + x1_displayed) / 2, y0_displayed, "n"),  # Top
                    (x1_displayed, y0_displayed, "ne"),  # Top-right
                    (x1_displayed, (y0_displayed + y1_displayed) / 2, "e"),  # Right
                    (x1_displayed, y1_displayed, "se"),  # Bottom-right
                    ((x0_displayed + x1_displayed) / 2, y1_displayed, "s"),  # Bottom
                    (x0_displayed, y1_displayed, "sw"),  # Bottom-left
                    (x0_displayed, (y0_displayed + y1_displayed) / 2, "w"),  # Left
                ]
                
                for handle_x, handle_y, handle_pos in handles:
                    self.canvas.create_rectangle(
                        handle_x - handle_size/2, handle_y - handle_size/2,
                        handle_x + handle_size/2, handle_y + handle_size/2,
                        outline="#0066cc", fill="#ffffff", width=2,
                        tags=(f"field_{field_name}", f"handle_{handle_pos}", "resize_handle")
                    )
            
            # Add field name label (only if field is large enough)
            field_width = x1_displayed - x0_displayed
            field_height = y1_displayed - y0_displayed
            if field_width > 50 and field_height > 15:
                label_x = (x0_displayed + x1_displayed) / 2
                label_y = y0_displayed - 12 if y0_displayed > 15 else y1_displayed + 12
                self.canvas.create_text(label_x, label_y, text=field_name,
                                       fill=outline, font=("Arial", 8, "bold"),
                                       tags=(f"field_{field_name}", "field_label"))
    
    def _get_field_coords(self, field_name: str, page_num: int) -> Tuple[float, float, float, float] | None:
        """Get field coordinates in canvas coordinate system.
        
        Returns:
            Tuple of (x0, y0, x1, y1) in canvas coordinates, or None if field not found
        """
        if field_name not in self.field_data:
            return None
        
        field_info = self.field_data[field_name]
        if field_info["page"] != page_num:
            return None
        
        if not self.pdf_images or page_num >= len(self.pdf_images):
            return None
        
        original_img = self.pdf_images[page_num]
        displayed_img = self.pdf_image_tk[0] if self.pdf_image_tk else None
        if not displayed_img:
            return None
        
        rect = field_info["rect"]
        
        # Get PDF page dimensions (from cached data)
        if page_num < len(self.pdf_page_dimensions):
            pdf_page_width, pdf_page_height = self.pdf_page_dimensions[page_num]
        else:
            # Fallback: estimate from image dimensions
            pdf_page_width = original_img.width / 2.0
            pdf_page_height = original_img.height / 2.0
        
        # Calculate conversion factor: original image pixels per PDF point
        # Image was rendered at 2x (200 DPI), so pixels = points * 2
        pdf_to_img_x = original_img.width / pdf_page_width if pdf_page_width > 0 else 2.0
        pdf_to_img_y = original_img.height / pdf_page_height if pdf_page_height > 0 else 2.0
        
        # Get displayed image dimensions (accounting for zoom)
        displayed_width = displayed_img.width()
        displayed_height = displayed_img.height()
        original_width = original_img.width
        original_height = original_img.height
        
        zoom_scale_x = displayed_width / original_width if original_width > 0 else 1.0
        zoom_scale_y = displayed_height / original_height if original_height > 0 else 1.0
        
        # Convert PDF coordinates to displayed image coordinates
        x0_pdf = rect.x0
        y0_pdf = rect.y0
        x1_pdf = rect.x1
        y1_pdf = rect.y1
        
        # Convert to image coordinates (flip Y axis)
        x0_img = x0_pdf * pdf_to_img_x
        y0_img = (pdf_page_height - y1_pdf) * pdf_to_img_y
        x1_img = x1_pdf * pdf_to_img_x
        y1_img = (pdf_page_height - y0_pdf) * pdf_to_img_y
        
        # Apply zoom
        x0 = x0_img * zoom_scale_x
        y0 = y0_img * zoom_scale_y
        x1 = x1_img * zoom_scale_x
        y1 = y1_img * zoom_scale_y
        
        return (x0, y0, x1, y1)
    
    def _get_resize_handle_at(self, x: float, y: float, field_name: str) -> str | None:
        """Get resize handle at the given position, if any.
        
        Returns:
            Handle position ("nw", "ne", "sw", "se", "n", "s", "e", "w") or None
        """
        coords = self._get_field_coords(field_name, self.current_page)
        if not coords:
            return None
        
        x0, y0, x1, y1 = coords
        handle_size = 8
        threshold = handle_size / 2 + 2  # Add some tolerance
        
        # Check each handle position
        handles = {
            "nw": (x0, y0),
            "n": ((x0 + x1) / 2, y0),
            "ne": (x1, y0),
            "e": (x1, (y0 + y1) / 2),
            "se": (x1, y1),
            "s": ((x0 + x1) / 2, y1),
            "sw": (x0, y1),
            "w": (x0, (y0 + y1) / 2),
        }
        
        for handle_pos, (hx, hy) in handles.items():
            if abs(x - hx) <= threshold and abs(y - hy) <= threshold:
                return handle_pos
        
        return None
    
    def _on_canvas_click(self, event) -> None:
        """Handle click on canvas."""
        # Get click position in canvas coordinates
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        
        # Check if we're waiting for deletion
        if self._waiting_for_delete:
            # Find clicked field
            clicked_field = None
            for field_name in self.field_data.keys():
                coords = self._get_field_coords(field_name, self.current_page)
                if not coords:
                    continue
                
                x0, y0, x1, y1 = coords
                
                # Check if click is within field rectangle
                if x0 <= x <= x1 and y0 <= y <= y1:
                    clicked_field = field_name
                    break
            
            if clicked_field:
                # Confirm deletion
                if messagebox.askyesno("Delete Field Selector", 
                                      f"Delete field selector '{clicked_field}'?\n"
                                      "This will remove the field from configuration."):
                    # Remove from field_config (don't remove from field_data - field still exists in PDF)
                    if clicked_field in self.field_config:
                        del self.field_config[clicked_field]
                    
                    # Clear selection if deleted field was selected
                    if clicked_field == self._selected_field:
                        self._selected_field = None
                    
                    # Redraw page
                    self._display_page(self.current_page)
                    self.status_label.config(text=f"Field selector '{clicked_field}' deleted.")
                    self.log_func(f"Deleted field selector: {clicked_field}", level="INFO")
            else:
                # Cancel deletion mode
                self.status_label.config(text="No field clicked. Deletion cancelled.")
            
            # Exit deletion mode
            self._waiting_for_delete = False
            self.canvas.config(cursor="")
            return
        
        # Check if we're waiting for manual positioning
        if self._waiting_for_manual_position:
            field_name = self._waiting_for_manual_position
            self._position_field_manually(field_name, x, y)
            self._waiting_for_manual_position = None
            self.canvas.config(cursor="")
            return
        
        # Check if clicking on a resize handle
        if self._selected_field:
            handle = self._get_resize_handle_at(x, y, self._selected_field)
            if handle:
                # Start resizing
                self._resizing_field = self._selected_field
                self._resize_handle = handle
                self._resize_start_pos = (x, y)
                coords = self._get_field_coords(self._selected_field, self.current_page)
                if coords:
                    self._resize_start_rect = coords
                # Store original PDF rect for resize
                if self._selected_field in self.field_data:
                    rect = self.field_data[self._selected_field]["rect"]
                    self._resize_start_pdf_rect = (rect.x0, rect.y0, rect.x1, rect.y1)
                return
        
        # Find clicked field
        clicked_field = None
        clicked_on_field = False
        
        for field_name in self.field_data.keys():
            coords = self._get_field_coords(field_name, self.current_page)
            if not coords:
                continue
            
            x0, y0, x1, y1 = coords
            
            # Check if click is within field rectangle
            if x0 <= x <= x1 and y0 <= y <= y1:
                clicked_field = field_name
                clicked_on_field = True
                break
        
        if clicked_field:
            # Check if clicking on a resize handle (only if field is selected and in select mode)
            if self._selected_field == clicked_field and self._select_mode:
                handle = self._get_resize_handle_at(x, y, clicked_field)
                if handle:
                    # Start resizing
                    self._resizing_field = clicked_field
                    self._resize_handle = handle
                    self._resize_start_pos = (x, y)
                    coords = self._get_field_coords(clicked_field, self.current_page)
                    if coords:
                        self._resize_start_rect = coords
                    # Store original PDF rect for resize
                    if clicked_field in self.field_data:
                        rect = self.field_data[clicked_field]["rect"]
                        self._resize_start_pdf_rect = (rect.x0, rect.y0, rect.x1, rect.y1)
                    return
            
            # If clicking on an existing selector (field in field_data)
            if clicked_field in self.field_data:
                # If in select mode, select the field for dragging/resizing
                if self._select_mode:
                    # Select the field (shows resize handles)
                    self._selected_field = clicked_field
                    self._display_page(self.current_page)  # Redraw to show selection
                    
                    # Start dragging if clicking on field (not on resize handle)
                    if not self._get_resize_handle_at(x, y, clicked_field):
                        self._dragging_field = clicked_field
                        self._drag_start_pos = (x, y)
                        # Get current field rect for drag offset calculation
                        coords = self._get_field_coords(clicked_field, self.current_page)
                        if coords:
                            x0, y0, x1, y1 = coords
                            self._drag_offset = (x - x0, y - y0)
                        # Store original PDF rect when drag starts
                        if clicked_field in self.field_data:
                            rect = self.field_data[clicked_field]["rect"]
                            self._drag_start_pdf_rect = (rect.x0, rect.y0, rect.x1, rect.y1)
                    
                    self.status_label.config(text=f"Selected: {clicked_field} - Drag to move, drag handles to resize")
                    return
                else:
                    # Not in select mode - open edit dialog immediately
                    self._open_edit_selector_dialog(clicked_field)
                    return
            
            # If field is not in field_data, it's a new field - select it for dragging/resizing
            # Select the field
            self._selected_field = clicked_field
            self._display_page(self.current_page)  # Redraw to show selection
            
            # Start dragging if clicking on field (not on resize handle)
            if not self._get_resize_handle_at(x, y, clicked_field):
                self._dragging_field = clicked_field
                self._drag_start_pos = (x, y)
                # Get current field rect for drag offset calculation
                coords = self._get_field_coords(clicked_field, self.current_page)
                if coords:
                    x0, y0, x1, y1 = coords
                    self._drag_offset = (x - x0, y - y0)
                # Store original PDF rect when drag starts
                if clicked_field in self.field_data:
                    rect = self.field_data[clicked_field]["rect"]
                    self._drag_start_pdf_rect = (rect.x0, rect.y0, rect.x1, rect.y1)
            
            self.status_label.config(text=f"Selected: {clicked_field} - Drag to move, drag handles to resize")
        else:
            # Deselect if clicking on empty space
            if self._selected_field and not self._select_mode:
                self._selected_field = None
                self._display_page(self.current_page)
            if self._select_mode:
                self.status_label.config(text="Select Mode: Click on a field to select it for dragging/resizing")
            else:
                self.status_label.config(text="Edit Mode: Click on a field to edit it")
    
    def _position_field_manually(self, field_name: str, canvas_x: float, canvas_y: float) -> None:
        """Position a field manually at the clicked location."""
        if field_name not in self.field_data:
            return
        
        if not self.pdf_images or self.current_page >= len(self.pdf_images):
            return
        
        original_img = self.pdf_images[self.current_page]
        displayed_img = self.pdf_image_tk[0] if self.pdf_image_tk else None
        if not displayed_img:
            return
        
        # Get PDF page dimensions
        if self.current_page < len(self.pdf_page_dimensions):
            pdf_page_width, pdf_page_height = self.pdf_page_dimensions[self.current_page]
        else:
            pdf_page_width = original_img.width / 2.0
            pdf_page_height = original_img.height / 2.0
        
        # Get displayed image dimensions (accounting for zoom)
        displayed_width = displayed_img.width()
        displayed_height = displayed_img.height()
        original_width = original_img.width
        original_height = original_img.height
        
        zoom_scale_x = displayed_width / original_width if original_width > 0 else 1.0
        zoom_scale_y = displayed_height / original_height if original_height > 0 else 1.0
        
        # Convert canvas coordinates back to original image coordinates
        img_x = canvas_x / zoom_scale_x if zoom_scale_x > 0 else canvas_x
        img_y = canvas_y / zoom_scale_y if zoom_scale_y > 0 else canvas_y
        
        # Convert image coordinates to PDF coordinates
        # Image was rendered at 2x (200 DPI), so pixels = points * 2
        pdf_to_img_x = original_width / pdf_page_width if pdf_page_width > 0 else 2.0
        pdf_to_img_y = original_height / pdf_page_height if pdf_page_height > 0 else 2.0
        
        # Convert to PDF coordinates (flip Y axis)
        pdf_x = img_x / pdf_to_img_x if pdf_to_img_x > 0 else img_x / 2.0
        pdf_y = pdf_page_height - (img_y / pdf_to_img_y if pdf_to_img_y > 0 else img_y / 2.0)
        
        # Get field type from waiting state
        field_type = getattr(self, '_waiting_field_type', "text")
        
        # Create a rectangle around the click point (size depends on field type)
        if field_type == "checkbox":
            # Checkboxes are typically square, around 12-15 points
            default_size = 12
            x0_pdf = pdf_x
            y0_pdf = pdf_y - default_size / 2
            x1_pdf = pdf_x + default_size
            y1_pdf = pdf_y + default_size / 2
        elif field_type == "radio":
            # Radio buttons are typically circular/square, around 10-12 points
            default_size = 10
            x0_pdf = pdf_x
            y0_pdf = pdf_y - default_size / 2
            x1_pdf = pdf_x + default_size
            y1_pdf = pdf_y + default_size / 2
        else:
            # Text fields are rectangular, typically 100x20 points
            default_width = 100
            default_height = 20
            x0_pdf = pdf_x
            y0_pdf = pdf_y - default_height / 2
            x1_pdf = pdf_x + default_width
            y1_pdf = pdf_y + default_height / 2
        
        # Create RectLike object
        class RectLike:
            def __init__(self, x0, y0, x1, y1):
                self.x0 = x0
                self.y0 = y0
                self.x1 = x1
                self.y1 = y1
        
        # Get field type from waiting state or existing field data
        field_type = getattr(self, '_waiting_field_type', self.field_data[field_name].get("field_type", "text"))
        
        # Update field data
        self.field_data[field_name]["rect"] = RectLike(x0_pdf, y0_pdf, x1_pdf, y1_pdf)
        self.field_data[field_name]["page"] = self.current_page
        self.field_data[field_name]["field_type"] = field_type
        if "needs_positioning" in self.field_data[field_name]:
            del self.field_data[field_name]["needs_positioning"]
        
        # Clear waiting state
        if hasattr(self, '_waiting_field_type'):
            delattr(self, '_waiting_field_type')
        
        # Redraw page to show the positioned field
        self._display_page(self.current_page)
        
        # Automatically open configuration dialog
        self._configure_field(field_name)
        
        self.log_func(f"Manually positioned {field_type} field '{field_name}' at ({pdf_x:.1f}, {pdf_y:.1f})", level="INFO")
    
    def _canvas_to_pdf_coords(self, canvas_x: float, canvas_y: float) -> Tuple[float, float]:
        """Convert canvas coordinates to PDF coordinates.
        
        Returns:
            Tuple of (pdf_x, pdf_y) in PDF coordinate system
        """
        if not self.pdf_images or self.current_page >= len(self.pdf_images):
            return (0, 0)
        
        original_img = self.pdf_images[self.current_page]
        displayed_img = self.pdf_image_tk[0] if self.pdf_image_tk else None
        if not displayed_img:
            return (0, 0)
        
        # Get PDF page dimensions
        if self.current_page < len(self.pdf_page_dimensions):
            pdf_page_width, pdf_page_height = self.pdf_page_dimensions[self.current_page]
        else:
            pdf_page_width = original_img.width / 2.0
            pdf_page_height = original_img.height / 2.0
        
        # Get displayed image dimensions (accounting for zoom)
        displayed_width = displayed_img.width()
        displayed_height = displayed_img.height()
        original_width = original_img.width
        original_height = original_img.height
        
        zoom_scale_x = displayed_width / original_width if original_width > 0 else 1.0
        zoom_scale_y = displayed_height / original_height if original_height > 0 else 1.0
        
        # Convert canvas coordinates back to original image coordinates
        img_x = canvas_x / zoom_scale_x if zoom_scale_x > 0 else canvas_x
        img_y = canvas_y / zoom_scale_y if zoom_scale_y > 0 else canvas_y
        
        # Convert image coordinates to PDF coordinates
        pdf_to_img_x = original_width / pdf_page_width if pdf_page_width > 0 else 2.0
        pdf_to_img_y = original_height / pdf_page_height if pdf_page_height > 0 else 2.0
        
        # Convert to PDF coordinates (flip Y axis)
        pdf_x = img_x / pdf_to_img_x if pdf_to_img_x > 0 else img_x / 2.0
        pdf_y = pdf_page_height - (img_y / pdf_to_img_y if pdf_to_img_y > 0 else img_y / 2.0)
        
        return (pdf_x, pdf_y)
    
    def _pdf_to_canvas_coords(self, pdf_x: float, pdf_y: float) -> Tuple[float, float]:
        """Convert PDF coordinates to canvas coordinates.
        
        Returns:
            Tuple of (canvas_x, canvas_y) in canvas coordinate system
        """
        if not self.pdf_images or self.current_page >= len(self.pdf_images):
            return (0, 0)
        
        original_img = self.pdf_images[self.current_page]
        displayed_img = self.pdf_image_tk[0] if self.pdf_image_tk else None
        if not displayed_img:
            return (0, 0)
        
        # Get PDF page dimensions
        if self.current_page < len(self.pdf_page_dimensions):
            pdf_page_width, pdf_page_height = self.pdf_page_dimensions[self.current_page]
        else:
            pdf_page_width = original_img.width / 2.0
            pdf_page_height = original_img.height / 2.0
        
        # Convert PDF coordinates to image coordinates
        pdf_to_img_x = original_img.width / pdf_page_width if pdf_page_width > 0 else 2.0
        pdf_to_img_y = original_img.height / pdf_page_height if pdf_page_height > 0 else 2.0
        
        # Convert to image coordinates (flip Y axis)
        img_x = pdf_x * pdf_to_img_x
        img_y = (pdf_page_height - pdf_y) * pdf_to_img_y
        
        # Get displayed image dimensions (accounting for zoom)
        displayed_width = displayed_img.width()
        displayed_height = displayed_img.height()
        original_width = original_img.width
        original_height = original_img.height
        
        zoom_scale_x = displayed_width / original_width if original_width > 0 else 1.0
        zoom_scale_y = displayed_height / original_height if original_height > 0 else 1.0
        
        # Apply zoom
        canvas_x = img_x * zoom_scale_x
        canvas_y = img_y * zoom_scale_y
        
        return (canvas_x, canvas_y)
    
    def _on_canvas_drag(self, event) -> None:
        """Handle mouse drag on canvas."""
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        
        # Handle resizing
        if self._resizing_field and self._resize_handle and self._resize_start_rect:
            self._handle_resize(x, y)
            return
        
        # Handle dragging
        if self._dragging_field and self._drag_start_pos:
            self._handle_drag(x, y)
            return
    
    def _handle_resize(self, canvas_x: float, canvas_y: float) -> None:
        """Handle field resizing."""
        if not self._resizing_field or not self._resize_handle or not self._resize_start_rect or not self._resize_start_pos:
            return
        
        field_name = self._resizing_field
        handle = self._resize_handle
        start_canvas_x, start_canvas_y = self._resize_start_pos
        
        # Get current field rect in PDF coordinates
        if field_name not in self.field_data:
            return
        
        field_info = self.field_data[field_name]
        rect = field_info["rect"]
        
        # Get original PDF coordinates (from when resize started)
        # We need to track the original PDF rect when resize started
        if not hasattr(self, '_resize_start_pdf_rect'):
            # Store original PDF rect if not already stored
            self._resize_start_pdf_rect = (rect.x0, rect.y0, rect.x1, rect.y1)
        
        orig_x0_pdf, orig_y0_pdf, orig_x1_pdf, orig_y1_pdf = self._resize_start_pdf_rect
        
        # Calculate delta in canvas coordinates
        delta_canvas_x = canvas_x - start_canvas_x
        delta_canvas_y = canvas_y - start_canvas_y
        
        # Convert delta to PDF coordinates
        # Convert start position to PDF
        start_pdf_x, start_pdf_y = self._canvas_to_pdf_coords(start_canvas_x, start_canvas_y)
        # Convert current position to PDF
        current_pdf_x, current_pdf_y = self._canvas_to_pdf_coords(canvas_x, canvas_y)
        
        # Calculate delta in PDF coordinates
        delta_pdf_x = current_pdf_x - start_pdf_x
        delta_pdf_y = start_pdf_y - current_pdf_y  # Y is inverted (canvas Y increases downward, PDF Y increases upward)
        
        # Calculate new rectangle based on handle
        new_x0_pdf = orig_x0_pdf
        new_y0_pdf = orig_y0_pdf
        new_x1_pdf = orig_x1_pdf
        new_y1_pdf = orig_y1_pdf
        
        if "w" in handle:
            new_x0_pdf = orig_x0_pdf + delta_pdf_x
        if "e" in handle:
            new_x1_pdf = orig_x1_pdf + delta_pdf_x
        if "n" in handle:
            # North means moving top edge (y1 in PDF, which is top)
            new_y1_pdf = orig_y1_pdf + delta_pdf_y
        if "s" in handle:
            # South means moving bottom edge (y0 in PDF, which is bottom)
            new_y0_pdf = orig_y0_pdf + delta_pdf_y
        
        # Ensure minimum size (5 PDF points)
        min_size = 5
        if abs(new_x1_pdf - new_x0_pdf) < min_size:
            if "w" in handle:
                new_x0_pdf = new_x1_pdf - min_size
            else:
                new_x1_pdf = new_x0_pdf + min_size
        if abs(new_y1_pdf - new_y0_pdf) < min_size:
            if "n" in handle:
                new_y1_pdf = new_y0_pdf + min_size
            else:
                new_y0_pdf = new_y1_pdf - min_size
        
        # Ensure x0 < x1 and y0 < y1 (in PDF coordinates)
        if new_x1_pdf < new_x0_pdf:
            new_x0_pdf, new_x1_pdf = new_x1_pdf, new_x0_pdf
        if new_y1_pdf < new_y0_pdf:
            new_y0_pdf, new_y1_pdf = new_y1_pdf, new_y0_pdf
        
        # Update field rectangle
        class RectLike:
            def __init__(self, x0, y0, x1, y1):
                self.x0 = x0
                self.y0 = y0
                self.x1 = x1
                self.y1 = y1
        
        self.field_data[field_name]["rect"] = RectLike(new_x0_pdf, new_y0_pdf, new_x1_pdf, new_y1_pdf)
        
        # Update resize start position for next resize event
        self._resize_start_pos = (canvas_x, canvas_y)
        
        # Redraw page
        self._display_page(self.current_page)
    
    def _handle_drag(self, canvas_x: float, canvas_y: float) -> None:
        """Handle field dragging."""
        if not self._dragging_field or not self._drag_start_pos or not self._drag_start_pdf_rect:
            return
        
        field_name = self._dragging_field
        start_canvas_x, start_canvas_y = self._drag_start_pos
        
        # Get original PDF coordinates (from when drag started)
        orig_x0_pdf, orig_y0_pdf, orig_x1_pdf, orig_y1_pdf = self._drag_start_pdf_rect
        
        # Calculate width and height in PDF coordinates (preserve size)
        pdf_width = orig_x1_pdf - orig_x0_pdf
        pdf_height = orig_y1_pdf - orig_y0_pdf
        
        # Convert start canvas position to PDF coordinates
        start_pdf_x, start_pdf_y = self._canvas_to_pdf_coords(start_canvas_x, start_canvas_y)
        # Convert current canvas position to PDF coordinates
        current_pdf_x, current_pdf_y = self._canvas_to_pdf_coords(canvas_x, canvas_y)
        
        # Calculate delta in PDF coordinates
        delta_pdf_x = current_pdf_x - start_pdf_x
        delta_pdf_y = start_pdf_y - current_pdf_y  # Y is inverted (canvas Y increases downward, PDF Y increases upward)
        
        # Apply delta to original position
        new_x0_pdf = orig_x0_pdf + delta_pdf_x
        new_y0_pdf = orig_y0_pdf + delta_pdf_y
        new_x1_pdf = new_x0_pdf + pdf_width
        new_y1_pdf = new_y0_pdf + pdf_height
        
        # Ensure y0 < y1 (in PDF coordinates, y increases upward)
        if new_y1_pdf < new_y0_pdf:
            new_y0_pdf, new_y1_pdf = new_y1_pdf, new_y0_pdf
        
        # Update field rectangle
        class RectLike:
            def __init__(self, x0, y0, x1, y1):
                self.x0 = x0
                self.y0 = y0
                self.x1 = x1
                self.y1 = y1
        
        self.field_data[field_name]["rect"] = RectLike(new_x0_pdf, new_y0_pdf, new_x1_pdf, new_y1_pdf)
        
        # Redraw page
        self._display_page(self.current_page)
    
    def _on_canvas_release(self, event) -> None:
        """Handle mouse release on canvas."""
        # Stop resizing
        if self._resizing_field:
            self.log_func(f"Finished resizing field: {self._resizing_field}", level="DEBUG")
            self._resizing_field = None
            self._resize_handle = None
            self._resize_start_pos = None
            self._resize_start_rect = None
            if hasattr(self, '_resize_start_pdf_rect'):
                delattr(self, '_resize_start_pdf_rect')
        
        # Stop dragging
        if self._dragging_field:
            self.log_func(f"Finished dragging field: {self._dragging_field}", level="DEBUG")
            self._dragging_field = None
            self._drag_start_pos = None
            self._drag_offset = None
            self._drag_start_pdf_rect = None
    
    def _on_canvas_double_click(self, event) -> None:
        """Handle double-click on canvas."""
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        
        # Find clicked field
        clicked_field = None
        for field_name in self.field_data.keys():
            coords = self._get_field_coords(field_name, self.current_page)
            if not coords:
                continue
            
            x0, y0, x1, y1 = coords
            
            # Check if click is within field rectangle
            if x0 <= x <= x1 and y0 <= y <= y1:
                clicked_field = field_name
                break
        
        if clicked_field:
            # Double-click opens full configuration dialog
            self._configure_field(clicked_field)
    
    def _delete_selected_field(self) -> None:
        """Delete the currently selected field."""
        if not self._selected_field:
            return
        
        field_name = self._selected_field
        if messagebox.askyesno("Delete Field Selector", 
                              f"Delete field selector '{field_name}'?\n"
                              "This will remove the field from configuration."):
            # Remove from field_config
            if field_name in self.field_config:
                del self.field_config[field_name]
            
            # Clear selection
            self._selected_field = None
            
            # Redraw page
            self._display_page(self.current_page)
            self.status_label.config(text=f"Field selector '{field_name}' deleted.")
            self.log_func(f"Deleted field selector: {field_name}", level="INFO")
    
    def _clear_selection(self) -> None:
        """Clear the current selection."""
        if self._selected_field:
            self._selected_field = None
            self._display_page(self.current_page)
            self.status_label.config(text="Selection cleared.")
    
    def _on_canvas_motion(self, event) -> None:
        """Handle mouse motion on canvas."""
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        
        # If dragging or resizing, return (don't change cursor)
        if self._dragging_field or self._resizing_field:
            return
        
        # Check if mouse is over a resize handle (only in select mode)
        if self._selected_field and self._select_mode:
            handle = self._get_resize_handle_at(x, y, self._selected_field)
            if handle:
                # Set cursor based on handle
                cursors = {
                    "nw": "size_nw_se", "ne": "size_ne_sw",
                    "sw": "size_ne_sw", "se": "size_nw_se",
                    "n": "size_n_s", "s": "size_n_s",
                    "e": "size_w_e", "w": "size_w_e"
                }
                self.canvas.config(cursor=cursors.get(handle, "crosshair"))
                self.status_label.config(text=f"Resize handle: {handle} - Drag to resize")
                return
        
        # Check if mouse is over a field
        over_field = None
        for field_name in self.field_data.keys():
            coords = self._get_field_coords(field_name, self.current_page)
            if not coords:
                continue
            
            x0, y0, x1, y1 = coords
            
            if x0 <= x <= x1 and y0 <= y <= y1:
                over_field = field_name
                break
        
        # Change cursor
        if over_field:
            self.canvas.config(cursor="hand2")
            # Check if field is configured
            is_configured = (over_field in self.field_config and 
                           self.field_config[over_field].get("enabled", False))
            is_selected = (over_field == self._selected_field)
            config_status = " (Enabled)" if is_configured else " (Disabled)"
            select_status = " [Selected]" if is_selected else ""
            
            if self._select_mode:
                self.status_label.config(text=f"Field: {over_field}{config_status}{select_status} - Click to select, right-click to edit")
            else:
                self.status_label.config(text=f"Field: {over_field}{config_status}{select_status} - Click to edit, double-click to configure")
        else:
            self.canvas.config(cursor="")
            if self._select_mode:
                if self._selected_field:
                    self.status_label.config(text=f"Selected: {self._selected_field} - Drag to move, drag handles to resize")
                else:
                    self.status_label.config(text="Select Mode: Click on a field to select it for dragging/resizing")
            else:
                if self._selected_field:
                    self.status_label.config(text=f"Selected: {self._selected_field} - Double-click to configure")
                else:
                    self.status_label.config(text="Edit Mode: Click on a field to edit it")
    
    def _open_edit_selector_dialog(self, field_name: str) -> None:
        """Open edit selector dialog for an existing field selector."""
        # Create edit dialog window
        edit_dialog = tk.Toplevel(self.window)
        edit_dialog.title(f"Edit Selector: {field_name}")
        edit_dialog.geometry("400x450")
        edit_dialog.transient(self.window)
        edit_dialog.grab_set()
        edit_dialog.configure(bg="#f0f0f0")
        
        main_frame = tk.Frame(edit_dialog, bg="#f0f0f0")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Field name
        tk.Label(main_frame, text=f"Field: {field_name}",
                font=("Segoe UI", 12, "bold"), bg="#f0f0f0", fg="#003366").pack(pady=(0, 20))
        
        # Get current configuration
        current_config = self.field_config.get(field_name, {
            "type": "static",
            "value": "",
            "enabled": False
        })
        is_enabled = current_config.get("enabled", False)
        
        # Get field type
        field_type = self.field_data.get(field_name, {}).get("field_type", "text")
        
        # Status label
        if is_enabled:
            status_text = "Status: Enabled (Green box)"
            status_color = "#28a745"
        else:
            status_text = "Status: Disabled (Yellow box)"
            status_color = "#ffc107"
        
        status_label = tk.Label(main_frame, text=status_text,
                               font=("Segoe UI", 10), bg="#f0f0f0", fg=status_color)
        status_label.pack(pady=(0, 20))
        
        # Enable/Disable checkbox (live update)
        enable_var = tk.BooleanVar(value=is_enabled)
        updating_enable = False  # Flag to prevent circular updates
        
        def enable_field():
            """Enable the field."""
            nonlocal updating_enable
            if updating_enable:
                return
            
            updating_enable = True
            if field_name not in self.field_config:
                # Create default configuration
                self.field_config[field_name] = {
                    "type": "static",
                    "value": "",
                    "enabled": True,
                    "field_type": field_type
                }
            else:
                self.field_config[field_name]["enabled"] = True
            
            # Update status label in dialog
            status_label.config(text="Status: Enabled (Green box)", fg="#28a745")
            enable_var.set(True)
            updating_enable = False
            
            self._display_page(self.current_page)
            self.status_label.config(text=f"Field '{field_name}' enabled.")
            self.log_func(f"Enabled field: {field_name}", level="INFO")
        
        def disable_field():
            """Disable the field."""
            nonlocal updating_enable
            if updating_enable:
                return
            
            updating_enable = True
            if field_name not in self.field_config:
                # Create default configuration with enabled=False
                self.field_config[field_name] = {
                    "type": "static",
                    "value": "",
                    "enabled": False,
                    "field_type": field_type
                }
            else:
                self.field_config[field_name]["enabled"] = False
            
            # Update status label in dialog
            status_label.config(text="Status: Disabled (Yellow box)", fg="#ffc107")
            enable_var.set(False)
            updating_enable = False
            
            self._display_page(self.current_page)
            self.status_label.config(text=f"Field '{field_name}' disabled.")
            self.log_func(f"Disabled field: {field_name}", level="INFO")
        
        def update_enable_status():
            """Update enable status when checkbox changes."""
            if enable_var.get():
                enable_field()
            else:
                disable_field()
        
        enable_check = tk.Checkbutton(main_frame, text="Enable this field",
                                     variable=enable_var, font=("Segoe UI", 10),
                                     bg="#f0f0f0", fg="#000000",
                                     command=update_enable_status)
        enable_check.pack(pady=(0, 20), anchor="w")
        
        # Buttons frame
        buttons_frame = tk.Frame(main_frame, bg="#f0f0f0")
        buttons_frame.pack(fill="x", pady=(10, 0))
        
        def configure_field():
            """Open full configuration dialog."""
            edit_dialog.destroy()
            self._configure_field(field_name)
        
        def remove_selector():
            """Remove selector from screen (delete from field_data)."""
            if messagebox.askyesno("Remove Selector", 
                                  f"Remove selector '{field_name}' from screen?\n\n"
                                  "This will remove the field box from the PDF view.\n"
                                  "You can add it back manually later."):
                # Remove from field_data (disappears from screen)
                if field_name in self.field_data:
                    del self.field_data[field_name]
                # Remove from config
                if field_name in self.field_config:
                    del self.field_config[field_name]
                
                # Clear selection if selected field was removed
                if self._selected_field == field_name:
                    self._selected_field = None
                
                # Redraw page
                self._display_page(self.current_page)
                self.status_label.config(text=f"Removed selector '{field_name}' from screen.")
                self.log_func(f"Removed selector: {field_name}", level="INFO")
                edit_dialog.destroy()
        
        def edit_coordinates():
            """Open coordinate editing dialog."""
            edit_dialog.destroy()
            self._edit_field_coordinates(field_name)
        
        # Action buttons
        tk.Button(buttons_frame, text="Enable",
                 command=enable_field,
                 bg="#28a745", fg="white", font=("Segoe UI", 10, "bold"),
                 padx=15, pady=10, cursor="hand2", relief="flat").pack(fill="x", pady=5)
        
        tk.Button(buttons_frame, text="Disable",
                 command=disable_field,
                 bg="#ffc107", fg="black", font=("Segoe UI", 10),
                 padx=15, pady=10, cursor="hand2", relief="flat").pack(fill="x", pady=5)
        
        tk.Button(buttons_frame, text="Configure Field...",
                 command=configure_field,
                 bg="#007bff", fg="white", font=("Segoe UI", 10),
                 padx=15, pady=10, cursor="hand2", relief="flat").pack(fill="x", pady=5)
        
        tk.Button(buttons_frame, text="Edit Coordinates...",
                 command=edit_coordinates,
                 bg="#17a2b8", fg="white", font=("Segoe UI", 10),
                 padx=15, pady=10, cursor="hand2", relief="flat").pack(fill="x", pady=5)
        
        tk.Button(buttons_frame, text="Remove Selector",
                 command=remove_selector,
                 bg="#dc3545", fg="white", font=("Segoe UI", 10, "bold"),
                 padx=15, pady=10, cursor="hand2", relief="flat").pack(fill="x", pady=5)
        
        # Close button
        tk.Button(buttons_frame, text="Close",
                 command=edit_dialog.destroy,
                 bg="#6c757d", fg="white", font=("Segoe UI", 10),
                 padx=15, pady=10, cursor="hand2", relief="flat").pack(fill="x", pady=(15, 0))
    
    def _edit_field_coordinates(self, field_name: str) -> None:
        """Open coordinate editing dialog for a field."""
        if field_name not in self.field_data:
            messagebox.showwarning("Field Not Found", f"Field '{field_name}' not found.")
            return
        
        # Create coordinate editing dialog
        coord_dialog = tk.Toplevel(self.window)
        coord_dialog.title(f"Edit Coordinates: {field_name}")
        coord_dialog.geometry("450x350")
        coord_dialog.transient(self.window)
        coord_dialog.grab_set()
        coord_dialog.configure(bg="#f0f0f0")
        
        main_frame = tk.Frame(coord_dialog, bg="#f0f0f0")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Field name
        tk.Label(main_frame, text=f"Field: {field_name}",
                font=("Segoe UI", 12, "bold"), bg="#f0f0f0", fg="#003366").pack(pady=(0, 20))
        
        # Get current coordinates
        field_rect = self.field_data[field_name].get("rect")
        if field_rect:
            current_x0 = field_rect.x0
            current_y0 = field_rect.y0
            current_x1 = field_rect.x1
            current_y1 = field_rect.y1
            current_width = current_x1 - current_x0
            current_height = current_y1 - current_y0
        else:
            current_x0 = current_y0 = current_x1 = current_y1 = 0.0
            current_width = current_height = 0.0
        
        # Position section
        position_frame = tk.LabelFrame(main_frame, text="Position (PDF Coordinates)", 
                                       font=("Segoe UI", 10, "bold"), bg="#f0f0f0", fg="#003366",
                                       padx=10, pady=10)
        position_frame.pack(fill="x", pady=(0, 15))
        
        # X0, Y0 (bottom-left)
        coord_frame1 = tk.Frame(position_frame, bg="#f0f0f0")
        coord_frame1.pack(fill="x", pady=(0, 5))
        tk.Label(coord_frame1, text="X0 (Left):", font=("Segoe UI", 9), bg="#f0f0f0", width=12, anchor="w").pack(side="left", padx=(0, 5))
        x0_var = tk.StringVar(value=f"{current_x0:.2f}")
        x0_entry = tk.Entry(coord_frame1, textvariable=x0_var, width=15, font=("Segoe UI", 9))
        x0_entry.pack(side="left", padx=(0, 10))
        
        tk.Label(coord_frame1, text="Y0 (Bottom):", font=("Segoe UI", 9), bg="#f0f0f0", width=12, anchor="w").pack(side="left", padx=(0, 5))
        y0_var = tk.StringVar(value=f"{current_y0:.2f}")
        y0_entry = tk.Entry(coord_frame1, textvariable=y0_var, width=15, font=("Segoe UI", 9))
        y0_entry.pack(side="left")
        
        # X1, Y1 (top-right)
        coord_frame2 = tk.Frame(position_frame, bg="#f0f0f0")
        coord_frame2.pack(fill="x", pady=(0, 5))
        tk.Label(coord_frame2, text="X1 (Right):", font=("Segoe UI", 9), bg="#f0f0f0", width=12, anchor="w").pack(side="left", padx=(0, 5))
        x1_var = tk.StringVar(value=f"{current_x1:.2f}")
        x1_entry = tk.Entry(coord_frame2, textvariable=x1_var, width=15, font=("Segoe UI", 9))
        x1_entry.pack(side="left", padx=(0, 10))
        
        tk.Label(coord_frame2, text="Y1 (Top):", font=("Segoe UI", 9), bg="#f0f0f0", width=12, anchor="w").pack(side="left", padx=(0, 5))
        y1_var = tk.StringVar(value=f"{current_y1:.2f}")
        y1_entry = tk.Entry(coord_frame2, textvariable=y1_var, width=15, font=("Segoe UI", 9))
        y1_entry.pack(side="left")
        
        # Width, Height
        size_frame = tk.Frame(position_frame, bg="#f0f0f0")
        size_frame.pack(fill="x", pady=(5, 0))
        tk.Label(size_frame, text="Width:", font=("Segoe UI", 9), bg="#f0f0f0", width=12, anchor="w").pack(side="left", padx=(0, 5))
        width_var = tk.StringVar(value=f"{current_width:.2f}")
        width_entry = tk.Entry(size_frame, textvariable=width_var, width=15, font=("Segoe UI", 9))
        width_entry.pack(side="left", padx=(0, 10))
        
        tk.Label(size_frame, text="Height:", font=("Segoe UI", 9), bg="#f0f0f0", width=12, anchor="w").pack(side="left", padx=(0, 5))
        height_var = tk.StringVar(value=f"{current_height:.2f}")
        height_entry = tk.Entry(size_frame, textvariable=height_var, width=15, font=("Segoe UI", 9))
        height_entry.pack(side="left")
        
        # Helper: Update width/height when coordinates change
        def update_size(*args):
            try:
                x0 = float(x0_var.get())
                y0 = float(y0_var.get())
                x1 = float(x1_var.get())
                y1 = float(y1_var.get())
                width_var.set(f"{abs(x1 - x0):.2f}")
                height_var.set(f"{abs(y1 - y0):.2f}")
            except ValueError:
                pass
        
        # Helper: Update coordinates when width/height change (maintains x0, y0)
        def update_coords_from_size(*args):
            try:
                x0 = float(x0_var.get())
                y0 = float(y0_var.get())
                width = float(width_var.get())
                height = float(height_var.get())
                x1_var.set(f"{x0 + width:.2f}")
                y1_var.set(f"{y0 + height:.2f}")
            except ValueError:
                pass
        
        x0_var.trace("w", update_size)
        y0_var.trace("w", update_size)
        x1_var.trace("w", update_size)
        y1_var.trace("w", update_size)
        width_var.trace("w", update_coords_from_size)
        height_var.trace("w", update_coords_from_size)
        
        # Buttons
        buttons_frame = tk.Frame(main_frame, bg="#f0f0f0")
        buttons_frame.pack(fill="x", pady=(15, 0))
        
        def save_coordinates():
            try:
                new_x0 = float(x0_var.get())
                new_y0 = float(y0_var.get())
                new_x1 = float(x1_var.get())
                new_y1 = float(y1_var.get())
                
                # Validate coordinates
                if new_x1 < new_x0:
                    messagebox.showerror("Error", "X1 must be greater than X0.")
                    return
                if new_y1 < new_y0:
                    messagebox.showerror("Error", "Y1 must be greater than Y0.")
                    return
                
                # Update field rectangle
                class RectLike:
                    def __init__(self, x0, y0, x1, y1):
                        self.x0 = x0
                        self.y0 = y0
                        self.x1 = x1
                        self.y1 = y1
                
                self.field_data[field_name]["rect"] = RectLike(new_x0, new_y0, new_x1, new_y1)
                self.log_func(f"Updated coordinates for '{field_name}': ({new_x0:.2f}, {new_y0:.2f}) to ({new_x1:.2f}, {new_y1:.2f})", level="INFO")
                
                # Refresh display
                self._display_page(self.current_page)
                self.status_label.config(text=f"Coordinates updated for '{field_name}'.")
                
                coord_dialog.destroy()
            except ValueError:
                messagebox.showerror("Error", "Invalid coordinate values. Please enter valid numbers.")
        
        tk.Button(buttons_frame, text="Save",
                 command=save_coordinates,
                 bg="#28a745", fg="white", font=("Segoe UI", 10, "bold"),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="left", padx=5)
        
        tk.Button(buttons_frame, text="Cancel",
                 command=coord_dialog.destroy,
                 bg="#6c757d", fg="white", font=("Segoe UI", 10),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="right", padx=5)
    
    def _configure_field(self, field_name: str) -> None:
        """Configure a field."""
        # Get field type from field_data
        field_type = self.field_data.get(field_name, {}).get("field_type", "text")
        
        # Get current configuration
        current_config = self.field_config.get(field_name, {
            "type": "static",
            "value": "",
            "enabled": False
        })
        
        # For checkboxes, default value is "checked" (Yes) if enabled, "unchecked" (Off) if not
        if field_type == "checkbox":
            if current_config.get("enabled", False):
                default_checkbox_value = current_config.get("value", "checked")
            else:
                default_checkbox_value = "unchecked"
        else:
            default_checkbox_value = current_config.get("value", "")
        
        # Get current field coordinates
        field_rect = None
        if field_name in self.field_data:
            field_rect = self.field_data[field_name].get("rect")
        
        # Create configuration dialog
        config_dialog = tk.Toplevel(self.window)
        config_dialog.title(f"Configure Field: {field_name}")
        config_dialog.geometry("550x550")
        config_dialog.transient(self.window)
        config_dialog.grab_set()
        config_dialog.configure(bg="#f0f0f0")
        
        # Create scrollable frame
        canvas_frame = tk.Canvas(config_dialog, bg="#f0f0f0", highlightthickness=0)
        scrollbar = tk.Scrollbar(config_dialog, orient="vertical", command=canvas_frame.yview)
        main_frame = tk.Frame(canvas_frame, bg="#f0f0f0")
        
        canvas_frame.configure(yscrollcommand=scrollbar.set)
        canvas_frame.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        canvas_frame.create_window((0, 0), window=main_frame, anchor="nw")
        main_frame.bind("<Configure>", lambda e: canvas_frame.configure(scrollregion=canvas_frame.bbox("all")))
        
        # Padding
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Field name and type
        field_info = f"Field: {field_name} ({field_type})"
        tk.Label(main_frame, text=field_info,
                font=("Segoe UI", 12, "bold"), bg="#f0f0f0", fg="#003366").pack(pady=(0, 15))
        
        # Enable checkbox
        enable_var = tk.BooleanVar(value=current_config.get("enabled", False))
        enable_check = tk.Checkbutton(main_frame, text="Enable this field",
                                     variable=enable_var, font=("Segoe UI", 10),
                                     bg="#f0f0f0", fg="#000000")
        enable_check.pack(pady=(0, 15), anchor="w")
        
        # Position section (manual coordinate editing)
        position_frame = tk.LabelFrame(main_frame, text="Position (PDF Coordinates)", 
                                       font=("Segoe UI", 10, "bold"), bg="#f0f0f0", fg="#003366",
                                       padx=10, pady=10)
        position_frame.pack(fill="x", pady=(0, 15))
        
        # Current coordinates
        if field_rect:
            current_x0 = field_rect.x0
            current_y0 = field_rect.y0
            current_x1 = field_rect.x1
            current_y1 = field_rect.y1
            current_width = current_x1 - current_x0
            current_height = current_y1 - current_y0
        else:
            current_x0 = current_y0 = current_x1 = current_y1 = 0.0
            current_width = current_height = 0.0
        
        # X0, Y0 (bottom-left)
        coord_frame1 = tk.Frame(position_frame, bg="#f0f0f0")
        coord_frame1.pack(fill="x", pady=(0, 5))
        tk.Label(coord_frame1, text="X0 (Left):", font=("Segoe UI", 9), bg="#f0f0f0", width=12, anchor="w").pack(side="left", padx=(0, 5))
        x0_var = tk.StringVar(value=f"{current_x0:.2f}")
        x0_entry = tk.Entry(coord_frame1, textvariable=x0_var, width=15, font=("Segoe UI", 9))
        x0_entry.pack(side="left", padx=(0, 10))
        
        tk.Label(coord_frame1, text="Y0 (Bottom):", font=("Segoe UI", 9), bg="#f0f0f0", width=12, anchor="w").pack(side="left", padx=(0, 5))
        y0_var = tk.StringVar(value=f"{current_y0:.2f}")
        y0_entry = tk.Entry(coord_frame1, textvariable=y0_var, width=15, font=("Segoe UI", 9))
        y0_entry.pack(side="left")
        
        # X1, Y1 (top-right)
        coord_frame2 = tk.Frame(position_frame, bg="#f0f0f0")
        coord_frame2.pack(fill="x", pady=(0, 5))
        tk.Label(coord_frame2, text="X1 (Right):", font=("Segoe UI", 9), bg="#f0f0f0", width=12, anchor="w").pack(side="left", padx=(0, 5))
        x1_var = tk.StringVar(value=f"{current_x1:.2f}")
        x1_entry = tk.Entry(coord_frame2, textvariable=x1_var, width=15, font=("Segoe UI", 9))
        x1_entry.pack(side="left", padx=(0, 10))
        
        tk.Label(coord_frame2, text="Y1 (Top):", font=("Segoe UI", 9), bg="#f0f0f0", width=12, anchor="w").pack(side="left", padx=(0, 5))
        y1_var = tk.StringVar(value=f"{current_y1:.2f}")
        y1_entry = tk.Entry(coord_frame2, textvariable=y1_var, width=15, font=("Segoe UI", 9))
        y1_entry.pack(side="left")
        
        # Width, Height
        size_frame = tk.Frame(position_frame, bg="#f0f0f0")
        size_frame.pack(fill="x", pady=(5, 0))
        tk.Label(size_frame, text="Width:", font=("Segoe UI", 9), bg="#f0f0f0", width=12, anchor="w").pack(side="left", padx=(0, 5))
        width_var = tk.StringVar(value=f"{current_width:.2f}")
        width_entry = tk.Entry(size_frame, textvariable=width_var, width=15, font=("Segoe UI", 9))
        width_entry.pack(side="left", padx=(0, 10))
        
        tk.Label(size_frame, text="Height:", font=("Segoe UI", 9), bg="#f0f0f0", width=12, anchor="w").pack(side="left", padx=(0, 5))
        height_var = tk.StringVar(value=f"{current_height:.2f}")
        height_entry = tk.Entry(size_frame, textvariable=height_var, width=15, font=("Segoe UI", 9))
        height_entry.pack(side="left")
        
        # Helper: Update width/height when coordinates change
        def update_size(*args):
            try:
                x0 = float(x0_var.get())
                y0 = float(y0_var.get())
                x1 = float(x1_var.get())
                y1 = float(y1_var.get())
                width_var.set(f"{abs(x1 - x0):.2f}")
                height_var.set(f"{abs(y1 - y0):.2f}")
            except ValueError:
                pass
        
        # Helper: Update coordinates when width/height change (maintains x0, y0)
        def update_coords_from_size(*args):
            try:
                x0 = float(x0_var.get())
                y0 = float(y0_var.get())
                width = float(width_var.get())
                height = float(height_var.get())
                x1_var.set(f"{x0 + width:.2f}")
                y1_var.set(f"{y0 + height:.2f}")
            except ValueError:
                pass
        
        x0_var.trace("w", update_size)
        y0_var.trace("w", update_size)
        x1_var.trace("w", update_size)
        y1_var.trace("w", update_size)
        width_var.trace("w", update_coords_from_size)
        height_var.trace("w", update_coords_from_size)
        
        # For checkbox fields, show checked/unchecked option
        if field_type == "checkbox":
            # Data source (for checkbox, can be Excel column that contains "yes"/"no" or static)
            source_frame = tk.Frame(main_frame, bg="#f0f0f0")
            source_frame.pack(fill="x", pady=(0, 15))
            
            tk.Label(source_frame, text="Data Source:", font=("Segoe UI", 10),
                    bg="#f0f0f0").pack(side="left", padx=(0, 10))
            
            source_var = tk.StringVar(value=current_config.get("type", "static"))
            source_dropdown = ttk.Combobox(source_frame, textvariable=source_var,
                                          values=["excel", "static"], state="readonly", width=15)
            source_dropdown.pack(side="left")
            
            # Value frame (for Excel column or static checked/unchecked)
            value_frame = tk.Frame(main_frame, bg="#f0f0f0")
            value_frame.pack(fill="x", pady=(0, 15))
            
            tk.Label(value_frame, text="Value:", font=("Segoe UI", 10),
                    bg="#f0f0f0").pack(side="left", padx=(0, 10))
            
            # Determine initial value based on config
            initial_value = current_config.get("value", "checked" if current_config.get("enabled") else "unchecked")
            value_var = tk.StringVar(value=initial_value)
            value_widget = None
            
            def update_checkbox_value_widget():
                nonlocal value_widget
                if value_widget:
                    value_widget.destroy()
                
                if source_var.get() == "excel":
                    # Excel: show dropdown for Excel columns
                    value_widget = ttk.Combobox(value_frame, textvariable=value_var,
                                              values=self.excel_columns, width=25, state="readonly")
                    value_widget.pack(side="left")
                    # Set value if it's a valid Excel column
                    if initial_value in self.excel_columns:
                        value_var.set(initial_value)
                    else:
                        value_var.set("")  # Clear if not a valid column
                else:
                    # Static: show radio buttons for checked/unchecked
                    static_frame = tk.Frame(value_frame, bg="#f0f0f0")
                    static_frame.pack(side="left")
                    tk.Radiobutton(static_frame, text="Checked (Yes)", variable=value_var,
                                  value="checked", font=("Segoe UI", 10), bg="#f0f0f0", fg="#000000").pack(side="left", padx=10)
                    tk.Radiobutton(static_frame, text="Unchecked (Off)", variable=value_var,
                                  value="unchecked", font=("Segoe UI", 10), bg="#f0f0f0", fg="#000000").pack(side="left", padx=10)
                    value_widget = static_frame
                    # Ensure value is set to checked/unchecked
                    if value_var.get() not in ["checked", "unchecked"]:
                        value_var.set("checked" if current_config.get("enabled") else "unchecked")
            
            update_checkbox_value_widget()
            source_dropdown.bind("<<ComboboxSelected>>", lambda e: update_checkbox_value_widget())
            
        else:
            # Text field or radio button
            # Data source
            source_frame = tk.Frame(main_frame, bg="#f0f0f0")
            source_frame.pack(fill="x", pady=(0, 15))
            
            tk.Label(source_frame, text="Data Source:", font=("Segoe UI", 10),
                    bg="#f0f0f0").pack(side="left", padx=(0, 10))
            
            source_var = tk.StringVar(value=current_config.get("type", "static"))
            source_dropdown = ttk.Combobox(source_frame, textvariable=source_var,
                                          values=["excel", "static"], state="readonly", width=15)
            source_dropdown.pack(side="left")
            
            # Value
            value_frame = tk.Frame(main_frame, bg="#f0f0f0")
            value_frame.pack(fill="x", pady=(0, 15))
            
            tk.Label(value_frame, text="Value:", font=("Segoe UI", 10),
                    bg="#f0f0f0").pack(side="left", padx=(0, 10))
            
            value_var = tk.StringVar(value=default_checkbox_value)
            value_widget = None
            
            def update_value_widget():
                nonlocal value_widget
                if value_widget:
                    value_widget.destroy()
                
                if source_var.get() == "excel":
                    value_widget = ttk.Combobox(value_frame, textvariable=value_var,
                                              values=self.excel_columns, width=25, state="readonly")
                else:
                    value_widget = tk.Entry(value_frame, textvariable=value_var, width=30)
                
                value_widget.pack(side="left")
            
            update_value_widget()
            source_dropdown.bind("<<ComboboxSelected>>", lambda e: update_value_widget())
        
        # Buttons
        buttons_frame = tk.Frame(main_frame, bg="#f0f0f0")
        buttons_frame.pack(fill="x", pady=(20, 0))
        
        def save_field_config():
            enabled = enable_var.get()
            source_type = source_var.get()
            value = value_var.get().strip()
            
            if enabled and not value:
                messagebox.showerror("Error", "Please provide a value for this field.")
                return
            
            # Update coordinates if manually edited
            try:
                new_x0 = float(x0_var.get())
                new_y0 = float(y0_var.get())
                new_x1 = float(x1_var.get())
                new_y1 = float(y1_var.get())
                
                # Validate coordinates
                if new_x1 < new_x0:
                    messagebox.showerror("Error", "X1 must be greater than X0.")
                    return
                if new_y1 < new_y0:
                    messagebox.showerror("Error", "Y1 must be greater than Y0.")
                    return
                
                # Update field rectangle
                class RectLike:
                    def __init__(self, x0, y0, x1, y1):
                        self.x0 = x0
                        self.y0 = y0
                        self.x1 = x1
                        self.y1 = y1
                
                if field_name in self.field_data:
                    self.field_data[field_name]["rect"] = RectLike(new_x0, new_y0, new_x1, new_y1)
                    self.log_func(f"Updated coordinates for '{field_name}': ({new_x0:.2f}, {new_y0:.2f}) to ({new_x1:.2f}, {new_y1:.2f})", level="INFO")
            except ValueError:
                messagebox.showerror("Error", "Invalid coordinate values. Please enter valid numbers.")
                return
            
            # Update configuration
            # For checkboxes with static values, store "checked" or "unchecked"
            # For checkboxes with Excel values, store the Excel column name
            # The PDF filling logic will handle the conversion to "Yes"/"Off"
            stored_value = value
            
            # Update configuration
            self.field_config[field_name] = {
                "type": source_type,
                "value": stored_value,
                "enabled": enabled,
                "field_type": field_type
            }
            
            # Refresh display
            self._display_page(self.current_page)
            
            # Update status
            self.status_label.config(text=f"Field '{field_name}' configured successfully.")
            
            config_dialog.destroy()
        
        def disable_field():
            self.field_config[field_name] = {
                "type": "static",
                "value": "",
                "enabled": False,
                "field_type": field_type
            }
            self._display_page(self.current_page)
            self.status_label.config(text=f"Field '{field_name}' disabled.")
            config_dialog.destroy()
        
        tk.Button(buttons_frame, text="Save",
                 command=save_field_config,
                 bg="#28a745", fg="white", font=("Segoe UI", 10, "bold"),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="left", padx=5)
        
        tk.Button(buttons_frame, text="Disable",
                 command=disable_field,
                 bg="#dc3545", fg="white", font=("Segoe UI", 10),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="left", padx=5)
        
        tk.Button(buttons_frame, text="Cancel",
                 command=config_dialog.destroy,
                 bg="#6c757d", fg="white", font=("Segoe UI", 10),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="right", padx=5)
    
    def _prev_page(self) -> None:
        """Go to previous page."""
        if self.current_page > 0:
            self._display_page(self.current_page - 1)
    
    def _next_page(self) -> None:
        """Go to next page."""
        if self.current_page < len(self.pdf_images) - 1:
            self._display_page(self.current_page + 1)
    
    def _set_zoom(self, factor: float) -> None:
        """Set zoom factor."""
        self.zoom_factor *= factor
        self.zoom_factor = max(0.5, min(3.0, self.zoom_factor))  # Limit zoom
        self._display_page(self.current_page)
    
    def _save_selectors(self) -> None:
        """Save field selectors to configuration file."""
        try:
            # Ensure field_type is included for all configured fields
            complete_config = {}
            for field_name, config in self.field_config.items():
                # Get field_type from field_data if not already in config
                field_type = config.get("field_type")
                if not field_type and field_name in self.field_data:
                    field_type = self.field_data[field_name].get("field_type", "text")
                
                # Create complete config entry
                complete_config[field_name] = {
                    "type": config.get("type", "static"),
                    "value": config.get("value", ""),
                    "enabled": config.get("enabled", False),
                    "field_type": field_type or "text"
                }
            
            self.on_config_changed(complete_config)
            messagebox.showinfo("Success", 
                              "Field selectors saved successfully!\n"
                              "Your configuration will be loaded automatically when you reopen the bot.")
            self.status_label.config(text="Field selectors saved successfully!")
            self.log_func("Field selectors saved from visual mapper.", level="INFO")
        except Exception as e:
            messagebox.showerror("Error", f"Error saving selectors:\n{e}")
            self.log_func(f"Error saving selectors: {e}", level="ERROR")
    
    def _add_field_manually(self) -> None:
        """Add a field manually by entering its name, type, and clicking to position it."""
        # Create a dialog to get field name and type
        field_dialog = tk.Toplevel(self.window)
        field_dialog.title("Add Field Manually - Version 3.1.0, Last Updated 12/04/2025")
        field_dialog.geometry("450x250")
        field_dialog.transient(self.window)
        field_dialog.grab_set()
        field_dialog.configure(bg="#f0f0f0")
        
        main_frame = tk.Frame(field_dialog, bg="#f0f0f0")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Field name
        tk.Label(main_frame, text="Field Name:",
                font=("Segoe UI", 10, "bold"), bg="#f0f0f0", fg="#003366").pack(anchor="w", pady=(0, 5))
        field_name_var = tk.StringVar()
        field_name_entry = tk.Entry(main_frame, textvariable=field_name_var, width=40, font=("Segoe UI", 10))
        field_name_entry.pack(fill="x", pady=(0, 15))
        field_name_entry.focus()
        
        # Field type
        tk.Label(main_frame, text="Field Type:",
                font=("Segoe UI", 10, "bold"), bg="#f0f0f0", fg="#003366").pack(anchor="w", pady=(0, 5))
        field_type_var = tk.StringVar(value="text")
        type_frame = tk.Frame(main_frame, bg="#f0f0f0")
        type_frame.pack(fill="x", pady=(0, 15))
        
        tk.Radiobutton(type_frame, text="Text Field", variable=field_type_var, value="text",
                      font=("Segoe UI", 10), bg="#f0f0f0", fg="#000000").pack(side="left", padx=10)
        tk.Radiobutton(type_frame, text="Checkbox", variable=field_type_var, value="checkbox",
                      font=("Segoe UI", 10), bg="#f0f0f0", fg="#000000").pack(side="left", padx=10)
        tk.Radiobutton(type_frame, text="Radio Button", variable=field_type_var, value="radio",
                      font=("Segoe UI", 10), bg="#f0f0f0", fg="#000000").pack(side="left", padx=10)
        
        # Instructions
        instructions = tk.Label(main_frame, 
                               text="After clicking 'Add', click on the PDF where this field should be positioned.",
                               font=("Segoe UI", 9, "italic"), bg="#f0f0f0", fg="#666666",
                               wraplength=400, justify="left")
        instructions.pack(pady=(0, 15))
        
        # Buttons
        buttons_frame = tk.Frame(main_frame, bg="#f0f0f0")
        buttons_frame.pack(fill="x", pady=(10, 0))
        
        def add_field():
            field_name = field_name_var.get().strip()
            field_type = field_type_var.get()
            
            if not field_name:
                messagebox.showerror("Error", "Please enter a field name.")
                return
            
            field_dialog.destroy()
            
            # Check if field already exists
            if field_name in self.field_data:
                if messagebox.askyesno("Field Exists", 
                                      f"Field '{field_name}' already exists.\n"
                                      "Do you want to reposition it by clicking on the PDF?"):
                    self.status_label.config(text=f"Click on the PDF where '{field_name}' should be positioned.")
                    self._waiting_for_manual_position = field_name
                    self._waiting_field_type = field_type
                    self.canvas.config(cursor="crosshair")
                return
            
            # Add field and wait for user to click position
            self.status_label.config(text=f"Click on the PDF where '{field_name}' should be positioned.")
            self._waiting_for_manual_position = field_name
            self._waiting_field_type = field_type
            
            # Create a dummy rect for now
            class RectLike:
                def __init__(self):
                    self.x0 = 0
                    self.y0 = 0
                    self.x1 = 0
                    self.y1 = 0
            
            self.field_data[field_name] = {
                "rect": RectLike(),
                "page": self.current_page,
                "field_type": field_type,
                "field_value": None,
                "needs_positioning": True
            }
            
            self.canvas.config(cursor="crosshair")
            self.log_func(f"Waiting for manual positioning of {field_type} field: {field_name}", level="INFO")
        
        tk.Button(buttons_frame, text="Add",
                 command=add_field,
                 bg="#28a745", fg="white", font=("Segoe UI", 10, "bold"),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="left", padx=5)
        
        tk.Button(buttons_frame, text="Cancel",
                 command=field_dialog.destroy,
                 bg="#6c757d", fg="white", font=("Segoe UI", 10),
                 padx=20, pady=8, cursor="hand2", relief="flat").pack(side="right", padx=5)
        
        # Bind Enter key to add field
        field_name_entry.bind("<Return>", lambda e: add_field())
    
    def _delete_field_selector(self) -> None:
        """Delete a field selector by clicking on it."""
        self.status_label.config(text="Click on a field selector to delete it, or press Escape to cancel.")
        self._waiting_for_delete = True
        self.canvas.config(cursor="pirate")
        self.log_func("Waiting for field selector to delete...", level="INFO")
    
    def _toggle_select_mode(self) -> None:
        """Toggle between select mode and edit mode."""
        self._select_mode = not self._select_mode
        
        if self._select_mode:
            self.select_mode_btn.config(text="Select Mode (On)", bg="#28a745")
            self.status_label.config(text="Select Mode: Click on fields to select them for dragging/resizing")
            self.log_func("Select Mode enabled - Click on fields to select them for dragging/resizing", level="INFO")
        else:
            self.select_mode_btn.config(text="Select Mode (Off)", bg="#6c757d")
            self.status_label.config(text="Edit Mode: Click on fields to edit them")
            # Clear selection when exiting select mode
            if self._selected_field:
                self._selected_field = None
                self._display_page(self.current_page)
            self.log_func("Edit Mode enabled - Click on fields to edit them", level="INFO")
    
    def _on_canvas_right_click(self, event) -> None:
        """Handle right-click on canvas (context menu)."""
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        
        # Find clicked field
        clicked_field = None
        for field_name in self.field_data.keys():
            coords = self._get_field_coords(field_name, self.current_page)
            if not coords:
                continue
            
            x0, y0, x1, y1 = coords
            
            # Check if click is within field rectangle
            if x0 <= x <= x1 and y0 <= y <= y1:
                clicked_field = field_name
                break
        
        if clicked_field:
            # Right-click opens edit dialog
            self._open_edit_selector_dialog(clicked_field)
    
    def _delete_all_selectors(self) -> None:
        """Delete all field selectors (remove yellow boxes from screen)."""
        # Count fields that are visible but not configured (yellow boxes)
        unconfigured_fields = []
        for field_name in list(self.field_data.keys()):  # Create a copy to iterate safely
            # Only count fields that are not configured (yellow boxes)
            if field_name not in self.field_config or not self.field_config[field_name].get("enabled", False):
                unconfigured_fields.append(field_name)
        
        if not unconfigured_fields:
            messagebox.showinfo("No Selectors", "No unconfigured field selectors to remove.")
            return
        
        count = len(unconfigured_fields)
        if messagebox.askyesno("Delete All Selectors", 
                              f"Remove all {count} unconfigured field selector(s) from screen?\n\n"
                              "This will remove the yellow boxes from the PDF view.\n"
                              "Configured fields (green boxes) will remain."):
            # Remove unconfigured fields from field_data (they disappear from screen)
            for field_name in unconfigured_fields:
                if field_name in self.field_data:
                    del self.field_data[field_name]
                # Also remove from config if present
                if field_name in self.field_config:
                    del self.field_config[field_name]
            
            # Clear selection if selected field was removed
            if self._selected_field and self._selected_field not in self.field_data:
                self._selected_field = None
            
            # Redraw page
            self._display_page(self.current_page)
            self.status_label.config(text=f"Removed {count} unconfigured field selector(s) from screen.")
            self.log_func(f"Removed {count} unconfigured field selector(s) from screen.", level="INFO")
    
    def _clear_all_fields(self) -> None:
        """Clear all field configurations."""
        if messagebox.askyesno("Clear All", 
                               "Clear all field configurations?\n"
                               "This will disable all fields and leave them blank."):
            self.field_config = {}
            self._selected_field = None
            self._display_page(self.current_page)
            self.status_label.config(text="All field configurations cleared.")
            self.log_func("All field configurations cleared.", level="INFO")
    
    def _close_window(self) -> None:
        """Close the visual mapper window."""
        # Ask if user wants to save before closing
        if messagebox.askyesno("Close Visual Mapper", 
                               "Save field selectors before closing?\n"
                               "(Selectors are automatically saved when you click 'Save Selectors')"):
            self._save_selectors()
        
        # Close PDF document
        if self.pdf_doc:
            try:
                self.pdf_doc.close()
            except Exception:
                pass
            self.pdf_doc = None
        
        self.window.destroy()


# -----------------------------------------------------------------------------
# Entry point
# -----------------------------------------------------------------------------
def main() -> None:
    bot = MedicareRefilingBot()
    bot.create_main_window()
    if bot.root:
        bot.root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nMedicare Refiling Bot interrupted by user.")

