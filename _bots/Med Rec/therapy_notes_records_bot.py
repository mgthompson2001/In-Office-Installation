#!/usr/bin/env python3
"""
TherapyNotes Records Bot
========================

High-level workflow (stubs in place for navigation):
1. User loads an insurer document (PDF/image/Excel) and specifies the desired manual date range.
2. DocumentParser extracts client metadata (name, DOB) using OCR or direct parsing.
3. TherapyNotesClientFetcher logs in to TherapyNotes and retrieves billable notes.
4. Dates of service are filtered against the manual range and exported to CSV/Excel.

Navigation-specific logic will be added later once the TherapyNotes HTML element map is available.
"""

from __future__ import annotations

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext, simpledialog

import logging
import threading
import time
import os
import json
import csv
from pathlib import Path
from datetime import datetime, date
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any, Tuple

# Optional dependencies -------------------------------------------------------
# (Mirrors medisoft_billing_bot.py so we can reuse OCR/document parsing stack.)

try:
    import pdfplumber  # type: ignore
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    pdfplumber = None  # type: ignore
    PDFPLUMBER_AVAILABLE = False

try:
    import pytesseract  # type: ignore
    from PIL import Image  # type: ignore
    OCR_AVAILABLE = True
except ImportError:
    pytesseract = None  # type: ignore
    Image = None  # type: ignore
    OCR_AVAILABLE = False

try:
    import pandas as pd  # type: ignore
    EXCEL_AVAILABLE = True
    try:
        import openpyxl  # type: ignore
        OPENPYXL_AVAILABLE = True
    except ImportError:
        OPENPYXL_AVAILABLE = False
    try:
        import pyxlsb  # type: ignore
        PYXLSB_AVAILABLE = True
    except ImportError:
        PYXLSB_AVAILABLE = False
    try:
        import xlrd  # type: ignore
        XLRD_AVAILABLE = True
    except ImportError:
        XLRD_AVAILABLE = False
except ImportError:
    pd = None  # type: ignore
    EXCEL_AVAILABLE = False
    OPENPYXL_AVAILABLE = False
    PYXLSB_AVAILABLE = False
    XLRD_AVAILABLE = False

try:
    from selenium import webdriver  # type: ignore
    from selenium.webdriver.chrome.options import Options as ChromeOptions  # type: ignore
    from selenium.webdriver.common.by import By  # type: ignore
    from selenium.webdriver.support.ui import WebDriverWait  # type: ignore
    from selenium.webdriver.support import expected_conditions as EC  # type: ignore
    from selenium.common.exceptions import TimeoutException, StaleElementReferenceException  # type: ignore
    from selenium.webdriver.common.keys import Keys  # type: ignore
    SELENIUM_AVAILABLE = True
except ImportError:
    webdriver = None  # type: ignore
    ChromeOptions = None  # type: ignore
    By = None  # type: ignore
    WebDriverWait = None  # type: ignore
    EC = None  # type: ignore
    TimeoutException = None  # type: ignore
    StaleElementReferenceException = None  # type: ignore
    Keys = None  # type: ignore
    SELENIUM_AVAILABLE = False

# Logging ---------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("therapy_notes_records_bot.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def _configure_ocr_paths() -> None:
    """Detect and configure Tesseract and Poppler paths on any Windows machine.
    
    Precedence (highest to lowest):
    1) Environment variables: TESSERACT_PATH, POPPLER_PATH
    2) Local vendor folder next to this script: vendor/Tesseract-OCR, vendor/poppler/Library/bin
    3) Common install locations (Program Files, LocalAppData, Conda)
    """
    if not (OCR_AVAILABLE and pytesseract):
        return
    try:
        script_dir = Path(__file__).parent
        # 1) Env vars
        tesseract_env = os.environ.get("TESSERACT_PATH")
        poppler_env = os.environ.get("POPPLER_PATH")
        if tesseract_env and Path(tesseract_env).exists():
            pytesseract.pytesseract.tesseract_cmd = tesseract_env
            logger.info("Configured Tesseract from env: %s", tesseract_env)
        if poppler_env and Path(poppler_env).exists():
            os.environ["POPPLER_PATH"] = poppler_env
            logger.info("Configured Poppler from env: %s", poppler_env)
        # 2) Vendor folder
        if not getattr(pytesseract.pytesseract, "tesseract_cmd", None):
            vend_tess = script_dir / "vendor" / "Tesseract-OCR" / "tesseract.exe"
            if vend_tess.exists():
                pytesseract.pytesseract.tesseract_cmd = str(vend_tess)
                logger.info("Configured Tesseract from vendor: %s", vend_tess)
        if "POPPLER_PATH" not in os.environ:
            vend_poppler = script_dir / "vendor" / "poppler" / "Library" / "bin"
            if vend_poppler.exists():
                os.environ["POPPLER_PATH"] = str(vend_poppler)
                logger.info("Configured Poppler from vendor: %s", vend_poppler)
        # 3) Common locations
        current_cmd = getattr(pytesseract.pytesseract, "tesseract_cmd", None)
        if not current_cmd or not Path(current_cmd).exists() if current_cmd else True:
            candidates = [
                Path("C:/Program Files/Tesseract-OCR/tesseract.exe"),
                Path("C:/Program Files (x86)/Tesseract-OCR/tesseract.exe"),
                Path.home() / "AppData/Local/Programs/Tesseract-OCR/tesseract.exe",
            ]
            conda_prefix = os.environ.get("CONDA_PREFIX")
            if conda_prefix:
                candidates.append(Path(conda_prefix) / "Library/bin/tesseract.exe")
            for c in candidates:
                if c.exists():
                    pytesseract.pytesseract.tesseract_cmd = str(c)
                    logger.info("Configured Tesseract from common path: %s", c)
                    break
            # Final check: verify tesseract is callable
            try:
                pytesseract.get_tesseract_version()
                logger.info("Tesseract verified successfully")
            except Exception as exc:  # noqa: BLE001
                logger.warning("Tesseract path configured but not working: %s", exc)
        if "POPPLER_PATH" not in os.environ:
            poppler_candidates = [
                Path("C:/Program Files/poppler/Library/bin"),
                Path("C:/Program Files (x86)/poppler/Library/bin"),
                Path.home() / "AppData/Local/poppler/Library/bin",
            ]
            conda_prefix = os.environ.get("CONDA_PREFIX")
            if conda_prefix:
                poppler_candidates.append(Path(conda_prefix) / "Library/bin")
            for d in poppler_candidates:
                if d.exists() and (d / "pdftoppm.exe").exists():
                    os.environ["POPPLER_PATH"] = str(d)
                    logger.info("Configured Poppler from common path: %s", d)
                    break
    except Exception as exc:  # noqa: BLE001
        logger.warning("OCR path auto-config failed: %s", exc)


_configure_ocr_paths()


# Data structures -------------------------------------------------------------

@dataclass
class ClientMetadata:
    name: str = ""
    dob: str = ""
    raw_text: str = ""
    source_file: Optional[Path] = None
    extras: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ServiceEntry:
    client_name: str
    client_dob: str
    service_date: datetime
    service_type: str
    notes: str = ""


# Document parsing ------------------------------------------------------------

class DocumentParser:
    """Parses insurer documents to extract client metadata."""

    def __init__(self) -> None:
        self.ocr_available = OCR_AVAILABLE and pytesseract is not None
        self.pdf_available = PDFPLUMBER_AVAILABLE and pdfplumber is not None
        self.excel_available = EXCEL_AVAILABLE and pd is not None

    def parse_document(self, file_path: Path) -> ClientMetadata:
        logger.info("Parsing document: %s", file_path)
        suffix = file_path.suffix.lower()
        if suffix in {".xlsx", ".xls", ".csv"}:
            return self._parse_tabular(file_path)
        if suffix in {".pdf"}:
            return self._parse_pdf(file_path)
        # Assume image or other OCR-eligible format
        return self._parse_image(file_path)

    def _parse_tabular(self, file_path: Path) -> ClientMetadata:
        if not self.excel_available:
            raise RuntimeError("pandas is required to parse Excel/CSV documents.")
        logger.info("Attempting to parse Excel/CSV for client info.")
        suffix = file_path.suffix.lower()
        df: Optional[pd.DataFrame] = None
        try:
            if suffix in {".xls", ".xlsx"}:
                df = pd.read_excel(file_path, engine="openpyxl")
            elif suffix == ".csv":
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)
        except Exception as exc:
            logger.warning("Standard Excel parsing failed (%s). Attempting fallback.", exc)
            df = self._fallback_load_workbook(file_path, suffix)

        if df is None or df.empty:
            text_blob = file_path.read_text(encoding="utf-8", errors="ignore")
            name = self._search_for_name(text_blob)
            dob = self._search_for_dob(text_blob)
            date_range = self._search_for_date_range(text_blob)
            extras: Dict[str, Any] = {}
            if date_range[0] or date_range[1]:
                extras["date_range"] = {"start": date_range[0], "end": date_range[1]}
            return ClientMetadata(name=name, dob=dob, raw_text=text_blob, source_file=file_path, extras=extras)

        text_blob = "\n".join(df.astype(str).fillna("").values.flatten())
        records, agg_name, agg_dob, agg_range = self._extract_tabular_records(df)
        name = agg_name or self._search_for_name(text_blob)
        # Only use text search for DOB if we didn't find it in tabular records
        # and avoid the fallback that picks first standalone date
        if agg_dob:
            dob = agg_dob
            logger.info(f"Found DOB from tabular records: {dob}")
            # Also log to console for visibility
            print(f"INFO: Found DOB from tabular records: {dob}")
        else:
            # Try text search but be more careful - don't use first standalone date fallback
            dob = self._search_for_dob_strict(text_blob)
            if dob:
                logger.info(f"Found DOB from text search: {dob}")
                print(f"INFO: Found DOB from text search: {dob}")
            else:
                logger.warning("Could not find DOB in document")
                print("WARNING: Could not find DOB in document")
                dob = ""
        date_range = agg_range or self._search_for_date_range(text_blob)

        extras: Dict[str, Any] = {"records": records}
        if date_range[0] or date_range[1]:
            extras["date_range"] = {"start": date_range[0], "end": date_range[1]}

        return ClientMetadata(name=name, dob=dob, raw_text=text_blob, source_file=file_path, extras=extras)

    def _fallback_load_workbook(self, file_path: Path, suffix: str) -> Optional[pd.DataFrame]:
        if pd is None:
            return None
        frames: List[pd.DataFrame] = []

        if PYXLSB_AVAILABLE:
            try:
                frames.append(pd.read_excel(file_path, engine="pyxlsb"))
            except Exception as exc:
                logger.warning("pyxlsb engine failed (%s).", exc)

        if XLRD_AVAILABLE:
            try:
                frames.append(pd.read_excel(file_path, engine="xlrd"))
            except Exception as exc:
                logger.warning("xlrd engine failed (%s).", exc)

        if OPENPYXL_AVAILABLE:
            try:
                import openpyxl  # type: ignore

                wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
                for sheet in wb.worksheets:
                    rows = [
                        [self._format_cell_value(cell) for cell in row]
                        for row in sheet.iter_rows(values_only=True)
                    ]
                    frame = self._rows_to_dataframe(rows)
                    if frame is not None and not frame.empty:
                        frames.append(frame)
            except Exception as exc:
                logger.warning("openpyxl manual load failed (%s).", exc)

        if frames:
            try:
                return pd.concat(frames, ignore_index=True)
            except Exception as exc:
                logger.warning("Concatenating fallback frames failed (%s).", exc)

        logger.warning("All workbook fallbacks failed; returning None.")
        return None

    @staticmethod
    def _format_cell_value(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%m/%d/%Y")
        from datetime import date

        if isinstance(value, date):
            return value.strftime("%m/%d/%Y")
        return str(value)

    @staticmethod
    def _rows_to_dataframe(rows: List[List[str]]) -> Optional[pd.DataFrame]:
        if pd is None or not rows:
            return None
        header: Optional[List[str]] = None
        data: List[List[str]] = []
        for row in rows:
            if not any(cell.strip() for cell in row):
                continue
            if header is None:
                header = [
                    cell.strip() if cell.strip() else f"Column {idx + 1}"
                    for idx, cell in enumerate(row)
                ]
            else:
                # Ensure row matches header length
                normalized = row + [""] * (len(header) - len(row))
                data.append([cell.strip() for cell in normalized[: len(header)]])
        if header is None:
            return None
        if not data:
            return pd.DataFrame(columns=header)
        return pd.DataFrame(data, columns=header)

    def _extract_tabular_records(
        self, df: pd.DataFrame
    ) -> Tuple[List[Dict[str, Any]], Optional[str], Optional[str], Tuple[Optional[str], Optional[str]]]:
        records: List[Dict[str, Any]] = []
        if df is None or df.empty:
            return records, None, None, (None, None)

        normalized_columns = {str(col).strip().lower(): col for col in df.columns}
        logger.debug(f"Available columns: {list(df.columns)}")

        def _find_column(keywords: List[str]) -> Optional[str]:
            for key, original in normalized_columns.items():
                if any(keyword in key for keyword in keywords):
                    logger.debug(f"Found column '{original}' matching keywords: {keywords}")
                    return original
            logger.debug(f"No column found matching keywords: {keywords}")
            return None

        name_col = _find_column(["patient name", "member name", "client name", "insured name"])
        # Try to find DOB column - be very explicit about "Patient Date of Birth"
        dob_col = _find_column(["patient date of birth", "date of birth", "dob", "birth date", "birthdate", "d.o.b", "d.o.b."])
        if not dob_col:
            # Fallback: try to find any column with "birth" in it
            for col in df.columns:
                col_lower = str(col).strip().lower()
                if "birth" in col_lower and "date" in col_lower:
                    dob_col = col
                    logger.info(f"Found DOB column via fallback: '{dob_col}'")
                    print(f"INFO: Found DOB column via fallback: '{dob_col}'")
                    break
        dos_col = _find_column(["dates of service", "date of service", "dos"])
        
        if not dob_col:
            logger.warning(f"DOB column not found. Available columns: {list(df.columns)}")
            print(f"WARNING: DOB column not found. Available columns: {list(df.columns)}")
        else:
            logger.info(f"Using DOB column: '{dob_col}'")
            print(f"INFO: Using DOB column: '{dob_col}'")

        overall_dates: List[datetime] = []
        has_present = False

        for _, row in df.iterrows():
            record: Dict[str, Any] = {}
            if name_col:
                name_val = str(row.get(name_col, "")).strip()
                if name_val and name_val.lower() not in {"nan", "none"}:
                    record["name"] = name_val
            if dob_col:
                dob_value = row.get(dob_col)
                # Handle pandas Timestamp objects directly
                if pd and isinstance(dob_value, pd.Timestamp):
                    normalized_dob = dob_value.strftime("%m/%d/%Y")
                    record["dob"] = normalized_dob
                    record["dob_raw"] = str(dob_value)
                    logger.info(f"Extracted DOB from Timestamp for {record.get('name', 'unknown')}: {normalized_dob}")
                    print(f"INFO: Extracted DOB from Timestamp for {record.get('name', 'unknown')}: {normalized_dob}")
                elif isinstance(dob_value, (datetime, date)):
                    # Handle Python datetime/date objects
                    normalized_dob = dob_value.strftime("%m/%d/%Y")
                    record["dob"] = normalized_dob
                    record["dob_raw"] = str(dob_value)
                    logger.info(f"Extracted DOB from datetime for {record.get('name', 'unknown')}: {normalized_dob}")
                    print(f"INFO: Extracted DOB from datetime for {record.get('name', 'unknown')}: {normalized_dob}")
                else:
                    # Handle string values
                    dob_raw = str(dob_value).strip() if dob_value is not None else ""
                    if dob_raw and dob_raw.lower() not in {"nan", "none", ""}:
                        normalized_dob = self._normalize_date(dob_raw) if dob_raw else ""
                        if normalized_dob:
                            record["dob"] = normalized_dob
                            record["dob_raw"] = dob_raw
                            logger.info(f"Extracted DOB for {record.get('name', 'unknown')}: raw='{dob_raw}' -> normalized='{normalized_dob}'")
                            print(f"INFO: Extracted DOB for {record.get('name', 'unknown')}: raw='{dob_raw}' -> normalized='{normalized_dob}'")
                        else:
                            logger.warning(f"Could not normalize DOB value '{dob_raw}' for {record.get('name', 'unknown')}")
                    elif dob_raw:
                        logger.debug(f"DOB value '{dob_raw}' was skipped (empty/nan/none)")
            if dos_col:
                dos_raw = str(row.get(dos_col, "")).strip()
                if dos_raw and dos_raw.lower() not in {"nan", "none"}:
                    record["dos_raw"] = dos_raw
                    dates, present_flag = self._parse_dos_string(dos_raw)
                    if dates:
                        record["dos_dates"] = [dt.strftime("%m/%d/%Y") for dt in dates]
                        overall_dates.extend(dates)
                    if present_flag:
                        record["dos_present"] = True
                        has_present = True
            # Include other useful columns for context
            for key_phrase in ["member id", "patient id", "chart id", "requester", "status"]:
                col = _find_column([key_phrase])
                if col:
                    val = str(row.get(col, "")).strip()
                    if val and val.lower() not in {"nan", "none"}:
                        record[col] = val
            if record:
                records.append(record)

        agg_name = next((r.get("name") for r in records if r.get("name")), None)
        agg_dob = next((r.get("dob") for r in records if r.get("dob")), None)

        start_str: Optional[str] = None
        end_str: Optional[str] = None
        if overall_dates:
            start_str = min(overall_dates).strftime("%m/%d/%Y")
            end_str = max(overall_dates).strftime("%m/%d/%Y")
        if has_present:
            end_str = "Present"

        return records, agg_name, agg_dob, (start_str, end_str)

    def _parse_dos_string(self, dos_raw: str) -> Tuple[List[datetime], bool]:
        import re

        if not dos_raw:
            return [], False
        text = dos_raw.strip()
        present_flag = any(word in text.lower() for word in ["present", "current", "today"])

        date_tokens: List[datetime] = []
        range_patterns = [
            r"([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{2,4})\s*(?:to|-|through|and)\s*([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{2,4})",
            r"([A-Za-z]{3,9}\s+\d{1,2}(?:st|nd|rd|th)?[,]?\s+\d{2,4})\s*(?:to|-|through|and)\s*([A-Za-z]{3,9}\s+\d{1,2}(?:st|nd|rd|th)?[,]?\s+\d{2,4})",
        ]
        for pattern in range_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                start_norm = self._normalize_date(match.group(1))
                end_norm = self._normalize_date(match.group(2))
                if start_norm:
                    dt = self._to_datetime(start_norm)
                    if dt:
                        date_tokens.append(dt)
                if end_norm:
                    dt = self._to_datetime(end_norm)
                    if dt:
                        date_tokens.append(dt)

        # Split on delimiters for individual dates
        for chunk in re.split(r"[;,|]", text):
            chunk = chunk.strip()
            if not chunk:
                continue
            normalized = self._normalize_date(chunk)
            if normalized:
                dt = self._to_datetime(normalized)
                if dt:
                    date_tokens.append(dt)

        # Deduplicate dates
        unique_dates = []
        seen = set()
        for dt in date_tokens:
            key = dt.strftime("%Y-%m-%d")
            if key not in seen:
                seen.add(key)
                unique_dates.append(dt)

        return unique_dates, present_flag

    @staticmethod
    def _to_datetime(date_str: str) -> Optional[datetime]:
        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        return None

    def _parse_pdf(self, file_path: Path) -> ClientMetadata:
        text_parts: List[str] = []
        tables_found: List[List[List[Optional[str]]]] = []
        is_scanned = False
        
        if self.pdf_available:
            try:
                with pdfplumber.open(file_path) as pdf:
                    total_text_length = 0
                    for page in pdf.pages:
                        text = page.extract_text() or ""
                        text_parts.append(text)
                        total_text_length += len(text)
                        # Also try to extract tables
                        page_tables = page.extract_tables()
                        if page_tables:
                            tables_found.extend(page_tables)
                    
                    # Detect if PDF is scanned: if very little or no text extracted, it's likely scanned
                    is_scanned = total_text_length < 200
                    if is_scanned:
                        logger.info("PDF appears to be scanned (minimal text extracted), will use OCR")
            except Exception as exc:  # noqa: BLE001
                logger.warning("pdfplumber failed (%s). Falling back to OCR if available.", exc)
                is_scanned = True
        
        text_blob = "\n".join(text_parts).strip()
        
        # If scanned or no text found, use OCR
        if is_scanned or not text_blob or len(text_blob) < 100:
            if not self.ocr_available:
                raise RuntimeError("PDF text extraction failed and OCR is unavailable.")
            logger.info("Using OCR to extract text from scanned PDF")
            text_blob = self._ocr_pdf(file_path)
            is_scanned = True
        
        # Try to extract from tables first (more reliable for structured data)
        # For scanned PDFs, try OCR-based table extraction on the last page (where client data usually is)
        if is_scanned and not tables_found and self.pdf_available:
            try:
                with pdfplumber.open(file_path) as pdf:
                    # Try to extract tables from last page (where client list usually is)
                    if len(pdf.pages) > 0:
                        last_page = pdf.pages[-1]
                        # Try different table extraction strategies
                        # Strategy 1: Default table extraction
                        page_tables = last_page.extract_tables()
                        if page_tables:
                            tables_found.extend(page_tables)
                        else:
                            # Strategy 2: Try with explicit table settings
                            try:
                                page_tables = last_page.extract_tables(
                                    table_settings={
                                        "vertical_strategy": "lines_strict",
                                        "horizontal_strategy": "lines_strict",
                                    }
                                )
                                if page_tables:
                                    tables_found.extend(page_tables)
                            except Exception:
                                pass
            except Exception:
                pass
        
        if tables_found:
            multi_client_records = self._extract_clients_from_tables(tables_found)
            if multi_client_records:
                logger.info(f"Extracted {len(multi_client_records)} clients from PDF tables")
                name = multi_client_records[0].get("name", "")
                dob = multi_client_records[0].get("dob", "")
                extras: Dict[str, Any] = {"records": multi_client_records}
                return ClientMetadata(name=name, dob=dob, raw_text=text_blob, source_file=file_path, extras=extras)
        
        # Try to extract multiple clients with per-client date ranges from text
        # For scanned PDFs, use improved OCR text parsing
        if is_scanned:
            multi_client_records = self._extract_clients_from_ocr_text(text_blob)
            # If OCR extraction failed, try a more lenient approach
            if not multi_client_records:
                logger.warning("Standard OCR extraction found no clients. Trying lenient extraction for garbled OCR...")
                multi_client_records = self._extract_clients_lenient_ocr(text_blob)
        else:
            multi_client_records = self._extract_multiple_clients_from_pdf(text_blob)
        
        if multi_client_records:
            # Multiple clients found with individual date ranges
            logger.info(f"Extracted {len(multi_client_records)} clients with per-client date ranges from PDF text")
            name = multi_client_records[0].get("name", "")
            dob = multi_client_records[0].get("dob", "")
            extras: Dict[str, Any] = {"records": multi_client_records}
            return ClientMetadata(name=name, dob=dob, raw_text=text_blob, source_file=file_path, extras=extras)

        # Fallback to single client extraction
        name = self._search_for_name(text_blob)
        dob = self._search_for_dob(text_blob)
        date_range = self._search_for_date_range(text_blob)
        extras: Dict[str, Any] = {}
        if date_range[0] or date_range[1]:
            extras["date_range"] = {
                "start": date_range[0],
                "end": date_range[1],
            }
        return ClientMetadata(name=name, dob=dob, raw_text=text_blob, source_file=file_path, extras=extras)

    def _extract_clients_from_tables(self, tables: List[List[List[Optional[str]]]]) -> List[Dict[str, Any]]:
        """
        Extract client data from PDF tables.
        Looks for tables with columns containing: Name, DOB, Date Range
        """
        import re
        records: List[Dict[str, Any]] = []
        
        for table in tables:
            if not table or len(table) < 2:  # Need at least header + 1 data row
                continue
            
            # Try to identify column indices
            header_row = table[0]
            name_col_idx = None
            dob_col_idx = None
            dos_col_idx = None
            
            # Look for header row with column names
            for col_idx, cell in enumerate(header_row):
                if not cell:
                    continue
                cell_upper = str(cell).upper()
                if any(keyword in cell_upper for keyword in ["NAME", "PATIENT", "MEMBER", "CLIENT", "BENEFICIARY"]):
                    name_col_idx = col_idx
                elif any(keyword in cell_upper for keyword in ["DOB", "DATE OF BIRTH", "BIRTH"]):
                    dob_col_idx = col_idx
                elif any(keyword in cell_upper for keyword in ["DOS", "DATE OF SERVICE", "DATES", "SERVICE DATE", "DATE RANGE"]):
                    dos_col_idx = col_idx
            
            # If no clear headers, try to infer from first few rows
            if name_col_idx is None:
                # Look for columns that contain names (2-4 words, proper case)
                for col_idx in range(len(header_row)):
                    name_count = 0
                    for row in table[1:min(6, len(table))]:  # Check first 5 data rows
                        if col_idx < len(row) and row[col_idx]:
                            cell_text = str(row[col_idx]).strip()
                            words = cell_text.split()
                            if 2 <= len(words) <= 5 and re.match(r"^[A-Z][A-Za-z',.\- ]+$", cell_text):
                                name_count += 1
                    if name_count >= 3:  # If 3+ rows look like names
                        name_col_idx = col_idx
                        break
            
            # Extract data rows
            for row_idx, row in enumerate(table[1:], start=1):  # Skip header
                if not row or len(row) == 0:
                    continue
                
                name = None
                dob = None
                dos_range = None
                
                # Extract name
                if name_col_idx is not None and name_col_idx < len(row) and row[name_col_idx]:
                    name_candidate = str(row[name_col_idx]).strip()
                    # Validate it looks like a name
                    words = name_candidate.split()
                    if 2 <= len(words) <= 5 and len(name_candidate) <= 60:
                        # Exclude common non-name text
                        if not any(exclude in name_candidate.upper() for exclude in 
                                  ["SERVICES", "LLC", "HEALTHCARE", "SOCIAL WORK", "USPS", 
                                   "CERTIFIED", "MAIL", "PLEASE", "SUBMIT", "MOLINA", "TIOT"]):
                            name = name_candidate
                else:
                    # If no name column identified, try to find name in any column
                    # Look for columns that contain text that looks like names
                    for col_idx, cell in enumerate(row):
                        if cell and col_idx != dob_col_idx and col_idx != dos_col_idx:
                            name_candidate = str(cell).strip()
                            words = name_candidate.split()
                            # Name should be 2-5 words, mostly letters
                            if (2 <= len(words) <= 5 and len(name_candidate) <= 60 and
                                not any(exclude in name_candidate.upper() for exclude in 
                                       ["SERVICES", "LLC", "HEALTHCARE", "SOCIAL", "USPS"])):
                                # Check if it's mostly letters (not just numbers/dates)
                                if sum(c.isalpha() for c in name_candidate) > len(name_candidate) * 0.5:
                                    name = name_candidate
                                    if name_col_idx is None:
                                        name_col_idx = col_idx
                                    break
                
                # Extract DOB
                if dob_col_idx is not None and dob_col_idx < len(row) and row[dob_col_idx]:
                    dob_candidate = str(row[dob_col_idx]).strip()
                    dob_normalized = self._normalize_date(dob_candidate)
                    if dob_normalized:
                        dob = dob_normalized
                else:
                    # Try to find DOB in any column
                    for col_idx, cell in enumerate(row):
                        if cell and col_idx != name_col_idx:
                            dob_candidate = str(cell).strip()
                            dob_normalized = self._normalize_date(dob_candidate)
                            if dob_normalized:
                                dob = dob_normalized
                                dob_col_idx = col_idx
                                break
                
                # Extract date range
                if dos_col_idx is not None and dos_col_idx < len(row) and row[dos_col_idx]:
                    dos_candidate = str(row[dos_col_idx]).strip()
                    dos_range = self._search_for_date_range(dos_candidate)
                else:
                    # Try to find date range in any column
                    for col_idx, cell in enumerate(row):
                        if cell and col_idx != name_col_idx and col_idx != dob_col_idx:
                            dos_candidate = str(cell).strip()
                            dos_range = self._search_for_date_range(dos_candidate)
                            if dos_range[0] or dos_range[1]:
                                dos_col_idx = col_idx
                                break
                
                # Only create record if we have a name AND (DOB or date range)
                if name and (dob or (dos_range and (dos_range[0] or dos_range[1]))):
                    record: Dict[str, Any] = {"name": name}
                    if dob:
                        record["dob"] = dob
                        record["dob_raw"] = dob
                    if dos_range and (dos_range[0] or dos_range[1]):
                        record["date_range"] = {
                            "start": dos_range[0],
                            "end": dos_range[1],
                        }
                    records.append(record)
                    logger.info(f"Extracted client from table: {name}, DOB: {dob or 'N/A'}, Date Range: {dos_range[0] if dos_range else 'N/A'} to {dos_range[1] if dos_range else 'N/A'}")
        
        # Deduplicate
        deduped: List[Dict[str, Any]] = []
        seen: set[Tuple[str, str]] = set()
        for record in records:
            key = (record["name"].lower(), (record.get("dob") or record.get("dob_raw") or "").lower())
            if key not in seen:
                seen.add(key)
                deduped.append(record)
        
        return deduped if len(deduped) > 1 else []

    def _extract_clients_from_ocr_text(self, text_blob: str) -> List[Dict[str, Any]]:
        """
        Extract clients from OCR text.
        For Molina Healthcare PDFs with PSM 12 OCR, the data is in a structured format where
        each field is on a separate line:
        - Member ID (like "NYM21289999710")
        - Member Full Name (like "ARROYO, JOSEPH A")
        - Date of Birth (like "2/24/1988")
        - DOS From (like "10/10/2023")
        - DOS To (like "6/24/2025")
        - Billing Provider (like "INTEGRITY SENIOR SERVICES")
        """
        import re
        records: List[Dict[str, Any]] = []
        lines = [line.strip() for line in text_blob.splitlines() if line.strip()]
        
        # Skip header lines
        skip_keywords = ["MEMBER ID", "MEMBER FULL NAME", "DATE OF BIRTH", "DOS FROM", "DOS TO", 
                        "BILLING PROVIDER", "NO.", "PROVIDER:", "MOLINA HEALTHCARE", 
                        "SPECIAL INVESTIGATION", "MEDICAL RECORDS REQUEST", "MEDICAL RECORDS DUE"]
        
        # First, identify all potential client names and their positions
        name_pattern = r"^([A-Z][A-Za-z]+(?:,\s+[A-Z][A-Za-z]+(?:\s+[A-Z])?)+)$"
        client_positions = []
        
        for i, line in enumerate(lines):
            # Skip header/footer lines
            if any(skip in line.upper() for skip in skip_keywords):
                continue
            
            # Skip if it's "SERVICES, LLC" or similar non-name patterns
            if "SERVICES" in line.upper() or "LLC" in line.upper():
                continue
            
            # Check if this line matches a name pattern
            name_match = re.match(name_pattern, line)
            if name_match:
                client_positions.append((i, line.strip()))
        
        # Now process each client position
        processed_names = set()
        
        for name_idx, name in client_positions:
            if name in processed_names:
                continue
            
            # Skip header/footer lines
            if any(skip in name.upper() for skip in skip_keywords):
                continue
            
            # Find the next client's position to know where to stop searching
            next_client_idx = len(lines)
            for next_idx, next_name in client_positions:
                if next_idx > name_idx:
                    next_client_idx = next_idx
                    break
            
            # Search context around this name (3 lines before, up to next client)
            # For the last client, search a bit further to ensure we capture all data
            search_start = max(0, name_idx - 3)
            if next_client_idx == len(lines):
                # Last client - search up to 15 lines ahead to ensure we get DOB
                search_end = min(len(lines), name_idx + 15)
            else:
                search_end = min(len(lines), next_client_idx)
            
            dob = None
            dos_start = None
            dos_end = None
            member_id = None
            
            # Collect all dates and member IDs in the context
            all_dates = []
            for j in range(search_start, search_end):
                context_line = lines[j]
                
                # Skip if we hit a header
                if any(skip in context_line.upper() for skip in skip_keywords):
                    continue
                
                # Skip if it's "INTEGRITY SENIOR SERVICES" (billing provider)
                if "INTEGRITY SENIOR SERVICES" in context_line.upper():
                    continue
                
                # Skip if it's another name (shouldn't happen since we stop at next_client_idx, but just in case)
                if j != name_idx and re.match(name_pattern, context_line):
                    continue
                
                # Look for dates in this line
                dates = re.findall(r"\d{1,2}/\d{1,2}/\d{4}", context_line)
                
                if dates:
                    for date_str in dates:
                        normalized = self._normalize_date(date_str)
                        if normalized:
                            try:
                                date_obj = datetime.strptime(normalized, "%m/%d/%Y")
                                all_dates.append((date_obj, normalized, j))
                            except Exception as e:
                                # If normalization worked but parsing failed, try direct parsing
                                try:
                                    # Try parsing the original date string directly
                                    date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                                    all_dates.append((date_obj, date_str, j))
                                except:
                                    # Try with normalized format
                                    try:
                                        date_obj = datetime.strptime(normalized, "%m/%d/%Y")
                                        all_dates.append((date_obj, normalized, j))
                                    except:
                                        logger.debug(f"Failed to parse date: {date_str} (normalized: {normalized})")
                                        pass
                
                # Check for Member ID pattern (NYM followed by numbers)
                if re.match(r"^NYM\d+$", context_line):
                    member_id = context_line
            
            # Process collected dates (outside the loop)
            # Separate DOB (before 2000) from DOS dates (2020+)
            dob_dates = []
            dos_dates = []
            
            for date_obj, normalized, line_idx in all_dates:
                if date_obj.year < 2000:
                    dob_dates.append((date_obj, normalized, line_idx))
                elif date_obj.year >= 2020:
                    dos_dates.append((date_obj, normalized, line_idx))
            
            # Use the DOB closest to the name (usually right after the name)
            if dob_dates:
                # Sort by distance from name line, prefer DOBs that come after the name
                dob_dates.sort(key=lambda x: (abs(x[2] - name_idx), -x[2] if x[2] > name_idx else x[2]))
                dob = dob_dates[0][1]
            else:
                # If no DOB found in the search window, try searching a bit further
                # This handles cases where DOB might be slightly separated from the name
                extended_search_start = max(0, name_idx - 5)
                extended_search_end = min(len(lines), name_idx + 20)
                for j in range(extended_search_start, extended_search_end):
                    if j < search_start or j >= search_end:
                        context_line = lines[j]
                        dates = re.findall(r"\d{1,2}/\d{1,2}/\d{4}", context_line)
                        if dates:
                            for date_str in dates:
                                normalized = self._normalize_date(date_str)
                                if normalized:
                                    try:
                                        date_obj = datetime.strptime(normalized, "%m/%d/%Y")
                                        if date_obj.year < 2000:
                                            dob = normalized
                                            logger.info(f"Found DOB for {name} in extended search: {dob}")
                                            break
                                    except:
                                        pass
                        if dob:
                            break
            
            # Sort DOS dates chronologically and use first as start, last as end
            if dos_dates:
                dos_dates.sort(key=lambda x: x[0])  # Sort by date object
                dos_start = dos_dates[0][1]
                if len(dos_dates) > 1:
                    dos_end = dos_dates[-1][1]
                else:
                    dos_end = dos_start
            
            # If we found a name and at least one date, create a record
            if name and (dob or dos_start):
                record: Dict[str, Any] = {"name": name}
                if dob:
                    record["dob"] = dob
                    record["dob_raw"] = dob
                if member_id:
                    record["member_id"] = member_id
                
                # Create date range - require at least start date
                if dos_start:
                    record["date_range"] = {
                        "start": dos_start,
                        "end": dos_end or dos_start,  # Use start as end if no end date
                    }
                    records.append(record)
                    processed_names.add(name)
                    logger.info(f"Extracted client from OCR: {name}, DOB: {dob or 'N/A'}, DOS: {dos_start} to {dos_end or dos_start}")
                elif dob:
                    # If we only have DOB but no DOS, still create record but warn
                    logger.warning(f"Found client {name} with DOB but no DOS dates")
                    processed_names.add(name)
        
        # Deduplicate - keep the best record for each client (one with most complete data)
        deduped: List[Dict[str, Any]] = []
        seen: dict[str, Dict[str, Any]] = {}  # Key by name only
        
        for record in records:
            name_key = record["name"].lower()
            
            # Skip if this is a duplicate name with no useful data
            if "services" in name_key or "llc" in name_key:
                continue
            
            if name_key not in seen:
                seen[name_key] = record
            else:
                # Keep the record with more complete data
                existing = seen[name_key]
                existing_score = (1 if existing.get("dob") else 0) + (1 if existing.get("date_range", {}).get("start") else 0) + (1 if existing.get("date_range", {}).get("end") else 0)
                new_score = (1 if record.get("dob") else 0) + (1 if record.get("date_range", {}).get("start") else 0) + (1 if record.get("date_range", {}).get("end") else 0)
                if new_score > existing_score:
                    seen[name_key] = record
        
        deduped = list(seen.values())
        return deduped if len(deduped) > 0 else []

    def _extract_clients_lenient_ocr(self, text_blob: str) -> List[Dict[str, Any]]:
        """
        Very lenient extraction for heavily garbled OCR text.
        Tries to extract any recognizable patterns that might be client data.
        This is a fallback when standard extraction fails.
        """
        import re
        records: List[Dict[str, Any]] = []
        lines = [line.strip() for line in text_blob.splitlines() if line.strip()]
        
        # Focus on the last page (usually where client data is)
        # Take last 20 lines as that's typically where the client table is
        focus_lines = lines[-20:] if len(lines) > 20 else lines
        
        # Look for lines that have multiple date-like patterns (likely client rows)
        # Even with OCR errors, dates often have recognizable patterns
        for line in focus_lines:
            # Skip obvious non-client lines
            if any(skip in line.upper() for skip in ["TIOT", "Wun uoNesysaau", "SERVICES", "LLC", "MOLINA"]):
                continue
            
            # Look for patterns that might be dates (even garbled)
            # Pattern: something/something/something (dates with OCR errors)
            date_like = re.findall(r"[0-9OoIl1S5Z2Ç][0-9OoIl1S5Z2Ç]?[/\\|Ç][0-9OoIl1S5Z2Ç][0-9OoIl1S5Z2Ç]?[/\\|Ç][0-9OoIl1S5Z2Ç][0-9OoIl1S5Z2Ç][0-9OoIl1S5Z2Ç]?[0-9OoIl1S5Z2Ç]?", line)
            
            if len(date_like) >= 2:  # If line has 2+ date-like patterns, might be a client entry
                # Try to extract and clean dates
                cleaned_dates = []
                for date_str in date_like[:2]:  # Take first 2
                    # Try common OCR error corrections
                    cleaned = (date_str.replace('O', '0').replace('o', '0')
                              .replace('I', '1').replace('l', '1')
                              .replace('Ç', '/').replace('Z', '2')
                              .replace('S', '5').replace('s', '5'))
                    normalized = self._normalize_date(cleaned)
                    if normalized:
                        cleaned_dates.append(normalized)
                
                if len(cleaned_dates) >= 1:
                    # Try to find a name in this line or nearby
                    # Look for words that might be names (sequences of letters)
                    words = line.split()
                    name_candidates = []
                    for word in words:
                        # Skip if it's clearly not a name
                        if (len(word) > 2 and word[0].isupper() and 
                            sum(c.isalpha() for c in word) > len(word) * 0.6 and
                            not any(skip in word.upper() for skip in ["SERVICES", "LLC", "HEALTHCARE"])):
                            name_candidates.append(word)
                    
                    # If we found potential name parts, combine them
                    if len(name_candidates) >= 2:
                        name = " ".join(name_candidates[:3])  # Take up to 3 words
                        
                        record: Dict[str, Any] = {"name": name}
                        record["date_range"] = {
                            "start": cleaned_dates[0],
                            "end": cleaned_dates[1] if len(cleaned_dates) > 1 else cleaned_dates[0],
                        }
                        records.append(record)
                        logger.info(f"Extracted client (lenient OCR): {name}, Date Range: {cleaned_dates[0]} to {cleaned_dates[1] if len(cleaned_dates) > 1 else cleaned_dates[0]}")
        
        # Deduplicate - keep the best record for each client (one with most complete data)
        deduped: List[Dict[str, Any]] = []
        seen: dict[str, Dict[str, Any]] = {}  # Key by name only
        
        for record in records:
            name_key = record["name"].lower()
            
            # Skip if this is a duplicate name with no useful data
            if "services" in name_key or "llc" in name_key:
                continue
            
            if name_key not in seen:
                seen[name_key] = record
            else:
                # Keep the record with more complete data
                existing = seen[name_key]
                existing_score = (1 if existing.get("dob") else 0) + (1 if existing.get("date_range", {}).get("start") else 0) + (1 if existing.get("date_range", {}).get("end") else 0)
                new_score = (1 if record.get("dob") else 0) + (1 if record.get("date_range", {}).get("start") else 0) + (1 if record.get("date_range", {}).get("end") else 0)
                if new_score > existing_score:
                    seen[name_key] = record
        
        deduped = list(seen.values())
        return deduped if len(deduped) > 0 else []

    def _extract_multiple_clients_from_pdf(self, text_blob: str) -> List[Dict[str, Any]]:
        """
        Extract multiple clients with their individual date ranges from PDF text.
        Looks for client entries that have BOTH a name AND a date range (strict requirement).
        """
        import re
        
        records: List[Dict[str, Any]] = []
        lines = [line.strip() for line in text_blob.splitlines() if line.strip()]
        
        # Common words/phrases to exclude (not client names)
        exclude_patterns = [
            r"^(SERVICES|LLC|INC|CORP|LTD|LLP)$",
            r"^(INTEGRITY|SOCIAL|WORK|HEALTHCARE|MOLINA|CARE|TOTAL)$",
            r"^(USPS|CERTIFIED|MAIL|POSTAGE|FIRST-CLASS|POSTAGEM)$",
            r"^(PLEASE|SUBMIT|COMPLETE|MEDICAL|RECORDS|FOR|ALL|OF|THE|MEMBERS)$",
            r"^(PROVIDER|ORDERS|TREATMENT|PLANS|SESSION|NOTES|PROGRESS|CONSULTATION)$",
            r"^(SIGNATURE|LOGS|MEDICATION|PROCEDURE|REPORTS|DISCHARGE|SUMMARY)$",
            r"^(PHOTOCOPY|EACH|RECORD|MAKE|SURE|ALL|COPIES|ARE|COMPLETE)$",
            r"^(LEGIBLE|CONTAIN|BOTH|SIDES|OF|EACH|PAGE|INCLUDING|PAGE|EDGES)$",
            r"^(COMPLETE|COPIES|SHOULD|INCLUDE|SPECIFIC|RECORDS|TO|SUPPORT)$",
            r"^(PROVIDED|BE|SEPARATED|BY|PATIENT|IN|CHRONOLOGICAL|ORDER)$",
            r"^(FAILURE|TO|SUBMIT|REQUESTED|DOCUMENTATION|COULD|RESULT)$",
            r"^(RETROSPECTIVE|DENIAL|OF|CLAIMS|AND|OTHER|SANCTIONS)$",
            r"^(MOLINA|DOES|NOT|REIMBURSE|PROVIDERS|FOR|CHARGES|ASSOCIATED)$",
            r"^(MAKING|COPIES|AS|RELATED|TO|CLAIMS|REVIEW)$",
            r"^(SHOULD|YOU|HAVE|QUESTIONS|REGARDING|THIS|REQUEST)$",
            r"^(YOU|MAY|CONTACT|ME|VIA|EMAIL|AT)$",
            r"^(SINCERELY|INVESTIGATOR|SPECIAL|INVESTIGATION|UNIT)$",
            r"^(THE|REVIEW|IS|TAKING|PLACE|IN|ORDER|TO|MONITOR)$",
            r"^(SUBSTANTIATE|PATIENT|CARE|AND|CLAIMS|PAYMENT|ACCURACY)$",
            r"^(YOUR|COOPERATION|IN|RESPONDING|TO|THIS|REQUEST)$",
            r"^(IS|NECESSARY|IN|ORDER|TO|COMPLETE|THIS|REVIEW)$",
            r"^(AND|HEALTH|CARE|OPERATIONS|MOLINA|IS|A|COVERED|ENTITY)$",
            r"^(AS|DEFINED|BY|HIPAA|HEALTH|PLAN|MEMBERS|UPON)$",
            r"^(ENROLLMENT|IN|OUR|HEALTH|PLAN|ARE|GIVEN|A|HIPAA)$",
            r"^(PRIVACY|NOTICE|DELINEATING|EXCEPTIONS|UNDER|HIPAA)$",
        ]
        
        def is_excluded_text(text: str) -> bool:
            """Check if text should be excluded as it's not a client name."""
            text_upper = text.upper()
            # Check against exclude patterns
            for pattern in exclude_patterns:
                if re.search(pattern, text_upper):
                    return True
            # Exclude very long text (likely document body, not names)
            if len(text) > 80:
                return True
            # Exclude text with too many words (likely sentences, not names)
            if len(text.split()) > 6:
                return True
            # Exclude text that looks like addresses or long descriptions
            if any(char in text for char in [',', ';', ':', '(', ')', '[', ']']) and len(text) > 40:
                return True
            return False
        
        # Look for client entries: must have BOTH name AND date range
        i = 0
        while i < len(lines):
            # Look for a date range first (more reliable indicator of client entry)
            # Also look for date patterns that might be garbled by OCR
            date_range = None
            date_range_start_idx = None
            
            # Check current and next 15 lines for a date range
            for j in range(i, min(i + 15, len(lines))):
                section_text = "\n".join(lines[j:j+5])
                found_range = self._search_for_date_range(section_text)
                if found_range[0] or found_range[1]:
                    date_range = found_range
                    date_range_start_idx = j
                    break
                else:
                    # Try to find garbled date patterns (OCR errors)
                    # Look for patterns like "S20Z/vT/b ÇZ02/Ç7/0T" which might be dates
                    line_text = lines[j] if j < len(lines) else ""
                    # Look for sequences that might be dates with OCR errors
                    # Pattern: something that looks like MM/DD/YYYY or similar
                    garbled_date_pattern = r"[0-9Oo][0-9Oo][/\\|][0-9Oo][0-9Oo][/\\|][0-9Oo][0-9Oo][0-9Oo][0-9Oo]"
                    if re.search(garbled_date_pattern, line_text):
                        # Try to clean and extract
                        # This is a fallback - the table extraction should handle it better
                        pass
            
            if not date_range or not (date_range[0] or date_range[1]):
                i += 1
                continue
            
            # Found a date range, now look for a name nearby (within 10 lines before or after)
            name = None
            name_line_idx = None
            
            # Look backwards and forwards from date range
            search_start = max(0, date_range_start_idx - 10)
            search_end = min(len(lines), date_range_start_idx + 10)
            
            for j in range(search_start, search_end):
                line = lines[j]
                # Skip if this line is excluded text
                if is_excluded_text(line):
                    continue
                
                # Look for name patterns
                # Must be 2-5 words, proper case or all caps, not too long
                words = line.split()
                if 2 <= len(words) <= 5 and len(line) <= 60:
                    # Check if it looks like a name (starts with capital, has letters)
                    if re.match(r"^[A-Z][A-Za-z',.\- ]+$", line):
                        # Additional validation: should not be all common words
                        if not all(word.upper() in ["THE", "AND", "OR", "FOR", "TO", "OF", "IN", "ON", "AT", "BY"] for word in words):
                            # Try to extract name using the name search function
                            candidate_name = self._search_for_name("\n".join(lines[max(0, j-2):j+3]))
                            if candidate_name and len(candidate_name.split()) >= 2 and not is_excluded_text(candidate_name):
                                name = candidate_name
                                name_line_idx = j
                                break
                
                # Also check for explicit name labels
                if any(keyword in line.upper() for keyword in ["PATIENT:", "MEMBER:", "CLIENT:", "NAME:", "BENEFICIARY:"]):
                    if ':' in line:
                        potential_name = line.split(':', 1)[1].strip()
                        if potential_name and 2 <= len(potential_name.split()) <= 5 and not is_excluded_text(potential_name):
                            name = potential_name
                            name_line_idx = j
                            break
            
            # Only create record if we found BOTH name AND date range
            if name and date_range and (date_range[0] or date_range[1]):
                # Also try to find DOB in the same section
                section_start = min(name_line_idx or date_range_start_idx, date_range_start_idx)
                section_end = max((name_line_idx or date_range_start_idx) + 5, date_range_start_idx + 5)
                section_text = "\n".join(lines[section_start:min(section_end, len(lines))])
                dob = self._search_for_dob_strict(section_text)
                
                record: Dict[str, Any] = {"name": name}
                if dob:
                    record["dob"] = dob
                    record["dob_raw"] = dob
                record["date_range"] = {
                    "start": date_range[0],
                    "end": date_range[1],
                }
                records.append(record)
                logger.info(f"Extracted client from PDF: {name}, DOB: {dob or 'N/A'}, Date Range: {date_range[0] or 'N/A'} to {date_range[1] or 'N/A'}")
                
                # Move past this entry
                i = max((name_line_idx or date_range_start_idx), date_range_start_idx) + 10
            else:
                i += 1
        
        # Deduplicate records by name+DOB
        deduped: List[Dict[str, Any]] = []
        seen: set[Tuple[str, str]] = set()
        for record in records:
            key = (record["name"].lower(), (record.get("dob") or record.get("dob_raw") or "").lower())
            if key not in seen:
                seen.add(key)
                deduped.append(record)
        
        return deduped if len(deduped) > 1 else []  # Only return if multiple clients found

    def _parse_image(self, file_path: Path) -> ClientMetadata:
        if not self.ocr_available:
            raise RuntimeError("OCR libraries (pytesseract + Pillow) are required for image parsing.")
        logger.info("Running OCR on image: %s", file_path)
        text = pytesseract.image_to_string(Image.open(file_path))  # type: ignore[arg-type]
        name = self._search_for_name(text)
        dob = self._search_for_dob(text)
        date_range = self._search_for_date_range(text)
        extras: Dict[str, Any] = {}
        if date_range[0] or date_range[1]:
            extras["date_range"] = {
                "start": date_range[0],
                "end": date_range[1],
            }
        return ClientMetadata(name=name, dob=dob, raw_text=text, source_file=file_path, extras=extras)

    def _ocr_pdf(self, file_path: Path) -> str:
        if not self.ocr_available:
            raise RuntimeError("OCR requested but pytesseract is unavailable.")
        from pdf2image import convert_from_path  # type: ignore[import-untyped]
        pages = convert_from_path(file_path)
        # Try different OCR modes for better accuracy
        # PSM 12 = Sparse text with OSD (works best for structured tables like Molina Healthcare)
        # PSM 6 = Assume a single uniform block of text (good for tables)
        # PSM 11 = Sparse text (good for documents with mixed content)
        text_blobs = []
        for page_num, page in enumerate(pages, 1):
            try:
                # For last page (where client data usually is), try PSM 12 first (best for structured tables)
                if page_num == len(pages):
                    # Try PSM 12 first (sparse text with OSD - works great for Molina Healthcare format)
                    try:
                        text = pytesseract.image_to_string(page, config='--psm 12')
                        if len(text.strip()) > 100:  # If we got substantial text
                            text_blobs.append(text)
                        else:
                            raise ValueError("PSM 12 didn't produce enough text")
                    except Exception:
                        # Fallback to PSM 6 (uniform block - good for tables)
                        text = pytesseract.image_to_string(page, config='--psm 6')
                        if len(text.strip()) > 50:
                            text_blobs.append(text)
                        else:
                            # Fallback to PSM 4 (single column)
                            text = pytesseract.image_to_string(page, config='--psm 4')
                            text_blobs.append(text)
                else:
                    # For other pages, use standard OCR
                    text = pytesseract.image_to_string(page, config='--psm 11')
                    text_blobs.append(text)
            except Exception:
                # Fallback to default
                text_blobs.append(pytesseract.image_to_string(page))
        return "\n".join(text_blobs)

    @staticmethod
    def _search_for_name(text_blob: str) -> str:
        logger.debug("Searching for client name in extracted text.")
        import re

        lines = [line.strip() for line in text_blob.splitlines() if line.strip()]
        name_patterns = [
            r"(?:member|patient|client|insured|recipient)\s*name[:\-\s]+(.+)",
            r"(?:patient\s+information|member\s+information|client\s+information)\W+([A-Z][A-Za-z',.\- ]{2,})",
            r"(?:name)\s*[:\-]\s*(.+)",
        ]
        stop_tokens = re.compile(r"\b(DOB|DATE OF BIRTH|MEMBER ID|ID|MRN|REQUEST|PHONE|FAX)\b", re.IGNORECASE)

        for line in lines:
            for pattern in name_patterns:
                match = re.search(pattern, line, re.IGNORECASE)
                if match:
                    candidate = match.group(1)
                    candidate = stop_tokens.split(candidate)[0]
                    candidate = candidate.strip(" :-|.,")
                    if candidate:
                        return candidate

        # Secondary heuristic: look for name following "Patient Information" block
        block_keywords = ("patient information", "member information", "client information")
        for idx, line in enumerate(lines):
            if any(keyword in line.lower() for keyword in block_keywords):
                for offset in range(1, 6):
                    if idx + offset >= len(lines):
                        break
                    candidate = lines[idx + offset].strip(" :-|.,")
                    if candidate:
                        return candidate
        return ""

    @staticmethod
    def _search_for_dob(text_blob: str) -> str:
        logger.debug("Searching for DOB in extracted text.")
        import re

        lines = [line.strip() for line in text_blob.splitlines() if line.strip()]
        flex_dob = r"D[.\s\-]*[0O][.\s\-]*B\.?"
        dob_patterns = [
            rf"(?:DOB|{flex_dob}|DATE OF BIRTH)[:\-\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
            rf"(?:DOB|{flex_dob}|DATE OF BIRTH)\W+([A-Za-z]{{3,9}}\s+\d{{1,2}}[,]?\s+\d{{2,4}})",
            rf"\({flex_dob}\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
            rf"\({flex_dob}\s*([A-Za-z]{{3,9}}\s+\d{{1,2}}[,]?\s+\d{{2,4}})",
        ]
        for line in lines:
            for pattern in dob_patterns:
                match = re.search(pattern, line, re.IGNORECASE)
                if match:
                    normalized = DocumentParser._normalize_date(match.group(1))
                    if normalized:
                        return normalized

        # Broader context search for DOB without colon
        wide_context = re.search(
            rf"(DOB|{flex_dob}|DATE OF BIRTH)[^\dA-Za-z]{{0,10}}([A-Za-z]{{3,9}}\s+\d{{1,2}}[,]?\s+\d{{2,4}}|\d{{1,2}}[/-]\d{{1,2}}[/-]\d{{2,4}})",
            text_blob,
            re.IGNORECASE
        )
        if wide_context:
            normalized = DocumentParser._normalize_date(wide_context.group(2))
            if normalized:
                return normalized

        # Fallback: first standalone date in document (WARNING: may be incorrect)
        match = re.search(r"\b(0?[1-9]|1[0-2])[/-](0?[1-9]|[12][0-9]|3[01])[/-](\d{2}|\d{4})\b", text_blob)
        if match:
            normalized = DocumentParser._normalize_date(match.group(0))
            if normalized:
                return normalized
        return ""

    @staticmethod
    def _search_for_dob_strict(text_blob: str) -> str:
        """Search for DOB in text - strict version that only looks for labeled DOB, no fallback to first date."""
        logger.debug("Searching for DOB in extracted text (strict mode).")
        import re

        lines = [line.strip() for line in text_blob.splitlines() if line.strip()]
        flex_dob = r"D[.\s\-]*[0O][.\s\-]*B\.?"
        dob_patterns = [
            rf"(?:DOB|{flex_dob}|DATE OF BIRTH)[:\-\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
            rf"(?:DOB|{flex_dob}|DATE OF BIRTH)\W+([A-Za-z]{{3,9}}\s+\d{{1,2}}[,]?\s+\d{{2,4}})",
            rf"\({flex_dob}\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
            rf"\({flex_dob}\s*([A-Za-z]{{3,9}}\s+\d{{1,2}}[,]?\s+\d{{2,4}})",
        ]
        for line in lines:
            for pattern in dob_patterns:
                match = re.search(pattern, line, re.IGNORECASE)
                if match:
                    normalized = DocumentParser._normalize_date(match.group(1))
                    if normalized:
                        logger.debug(f"Found DOB via pattern match: {normalized}")
                        return normalized

        # Broader context search for DOB without colon
        wide_context = re.search(
            rf"(DOB|{flex_dob}|DATE OF BIRTH)[^\dA-Za-z]{{0,10}}([A-Za-z]{{3,9}}\s+\d{{1,2}}[,]?\s+\d{{2,4}}|\d{{1,2}}[/-]\d{{1,2}}[/-]\d{{2,4}})",
            text_blob,
            re.IGNORECASE
        )
        if wide_context:
            normalized = DocumentParser._normalize_date(wide_context.group(2))
            if normalized:
                logger.debug(f"Found DOB via wide context search: {normalized}")
                return normalized
        
        # No fallback to first standalone date - return empty if not found
        logger.debug("No DOB found with strict search (no fallback to first date)")
        return ""

    @staticmethod
    def _normalize_date(raw_date: str) -> Optional[str]:
        raw = raw_date.strip().replace("  ", " ")
        # Try common date formats including YYYY-MM-DD (pandas Timestamp format)
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                dt = datetime.strptime(raw, fmt)
                return dt.strftime("%m/%d/%Y")
            except ValueError:
                continue
        # Try formats with time component (from pandas Timestamp string representation)
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f"):
            try:
                dt = datetime.strptime(raw, fmt)
                return dt.strftime("%m/%d/%Y")
            except ValueError:
                continue
        for fmt in ("%B %d, %Y", "%B %d %Y", "%b %d, %Y", "%b %d %Y"):
            try:
                dt = datetime.strptime(raw.title(), fmt)
                return dt.strftime("%m/%d/%Y")
            except ValueError:
                continue
        # handle ordinal suffixes (July 1st, 2024)
        import re
        cleaned = re.sub(r"(\d+)(st|nd|rd|th)", r"\1", raw, flags=re.IGNORECASE)
        for fmt in ("%B %d, %Y", "%B %d %Y", "%b %d, %Y", "%b %d %Y"):
            try:
                dt = datetime.strptime(cleaned.title(), fmt)
                return dt.strftime("%m/%d/%Y")
            except ValueError:
                continue
        return None

    @staticmethod
    def _search_for_date_range(text_blob: str) -> Tuple[Optional[str], Optional[str]]:
        import re

        lines = [line.strip() for line in text_blob.splitlines() if line.strip()]
        date_pattern_numeric = r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}"
        date_pattern_text = r"[A-Za-z]{3,9}\s+\d{1,2}(?:st|nd|rd|th)?[,]?\s+\d{2,4}"
        any_date_pattern = rf"(?:{date_pattern_numeric}|{date_pattern_text})"
        dash_class = r"\-–—"
        present_words = r"(?:present|current|today)"

        pattern_definitions = [
            (
                ["DATE OF SERVICE", "DATES OF SERVICE", "DOS", "SERVICE DATES", "DATE RANGE", "BETWEEN"],
                [
                    rf"(?P<start>{any_date_pattern})\s*(?:to|[{dash_class}]|through|and)\s*(?P<end>{present_words}|{any_date_pattern})",
                    rf"(?:from|between)\s+(?P<start>{any_date_pattern})\s*(?:to|[{dash_class}]|through|and)\s*(?P<end>{present_words}|{any_date_pattern})",
                ],
            ),
            (
                ["RECORDS", "RECORD"],
                [
                    rf"(?:records?|record)\s+(?:for|from)\s+(?P<start>{any_date_pattern})(?:\s*(?:to|[{dash_class}]|through|and)\s*(?P<end>{present_words}|{any_date_pattern}))?",
                ],
            ),
            (
                ["FROM", "SINCE"],
                [
                    rf"(?:from|since)\s+(?P<start>{any_date_pattern})(?:\s*(?:to|[{dash_class}]|through|and)\s*(?P<end>{present_words}|{any_date_pattern}))?",
                ],
            ),
            (
                ["DATE OF SERVICE", "DOS"],
                [
                    rf"(?:date\s+of\s+service|DOS)[:\-\s]+(?P<start>{any_date_pattern})(?:\s*(?:to|[{dash_class}]|through|and)\s*(?P<end>{present_words}|{any_date_pattern}))?",
                ],
            ),
        ]

        for keywords, patterns in pattern_definitions:
            for line in lines:
                upper = line.upper()
                if any(keyword in upper for keyword in keywords):
                    for pattern in patterns:
                        match = re.search(pattern, line, re.IGNORECASE)
                        if match:
                            start_raw = match.group("start")
                            end_raw = match.group("end") if "end" in match.groupdict() else None
                            start = DocumentParser._normalize_date(start_raw) if start_raw else None
                            end = None
                            if end_raw:
                                end_key = end_raw.lower().strip(" .,:;")
                                if end_key in {"present", "current", "today"}:
                                    end = "Present"
                                else:
                                    end = DocumentParser._normalize_date(end_raw)
                            if start or end:
                                return start, end
        # Broad fallback scans
        fallback_patterns = [
            rf"(?:all\s+)?records?\s+(?:needed[:\s]*)?(?:for|from)?\s*(?P<start>{any_date_pattern})\s*(?:to|[{dash_class}]|through|and)\s*(?P<end>{present_words}|{any_date_pattern})",
            rf"(?P<start>{any_date_pattern})\s*(?:to|[{dash_class}]|through|and)\s*(?P<end>{present_words}|{any_date_pattern})\s+(?:records?|documentation)",
        ]
        for pattern in fallback_patterns:
            match = re.search(pattern, text_blob, re.IGNORECASE)
            if match:
                start_raw = match.group("start")
                end_raw = match.group("end") if "end" in match.groupdict() else None
                start = DocumentParser._normalize_date(start_raw) if start_raw else None
                end = None
                if end_raw:
                    end_key = end_raw.lower().strip(" .,:;")
                    if end_key in {"present", "current", "today"}:
                        end = "Present"
                    else:
                        end = DocumentParser._normalize_date(end_raw)
                if start or end:
                    return start, end

        return None, None


# TherapyNotes navigation placeholder -----------------------------------------

DEFAULT_INS_STATUS_FILTERS = [
    "Pending Secondary",
    "Submitted External",
    "Submitted Claim",
    "Submitted",
    "Paid",
    "Forwarded to Secondary",
]

PRIMARY_PAYER_OPTIONS = [
    "1199 National Benefit Fund",
    "AARP Medicare Supplement Plans",
    "Aetna",
    "AgeWell New York",
    "ANTHEM BLUE CROSS AND BLUE SHIELD HP",
    "APWU",
    "Beacon Health Strategies",
    "Carefirst",
    "CARELON BEHAVIORAL HEALTH",
    "CARELON BEHAVIORAL HEALTH - VCM",
    "CDPHP",
    "Centerlight",
    "Centers Plan for Healthy Living",
    "ChampVA",
    "CIGNA",
    "ElderPlan Inc.",
    "Emblem Health GHI New York Group Health Inc",
    "Empire Blue Cross and Blue Shield of New York",
    "Fidelis Care New York",
    "GHI HMO",
    "Globe Life",
    "Government Employees Health Association (GEHA)",
    "Healthfirst Inc. (New York)",
    "Humana Inc.",
    "Magellan Health Services",
    "Magnacare Administrative Services",
    "MetroPlus Health Plan",
    "Molina Healthcare of NY (formerly Total Care NY)",
    "MVP Health Plan of New York",
    "National Association of Letter Carriers",
    "New York Medicaid",
    "NY Medicare Part B Downstate",
    "OptumHealth",
    "Oscar Health",
    "Oxford",
    "Parnters Health Plan",
    "Railroad Medicare (PGBA)",
    "RiverSpring Health Plans",
    "Tricare for Life",
    "Trustmark - Lumincare",
    "UnitedHealthcare",
    "Value Options - Commercial",
    "VillageCareMAX",
    "WebTPA Employer Services LLC.",
    "Wellcare",
    "Archcare",
    "First United American Life",
    "Transamerica Financial Life Insurance Company",
]


class TherapyNotesClientFetcher:
    """Handles Selenium-based navigation to TherapyNotes."""

    DEFAULT_LOGIN_URL = "https://www.therapynotes.com/app/login/IntegritySWS/"
    USERNAME_INPUT_ID = "Login__UsernameField"
    PASSWORD_INPUT_ID = "Login__Password"
    SUBMIT_BUTTON_ID = "Login__LogInButton"

    def __init__(self, credentials: Dict[str, str], log_callback=None) -> None:
        self.credentials = credentials
        self.driver = None
        self.log_callback = log_callback

    def _log(self, message: str) -> None:
        if self.log_callback:
            try:
                self.log_callback(message)
            except Exception:
                logger.info(message)
        else:
            logger.info(message)

    def authenticate(self) -> None:
        if not SELENIUM_AVAILABLE:
            raise RuntimeError(
                "Selenium is required for TherapyNotes automation. Install the selenium package and retry."
            )
        username = self.credentials.get("username")
        password = self.credentials.get("password")
        if not username or not password:
            raise ValueError("TherapyNotes credentials missing username or password.")

        driver = self._ensure_driver()
        login_url = self.credentials.get("login_url") or self.DEFAULT_LOGIN_URL
        self._log(f"Navigating to TherapyNotes login: {login_url}")
        driver.get(login_url)

        wait = WebDriverWait(driver, 20)
        user_field = wait.until(EC.presence_of_element_located((By.ID, self.USERNAME_INPUT_ID)))
        user_field.clear()
        user_field.send_keys(username)

        password_field = wait.until(EC.presence_of_element_located((By.ID, self.PASSWORD_INPUT_ID)))
        password_field.clear()
        password_field.send_keys(password)

        submit_button = wait.until(EC.element_to_be_clickable((By.ID, self.SUBMIT_BUTTON_ID)))
        current_url = driver.current_url
        submit_button.click()

        try:
            wait.until(EC.url_changes(current_url))
            self._log("TherapyNotes login successful.")
        except TimeoutException:
            self._log("TherapyNotes login did not redirect. Check credentials or MFA requirements.")
            raise RuntimeError("TherapyNotes login failed or requires additional verification.")

    def open_client_chart(self, client: ClientMetadata) -> None:
        if not self.driver:
            raise RuntimeError("Driver not initialized. Call authenticate() first.")
        self._log(f"Navigating to Patients section to locate {client.name}.")
        wait = WebDriverWait(self.driver, 20)
        patients_link = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "a.sidebar-link[href='/app/patients/']"))
        )
        patients_link.click()
        try:
            wait.until(EC.url_contains("/app/patients"))
        except TimeoutException:
            self._log("[NAV][WARN] Patients page navigation did not complete as expected.")

    def search_and_open_patient(self, name: str, dob: Optional[str]) -> bool:
        if not self.driver:
            raise RuntimeError("Driver not initialized. Call authenticate() first.")
        wait = WebDriverWait(self.driver, 20)
        normalized_target = DocumentParser._normalize_date(dob) if dob else None
        search_names = self._build_search_names(name)
        self._log(
            f"[SEARCH] Starting lookup for '{name}' (DOB hint: {normalized_target or 'N/A'}) | variants={search_names}"
        )

        for search_name in search_names:
            self._log(f"[SEARCH] Attempting search with name variant '{search_name}'.")
            search_box = self._locate_search_box(wait)
            if not search_box:
                self._log("[SEARCH][WARN] Could not locate search box; aborting search.")
                return False

            try:
                search_box.clear()
            except StaleElementReferenceException:
                self._log("[SEARCH][WARN] Search box became stale during clear; retrying locator.")
                continue

            search_box.send_keys(search_name)
            matched = self._select_dropdown_entry(wait, search_name, normalized_target)
            if matched:
                self._wait_for_patient_detail()
                return True

        self._log(f"[SEARCH][WARN] Unable to find dropdown entry for '{name}' with DOB {normalized_target or 'N/A'}")
        return False

    def _locate_search_box(self, wait: WebDriverWait):
        search_box_selectors = [
            (By.ID, "ctl00_BodyContent_TextBoxSearchPatientName"),
            (By.NAME, "ctl00$BodyContent$TextBoxSearchPatientName"),
            (By.XPATH, "//input[@placeholder='Name, Acct #, Phone, or Ins ID']"),
            (By.XPATH, "//input[@name='ctl00$BodyContent$TextBoxSearchPatientName']"),
        ]

        for by, sel in search_box_selectors:
            try:
                element = wait.until(EC.element_to_be_clickable((by, sel)))
                return element
            except Exception:
                continue
        return None

    @staticmethod
    def _build_search_names(name: str) -> List[str]:
        cleaned = name.strip()
        if not cleaned:
            return []
        variants = set()
        if "," in cleaned:
            last, *rest = [part.strip() for part in cleaned.split(",", 1)]
            rest_part = rest[0] if rest else ""
            first_tokens = rest_part.split()
            if first_tokens:
                first = first_tokens[0]
                remaining = " ".join(first_tokens[1:])
            else:
                first = rest_part
                remaining = ""
            base_variants = [
                f"{first} {last}".strip(),
                f"{last} {first}".strip(),
                f"{first} {remaining} {last}".strip(),
                cleaned.replace(",", "").strip(),
            ]
        else:
            tokens = cleaned.split()
            if len(tokens) >= 2:
                first = tokens[0]
                last = tokens[-1]
                base_variants = [
                    " ".join(tokens).strip(),
                    f"{first} {last}".strip(),
                    f"{last} {first}".strip(),
                ]
            else:
                base_variants = [cleaned]

        for variant in base_variants:
            if variant:
                variants.add(" ".join(variant.split()))

        return [v for v in variants if v]

    def _select_dropdown_entry(self, wait: WebDriverWait, search_name: str, normalized_dob: Optional[str]) -> bool:
        try:
            results_container = WebDriverWait(self.driver, 6).until(
                EC.visibility_of_element_located((By.ID, "ContentBubbleResultsContainer"))
            )
            self._log("[SEARCH] Dropdown container appeared.")
        except TimeoutException:
            self._log(f"[SEARCH] Dropdown container did not appear for '{search_name}'.")
            return False

        results = self._gather_dropdown_entries(results_container, timeout=4.0)
        if not results:
            self._log("[SEARCH][WARN] No results in dropdown container after polling.")
            return False

        self._log(f"[SEARCH] Total dropdown results found: {len(results)}.")

        matched_patient = None
        if normalized_dob:
            dob_str = normalized_dob
            dob_normalized = dob_str.replace("/", "").replace("-", "").replace(" ", "")
            year_digits = dob_str[-2:] if len(dob_str) >= 2 else None
            self._log(
                f"[SEARCH] Looking for DOB match: normalized={dob_normalized} year_suffix={year_digits}."
            )

            for result in results:
                try:
                    result_text = result.text.strip()
                except StaleElementReferenceException:
                    continue
                if not result_text:
                    continue
                result_normalized = (
                    result_text.replace("/", "").replace("-", "").replace(" ", "")
                )
                self._log(f"[SEARCH] Checking dropdown entry: '{result_text}'.")
                if dob_normalized and dob_normalized in result_normalized:
                    matched_patient = result
                    self._log("[SEARCH] Matched by full DOB in dropdown entry.")
                    break
                if year_digits and year_digits in result_text:
                    matched_patient = result
                    self._log("[SEARCH] Matched by DOB year in dropdown entry.")
                    break

        if matched_patient is None and results:
            matched_patient = results[0]
            self._log("[SEARCH] No DOB match; selecting first dropdown result.")

        if not matched_patient:
            self._log("[SEARCH][WARN] Unable to select any dropdown entry.")
            return False

        try:
            matched_patient.click()
            self._log("[SEARCH] Clicked dropdown entry successfully.")
            # Reduced wait - the _wait_for_patient_detail will handle waiting for page load
            time.sleep(0.5)
            return True
        except Exception as exc:
            self._log(f"[SEARCH][WARN] Failed to click dropdown entry: {exc}")
            return False

    def _gather_dropdown_entries(self, container, timeout: float = 3.0) -> List:
        selectors = [
            "div.ui-menu-item",
            ".ui-menu-item",
            "li.ui-menu-item",
            "a.ui-menu-item",
            "*[data-testid*='search']",
            "div[data-testid*='incremental']",
            "*[class*='IncrementalSearch']",
            "*[role='option']",
        ]
        end_time = time.time() + max(timeout, 0.5)
        last_visible: List = []
        while time.time() < end_time:
            if container.text and container.text.strip():
                entries = [el for el in container.find_elements(By.XPATH, ".//*") if el.text.strip()]
                if entries:
                    self._log(
                        f"[SEARCH] Container text populated with {len(entries)} raw entries (fast path)."
                    )
                    return entries
            for sel in selectors:
                try:
                    elements = container.find_elements(By.CSS_SELECTOR, sel)
                except Exception:
                    continue
                with_text = [el for el in elements if el.text.strip()]
                if with_text:
                    self._log(f"[SEARCH] Selector '{sel}' yielded {len(with_text)} results.")
                    return with_text
                last_visible = with_text or last_visible
            time.sleep(0.15)

        if not last_visible:
            try:
                fallback = [
                    el for el in container.find_elements(By.TAG_NAME, "*")
                    if el.text.strip()
                ]
            except Exception:
                fallback = []
            if fallback:
                limited = fallback[:5]
                self._log(
                    f"[SEARCH] Fallback traversal yielded {len(limited)} dropdown results."
                )
                return limited
        return last_visible

    @staticmethod
    def _normalize_name_for_compare(value: str) -> str:
        import re

        return re.sub(r"[^A-Za-z0-9]", "", value).upper()

    def _generate_name_variants(self, name: str) -> List[str]:
        import re

        cleaned = name.strip()
        if not cleaned:
            return []
        parts: List[str] = []
        if "," in cleaned:
            last, *rest = [p.strip() for p in cleaned.split(",", 1)]
            parts.append(last)
            if rest:
                parts.extend(rest[0].split())
        else:
            parts = cleaned.split()
        parts = [re.sub(r"[^A-Za-z0-9]", "", token).upper() for token in parts if token]
        if not parts:
            return []
        variants = set()
        concatenated = "".join(parts)
        variants.add(concatenated)
        if len(parts) > 1:
            variants.add("".join(reversed(parts)))
            variants.add(parts[0] + parts[1])
            variants.add(parts[1] + parts[0])
        for token in parts:
            variants.add(token)
        initials = "".join(token[0] for token in parts if token)
        if initials:
            variants.add(initials)
            variants.add("".join(reversed(list(initials))))
        return [v for v in variants if v]

    def _collect_dropdown_results(self, container, retries: int = 1) -> List:
        selectors = [
            "div.ui-menu-item",
            ".ui-menu-item",
            "li.ui-menu-item",
            "a.ui-menu-item",
            ".ui-menu-item-wrapper",
            "[data-testid*='search']",
            "*[role='option']",
        ]
        for _ in range(retries):
            for selector in selectors:
                try:
                    elements = container.find_elements(By.CSS_SELECTOR, selector)
                    filtered = [el for el in elements if el.is_displayed() and el.text.strip()]
                    if filtered:
                        self._log(f"[SEARCH] Selector '{selector}' yielded {len(filtered)} results.")
                        return filtered
                except Exception:
                    continue
            time.sleep(0.2)
        try:
            elements = [
                el for el in container.find_elements(By.XPATH, ".//*")
                if el.is_displayed() and el.text.strip()
            ]
            if elements:
                self._log(f"[SEARCH] Fallback traversal yielded {len(elements)} dropdown results.")
            return elements
        except Exception:
            return []

    def _match_candidate(
        self,
        candidates: List,
        name_variants: List[str],
        normalized_dob: Optional[str],
    ):
        import re

        fallback = None
        display_name = name_variants[0] if name_variants else "client"
        year_fragment = normalized_dob[-2:] if normalized_dob else None
        for candidate in candidates:
            try:
                text = candidate.text.strip()
            except StaleElementReferenceException:
                continue
            if not text:
                continue

            norm_text = self._normalize_name_for_compare(text)
            self._log(f"[SEARCH] Evaluating dropdown entry: '{text}' | normalized='{norm_text}'")
            if not any(variant in norm_text for variant in name_variants):
                self._log(f"[SEARCH] Entry rejected (name mismatch) for variants {name_variants}.")
                continue

            candidate_dob = None
            dob_match = re.search(r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", text)
            if dob_match:
                candidate_dob = DocumentParser._normalize_date(dob_match.group(1))
                self._log(f"[SEARCH] Found DOB in entry: raw={dob_match.group(1)} normalized={candidate_dob}")

            if normalized_dob and candidate_dob == normalized_dob:
                self._log(f"[SEARCH] Matched '{display_name}' by exact DOB with dropdown entry '{text}'.")
                return candidate
            if normalized_dob and year_fragment and year_fragment in text:
                self._log(f"[SEARCH] Matched '{display_name}' by DOB year ({year_fragment}) with dropdown entry '{text}'.")
                return candidate
            if fallback is None:
                self._log(f"[SEARCH] Storing fallback entry '{text}'.")
                fallback = candidate

        if fallback:
            if normalized_dob:
                self._log(f"[SEARCH] No DOB match; selecting first matching name entry '{fallback.text.strip()}'.")
            else:
                self._log(f"[SEARCH] Selecting name match (no DOB constraint) using '{fallback.text.strip()}'.")
            return fallback

        return None

    def _clear_and_type(self, element, value: str) -> bool:
        try:
            if Keys:
                element.send_keys(Keys.CONTROL, "a")
                element.send_keys(Keys.DELETE)
            else:
                element.clear()
            element.send_keys(value)
            return True
        except StaleElementReferenceException:
            self._log(f"[SEARCH] Search box became stale while typing '{value}'.")
            return False

    def _derive_search_terms(self, name: str) -> List[str]:
        cleaned = name.strip()
        if not cleaned:
            return []

        terms: List[str] = []
        seen: set[str] = set()

        def _add(term: str) -> None:
            stripped = term.strip()
            key = stripped.lower()
            if stripped and key not in seen:
                seen.add(key)
                terms.append(stripped)

        _add(cleaned)

        if "," in cleaned:
            last, rest = [segment.strip() for segment in cleaned.split(",", 1)]
            first_parts = [part for part in rest.split() if part]
            if first_parts:
                first = first_parts[0]
                remainder = " ".join(first_parts[1:])
                _add(f"{first} {last}")
                _add(f"{first} {remainder} {last}" if remainder else f"{first} {last}")
                _add(f"{last} {first}")
                _add(f"{last}, {first}")
        else:
            parts = cleaned.split()
            if len(parts) >= 2:
                first = parts[0]
                last = parts[-1]
                middle = " ".join(parts[1:-1])
                _add(f"{first} {last}")
                _add(f"{first} {middle} {last}" if middle else f"{first} {last}")
                _add(f"{last} {first}")
                _add(f"{last}, {first}")

        return terms or [cleaned]

    def fetch_billable_dates(
        self,
        client: ClientMetadata,
        start: datetime,
        end: datetime,
        payer_filters: List[str],
        status_filters: List[str]
    ) -> List[ServiceEntry]:
        if not self.driver:
            raise RuntimeError("Driver not initialized. Call authenticate() first.")
        wait = WebDriverWait(self.driver, 20)
        try:
            wait.until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "[data-testid='billingstatementtable-paymentdate-link']")
                )
            )
        except TimeoutException:
            self._log(f"[BILLING] No billing transactions found for {client.name}.")
            return []

        start_date = start.date()
        end_date = end.date()
        # Use exact matching for payer filters (case-insensitive)
        payer_filters_norm = [p.strip().lower() for p in payer_filters if p.strip()]
        status_filters_norm = [s.lower() for s in status_filters]

        self._log(
            f"[BILLING] Scanning transactions for {client.name} between {start_date} and {end_date}."
        )

        # Wait a moment to ensure the table is fully loaded after filter application
        time.sleep(0.5)
        
        # First, scroll to load all rows in the table
        self._log("[BILLING] Scrolling to load all billing table rows...")
        self._scroll_to_load_all_rows()
        
        # Wait again after scrolling to ensure all rows are loaded
        time.sleep(0.5)

        # Collect all rows with their data
        results: List[ServiceEntry] = []
        processed_rows = set()  # Track processed rows by a unique identifier

        max_attempts = 3
        for attempt in range(max_attempts):
            try:
                # Get all date links after scrolling
                date_links = self.driver.find_elements(
                    By.CSS_SELECTOR,
                    "[data-testid='billingstatementtable-paymentdate-link']"
                )
                
                if not date_links:
                    if attempt == 0:
                        self._log(f"[BILLING] Billing table empty for {client.name}.")
                    break

                self._log(f"[BILLING] Found {len(date_links)} date links (attempt {attempt + 1})")

                # Process each row
                processed_all = True
                for link_idx, link in enumerate(date_links):
                    try:
                        # Get the row element
                        try:
                            row = link.find_element(By.XPATH, "./ancestor::tr")
                        except Exception:
                            continue

                        # Extract date first (needed for row ID)
                        try:
                            raw_date = link.text.strip()
                        except Exception:
                            continue

                        if not raw_date:
                            continue

                        # Create a unique identifier for this row to avoid duplicates
                        # Use a combination of date, payer, and row index for stable identification
                        try:
                            row_id = row.get_attribute("data-testid") or row.get_attribute("id")
                            if not row_id:
                                # Use date text + payer text + index as identifier
                                try:
                                    payer_elem = row.find_elements(By.CSS_SELECTOR, "[data-testid='billingstatementtable-payer-link']")
                                    payer_text_for_id = payer_elem[0].text.strip() if payer_elem else ""
                                except Exception:
                                    payer_text_for_id = ""
                                row_id = f"{raw_date}|{payer_text_for_id}|{link_idx}"
                        except Exception:
                            # Last resort: use index and timestamp
                            row_id = f"{link_idx}_{time.time()}"

                        # Skip if we've already processed this row
                        if row_id in processed_rows:
                            continue
                        processed_rows.add(row_id)

                        # Parse and validate date
                        normalized_date = DocumentParser._normalize_date(raw_date)
                        if not normalized_date:
                            self._log(f"[BILLING] Skipping row with unrecognized date '{raw_date}'.")
                            continue

                        dt = DocumentParser._to_datetime(normalized_date)
                        if not dt:
                            self._log(f"[BILLING] Unable to parse date '{normalized_date}'.")
                            continue

                        # Filter by date range (inclusive)
                        if dt.date() < start_date or dt.date() > end_date:
                            continue  # Skip dates outside range silently (we'll log summary)

                        # Extract payer
                        payer_text = ""
                        try:
                            payer_elem = row.find_elements(By.CSS_SELECTOR, "[data-testid='billingstatementtable-payer-link']")
                            if payer_elem:
                                payer_text = payer_elem[0].text.strip()
                        except Exception:
                            payer_text = ""

                        # Apply payer filter
                        if payer_filters_norm:
                            payer_lower = payer_text.strip().lower()
                            if payer_lower not in payer_filters_norm:
                                continue  # Skip if payer doesn't match filter

                        # Extract status
                        status_text = ""
                        try:
                            # Try multiple selectors to find the status
                            status_elem = row.find_elements(
                                By.CSS_SELECTOR,
                                "[data-testid='billingstatementtable-insurancestatus-container']"
                            )
                            if status_elem:
                                status_text = status_elem[0].text.strip()
                            
                            # If empty, try the batchclaimcreator link
                            if not status_text:
                                batch_elem = row.find_elements(
                                    By.CSS_SELECTOR,
                                    "[data-testid='billingstatementtable-batchclaimcreator-link']"
                                )
                                if batch_elem:
                                    status_text = batch_elem[0].text.strip()
                            
                            # If still empty, try any action-link in the row
                            if not status_text:
                                action_links = row.find_elements(By.CSS_SELECTOR, "a.action-link")
                                for action_link in action_links:
                                    link_text = action_link.text.strip()
                                    if link_text and link_text not in ["", "View", "Edit"]:
                                        status_text = link_text
                                        break
                            
                            # If still empty, try looking for status keywords in row text
                            if not status_text:
                                row_text = row.text
                                status_keywords = ["Pending Resubmit", "Pending", "Submitted", "Paid", "Forwarded"]
                                import re
                                for keyword in status_keywords:
                                    if keyword in row_text:
                                        match = re.search(rf"\b{re.escape(keyword)}(?:\s+\w+)?\b", row_text)
                                        if match:
                                            status_text = match.group(0).strip()
                                            break
                        except Exception as e:
                            status_text = ""
                            self._log(f"[BILLING][DEBUG] Exception extracting status for date {dt.date()}: {e}")

                        # Apply status filter
                        if status_filters_norm:
                            if status_text.lower() not in status_filters_norm:
                                continue  # Skip if status doesn't match filter

                        # Add to results
                        results.append(
                            ServiceEntry(
                                client_name=client.name,
                                client_dob=client.dob,
                                service_date=dt,
                                service_type=payer_text,
                                notes=status_text
                            )
                        )
                        self._log(
                            f"[BILLING] Captured DOS {dt.strftime('%m/%d/%Y')} | payer='{payer_text}' status='{status_text}'."
                        )

                    except StaleElementReferenceException:
                        # Row became stale, re-fetch all links and retry
                        self._log(f"[BILLING] Stale element detected, re-fetching rows...")
                        processed_all = False
                        time.sleep(0.3)
                        break  # Break inner loop to re-fetch
                    except Exception as e:
                        self._log(f"[BILLING][WARN] Error processing row {link_idx}: {e}")
                        continue

                # If we processed all rows without stale element issues, we're done
                if processed_all:
                    break

            except Exception as e:
                self._log(f"[BILLING][WARN] Error in fetch attempt {attempt + 1}: {e}")
                if attempt < max_attempts - 1:
                    time.sleep(0.5)
                    continue
                else:
                    break

        self._log(
            f"[BILLING] Collected {len(results)} transactions for {client.name} within {start_date} - {end_date}."
        )
        return results

    def _scroll_to_load_all_rows(self) -> None:
        """Scroll through the billing table to ensure all rows are loaded."""
        if not self.driver:
            return
        
        try:
            # Find the table container or scrollable element
            # Try to find the table or its container
            table_selectors = [
                "table",
                "[data-testid*='billing']",
                ".table-container",
                ".billing-table",
                "div[role='table']",
            ]
            
            scrollable_element = None
            for selector in table_selectors:
                try:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    for elem in elements:
                        # Check if element contains billing date links
                        if elem.find_elements(By.CSS_SELECTOR, "[data-testid='billingstatementtable-paymentdate-link']"):
                            scrollable_element = elem
                            break
                    if scrollable_element:
                        break
                except Exception:
                    continue
            
            # If no specific table found, use the body or a main container
            if not scrollable_element:
                try:
                    scrollable_element = self.driver.find_element(By.TAG_NAME, "body")
                except Exception:
                    return

            last_count = 0
            scroll_attempts = 0
            max_scroll_attempts = 50  # Prevent infinite scrolling
            
            while scroll_attempts < max_scroll_attempts:
                # Get current count of date links
                current_links = self.driver.find_elements(
                    By.CSS_SELECTOR,
                    "[data-testid='billingstatementtable-paymentdate-link']"
                )
                current_count = len(current_links)
                
                # If count hasn't changed, we've loaded all rows
                if current_count == last_count and scroll_attempts > 0:
                    break
                
                last_count = current_count
                
                # Scroll to the last visible date link
                if current_links:
                    try:
                        # Scroll the last link into view
                        self.driver.execute_script(
                            "arguments[0].scrollIntoView({behavior: 'smooth', block: 'end'});",
                            current_links[-1]
                        )
                        time.sleep(0.3)  # Wait for potential lazy loading
                    except Exception:
                        # Fallback: scroll the page down
                        self.driver.execute_script(
                            "window.scrollBy(0, 500);"
                        )
                        time.sleep(0.3)
                else:
                    # No links found, try scrolling the page
                    self.driver.execute_script("window.scrollBy(0, 500);")
                    time.sleep(0.3)
                
                scroll_attempts += 1
                
                # Also try scrolling the scrollable element if it's different from body
                if scrollable_element and scrollable_element.tag_name != "body":
                    try:
                        self.driver.execute_script(
                            "arguments[0].scrollTop = arguments[0].scrollHeight;",
                            scrollable_element
                        )
                        time.sleep(0.2)
                    except Exception:
                        pass
            
            # Scroll back to top to ensure we can access all elements
            try:
                self.driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(0.2)
            except Exception:
                pass
            
            final_count = len(self.driver.find_elements(
                By.CSS_SELECTOR,
                "[data-testid='billingstatementtable-paymentdate-link']"
            ))
            self._log(f"[BILLING] Scrolling complete. Found {final_count} total date links after scrolling.")
            
        except Exception as e:
            self._log(f"[BILLING][WARN] Error during scrolling: {e}. Continuing with available rows...")

    def close(self) -> None:
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None

    def _ensure_driver(self):
        if self.driver:
            return self.driver
        if not SELENIUM_AVAILABLE:
            raise RuntimeError("Selenium is required to instantiate a browser driver.")
        if ChromeOptions is None or webdriver is None:
            raise RuntimeError("Selenium Chrome dependencies are unavailable.")

        options = ChromeOptions()
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-infobars")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--no-sandbox")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)

        self.driver = webdriver.Chrome(options=options)
        self.driver.maximize_window()
        self.driver.implicitly_wait(5)
        return self.driver

    def _wait_for_patient_detail(self) -> None:
        """Wait for patient detail page to load and navigate to Billing tab."""
        if not self.driver:
            return
        
        wait = WebDriverWait(self.driver, 15)
        
        # Wait for dropdown to disappear (indicates navigation started)
        try:
            wait.until(EC.invisibility_of_element_located((By.ID, "ContentBubbleResultsContainer")))
        except TimeoutException:
            self._log("[NAV] Dropdown did not hide after selection; continuing.")
        
        # Wait for patient detail page to load - look for any tab element as indicator
        try:
            # Wait for tabs to appear (more reliable than waiting for specific tab)
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "a[data-tab-id]")))
        except TimeoutException:
            self._log("[NAV][WARN] Patient detail tabs did not appear after opening patient chart.")
            return
        
        # Now find and click the Billing tab
        try:
            # Use a shorter wait since we know tabs are present
            billing_wait = WebDriverWait(self.driver, 5)
            billing_tab = billing_wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "a[data-tab-id='Patient Billing']"))
            )
            
            # Scroll into view if needed and click
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'instant'});", billing_tab)
            time.sleep(0.1)
            billing_tab.click()
            
            # Wait for URL to change (indicates tab loaded)
            wait.until(EC.url_contains("Patient+Billing"))
            self._log("[NAV] Billing tab opened.")
            
            # Reduced wait - just enough for page to stabilize
            time.sleep(0.3)
            self._activate_all_items_filter()
        except TimeoutException:
            self._log("[NAV][WARN] Billing tab did not become clickable or URL did not change.")
        except Exception as exc:
            self._log(f"[NAV][WARN] Failed to activate Billing tab: {exc}")

    def _activate_all_items_filter(self) -> None:
        """Activate the 'All Items' filter on billing transactions.
        
        This function checks if the filter is already active, waits briefly for dialogs to clear,
        and then clicks the filter if needed.
        """
        if not self.driver:
            return
        
        max_retries = 3
        
        for attempt in range(max_retries):
            try:
                # Brief wait for page to stabilize
                time.sleep(0.5)
                
                # Quick check for blocking dialogs - but don't wait forever
                try:
                    # Check if Dialog element exists and is visible (with short timeout)
                    dialog_wait = WebDriverWait(self.driver, 2)
                    dialog = self.driver.find_element(By.CSS_SELECTOR, "div.Dialog")
                    if dialog.is_displayed():
                        # Dialog is visible, try to close it with Escape
                        self.driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                        time.sleep(0.5)
                        # Wait briefly for it to disappear
                        try:
                            dialog_wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.Dialog")))
                        except TimeoutException:
                            # Dialog didn't disappear, but continue anyway
                            pass
                except Exception:
                    # No dialog found or other issue - that's fine, continue
                    pass
                
                # Find the All Items element
                wait = WebDriverWait(self.driver, 10)
                all_items = wait.until(
                    EC.presence_of_element_located((By.ID, "SearchBillingTransactionsFilter__AllItems"))
                )
                
                # Check if it's already selected (has "selected" class)
                classes = all_items.get_attribute("class") or ""
                if "selected" in classes:
                    self._log("[NAV] 'All Items' filter is already active.")
                    return
                
                # Scroll element into view
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'instant'});", all_items)
                time.sleep(0.2)
                
                # Try to click - use multiple strategies
                clicked = False
                
                # Strategy 1: Regular click
                try:
                    wait_clickable = WebDriverWait(self.driver, 3)
                    clickable_element = wait_clickable.until(
                        EC.element_to_be_clickable((By.ID, "SearchBillingTransactionsFilter__AllItems"))
                    )
                    clickable_element.click()
                    clicked = True
                    self._log("[NAV] Applied 'All Items' filter on billing transactions.")
                except Exception as click_exc:
                    # Strategy 2: JavaScript click
                    try:
                        self.driver.execute_script("arguments[0].click();", all_items)
                        clicked = True
                        self._log("[NAV] Applied 'All Items' filter using JavaScript click.")
                    except Exception as js_exc:
                        if attempt < max_retries - 1:
                            self._log(f"[NAV] Click failed, retrying... (attempt {attempt + 1}/{max_retries}): {js_exc}")
                            time.sleep(0.5)
                            continue
                        else:
                            self._log(f"[NAV][WARN] Failed to click 'All Items' filter after all retries: {js_exc}")
                            return
                
                if clicked:
                    # Wait for the page to reload after applying filter
                    # The table needs time to refresh with the new filter applied
                    try:
                        wait_table = WebDriverWait(self.driver, 15)
                        
                        # First, wait for any loading indicators to disappear
                        try:
                            # Wait for common loading indicators to be gone
                            wait_table.until_not(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "[class*='loading'], [class*='spinner'], [aria-busy='true']"))
                            )
                        except (TimeoutException, Exception):
                            # No loading indicator found or it's already gone - that's fine
                            pass
                        
                        # Wait a moment for the filter to take effect
                        time.sleep(0.5)
                        
                        # Wait for the billing table to be present (it might be empty, that's okay)
                        try:
                            wait_table.until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "table, [data-testid*='billing'], [class*='table'], [class*='billing']"))
                            )
                        except TimeoutException:
                            # Table not found, but continue anyway
                            pass
                        
                        # Additional wait to ensure the DOM has updated after filter application
                        # This gives the page time to reload the table content
                        time.sleep(1.5)
                        
                        # Verify the table is ready by checking if we can find table elements
                        # (even if empty, the structure should be there)
                        try:
                            # Check if table structure exists
                            table_elements = self.driver.find_elements(By.CSS_SELECTOR, "table, [data-testid*='billing']")
                            if table_elements:
                                self._log("[NAV] Billing table structure ready after applying 'All Items' filter.")
                            else:
                                self._log("[NAV][WARN] Billing table structure not found, but continuing...")
                        except Exception:
                            pass
                            
                    except Exception as wait_exc:
                        # If waiting fails, log but continue anyway with a fallback wait
                        self._log(f"[NAV][WARN] Could not verify table load after filter: {wait_exc}")
                        time.sleep(2.0)  # Longer fallback wait
                    
                    return
                    
            except TimeoutException:
                if attempt < max_retries - 1:
                    self._log(f"[NAV] 'All Items' filter button not found, retrying... (attempt {attempt + 1}/{max_retries})")
                    time.sleep(0.5)
                    continue
                else:
                    self._log("[NAV][WARN] 'All Items' filter button not found on billing tab after all retries.")
                    return
            except Exception as exc:
                if attempt < max_retries - 1:
                    self._log(f"[NAV] Error activating 'All Items' filter, retrying... (attempt {attempt + 1}/{max_retries}): {exc}")
                    time.sleep(0.5)
                    continue
                else:
                    self._log(f"[NAV][WARN] Failed to activate 'All Items' filter after all retries: {exc}")
                    return

    def return_to_patients_sidebar(self) -> None:
        """Return to the Patients list page quickly using the fastest method available."""
        if not self.driver:
            return
        
        # Check if we're already on the patients LIST page (not a patient detail page)
        # Only skip navigation if URL ends with /app/patients/ or /app/patients (exact list page)
        try:
            current_url = self.driver.current_url.rstrip('/')
            if current_url.endswith("/app/patients"):
                self._log("[NAV] Already on Patients list page.")
                return
        except Exception:
            pass
        
        try:
            # Strategy 1: Direct URL navigation (fastest - no waiting for elements)
            try:
                # Extract base URL and navigate directly
                current_url = self.driver.current_url
                if "/app/" in current_url:
                    base_url = current_url.split("/app/")[0]
                    patients_url = base_url + "/app/patients/"
                    self.driver.get(patients_url)
                    # Wait briefly for navigation to start
                    time.sleep(0.1)
                else:
                    raise Exception("Could not extract base URL")
            except Exception as nav_exc:
                # Strategy 2: Try sidebar link click (fallback)
                self._log("[NAV] Direct navigation failed, trying sidebar link...")
                wait = WebDriverWait(self.driver, 3)
                patients_link = wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "a.sidebar-link[href='/app/patients/']"))
                )
                # Use JavaScript click for faster execution
                self.driver.execute_script("arguments[0].click();", patients_link)
            
            # Wait for URL to change to the patients list page
            url_wait = WebDriverWait(self.driver, 5)
            url_wait.until(lambda driver: driver.current_url.rstrip('/').endswith("/app/patients"))
            self._log("[NAV] Returned to Patients list.")
            
            # Minimal wait - just enough for page to be ready
            time.sleep(0.1)
            
        except TimeoutException:
            # Even if timeout, check if we're on the right page now
            try:
                current_url = self.driver.current_url.rstrip('/')
                if current_url.endswith("/app/patients"):
                    self._log("[NAV] Returned to Patients list (timeout but URL correct).")
                else:
                    self._log("[NAV][WARN] Unable to return to Patients list; continuing anyway.")
            except Exception:
                self._log("[NAV][WARN] Unable to return to Patients list; continuing anyway.")
        except Exception as exc:
            self._log(f"[NAV][WARN] Error while returning to Patients list: {exc}")


# Main bot --------------------------------------------------------------------

class TherapyNotesRecordsBot:
    """GUI entry point orchestrating document parsing and TherapyNotes fetching."""

    def __init__(self) -> None:
        self.root: Optional[tk.Tk] = None
        self.log_widget: Optional[scrolledtext.ScrolledText] = None
        self.status_label: Optional[tk.Label] = None
        self.user_dropdown: Optional[ttk.Combobox] = None
        self.username_entry: Optional[ttk.Entry] = None
        self.password_entry: Optional[ttk.Entry] = None

        self.document_path_var: Optional[tk.StringVar] = None
        self.start_date_var: Optional[tk.StringVar] = None
        self.end_date_var: Optional[tk.StringVar] = None
        self.username_var: Optional[tk.StringVar] = None
        self.password_var: Optional[tk.StringVar] = None
        self.selected_user_var: Optional[tk.StringVar] = None
        self.output_dir_var: Optional[tk.StringVar] = None
        self.status_var: Optional[tk.StringVar] = None
        self.payer_filters_var: Optional[tk.StringVar] = None  # Deprecated - kept for backward compatibility
        self.payer_filter_vars: Dict[str, tk.BooleanVar] = {}
        self.status_filter_vars: Dict[str, tk.BooleanVar] = {}
        self.status_filters_frame: Optional[tk.Frame] = None
        self.payer_filter_popup: Optional[tk.Toplevel] = None

        self.parser = DocumentParser()
        self.client_fetcher: Optional[TherapyNotesClientFetcher] = None
        self.current_client: Optional[ClientMetadata] = None
        self.service_entries: List[ServiceEntry] = []
        self.client_record_queue: List[Dict[str, Any]] = []
        self.active_record: Optional[Dict[str, Any]] = None

        self.settings_path = Path(__file__).parent / "therapy_notes_records_settings.json"
        self.users_file = Path(__file__).parent / "therapy_notes_records_users.json"
        self.settings: Dict[str, Any] = self._load_settings()
        self.users: Dict[str, Dict[str, str]] = self._load_users()

        self._stop_event = threading.Event()
        self._worker_thread: Optional[threading.Thread] = None
        self._run_button: Optional[tk.Button] = None
        self._stop_button: Optional[tk.Button] = None

    # GUI construction ----------------------------------------------------
    def launch(self) -> None:
        self.root = tk.Tk()
        self.root.title("TherapyNotes Medical Records Billing Log Bot - Version 3.1.0, Last Updated 12/04/2025")
        self.root.configure(bg="#f0f0f0")
        self.root.minsize(1150, 680)
        self.root.geometry("1300x700")

        last_user = self.settings.get("last_user")
        default_username = ""
        default_password = ""
        if last_user and last_user in self.users:
            default_username = self.users[last_user].get("username", "")
            default_password = self.users[last_user].get("password", "")
        else:
            default_username = self.settings.get("last_username", "")

        self.document_path_var = tk.StringVar(self.root, value="")
        self.start_date_var = tk.StringVar(self.root, value="")
        self.end_date_var = tk.StringVar(self.root, value="")
        self.username_var = tk.StringVar(self.root, value=default_username)
        self.password_var = tk.StringVar(self.root, value=default_password)
        self.selected_user_var = tk.StringVar(self.root, value=last_user or "")
        self.output_dir_var = tk.StringVar(
            self.root,
            value=self.settings.get(
                "output_dir",
                str((Path(__file__).parent / "outputs").resolve())
            )
        )
        self.status_var = tk.StringVar(self.root, value="Ready. Configure inputs and press Run.")
        self.payer_filters_var = tk.StringVar(
            self.root,
            value=self.settings.get("payer_filters", "")
        )
        self._initialize_payer_filters()
        default_login_url = self.settings.get(
            "login_url",
            TherapyNotesClientFetcher.DEFAULT_LOGIN_URL
        )
        self.login_url_var = tk.StringVar(self.root, value=default_login_url)
        self._initialize_status_filters()

        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self._build_gui()
        self._update_user_dropdown(select_name=self.selected_user_var.get() or None)
        self._update_payer_filter_status()
        self._append_log("TherapyNotes Records Bot ready.")
        self._set_status("Ready. Configure inputs and press Run.")
        self.root.mainloop()

    def _build_gui(self) -> None:
        assert self.root is not None
        assert self.document_path_var is not None
        assert self.start_date_var is not None
        assert self.end_date_var is not None
        assert self.username_var is not None
        assert self.password_var is not None
        assert self.output_dir_var is not None
        assert self.status_var is not None
        assert self.selected_user_var is not None
        assert self.payer_filters_var is not None

        style = ttk.Style(self.root)
        style.configure("Card.TLabelframe", font=("Segoe UI", 11, "bold"))
        style.configure("Card.TLabelframe.Label", font=("Segoe UI", 11, "bold"))
        style.configure("Card.TFrame", background="white")
        style.configure("Accent.TButton", font=("Segoe UI Semibold", 10))

        header = tk.Frame(self.root, bg="#660000")
        header.pack(fill="x")
        tk.Label(
            header,
            text="TherapyNotes Medical Records Billing Log Bot",
            bg="#660000",
            fg="white",
            font=("Segoe UI", 16, "bold"),
            pady=10
        ).pack(side="left", padx=20)

        content = tk.Frame(self.root, bg="#f0f0f0")
        content.pack(fill="both", expand=True, padx=20, pady=15)
        content.columnconfigure(0, weight=3)
        content.columnconfigure(1, weight=2)

        controls_container = tk.Frame(content, bg="#f0f0f0")
        controls_container.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        controls_container.columnconfigure(0, weight=1)
        controls_container.rowconfigure(0, weight=1)

        controls_canvas = tk.Canvas(controls_container, bg="#f0f0f0", highlightthickness=0)
        controls_canvas.grid(row=0, column=0, sticky="nsew")
        controls_scrollbar = ttk.Scrollbar(controls_container, orient="vertical", command=controls_canvas.yview)
        controls_scrollbar.grid(row=0, column=1, sticky="ns")
        controls_canvas.configure(yscrollcommand=controls_scrollbar.set)

        controls_frame = tk.Frame(controls_canvas, bg="#f0f0f0")
        # Create window with padding to ensure headers aren't cut off
        controls_canvas.create_window((0, 0), window=controls_frame, anchor="nw")
        
        # Update scroll region when frame size changes
        def update_scroll_region(event=None):
            controls_canvas.configure(scrollregion=controls_canvas.bbox("all"))
        controls_frame.bind("<Configure>", update_scroll_region)

        def _on_mousewheel(event):
            controls_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        controls_frame.bind_all("<MouseWheel>", _on_mousewheel)

        log_frame = tk.Frame(content, bg="#f0f0f0")
        log_frame.grid(row=0, column=1, sticky="nsew")
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(1, weight=1)

        doc_card = tk.LabelFrame(
            controls_frame,
            text="Insurer Document",
            bg="#ffffff",
            fg="#660000",
            font=("Segoe UI", 12, "bold"),
            padx=15,
            pady=12,
            labelanchor="n"  # Ensure label is at the top
        )
        doc_card.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        doc_card.columnconfigure(0, weight=1)

        entry_doc = ttk.Entry(doc_card, textvariable=self.document_path_var, font=("Segoe UI", 10))
        entry_doc.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        ttk.Button(
            doc_card,
            text="Browse for Document",
            command=self._browse_document,
            style="Accent.TButton"
        ).grid(row=1, column=0, sticky="w")

        # PDF Reading Status Checker
        pdf_status_card = tk.LabelFrame(
            controls_frame,
            text="PDF Reading Status",
            bg="#ffffff",
            fg="#660000",
            font=("Segoe UI", 12, "bold"),
            padx=15,
            pady=12
        )
        pdf_status_card.grid(row=1, column=0, sticky="ew", pady=12)
        pdf_status_card.columnconfigure(0, weight=1)

        self.pdf_status_frame = tk.Frame(pdf_status_card, bg="#ffffff")
        self.pdf_status_frame.grid(row=0, column=0, sticky="ew")
        self.pdf_status_frame.columnconfigure(1, weight=1)

        # Status labels (will be populated by _check_pdf_reading_status)
        self.pdfplumber_status_label = tk.Label(
            self.pdf_status_frame,
            text="Checking...",
            font=("Segoe UI", 9),
            bg="#ffffff",
            anchor="w"
        )
        self.pdfplumber_status_label.grid(row=0, column=0, sticky="w", pady=2)

        self.pytesseract_status_label = tk.Label(
            self.pdf_status_frame,
            text="Checking...",
            font=("Segoe UI", 9),
            bg="#ffffff",
            anchor="w"
        )
        self.pytesseract_status_label.grid(row=1, column=0, sticky="w", pady=2)

        self.tesseract_status_label = tk.Label(
            self.pdf_status_frame,
            text="Checking...",
            font=("Segoe UI", 9),
            bg="#ffffff",
            anchor="w"
        )
        self.tesseract_status_label.grid(row=2, column=0, sticky="w", pady=2)

        self.poppler_status_label = tk.Label(
            self.pdf_status_frame,
            text="Checking...",
            font=("Segoe UI", 9),
            bg="#ffffff",
            anchor="w"
        )
        self.poppler_status_label.grid(row=3, column=0, sticky="w", pady=2)

        self.pdf_overall_status_label = tk.Label(
            pdf_status_card,
            text="Checking...",
            font=("Segoe UI", 10, "bold"),
            bg="#ffffff",
            anchor="w"
        )
        self.pdf_overall_status_label.grid(row=1, column=0, sticky="w", pady=(8, 0))

        ttk.Button(
            pdf_status_card,
            text="Refresh Status",
            command=self._check_pdf_reading_status,
            style="Accent.TButton"
        ).grid(row=2, column=0, sticky="w", pady=(8, 0))

        # Initial status check (defer until after GUI is fully built)
        self.root.after(100, self._check_pdf_reading_status)

        date_card = tk.LabelFrame(
            controls_frame,
            text="Date Range (Manual Override)",
            bg="#ffffff",
            fg="#660000",
            font=("Segoe UI", 12, "bold"),
            padx=15,
            pady=12
        )
        date_card.grid(row=2, column=0, sticky="ew", pady=12)
        for i in range(2):
            date_card.columnconfigure(i, weight=1)

        ttk.Label(date_card, text="Start Date (MM/DD/YYYY):", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w")
        ttk.Entry(date_card, textvariable=self.start_date_var, font=("Segoe UI", 10)).grid(row=1, column=0, sticky="ew", padx=(0, 10))

        ttk.Label(date_card, text="End Date (MM/DD/YYYY):", font=("Segoe UI", 10)).grid(row=0, column=1, sticky="w")
        ttk.Entry(date_card, textvariable=self.end_date_var, font=("Segoe UI", 10)).grid(row=1, column=1, sticky="ew")

        output_card = tk.LabelFrame(
            controls_frame,
            text="Output Location",
            bg="#ffffff",
            fg="#660000",
            font=("Segoe UI", 12, "bold"),
            padx=15,
            pady=12
        )
        output_card.grid(row=3, column=0, sticky="ew")
        output_card.columnconfigure(0, weight=1)

        ttk.Entry(output_card, textvariable=self.output_dir_var, font=("Segoe UI", 10)).grid(row=0, column=0, sticky="ew", pady=(0, 10))
        ttk.Button(
            output_card,
            text="Browse Output Folder",
            command=self._browse_output_directory,
            style="Accent.TButton"
        ).grid(row=1, column=0, sticky="w")

        payer_card = tk.LabelFrame(
            controls_frame,
            text="Primary Payer Filters",
            bg="#ffffff",
            fg="#660000",
            font=("Segoe UI", 12, "bold"),
            padx=15,
            pady=12
        )
        payer_card.grid(row=4, column=0, sticky="ew", pady=12)
        payer_card.columnconfigure(0, weight=1)

        tk.Label(
            payer_card,
            text="Select primary payers to include. Leave all unchecked to include all payers:",
            font=("Segoe UI", 10),
            bg="#ffffff",
            anchor="w"
        ).grid(row=0, column=0, sticky="w", pady=(0, 6))

        self.payer_filter_status_label = tk.Label(
            payer_card,
            text="No payers selected (all will be included)",
            font=("Segoe UI", 9),
            bg="#ffffff",
            fg="#666666",
            anchor="w"
        )
        self.payer_filter_status_label.grid(row=1, column=0, sticky="w", pady=(0, 6))

        tk.Button(
            payer_card,
            text="Configure Primary Payer Filters",
            command=self._open_payer_filter_popup,
            bg="#800000",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            padx=15,
            pady=8,
            cursor="hand2"
        ).grid(row=2, column=0, sticky="w")

        status_card = tk.LabelFrame(
            controls_frame,
            text="Insurance Status Filters",
            bg="#ffffff",
            fg="#660000",
            font=("Segoe UI", 12, "bold"),
            padx=15,
            pady=12,
            labelanchor="n"  # Ensure label is at the top
        )
        status_card.grid(row=5, column=0, sticky="ew", pady=12)
        status_card.columnconfigure(0, weight=1)

        self.status_filters_frame = tk.Frame(status_card, bg="#ffffff")
        self.status_filters_frame.grid(row=0, column=0, sticky="ew", pady=(5, 0))  # Add top padding to avoid header overlap
        self._rebuild_status_checkboxes()

        tk.Button(
            status_card,
            text="Add Status",
            command=self._add_ins_status,
            bg="#800000",
            fg="white",
            font=("Segoe UI", 9, "bold"),
            relief="flat",
            padx=10,
            pady=4,
            cursor="hand2"
        ).grid(row=1, column=0, sticky="e", pady=(8, 0))

        cred_card = tk.LabelFrame(
            controls_frame,
            text="TherapyNotes Credentials",
            bg="#ffffff",
            fg="#660000",
            font=("Segoe UI", 12, "bold"),
            padx=15,
            pady=12,
            labelanchor="n"  # Ensure label is at the top
        )
        cred_card.grid(row=6, column=0, sticky="ew", pady=12)
        cred_card.columnconfigure(0, weight=1)
        cred_card.columnconfigure(1, weight=1)

        ttk.Label(cred_card, text="Login URL:", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w", pady=(0, 4))
        ttk.Entry(cred_card, textvariable=self.login_url_var, font=("Segoe UI", 10)).grid(row=0, column=1, sticky="ew", pady=(0, 4))

        user_row = tk.Frame(cred_card, bg="#ffffff")
        user_row.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        user_row.columnconfigure(1, weight=1)

        tk.Label(
            user_row,
            text="Saved User:",
            font=("Segoe UI", 10),
            bg="#ffffff"
        ).grid(row=0, column=0, sticky="w")

        self.user_dropdown = ttk.Combobox(
            user_row,
            font=("Segoe UI", 10),
            width=25,
            state="readonly"
        )
        self.user_dropdown.grid(row=0, column=1, sticky="ew", padx=(10, 10))
        self.user_dropdown.bind("<<ComboboxSelected>>", lambda _: self._on_user_selected())

        add_user_btn = tk.Button(
            user_row,
            text="Add User",
            command=self._add_user,
            bg="#800000",
            fg="white",
            font=("Segoe UI", 9),
            padx=10,
            pady=3,
            cursor="hand2",
            relief="flat",
            bd=0
        )
        add_user_btn.grid(row=0, column=2, padx=(0, 5))

        update_user_btn = tk.Button(
            user_row,
            text="Update User",
            command=self._update_user,
            bg="#666666",
            fg="white",
            font=("Segoe UI", 9),
            padx=10,
            pady=3,
            cursor="hand2",
            relief="flat",
            bd=0
        )
        update_user_btn.grid(row=0, column=3, padx=(0, 5))

        delete_user_btn = tk.Button(
            user_row,
            text="Delete User",
            command=self._delete_user,
            bg="#990000",
            fg="white",
            font=("Segoe UI", 9),
            padx=10,
            pady=3,
            cursor="hand2",
            relief="flat",
            bd=0
        )
        delete_user_btn.grid(row=0, column=4, padx=(0, 5))

        ttk.Label(cred_card, text="Username:", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="w", pady=(5, 0))
        self.username_entry = ttk.Entry(cred_card, textvariable=self.username_var, font=("Segoe UI", 10))
        self.username_entry.grid(row=4, column=0, sticky="ew", pady=(0, 10), padx=(0, 10))

        ttk.Label(cred_card, text="Password:", font=("Segoe UI", 10)).grid(row=3, column=1, sticky="w", pady=(5, 0))
        self.password_entry = ttk.Entry(cred_card, textvariable=self.password_var, font=("Segoe UI", 10), show="•")
        self.password_entry.grid(row=4, column=1, sticky="ew", pady=(0, 10))

        run_frame = tk.Frame(controls_frame, bg="#f0f0f0")
        run_frame.grid(row=7, column=0, sticky="ew", pady=(10, 0))
        self._run_button = tk.Button(
            run_frame,
            text="🚀 Run Billing Log Collection",
            command=self._on_run_clicked,
            bg="#800000",
            fg="white",
            font=("Segoe UI", 12, "bold"),
            relief="raised",
            padx=20,
            pady=10,
            cursor="hand2"
        )
        self._run_button.pack(fill="x")

        self._stop_button = tk.Button(
            run_frame,
            text="⛔ Stop & Export",
            command=self._on_stop_clicked,
            bg="#a52a2a",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            relief="raised",
            padx=20,
            pady=8,
            cursor="hand2",
            state=tk.DISABLED
        )
        self._stop_button.pack(fill="x", pady=(8, 0))

        self.status_label = tk.Label(
            run_frame,
            textvariable=self.status_var,
            font=("Segoe UI", 10),
            bg="#f0f0f0",
            fg="gray"
        )
        self.status_label.pack(anchor="w", pady=(8, 0))

        tk.Label(
            log_frame,
            text="Run Log",
            font=("Segoe UI", 12, "bold"),
            bg="#f0f0f0",
            fg="#660000"
        ).grid(row=0, column=0, sticky="w", pady=(0, 5))

        log_container = tk.Frame(log_frame, bg="#ffffff", bd=1, relief="solid")
        log_container.grid(row=1, column=0, sticky="nsew")
        log_container.rowconfigure(0, weight=1)
        log_container.columnconfigure(0, weight=1)

        self.log_widget = scrolledtext.ScrolledText(
            log_container,
            width=70,
            height=25,
            state=tk.DISABLED,
            font=("Consolas", 10),
            bg="#ffffff"
        )
        self.log_widget.grid(row=0, column=0, sticky="nsew")

        help_label = tk.Label(
            log_frame,
            text="Tip: Use saved users to store credentials securely on this workstation.",
            font=("Segoe UI", 9),
            bg="#f0f0f0",
            fg="gray"
        )
        help_label.grid(row=2, column=0, sticky="w", pady=(6, 0))

    # Settings & user management -----------------------------------------
    def _load_settings(self) -> Dict[str, Any]:
        try:
            if self.settings_path.exists():
                with self.settings_path.open("r", encoding="utf-8") as fh:
                    data = json.load(fh)
                    if isinstance(data, dict):
                        return data
        except Exception as exc:  # noqa: BLE001
            logger.warning("Failed to load settings: %s", exc)
        return {}

    def _save_settings(self) -> None:
        try:
            payload: Dict[str, Any] = {}
            if self.document_path_var is not None:
                payload["last_document"] = self.document_path_var.get()
            if self.start_date_var is not None:
                payload["start_date"] = self.start_date_var.get()
            if self.end_date_var is not None:
                payload["end_date"] = self.end_date_var.get()
            if self.output_dir_var is not None:
                payload["output_dir"] = self.output_dir_var.get()
            if self.selected_user_var is not None:
                payload["last_user"] = self.selected_user_var.get()
            if self.username_var is not None:
                payload["last_username"] = self.username_var.get()
            if self.payer_filters_var is not None:
                payload["payer_filters"] = self.payer_filters_var.get()
            # Save selected primary payers
            payload["primary_payers_selected"] = [
                payer for payer, var in self.payer_filter_vars.items() if var.get()
            ]
            if self.login_url_var is not None:
                payload["login_url"] = self.login_url_var.get()
            payload["insurance_statuses"] = sorted(self.status_filter_vars.keys())
            payload["insurance_status_selected"] = [
                status for status, var in self.status_filter_vars.items() if var.get()
            ]
            with self.settings_path.open("w", encoding="utf-8") as fh:
                json.dump(payload, fh, indent=2)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Failed to save settings: %s", exc)

    def _load_users(self) -> Dict[str, Dict[str, str]]:
        try:
            if self.users_file.exists():
                with self.users_file.open("r", encoding="utf-8") as fh:
                    data = json.load(fh)
                    if isinstance(data, dict):
                        return {k: v for k, v in data.items() if isinstance(v, dict)}
        except Exception as exc:  # noqa: BLE001
            logger.warning("Failed to load users: %s", exc)
        return {}

    def _save_users(self) -> None:
        try:
            with self.users_file.open("w", encoding="utf-8") as fh:
                json.dump(self.users, fh, indent=2)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Failed to save users: %s", exc)

    def _build_record_queue(self) -> List[Dict[str, Any]]:
        queue: List[Dict[str, Any]] = []
        if not self.current_client:
            return queue
        extras = getattr(self.current_client, "extras", {}) or {}
        raw_records = extras.get("records") or []
        for entry in raw_records:
            if not isinstance(entry, dict):
                continue
            name = (entry.get("name")
                    or entry.get("Patient Name")
                    or entry.get("Member Name")
                    or entry.get("Client Name")
                    or "").strip()
            if not name:
                continue
            dob_raw = (entry.get("dob")
                       or entry.get("Patient Date of Birth")
                       or entry.get("DOB")
                       or entry.get("Date of Birth")
                       or "").strip()
            dob_norm = DocumentParser._normalize_date(dob_raw) if dob_raw else ""
            record: Dict[str, Any] = {
                "name": name,
                "dob": dob_norm,
                "dob_raw": dob_raw,
                "source": entry,
            }
            # Include date range if present in the record
            if "date_range" in entry:
                record["date_range"] = entry["date_range"]
            queue.append(record)
        deduped: List[Dict[str, Any]] = []
        seen: set[Tuple[str, str]] = set()
        for record in queue:
            key = (record["name"].lower(), (record.get("dob") or record.get("dob_raw") or "").lower())
            if key in seen:
                continue
            seen.add(key)
            deduped.append(record)
        return deduped

    def _initialize_payer_filters(self) -> None:
        """Initialize payer filter checkboxes from settings."""
        saved_selected = self.settings.get("primary_payers_selected") or []
        selected_set = {str(payer).strip() for payer in saved_selected}
        for payer in PRIMARY_PAYER_OPTIONS:
            self.payer_filter_vars[payer] = tk.BooleanVar(
                value=payer in selected_set
            )

    def _get_payer_filters(self) -> List[str]:
        """Get list of selected primary payers from checkboxes."""
        return [
            payer
            for payer, var in self.payer_filter_vars.items()
            if var.get()
        ]

    def _get_selected_ins_statuses(self) -> List[str]:
        return [
            status
            for status, var in self.status_filter_vars.items()
            if var.get()
        ]

    def _initialize_status_filters(self) -> None:
        saved_statuses = self.settings.get("insurance_statuses")
        if not isinstance(saved_statuses, list) or not saved_statuses:
            saved_statuses = DEFAULT_INS_STATUS_FILTERS.copy()
        saved_selected = self.settings.get("insurance_status_selected") or []
        selected_set = {str(status).strip() for status in saved_selected}
        for status in saved_statuses:
            text = str(status).strip()
            if text:
                self._ensure_status_var(text, text in selected_set if selected_set else True)
        for default_status in DEFAULT_INS_STATUS_FILTERS:
            if default_status not in self.status_filter_vars:
                self._ensure_status_var(default_status, True)

    def _ensure_status_var(self, status: str, selected: bool = True) -> None:
        if status not in self.status_filter_vars:
            self.status_filter_vars[status] = tk.BooleanVar(value=selected)
        else:
            if selected:
                self.status_filter_vars[status].set(True)

    def _rebuild_status_checkboxes(self) -> None:
        if not self.status_filters_frame:
            return
        for child in self.status_filters_frame.winfo_children():
            child.destroy()
        for row_index, status in enumerate(sorted(self.status_filter_vars.keys(), key=str.lower)):
            var = self.status_filter_vars[status]
            ttk.Checkbutton(
                self.status_filters_frame,
                text=status,
                variable=var
            ).grid(row=row_index, column=0, sticky="w", pady=2)

    def _add_ins_status(self) -> None:
        if self.root is None:
            return
        status = simpledialog.askstring("Add Insurance Status", "Enter insurance status label:", parent=self.root)
        if not status:
            return
        status = status.strip()
        if not status:
            messagebox.showwarning("Invalid Status", "Status label cannot be empty.")
            return
        if status in self.status_filter_vars:
            messagebox.showinfo("Exists", f"'{status}' is already in the list.")
            return
        self._ensure_status_var(status, True)
        self._rebuild_status_checkboxes()
        self._append_log(f"Added insurance status filter '{status}'.")

    def _open_payer_filter_popup(self) -> None:
        """Open popup window for selecting primary payer filters."""
        if self.root is None:
            return
        
        # Close existing popup if open
        if self.payer_filter_popup is not None:
            try:
                self.payer_filter_popup.destroy()
            except Exception:
                pass
        
        # Create popup window
        popup = tk.Toplevel(self.root)
        popup.title("Primary Payer Filters - Version 3.1.0, Last Updated 12/04/2025")
        popup.configure(bg="#f0f0f0")
        popup.geometry("600x700")
        popup.transient(self.root)
        popup.grab_set()
        
        self.payer_filter_popup = popup
        
        # Header
        header_frame = tk.Frame(popup, bg="#660000")
        header_frame.pack(fill="x", padx=0, pady=0)
        tk.Label(
            header_frame,
            text="Select Primary Payers to Include",
            font=("Segoe UI", 14, "bold"),
            bg="#660000",
            fg="white",
            padx=15,
            pady=10
        ).pack()
        
        # Instructions
        instructions = tk.Label(
            popup,
            text="Check the primary payers you want to include. Leave all unchecked to include all payers.",
            font=("Segoe UI", 9),
            bg="#f0f0f0",
            fg="#333333",
            wraplength=550,
            justify="left"
        )
        instructions.pack(pady=10, padx=15, anchor="w")
        
        # Scrollable frame for checkboxes
        canvas = tk.Canvas(popup, bg="#ffffff", highlightthickness=0)
        scrollbar = ttk.Scrollbar(popup, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#ffffff")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Add checkboxes for each payer
        for payer in sorted(PRIMARY_PAYER_OPTIONS, key=str.lower):
            var = self.payer_filter_vars.get(payer)
            if var is None:
                var = tk.BooleanVar(value=False)
                self.payer_filter_vars[payer] = var
            
            checkbox = ttk.Checkbutton(
                scrollable_frame,
                text=payer,
                variable=var
            )
            checkbox.pack(anchor="w", padx=15, pady=2)
        
        canvas.pack(side="left", fill="both", expand=True, padx=(15, 0), pady=(0, 15))
        scrollbar.pack(side="right", fill="y", pady=(0, 15))
        
        # Button frame
        button_frame = tk.Frame(popup, bg="#f0f0f0")
        button_frame.pack(fill="x", padx=15, pady=10)
        
        # Select All / Deselect All buttons
        select_all_btn = tk.Button(
            button_frame,
            text="Select All",
            command=lambda: self._select_all_payers(True),
            bg="#4a4a4a",
            fg="white",
            font=("Segoe UI", 9),
            relief="flat",
            padx=10,
            pady=5,
            cursor="hand2"
        )
        select_all_btn.pack(side="left", padx=(0, 5))
        
        deselect_all_btn = tk.Button(
            button_frame,
            text="Deselect All",
            command=lambda: self._select_all_payers(False),
            bg="#4a4a4a",
            fg="white",
            font=("Segoe UI", 9),
            relief="flat",
            padx=10,
            pady=5,
            cursor="hand2"
        )
        deselect_all_btn.pack(side="left")
        
        # Apply and Cancel buttons
        button_frame2 = tk.Frame(popup, bg="#f0f0f0")
        button_frame2.pack(fill="x", padx=15, pady=(0, 15))
        
        def apply_filters():
            self._update_payer_filter_status()
            self._save_settings()
            self.payer_filter_popup = None
            popup.destroy()
        
        def cancel_popup():
            # Restore previous selections
            saved_selected = self.settings.get("primary_payers_selected") or []
            selected_set = {str(payer).strip() for payer in saved_selected}
            for payer, var in self.payer_filter_vars.items():
                var.set(payer in selected_set)
            self.payer_filter_popup = None
            popup.destroy()
        
        cancel_btn = tk.Button(
            button_frame2,
            text="Cancel",
            command=cancel_popup,
            bg="#666666",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2"
        )
        cancel_btn.pack(side="right", padx=(5, 0))
        
        apply_btn = tk.Button(
            button_frame2,
            text="Apply",
            command=apply_filters,
            bg="#800000",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            padx=20,
            pady=8,
            cursor="hand2"
        )
        apply_btn.pack(side="right")
        
        # Make canvas scrollable with mouse wheel
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
        def _on_popup_close():
            # Clean up bindings when popup closes
            try:
                popup.unbind_all("<MouseWheel>")
            except Exception:
                pass
            self.payer_filter_popup = None
        
        popup.bind("<MouseWheel>", _on_mousewheel)
        canvas.bind("<MouseWheel>", _on_mousewheel)
        popup.protocol("WM_DELETE_WINDOW", _on_popup_close)
        
        # Focus on popup
        popup.focus_set()

    def _select_all_payers(self, select: bool) -> None:
        """Select or deselect all payer checkboxes."""
        for var in self.payer_filter_vars.values():
            var.set(select)

    def _update_payer_filter_status(self) -> None:
        """Update the status label showing selected payers."""
        if not hasattr(self, 'payer_filter_status_label'):
            return
        selected = self._get_payer_filters()
        if not selected:
            self.payer_filter_status_label.config(
                text="No payers selected (all will be included)",
                fg="#666666"
            )
        elif len(selected) == 1:
            self.payer_filter_status_label.config(
                text=f"1 payer selected: {selected[0]}",
                fg="#006600"
            )
        else:
            self.payer_filter_status_label.config(
                text=f"{len(selected)} payers selected",
                fg="#006600"
            )

    def _update_user_dropdown(self, select_name: Optional[str] = None) -> None:
        if not self.user_dropdown or not self.selected_user_var:
            return
        user_names = sorted(self.users.keys())
        placeholder = "Select saved user..."
        display_values = [placeholder] + user_names if user_names else [placeholder]
        current = select_name or self.selected_user_var.get()

        self.user_dropdown["values"] = display_values

        if current in user_names:
            self.user_dropdown.set(current)
        else:
            self.user_dropdown.set(placeholder)
            self.selected_user_var.set("")
            current = ""

        if current:
            self._populate_credentials_from_user(current)
        else:
            self._save_settings()

    def _populate_credentials_from_user(self, user_name: str) -> None:
        if self.username_var is None or self.password_var is None:
            return
        creds = self.users.get(user_name, {})
        self.username_var.set(creds.get("username", ""))
        self.password_var.set(creds.get("password", ""))
        if self.login_url_var is not None:
            login_url = creds.get("login_url") or TherapyNotesClientFetcher.DEFAULT_LOGIN_URL
            self.login_url_var.set(login_url)
        if self.selected_user_var:
            self.selected_user_var.set(user_name)
        self._save_settings()

    def _on_user_selected(self) -> None:
        if not self.selected_user_var or not self.user_dropdown:
            return
        selected_display = self.user_dropdown.get().strip()
        placeholder = "Select saved user..."
        if selected_display == placeholder:
            self.selected_user_var.set("")
            if self.username_var is not None:
                self.username_var.set("")
            if self.password_var is not None:
                self.password_var.set("")
            self._append_log("Saved user selection cleared.")
            self._save_settings()
            return

        selected = selected_display
        self.selected_user_var.set(selected)
        if selected and selected in self.users:
            self._populate_credentials_from_user(selected)
            self._append_log(f"Loaded saved user '{selected}'.")
        else:
            self._append_log("Saved user selection cleared.")

    def _add_user(self) -> None:
        if self.root is None:
            return
        name = simpledialog.askstring("Add Saved User", "Enter a nickname for this credential set:", parent=self.root)
        if not name:
            return
        name = name.strip()
        if not name:
            messagebox.showwarning("Invalid Name", "User nickname cannot be empty.")
            return
        username = simpledialog.askstring("Username", "Enter TherapyNotes username:", parent=self.root)
        if username is None:
            return
        password = simpledialog.askstring("Password", "Enter TherapyNotes password:", parent=self.root, show="*")
        if password is None:
            return
        login_url = self.login_url_var.get().strip() if self.login_url_var else TherapyNotesClientFetcher.DEFAULT_LOGIN_URL
        self.users[name] = {"username": username.strip(), "password": password, "login_url": login_url}
        self._save_users()
        self._update_user_dropdown(select_name=name)
        self._append_log(f"Saved credentials for '{name}'.")

    def _update_user(self) -> None:
        if self.root is None or not self.selected_user_var:
            return
        selected = self.selected_user_var.get()
        if not selected:
            messagebox.showinfo("Select User", "Please select a saved user to update.")
            return
        username = simpledialog.askstring(
            "Update Username",
            "Update TherapyNotes username:",
            parent=self.root,
            initialvalue=self.users.get(selected, {}).get("username", "")
        )
        if username is None:
            return
        password = simpledialog.askstring(
            "Update Password",
            "Update TherapyNotes password:",
            parent=self.root,
            show="*",
            initialvalue=self.users.get(selected, {}).get("password", "")
        )
        if password is None:
            return
        login_url = self.login_url_var.get().strip() if self.login_url_var else TherapyNotesClientFetcher.DEFAULT_LOGIN_URL
        self.users[selected] = {"username": username.strip(), "password": password, "login_url": login_url}
        self._save_users()
        self._populate_credentials_from_user(selected)
        self._append_log(f"Updated credentials for '{selected}'.")

    def _delete_user(self) -> None:
        if self.root is None or not self.selected_user_var:
            return
        selected = self.selected_user_var.get()
        if not selected:
            messagebox.showinfo("Select User", "Please select a saved user to delete.")
            return
        if not messagebox.askyesno("Delete Saved User", f"Remove saved credentials for '{selected}'?"):
            return
        self.users.pop(selected, None)
        self._save_users()
        self._update_user_dropdown()
        self._save_settings()
        self._append_log(f"Deleted saved user '{selected}'.")

    def _check_pdf_reading_status(self) -> None:
        """Check and display the status of all PDF reading dependencies."""
        try:
            if not hasattr(self, 'pdfplumber_status_label'):
                return
            
            statuses = []
            
            # Check pdfplumber
            if PDFPLUMBER_AVAILABLE:
                try:
                    # Verify pdfplumber is actually available
                    if pdfplumber is not None:
                        self.pdfplumber_status_label.config(
                            text="✓ pdfplumber: Installed",
                            fg="#006600"
                        )
                        statuses.append(True)
                    else:
                        self.pdfplumber_status_label.config(
                            text="✗ pdfplumber: Import failed",
                            fg="#cc0000"
                        )
                        statuses.append(False)
                except Exception as exc:
                    self.pdfplumber_status_label.config(
                        text=f"✗ pdfplumber: Error - {str(exc)[:50]}",
                        fg="#cc0000"
                    )
                    statuses.append(False)
            else:
                self.pdfplumber_status_label.config(
                    text="✗ pdfplumber: Not installed (pip install pdfplumber)",
                    fg="#cc0000"
                )
                statuses.append(False)
            
            # Check pytesseract
            if OCR_AVAILABLE:
                try:
                    # Verify pytesseract is actually available
                    if pytesseract is not None:
                        self.pytesseract_status_label.config(
                            text="✓ pytesseract: Installed",
                            fg="#006600"
                        )
                        statuses.append(True)
                    else:
                        self.pytesseract_status_label.config(
                            text="✗ pytesseract: Import failed",
                            fg="#cc0000"
                        )
                        statuses.append(False)
                except Exception as exc:
                    self.pytesseract_status_label.config(
                        text=f"✗ pytesseract: Error - {str(exc)[:50]}",
                        fg="#cc0000"
                    )
                    statuses.append(False)
            else:
                self.pytesseract_status_label.config(
                    text="✗ pytesseract: Not installed (pip install pytesseract)",
                    fg="#cc0000"
                )
                statuses.append(False)
            
            # Check Tesseract OCR executable
            tesseract_ok = False
            if OCR_AVAILABLE and pytesseract is not None:
                try:
                    tesseract_cmd = getattr(pytesseract.pytesseract, 'tesseract_cmd', None)
                    if tesseract_cmd and Path(tesseract_cmd).exists():
                        # Try to get version to verify it works
                        version = pytesseract.get_tesseract_version()
                        self.tesseract_status_label.config(
                            text=f"✓ Tesseract OCR: Found (v{version})",
                            fg="#006600"
                        )
                        tesseract_ok = True
                    else:
                        # Try to find it
                        candidates = [
                            Path("C:/Program Files/Tesseract-OCR/tesseract.exe"),
                            Path("C:/Program Files (x86)/Tesseract-OCR/tesseract.exe"),
                            Path.home() / "AppData/Local/Programs/Tesseract-OCR/tesseract.exe",
                        ]
                        found = False
                        for candidate in candidates:
                            if candidate.exists():
                                self.tesseract_status_label.config(
                                    text=f"⚠ Tesseract OCR: Found at {candidate.parent} (not configured)",
                                    fg="#cc6600"
                                )
                                found = True
                                break
                        if not found:
                            self.tesseract_status_label.config(
                                text="✗ Tesseract OCR: Not found (install from https://github.com/tesseract-ocr/tesseract/wiki)",
                                fg="#cc0000"
                            )
                except Exception as exc:
                    self.tesseract_status_label.config(
                        text=f"✗ Tesseract OCR: Error - {str(exc)[:50]}",
                        fg="#cc0000"
                    )
            else:
                self.tesseract_status_label.config(
                    text="✗ Tesseract OCR: pytesseract not available",
                    fg="#cc0000"
                )
            statuses.append(tesseract_ok)
            
            # Check Poppler (for pdf2image)
            poppler_ok = False
            try:
                # First check if POPPLER_PATH is set in environment
                poppler_path = os.environ.get("POPPLER_PATH")
                if poppler_path and Path(poppler_path).exists() and (Path(poppler_path) / "pdftoppm.exe").exists():
                    self.poppler_status_label.config(
                        text=f"✓ Poppler: Found at {poppler_path}",
                        fg="#006600"
                    )
                    poppler_ok = True
                else:
                    # Try to find it in common locations
                    script_dir = Path(__file__).parent
                    candidates = [
                        script_dir / "vendor" / "poppler" / "Library" / "bin",  # Check vendor folder first
                        Path("C:/Program Files/poppler/Library/bin"),
                        Path("C:/Program Files (x86)/poppler/Library/bin"),
                        Path.home() / "AppData/Local/poppler/Library/bin",
                    ]
                    conda_prefix = os.environ.get("CONDA_PREFIX")
                    if conda_prefix:
                        candidates.append(Path(conda_prefix) / "Library/bin")
                    
                    found = False
                    found_path = None
                    for candidate in candidates:
                        if candidate.exists() and (candidate / "pdftoppm.exe").exists():
                            found_path = candidate
                            found = True
                            break
                    
                    if found and found_path:
                        # If found but not in environment, try to set it automatically
                        if not poppler_path or not Path(poppler_path).exists():
                            try:
                                os.environ["POPPLER_PATH"] = str(found_path)
                                logger.info("Auto-configured Poppler from: %s", found_path)
                            except Exception:
                                pass
                        
                        if found_path == script_dir / "vendor" / "poppler" / "Library" / "bin":
                            self.poppler_status_label.config(
                                text=f"✓ Poppler: Found in vendor folder",
                                fg="#006600"
                            )
                        else:
                            self.poppler_status_label.config(
                                text=f"✓ Poppler: Found at {found_path}",
                                fg="#006600"
                            )
                        poppler_ok = True
                    else:
                        # Also check if pdf2image can find it via PATH
                        try:
                            import subprocess
                            result = subprocess.run(
                                ["pdftoppm", "-v"],
                                capture_output=True,
                                text=True,
                                timeout=2
                            )
                            if result.returncode == 0 or "pdftoppm" in result.stderr.lower():
                                self.poppler_status_label.config(
                                    text="✓ Poppler: Found in system PATH",
                                    fg="#006600"
                                )
                                poppler_ok = True
                            else:
                                raise FileNotFoundError
                        except (FileNotFoundError, subprocess.TimeoutExpired, Exception):
                            self.poppler_status_label.config(
                                text="✗ Poppler: Not found (required for scanned PDFs)\n   Download: https://github.com/oschwartz10612/poppler-windows/releases/\n   Extract and set POPPLER_PATH to Library/bin folder",
                                fg="#cc0000"
                            )
            except Exception as exc:
                self.poppler_status_label.config(
                    text=f"✗ Poppler: Error - {str(exc)[:50]}",
                    fg="#cc0000"
                )
            statuses.append(poppler_ok)
            
            # Overall status
            all_required = statuses[0]  # pdfplumber is required
            all_optional = all(statuses[1:])  # OCR components are optional but recommended
            
            if all_required and all_optional:
                self.pdf_overall_status_label.config(
                    text="✓ PDF Reading: Fully Ready (text and scanned PDFs supported)",
                    fg="#006600"
                )
            elif all_required:
                self.pdf_overall_status_label.config(
                    text="⚠ PDF Reading: Partially Ready (text PDFs only, scanned PDFs may fail)",
                    fg="#cc6600"
                )
            else:
                self.pdf_overall_status_label.config(
                    text="✗ PDF Reading: Not Ready (install pdfplumber: pip install pdfplumber)",
                    fg="#cc0000"
                )
        except Exception as exc:
            # If status check fails, log it but don't crash the app
            logger.warning("Error checking PDF reading status: %s", exc)
            if hasattr(self, 'pdf_overall_status_label'):
                try:
                    self.pdf_overall_status_label.config(
                        text=f"⚠ Status check error: {str(exc)[:50]}",
                        fg="#cc6600"
                    )
                except Exception:
                    pass  # If even this fails, just ignore it

    def _browse_document(self) -> None:
        filetypes = [
            ("All Supported", "*.pdf *.png *.jpg *.jpeg *.tif *.tiff *.xls *.xlsx *.csv"),
            ("PDF", "*.pdf"),
            ("Images", "*.png *.jpg *.jpeg *.tif *.tiff"),
            ("Excel", "*.xls *.xlsx"),
            ("CSV", "*.csv"),
            ("All files", "*.*"),
        ]
        filename = filedialog.askopenfilename(filetypes=filetypes)
        if filename and self.document_path_var is not None:
            self.document_path_var.set(filename)
            self._append_log(f"Selected document: {filename}")
            self._set_status("Insurer document selected.", "blue")
            # Refresh PDF status to show if the document can be read
            self._check_pdf_reading_status()
            self._save_settings()

    def _browse_output_directory(self) -> None:
        directory = filedialog.askdirectory()
        if directory and self.output_dir_var is not None:
            self.output_dir_var.set(directory)
            self._append_log(f"Output directory set to: {directory}")
            self._set_status("Updated export folder.", "blue")
            self._save_settings()

    def _append_log(self, message: str) -> None:
        logger.info(message)
        if self.log_widget:
            self.log_widget.configure(state=tk.NORMAL)
            self.log_widget.insert(tk.END, f"{datetime.now():%Y-%m-%d %H:%M:%S} - {message}\n")
            self.log_widget.configure(state=tk.DISABLED)
            self.log_widget.see(tk.END)

    def _set_running_state(self, running: bool) -> None:
        state_run = tk.DISABLED if running else tk.NORMAL
        state_stop = tk.NORMAL if running else tk.DISABLED
        if self._run_button is not None:
            self._run_button.config(state=state_run)
        if self._stop_button is not None:
            self._stop_button.config(state=state_stop)

    def _signal_run_finished(self) -> None:
        def _apply():
            self._set_running_state(False)
            self._stop_event.clear()
            self._worker_thread = None

        if self.root:
            self.root.after(0, _apply)
        else:
            _apply()

    def _on_stop_clicked(self) -> None:
        if not self._stop_event.is_set():
            self._stop_event.set()
            if self._stop_button is not None:
                self._stop_button.config(state=tk.DISABLED)
            self._append_log("Stop requested by user. Finalizing current progress...")
            self._set_status("Stop requested. Finishing current client...", "orange")

    # Run flow ------------------------------------------------------------
    def _on_run_clicked(self) -> None:
        assert self.document_path_var is not None
        assert self.start_date_var is not None
        assert self.end_date_var is not None
        assert self.username_var is not None
        assert self.password_var is not None
        assert self.output_dir_var is not None

        if not self.document_path_var.get():
            messagebox.showwarning("Missing Document", "Please select an insurer document first.")
            return
        # Allow running without date range - will extract from document
        start_date = None
        end_date = None
        if self.start_date_var.get() and self.end_date_var.get():
            try:
                start_date, end_date = self._parse_date_range(
                    self.start_date_var.get(), self.end_date_var.get()
                )
            except ValueError as exc:
                messagebox.showerror("Invalid Dates", str(exc))
                return
        elif self.start_date_var.get() or self.end_date_var.get():
            messagebox.showwarning("Missing Date Range", "Please enter both start and end dates, or leave both empty to extract from document.")
            return
        if not self.username_var.get() or not self.password_var.get():
            messagebox.showwarning("Missing Credentials", "Please enter TherapyNotes credentials.")
            return
        if not self.output_dir_var.get():
            messagebox.showwarning("Missing Output Folder", "Please choose an output folder for the export.")
            return

        if self._worker_thread and self._worker_thread.is_alive():
            messagebox.showinfo("Run in Progress", "A workflow is already running. Please wait or press Stop.")
            return

        self._stop_event.clear()
        self._set_running_state(True)
        self._append_log("Starting workflow...")
        self._set_status("Running billing log collection...", "blue")
        self._save_settings()
        thread = threading.Thread(
            target=self._run_workflow,
            args=(Path(self.document_path_var.get()), start_date, end_date),
            daemon=True
        )
        self._worker_thread = thread
        thread.start()

    def _parse_date_range(self, start: str, end: str) -> Tuple[datetime, datetime]:
        fmt = "%m/%d/%Y"
        start_dt = datetime.strptime(start, fmt)
        end_dt = datetime.strptime(end, fmt)
        if start_dt > end_dt:
            raise ValueError("Start date must be on or before end date.")
        return start_dt, end_dt

    def _run_workflow(self, document_path: Path, start: Optional[datetime], end: Optional[datetime]) -> None:
        try:
            assert self.username_var is not None
            assert self.password_var is not None
            assert self.output_dir_var is not None

            self._append_log("Parsing document for client metadata...")
            self.current_client = self.parser.parse_document(document_path)
            self._append_log(
                f"Found client: {self.current_client.name or 'UNKNOWN'}, DOB: {self.current_client.dob or 'UNKNOWN'}"
            )
            self._set_status("Client metadata extracted successfully.", "blue")

            self.client_record_queue = self._build_record_queue()
            if self.client_record_queue:
                self._append_log(f"Loaded {len(self.client_record_queue)} client record(s) from document.")
            else:
                fallback_record: Dict[str, Any] = {}
                if self.current_client.name:
                    fallback_record["name"] = self.current_client.name
                if self.current_client.dob:
                    normalized = DocumentParser._normalize_date(self.current_client.dob)
                    fallback_record["dob"] = normalized or self.current_client.dob
                    fallback_record["dob_raw"] = self.current_client.dob
                if fallback_record.get("name"):
                    self.client_record_queue = [fallback_record]
                    self._append_log("No tabular records detected; using single client metadata.")
                else:
                    self._append_log("Warning: No client records found in the document.")
            target_payers = self._get_payer_filters()
            if target_payers:
                self._append_log(f"Payer filters active: {', '.join(target_payers)}")
            else:
                self._append_log("No payer filters specified; all payers will be considered.")
            ins_statuses = self._get_selected_ins_statuses()
            if ins_statuses:
                self._append_log(f"Insurance statuses included: {', '.join(ins_statuses)}")
            else:
                self._append_log("No insurance status filters selected; all statuses will be considered.")

            self._append_log("Connecting to TherapyNotes (placeholder)...")
            self.client_fetcher = TherapyNotesClientFetcher(
                {"username": self.username_var.get(), "password": self.password_var.get()},
                log_callback=self._append_log
            )
            self.client_fetcher.authenticate()
            self.client_fetcher.open_client_chart(self.current_client)
            self._set_status("TherapyNotes patients view loaded.", "blue")

            output_dir = Path(self.output_dir_var.get()).expanduser()
            batch_results: List[Dict[str, Any]] = []
            stop_requested = False

            for index, record in enumerate(self.client_record_queue, start=1):
                if self._stop_event.is_set():
                    stop_requested = True
                    self._append_log("Stop requested. Halting client queue and preparing export...")
                    break
                name = record.get("name", "").strip()
                if not name:
                    continue
                dob_hint = record.get("dob") or record.get("dob_raw") or ""
                
                # Determine date range for this client
                client_start = start
                client_end = end
                if start is None or end is None:
                    # No manual date range provided, try to use per-client date range
                    record_date_range = record.get("date_range")
                    if record_date_range:
                        start_str = record_date_range.get("start")
                        end_str = record_date_range.get("end")
                        if start_str:
                            try:
                                client_start = datetime.strptime(start_str, "%m/%d/%Y")
                            except ValueError:
                                self._append_log(f"Warning: Could not parse start date '{start_str}' for {name}, skipping date range.")
                                client_start = None
                        if end_str:
                            if end_str.upper() == "PRESENT":
                                # Use a far future date for "Present"
                                client_end = datetime(2099, 12, 31)
                            else:
                                try:
                                    client_end = datetime.strptime(end_str, "%m/%d/%Y")
                                except ValueError:
                                    self._append_log(f"Warning: Could not parse end date '{end_str}' for {name}, skipping date range.")
                                    client_end = None
                        if client_start and client_end:
                            self._append_log(f"Using date range from document for {name}: {start_str} to {end_str}")
                    else:
                        self._append_log(f"Warning: No date range found for {name} and no manual date range provided. Skipping this client.")
                        batch_results.append(
                            {
                                "name": name,
                                "dob": dob_hint,
                                "entries": [],
                                "skip_note": "Skipped - no date range available",
                            }
                        )
                        continue
                
                if client_start is None or client_end is None:
                    self._append_log(f"Warning: Invalid date range for {name}. Skipping this client.")
                    batch_results.append(
                        {
                            "name": name,
                            "dob": dob_hint,
                            "entries": [],
                            "skip_note": "Skipped - invalid date range",
                        }
                    )
                    continue
                
                if index > 1:
                    self.client_fetcher.return_to_patients_sidebar()
                self._append_log(
                    f"[{index}/{len(self.client_record_queue)}] Searching TherapyNotes for {name} ({dob_hint or 'DOB unknown'})..."
                )
                record_client = ClientMetadata(name=name, dob=dob_hint)
                if self._stop_event.is_set():
                    stop_requested = True
                    self._append_log("Stop requested before searching next client. Preparing export...")
                    break
                try:
                    found = self.client_fetcher.search_and_open_patient(name, dob_hint)
                except Exception as nav_exc:
                    self._append_log(f"Error while searching for patient {name}: {nav_exc}")
                    found = False
                if not found:
                    self._append_log(f"Skipped {name}: could not find client in TherapyNotes.")
                    batch_results.append(
                        {
                            "name": name,
                            "dob": dob_hint,
                            "entries": [],
                            "skip_note": "Skipped - could not find client in TN",
                        }
                    )
                    if self._stop_event.is_set():
                        stop_requested = True
                        self._append_log("Stop requested after skip. Preparing export...")
                        break
                    continue

                self.active_record = record
                self._append_log("Fetching billable service dates (placeholder)...")
                if self._stop_event.is_set():
                    stop_requested = True
                    self._append_log("Stop requested before billing fetch. Preparing export...")
                    break
                try:
                    self.service_entries = self.client_fetcher.fetch_billable_dates(
                        record_client,
                        client_start,
                        client_end,
                        target_payers,
                        ins_statuses
                    )
                except Exception as fetch_exc:
                    self._append_log(f"Error retrieving billing transactions for {name}: {fetch_exc}")
                    batch_results.append(
                        {
                            "name": name,
                            "dob": dob_hint,
                            "entries": [],
                            "skip_note": "Skipped - error retrieving billing data",
                        }
                    )
                    self.client_fetcher.return_to_patients_sidebar()
                    if self._stop_event.is_set():
                        stop_requested = True
                        self._append_log("Stop requested after billing error. Preparing export...")
                        break
                    continue

                filtered = [entry for entry in self.service_entries if client_start <= entry.service_date <= client_end]
                date_range_str = f"{client_start.strftime('%m/%d/%Y')} to {client_end.strftime('%m/%d/%Y')}" if client_end.year < 2099 else f"{client_start.strftime('%m/%d/%Y')} to Present"
                self._append_log(f"Found {len(filtered)} service entries within date range ({date_range_str}) for {name}.")
                if not filtered:
                    batch_results.append(
                        {
                            "name": name,
                            "dob": dob_hint,
                            "entries": [],
                            "skip_note": "Skipped - no matching DOS within range",
                        }
                    )
                    self.client_fetcher.return_to_patients_sidebar()
                    if self._stop_event.is_set():
                        stop_requested = True
                        self._append_log("Stop requested after processing client with no results. Preparing export...")
                        break
                    continue

                # Use dob_hint which is already correctly extracted from the current record
                batch_results.append(
                    {
                        "name": name,
                        "dob": dob_hint,
                        "entries": filtered,
                        "skip_note": None,
                    }
                )
                self._append_log(f"Queued {len(filtered)} DOS entries for {name}.")
                self._set_status(f"Queued {len(filtered)} entries for {name}.", "green")
                self.client_fetcher.return_to_patients_sidebar()
                if self._stop_event.is_set():
                    stop_requested = True
                    self._append_log("Stop requested. Wrapping up after current client.")
                    break

            stop_requested = stop_requested or self._stop_event.is_set()

            if batch_results:
                self._append_log("Exporting consolidated billing log...")
                csv_path, xlsx_path = self._export_batch_results(batch_results, output_dir)
                self._append_log(f"Consolidated CSV saved to {csv_path}")
                if xlsx_path:
                    self._append_log(f"Consolidated Excel saved to {xlsx_path}")
                if stop_requested:
                    self._set_status("Run stopped. Partial export saved.", "orange")
                else:
                    self._set_status("Export complete.", "green")
            else:
                self._append_log("No client data to export.")
                if stop_requested:
                    self._set_status("Run stopped before collecting data.", "orange")
                else:
                    self._set_status("No client data to export.", "yellow")
        except Exception as exc:  # noqa: BLE001
            logger.exception("Workflow failed: %s", exc)
            self._append_log(f"Error: {exc}")
            self._set_status("Workflow encountered an error. Check log for details.", "red")
        finally:
            if self.client_fetcher:
                try:
                    self.client_fetcher.close()
                except Exception:
                    pass
            self._signal_run_finished()
            self._save_settings()

    def _export_batch_results(
        self,
        batch_results: List[Dict[str, Any]],
        output_dir: Path
    ) -> Tuple[Path, Optional[Path]]:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir.mkdir(parents=True, exist_ok=True)
        csv_path = output_dir / f"therapy_notes_log_{timestamp}.csv"
        xlsx_path = csv_path.with_suffix(".xlsx") if OPENPYXL_AVAILABLE else None

        headers: List[str] = []
        columns: List[List[str]] = []
        highlight_map: List[bool] = []

        for result in batch_results:
            name = result.get("name", "")
            dob = result.get("dob", "")
            entries: List[ServiceEntry] = result.get("entries", [])
            skip_note = result.get("skip_note")

            header = f"{name} (DOB: {dob})" if dob else name
            headers.append(header)

            column_values: List[str] = []
            highlight_flags: List[bool] = []

            if entries:
                sorted_entries = sorted(entries, key=lambda e: e.service_date)
                for entry in sorted_entries:
                    formatted = f"{entry.service_date.strftime('%m/%d/%Y')} - {entry.notes or ''}".strip(" -")
                    column_values.append(formatted)
                    highlight_flags.append(
                        bool(entry.notes and "submitted external" in entry.notes.lower().strip())
                    )
            else:
                note = skip_note or "Skipped - reason unknown"
                column_values.append(note)
                highlight_flags.append(False)

            columns.append(column_values)
            highlight_map.append(highlight_flags)

        max_rows = max((len(col) for col in columns), default=0)

        csv_rows: List[List[str]] = [headers]
        for row_idx in range(max_rows):
            row: List[str] = []
            for col in columns:
                row.append(col[row_idx] if row_idx < len(col) else "")
            csv_rows.append(row)

        with csv_path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerows(csv_rows)

        created_xlsx: Optional[Path] = None
        if OPENPYXL_AVAILABLE:
            try:
                from openpyxl import Workbook  # type: ignore
                from openpyxl.styles import PatternFill  # type: ignore
            except Exception:
                return csv_path, None

            wb = Workbook()
            ws = wb.active
            ws.title = "Billing Log"

            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx).value = header

            highlight_fill = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")

            for col_idx, col_values in enumerate(columns, start=1):
                flags = highlight_map[col_idx - 1]
                for row_offset, value in enumerate(col_values, start=2):
                    cell = ws.cell(row=row_offset, column=col_idx)
                    cell.value = value
                    if highlight_fill and flags[row_offset - 2]:
                        cell.fill = highlight_fill

            wb.save(xlsx_path)
            created_xlsx = xlsx_path
            self._append_log(f"Excel copy saved to {xlsx_path}")

        return csv_path, created_xlsx

    def _set_status(self, message: str, color: str = "gray") -> None:
        if self.status_var is not None:
            self.status_var.set(message)
        if self.status_label is not None:
            self.status_label.configure(fg=color)

    def _on_close(self) -> None:
        self._save_settings()
        if self.client_fetcher:
            try:
                self.client_fetcher.close()
            except Exception:
                pass
        if self.root is not None:
            self.root.destroy()


def main() -> None:
    bot = TherapyNotesRecordsBot()
    bot.launch()


if __name__ == "__main__":
    main()

