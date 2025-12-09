# integrity_medical_records_bot.py
import os, re, sys, time, glob, argparse, datetime
from datetime import datetime as DT

# =============== Logging & small utils ===============
def log(msg):
    print(time.strftime("[%H:%M:%S] "), msg, flush=True)

def sanitize_filename(s):
    return re.sub(r'[<>:"/\\|?*]+', "_", str(s or "")).strip()

def to_mmddyyyy(val):
    if val is None:
        return ""
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.strftime("%m/%d/%Y")
    s = str(val).strip()
    try:
        # Excel serial
        if isinstance(val, (int, float)) and float(val) > 20000:
            base = datetime.datetime(1899, 12, 30)
            return (base + datetime.timedelta(days=float(val))).strftime("%m/%d/%Y")
    except Exception:
        pass
    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return DT.strptime(s, fmt).strftime("%m/%d/%Y")
        except Exception:
            pass
    m = re.search(r"\b(\d{1,2}/\d{1,2}/\d{4})\b", s)
    return m.group(1) if m else s

def canonical_mmddyyyy(s):
    """Return mm/dd/YYYY if parseable, else None."""
    if not s: return None
    s = s.strip()
    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%m/%d/%y", "%-m/%-d/%Y", "%-m/%-d/%y"):
        try:
            return DT.strptime(s, fmt).strftime("%m/%d/%Y")
        except Exception:
            pass
    # try to normalize crude m/d/yy or m/d/yyyy
    m = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b", s)
    if m:
        mm = int(m.group(1)); dd = int(m.group(2)); yy = m.group(3)
        if len(yy) == 2:
            # datetime default: 69->2069 wrap; we’ll just defer to strptime for clarity
            try:
                return DT.strptime(f"{mm}/{dd}/{yy}", "%m/%d/%y").strftime("%m/%d/%Y")
            except Exception:
                return None
        else:
            try:
                return DT.strptime(f"{mm}/{dd}/{yy}", "%m/%d/%Y").strftime("%m/%d/%Y")
            except Exception:
                return None
    return None

def extract_first_date_mmddyyyy(text):
    """Find first date-ish string in text and return mm/dd/YYYY canon or None."""
    if not text: return None
    m = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", text)
    if not m: return None
    return canonical_mmddyyyy(m.group(1))

def ensure_client_dirs(client_name):
    base = os.path.join(os.path.expanduser("~"), "Desktop", "Medical Records Bot", sanitize_filename(client_name))
    dl = os.path.join(base, "downloads")
    os.makedirs(dl, exist_ok=True)
    return base, dl

def current_files(dirpath):
    return set(glob.glob(os.path.join(dirpath, "*")))

def wait_for_downloads(dirpath, timeout=90):
    end = time.time() + timeout
    while time.time() < end:
        if not glob.glob(os.path.join(dirpath, "*.crdownload")):
            return True
        time.sleep(0.12)
    return False

def wait_for_new_files(dirpath, before_set, timeout=40):
    end = time.time() + timeout
    new_files = set()
    while time.time() < end:
        now = current_files(dirpath)
        diff = {p for p in (now - before_set) if not p.endswith(".crdownload")}
        if diff:
            new_files = diff
            if wait_for_downloads(dirpath, timeout=max(1, int(end - time.time()))):
                return True, new_files
        time.sleep(0.12)
    return False, new_files

# =============== Excel ===============
def _color_is_yellow(openpyxl_color) -> bool:
    try:
        rgb = getattr(openpyxl_color, "rgb", None)
    except Exception:
        rgb = None
    if not rgb:
        return False
    if len(rgb) == 8:  # ARGB
        rgb = rgb[2:]
    if len(rgb) != 6:
        return False
    try:
        r = int(rgb[0:2], 16); g = int(rgb[2:4], 16); b = int(rgb[4:6], 16)
    except Exception:
        return False
    return (r >= 220 and g >= 200 and b <= 80)

def read_matrix_with_ia_and_colors(xlsx_path, sheet_name):
    """
    {
      'Client Name': {
         'ia': 'mm/dd/yyyy' or None,
         'progress': [(date_str, is_yellow_bool), ...]
      }, ...
    }
    Row 1 = client names; below = dates/labels.
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    max_col = ws.max_column or 1
    max_row = ws.max_row or 1
    out = {}

    for col in range(1, max_col + 1):
        name = str(ws.cell(row=1, column=col).value or "").strip()
        if not name:
            continue
        ia_date = None
        progress = []

        EMPTY_RUN_LIMIT = 40
        empty_run = 0

        for r in range(2, max_row + 1):
            c = ws.cell(row=r, column=col)
            raw = c.value

            if raw is None or str(raw).strip() == "":
                empty_run += 1
                if empty_run >= EMPTY_RUN_LIMIT:
                    break
                continue
            empty_run = 0

            s = str(raw).strip()
            d = to_mmddyyyy(raw)
            if not re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", d):
                continue

            label_left = str(ws.cell(row=r, column=1).value or "")
            is_ia = ("ia" in s.lower()) or ("intake" in s.lower()) or ("ia" in label_left.lower()) or ("intake" in label_left.lower())

            yellow = False
            try:
                if getattr(c.fill, "fill_type", None) == "solid":
                    yellow = _color_is_yellow(c.fill.start_color)
            except Exception:
                yellow = False

            if is_ia and not ia_date:
                ia_date = d
            else:
                progress.append((d, yellow))

        if ia_date or progress:
            out[name] = {"ia": ia_date, "progress": progress}
    return out

# =============== Selenium setup & helpers ===============

def go_to_patients_list(driver, max_tries=3):
    """
    Ensure we're on Patients list. Prefer clicking the top 'Patients' link to avoid permission redirects.
    Falls back to direct URL if needed. Waits for the search box.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    tries = 0
    while tries < max_tries:
        tries += 1
        try:
            # Try clicking top nav 'Patients' if present
            links = driver.find_elements(By.LINK_TEXT, "Patients")
            if not links:
                links = driver.find_elements(By.XPATH, "//a[normalize-space(.)='Patients']")
            if links:
                try:
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", links[0])
                except Exception:
                    pass
                try:
                    links[0].click()
                except Exception:
                    try:
                        driver.execute_script('arguments[0].click();', links[0])
                    except Exception:
                        pass
            else:
                # fallback: direct URL
                driver.get("https://www.therapynotes.com/app/patients/list")
            # wait for patients search box
            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Name, Acct #, Phone, or Ins ID']"))
            )
            return True
        except Exception:
            # small retry
            try:
                driver.get("https://www.therapynotes.com/app/patients/list")
            except Exception:
                pass
            try:
                WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Name, Acct #, Phone, or Ins ID']"))
                )
                return True
            except Exception:
                continue
    return False


def build_driver(download_dir, headless=False, fast=True):
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    log("Preparing ChromeDriver (first run may download the driver)…")

    opts = webdriver.ChromeOptions()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-features=Translate,CalculateNativeWinOcclusion")
    if fast:
        opts.page_load_strategy = "eager"
        opts.add_argument("--blink-settings=imagesEnabled=false")

    prefs = {
        "download.default_directory": os.path.abspath(download_dir),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True,
        "profile.default_content_setting_values.notifications": 2,
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
    }
    opts.add_experimental_option("prefs", prefs)
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opts)
    try:
        driver.set_page_load_timeout(20)
    except Exception:
        pass
    log("ChromeDriver ready.")
    return driver

def set_download_dir(driver, dirpath):
    try:
        driver.execute_cdp_cmd("Page.setDownloadBehavior", {
            "behavior": "allow",
            "downloadPath": os.path.abspath(dirpath),
        })
        return True
    except Exception:
        return False

def login_therapynotes(driver, username, password):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    login_url = "https://www.therapynotes.com/app/login/IntegritySWS/"
    driver.get(login_url)
    log("Opening login page…")
    wait = WebDriverWait(driver, 15)
    uf = wait.until(EC.presence_of_element_located((By.ID, "Login__UsernameField")))
    pf = wait.until(EC.presence_of_element_located((By.ID, "Login__Password")))
    uf.clear(); uf.send_keys(username)
    pf.clear(); pf.send_keys(password)
    wait.until(EC.element_to_be_clickable((By.ID, "Login__LogInButton"))).click()
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.LINK_TEXT, "Patients")))
    log("Login successful.")


def safe_return_to_patients(driver):
    try:
        return go_to_patients_list(driver, max_tries=2)
    except Exception:
        return False
def open_documents_tab(driver):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    locs = [
        (By.CSS_SELECTOR, "a[data-tab-id='Documents']"),
        (By.XPATH, "//a[@data-tab-id='Documents']"),
        (By.XPATH, "//a[normalize-space(text())='Documents']"),
    ]
    for by, sel in locs:
        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((by, sel))).click()
            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.XPATH, "//a[normalize-space(.)='Download Multiple'] | //span[contains(@class,'documentNameSpan')]"))
            )
            time.sleep(0.15)
            return True
        except Exception:
            continue
    return False

# ========== Patient selection ==========
def select_patient_by_name(driver, full_name, allow_loose=True):
    import re, time
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    def norm(s: str) -> str:
        s = (s or "").lower()
        s = re.sub(r"[^a-z0-9 ]+", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def tokens_in_order(candidate_norm: str, target_tokens):
        pos = 0
        for tok in target_tokens:
            i = candidate_norm.find(tok, pos)
            if i == -1: return False
            pos = i + len(tok)
        return True

    s = " ".join(str(full_name or "").split()).strip()
    first_last = None; first_middle_last = None
    m = re.match(r'^([^,]+),\s*(.+)$', s)
    if m:
        last = m.group(1).strip()
        rest = m.group(2).strip()
        parts = rest.split()
        if parts:
            first = parts[0]; middle = " ".join(parts[1:])
            first_last = f"{first} {last}"
            if middle: first_middle_last = f"{first} {middle} {last}"
    else:
        s2 = re.sub(r",", " ", s); s2 = re.sub(r"\s+", " ", s2).strip()
        first_last = s2

    def drop_single_middle(q: str) -> str:
        return re.sub(r'^(\S+)\s+[A-Za-z]\s+(\S+.*)$', r'\1 \2', q).strip()

    query_variants = []
    if first_last:
        query_variants.append(first_last)
        no_mid = drop_single_middle(first_last)
        if no_mid != first_last: query_variants.append(no_mid)
    if first_middle_last:
        query_variants.append(first_middle_last)
        no_mid2 = drop_single_middle(first_middle_last)
        if no_mid2 not in query_variants: query_variants.append(no_mid2)

    seen = set(); tried_queries = []
    for q in query_variants:
        if q and q.lower() not in seen:
            seen.add(q.lower()); tried_queries.append(q)

    target_norm = norm(first_last or first_middle_last or s)
    target_tokens = [t for t in target_norm.split(" ") if t]

    if not safe_return_to_patients(driver): return False
    wait = WebDriverWait(driver, 10)
    try:
        sb = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='Name, Acct #, Phone, or Ins ID']")))
    except Exception:
        return False

    for q in tried_queries:
        try:
            clean_q = re.sub(r",", " ", q)
            clean_q = re.sub(r"\s+", " ", clean_q)
            sb.clear(); sb.send_keys(clean_q)
            time.sleep(0.6)
            try:
                wait.until(EC.visibility_of_element_located((By.ID, "ContentBubbleResultsContainer")))
            except Exception:
                pass

            candidates = driver.find_elements(By.CSS_SELECTOR, "#ContentBubbleResultsContainer > div")

            for rdiv in candidates:
                try:
                    cand_text = norm(rdiv.text)
                    if cand_text == target_norm:
                        rdiv.click(); time.sleep(0.25); return True
                except Exception:
                    continue

            if allow_loose:
                for rdiv in candidates:
                    try:
                        cand_text = norm(rdiv.text)
                        if tokens_in_order(cand_text, target_tokens):
                            rdiv.click(); time.sleep(0.25); return True
                    except Exception:
                        continue
        except Exception:
            continue
    return False

# -------- "Download Multiple" utilities ----------
def click_download_multiple(driver):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        el = WebDriverWait(driver, 8).until(
            EC.element_to_be_clickable((By.XPATH, "//a[normalize-space(.)='Download Multiple']"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        el.click()
        WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.XPATH, "//input[@type='checkbox']"))
        )
        log("   Entered multi-select mode.")
        return True
    except Exception:
        log("   - Could not open 'Download Multiple'.")
        return False


def exit_multi_mode_via_cancel(driver):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        el = WebDriverWait(driver, 8).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space(.)='Cancel' and @href='#']")))
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        except Exception:
            pass
        try:
            el.click()
            return True
        except Exception:
            try:
                driver.execute_script("arguments[0].click();", el)
                return True
            except Exception:
                return False
    except Exception:
        return False

def _find_batch_tiles(driver):
    from selenium.webdriver.common.by import By
    try:
        return driver.find_elements(
            By.XPATH,
            "//div[contains(@class,'ellipsis') and contains(normalize-space(.),'Batch of Documents')]"
        )
    except Exception:
        return []

def click_batch_tile_and_wait(driver, download_dir, timeout=90):
    """
    After pressing 'Download Selected Documents', TherapyNotes shows a bottom tile 'Batch of Documents'.
    We must click that tile to initiate the actual download. Then wait until files show up.
    """
    before = current_files(download_dir)
    import time as _t
    end = time.time() + timeout
    while time.time() < end:
        tiles = _find_batch_tiles(driver)
        if tiles:
            tile = tiles[0]
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'end'});", tile)
            except Exception:
                pass
            try:
                tile.click()
            except Exception:
                try:
                    driver.execute_script("arguments[0].click();", tile)
                except Exception:
                    pass
            ok, new_files = wait_for_new_files(download_dir, before, timeout=10)
            if ok:
                return True, new_files
        _t.sleep(0.4)
    return False, set()

def click_download_selected(driver):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    xp = "//input[@type='button' and @value='Download Selected Documents']"
    try:
        btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xp)))
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        except Exception:
            pass
        try:
            btn.click()
            return True
        except Exception:
            try:
                driver.execute_script("arguments[0].click();", btn)
                return True
            except Exception:
                return False
    except Exception:
        return False


def wait_for_doc_checkboxes(driver, timeout=12):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, "//tr[.//input[@type='checkbox']]//input[@type='checkbox']"))
        )
        return True
    except Exception:
        return False

def find_doc_rows(driver):
    from selenium.webdriver.common.by import By
    try:
        return driver.find_elements(By.XPATH, "//tr[.//input[@type='checkbox']]")
    except Exception:
        return []

def row_text_and_name(tr):
    from selenium.webdriver.common.by import By
    try:
        name = tr.find_element(By.XPATH, ".//span[contains(@class,'documentNameSpan')]").text.lower()
    except Exception:
        name = ""
    try:
        txt = tr.text or ""
    except Exception:
        txt = ""
    return name, txt

def row_has_pcp_blurb(tr):
    from selenium.webdriver.common.by import By
    try:
        el = tr.find_element(By.XPATH, ".//div[contains(@class,'extra-info') and normalize-space(.)='PCP']")
        return el is not None
    except Exception:
        return False


def row_matches_intake(tr, target_mmddyyyy):
    name, txt = row_text_and_name(tr)
    if not (("intake" in name) or ("intake" in (txt or "").lower())):
        return False
    d = extract_first_date_mmddyyyy(txt)
    return d == target_mmddyyyy


def row_is_progress_with_date(tr, target_dates_set):
    name, txt = row_text_and_name(tr)
    low = (name + " " + (txt or "")).lower()
    if "intake" in low or "assessment" in low:
        return None
    if not (("progress note" in low) or ("session note" in low)):
        return None
    d = extract_first_date_mmddyyyy(txt)
    return d if (d and d in target_dates_set) else None


def row_is_tp_or_contact_pcp_in_range(tr, start_dt, end_dt):
    name, txt = row_text_and_name(tr)
    low = (name + " " + (txt or "")).lower()
    is_tp = "treatment plan" in low
    is_contact_pcp = ("contact note" in low) and ("pcp" in low)
    if not (is_tp or is_contact_pcp):
        return False
    d = extract_first_date_mmddyyyy(txt)
    if not d:
        return False
    try:
        dt = DT.strptime(d, "%m/%d/%Y")
    except Exception:
        return False
    return start_dt <= dt <= end_dt

def tick_checkbox_in_row(driver, tr):
    from selenium.webdriver.common.by import By
    from selenium.common.exceptions import ElementClickInterceptedException, StaleElementReferenceException
    try:
        cb = tr.find_element(By.XPATH, ".//input[@type='checkbox']")
    except Exception:
        return False
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cb)
    except Exception:
        pass
    try:
        cb.click()
        return True
    except (ElementClickInterceptedException, StaleElementReferenceException, Exception):
        try:
            driver.execute_script("arguments[0].click();", cb)
            return True
        except Exception:
            return False

def try_click_older_or_next(driver):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    xps = [
        "//a[normalize-space(.)='Older']",
        "//a[normalize-space(.)='Next']",
        "//a[normalize-space(.)='Next ›']",
        "//a[normalize-space(.)='›']",
        "//a[contains(@class,'next')]",
    ]
    for xp in xps:
        try:
            el = WebDriverWait(driver, 2.0).until(EC.element_to_be_clickable((By.XPATH, xp)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            el.click()
            time.sleep(0.5)
            return True
        except Exception:
            continue
    return False

def scroll_page_for_more(driver):
    try:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(0.25)
        driver.execute_script("window.scrollBy(0, -150);")
        time.sleep(0.1)
    except Exception:
        time.sleep(0.15)



def scroll_until_stable(driver, max_passes=6, pause=0.12):
    # Scroll to bottom in small steps until no new rows appear or max_passes reached.
    # Returns total rows found on the page.
    seen = 0
    unchanged_passes = 0
    for _ in range(max_passes):
        rows = find_doc_rows(driver)
        n = len(rows)
        if n <= seen:
            unchanged_passes += 1
        else:
            seen = n
            unchanged_passes = 0
        try:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        except Exception:
            pass
        import time as _t
        _t.sleep(pause)
        try:
            driver.execute_script("window.scrollBy(0, -150);")
        except Exception:
            pass
        _t.sleep(pause)
        if unchanged_passes >= 2:
            break
    return seen
# -------- Faster selection routines --------
def select_progress_fast(driver, pn_dates, page_limit=12, per_page_scrolls=2):
    """Fast path: build a set of canonical target dates; scan rows once and match by parsed date."""
    target = {canonical_mmddyyyy(d) for d in pn_dates if canonical_mmddyyyy(d)}
    selected = 0
    visited_pages = 0

    while target and visited_pages < page_limit:
        for _ in range(per_page_scrolls):
            scroll_until_stable(driver, max_passes=6, pause=0.12)
            rows = find_doc_rows(driver)
            if not rows:
                scroll_page_for_more(driver); continue
            # one pass across rows:
            to_remove = set()
            for tr in rows:
                try:
                    if row_is_tp_or_contact_pcp_in_range(tr, start_dt, end_dt):
                        if tick_checkbox_in_row(driver, tr):
                            selected += 1
                except Exception:
                    continue

        if target and try_click_older_or_next(driver):
            visited_pages += 1
        else:
            break

    return selected

def select_tp_and_pcp_in_range(driver, start_dt, end_dt, page_limit=10, per_page_scrolls=2):
    selected = 0
    visited_pages = 0
    while visited_pages < page_limit:
        for _ in range(per_page_scrolls):
            scroll_until_stable(driver, max_passes=6, pause=0.12)
            rows = find_doc_rows(driver)
            if not rows:
                scroll_page_for_more(driver); continue
            for tr in rows:
                try:
                    if row_is_tp_or_contact_pcp_in_range(tr, start_dt, end_dt):
                        if tick_checkbox_in_row(driver, tr):
                            selected += 1
                except Exception:
                    continue
        if try_click_older_or_next(driver):
            visited_pages += 1
        else:
            break
    return selected


def do_download_selected_and_cancel(driver, download_dir, label="selection", max_wait_secs=30):
    before = current_files(download_dir)
    if not click_download_selected(driver):
        log(f"   - Could not click 'Download Selected Documents' for {label}.")
        exit_multi_mode_via_cancel(driver)
        return False
    # Click bottom "Batch of Documents" tile and wait for files
    ok_tile, new_files = click_batch_tile_and_wait(driver, download_dir, timeout=max(40, max_wait_secs))
    if not ok_tile:
        # Fallback to plain watcher
        ok2, new_files2 = wait_for_new_files(download_dir, before, timeout=max_wait_secs)
        if not ok2:
            log(f"   - No new files detected for {label} in {max_wait_secs}s (continuing).")
        else:
            new_files = new_files2
    # Always cancel to clear selections and unblur the page before the next step
    exit_multi_mode_via_cancel(driver)
    return True if new_files else False
def download_via_multi_for_client(driver, ia_date, pn_dates, download_dir, stop_event=None):
    # ---------- Intake (select -> Download Selected -> Cancel) ----------
    if stop_event and stop_event.is_set(): return
    ia_canon = canonical_mmddyyyy(ia_date) if ia_date else None
    if ia_canon and click_download_multiple(driver):
        if not wait_for_doc_checkboxes(driver, timeout=12):
            log("   - No checkboxes visible after entering multi-select; cancelling.")
            exit_multi_mode_via_cancel(driver)
        else:
            # stabilize then scan
            try: scroll_until_stable(driver, max_passes=6, pause=0.12)
            except Exception: pass
            found = False; page_hops = 0
            while page_hops < 10 and not found:
                rows = find_doc_rows(driver)
                for tr in rows:
                    try:
                        if row_matches_intake(tr, ia_canon):
                            if tick_checkbox_in_row(driver, tr):
                                log(f"   ✓ [IA] Checked Intake for {ia_canon}")
                                found = True; break
                    except Exception:
                        continue
                if not found and try_click_older_or_next(driver):
                    page_hops += 1
                    try: scroll_until_stable(driver, max_passes=4, pause=0.10)
                    except Exception: pass
                else:
                    break
            if found:
                do_download_selected_and_cancel(driver, download_dir, label=f"IA {ia_canon}", max_wait_secs=30)
            else:
                log(f"   - Intake note for {ia_canon} not found; cancelling.")
                exit_multi_mode_via_cancel(driver)

    # ---------- Progress Notes (select all by Excel dates -> Download Selected -> Cancel) ----------
    if stop_event and stop_event.is_set(): return
    pn_targets = set()
    if isinstance(pn_dates, list):
        for d in pn_dates:
            cd = canonical_mmddyyyy(d)
            if cd: pn_targets.add(cd)
    else:
        for d in (pn_dates or []):
            cd = canonical_mmddyyyy(d)
            if cd: pn_targets.add(cd)

    if pn_targets:
        # One batch per pages across pagination
        if click_download_multiple(driver):
            if not wait_for_doc_checkboxes(driver, timeout=12):
                log("   - No checkboxes visible in Progress selection; cancelling.")
                exit_multi_mode_via_cancel(driver)
            else:
                selected_total = 0
                pages = 0
                while pn_targets and pages < 24:
                    try: scroll_until_stable(driver, max_passes=6, pause=0.12)
                    except Exception: pass
                    rows = find_doc_rows(driver)
                    any_this_page = 0
                    to_remove = set()
                    for tr in rows:
                        try:
                            hit = row_is_progress_with_date(tr, pn_targets)
                            if hit and tick_checkbox_in_row(driver, tr):
                                any_this_page += 1
                                to_remove.add(hit)
                                log(f"   ✓ [PN] Checked {hit}")
                        except Exception:
                            continue
                    if to_remove:
                        pn_targets -= to_remove
                    if any_this_page > 0:
                        # download this page batch then continue
                        before_label = f"{any_this_page} PN (page {pages+1})"
                        ok = do_download_selected_and_cancel(driver, download_dir, label=before_label, max_wait_secs=40)
                        log(f"   ↳ Downloaded batch: {before_label} ({'ok' if ok else 'timeout'})")
                        selected_total += any_this_page
                        # re-enter multi-select for next page
                        click_download_multiple(driver)
                        if not wait_for_doc_checkboxes(driver, timeout=10):
                            break
                    # move to next page
                    if pn_targets and try_click_older_or_next(driver):
                        pages += 1
                    else:
                        break
                if selected_total == 0:
                    log("   - No progress notes matched within available pages.")
                else:
                    log(f"   ✓ Downloaded {selected_total} progress note(s) total.")
                exit_multi_mode_via_cancel(driver)

    # ---------- Treatment Plans & PCP Contact Notes (select in range -> Download Selected -> Cancel) ----------
    if stop_event and stop_event.is_set(): return
    # Determine range from IA..last PN (inclusive)
    start_dt = None; end_dt = None
    if ia_canon:
        start_dt = DT.strptime(ia_canon, "%m/%d/%Y")
    pn_all = []
    if isinstance(pn_dates, list):
        pn_all = [canonical_mmddyyyy(d) for d in pn_dates if canonical_mmddyyyy(d)]
    if pn_all:
        pn_dts = sorted(DT.strptime(d, "%m/%d/%Y") for d in pn_all)
        if not start_dt: start_dt = pn_dts[0]
        end_dt = pn_dts[-1]
    if start_dt and end_dt and click_download_multiple(driver):
        if not wait_for_doc_checkboxes(driver, timeout=12):
            log("   - No checkboxes visible in TP/PCP selection; cancelling.")
            exit_multi_mode_via_cancel(driver)
        else:
            selected_total2 = 0
            pages2 = 0
            while pages2 < 20:
                try: scroll_until_stable(driver, max_passes=6, pause=0.12)
                except Exception: pass
                rows = find_doc_rows(driver)
                any_this_page = 0
                for tr in rows:
                    try:
                        if row_is_tp_or_contact_pcp_in_range(tr, start_dt, end_dt):
                            if tick_checkbox_in_row(driver, tr):
                                any_this_page += 1
                                log("   ✓ [TP/PCP] Checked a doc in range")
                    except Exception:
                        continue
                if any_this_page > 0:
                    ok = do_download_selected_and_cancel(driver, download_dir, label=f"{any_this_page} TP/PCP docs", max_wait_secs=40)
                    log(f"   ↳ Downloaded TP/PCP batch ({'ok' if ok else 'timeout'})")
                    # re-enter multi for more
                    click_download_multiple(driver)
                    if not wait_for_doc_checkboxes(driver, timeout=10):
                        break
                if try_click_older_or_next(driver):
                    pages2 += 1
                else:
                    break
            if selected_total2 == 0 and pages2 == 0:
                log("   - No Treatment Plan or PCP Contact Notes in range.")
            exit_multi_mode_via_cancel(driver)

def run(xlsx_path, sheet_name, tn_user, tn_pass, headless=False, fast=True, limit=None, stop_event=None):
    t0 = time.time()
    log(f"Reading Excel… ({os.path.basename(xlsx_path)} | sheet='{sheet_name}')")
    try:
        clients = read_matrix_with_ia_and_colors(xlsx_path, sheet_name)
    except Exception as e:
        log(f"[FATAL] Could not read Excel/sheet: {e}")
        return
    log(f"Excel parsed in {time.time()-t0:.1f}s. Clients with data: {len(clients)}")
    if not clients:
        log("No client columns found with IA/progress data.")
        return

    scratch_dl = os.path.join(os.path.expanduser("~"), "Desktop", "Medical Records Bot", "_downloads")
    os.makedirs(scratch_dl, exist_ok=True)
    driver = None
    try:
        if stop_event and stop_event.is_set(): return
        log(f"Starting Chrome… (headless={headless}, fast_mode={fast})")
        driver = build_driver(scratch_dl, headless=headless, fast=fast)

        if stop_event and stop_event.is_set(): return
        log("Logging into TherapyNotes…")
        login_therapynotes(driver, tn_user, tn_pass)

        count = 0
        for client_name, data in clients.items():
            if stop_event and stop_event.is_set():
                log("Stop requested — ending early."); break
            if limit and count >= limit: break
            count += 1

            base_dir, dl_dir = ensure_client_dirs(client_name)
            set_ok = set_download_dir(driver, dl_dir)
            log(f"→ {client_name}  (download dir set: {bool(set_ok)})")

            if not select_patient_by_name(driver, client_name, allow_loose=True):
                log(f"— {client_name}: not found in TN search; skipping."); continue
            if not open_documents_tab(driver):
                log(f"— {client_name}: cannot open Documents tab; skipping."); continue

            ia = data.get("ia")
            progress_all = data.get("progress") or []
            pn_dates = [d for (d, is_yellow) in progress_all if not is_yellow]
            skipped_yellow = [d for (d, is_yellow) in progress_all if is_yellow]
            if skipped_yellow:
                log(f"   (skipped yellow dates: {', '.join(skipped_yellow)})")
            log(f"   IA: {ia or 'None'} | progress dates: {len(pn_dates)}")

            download_via_multi_for_client(driver, ia, pn_dates, dl_dir, stop_event=stop_event)
            safe_return_to_patients(driver)

        log("All requested clients processed.")
    finally:
        try:
            if driver: driver.quit()
        except Exception:
            pass

# =========================== GUI (maroon, like your other bots) ===========================
class MedicalRecordsGUI:
    MAROON = "#800000"

    def __init__(self, root):
        import tkinter as tk
        from tkinter import ttk, filedialog, scrolledtext, messagebox

        self.root = root
        self.root.title("Medical Records Bot - Version 3.1.0, Last Updated 12/04/2025")
        self.root.geometry("1020x760")
        self.root.configure(bg=self.MAROON)

        style = ttk.Style()
        try: style.theme_use("clam")
        except Exception: pass
        style.configure('TLabel', background=self.MAROON, foreground='#ffffff', font=("Helvetica", 11))
        style.configure('Header.TLabel', background=self.MAROON, foreground='#ffffff', font=("Helvetica", 22, 'bold'))
        style.configure('TButton', font=("Helvetica", 11, 'bold'))
        style.configure('Card.TFrame', background='#faf7f7', relief='flat')
        style.configure('TCheckbutton', background=self.MAROON, foreground='#ffffff', font=("Helvetica", 11, 'bold'))

        ttk.Label(self.root, text="Medical Records Bot", style='Header.TLabel', anchor='center').pack(pady=(14, 6), fill='x')

        container = ttk.Frame(self.root); container.pack(fill="both", expand=True)
        self.canvas = tk.Canvas(container, background=self.MAROON, highlightthickness=0)
        vscroll = ttk.Scrollbar(container, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=vscroll.set)
        self.canvas.pack(side="left", fill="both", expand=True); vscroll.pack(side="right", fill="y")
        self.body = ttk.Frame(self.canvas, style='Card.TFrame')
        self.win = self.canvas.create_window((0, 0), window=self.body, anchor="nw")

        def _on_frame_configure(_): self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        def _on_canvas_configure(event): self.canvas.itemconfig(self.win, width=event.width)
        self.body.bind("<Configure>", _on_frame_configure); self.canvas.bind("<Configure>", _on_canvas_configure)

        def card(parent):
            f = ttk.Frame(parent, style='Card.TFrame'); f.configure(padding=(16, 14)); return f

        import tkinter as tk
        self.xlsx_path = ""
        self.selected_sheet = tk.StringVar()
        self.tn_user = tk.StringVar()
        self.tn_pass = tk.StringVar()
        self.headless = tk.BooleanVar(value=False)
        self.fast = tk.BooleanVar(value=True)
        self.limit = tk.StringVar(value="")
        self._stop_event = None
        self._worker_thread = None

        files = card(self.body); files.pack(padx=14, pady=8, fill='x')
        ttk.Button(files, text="Select Excel", command=self.pick_excel).grid(row=0, column=0, padx=6, pady=6, sticky='w')
        self.excel_label = ttk.Label(files, text="No Excel selected"); self.excel_label.grid(row=0, column=1, sticky='w', padx=8)
        ttk.Label(files, text="Sheet:").grid(row=1, column=0, sticky='e', padx=6, pady=4)
        self.sheet_combo = ttk.Combobox(files, textvariable=self.selected_sheet, width=36, state="readonly")
        self.sheet_combo.grid(row=1, column=1, sticky='w', padx=6, pady=4)

        cred = card(self.body); cred.pack(padx=14, pady=8, fill='x')
        ttk.Label(cred, text="TN Username:").grid(row=0, column=0, sticky='e', padx=6, pady=4)
        ttk.Entry(cred, textvariable=self.tn_user, width=35).grid(row=0, column=1, sticky='w', padx=6, pady=4)
        ttk.Label(cred, text="TN Password:").grid(row=1, column=0, sticky='e', padx=6, pady=4)
        ttk.Entry(cred, textvariable=self.tn_pass, show='*', width=35).grid(row=1, column=1, sticky='w', padx=6, pady=4)

        opts = card(self.body); opts.pack(padx=14, pady=8, fill='x')
        ttk.Checkbutton(opts, text="Headless Chrome", variable=self.headless).grid(row=0, column=0, sticky='w', padx=6, pady=2)
        ttk.Checkbutton(opts, text="Fast mode (block images)", variable=self.fast).grid(row=0, column=1, sticky='w', padx=12, pady=2)
        ttk.Label(opts, text="Limit clients:").grid(row=1, column=0, sticky='e', padx=6, pady=2)
        ttk.Entry(opts, textvariable=self.limit, width=10).grid(row=1, column=1, sticky='w', padx=6, pady=2)

        ctl = card(self.body); ctl.pack(padx=14, pady=8, fill='x')
        self.start_btn = ttk.Button(ctl, text="Start", command=self.start); self.start_btn.grid(row=0, column=0, padx=6)
        self.stop_btn = ttk.Button(ctl, text="Stop", command=self.stop, state="disabled"); self.stop_btn.grid(row=0, column=1, padx=6)

        log_card = card(self.body); log_card.pack(padx=14, pady=10, fill='both', expand=True)
        self.log_area = scrolledtext.ScrolledText(log_card, width=120, height=26, bg="#f5f5f5", fg="#000000", font=("Courier New", 10))
        self.log_area.pack(padx=4, pady=4, fill="both", expand=True)

        self._log_queue = []
        self.root.after(100, self.flush_log_queue)
        self._openpyxl_ok = self._check_openpyxl()

    # ----- GUI helpers -----
    def _check_openpyxl(self):
        try:
            import openpyxl  # noqa
            return True
        except Exception:
            return False

    def enqueue_log(self, msg):
        self._log_queue.append(msg)

    def flush_log_queue(self):
        while self._log_queue:
            msg = self._log_queue.pop(0)
            try:
                self.log_area.insert("end", msg + "\n"); self.log_area.see("end")
            except Exception:
                pass
        self.root.after(100, self.flush_log_queue)

    def pick_excel(self):
        from tkinter import filedialog, messagebox
        try:
            from openpyxl import load_workbook
        except Exception:
            self.enqueue_log("[ERROR] openpyxl is not installed. Run: pip install openpyxl")
            return
        f = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not f: return
        self.xlsx_path = f
        self.excel_label.config(text=f)
        self.enqueue_log(f"Selected Excel: {f}")
        try:
            wb = load_workbook(f, data_only=True)
            names = wb.sheetnames
            self.sheet_combo["values"] = names
            if names: self.selected_sheet.set(names[0])
        except Exception as e:
            messagebox.showerror("Workbook error", f"Could not read workbook:\n{e}")

    def start(self):
        import threading, time
        from tkinter import messagebox
        if not self.xlsx_path or not self.selected_sheet.get() or not self.tn_user.get() or not self.tn_pass.get():
            messagebox.showwarning("Missing info", "Please select Excel + sheet and enter TN credentials.")
            return
        def gui_log(m): self.enqueue_log(time.strftime("[%H:%M:%S] ") + str(m))
        globals()["log"] = gui_log

        self.start_btn.config(state="disabled"); self.stop_btn.config(state="normal")
        self.enqueue_log("Starting…")
        self.enqueue_log(f"  Sheet: {self.selected_sheet.get()} | Headless: {self.headless.get()} | Fast mode: {self.fast.get()}")

        import threading
        self._stop_event = threading.Event()
        def worker():
            try:
                lim = None
                s = self.limit.get().strip()
                if s:
                    try: lim = int(s)
                    except Exception:
                        self.enqueue_log("Limit is not an integer; ignoring.")
                        lim = None
                run(
                    xlsx_path=self.xlsx_path,
                    sheet_name=self.selected_sheet.get(),
                    tn_user=self.tn_user.get().strip(),
                    tn_pass=self.tn_pass.get().strip(),
                    headless=self.headless.get(),
                    fast=self.fast.get(),
                    limit=lim,
                    stop_event=self._stop_event
                )
            except Exception as e:
                self.enqueue_log(f"[FATAL] {e}")
            finally:
                self.enqueue_log("Process finished.")
                try:
                    self.start_btn.config(state="normal"); self.stop_btn.config(state="disabled")
                except Exception:
                    pass
        self._worker_thread = threading.Thread(target=worker, daemon=True)
        self._worker_thread.start()

    def stop(self):
        if hasattr(self, "_stop_event") and self._stop_event:
            self.enqueue_log("Stop requested…")
            self._stop_event.set()

if __name__ == "__main__":
    if len(sys.argv) > 1:
        parser = argparse.ArgumentParser(description="ISWS Medical Records Bot (Progress, IA, TP & PCP Contact Notes)")
        parser.add_argument("--excel", required=True, help="Path to the Excel workbook (.xlsx)")
        parser.add_argument("--sheet", required=True, help="Worksheet name")
        parser.add_argument("--user", required=True, help="TherapyNotes username")
        parser.add_argument("--password", required=True, help="TherapyNotes password")
        parser.add_argument("--headless", action="store_true", help="Run Chrome headless")
        parser.add_argument("--slow", action="store_true", help="Disable fast mode (images etc.)")
        parser.add_argument("--limit", type=int, default=None, help="Max number of clients to run")
        args = parser.parse_args()
        run(
            xlsx_path=args.excel,
            sheet_name=args.sheet,
            tn_user=args.user,
            tn_pass=args.password,
            headless=args.headless,
            fast=not args.slow,
            limit=args.limit,
        )
    else:
        import tkinter as tk
        root = tk.Tk()
        app = MedicalRecordsGUI(root)
        root.mainloop()
